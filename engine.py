"""
Artemis Document Engine
-----------------------
Fills a brand's QTN / INV / DO Excel templates with data, names files using the
company convention, converts to PDF, and indexes an existing folder of documents.

Filename convention (new docs, tagged with the issuing brand):
    BRAND_TYPE_NUMBER_REV_COMPANY_PROJECT_DATE.ext
e.g. ARTEMIS_QTN_0042_R0_Resinal-Developments_Facade-Lighting_2026-06-30.xlsx

Legacy convention (files created before multi-brand support, no brand tag):
    TYPE_NUMBER_REV_COMPANY_PROJECT_DATE.ext
"""

import os, sys, re, copy, shutil, subprocess, datetime, glob, base64, json, zipfile, html, time
from io import BytesIO
from xml.etree import ElementTree as ET
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font

# BASE / DATA_BASE split exists for exactly one reason: a PyInstaller-frozen
# .exe extracts its bundled files (this code, templates_html, static/) into
# a fresh temp directory every single launch (sys._MEIPASS) — anything the
# app WRITES there (config.json, drafts, submissions, generated caches)
# would silently vanish the moment the app closes. BASE stays the
# read-only bundled-resource root (templates/*.xlsx, and — via
# html_engine's own identical split — templates_html/, static/fonts,
# static/doc_html); DATA_BASE is where persistent read/write data lives,
# anchored to the real, permanent folder the .exe itself sits in so an
# existing config.json/drafts/submissions sitting next to a freshly built
# .exe (this app's own folder, unchanged) is picked up with zero
# migration. Both collapse to the same plain "this file's own directory"
# in normal (non-frozen) `python app.py` use — dev behavior is unchanged.
def _resource_base():
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

def _data_base():
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))

BASE = _resource_base()
DATA_BASE = _data_base()
TEMPLATES = os.path.join(BASE, "templates")

DOC_TYPES = ["INV", "DO", "PI", "RV", "PO", "LPO", "RFQ", "CN", "DN", "QTN2", "CAT", "EXP"]  # extend freely
TYPE_LABEL = {
    "INV": "Tax Invoice", "DO": "Delivery Order", "PI": "Proforma Invoice", "RV": "Payment Receipt",
    "PO": "Purchase Order", "LPO": "Local Purchase Order", "RFQ": "Request for Quotation",
    "CN": "Credit Note", "DN": "Debit Note",
    "QTN2": "Quotation",
    "CAT": "Sololuce Datasheet",
    "EXP": "Expense Report",
}

# Doc types rendered via html_engine.py (Playwright + Jinja2, pixel-fidelity
# HTML/CSS design pack) instead of the openpyxl/LibreOffice xlsx pipeline —
# PDF-only, no .xlsx sibling. See html_engine.RENDERERS. PI/RV/CN (Proforma
# Invoice, Payment Receipt, Credit Note) are generate-only-from-another-
# document for now (see the "Generate ▾" menu on a Quotation/Proforma
# Invoice/Delivery Order/Tax Invoice row) — same pixel-fidelity
# design-handoff pipeline as the others, just no dedicated sidebar "start
# blank" entry yet.
HTML_DOC_TYPES = {"QTN2", "CAT", "EXP", "PI", "RV", "CN"}

# INV/DO are NOT in HTML_DOC_TYPES — they keep their editable .xlsx (some
# workflows/accounting depend on it existing), so every other xlsx-pipeline
# behavior (FILLERS, sidecar-free item reading, etc.) still applies to them.
# This second set only decides which pipeline renders their PDF: the .xlsx
# is still filled by FILLERS as always, but the customer-facing PDF is
# rendered through html_engine's pixel-fidelity pipeline (see generate()
# below) so it visually matches QTN2 instead of a plain LibreOffice-
# converted spreadsheet. See html_engine.RENDERERS for the render functions.
HTML_PDF_DOC_TYPES = {"INV", "DO"}

# ----------------------------------------------------------------------------
# Brands — each has its own subfolder under templates/ with its own QTN/INV/DO
# Excel files. Until real templates are supplied per brand, new brands fall
# back to the Artemis templates so the switcher is fully usable end to end.
# ----------------------------------------------------------------------------
DEFAULT_BRAND = "ARTEMIS"
BRANDS = {
    "ARTEMIS": "Artemis Lightings",
    "SOLOLUCE": "Sololuce Lightings",
    "ADS": "ADS Lightings",
    "WATT": "Watt Electricals",
}

EXTERNAL_TEMPLATES = {}  # {brand: optional external templates folder}, set via Settings

def template_path(brand, doc_type):
    """Path to a brand's template for a doc type. Checks that brand's
    user-configured external templates folder first (if set), then the app's
    own bundled templates, falling back to Artemis's template if that brand
    has no template of its own yet."""
    brand = (brand or DEFAULT_BRAND).upper()
    doc_type = doc_type.upper()
    roots = [r for r in (EXTERNAL_TEMPLATES.get(brand), TEMPLATES) if r]
    for root in roots:
        p = os.path.join(root, brand, f"{doc_type}.xlsx")
        if os.path.exists(p):
            return p
    for root in roots:
        p = os.path.join(root, DEFAULT_BRAND, f"{doc_type}.xlsx")
        if os.path.exists(p):
            return p
    return os.path.join(TEMPLATES, DEFAULT_BRAND, f"{doc_type}.xlsx")

# ----------------------------------------------------------------------------
# Filename convention
# ----------------------------------------------------------------------------
def _slug(s):
    """Make a token safe for a filename: spaces->hyphens, strip separators.
    Strips HTML tags wholesale first (not just angle brackets) — QTN2's rich
    text fields can contain <b>/<i>/<font color=...> markup, and stripping
    only '<'/'>' would leave the tag *name* behind (e.g. "<b>Foo</b>" would
    otherwise slug to "bFoo-b", not "Foo")."""
    s = (s or "").strip()
    s = re.sub(r"<[^>]*>", "", s)
    s = re.sub(r"[_/\\]+", "-", s)
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"[^A-Za-z0-9\-.]", "", s)
    return s or "NA"

def build_filename(doc_type, number, rev, company, project, date_iso, brand=None, ext="xlsx"):
    """Compose a filename following the company convention. brand=None omits
    the brand tag (used only for backward-compat testing; generate() always
    passes a brand for new files)."""
    num = str(number).strip()
    if num.isdigit():
        num = num.zfill(4)
    rev_token = f"R{re.sub('[^0-9]', '', str(rev)) or '0'}"
    prefix = f"{brand.upper()}_" if brand else ""
    return (f"{prefix}{doc_type.upper()}_{num}_{rev_token}_"
            f"{_slug(company)}_{_slug(project)}_{date_iso}.{ext}")

_FN_RE_BRANDED = re.compile(
    r"^(?P<brand>[A-Za-z]+)_(?P<type>[A-Za-z]+\d*)_(?P<number>[^_]+)_R(?P<rev>\d+)_"
    r"(?P<company>[^_]+)_(?P<project>[^_]+)_(?P<date>\d{4}-\d{2}-\d{2})\.(?P<ext>\w+)$"
)
_FN_RE = re.compile(
    r"^(?P<type>[A-Za-z]+\d*)_(?P<number>[^_]+)_R(?P<rev>\d+)_"
    r"(?P<company>[^_]+)_(?P<project>[^_]+)_(?P<date>\d{4}-\d{2}-\d{2})\.(?P<ext>\w+)$"
)

def parse_filename(fname):
    """Parse a filename into structured fields. Returns dict or None if it doesn't
    match either the branded (new) or legacy (pre-multi-brand) convention."""
    base = os.path.basename(fname)
    m = _FN_RE_BRANDED.match(base)
    if m and m.group("brand").upper() in BRANDS:
        d = m.groupdict()
        d["rev"] = int(d["rev"])
        d["company_label"] = d["company"].replace("-", " ")
        d["project_label"] = d["project"].replace("-", " ")
        d["type_label"] = TYPE_LABEL.get(d["type"].upper(), d["type"].upper())
        d["brand"] = d["brand"].upper()
        d["brand_label"] = BRANDS[d["brand"]]
        return d
    m = _FN_RE.match(base)
    if not m:
        return None
    d = m.groupdict()
    d["rev"] = int(d["rev"])
    d["company_label"] = d["company"].replace("-", " ")
    d["project_label"] = d["project"].replace("-", " ")
    d["type_label"] = TYPE_LABEL.get(d["type"].upper(), d["type"].upper())
    d["brand"] = None
    d["brand_label"] = None
    return d

# ----------------------------------------------------------------------------
# Folder indexer  (gives the software its view of "all documents")
# ----------------------------------------------------------------------------
def index_folder(folder, recursive=True):
    """Scan a folder, parse every file that matches the convention, return sorted records."""
    pattern = "**/*" if recursive else "*"
    records = []
    for path in glob.glob(os.path.join(folder, pattern), recursive=recursive):
        if not os.path.isfile(path):
            continue
        meta = parse_filename(path)
        if not meta:
            continue
        meta["path"] = path
        meta["ext_is_pdf"] = meta["ext"].lower() == "pdf"
        records.append(meta)
    # newest first, by date then number
    records.sort(key=lambda r: (r["date"], r["number"]), reverse=True)
    return records

def previous_for_company(folder, company, doc_type="QTN2", limit=3):
    """Last N documents of a type for a given company (by filename match).
    xlsx-pipeline types (QTN/INV/DO) are keyed on their .xlsx (the canonical,
    editable file); HTML_DOC_TYPES (QTN2, ...) have no .xlsx twin, so their
    .pdf is canonical instead."""
    target = _slug(company).lower()
    canonical_ext = "pdf" if doc_type.upper() in HTML_DOC_TYPES else "xlsx"
    recs = [r for r in index_folder(folder)
            if r["type"].upper() == doc_type.upper()
            and _slug(r["company_label"]).lower() == target
            and r["ext"].lower() == canonical_ext]
    # de-duplicate by number+rev (xlsx/pdf pairs share the stem)
    seen, out = set(), []
    for r in recs:
        key = (r["number"], r["rev"])
        if key in seen:
            continue
        seen.add(key); out.append(r)
    return out[:limit]

# ----------------------------------------------------------------------------
# HTML_DOC_TYPES sidecar — QTN2 etc. have no .xlsx to read line items back
# out of (read_items_from_doc needs openpyxl/a real workbook), so a small
# JSON file is saved next to the .pdf with the exact data used to generate
# it. Powers "previous quotations for this company"; also lays groundwork
# for round-trip editing later, though that's not wired up yet.
# ----------------------------------------------------------------------------
def sidecar_path(doc_path):
    return os.path.splitext(doc_path)[0] + ".json"

def save_sidecar(doc_path, data):
    try:
        with open(sidecar_path(doc_path), "w", encoding="utf-8") as f:
            json.dump(data, f)
    except OSError:
        pass

def read_sidecar(doc_path):
    """Full generation payload (header fields + items + discount/vat), or {}
    if there's no sidecar / it's unreadable."""
    try:
        with open(sidecar_path(doc_path), encoding="utf-8") as f:
            return json.load(f)
    except (OSError, ValueError):
        return {}

def read_sidecar_items(doc_path):
    return read_sidecar(doc_path).get("items", [])

# ----------------------------------------------------------------------------
# Markdown companion — a plain-text rendition of a document's full generation
# payload (same source data as the sidecar above), saved as a same-name .md
# sibling next to every generated document. The JSON sidecar already has
# everything, but it's bulky (product photos alone are base64 blobs hundreds
# of KB long) and not meant for a human or an AI assistant to just open and
# read — this is: every real field, in full, as plain readable text/tables,
# photos noted by name rather than embedded, so the document's complete
# content can be read in one pass and edited directly in a text editor
# without wading through JSON or opening the PDF. Refreshed on every
# Generate, same lifecycle as the sidecar.
# ----------------------------------------------------------------------------
_MD_BR_RE = re.compile(r"<br\s*/?>", re.I)
_MD_TAG_RE = re.compile(r"<[^>]+>")

def _md_html_to_text(html_str):
    """Best-effort plain text for the small amount of rich-text HTML this
    app's contenteditable fields produce (Attn/Address boxes — mainly <br>
    line breaks, occasionally <b>/<i>). Not a general HTML parser."""
    if not html_str:
        return ""
    s = _MD_BR_RE.sub("\n", html_str)
    s = _MD_TAG_RE.sub("", s)
    return html.unescape(s).strip()

def _md_table(headers, rows):
    rows = [r for r in rows if any(str(c or "").strip() for c in r)]
    if not rows:
        return "_(none)_"
    esc = lambda v: str(v if v is not None else "").replace("|", "\\|").replace("\n", "<br>")
    lines = ["| " + " | ".join(headers) + " |", "|" + "|".join([" --- "] * len(headers)) + "|"]
    for r in rows:
        lines.append("| " + " | ".join(esc(c) for c in r) + " |")
    return "\n".join(lines)

def _md_money_rule(d):
    if not d or not d.get("enabled"):
        return "Off"
    mode, val = d.get("mode", "amount"), d.get("value", 0)
    if mode == "percent":
        return f"{val}%"
    if mode == "target":
        return f"target total {val}"
    return str(val)

def build_markdown(doc_type, data, brand=None, badge_library=None):
    """Returns the full Markdown text for one generated document's data."""
    label = TYPE_LABEL.get(doc_type, doc_type)
    brand_name = BRANDS.get(brand or DEFAULT_BRAND, brand or DEFAULT_BRAND)
    L = [f"# {label} {data.get('number', '')}".rstrip(), "",
         f"_{brand_name} — Office Tool_", ""]

    if doc_type == "CAT":
        L.append(f"**Product:** {data.get('company', '')}")
        L.append(f"**Series / Category:** {data.get('project') or data.get('series', '')}")
        if data.get("family"):
            L.append(f"**Family:** {data['family']}")
        L.append(f"**Product Type:** {data.get('product_type', '')}")
        L.append(f"**Page Number:** {data.get('page_number', '')}")
        L.append("")
        if data.get("description"):
            L += ["## Description", "", data["description"], ""]
        specs = data.get("specs") or []
        L += ["## Technical Specifications", "",
              _md_table(["Label", "Value"], [[s.get("label", ""), s.get("value", "")] for s in specs]), ""]
        if data.get("note"):
            L += ["_" + data["note"] + "_", ""]
        finish = data.get("finish_colors") or []
        if finish:
            L += ["## Finish Colors", "",
                  ", ".join(f"{f.get('label', '')} ({f.get('hex', '')})" for f in finish), ""]
        badges = data.get("badges") or []
        if badges:
            badge_labels = {b.get("key"): b.get("label") for b in (badge_library or [])}
            L += ["## Badges", "",
                  ", ".join(badge_labels.get(b.get("key"), b.get("key", "")) for b in badges if b.get("key")), ""]
        cols = data.get("ordering_columns") or []
        if cols:
            n = max((len(c.get("values") or []) for c in cols), default=0)
            headers = [c.get("label", "") or "(untitled)" for c in cols]
            rows = [[(cols[ci].get("values") or [""] * n)[vi] if vi < len(cols[ci].get("values") or []) else ""
                      for ci in range(len(cols))] for vi in range(n)]
            L.append("## Ordering Table")
            L.append("")
            if data.get("ordering_code_example"):
                L.append(f"Ordering Code Example: {data['ordering_code_example']}")
                L.append("")
            L.append(_md_table(headers, rows))
            L.append("")
        photos = [("Main Product Photo", data.get("main_photo")), ("Application Photo", data.get("lifestyle_photo")),
                  ("Dimension Diagram", data.get("dimension_diagram"))]
        attached = [name for name, v in photos if v]
        if attached:
            L += ["## Photos", "", "Attached in the PDF (not embedded here): " + ", ".join(attached), ""]

    elif doc_type == "EXP":
        L.append(f"**Employee:** {data.get('company', '')}")
        L.append(f"**Category:** {data.get('category', '')}")
        L.append(f"**Currency:** {data.get('currency', 'AED')}")
        if data.get("project"):
            L.append(f"**Period:** {data['project'].replace('_to_', ' to ')}")
        L.append("")
        rows = data.get("rows") or []
        total = sum(float(r.get("amount") or 0) for r in rows)
        L += ["## Expense Lines", "",
              _md_table(["Date", "Product", "Description", "Payment Method", "Amount"],
                        [[r.get("date", ""), r.get("product", ""), r.get("description", ""),
                          r.get("payment_method", ""), r.get("amount", "")] for r in rows]),
              "", f"**Total: {total:,.2f} {data.get('currency', 'AED')}**", ""]

    elif doc_type == "QTN2":
        L.append(f"**Company:** {data.get('company', '')}")
        L.append(f"**Project:** {_md_html_to_text(data.get('project', ''))}")
        if data.get("area"):
            L.append(f"**Area:** {_md_html_to_text(data['area'])}")
        L.append(f"**Revision:** R{data.get('rev', 0)}")
        L.append(f"**Status:** {data.get('status', 'Draft')}")
        L.append("")
        attn = _md_html_to_text(data.get("customer_attn", ""))
        addr = _md_html_to_text(data.get("customer_address", ""))
        if attn or addr:
            L.append("## Customer")
            L.append("")
            if attn:
                L.append(f"Attn: {attn}")
            if addr:
                L.append(addr)
            L.append("")
        items = data.get("items") or []
        L += ["## Items", "",
              _md_table(["Type", "Description", "Unit", "Qty", "Price"],
                        [[it.get("type", ""), it.get("description", ""), it.get("unit", "PCS"),
                          it.get("qty", ""), it.get("price", "")] for it in items]), ""]
        disc, vat = data.get("discount") or {}, data.get("vat") or {}
        L += [f"- Discount: {_md_money_rule(disc)}", f"- VAT: {_md_money_rule(vat)}", ""]
        terms = data.get("terms") or {}
        if terms:
            L += ["## Terms", "",
                  f"- Delivery: {terms.get('delivery', '')}",
                  f"- Payment: {terms.get('payment', '')}",
                  f"- Warranty: {terms.get('warranty', '')}", ""]

    elif doc_type in ("INV", "DO"):
        L.append(f"**Company:** {data.get('company', '')}")
        if data.get("project"):
            L.append(f"**Project:** {data['project']}")
        if doc_type == "INV":
            if data.get("qtn_number"):
                L.append(f"**QTN Number:** {data['qtn_number']}")
            if data.get("type"):
                L.append(f"**Type:** {data['type']}")
        if data.get("lpo_number"):
            L.append(f"**LPO Number:** {data['lpo_number']}")
        L.append("")
        if data.get("customer_block"):
            L += ["## Customer", "", data["customer_block"], ""]
        items = data.get("items") or []
        if doc_type == "INV":
            L += ["## Items", "",
                  _md_table(["Description", "Unit", "Qty", "Price"],
                            [[it.get("description", ""), it.get("unit", "PCS"),
                              it.get("qty", ""), it.get("price", "")] for it in items]), ""]
            disc, vat = data.get("discount") or {}, data.get("vat") or {}
            L += [f"- Discount: {_md_money_rule(disc)}", f"- VAT: {_md_money_rule(vat)}", ""]
        else:
            L += ["## Items", "",
                  _md_table(["Description", "Unit", "LPO Qty", "Prev. Delivery", "Delivered"],
                            [[it.get("description", ""), it.get("unit", "PCS"), it.get("lpo_qty", ""),
                              it.get("prev_delivery", ""), it.get("delivered", "")] for it in items]), ""]
    else:
        # Generic fallback for any doc type without dedicated formatting above.
        L.append(f"**Company:** {data.get('company', '')}")
        if data.get("project"):
            L.append(f"**Project:** {data['project']}")
        L.append("")
        items = data.get("items") or []
        if items:
            keys = sorted({k for it in items for k in it.keys() if k != "photo"})
            L += ["## Items", "", _md_table(keys, [[it.get(k, "") for k in keys] for it in items]), ""]

    return "\n".join(L).rstrip() + "\n"

def markdown_path(doc_path):
    return os.path.splitext(doc_path)[0] + ".md"

def save_markdown(doc_path, doc_type, data, brand=None, badge_library=None):
    """Writes the Markdown companion next to a generated document. Never
    raises — a failed write here shouldn't block the real PDF/xlsx from
    being saved, same defensive stance as save_sidecar."""
    try:
        with open(markdown_path(doc_path), "w", encoding="utf-8") as f:
            f.write(build_markdown(doc_type, data, brand, badge_library))
    except OSError:
        pass

# ----------------------------------------------------------------------------
# Broad importer — reads EVERY file in the folder (not just convention-named
# ones), grouping by the company subfolder it lives in. Legacy files get their
# type/date/number guessed from the filename so they still show up and sort
# correctly; only convention-named files are marked "editable" (inline edit
# needs to know the exact cell layout, which we only know for those).
# ----------------------------------------------------------------------------
DOC_EXT = {".xlsx", ".xls", ".pdf", ".doc", ".docx"}

_TYPE_KEYWORDS = [
    ("QTN", re.compile(r"(?<![A-Za-z])QTN(?![A-Za-z])|QUOT", re.I)),
    ("INV", re.compile(r"(?<![A-Za-z])INV(?![A-Za-z])|INVOICE", re.I)),
    ("DO", re.compile(r"(?<![A-Za-z])D\.?O\.?(?![A-Za-z])|DELIVERY", re.I)),
    ("LPO", re.compile(r"(?<![A-Za-z])LPO(?![A-Za-z])", re.I)),
    ("PO", re.compile(r"(?<![A-Za-z])P\.?O\.?(?![A-Za-z])|PURCHASE", re.I)),
]
_DATE_PATTERNS = [
    (re.compile(r"(20\d{2})[-_.](\d{1,2})[-_.](\d{1,2})"), lambda m: (int(m[1]), int(m[2]), int(m[3]))),
    (re.compile(r"(\d{1,2})[-_.](\d{1,2})[-_.](20\d{2})"), lambda m: (int(m[3]), int(m[2]), int(m[1]))),
]
_NUMBER_RE = re.compile(r"(\d{3,6})")

def _guess_type(name):
    for t, pat in _TYPE_KEYWORDS:
        if pat.search(name):
            return t
    return "DOC"

# Common catch-all folder names that show up in real archives but are never
# themselves a client (year bins, misc buckets, doc-type buckets like
# "ABU DHABI QTN"). Used by the client-import feature to keep scan_all's
# folder-name-derived company labels from filling the client list with junk —
# heuristic, not exhaustive, so the user still reviews/prunes after import.
_NON_CLIENT_LABELS = {
    "cash", "misc", "miscellaneous", "others", "other", "temp", "temporary",
    "old", "new", "archive", "backup", "copy", "draft", "drafts", "general",
    "sample", "samples", "test", "root",
}

def looks_like_non_client_label(label):
    s = (label or "").strip()
    if not s:
        return True
    if s.replace(" ", "").isdigit():
        return True
    if s.lower() in _NON_CLIENT_LABELS:
        return True
    for _, pat in _TYPE_KEYWORDS:
        if pat.search(s):
            return True
    return False

# ----------------------------------------------------------------------------
# Legacy contact-block extraction (client-import "get the rest of the
# details too" feature). Pre-app QTN/INV/DO xlsx files were hand-built in
# Excel with the customer's company/attn/phone/address/email typed into a
# free-floating "letterhead" text box (a DrawingML shape) rather than real
# cell values — openpyxl's normal cell reader can't see this at all (it
# silently drops shape/drawing content), so this reads the raw drawing XML
# straight out of the xlsx zip. Best-effort and heuristic: real files vary a
# lot in wording/layout over the years, so this returns {} rather than
# guessing whenever it can't confidently isolate a single customer-info shape.
# ----------------------------------------------------------------------------
_CONTACT_LABEL_SHAPES = {
    "from", "to", "buyer", "seller", "buyer:", "seller:", "bill to", "ship to",
    "tax invoice", "quotation", "delivery note", "delivery order", "invoice",
    "d.o.", "do",
}
_OWN_BRAND_TOKENS = ["artemis electric", "ads lighting", "sololuce", "watt electric"]

_MOBILE_LABEL_RE = re.compile(r"^(mob(ile)?|phone|cell|contact\s*no\.?)\s*(no\.?|number)?\s*[:.]?\s*", re.I)
_LANDLINE_LABEL_RE = re.compile(r"^(tel(ephone)?|landline)\s*(no\.?|number)?\s*[:.]?\s*", re.I)
_FAX_LABEL_RE = re.compile(r"^fax\s*(no\.?|number)?\s*[:.]?\s*", re.I)
_EMAIL_LABEL_RE = re.compile(r"^(email|e-?mail)\s*[:.]?\s*$", re.I)
_WEBSITE_LABEL_RE = re.compile(r"^web(site)?\s*[:.]?\s*$", re.I)
_EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+")
_WEBSITE_RE = re.compile(r"(www\.|https?://)", re.I)
_ADDRESS_RE = re.compile(
    r"(p\.?\s*o\.?\s*box|dubai|abu dhabi|sharjah|\bshj\b|ajman|fujairah|"
    r"ras al khaimah|umm al quwain|u\.?a\.?e\.?)", re.I)
_TRN_RE = re.compile(r"\btrn\b", re.I)
_PHONE_VALUE_RE = re.compile(r"^[+\d][\d\s\-]{5,}$")
# Stricter than _PHONE_VALUE_RE: only for a fragment with NO preceding
# "Tel:"/"Mob:" label at all, so there's no context clue it's a phone number
# — requires a leading "+" or an internal separator (how people actually
# type phone numbers), rejecting a solid unbroken digit run. Without this,
# a bare 15-digit TRN/reference number (no spaces/dashes at all) matched the
# same as a real phone number and got misfiled as one.
_BARE_PHONE_RE = re.compile(r"^(\+[\d\s\-]{5,}|\d{2,4}[\s\-][\d\s\-]{4,})$")
_PO_BOX_RE = re.compile(r"p\.?\s*o\.?\s*box[:\s]*(\d{2,8})", re.I)
_UAE_CITY_RE = re.compile(
    r"\b(dubai|abu dhabi|sharjah|ajman|fujairah|ras al khaimah|umm al quwain|al ain)\b", re.I)
_UAE_COUNTRY_RE = re.compile(r"\bu\.?a\.?e\.?\b|united arab emirates", re.I)

# Words like "unit"/"floor"/"building"/"industrial" show up constantly
# inside ordinary company names too ("AL GHAITH BUILDING CONSTRUCTION",
# "GULF INDUSTRIAL SERVICES") — a bare keyword search misclassified both of
# those company names as an address. Requiring a number right next to the
# word (a real "Unit 608"/"Villa 196"/"Plot 942" reference) is what actually
# distinguishes an address line from those company names, so that's the bar
# here; only "business bay"/"free zone" (specific place names, not generic
# words) are kept as bare phrase matches.
_BUILDING_WORD_RE = re.compile(
    r"\b(unit|floor|tower|building|bldg\.?|villa|suite|plot)\s*[:#]?\s*\d|"
    r"\d+[\s,]*(st\.?|street|rd\.?|road)\b|"
    r"\bbusiness bay\b|\bfree zone\b", re.I)
_DRAWING_PART_RE = re.compile(r"xl/drawings/drawing\d+\.xml$")
_SHAPE_RE = re.compile(r"<xdr:sp[ >].*?</xdr:sp>", re.S)
_TEXT_RUN_RE = re.compile(r"<a:t>(.*?)</a:t>", re.S)

def _contact_shape_frags(xlsx_path):
    """One text-fragment list per DrawingML text-box shape, across every
    drawing part in the workbook."""
    try:
        z = zipfile.ZipFile(xlsx_path)
    except (OSError, zipfile.BadZipFile):
        return []
    out = []
    for name in z.namelist():
        if not _DRAWING_PART_RE.search(name):
            continue
        try:
            data = z.read(name).decode("utf-8", errors="replace")
        except (KeyError, OSError):
            continue
        for sp in _SHAPE_RE.findall(data):
            frags = [t.strip() for t in _TEXT_RUN_RE.findall(sp) if t.strip()]
            if frags:
                out.append(frags)
    return out

def _clean_frag(raw):
    return re.sub(r"\s+", " ", html.unescape(raw).replace("\xa0", " ")).strip()

def _is_contact_label_shape(frags):
    joined = re.sub(r"\s+", " ", " ".join(frags)).strip().lower().rstrip(":")
    return joined in _CONTACT_LABEL_SHAPES

def _is_own_letterhead_shape(frags):
    first = frags[0].lower()
    return any(tok in first for tok in _OWN_BRAND_TOKENS)

def _strip_known_company_prefix(text, company_name):
    """If `text` starts with the already-trusted company name (matched as a
    literal string, case-insensitive, tolerant of extra punctuation/spacing
    between words), remove exactly that prefix. Regex-anchored (`re.match`)
    rather than the word-by-word accumulation this used to do — that greedy
    approach had a bug where, once the accumulated text was merely a prefix
    of the target, it kept accepting every further word forever (since a
    longer string still "starts with" a shorter one), silently swallowing
    the real attn/contact-person text into the "already matched" pile."""
    if not company_name:
        return text
    pattern = re.sub(r"\\ ", r"[\\s,.-]*", re.escape(company_name.strip()))
    m = re.match(pattern, text, re.I)
    if m and m.end() < len(text):
        return text[m.end():].strip(" ,.-")
    return text

def _classify_contact_frags(frags, company_name=None):
    """Field classification is prefix-anchored for phone/landline/fax (the
    label word must START the fragment, e.g. "Tel: 04-..." or a bare "Mob:"
    on its own) rather than a free substring search — "cell" as a raw
    substring search matches inside ordinary words like "Excellence", which
    silently corrupted the phone field for at least one real client before
    this was anchored. Email/website/TRN stay substring-based since those
    tokens (an "@", "www.", "TRN") essentially never appear as false
    positives inside unrelated running text.
    Address-like fragments (PO Box, a UAE city/country name, or a building/
    unit/floor/tower word) are pulled into address/po_box/city/country
    instead of falling through to the leftover "attn" text — a fragment
    like "Unit 608 Business Point Tower" used to land in attn because
    nothing there matched the (narrower) old address check."""
    result, leading, pending = {}, [], None
    for raw in frags:
        frag = _clean_frag(raw)
        if not frag:
            continue

        if _EMAIL_LABEL_RE.match(frag) or _WEBSITE_LABEL_RE.match(frag):
            pending = None; continue  # bare "Email:"/"Website:" — value self-identifies via regex wherever it lands

        m_fax = _FAX_LABEL_RE.match(frag)
        m_land = _LANDLINE_LABEL_RE.match(frag)
        m_mob = _MOBILE_LABEL_RE.match(frag)
        if m_fax or m_land or m_mob:
            m = m_fax or m_land or m_mob
            value = frag[m.end():].strip()
            kind = "fax" if m_fax else ("landline" if m_land else "phone")
            if value:
                if kind != "fax":
                    result.setdefault(kind, value)
                pending = None
            else:
                pending = kind  # bare label — value is in the next fragment
            continue

        m = _EMAIL_RE.search(frag)
        if m:
            email = m.group(0)
            # a shape's adjacent text runs occasionally have no space between
            # them in the source file (e.g. "...ae" + "Website" glued into
            # one run as "...aeWebsite") — the greedy domain match then eats
            # the next label too. Real emails here are consistently
            # lowercase, so trim from the first uppercase letter onward.
            trimmed = re.split(r"(?=[A-Z])", email, maxsplit=1)[0]
            if "@" in trimmed and "." in trimmed.split("@", 1)[1]:
                email = trimmed
            result.setdefault("email", email); pending = None; continue
        if _WEBSITE_RE.search(frag):
            stripped = re.sub(r"^web(site)?\s*[:.]?\s*", "", frag, flags=re.I).strip()
            result.setdefault("website", stripped or frag); pending = None; continue
        if _TRN_RE.search(frag):
            m2 = re.search(r"[\d ]{6,}", frag)
            if m2:
                result.setdefault("trn", m2.group(0).strip())
            pending = None; continue

        m_po = _PO_BOX_RE.search(frag)
        if m_po:
            result.setdefault("po_box", m_po.group(1))
            remainder = (frag[:m_po.start()] + " " + frag[m_po.end():]).strip(" ,.-")
            if remainder:
                result["address"] = (result.get("address", "") + " " + remainder).strip(" ,")
            pending = None; continue

        m_country = _UAE_COUNTRY_RE.search(frag)
        m_city = _UAE_CITY_RE.search(frag)
        if m_country or m_city:
            if m_country:
                result.setdefault("country", "AE")
            if m_city:
                result.setdefault("city", m_city.group(1).title())
            remainder = _UAE_COUNTRY_RE.sub("", _UAE_CITY_RE.sub("", frag)).strip(" ,.-")
            if remainder:
                result["address"] = (result.get("address", "") + " " + remainder).strip(" ,")
            pending = None; continue

        if _ADDRESS_RE.search(frag) or _BUILDING_WORD_RE.search(frag):
            result["address"] = (result.get("address", "") + " " + frag).strip()
            pending = None; continue
        if pending in ("phone", "landline") and _PHONE_VALUE_RE.match(frag):
            result.setdefault(pending, frag); pending = None; continue
        if pending == "fax":
            pending = None; continue
        # a bare number-only fragment with no preceding "Tel:"/"Mob:" label at
        # all (some letterheads just drop the number in on its own line) —
        # safe to assume phone rather than attn text, but only if it's
        # actually formatted like a phone number (see _BARE_PHONE_RE) — a
        # solid unbroken digit run here is more likely a TRN/reference
        # number that just didn't have the word "TRN" next to it.
        if _BARE_PHONE_RE.match(frag) and not result.get("phone"):
            result["phone"] = frag; pending = None; continue
        leading.append(frag)
        pending = None
    if leading:
        text = _strip_known_company_prefix(" ".join(leading).strip(), company_name)
        if text:
            result["attn"] = text
    return result

def extract_legacy_contact(xlsx_path, company_name=None):
    """Best-effort customer contact details (attn/phone/landline/email/
    website/trn/address/po_box/city/country) from a legacy QTN/INV/DO
    file's letterhead text boxes. Returns {} if the shapes don't resolve to
    exactly one un-claimed customer-info block (own letterhead and
    "From"/"To"/"Buyer"-style labels are excluded first) — deliberately
    gives up rather than guessing wrong when a file's layout doesn't match
    the common pattern. `company_name`, when given, is stripped from the
    front of the leftover "attn" text if it's an exact (case-insensitive)
    prefix — real letterheads often restate the company name before the
    contact person's name."""
    shapes = _contact_shape_frags(xlsx_path)
    candidates = [f for f in shapes if not _is_contact_label_shape(f) and not _is_own_letterhead_shape(f)]
    if len(candidates) != 1:
        return {}
    return _classify_contact_frags(candidates[0], company_name=company_name)

# ----------------------------------------------------------------------------
# Client-database export (Settings' "Export All Client Profiles" and the
# Clients tab's per-card "Export" button both call this — same builder, just
# a different `clients` list, so the two stay identical in layout).
# ----------------------------------------------------------------------------
# Mirrors the client-side `COUNTRIES` list in app.py (the Country picker) —
# kept in sync manually since this is a single-file-frontend app with no
# shared-data mechanism between the embedded JS and Python. Needed on this
# side for two things: a readable Country column when exporting, and (more
# importantly, now that the client database itself can live in one of these
# workbooks — see build_clients_workbook/read_clients_workbook) turning a
# human-typed country name back into the 2-letter code the Country picker
# and countryFlag() expect.
COUNTRY_NAMES = {
    "AE": "United Arab Emirates", "AF": "Afghanistan", "AL": "Albania", "DZ": "Algeria",
    "AD": "Andorra", "AO": "Angola", "AG": "Antigua and Barbuda", "AR": "Argentina",
    "AM": "Armenia", "AU": "Australia", "AT": "Austria", "AZ": "Azerbaijan",
    "BS": "Bahamas", "BH": "Bahrain", "BD": "Bangladesh", "BB": "Barbados", "BY": "Belarus",
    "BE": "Belgium", "BZ": "Belize", "BJ": "Benin", "BT": "Bhutan", "BO": "Bolivia",
    "BA": "Bosnia and Herzegovina", "BW": "Botswana", "BR": "Brazil", "BN": "Brunei",
    "BG": "Bulgaria", "BF": "Burkina Faso", "BI": "Burundi", "KH": "Cambodia",
    "CM": "Cameroon", "CA": "Canada", "CV": "Cabo Verde", "CF": "Central African Republic",
    "TD": "Chad", "CL": "Chile", "CN": "China", "CO": "Colombia", "KM": "Comoros",
    "CG": "Congo", "CD": "Congo (DRC)", "CR": "Costa Rica", "HR": "Croatia", "CU": "Cuba",
    "CY": "Cyprus", "CZ": "Czechia", "DK": "Denmark", "DJ": "Djibouti", "DM": "Dominica",
    "DO": "Dominican Republic", "EC": "Ecuador", "EG": "Egypt", "SV": "El Salvador",
    "GQ": "Equatorial Guinea", "ER": "Eritrea", "EE": "Estonia", "SZ": "Eswatini",
    "ET": "Ethiopia", "FJ": "Fiji", "FI": "Finland", "FR": "France", "GA": "Gabon",
    "GM": "Gambia", "GE": "Georgia", "DE": "Germany", "GH": "Ghana", "GR": "Greece",
    "GD": "Grenada", "GT": "Guatemala", "GN": "Guinea", "GW": "Guinea-Bissau",
    "GY": "Guyana", "HT": "Haiti", "HN": "Honduras", "HK": "Hong Kong", "HU": "Hungary",
    "IS": "Iceland", "IN": "India", "ID": "Indonesia", "IR": "Iran", "IQ": "Iraq",
    "IE": "Ireland", "IL": "Israel", "IT": "Italy", "JM": "Jamaica", "JP": "Japan",
    "JO": "Jordan", "KZ": "Kazakhstan", "KE": "Kenya", "KI": "Kiribati", "KW": "Kuwait",
    "KG": "Kyrgyzstan", "LA": "Laos", "LV": "Latvia", "LB": "Lebanon", "LS": "Lesotho",
    "LR": "Liberia", "LY": "Libya", "LI": "Liechtenstein", "LT": "Lithuania",
    "LU": "Luxembourg", "MO": "Macao", "MG": "Madagascar", "MW": "Malawi",
    "MY": "Malaysia", "MV": "Maldives", "ML": "Mali", "MT": "Malta", "MR": "Mauritania",
    "MU": "Mauritius", "MX": "Mexico", "MD": "Moldova", "MC": "Monaco", "MN": "Mongolia",
    "ME": "Montenegro", "MA": "Morocco", "MZ": "Mozambique", "MM": "Myanmar",
    "NA": "Namibia", "NP": "Nepal", "NL": "Netherlands", "NZ": "New Zealand",
    "NI": "Nicaragua", "NE": "Niger", "NG": "Nigeria", "MK": "North Macedonia",
    "NO": "Norway", "OM": "Oman", "PK": "Pakistan", "PA": "Panama",
    "PG": "Papua New Guinea", "PY": "Paraguay", "PE": "Peru", "PH": "Philippines",
    "PL": "Poland", "PT": "Portugal", "QA": "Qatar", "RO": "Romania", "RU": "Russia",
    "RW": "Rwanda", "SA": "Saudi Arabia", "SN": "Senegal", "RS": "Serbia",
    "SC": "Seychelles", "SL": "Sierra Leone", "SG": "Singapore", "SK": "Slovakia",
    "SI": "Slovenia", "SO": "Somalia", "ZA": "South Africa", "KR": "South Korea",
    "SS": "South Sudan", "ES": "Spain", "LK": "Sri Lanka", "SD": "Sudan",
    "SR": "Suriname", "SE": "Sweden", "CH": "Switzerland", "SY": "Syria",
    "TW": "Taiwan", "TJ": "Tajikistan", "TZ": "Tanzania", "TH": "Thailand",
    "TL": "Timor-Leste", "TG": "Togo", "TO": "Tonga", "TT": "Trinidad and Tobago",
    "TN": "Tunisia", "TR": "Turkey", "TM": "Turkmenistan", "UG": "Uganda",
    "UA": "Ukraine", "GB": "United Kingdom", "US": "United States", "UY": "Uruguay",
    "UZ": "Uzbekistan", "VU": "Vanuatu", "VA": "Vatican City", "VE": "Venezuela",
    "VN": "Vietnam", "YE": "Yemen", "ZM": "Zambia", "ZW": "Zimbabwe",
}
_COUNTRY_NAME_TO_CODE = {name.lower(): code for code, name in COUNTRY_NAMES.items()}

def _country_display(code):
    """Full name for the Country cell — plain, no flag: this is written
    into a real .xlsx a human may open directly, and an emoji column looks
    odd/renders inconsistently in Excel itself (flags are a web/picker-only
    nicety in this app, via countryFlag() in app.py)."""
    if not code:
        return ""
    code = code.strip().upper()
    return COUNTRY_NAMES.get(code, code)

def _country_code_from_display(text):
    """Reverse of _country_display — turns a Country cell's text back into
    a 2-letter code. Falls back to returning the raw text unchanged if it
    doesn't match a known country name (e.g. a user hand-typed something
    non-standard directly into the spreadsheet) — still usable as free
    text even though it won't resolve to a flag in the picker."""
    if not text:
        return ""
    text = text.strip()
    if len(text) == 2 and text.upper() in COUNTRY_NAMES:
        return text.upper()
    return _COUNTRY_NAME_TO_CODE.get(text.lower(), text)

# One workbook layout serves three purposes now: the Settings-tab bulk
# export, the Clients-tab per-card export, AND (see read_clients_workbook
# below) the client database's actual live storage file, when the user
# points "Clients Spreadsheet" at one — same idea as every other document
# folder in this app: the data lives outside the app as a real file the
# user can open, back up, or move, not in some internal store. Column A
# (ID) is hidden — it's how the app recognizes "this row is the same
# client as before" across saves even if the user reorders or edits other
# columns by hand; a human opening the file sees a normal-looking sheet
# starting at Company Name.
CLIENT_SHEET_HEADERS = ["ID", "Company Name", "Section", "Attn", "Address", "PO Box", "City",
                         "Country", "Phone", "Landline", "Email", "Website", "TRN", "Notes",
                         "Last Updated", "Logo"]
CLIENT_EXPORT_HEADERS = CLIENT_SHEET_HEADERS
_CLIENT_SHEET_WIDTHS = [4, 26, 16, 24, 32, 12, 16, 22, 16, 16, 26, 22, 16, 32, 12, 10]

def build_clients_workbook(clients):
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    ws.append(CLIENT_SHEET_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, c in enumerate(clients):
        row = i + 2
        ws.cell(row, 1, c.get("id") or f"c_{int(datetime.datetime.now().timestamp() * 1000)}_{i}")
        ws.cell(row, 2, c.get("name", ""))
        ws.cell(row, 3, c.get("category", ""))
        ws.cell(row, 4, c.get("attn", ""))
        ws.cell(row, 5, c.get("address", ""))
        ws.cell(row, 6, c.get("po_box", ""))
        ws.cell(row, 7, c.get("city", ""))
        ws.cell(row, 8, _country_display(c.get("country", "")))
        ws.cell(row, 9, c.get("phone", ""))
        ws.cell(row, 10, c.get("landline", ""))
        ws.cell(row, 11, c.get("email", ""))
        ws.cell(row, 12, c.get("website", ""))
        ws.cell(row, 13, c.get("trn", ""))
        ws.cell(row, 14, c.get("notes", ""))
        ws.cell(row, 15, c.get("updated", ""))
        if c.get("logo"):
            _insert_photo(ws, c["logo"], row, 16)
    ws.column_dimensions["A"].hidden = True
    for idx, w in enumerate(_CLIENT_SHEET_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    return wb

def read_clients_workbook(path):
    """Reverse of build_clients_workbook — reads the client database back
    from the .xlsx a user has pointed "Clients Spreadsheet" at. Best-effort:
    any failure to open it (locked because it's open in Excel, corrupted,
    wrong format) returns [] rather than raising, since this runs on every
    Clients-tab load and one bad file shouldn't crash the app; a row with no
    Company Name is skipped, since a human may hand-edit this file (leaving
    blank/partial rows) and this isn't the place to validate their input."""
    try:
        wb = load_workbook(path)
        ws = wb.active
    except Exception:
        return []
    images_by_row = {}
    for img in getattr(ws, "_images", []):
        try:
            row_num = img.anchor._from.row + 1
            images_by_row[row_num] = f"data:image/{img.format};base64,{base64.b64encode(img._data()).decode('ascii')}"
        except Exception:
            continue
    clients = []
    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        if not any(vals):
            continue
        def g(idx):
            v = vals[idx] if idx < len(vals) else None
            return str(v).strip() if v is not None else ""
        name = g(1)
        if not name:
            continue
        row_num = row[0].row
        clients.append({
            "id": g(0) or f"c_row{row_num}",
            "name": name,
            "category": g(2),
            "attn": g(3),
            "address": g(4),
            "po_box": g(5),
            "city": g(6),
            "country": _country_code_from_display(g(7)),
            "phone": g(8),
            "landline": g(9),
            "email": g(10),
            "website": g(11),
            "trn": g(12),
            "notes": g(13),
            "updated": g(14),
            "logo": images_by_row.get(row_num, ""),
        })
    return clients

def _guess_number(name):
    m = _NUMBER_RE.search(name)
    return m.group(1) if m else ""

def _guess_date(name, fallback_ts):
    for pat, conv in _DATE_PATTERNS:
        m = pat.search(name)
        if m:
            try:
                y, mo, d = conv(m)
                return datetime.date(y, mo, d).isoformat()
            except ValueError:
                continue
    try:
        return datetime.date.fromtimestamp(fallback_ts).isoformat()
    except (OSError, OverflowError, ValueError):
        return ""

def scan_all(folder):
    """Recursively read every document in the folder, matched or not, grouped
    by top-level subfolder (company) and sorted company A-Z, newest date first."""
    records = []
    for root, _dirs, files in os.walk(folder):
        rel_root = os.path.relpath(root, folder)
        company = rel_root.split(os.sep)[0] if rel_root != "." else ""
        for fn in files:
            if fn.startswith("~$"):
                continue  # Office's transient lock file for a currently-open document
            ext = os.path.splitext(fn)[1].lower()
            if ext not in DOC_EXT:
                continue
            path = os.path.join(root, fn)
            meta = parse_filename(fn)
            editable = bool(meta and meta["type"].upper() in FILLERS)
            if meta:
                rec = dict(meta)
                if company:
                    rec["company_label"] = company.replace("_", " ")
            else:
                try:
                    mtime = os.path.getmtime(path)
                except OSError:
                    mtime = 0
                rec = {
                    "type": _guess_type(fn), "number": _guess_number(fn), "rev": 0,
                    "company": company, "company_label": company.replace("_", " ") if company else "(root)",
                    "project": "", "project_label": os.path.splitext(fn)[0],
                    "date": _guess_date(fn, mtime),
                }
            rec["ext"] = ext.lstrip(".")
            rec["path"] = path
            rec["editable"] = editable
            rec["folder_company"] = company  # the real on-disk subfolder — distinct from
                                              # "company", which for matched files is the
                                              # client name embedded in the filename
            records.append(rec)

    # group xlsx/pdf pairs that share the same stem; the xlsx (if any) is the
    # canonical record since that's the only one that can be edited inline
    grouped = {}
    for r in records:
        stem = os.path.splitext(r["path"])[0]
        g = grouped.get(stem)
        if g is None:
            grouped[stem] = g = {**r, "formats": []}
        elif r["ext"].lower() == "xlsx":
            fmts = g["formats"]
            g.clear(); g.update(r); g["formats"] = fmts
        g["formats"].append(r["ext"].lower())

    out = list(grouped.values())
    out.sort(key=lambda r: (r.get("date", ""), r.get("number", "")), reverse=True)
    out.sort(key=lambda r: r.get("company_label", "").lower())
    return out

HEADER_CELLS = {
    "INV": {"qtn_number": "J5", "lpo_number": "J6", "type": "J8"},
    "DO": {"lpo_number": "I6"},
}
# The "To" customer/company block in the app's own bundled templates is one
# free-text merged cell (not separate name/address/PO-box/etc fields) — this
# is where the Build form's composed customer block gets written on Generate,
# and read back from on reopen.
CUSTOMER_CELL = {"INV": "C3", "DO": "F3"}
ITEM_START = {"INV": 18, "DO": 16}
ITEM_COLS = {
    "INV": [("description", 5), ("unit", 6), ("qty", 7), ("price", 8)],
    "DO": [("description", 5), ("lpo_qty", 6), ("prev_delivery", 7), ("delivered", 8)],
}

def _write_customer_block(ws, doc_type, text):
    """Writes the composed customer block into the known cell, wrapped so a
    multi-line block actually shows as multiple lines instead of one
    overflowing line clipped at the merge boundary (confirmed by rendering a
    test file — without wrap_text the text is centered and gets clipped
    equally from both ends). Skips writing entirely when blank, so an
    unfilled field doesn't blank out whatever border/placeholder design the
    template already has there. Also grows the merged block's row heights to
    fit however many lines the block actually needs — confirmed by rendering
    a real 7-line customer block into the bundled template's default 3-row-
    tall merge, where everything past line ~4 got clipped at the merge
    boundary (Excel doesn't auto-grow a merged cell's row height for you)."""
    if not (text or "").strip():
        return
    cell_ref = CUSTOMER_CELL.get(doc_type, "")
    if not cell_ref:
        return
    cell = ws[cell_ref]
    cell.value = text
    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    merge_range = next((rng for rng in ws.merged_cells.ranges
                         if rng.min_row <= cell.row <= rng.max_row and rng.min_col <= cell.column <= rng.max_col),
                        None)
    if not merge_range:
        return
    n_lines = text.count("\n") + 1
    needed_pt = n_lines * 15 + 6
    rows = range(merge_range.min_row, merge_range.max_row + 1)
    current_pt = sum(ws.row_dimensions[r].height or 15 for r in rows)
    if needed_pt > current_pt:
        extra = (needed_pt - current_pt) / len(rows)
        for r in rows:
            ws.row_dimensions[r].height = (ws.row_dimensions[r].height or 15) + extra

# ----------------------------------------------------------------------------
# Best-effort read of the "To" customer block from a legacy xlsx that draws it
# as a floating Excel shape (a textbox) rather than a cell value — very common
# in this business's real historical documents (confirmed by unzipping one
# and finding the customer's name/address/TRN in xl/drawings/drawingN.xml,
# not in any cell — openpyxl's normal cell API can't see it at all). Finds a
# shape whose own text is just a "To"/"Bill To"/"Customer" label, then
# returns the text of the nearest shape positioned at or after it (same
# column range, same row or below) — same "find the label, take what's next
# to it" idea as the PDF header reader, applied to shape geometry instead of
# line order. Best-effort/heuristic: returns "" if it can't confidently find
# a label shape at all, rather than guessing which shape is the right one.
# ----------------------------------------------------------------------------
_XLSX_CUSTOMER_LABEL_RE = re.compile(r"^(to|bill to|customer|client)\s*:?$", re.I)

def read_xlsx_customer_shape(path):
    try:
        drawing_names = [n for n in zipfile.ZipFile(path).namelist()
                          if re.match(r"xl/drawings/drawing\d+\.xml$", n)]
        if not drawing_names:
            return ""
        ns = {"xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
              "a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
        shapes = []
        with zipfile.ZipFile(path) as z:
            for name in drawing_names:
                root = ET.fromstring(z.read(name))
                for anchor in root:
                    frm = anchor.find("xdr:from", ns)
                    if frm is None:
                        continue
                    row = int(frm.find("xdr:row", ns).text)
                    col = int(frm.find("xdr:col", ns).text)
                    text = "\n".join(t.strip() for t in
                                      (el.text for el in anchor.findall(".//a:t", ns)) if t and t.strip())
                    shapes.append((row, col, text))
    except Exception:
        return ""
    label = next((s for s in shapes if _XLSX_CUSTOMER_LABEL_RE.match(s[2])), None)
    if not label:
        return ""
    lrow, lcol, _ = label
    candidates = [s for s in shapes if s is not label and lcol - 1 <= s[1] <= lcol + 3 and s[0] >= lrow - 1]
    if not candidates:
        return ""
    candidates.sort(key=lambda s: (abs(s[0] - lrow), abs(s[1] - lcol)))
    return candidates[0][2]

# ----------------------------------------------------------------------------
# Generic header-label scanner for xlsx best-effort import — mirrors the PDF
# DO reader's "find the label, take the adjacent value" approach so files
# that don't share the app's own template layout still get their header
# fields read instead of left blank. Confirmed against a real historical Tax
# Invoice that turned out to actually share the app's own cell layout almost
# exactly (just a different filename convention) — this scanner finds "QTN
# Number"/"LPO Number"/etc wherever they sit, so it works whether or not the
# layout happens to match.
# ----------------------------------------------------------------------------
_XLSX_HEADER_LABELS = {
    "INV": {"date": "date", "inv no": "number", "inv number": "number",
            "invoice no": "number", "invoice number": "number",
            "qtn no": "qtn_number", "qtn number": "qtn_number", "quotation no": "qtn_number",
            "lpo no": "lpo_number", "lpo number": "lpo_number",
            "project": "project", "type": "type"},
    "DO": {"date": "date", "do no": "number", "delivery order no": "number",
           "lpo no": "lpo_number", "lpo number": "lpo_number",
           "project": "project", "rev": "rev"},
}

def read_xlsx_header_labels(ws, label_keys, max_row=40):
    out = {}
    max_col = min(ws.max_column, 20)
    for r in range(1, min(ws.max_row, max_row) + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            key = label_keys.get(_norm_pdf_label(v))
            if not key or key in out:
                continue
            val = None
            for c2 in range(c + 1, min(c + 4, max_col + 1)):
                cv = ws.cell(r, c2).value
                if cv in (None, ""):
                    continue
                if isinstance(cv, str) and _norm_pdf_label(cv) in label_keys:
                    break  # ran into the next field's label — this one was left blank
                val = cv
                break
            if val is None:
                cv = ws.cell(r + 1, c).value
                if cv not in (None, "") and not (isinstance(cv, str) and _norm_pdf_label(cv) in label_keys):
                    val = cv
            if val is not None:
                if isinstance(val, datetime.datetime):
                    val = val.date().isoformat()
                elif isinstance(val, datetime.date):
                    val = val.isoformat()
                out[key] = str(val).strip()
    return out

def read_full_record(path, doc_type):
    """Read every header field and line item back out of a generated xlsx so the
    Build form can be fully repopulated for editing."""
    doc_type = doc_type.upper()
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    out = {}
    for key, cell in HEADER_CELLS.get(doc_type, {}).items():
        v = ws[cell].value
        out[key] = "" if v is None else str(v)
    customer_cell = CUSTOMER_CELL.get(doc_type)
    cell_block = (str(ws[customer_cell].value).strip()
                  if customer_cell and ws[customer_cell].value else "")
    out["customer_block"] = read_xlsx_customer_shape(path) or cell_block
    cols = ITEM_COLS.get(doc_type, [])
    desc_idx = next((c for k, c in cols if k == "description"), None)
    photo_col = PHOTO_COL.get(doc_type)
    images_by_row = {}
    if photo_col:
        photo_col_0 = photo_col - 1
        for img in getattr(ws, "_images", []):
            frm = getattr(getattr(img, "anchor", None), "_from", None)
            if frm and frm.col == photo_col_0:
                try:
                    images_by_row[frm.row] = img._data()
                except Exception as e:
                    print(f"read_xlsx_items: could not read embedded image at row {frm.row}: {e}")
    items = []
    if desc_idx:
        for r in range(ITEM_START.get(doc_type, 1), ws.max_row + 1):
            desc = ws.cell(r, desc_idx).value
            if desc is None or not str(desc).strip():
                break
            item = {}
            for k, c in cols:
                v = ws.cell(r, c).value
                item[k] = "" if v is None else v
            photo_bytes = images_by_row.get(r - 1)
            if photo_bytes:
                item["photo"] = "data:image/png;base64," + base64.b64encode(photo_bytes).decode("ascii")
            items.append(item)
    out["items"] = items
    return out

_ITEM_COL_LABELS = {
    "unit": "unit", "units": "unit", "uom": "unit",
    "qty": "qty", "quantity": "qty", "qnty": "qty",
    "price": "price", "rate": "price", "unit price": "price", "unit rate": "price",
    "type": "type",
}

def read_items_from_doc(path):
    """Read line items back out of a generated/legacy xlsx so they can be
    reused. Only ever extracted `description` until this was caught by the
    Submissions feature: a picked quotation showed every item's quantity
    blank, because this function never looked for Unit/Qty/Price at all —
    confirmed by inspecting a real file that turned out to use this app's
    own exact column labels (No/Type/Photo/Item Description/Unit/Qty/
    Price/Vat/Amount) despite not matching its filename convention. Now
    scans the SAME header row the description column was found on for
    cells labeled Unit/Qty/Price/Type (case-insensitive, a few label
    variants each) and reads those columns per row too — each one is
    independently optional, so a layout missing one or more of them still
    gets whatever it does have rather than nothing at all."""
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
    except Exception:
        return []
    # locate the header row containing "Item Description"
    header_row = None; desc_col = None
    for r in range(1, min(ws.max_row, 40) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "item description" in v.strip().lower():
                header_row, desc_col = r, c
                break
        if header_row:
            break
    if not header_row:
        return []
    # other item columns (Unit/Qty/Price/Type) live somewhere else in this
    # same header row — best-effort, found by their own label text rather
    # than assumed fixed positions, since layouts vary.
    extra_cols = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if not isinstance(v, str):
            continue
        key = _ITEM_COL_LABELS.get(v.strip().lower())
        if key and key not in extra_cols:
            extra_cols[key] = c
    # embedded product photos live one column left of the description column;
    # index them by their anchor row (0-indexed) so they can be matched to items
    photo_col_0 = desc_col - 2
    images_by_row = {}
    if photo_col_0 >= 0:
        for img in getattr(ws, "_images", []):
            frm = getattr(getattr(img, "anchor", None), "_from", None)
            if frm and frm.col == photo_col_0:
                try:
                    images_by_row[frm.row] = img._data()
                except Exception as e:
                    print(f"read_xlsx_items: could not read embedded image at row {frm.row}: {e}")
    items = []
    stop = {"subtotal", "vat 5%", "vat%", "total amount", "amount in words:",
            "terms & conditions", "delivered by:", "name:"}
    for r in range(header_row + 1, ws.max_row + 1):
        desc = ws.cell(row=r, column=desc_col).value
        no_cell = ws.cell(row=r, column=3).value  # col C: running No. or a label
        c_label = str(no_cell).strip().lower() if no_cell else ""
        d_label = str(desc).strip().lower() if desc else ""
        if c_label in stop or d_label in stop:
            break
        if desc and str(desc).strip():
            item = {"description": str(desc).strip()}
            for key, col in extra_cols.items():
                v = ws.cell(row=r, column=col).value
                if v not in (None, ""):
                    item[key] = v
            photo_bytes = images_by_row.get(r - 1)
            if photo_bytes:
                item["photo"] = "data:image/png;base64," + base64.b64encode(photo_bytes).decode("ascii")
            items.append(item)
    return items

# ----------------------------------------------------------------------------
# Best-effort read of a Delivery Order that only exists as a PDF (no xlsx
# twin). Common specifically for DOs — after physical delivery/signature the
# business often only keeps the printed/scanned PDF. This only works when the
# PDF has a real, selectable text layer (i.e. it was printed straight from
# Excel, not scanned as a flattened image) — a signed/scanned copy has no
# text at all and this correctly returns no items rather than guessing via
# OCR, same "best-effort, never guess" philosophy as the datasheet/legacy-
# xlsx importers elsewhere in this file. Sampled 6 real DO PDFs across
# different companies/years before writing this — the "Date:/DO No:/
# Project:/LPO no:/[Rev:]" label block followed by "SN No. / Photo / Item
# Description / LPO Quantity / Previous Delivery" and sequential 1/2/3…
# row-number lines was consistent across all of them.
# ----------------------------------------------------------------------------
_PDF_DO_HEADER_LABELS = {
    "date": "date", "do no": "number", "project": "project",
    "lpo no": "lpo_number", "rev": "rev",
}
_PDF_ROW_NUM_RE = re.compile(r"^\d{1,3}$")
_PDF_QTY_UNIT_RE = re.compile(r"^(\d+(?:\.\d+)?)\s*([A-Za-z]+)$")
_PDF_ITEM_TRAILERS = ("delivered by:",)
# NOT "delivery order": on multi-page DOs the printed letterhead/title text
# ("DELIVERY ORDER" + company block) repeats after the item table on EVERY
# page (an artifact of the template's page layout, not a true end-of-items
# marker) — treating it as a stop marker silently dropped every item after
# page 1. "Delivered By:" only ever appears once, on the true last page.

def _norm_pdf_label(s):
    """'Rev:', 'Rev', 'REV :' all normalize the same — real DO templates
    vary on whether header labels carry a trailing colon."""
    return s.strip().lower().rstrip(":").strip()

def _read_pdf_label_block(lines, label_keys, stop_at):
    """Generic 'Label:' / 'value' reader for a PDF's text layer. A value can
    span multiple lines (some templates wrap long project names across 2-3
    lines) — everything non-empty up to the next known label or the stop
    marker is joined with a space. A label whose very next line is itself
    another known label (or the stop marker) means that field was left
    blank on the original document — correctly omits it instead of
    accidentally swallowing the next label's own text as the value."""
    out = {}
    n = len(lines)
    i = 0
    while i < n:
        low = _norm_pdf_label(lines[i])
        if low == stop_at:
            break
        key = label_keys.get(low)
        if not key:
            i += 1
            continue
        j = i + 1
        parts = []
        while j < n:
            nxt = lines[j].strip()
            if not nxt or _norm_pdf_label(nxt) in label_keys or _norm_pdf_label(nxt) == stop_at:
                break
            parts.append(nxt)
            j += 1
        if parts:
            out[key] = " ".join(parts)
        i = j
    return out

def _parse_ddmmyyyy(s):
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", (s or "").strip())
    if not m:
        return ""
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return datetime.date(y, mo, d).isoformat()
    except ValueError:
        return ""

def read_do_pdf(path):
    """Returns {"header": {...}, "items": [...]}, or {} if the PDF has no
    text layer at all (a scanned/signed copy) or doesn't match the expected
    shape."""
    try:
        import fitz
        doc = fitz.open(path)
    except Exception:
        return {}
    lines = []
    for page in doc:
        lines.extend(l for l in (s.strip() for s in page.get_text().split("\n")))
    header = _read_pdf_label_block(lines, _PDF_DO_HEADER_LABELS, "sn no.")
    start = None
    for i, l in enumerate(lines):
        if l.lower() == "item description":
            start = i + 1
            break
    if start is None:
        return {"header": header, "items": []}
    end = len(lines)
    for i in range(start, len(lines)):
        if lines[i].lower() in _PDF_ITEM_TRAILERS:
            end = i
            break
    body = [l for l in lines[start:end] if l]
    rows, current, expect = [], None, 1
    for l in body:
        if _PDF_ROW_NUM_RE.match(l) and int(l) == expect:
            if current is not None:
                rows.append(current)
            current, expect = [], expect + 1
            continue
        if current is not None:
            current.append(l)
        # else: leftover column-header text before row "1" — discarded
    if current is not None:
        rows.append(current)
    items = []
    for row in rows:
        qty_idx = next((i for i in range(len(row) - 1, -1, -1) if _PDF_QTY_UNIT_RE.match(row[i])), None)
        if qty_idx is not None:
            m = _PDF_QTY_UNIT_RE.match(row[qty_idx])
            qty, unit, desc_lines = m.group(1), m.group(2).upper(), row[:qty_idx]
        else:
            qty, unit, desc_lines = "", "", row
        desc = "\n".join(desc_lines).strip()
        if not desc:
            continue
        items.append({"description": desc, "unit": unit, "lpo_qty": qty})
    return {"header": header, "items": items}

# ----------------------------------------------------------------------------
# Style helpers for inserting extra item rows
# ----------------------------------------------------------------------------
def _clone_style(src, dst):
    if src.has_style:
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy.copy(src.protection)

def _expand_rows(ws, model_row, insert_at, n_extra):
    """Insert n_extra rows at insert_at, shifting merged cells below and copying
    the style of model_row. openpyxl does NOT move merges on insert, so we do it."""
    if n_extra <= 0:
        return
    from openpyxl.utils import get_column_letter
    # snapshot + unmerge every merge that starts at or below the insertion point
    to_shift = [str(m) for m in list(ws.merged_cells.ranges) if m.min_row >= insert_at]
    for rng in to_shift:
        ws.unmerge_cells(rng)
    ws.insert_rows(insert_at, n_extra)
    # re-merge shifted down by n_extra
    for rng in to_shift:
        from openpyxl.utils.cell import range_boundaries
        c1, r1, c2, r2 = range_boundaries(rng)
        ws.merge_cells(start_row=r1 + n_extra, start_column=c1,
                       end_row=r2 + n_extra, end_column=c2)
    # style + height for the new rows
    h = ws.row_dimensions[model_row].height
    for i in range(n_extra):
        tgt = insert_at + i
        if h:
            ws.row_dimensions[tgt].height = h
        for col in range(1, ws.max_column + 1):
            _clone_style(ws.cell(row=model_row, column=col), ws.cell(row=tgt, column=col))

def _remove_rows(ws, remove_at, n_remove):
    """Delete n_remove rows starting at remove_at, shifting merged cells below
    up to close the gap (mirrors _expand_rows). Merges that overlapped the
    removed rows are simply dropped; merges entirely below shift up."""
    if n_remove <= 0:
        return
    from openpyxl.utils.cell import range_boundaries
    remove_end = remove_at + n_remove  # exclusive
    to_shift = [str(m) for m in list(ws.merged_cells.ranges) if m.min_row >= remove_at]
    for rng in to_shift:
        ws.unmerge_cells(rng)
    ws.delete_rows(remove_at, n_remove)
    for rng in to_shift:
        c1, r1, c2, r2 = range_boundaries(rng)
        if r1 >= remove_end:  # entirely below the removed block -> shift up
            ws.merge_cells(start_row=r1 - n_remove, start_column=c1,
                           end_row=r2 - n_remove, end_column=c2)
        # else: overlapped the removed rows -> dropped, nothing to re-merge

# ----------------------------------------------------------------------------
# Product photos — each template has a "Photo" column just left of the Item
# Description column (confirmed against the real templates).
# ----------------------------------------------------------------------------
PHOTO_COL = {"INV": 4, "DO": 4}  # 1-indexed column letter position

# ----------------------------------------------------------------------------
# Discount / VAT summary block (INV only — DO has no price column at all).
# The template reserves a Subtotal/Vat/Total block with the label merged
# C{row}:{merge_end}{row} and the value in a single cell; the summary block
# grows/shrinks (2-5 rows) depending on which of Discount/VAT are on.
# ----------------------------------------------------------------------------
SUMMARY_VALUE_COL = {"INV": 10}     # J
SUMMARY_MERGE_END_COL = {"INV": 9}  # I
VAT_COL = {"INV": 9}                # per-line "Vat 5%" column
VAT_HEADER_CELL = {"INV": (17, 9)}

# ----------------------------------------------------------------------------
# Product photo catalog — matches a line item's description text against a
# folder of product PNGs (named after the product, e.g. STAGNA.png) so photos
# can be attached automatically instead of uploaded by hand every time.
# ----------------------------------------------------------------------------
_PHOTO_CATALOG_CACHE = {"folder": None, "index": {}}

def _build_photo_catalog(folder):
    index = {}
    for root, _dirs, files in os.walk(folder):
        for fn in files:
            if fn.lower().endswith(".png"):
                name = os.path.splitext(fn)[0].strip().upper()
                if name:
                    index[name] = os.path.join(root, fn)
    return index

def load_photo_catalog(folder, force=False):
    """Cached folder scan — rebuilt only when the folder changes or force=True."""
    if force or _PHOTO_CATALOG_CACHE["folder"] != folder:
        _PHOTO_CATALOG_CACHE["index"] = _build_photo_catalog(folder) if folder and os.path.isdir(folder) else {}
        _PHOTO_CATALOG_CACHE["folder"] = folder
    return _PHOTO_CATALOG_CACHE["index"]

def match_product_photo(description, folder):
    """Find the best-matching product PNG for a line item's description text.
    Longer/more specific product names (e.g. STAGNA-MS) are tried before
    shorter ones (STAGNA) so the more precise variant wins."""
    index = load_photo_catalog(folder)
    if not index or not description:
        return None
    text = description.upper()
    for name in sorted(index.keys(), key=len, reverse=True):
        if re.search(r"(?<![A-Z0-9])" + re.escape(name) + r"(?![A-Z0-9])", text):
            try:
                with open(index[name], "rb") as f:
                    data = f.read()
                return "data:image/png;base64," + base64.b64encode(data).decode("ascii")
            except OSError:
                continue
    return None

# ----------------------------------------------------------------------------
# Product datasheets — a product can have more than one PDF (e.g. STAGNA.pdf
# and "STAGNA with MOTION SENSOR ON.OFF.pdf"), so datasheets are grouped by
# their leading product-name word and every match for that product is
# returned, unlike photos where only the single best match is used.
# ----------------------------------------------------------------------------
_DATASHEET_CATALOG_CACHE = {"folder": None, "index": {}}

def _build_datasheet_catalog(folder):
    index = {}
    seen_stems = {}  # (dirpath, stem) -> True, so a .pdf and .ai of the same
                      # product in the same folder don't both get listed
    for root, _dirs, files in os.walk(folder):
        # process .pdf before .ai so a same-name .ai is reliably deduped
        # against its .pdf sibling regardless of filesystem listing order
        for fn in sorted(files, key=lambda f: 0 if f.lower().endswith(".pdf") else 1):
            ext = os.path.splitext(fn)[1].lower()
            if ext not in (".pdf", ".ai"):
                continue
            stem = os.path.splitext(fn)[0].strip()
            if not stem:
                continue
            dedupe_key = (root, stem.upper())
            if ext == ".ai" and dedupe_key in seen_stems:
                continue  # a .pdf sibling already covers this same product
            seen_stems[dedupe_key] = True
            key = re.split(r"[\s_-]+", stem)[0].upper()
            index.setdefault(key, []).append({"name": stem, "path": os.path.join(root, fn)})
    return index

def load_datasheet_catalog(folder, force=False):
    if force or _DATASHEET_CATALOG_CACHE["folder"] != folder:
        _DATASHEET_CATALOG_CACHE["index"] = _build_datasheet_catalog(folder) if folder and os.path.isdir(folder) else {}
        _DATASHEET_CATALOG_CACHE["folder"] = folder
    return _DATASHEET_CATALOG_CACHE["index"]

def list_datasheet_products(folder):
    """Flat, name-sorted list of every distinct product in the datasheet
    catalog — for the Build tab's "Find Product" finder (browse/search to
    pick one deliberately, rather than only auto-matching whatever text
    happens to already be typed in the description)."""
    index = load_datasheet_catalog(folder)
    seen, out = set(), []
    for entries in index.values():
        for e in entries:
            if e["path"] in seen:
                continue
            seen.add(e["path"])
            out.append(e)
    out.sort(key=lambda e: e["name"].lower())
    return out

def match_datasheets(description, folder):
    """Return every datasheet whose product key appears in the description —
    can be more than one file for the same product."""
    index = load_datasheet_catalog(folder)
    if not index or not description:
        return []
    text = description.upper()
    matches, seen = [], set()
    for key in sorted(index.keys(), key=len, reverse=True):
        if re.search(r"(?<![A-Z0-9])" + re.escape(key) + r"(?![A-Z0-9])", text):
            for entry in index[key]:
                if entry["path"] not in seen:
                    seen.add(entry["path"])
                    matches.append(entry)
    return matches

# ----------------------------------------------------------------------------
# Product option extraction (the "Product Builder" — pick Wattage/CCT/Beam
# Angle/Controls/Color for a line item straight from its matched datasheet,
# instead of typing a code from memory or re-reading the PDF by hand).
#
# The catalog's datasheets all follow the same real-world convention (sampled
# ~5 across different product families/categories to confirm this before
# writing any regex): a "Technical Specifications" block with labeled option
# groups (Wattage/CCT/Beam Angle/Controls/Finish Color/IP Rating), and an
# "Ordering Code Example: SLAQU-5W-D2-30-12-22-ND-WH" line that gives the
# product's model-code prefix directly. PyMuPDF's plain-text extraction turns
# each table cell into its own line (no column structure survives), which is
# actually convenient here — a spec table's "code / value" pairs end up as
# two adjacent lines almost every time, which is what every regex below
# anchors on.
#
# This is heuristic, not a real PDF-table parser: some datasheets extract
# cleanly, a few don't (one sampled file had a garbled character-by-character
# first section, likely an embedded-font quirk in that specific source PDF) —
# whatever a given regex doesn't find for a given product is just omitted
# from the result rather than guessed at, same "best-effort, user reviews"
# philosophy as extract_legacy_contact.
# ----------------------------------------------------------------------------
_ORDER_CODE_RE = re.compile(r"Ordering Code Example:?\s*([A-Za-z0-9]+(?:-[A-Za-z0-9]+)+)", re.I)
_WATTAGE_LIST_RE = re.compile(r"Wattage\s*\n\s*([\d\s/*+xX]+?)\s*W\b", re.I)
_CCT_PAIR_RE = re.compile(r"^(\d{2,3})\s*$\n^\s*(\d{3,5})\s*K\b", re.M)
_BEAM_PAIR_RE = re.compile(r"^(\d{1,3})\s*$\n^\s*(\d{1,3})\s*[°˚]", re.M)
_CONTROL_PAIR_RE = re.compile(r"^(ND|D\d{1,2}|DA|DI)\s*$\n^(.*(?:dimming|dali|non-dim).*)$", re.I | re.M)
_IP_RE = re.compile(r"IP\s*Rating\D{0,15}IP\s*(\d{2})", re.I)
_COLOR_WORDS = ("white", "black", "grey", "gray", "gold", "silver", "bronze",
                "ivory", "champagne", "chrome", "brass", "beige", "sandy")
_COLOR_PAIR_RE = re.compile(r"^([A-Z]{2})\s*$\n^(" + "|".join(_COLOR_WORDS) + r")\b", re.I | re.M)
_SIZE_RE = re.compile(r"\b(\d{2,4}\s*[*x]\s*\d{2,4}(?:\s*[*x]\s*\d{2,4})?\s*mm)\b", re.I)

def _dedupe_keep_order(seq):
    seen, out = set(), []
    for x in seq:
        if x not in seen:
            seen.add(x); out.append(x)
    return out

def extract_product_options(datasheet_path):
    """Best-effort parse of a datasheet PDF's spec/ordering-code tables into
    pickable option lists for the Product Builder. Returns {} if the file
    can't be opened at all; otherwise returns whatever categories were
    found (never guesses a category that isn't there)."""
    try:
        import fitz
        doc = fitz.open(datasheet_path)
        text = "\n".join(page.get_text() for page in doc)
        doc.close()
    except Exception:
        return {}

    result = {}

    m = _ORDER_CODE_RE.search(text)
    order_code_example = m.group(1) if m else ""
    if order_code_example:
        result["order_code_example"] = order_code_example
        result["model"] = order_code_example.split("-")[0]
    else:
        result["model"] = os.path.splitext(os.path.basename(datasheet_path))[0].strip().upper()

    m = _WATTAGE_LIST_RE.search(text)
    if m:
        watts = re.split(r"[/*+xX]", m.group(1))
        result["wattage"] = _dedupe_keep_order(f"{w.strip()}W" for w in watts if w.strip())

    cct = [(code, f"{val}K") for code, val in _CCT_PAIR_RE.findall(text)]
    if cct:
        result["cct"] = _dedupe_keep_order(cct)

    beam = [(code, f"{val}°") for code, val in _BEAM_PAIR_RE.findall(text)]
    if beam:
        result["beam_angle"] = _dedupe_keep_order(beam)

    controls = [(code.upper(), label.strip()) for code, label in _CONTROL_PAIR_RE.findall(text)]
    if controls:
        result["controls"] = _dedupe_keep_order(controls)

    colors = [(code.upper(), label.strip().title()) for code, label in _COLOR_PAIR_RE.findall(text)]
    if colors:
        result["color"] = _dedupe_keep_order(colors)

    m = _IP_RE.search(text)
    if m:
        result["ip_rating"] = f"IP{m.group(1)}"

    sizes = _SIZE_RE.findall(text)
    if sizes:
        result["size"] = _dedupe_keep_order(s.replace(" ", "").upper() for s in sizes)

    return result

def _insert_photo(ws, data_url, row, col):
    """Decode a data-URL image and anchor it into a cell. Failures are
    swallowed — a bad/corrupt photo shouldn't break the whole document."""
    try:
        _header, b64data = data_url.split(",", 1)
        raw = base64.b64decode(b64data)
        img = XLImage(BytesIO(raw))
        max_dim = 60
        if img.width and img.height:
            scale = min(max_dim / img.width, max_dim / img.height, 1)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
        needed_pt = img.height / 1.333 + 4
        if not ws.row_dimensions[row].height or ws.row_dimensions[row].height < needed_pt:
            ws.row_dimensions[row].height = needed_pt
        ws.add_image(img, f"{get_column_letter(col)}{row}")
    except Exception as e:
        print(f"_insert_photo: could not embed photo at row {row}: {e}")

# ----------------------------------------------------------------------------
# Template fillers
# ----------------------------------------------------------------------------
def _fmt_date(date_iso, style="dd.mm.yyyy"):
    d = datetime.date.fromisoformat(date_iso)
    return d.strftime("%d.%m.%Y") if style == "dd.mm.yyyy" else d.isoformat()

def _lit(v):
    """Format a number for embedding into an Excel formula literal."""
    try:
        f = float(v)
    except (TypeError, ValueError):
        f = 0.0
    return f"{f:.6f}".rstrip("0").rstrip(".") or "0"

def _norm_discount(d):
    d = d or {}
    mode = d.get("mode") if d.get("mode") in ("percent", "fixed", "target") else "percent"
    try:
        value = float(d.get("value") or 0)
    except (TypeError, ValueError):
        value = 0.0
    return {"enabled": bool(d.get("enabled")), "mode": mode, "value": value}

def _norm_vat(d):
    if d is None:
        return {"enabled": True, "mode": "percent", "value": 5.0}
    mode = d.get("mode") if d.get("mode") in ("percent", "fixed") else "percent"
    try:
        value = float(d.get("value") or 0)
    except (TypeError, ValueError):
        value = 0.0
    return {"enabled": bool(d.get("enabled")), "mode": mode, "value": value}

def _write_summary_block(ws, doc_type, sub_row, discount, vat):
    """Writes Discount / Price-After-Discount / Vat / Total below an
    already-written Subtotal row, growing/shrinking the block (the template
    ships with exactly 3 rows: Subtotal/Vat/Total) to fit whichever of
    Discount/VAT are enabled. Returns the row used for each present role."""
    value_col = SUMMARY_VALUE_COL[doc_type]
    merge_end = SUMMARY_MERGE_END_COL[doc_type]
    vcol = get_column_letter(value_col)

    roles = ["subtotal"]
    if discount["enabled"]:
        roles += ["discount", "price_after_discount"]
    if vat["enabled"]:
        roles.append("vat")
    roles.append("total")

    extra = len(roles) - 3
    if extra > 0:
        _expand_rows(ws, model_row=sub_row, insert_at=sub_row + 1, n_extra=extra)
    elif extra < 0:
        _remove_rows(ws, remove_at=sub_row + 1, n_remove=-extra)

    row_of = {role: sub_row + i for i, role in enumerate(roles)}

    # every non-subtotal row needs its label merge re-created — _expand_rows
    # copies style but not merges, and _remove_rows may have dropped one
    existing = {str(m) for m in ws.merged_cells.ranges}
    for role, r in row_of.items():
        if role == "subtotal":
            continue
        rng = f"C{r}:{get_column_letter(merge_end)}{r}"
        if rng not in existing:
            ws.merge_cells(rng)

    if "discount" in row_of:
        r = row_of["discount"]
        ws.cell(r, 3, "Discount")
        if discount["mode"] == "percent":
            ws.cell(r, value_col, f"={vcol}{sub_row}*({_lit(discount['value'])}/100)")
        elif discount["mode"] == "target":
            ws.cell(r, value_col, f"=MAX({vcol}{sub_row}-{_lit(discount['value'])},0)")
        else:  # fixed
            ws.cell(r, value_col, f"={_lit(discount['value'])}")

    if "price_after_discount" in row_of:
        r = row_of["price_after_discount"]
        ws.cell(r, 3, "Price After Discount")
        ws.cell(r, value_col, f"={vcol}{sub_row}-{vcol}{row_of['discount']}")

    pre_vat_row = row_of.get("price_after_discount", sub_row)
    if "vat" in row_of:
        r = row_of["vat"]
        if vat["mode"] == "percent":
            label = f"Vat {vat['value']:g}%"
            ws.cell(r, value_col, f"={vcol}{pre_vat_row}*({_lit(vat['value'])}/100)")
        else:
            label = "Vat"
            ws.cell(r, value_col, f"={_lit(vat['value'])}")
        ws.cell(r, 3, label)

    total_r = row_of["total"]
    ws.cell(total_r, 3, "TOTAL AMOUNT")
    if "vat" in row_of:
        ws.cell(total_r, value_col, f"={vcol}{pre_vat_row}+{vcol}{row_of['vat']}")
    else:
        ws.cell(total_r, value_col, f"={vcol}{pre_vat_row}")

    return row_of

def fill_invoice(data, out_path, brand=None):
    """data keys: number, date, qtn_number, lpo_number, project, type, company, items, discount, vat, ..."""
    wb = load_workbook(template_path(brand, "INV"))
    ws = wb.active
    ws["J3"] = _fmt_date(data["date"]); ws["J3"].number_format = "@"
    ws["J4"] = str(data.get("number", ""))
    ws["J5"] = str(data.get("qtn_number", ""))
    ws["J6"] = str(data.get("lpo_number", ""))
    ws["J7"] = data.get("project", "")
    ws["J8"] = data.get("type", "")
    _write_customer_block(ws, "INV", data.get("customer_block", ""))

    discount = _norm_discount(data.get("discount"))
    vat = _norm_vat(data.get("vat"))
    line_vat_pct = vat["value"] if (vat["enabled"] and vat["mode"] == "percent") else None
    hr, hc = VAT_HEADER_CELL["INV"]
    ws.cell(hr, hc, f"Vat {line_vat_pct:g}%" if line_vat_pct is not None else ("Vat" if vat["enabled"] else ""))

    items = data.get("items", []) or []
    first, n_template = 18, 1
    if len(items) > n_template:
        _expand_rows(ws, model_row=first, insert_at=first + n_template, n_extra=len(items) - n_template)
    last = first + max(len(items), 1) - 1

    for i, it in enumerate(items):
        r = first + i
        ws.cell(r, 3, i + 1)                       # C No
        desc_cell = ws.cell(r, 5, it.get("description", ""))   # E Item Description
        desc_cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(r, 6, it.get("unit", "PCS"))       # F Unit
        ws.cell(r, 7, _num(it.get("qty")))         # G Qty
        ws.cell(r, 8, _num(it.get("price")))       # H Price
        ws.cell(r, 10, f"=G{r}*H{r}")              # J Amount
        ws.cell(r, 9, f"=J{r}*({_lit(line_vat_pct)}/100)" if line_vat_pct is not None else None)  # I Vat
        if it.get("photo"):
            _insert_photo(ws, it["photo"], r, PHOTO_COL["INV"])

    shift = max(len(items), n_template) - n_template
    sub = 19 + shift
    ws.cell(sub, 10, f"=SUM(J{first}:J{last})")
    _write_summary_block(ws, "INV", sub, discount, vat)
    wb.save(out_path)
    return out_path

def fill_delivery_order(data, out_path, brand=None):
    """data keys: number, date, project, lpo_number, company, items[{description,unit,lpo_qty,prev_delivery,delivered}]"""
    wb = load_workbook(template_path(brand, "DO"))
    ws = wb.active
    ws["I3"] = _fmt_date(data["date"]); ws["I3"].number_format = "@"
    ws["I4"] = str(data.get("number", ""))
    ws["I5"] = data.get("project", "")
    ws["I6"] = str(data.get("lpo_number", ""))
    _write_customer_block(ws, "DO", data.get("customer_block", ""))

    items = data.get("items", []) or []
    first, n_template = 16, 1
    if len(items) > n_template:
        _expand_rows(ws, model_row=first, insert_at=first + n_template, n_extra=len(items) - n_template)

    for i, it in enumerate(items):
        r = first + i
        ws.cell(r, 3, i + 1)                                            # C SN No.
        desc_cell = ws.cell(r, 5, it.get("description", ""))            # E Item Description
        desc_cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(r, 6, f" {it.get('unit','PCS')}" if it.get('qty') in (None,'') else f"{_num(it.get('qty'))} {it.get('unit','PCS')}")
        ws.cell(r, 7, _num(it.get("lpo_qty")))                          # F? actually F=LPO Quantity (col 6) -> fix below
    # NOTE: DO columns: C SN, D Photo, E Desc, F LPO Quantity, G Previous Delivery, H Delivered
    for i, it in enumerate(items):
        r = first + i
        ws.cell(r, 6, _num(it.get("lpo_qty")))        # F LPO Quantity
        ws.cell(r, 7, _num(it.get("prev_delivery")))  # G Previous Delivery
        ws.cell(r, 8, _num(it.get("delivered")))      # H Delivered
        if it.get("photo"):
            _insert_photo(ws, it["photo"], r, PHOTO_COL["DO"])
    wb.save(out_path)
    return out_path

def _num(v):
    if v in (None, ""):
        return None
    try:
        f = float(v)
        return int(f) if f == int(f) else f
    except (TypeError, ValueError):
        return v

FILLERS = {"INV": fill_invoice, "DO": fill_delivery_order}

# ----------------------------------------------------------------------------
# PDF conversion
# ----------------------------------------------------------------------------
_SOFFICE_FALLBACKS = [
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    "/usr/bin/soffice", "/usr/bin/libreoffice",
]

def _find_soffice():
    return (shutil.which("soffice") or shutil.which("libreoffice")
            or next((p for p in _SOFFICE_FALLBACKS if os.path.exists(p)), None))

def to_pdf(xlsx_path, out_dir=None):
    """Convert an xlsx to PDF using LibreOffice headless. Returns the pdf path.
    Retries once after a short delay on failure — LibreOffice headless mode
    holds a per-user-profile lock for a moment after each invocation exits,
    so firing a second conversion immediately after the first (e.g.
    Submissions generating a DO and an Invoice back-to-back) can fail with a
    plain non-zero exit and no useful stderr. Caught by exactly this
    scenario in testing: the very same conversion that failed once succeeded
    immediately when retried by hand a few seconds later."""
    out_dir = out_dir or os.path.dirname(xlsx_path)
    soffice = _find_soffice()
    if not soffice:
        raise RuntimeError("LibreOffice not found. Install it to enable PDF export.")
    pdf = os.path.join(out_dir, os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf")
    last_err = None
    for attempt in range(2):
        try:
            subprocess.run([soffice, "--headless", "--convert-to", "pdf", "--outdir",
                            out_dir, xlsx_path], check=True,
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=120)
            if os.path.exists(pdf):
                return pdf
        except subprocess.CalledProcessError as e:
            last_err = e
        if attempt == 0:
            time.sleep(2)
    if last_err:
        raise RuntimeError(f"PDF conversion failed: {last_err}")
    raise RuntimeError("PDF conversion failed.")

def to_png_pages(pdf_path, out_dir, prefix, dpi=170):
    """Render every page of a PDF to <out_dir>/<prefix>_<n>.png (1-indexed, in
    order). Used for the multi-page live preview — LibreOffice's own png
    export only ever produces page 1, so this rasterizes the (already
    LibreOffice-generated) PDF directly via PyMuPDF instead. Returns the
    number of pages rendered."""
    import fitz
    doc = fitz.open(pdf_path)
    try:
        zoom = dpi / 72
        mat = fitz.Matrix(zoom, zoom)
        for i in range(doc.page_count):
            pix = doc[i].get_pixmap(matrix=mat)
            pix.save(os.path.join(out_dir, f"{prefix}_{i + 1}.png"))
        return doc.page_count
    finally:
        doc.close()

def _hex_to_unit_rgb(hexstr):
    """'#rrggbb' -> (r,g,b) as 0..1 floats, the format PyMuPDF's draw_rect wants."""
    hexstr = hexstr.lstrip("#")
    return tuple(int(hexstr[i:i + 2], 16) / 255 for i in (0, 2, 4))

# Must match sololuce_datasheet.html's own <doc-page margin="..."> — see
# stamp_catalogue_page_numbers' own use of this below. 0.5in = 36pt exactly.
_PAGE_MARGIN_PT = 0.5 * 72

def stamp_catalogue_page_numbers(pdf_path, start_page, tab_color=None, total_pages=None):
    """Stamps a running catalog page number onto the bottom-right of every
    physical page of an already-rendered PDF, starting at start_page and
    incrementing by 1 per page. A Sololuce Datasheet is one of many
    separately-generated PDFs eventually bound into one physical multi-
    hundred-page catalog, so its printed page number(s) need to continue on
    from wherever the previous datasheet left off rather than restarting at
    1 — something neither Chromium's own print-to-PDF page numbering nor
    CSS Paged Media counters can do (both always start at 1 for whatever
    single PDF is currently being produced), so this stamps plain text
    directly onto the rendered pages afterward instead, the same
    already-proven approach `to_png_pages` above uses for rasterizing.
    Written to a temp file and atomically swapped in, rather than an
    incremental save, to avoid PyMuPDF's incremental-save edge cases.

    If tab_color ('#rrggbb', from the Series/Category's auto-assigned color —
    see app.py's _assign_series_color) and total_pages (the eventual bound-
    catalog size, set once in Settings) are both given, also paints a solid
    color index tab flush against the right edge of every page, positioned
    top-to-bottom in proportion to that page's absolute position in the
    catalog — the "thumb index" convention printed dictionaries/manuals use,
    so flipping through the bound catalog gets you to the right color band
    fast. Sits flush against the true page edge while the page number stops
    _PAGE_MARGIN_PT short of it, so the two never collide horizontally even
    though both live in the bottom-right corner. Silently skipped
    if either value is missing (no Series/Category on this sheet, or Total
    Catalog Pages hasn't been set yet) rather than guessing at a position.

    Returns the PDF's page count, so the caller can advance whatever
    counter start_page came from by exactly that amount."""
    import fitz
    rgb = _hex_to_unit_rgb(tab_color) if tab_color else None
    doc = fitz.open(pdf_path)
    try:
        page_count = doc.page_count
        for i in range(page_count):
            page = doc[i]
            # Height matters here, not just cosmetically — insert_textbox
            # silently drops the text (no exception) and returns a negative
            # number if the box is too short for even one line at this font
            # size; 22pt of vertical room is comfortably enough for 9pt text
            # while a tighter box (confirmed by testing) was not.
            # Right-aligned against the content's own right edge (not
            # centered across the full page) so it lands in the footer's
            # right column, which sololuce_datasheet.html's own footer
            # deliberately leaves empty for exactly this — the number and
            # the ©/website line share the same right margin this way.
            # _PAGE_MARGIN_PT must track that template's own <doc-page
            # margin="..."> value (currently 7.2mm) for the two to actually
            # line up.
            # Vertical position — height-22.5 to height-0.5 (was height-30
            # to height-8) — tracks that same template's footer having been
            # moved 1.6% of the page height lower (margin-bottom:-13.62pt
            # on sololuce_datasheet.html's [slot=footer]) per explicit
            # request. The two 22pt-tall boxes (this one unchanged in
            # HEIGHT, just repositioned) sit back-to-back with only a
            # 0.5pt gap and a 0.5pt margin off the true page edge — that
            # tight fit is a hard ceiling, not a stylistic choice: the
            # footer's own request was originally for a full 2%/16.84pt
            # move, which was calculated to leave only 13.45pt of room
            # here, short of the 22pt this box needs, so the footer's own
            # shift was itself capped to make this box's fit possible.
            # Whenever either the footer's own CSS shift or this box's own
            # 22pt height requirement changes, re-derive both together —
            # they're coupled, not independent.
            rect = fitz.Rect(page.rect.width / 2, page.rect.height - 22.5,
                              page.rect.width - _PAGE_MARGIN_PT, page.rect.height - 0.5)
            fit = page.insert_textbox(rect, str(start_page + i), fontsize=9, fontname="helv",
                                       color=(0.137, 0.122, 0.125), align=fitz.TEXT_ALIGN_RIGHT)
            if fit < 0:
                raise RuntimeError(f"Page number text did not fit on page {i + 1} (insert_textbox returned {fit})")
            if rgb and total_pages and total_pages > 0:
                tab_w, tab_h = 12, 26
                # Bounds are on the tab's own top/bottom EDGES (center +/-
                # tab_h/2), not the center point itself — clamping the center
                # alone left the tab's bottom edge overlapping the page-number
                # band above by tab_h/2 whenever a page landed at or near the
                # very end of the catalog (frac close to 1.0).
                center_min = 20 + tab_h / 2
                center_max = (page.rect.height - 40) - tab_h / 2
                frac = min(max((start_page + i) / total_pages, 0.0), 1.0)
                center = center_min + frac * (center_max - center_min)
                tab_rect = fitz.Rect(page.rect.width - tab_w, center - tab_h / 2,
                                      page.rect.width, center + tab_h / 2)
                page.draw_rect(tab_rect, color=rgb, fill=rgb)
        tmp_path = pdf_path + ".tmp"
        doc.save(tmp_path)
    finally:
        doc.close()
    os.replace(tmp_path, pdf_path)
    return page_count

def to_png(src_path, out_dir=None):
    """Render the first page/sheet of any LibreOffice-openable document (xlsx,
    pdf, doc, docx, ...) as a PNG, for in-app preview thumbnails. Returns the
    png path.

    .pdf sources render directly via PyMuPDF (fitz) instead of shelling out
    to LibreOffice — LibreOffice's own PDF import doesn't reliably honor
    certain clip+transform constructs Chromium's print-to-PDF output uses,
    and every photo slot in sololuce_datasheet.html is exactly that (a
    photo masked via CSS overflow:hidden around a CSS scale()+translate()'d
    <img>): confirmed directly against a real generated CAT datasheet — its
    Application Photo rendered correctly, fully clipped to its own column,
    via fitz's own get_pixmap(), while the *same file* run through
    LibreOffice's PNG export instead showed that same photo bleeding most
    of the way across the page, well past the page's own left edge,
    ignoring the CSS clip entirely (LibreOffice's PDF import doesn't
    reliably apply this specific clip+transform combination the way a
    spec-compliant PDF renderer does). fitz is already this app's own PDF
    engine everywhere else that matters (to_png_pages, the Full Catalog
    Builder's page measurement, stamp_catalogue_page_numbers, ...) and
    skips LibreOffice's whole process-spawn/convert round-trip, so this is
    also simply faster for what's by far the common case here — a
    document's own PDF, not one of its xlsx-only siblings, which still
    need LibreOffice below since fitz can't open those directly. 170dpi
    matches to_png_pages' own default, so a document's live multi-page
    preview and its All Docs/vendor-datasheet thumbnail share one
    resolution convention rather than two arbitrary ones."""
    out_dir = out_dir or os.path.dirname(src_path)
    if src_path.lower().endswith(".pdf"):
        import fitz
        os.makedirs(out_dir, exist_ok=True)
        png = os.path.join(out_dir, os.path.splitext(os.path.basename(src_path))[0] + ".png")
        doc = fitz.open(src_path)
        try:
            if doc.page_count == 0:
                raise RuntimeError("PDF has no pages.")
            pix = doc[0].get_pixmap(matrix=fitz.Matrix(170 / 72, 170 / 72))
            pix.save(png)
        finally:
            doc.close()
        return png
    soffice = _find_soffice()
    if not soffice:
        raise RuntimeError("LibreOffice not found. Install it to enable previews.")
    subprocess.run([soffice, "--headless", "--convert-to", "png", "--outdir",
                    out_dir, src_path], check=True,
                   stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=120)
    png = os.path.join(out_dir, os.path.splitext(os.path.basename(src_path))[0] + ".png")
    if not os.path.exists(png):
        raise RuntimeError("Preview image conversion failed.")
    return png

# ----------------------------------------------------------------------------
# One-call generate: writes BOTH xlsx and pdf with the convention name
# ----------------------------------------------------------------------------
def generate(doc_type, data, dest_folder, brand=None, make_pdf=True):
    doc_type = doc_type.upper()
    brand = (brand or DEFAULT_BRAND).upper()
    os.makedirs(dest_folder, exist_ok=True)
    stem = build_filename(doc_type, data["number"], data.get("rev", 0),
                          data["company"], data.get("project", ""), data["date"], brand=brand, ext="")[:-1]
    xlsx_path = os.path.join(dest_folder, stem + ".xlsx")
    FILLERS[doc_type](data, xlsx_path, brand=brand)
    result = {"xlsx": xlsx_path}
    # Markdown companion (see save_markdown's own docstring) — centralized
    # here so every caller of generate() (the main Build form, and the
    # Quotation-approval/submittal pipeline's own direct DO+INV generation)
    # gets one for free, rather than every call site remembering to add it.
    save_markdown(xlsx_path, doc_type, data, brand=brand)
    if make_pdf:
        pdf_path = os.path.join(dest_folder, stem + ".pdf")
        if doc_type in HTML_PDF_DOC_TYPES:
            # Deferred import: html_engine imports this module (for
            # _fmt_date etc.), so importing it back at module level here
            # would be circular. By the time generate() actually runs both
            # modules have already finished loading, so a local import
            # resolves cleanly and costs nothing extra beyond the first
            # call (Python caches the module).
            import html_engine
            html_engine.RENDERERS[doc_type](data, pdf_path, brand=brand)
        else:
            pdf_tmp = to_pdf(xlsx_path, dest_folder)
            if pdf_tmp != pdf_path and os.path.exists(pdf_tmp):
                shutil.move(pdf_tmp, pdf_path)
        result["pdf"] = pdf_path
    return result

def build_submittal_pdf(parts, out_path):
    """Merge the Quotation, LPO, scanned Delivery Order, and Invoice — any
    mix of PDF and image (jpg/png) files — into one combined PDF, the
    literal package handed to a client/consultant. `fitz.Document.insert_pdf`
    only accepts genuine PDF sources (confirmed by trying it directly on an
    image file first — raises "source or target not a PDF"), so an image
    part gets its own new page sized to the image instead of being inserted
    as pages. Skips any part that's missing/blank rather than failing the
    whole merge — a submittal with 3 of 4 parts is still useful; refusing
    to build it at all over one missing file would not be."""
    import fitz
    out = fitz.open()
    for p in parts:
        if not p or not os.path.exists(p):
            continue
        ext = os.path.splitext(p)[1].lower()
        try:
            if ext == ".pdf":
                with fitz.open(p) as src:
                    out.insert_pdf(src)
            else:
                img_doc = fitz.open(p)
                rect = img_doc[0].rect
                img_doc.close()
                page = out.new_page(width=rect.width, height=rect.height)
                page.insert_image(rect, filename=p)
        except Exception:
            continue
    if out.page_count == 0:
        out.close()
        return None
    out.save(out_path)
    out.close()
    return out_path

"""
Office Tool — local app
Run:  python app.py   then open the browser link it prints.
Because it runs on your own PC it can read your documents folder and save
Excel + PDF files there. The on-screen preview is the generated PDF.
"""
import os, io, re, json, glob, shutil, datetime, traceback, base64, mimetypes, uuid, colorsys, random, threading, queue, time
from flask import Flask, request, jsonify, send_file, Response, session
from openpyxl import load_workbook
import engine
import html_engine
import pdf_extract
import catalog_builder
import update_checker
import accounts
import photo_store
import scanner
from version import APP_VERSION

# static_folder is explicit (not Flask's own __name__-relative default) so
# it still finds the bundled logos/fonts when frozen into a PyInstaller
# .exe, whose __file__/root_path point somewhere temporary — see engine.BASE.
app = Flask(__name__, static_folder=os.path.join(engine.BASE, "static"))
CONFIG = os.path.join(engine.DATA_BASE, "config.json")

# A random secret_key would invalidate every session (force re-login) on
# every single app restart, which is needlessly annoying for daily use —
# persist it once per install instead, same DATA_BASE as everything else
# this app writes at runtime (see engine.py's BASE/DATA_BASE split).
_SECRET_KEY_PATH = os.path.join(engine.DATA_BASE, "secret_key.txt")
try:
    with open(_SECRET_KEY_PATH, "r", encoding="utf-8") as _f:
        app.secret_key = _f.read().strip()
    if not app.secret_key:
        raise ValueError
except Exception:
    app.secret_key = os.urandom(32).hex()
    try:
        with open(_SECRET_KEY_PATH, "w", encoding="utf-8") as _f:
            _f.write(app.secret_key)
    except Exception:
        pass  # worst case: sessions won't survive a restart, not fatal

# "Remember me for 30 days" checkbox on login — see /api/login's own
# comment on session.permanent for how this pairs with it.
app.config["PERMANENT_SESSION_LIFETIME"] = datetime.timedelta(days=30)

# Endpoints reachable with no session at all: the SPA shell itself (so the
# login screen has something to render into), static assets, login itself,
# and update-checking (harmless to expose, and the update banner should
# still work pre-login).
_PUBLIC_PATHS = {"/", "/api/login", "/api/current-user", "/api/check-update"}
# Tool-block enforcement: which URL prefixes are hard-blocked server-side
# for each blockable view (see accounts.py's BLOCKABLE_TOOLS). This list is
# deliberately narrow — audited call-site by call-site — because this
# app's API is NOT cleanly separated by screen: e.g. /api/config,
# /api/index, /api/clients (GET/POST), /api/submissions (POST), and
# /api/units all get called during ordinary document editing (auto-filling
# the next doc number, company/project autocomplete, the inline "save as
# client" button, auto-creating a submission record when a DO/INV
# generates, the line-item unit picker) — blocking those would break core
# document generation for a restricted user, not just hide a screen. Only
# endpoints confirmed exclusive to one view's own bulk/delete/export
# actions are blocked here; "settings" isn't hard-blocked at all for the
# same reason (its endpoints are shared with inline saves elsewhere, e.g.
# Full Catalog Builder's own output-folder field uses /api/settings too).
# So: brand_lock and this list are real security boundaries; every
# blocked_tools entry (including "settings") ALSO hides its rail nav
# button in the UI (see applyAccessRestrictions()) — for "settings" that
# UI hide is the only restriction, a convenience rather than a hard wall.
_TOOL_PREFIXES = {
    "clients": ("/api/clients-import", "/api/clients-delete", "/api/clients-export"),
    "submissions": ("/api/submissions-delete", "/api/submissions-link-scanned-do", "/api/submissions-build-submittal",
                     "/submission-lpo", "/submission-submittal", "/api/browse-scanned-do", "/open-scanned-do"),
    "statement": ("/api/finance",),
    "alldocs": ("/api/alldocs-clone", "/api/alldocs-delete", "/api/alldocs-move"),
}

@app.before_request
def _require_login():
    path = request.path
    if path in _PUBLIC_PATHS or path.startswith("/static/"):
        return
    user = session.get("user")
    if not user:
        return jsonify({"error": "Not logged in."}), 401
    if path.startswith("/api/accounts") and session.get("role") != "admin":
        return jsonify({"error": "Admin access required."}), 403
    # Cloudflare/R2 settings are admin-only to even SEE, not just change
    # (explicit request) — non-admins get zero fields, silently using the
    # bundled read-only key instead (see photo_store.py's two-tier
    # comment). /api/photostore-list and -fetch stay open to everyone —
    # read-only browsing (the cloud photo gallery picker — see
    # openCloudPhotoPicker()) is a different thing from managing the
    # connection itself.
    _PHOTOSTORE_ADMIN_ONLY = ("/api/photostore-upload", "/api/photostore-delete", "/api/photostore-config")
    if any(path.startswith(p) for p in _PHOTOSTORE_ADMIN_ONLY) and session.get("role") != "admin":
        return jsonify({"error": "Admin access required."}), 403
    for tool, prefixes in _TOOL_PREFIXES.items():
        if tool in (session.get("blocked_tools") or []) and any(path.startswith(p) for p in prefixes):
            return jsonify({"error": "You don't have access to this."}), 403

@app.post("/api/login")
def api_login():
    data = request.json or {}
    u = accounts.verify_login(data.get("username", ""), data.get("password", ""))
    if not u:
        return jsonify({"ok": False, "error": "Wrong username or password."}), 401
    session.clear()
    session["user"] = u["username"]
    session["role"] = u["role"]
    session["brand_lock"] = u["brand_lock"]
    session["blocked_tools"] = u["blocked_tools"]
    # "Remember me" checkbox (defaults checked): permanent=True gives the
    # cookie a real expiry (PERMANENT_SESSION_LIFETIME, set below to 30
    # days) so it survives closing/reopening the app; unchecked makes it a
    # plain session cookie most browsers/WebView2 drop once the window closes.
    session.permanent = bool(data.get("remember", True))
    # A brand-locked user always operates in their locked brand, regardless
    # of whatever brand this install was last left on.
    if u["brand_lock"]:
        cfg = load_cfg(); cfg["brand"] = u["brand_lock"]; save_cfg(cfg)
    return jsonify({"ok": True, **u})

@app.post("/api/logout")
def api_logout():
    session.clear()
    return jsonify({"ok": True})

@app.get("/api/current-user")
def api_current_user():
    if not session.get("user"):
        return jsonify({"logged_in": False})
    return jsonify({"logged_in": True, "username": session["user"], "role": session["role"],
                     "brand_lock": session.get("brand_lock"), "blocked_tools": session.get("blocked_tools") or []})

# ---- Admin-only user management (Settings > Users) — see accounts.py ----
@app.get("/api/accounts")
def api_accounts_list():
    return jsonify({"users": accounts.list_users(), "brands": list(engine.BRANDS.keys()),
                     "blockable_tools": list(accounts.BLOCKABLE_TOOLS)})

@app.post("/api/accounts-save")
def api_accounts_save():
    data = request.json or {}
    try:
        u = accounts.upsert_user(data.get("username", ""), data.get("role", "user"),
                                  data.get("brand_lock") or None, data.get("blocked_tools") or [],
                                  password=data.get("password") or None)
        return jsonify({"ok": True, "user": u})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400

@app.post("/api/accounts-delete")
def api_accounts_delete():
    data = request.json or {}
    if data.get("username", "").strip().lower() == session.get("user", "").strip().lower():
        return jsonify({"ok": False, "error": "You can't delete your own account while logged in."}), 400
    try:
        accounts.delete_user(data.get("username", ""))
        return jsonify({"ok": True})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400

@app.post("/api/accounts-publish")
def api_accounts_publish():
    return jsonify(accounts.publish_to_cloud())

SETTINGS_FIELDS = ("inv_folder", "do_folder", "qtn2_folder", "pi_folder", "rv_folder", "cn_folder",
                    "scanned_do_folder", "product_photos_folder", "datasheets_folder", "templates_folder",
                    "clients_file", "catalogue_folder", "expense_folder", "full_catalog_folder")
FILE_SETTINGS_FIELDS = {"clients_file"}  # picks a file, not a folder (native dialog + validation differ)

def _empty_brand_settings():
    return {k: "" for k in SETTINGS_FIELDS}

def load_cfg():
    try:
        with open(CONFIG, encoding="utf-8") as f: c = json.load(f)
    except Exception:
        c = {}
    # html_engine.py (the PDF-rendering side) reads config.json's badge
    # library straight off disk rather than through this function, since it
    # can't import app.py — so a freshly-seeded default that only exists
    # in-memory here would render as "no badges" until *something* happened
    # to POST-save the file first (e.g. adding a custom badge). Persist once,
    # immediately, the first time this key is missing, so the seeded library
    # is always actually on disk for html_engine to find.
    needs_save = "cat_badge_library" not in c
    c.setdefault("brand", engine.DEFAULT_BRAND)
    c.setdefault("units", ["PCS", "MTR", "PAIRS", "ROLLS"])
    # Standard Technical Specifications row labels for a brand-new Sololuce
    # Datasheet — matches the real catalogue's own convention. Grows over
    # time: any custom label the user types gets remembered here too (see
    # /api/cat-spec-labels-add), same "remember what was typed" pattern as
    # the Unit dropdown's own custom-entry list above.
    c.setdefault("cat_spec_labels", [
        "Power", "Lifespan", "Light Source", "Luminare Efficacy", "Power Factor",
        "Ambient Temperature", "Body Material", "Diffuser", "Mounting Type", "IP Rating", "Driver",
    ])
    # "Wattage" was the original label; renamed to "Power" per user request.
    # Existing config.json already has the old label persisted, so rename it
    # in place (idempotent — only touches an exact "Wattage" entry, leaves
    # any other custom label the user typed untouched).
    if "Wattage" in c["cat_spec_labels"]:
        c["cat_spec_labels"] = ["Power" if l == "Wattage" else l for l in c["cat_spec_labels"]]
    # Series/Category dropdown — seeded with only the two confirmed-real
    # values (from AURA ECO and AGATA) rather than guessing the rest of
    # Sololuce's naming convention from folder names (which don't map 1:1 to
    # the marketing "Series" label anyway, e.g. AGATA lives under the
    # "PENDANT LIGHTS" folder but its Series is "Lamp Suspensions", not
    # "Pendant Lights"). Grows the same way cat_spec_labels does — see
    # /api/cat-series-add.
    c.setdefault("cat_series_labels", ["Downlight Series", "Lamp Suspensions"])
    # Auto-generated (never hand-picked) color per Series/Category value, for
    # the printed catalog's colored index tab — see _assign_series_color and
    # engine.stamp_catalogue_page_numbers. Backfilled here so the two seeded
    # labels above always have one, same reasoning as cat_badge_library's own
    # needs_save backfill: html_engine.py reads config.json straight off disk
    # and can't call into app.py to generate one on demand.
    c.setdefault("cat_series_colors", {})
    for _lbl in c["cat_series_labels"]:
        if _lbl not in c["cat_series_colors"]:
            _assign_series_color(c, _lbl)
            needs_save = True
    # Outdoor/Indoor/Striplight tag per category — purely a display grouping
    # for the Category Order screen (see /api/cat-series-section); a
    # product's real section always comes from its own Product Type field,
    # this changes nothing about the actual build. Absent = shows in that
    # screen's own "Unassigned" group until the user tags it.
    c.setdefault("cat_series_sections", {})
    # Family Tree — remembered family names, grows the same "type it once,
    # reuse it after" way cat_series_labels does (see /api/cat-family-add).
    # No color needed (unlike Series/Category): a family shows up in the
    # printed catalog as its own illustrated divider page, not a colored tab.
    c.setdefault("cat_family_labels", [])
    # Ordering Table "Fill Standard Information" button — which of the
    # known dropdown-backed field types it fills with every one of their
    # own remembered preset options (one option per row), one click instead
    # of manually adding a row per CCT/Controls/etc value. User-editable
    # (see /api/cat-standard-fill-fields) — "programmable" per explicit
    # request, not hardcoded to just CCT/Controls forever even though
    # that's the default set.
    c.setdefault("cat_standard_fill_fields", ["cct", "controls"])
    # Display/fill order for ALL 9 possible fields (not just the selected
    # ones) — separate from cat_standard_fill_fields above, which is only
    # ever the SELECTED subset. Lets the Configure popover's checklist (and
    # therefore the order fillCatOrdStandardInfo() processes them in) be
    # rearranged via /api/cat-standard-fill-order-move, same swap-with-
    # neighbor pattern as /api/cat-series-move. Defaults to
    # CAT_STANDARD_FILL_KEYS' own built-in order.
    c.setdefault("cat_standard_fill_order", list(CAT_STANDARD_FILL_KEYS))
    # Per-field SUBSET + sequence of preset VALUES actually used when Fill
    # Standard Information runs — one level deeper than cat_standard_fill_
    # order above (which only controls which FIELDS fill and in what
    # sequence). A field with no key here means "use its full current
    # preset list," exactly the original behavior — this dict only ever
    # holds fields the user has deliberately customized via /api/cat-
    # standard-fill-values(-move/-reset), so nothing regresses for anyone
    # who never touches this.
    c.setdefault("cat_standard_fill_values", {})
    # Standard Ordering Table column widths, keyed by column LABEL ("Power",
    # "Model No.", ...) so the same field gets the same width across every
    # future datasheet, not just the one it was set on. A label with no
    # entry here falls through to the existing content-based auto-sizing
    # (see build_ordering_table, html_engine.py) — this dict only ever
    # holds labels the user explicitly saved via "Save as Standard" in the
    # Column Widths widget, so nothing regresses for anyone who never
    # touches this. A column's own per-datasheet manual drag (c.width)
    # still wins over this when both are present — this is the fallback
    # BASELINE new/reset columns start from, not an override of a deliberate
    # per-datasheet adjustment.
    c.setdefault("cat_ordering_default_widths", {})
    # Full Catalog Builder settings — combines every generated Sololuce
    # Datasheet into one bound book (Index, per-category Pre-index, the
    # datasheets themselves grouped/sorted) plus whatever front/back matter
    # the user uploads (see catalog_extras below) — this app never invents
    # that content itself, only combines + paginates + tabs it. Section
    # order is Outdoor/Indoor/Striplight by default but user-reorderable;
    # category order within each section reuses cat_series_labels's own
    # array order rather than a second parallel list that could disagree
    # with it.
    c.setdefault("catalog_section_order", ["Outdoor", "Indoor", "Striplight"])
    # Display label + optional color per section — the underlying values
    # (Outdoor/Indoor/Striplight) stay fixed (Product Type dropdown, Class
    # 1/2/3 badge auto-assignment, and grouping all key off them literally),
    # but what actually PRINTS in the book is fully renamable. "Strip Light
    # & Neon Flex" replaces the old "Striplight / Neon Flex" wording per
    # explicit request — IP20 vs IP65+ variants within it stay one unified
    # section/range, distinguished by each product's own IP Rating badge
    # rather than being split across Indoor/Outdoor.
    c.setdefault("catalog_section_labels", {"Outdoor": "Outdoor", "Indoor": "Indoor", "Striplight": "Strip Light & Neon Flex"})
    c.setdefault("catalog_section_colors", {})
    # Superseded by catalog_extras (uploaded Cover/Introduction/Ending/
    # per-family-divider files) — these older auto-composed-content fields
    # have no upload-file equivalent to migrate their content into, so
    # they're just dropped rather than carried forward.
    for _old_key in ("catalog_intro_title", "catalog_intro_body", "catalog_intro_blocks",
                      "catalog_ending_title", "catalog_ending_body"):
        if _old_key in c:
            del c[_old_key]
            needs_save = True
    # Cover is its own single fixed slot (catalog_extras["cover"]), always
    # the FIRST page — like Ending, not reorderable, since a cover can't
    # meaningfully be dragged to a non-first position. Introduction and any
    # custom pages added via "+ Add Custom Page" live together in one
    # reorderable "front matter" list (catalog_extras["front_matter"]),
    # inserted verbatim in list order right after Cover (see
    # catalog_builder.py's build_full_catalog Phase B). "Introduction" is
    # the list's one permanent builtin row (id "introduction" — never
    # removable, only reorderable/uploadable/file-clearable); custom rows
    # have a free-typed label, a generated id, and can be deleted outright.
    # Ending stays its own single fixed slot too (catalog_extras["ending"])
    # — always the last page. family_dividers is {family_name: {filename,
    # stored_as}} for the same per-family. Any item/slot with no stored_as
    # means "nothing uploaded" — contributes zero pages, not a placeholder.
    # Actual files live in catalog_builder.CATALOG_EXTRAS_DIR, named by a
    # generated uuid (stored_as), never by user input — see
    # _save_catalog_extra_pdf.
    extras = c.setdefault("catalog_extras", {})
    extras.setdefault("family_dividers", {})
    fm = extras.setdefault("front_matter", [])
    cover_in_list = next((it for it in fm if it.get("id") == "cover"), None)
    if cover_in_list is not None:
        # migrate Cover back out of the reorderable list (an earlier config
        # shape had it as a draggable builtin row there) into its own fixed
        # slot, carrying over any uploaded file
        fm[:] = [it for it in fm if it.get("id") != "cover"]
        if cover_in_list.get("stored_as") and not extras.get("cover", {}).get("stored_as"):
            extras["cover"] = {"filename": cover_in_list.get("filename"), "stored_as": cover_in_list.get("stored_as")}
        needs_save = True
    extras.setdefault("cover", {})
    if not any(it.get("id") == "introduction" for it in fm):
        fm.append({"id": "introduction", "label": "Introduction", "builtin": True})
        needs_save = True
    c.setdefault("catalog_last_build", {})
    # Index Order — per-category product order override + a global
    # exclude list for the photo-grid Index (see catalog_builder.py's
    # compute_index_rows). A product missing from
    # catalog_index_order[category] just sorts alphabetically after
    # whatever IS explicitly ordered there — never persisted until the
    # user actually reorders something in that category, same lazy
    # pattern cat_series_labels-style lists use elsewhere. Excluding a
    # product hides it from the Index grid only — it still gets a real
    # stamped page in the book, same as before.
    c.setdefault("catalog_index_order", {})
    c.setdefault("catalog_index_excluded", [])
    # Ordering/Variant Table standard column categories — matches the real
    # catalogue's own convention (Model No., Power, Size, CCT, ...). "Size"
    # has two variants since round fittings are specified by diameter*height
    # and rectangular ones by length*width*height — kept as two separate
    # pickable entries rather than one field with an internal toggle, so
    # picking either is just "pick this instead of that" like every other
    # category. Grows the same way cat_spec_labels/cat_series_labels do: any
    # custom column name the user types is remembered here too (see
    # /api/cat-ordering-categories-add).
    c.setdefault("cat_ordering_categories", [
        "Model No.", "Power", "Size (Ø×H)", "Size (L×W×H)", "Cut Out", "CCT", "Beam Angle",
        "Input Voltage", "Controls", "Finish Options", "Options", "Lumen",
    ])
    if "Cut Out" not in c["cat_ordering_categories"]:
        c["cat_ordering_categories"].append("Cut Out")
    # Standard CCT (color temperature) presets for the Ordering Table's CCT
    # column dropdown — any custom value the user types gets appended here
    # too (see /api/cat-cct-options-add), same "remember what was typed"
    # pattern as cat_spec_labels/cat_ordering_categories above.
    c.setdefault("cat_cct_options", ["2700", "3000", "3500", "4000", "5000", "6000", "6500"])
    # Standard dimming/control protocols for the Ordering Table's Controls
    # column dropdown — any custom value the user types gets appended here
    # too (see /api/cat-controls-options-add), same "remember what was
    # typed" pattern as cat_cct_options/cat_ordering_categories above.
    c.setdefault("cat_controls_options", ["DALI", "0-10V", "Phase Dim", "Non-Dimmable"])
    # Standard input-voltage presets for the Ordering Table's Input Voltage
    # column dropdown — covers common low-voltage DC driver/module inputs
    # and AC mains ranges so nothing typical is left out by default. Any
    # custom value the user types gets appended here too (see
    # /api/cat-voltage-options-add), same "remember what was typed" pattern
    # as cat_cct_options/cat_controls_options above.
    c.setdefault("cat_voltage_options", ["12V DC", "24V DC", "48V DC", "100-240V", "110-120V", "220-240V", "380-415V"])
    # Wattage presets for the Ordering Table's Power column dropdown — every
    # whole watt from 1-40W, per explicit request (not just a curated
    # spread), so any value in that range is a clean dropdown pick rather
    # than falling to Custom. Any value outside 1-40W the user types gets
    # appended here too (see /api/cat-power-options-add), same "remember
    # what was typed" pattern as cat_cct_options/cat_voltage_options above.
    c.setdefault("cat_power_options", [f"{n}W" for n in range(1, 41)])
    # Standard beam-angle presets for the Ordering Table's Beam Angle column
    # dropdown — the common commercial spread from narrow spot to wide
    # flood. Any custom value the user types gets appended here too (see
    # /api/cat-beamangle-options-add), same "remember what was typed"
    # pattern as cat_cct_options/cat_power_options above.
    c.setdefault("cat_beamangle_options", ["15°", "24°", "36°", "38°", "45°", "60°", "90°", "100°", "120°"])
    # Catalogue-wide Size preset list for the Ordering Table's Size column
    # suggestion dropdown — plain size text only, growing as new sizes are
    # typed (see /api/cat-size-index-add). NOT where a size's printed "D{n}"
    # code comes from — that's local to each datasheet's own Size column
    # (see app.py's front-end recomputeCatOrdSizeDNumbers), always starting
    # at D1 for whichever size that one product uses first, regardless of
    # this list's own order or of any D-number the same size text happens
    # to carry on some other product's sheet.
    c.setdefault("cat_size_index", [])
    # Model No./Cut Out/Options presets for the Ordering Table's remaining
    # free-text columns — no curated seed data exists for these (unlike
    # Beam Angle/CCT), so they simply start empty and grow from whatever the
    # user actually types (see /api/cat-modelno-options-add etc.).
    c.setdefault("cat_modelno_options", [])
    c.setdefault("cat_cutout_options", [])
    c.setdefault("cat_options_options", [])
    # Payment Method presets for the Expense Report line-items table — any
    # custom value the user types gets appended here too (see
    # /api/expense-payment-methods-add), same "remember what was typed"
    # pattern as the CAT preset lists above.
    c.setdefault("expense_payment_methods", ["CASH", "CARD", "BANK TRANSFER"])
    # Employee, Category, Product, and Description presets for the Expense
    # Report — same "remember what was typed" pattern as expense_payment_methods
    # above (see /api/expense-employees-add etc.). Category/Product/Description
    # seeded from the real reference report's own values (see
    # project_expense_report_tool memory) plus a few complementary ones in the
    # same theme so the dropdown isn't just the 3-4 examples the user gave.
    c.setdefault("expense_employees", ["Edgar Kagramanyan", "Suraj Mathews", "Lea Galleato"])
    c.setdefault("expense_categories", ["PETROL", "SALIK", "MATERIAL", "REPAIR", "MAINTENANCE",
                                         "PARKING", "OFFICE SUPPLIES", "TRAVEL", "ACCOMMODATION", "MISCELLANEOUS"])
    c.setdefault("expense_products", ["Petrol", "Salik", "Material", "Repair"])
    c.setdefault("expense_descriptions", ["VEHICLE 36533", "CCT CHANGE", "PURCHASE FOR COMPANY"])
    # Per-label memory for Technical Specification VALUES (distinct from
    # cat_spec_labels, which remembers the label text itself) — keyed by
    # normalized label (see _norm_label_key) since "Body Material" and
    # "Driver" need separate remembered pools, not one shared list. Light
    # Source/Power Factor used to be a hardcoded, non-persistent curated
    # dropdown (CAT_SPEC_DROPDOWNS in the frontend) — migrated into this
    # same generic system so their old curated options survive and any
    # custom value the user types from now on is remembered too.
    c.setdefault("cat_spec_values", {})
    c["cat_spec_values"].setdefault("lightsource", ["COB", "SMD"])
    c["cat_spec_values"].setdefault("powerfactor", [">0.9"])
    # IP ratings matching the badge library's own IP badges (IP20/44/54/
    # 65/66/67/68), so picking a spec value and getting the matching
    # auto-selected badge always line up.
    c["cat_spec_values"].setdefault("iprating", ["IP20", "IP44", "IP54", "IP65", "IP66", "IP67", "IP68"])
    c["cat_spec_values"].setdefault("bodymaterial", ["Die Cast", "Aluminium", "GRP", "Polycarbonate", "Stainless Steel"])
    c["cat_spec_values"].setdefault("diffuser", ["Polycarbonate", "Tempered Glass", "Acrylic", "Aluminium", "PMMA Opal"])
    c["cat_spec_values"].setdefault("mountingtype", ["Surface Mounted", "Recessed", "Suspended", "Wall Mounted", "Track Mounted", "Pole Mounted", "Ground Recessed", "Pendant"])
    c["cat_spec_values"].setdefault("ambienttemperature", ["± 5 °C", "± 10 °C", "± 20 °C", "± 25 °C"])
    c["cat_spec_values"].setdefault("luminareefficacy", ["100 lm/W", "110 lm/W", "120 lm/W", "130 lm/W", "150 lm/W"])
    c["cat_spec_values"].setdefault("lifespan", ["30,000 Hrs", "50,000 Hrs", "70,000 Hrs"])
    # Standard Finish Color palette — seeded from the real catalogue's own
    # AGATA options (Finish Color Options: Black/White/Mustard Yellow/Mint
    # Blue/Rose) plus "Grey". Only Black/White/Grey show directly on the DSB
    # form (CAT_FINISH_QUICK in the frontend) — every other saved color,
    # including these presets and any custom/RAL color the user adds, lives
    # in the "More Colors" popup instead of cluttering the main form. Any
    # brand-new custom color the user adds gets remembered here too (see
    # /api/cat-finish-colors-add), same "grows over time" pattern as
    # cat_spec_labels/units above.
    c.setdefault("cat_finish_colors", [
        {"label": "Black", "hex": "#000000"},
        {"label": "White", "hex": "#ffffff"},
        {"label": "Grey", "hex": "#808080"},
        {"label": "Mustard Yellow", "hex": "#c9a227"},
        {"label": "Mint Blue", "hex": "#8fbfd6"},
        {"label": "Rose", "hex": "#e6a8b0"},
    ])
    # Safety net for configs saved before "Grey" was added to the seed list
    # above — setdefault() only fires when the key is missing entirely, so
    # an existing config.json with the old 5-color list would never pick up
    # Grey otherwise, and Grey must exist for the DSB's fixed 3-color quick
    # row (Black/White/Grey) to have something real to render.
    if not any(c2.get("label") == "Grey" for c2 in c["cat_finish_colors"]):
        c["cat_finish_colors"].append({"label": "Grey", "hex": "#808080"})
    # RAL labels used to be saved as "RAL 1234 – Name" (with an en-dash) —
    # dropped the dash for a cleaner look, so strip it from anything already
    # saved under the old format too, not just new additions going forward.
    for c2 in c["cat_finish_colors"]:
        c2["label"] = re.sub(r"^(RAL\s*\d{4})\s*–\s*", r"\1 ", c2.get("label", ""))
    # Spec-badge image library — the real 39 icon PNGs Sololuce actually uses
    # (copied read-only from F:\...\4. SOLOLUCE\1. CATALOGUE\Edgar\4. Back
    # Pages\ICONS into static/cat_badges/), each fully self-contained (icon +
    # caption + border baked in by the original designer — no per-product
    # value overlay needed, unlike the app's first hand-drawn-icon attempt).
    # Grows the same way cat_spec_labels/cat_finish_colors do: any badge the
    # user uploads via "+ Add Custom Badge" (see /api/cat-badges-add) is
    # appended here permanently, indistinguishable from the original 39.
    c.setdefault("cat_badge_library", [
        {"key": "3-color-access", "label": "3 Color access", "filename": "3-color-access.png"},
        {"key": "adjustable-beam-angle", "label": "Adjustable Beam Angle", "filename": "adjustable-beam-angle.png"},
        {"key": "adjustable-orientation", "label": "Adjustable Orientation", "filename": "adjustable-orientation.png"},
        {"key": "adjustable-rotation", "label": "Adjustable Rotation", "filename": "adjustable-rotation.png"},
        {"key": "class-1-fixture", "label": "Class 1 Fixture", "filename": "class-1-fixture.png"},
        {"key": "class-2-fixture", "label": "Class 2 Fixture", "filename": "class-2-fixture.png"},
        {"key": "class-3-fixture", "label": "Class 3 Fixture", "filename": "class-3-fixture.png"},
        {"key": "client-supplied-materials-supply-and-installation", "label": "Client-supplied materials, supply and installation", "filename": "client-supplied-materials-supply-and-installation.png"},
        {"key": "connected-devices-communicating-via-internet", "label": "Connected devices communicating via internet", "filename": "connected-devices-communicating-via-internet.png"},
        {"key": "coverage-against-manufacturing-defects", "label": "Coverage against manufacturing defects", "filename": "coverage-against-manufacturing-defects.png"},
        {"key": "dali-dimmable-system", "label": "DALI Dimmable System", "filename": "dali-dimmable-system.png"},
        {"key": "designed-for-use-inside-buildings-only", "label": "Designed for use inside buildings only", "filename": "designed-for-use-inside-buildings-only.png"},
        {"key": "designed-for-use-outside-buildings-only", "label": "Designed for use outside buildings only", "filename": "designed-for-use-outside-buildings-only.png"},
        {"key": "dispose-product-through-special-recycling", "label": "Dispose product through special recycling", "filename": "dispose-product-through-special-recycling.png"},
        {"key": "dust-protected-water-splash-resistant", "label": "Dust protected, water splash resistant", "filename": "dust-protected-water-splash-resistant.png"},
        {"key": "dust-tight-protected-against-jets", "label": "Dust tight, protected against jets", "filename": "dust-tight-protected-against-jets.png"},
        {"key": "dust-tight-strong-water-jets", "label": "Dust tight, strong water jets", "filename": "dust-tight-strong-water-jets.png"},
        {"key": "emergency-lighting-module-for-power-failure", "label": "Emergency lighting module for power failure", "filename": "emergency-lighting-module-for-power-failure.png"},
        {"key": "european-conformity-safety-marking", "label": "European conformity safety marking", "filename": "european-conformity-safety-marking.png"},
        {"key": "excellent-color-rendering-very-natural-light", "label": "Excellent color rendering, very natural light", "filename": "excellent-color-rendering-very-natural-light.png"},
        {"key": "fire-rated", "label": "Fire Rated", "filename": "fire-rated.png"},
        {"key": "general-performance-or-summary-rating", "label": "General performance or summary rating", "filename": "general-performance-or-summary-rating.png"},
        {"key": "limits-hazardous-substances-in-electronics", "label": "Limits hazardous substances in electronics", "filename": "limits-hazardous-substances-in-electronics.png"},
        {"key": "no-direct-eye-exposure", "label": "No Direct Eye Exposure", "filename": "no-direct-eye-exposure.png"},
        {"key": "office-standard", "label": "Office standard", "filename": "office-standard.png"},
        {"key": "power-and-data-via-ethernet", "label": "Power and data via Ethernet", "filename": "power-and-data-via-ethernet.png"},
        {"key": "protected-against-splashing-water-ingress", "label": "Protected against splashing water ingress", "filename": "protected-against-splashing-water-ingress.png"},
        {"key": "protected-against-vertically-falling-water", "label": "Protected against vertically falling water", "filename": "protected-against-vertically-falling-water.png"},
        {"key": "restricts-hazardous-substances-in-electronics", "label": "Restricts hazardous substances in electronics", "filename": "restricts-hazardous-substances-in-electronics.png"},
        {"key": "safe-for-continuous-water-immersion", "label": "Safe for continuous water immersion", "filename": "safe-for-continuous-water-immersion.png"},
        {"key": "safe-for-temporary-water-immersion", "label": "Safe for temporary water immersion", "filename": "safe-for-temporary-water-immersion.png"},
        {"key": "sdcm-3", "label": "SDCM 3", "filename": "sdcm-3.png"},
    ])
    # IP badges were originally labeled with the *descriptive* text baked
    # into the artwork (e.g. "Dust tight, protected against jets") with no
    # literal "IP65" anywhere in the label — fine for browsing, useless for
    # matching against a typed IP Rating spec value. Confirmed the real
    # printed number on each PNG directly (static/cat_badges/*.png) rather
    # than guessing from the filename, then prefixed each label with its
    # real code so /IP\s*(\d{2})/ can find the right one unambiguously.
    # "Designed for use inside/outside buildings only" renamed to the
    # shorter Indoor/Outdoor per explicit request — key/filename unchanged,
    # only the display label changes (avoids invalidating anything that
    # already references the old key).
    _badge_renames = {
        "protected-against-vertically-falling-water": "IP20 - Protected against vertically falling water",
        "protected-against-splashing-water-ingress": "IP44 - Protected against splashing water ingress",
        "dust-protected-water-splash-resistant": "IP54 - Dust protected, water splash resistant",
        "dust-tight-protected-against-jets": "IP65 - Dust tight, protected against jets",
        "dust-tight-strong-water-jets": "IP66 - Dust tight, strong water jets",
        "safe-for-temporary-water-immersion": "IP67 - Safe for temporary water immersion",
        "safe-for-continuous-water-immersion": "IP68 - Safe for continuous water immersion",
        "designed-for-use-inside-buildings-only": "Indoor",
        "designed-for-use-outside-buildings-only": "Outdoor",
    }
    for _b in c["cat_badge_library"]:
        if _b.get("key") in _badge_renames:
            _b["label"] = _badge_renames[_b["key"]]
    # SDCM 1-5 and UGR badges retired per explicit request (replaced by a
    # single "SDCM 3" badge with the number pulled out in orange) — removed
    # from any already-persisted library too, not just the fresh-install
    # default above. RoHS ("restricts-hazardous-substances-in-electronics")
    # retired the same way per a later explicit request — the similarly
    # named "limits-hazardous-substances-in-electronics" is a different
    # badge and stays.
    _retired_badge_keys = {
        "almost-identical-color", "extremely-tight-color-match", "good-color-consistency",
        "high-color-consistency", "visible-color-variation", "low-glare", "minimal-glare", "very-low-glare",
        "restricts-hazardous-substances-in-electronics",
    }
    _before_count = len(c["cat_badge_library"])
    c["cat_badge_library"] = [b for b in c["cat_badge_library"] if b.get("key") not in _retired_badge_keys]
    if len(c["cat_badge_library"]) != _before_count:
        needs_save = True
    if not any(b.get("key") == "sdcm-3" for b in c["cat_badge_library"]):
        c["cat_badge_library"].append({"key": "sdcm-3", "label": "SDCM 3", "filename": "sdcm-3.png"})
        needs_save = True
    if "brand_settings" not in c:
        # one-time migration from the old shared (non-per-brand) settings —
        # everything lands under Artemis, since that was the original brand
        legacy = _empty_brand_settings()
        c.pop("qtn_folder", None)  # legacy Excel quotations, removed for good — nothing left to migrate it into
        legacy["inv_folder"] = c.pop("inv_folder", c.get("folder", ""))
        legacy["do_folder"] = c.pop("do_folder", c.get("folder", ""))
        legacy["product_photos_folder"] = c.pop("product_photos_folder", "")
        legacy["datasheets_folder"] = c.pop("datasheets_folder", "")
        legacy["templates_folder"] = c.pop("templates_folder", "")
        c.pop("folder", None)
        c["brand_settings"] = {engine.DEFAULT_BRAND: legacy}
    for brand in engine.BRANDS:
        c["brand_settings"].setdefault(brand, _empty_brand_settings())
        for k in SETTINGS_FIELDS:
            c["brand_settings"][brand].setdefault(k, "")
    engine.EXTERNAL_TEMPLATES = {b: s.get("templates_folder") or None for b, s in c["brand_settings"].items()}
    if needs_save:
        save_cfg(c)
    return c
def save_cfg(c):
    with open(CONFIG, "w", encoding="utf-8") as f: json.dump(c, f, indent=2, ensure_ascii=False)

photo_store.configure(load_cfg, save_cfg)

def _assign_series_color(cfg, label):
    """Deterministically assigns (and persists into cfg) a distinct color for
    a Series/Category label, auto-generated rather than picked by hand —
    per explicit request, since the list grows freely via Custom… ("remember
    what was typed", see cat_series_labels) so a fixed manual map would
    always be one entry behind. Hue advances by the golden angle (137.508°)
    per already-assigned color, which spreads any number of hues about as
    evenly as possible with no upfront cap and never lands two different
    colors on exactly the same hue; lightness/saturation are pinned to a
    vivid mid-tone band so a color is never close to black, white, or gray
    regardless of hue — a hard requirement for a printed index tab to
    actually read against the page edge. The same label always resolves to
    the same stored color, including if it's removed and re-added later."""
    colors = cfg.setdefault("cat_series_colors", {})
    if label in colors:
        return colors[label]
    hue = (len(colors) * 137.508) % 360
    r, g, b = colorsys.hls_to_rgb(hue / 360, 0.46, 0.62)
    hexval = "#%02x%02x%02x" % (round(r * 255), round(g * 255), round(b * 255))
    colors[label] = hexval
    return hexval

# Every plain-string "grows over time, remembers what you typed" list in the
# app, in one place, so Settings can offer a single generic manage/delete UI
# instead of one-off screens per list. Finish Colors and the Badge Library
# aren't here — they hold richer objects (color hex, image files) and keep
# their own dedicated add/remove routes below; Finish Colors' existing
# /api/cat-finish-colors-remove is still logged the same way (see
# _log_change calls near it), it's just not driven through this registry.
MANAGED_STRING_LISTS = {
    "units": "Units",
    "cat_spec_labels": "Technical Spec Labels",
    "cat_series_labels": "Series / Category",
    "cat_family_labels": "Family Names",
    "cat_ordering_categories": "Ordering Table Column Names",
    "cat_modelno_options": "Model No. Presets",
    "cat_cct_options": "CCT Presets",
    "cat_controls_options": "Controls Presets",
    "cat_voltage_options": "Input Voltage Presets",
    "cat_power_options": "Power Presets",
    "cat_beamangle_options": "Beam Angle Presets",
    "cat_cutout_options": "Cut Out Presets",
    "cat_options_options": "Options Presets",
    "cat_size_index": "Size Presets",
    "expense_payment_methods": "Payment Methods",
    "expense_employees": "Employees",
    "expense_categories": "Expense Categories",
    "expense_products": "Expense Products",
    "expense_descriptions": "Expense Descriptions",
}
# Per-label Technical Specification value memory (cat_spec_values, a dict
# keyed by normalized label, not a flat list) is deliberately NOT in the
# registry above — its "one list per label, growing dynamically" shape
# doesn't fit the generic by-index route, same reasoning that keeps Finish
# Colors and the Badge Library out too. Logged via _log_change with a
# synthetic per-label list_key ("cat_spec_values:<key>") so Activity Log
# still shows it, but there's no delete UI for it yet.

# Shared audit trail for every add/remove against any managed list (called
# right before save_cfg, so the log entry lands in the same write as the
# change it describes). Kept on the config itself rather than a separate
# file since it's small, human-readable JSON either way — capped at 200
# entries so it can't grow unbounded over the life of the app.
def _log_change(cfg, action, list_key, list_label, value):
    log = cfg.setdefault("audit_log", [])
    log.append({
        "ts": datetime.datetime.now().isoformat(timespec="seconds"),
        "action": action,
        "list_key": list_key,
        "list_label": list_label,
        "value": value,
    })
    if len(log) > 200:
        cfg["audit_log"] = log[-200:]

# Mirrors the frontend's normSpecLabel(): case/whitespace-insensitive key so
# "Light Source" and "light source" share the same remembered-value bucket.
def _norm_label_key(label):
    return re.sub(r"\s+", "", (label or "").strip().lower())

# The 1-40W baseline never counts toward the custom-wattage cap (see
# api_cat_power_options_add) — normalized so a non-breaking-space-glued
# custom entry like "9\xa0W" (from appendCatSpecSuffix) still matches the
# compact seeded "9W".
CAT_POWER_BASELINE_NORM = {_norm_label_key(f"{n}W") for n in range(1, 41)}

def current_brand():
    return load_cfg().get("brand", engine.DEFAULT_BRAND)

def brand_settings(brand=None):
    return load_cfg()["brand_settings"].get(brand or current_brand(), _empty_brand_settings())

# ---------------------------------------------------------------- client database
# Same philosophy as every document folder in this app: the data lives
# outside the app, in a real file the user points to (Settings > Company
# Profiles > Clients Spreadsheet, the `clients_file` setting) — not inside
# some internal store. `clients/<BRAND>.json` still exists as a fallback for
# brands that haven't configured a spreadsheet location yet (keeps the
# feature usable out of the box, and is also the one-time migration source
# — see api_set_settings) but is never the source of truth once a location
# is set.
CLIENTS_DIR = os.path.join(engine.DATA_BASE, "clients")

def _legacy_clients_json_path(brand):
    return os.path.join(CLIENTS_DIR, f"{brand}.json")

CLIENT_FIELDS = ("id", "name", "category", "attn", "address", "po_box", "city", "country",
                  "phone", "landline", "email", "website", "trn", "notes", "logo", "updated")

def _load_legacy_clients_json(brand):
    try:
        with open(_legacy_clients_json_path(brand), encoding="utf-8") as f:
            clients = json.load(f)
    except (OSError, ValueError):
        return []
    # older records (saved before category/landline/website/trn existed)
    # won't have those keys yet — fill them in so every caller can rely on
    # the full field set being present, instead of scattering .get(...,'')
    # everywhere.
    for c in clients:
        for f in CLIENT_FIELDS:
            c.setdefault(f, "")
    return clients

def clients_file_path(brand=None):
    return (brand_settings(brand).get("clients_file") or "").strip()

def load_clients(brand=None):
    brand = (brand or current_brand()).upper()
    path = clients_file_path(brand)
    if path:
        return engine.read_clients_workbook(path) if os.path.exists(path) else []
    return _load_legacy_clients_json(brand)

def save_clients(brand, clients):
    brand = brand.upper()
    path = clients_file_path(brand)
    if path:
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        engine.build_clients_workbook(clients).save(path)
        return
    os.makedirs(CLIENTS_DIR, exist_ok=True)
    with open(_legacy_clients_json_path(brand), "w", encoding="utf-8") as f:
        json.dump(clients, f, indent=2)

def _maybe_migrate_clients_to_file(brand, old_path, new_path):
    """The first time a brand's Clients Spreadsheet location gets set: if
    the target file doesn't exist yet, seed it from the legacy JSON store
    so switching to file-based storage doesn't look like data loss. If the
    target already exists (the user pointed at a spreadsheet they already
    had — e.g. from a previous Export), leave it alone; that file is the
    authoritative data now."""
    if not new_path or new_path == old_path or os.path.exists(new_path):
        return
    existing = _load_legacy_clients_json(brand)
    if existing:
        os.makedirs(os.path.dirname(new_path) or ".", exist_ok=True)
        engine.build_clients_workbook(existing).save(new_path)

# ---------------------------------------------------------------- drafts (save in-progress work before Generate, any doc type)
DRAFTS_DIR = os.path.join(engine.DATA_BASE, "drafts")

def _drafts_path(brand):
    return os.path.join(DRAFTS_DIR, f"{brand}.json")

def load_drafts(brand=None):
    brand = (brand or current_brand()).upper()
    path = _drafts_path(brand)
    if not os.path.exists(path):
        return []  # genuinely no drafts saved yet for this brand
    with open(path, encoding="utf-8") as f:
        return json.load(f)
    # Deliberately NOT catching ValueError (corrupt JSON) here — a file that
    # exists but fails to parse is not the same thing as "no drafts", and
    # treating it that way silently is exactly what let one real draft get
    # permanently wiped: a caller that reads "[]" back on a transient/bad
    # read, then saves that "[]" (plus whatever it's adding) right back,
    # overwriting the real content that was actually still sitting in the
    # file the whole time. Let it raise instead — a loud 500 leaves the file
    # on disk untouched for a human to look at, instead of quietly destroying
    # it.

def save_drafts(brand, drafts):
    os.makedirs(DRAFTS_DIR, exist_ok=True)
    path = _drafts_path(brand.upper())
    # One rolling backup of whatever was there immediately before this write
    # — cheap insurance against exactly the failure mode above (or any other
    # bug that computes a wrong `drafts` list): the previous state is always
    # one file away, not gone the moment a bad save lands.
    if os.path.exists(path):
        try:
            shutil.copyfile(path, path + ".bak")
        except OSError:
            pass
    # Write-then-rename so a save that's interrupted partway (crash, killed
    # process) can never leave a half-written, unreadable JSON file behind —
    # same pattern as engine.stamp_catalogue_page_numbers's tmp_path swap.
    tmp_path = path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(drafts, f, indent=2)
    os.replace(tmp_path, path)

# ---------------------------------------------------------------- submissions (QTN -> LPO -> DO -> scanned DO -> INV, bundled and tracked together)
SUBMISSIONS_DIR = os.path.join(engine.DATA_BASE, "submissions")

def _submissions_path(brand):
    return os.path.join(SUBMISSIONS_DIR, f"{brand}.json")

def load_submissions(brand=None):
    brand = (brand or current_brand()).upper()
    try:
        with open(_submissions_path(brand), encoding="utf-8") as f:
            return json.load(f)
    except (OSError, ValueError):
        return []

def save_submissions(brand, subs):
    os.makedirs(SUBMISSIONS_DIR, exist_ok=True)
    with open(_submissions_path(brand.upper()), "w", encoding="utf-8") as f:
        json.dump(subs, f, indent=2)

# ---------------------------------------------------------------- finance ledger (Statement of Account)
# Deliberately NOT backfilled from the thousands of real historical
# invoices — there's no reliable record of what was actually paid for any
# of them, so treating them all as "unpaid" would make the outstanding
# balance look enormous and mostly wrong. Per the user's own call, this
# only tracks invoices generated from today forward; older ones simply
# don't appear here.
FINANCE_DIR = os.path.join(engine.DATA_BASE, "finance")

def _finance_path(brand):
    return os.path.join(FINANCE_DIR, f"{brand}.json")

def load_ledger(brand=None):
    brand = (brand or current_brand()).upper()
    try:
        with open(_finance_path(brand), encoding="utf-8") as f:
            return json.load(f)
    except (OSError, ValueError):
        return []

def save_ledger(brand, entries):
    os.makedirs(FINANCE_DIR, exist_ok=True)
    with open(_finance_path(brand.upper()), "w", encoding="utf-8") as f:
        json.dump(entries, f, indent=2)

def record_invoice_in_ledger(data, xlsx_path, brand, inv_folder):
    """Called every time a real Invoice xlsx is generated — whether through
    the normal Build/Generate flow or Submissions' auto-generation — so the
    ledger never depends on re-reading totals back out of Excel formula
    cells (unreliable for a file this app just wrote and nobody has
    opened in Excel yet to compute them). Totals are computed directly in
    Python from the same items/discount/vat the xlsx was just filled
    with, via the same compute_totals() QTN2 already uses for its own
    printed totals — one shared source of truth, not two."""
    try:
        totals = html_engine.compute_totals(data.get("items") or [], data.get("discount"), data.get("vat"))
    except Exception:
        return
    rel = os.path.relpath(xlsx_path, inv_folder).replace(os.sep, "/")
    entries = load_ledger(brand)
    entry = next((e for e in entries if e["rel"] == rel), None)
    now = datetime.datetime.now().isoformat()
    if entry is None:
        entry = {"rel": rel, "paid": False, "paid_date": None, "created": now}
        entries.append(entry)
    entry.update({
        "number": data.get("number", ""), "company": data.get("company", ""),
        "date": data.get("date", ""), "project": data.get("project", ""),
        "subtotal": round(totals["subtotal"], 2), "vat": round(totals["vat_amount"], 2),
        "discount": round(totals["discount_amount"], 2), "total": round(totals["total"], 2),
        "updated": now,
    })
    save_ledger(brand, entries)

def _brand_matches(rec, brand):
    """A record counts toward `brand` if it's tagged with that brand, or if
    it's untagged legacy data and `brand` is the original default (Artemis) —
    legacy docs count as Artemis for numbering/history continuity."""
    b = rec.get("brand")
    return b == brand if b else brand == engine.DEFAULT_BRAND

FOLDER_KEYS = {"INV": "inv_folder", "DO": "do_folder", "QTN2": "qtn2_folder", "PI": "pi_folder",
               "RV": "rv_folder", "CN": "cn_folder", "CAT": "catalogue_folder", "EXP": "expense_folder"}

def folder_for(doc_type, brand=None):
    return brand_settings(brand).get(FOLDER_KEYS.get((doc_type or "").upper(), ""), "")

def all_doc_folders(brand=None):
    """Every distinct configured document folder (INV/DO/QTN2) for a brand,
    deduplicated — they may all point at the same place, or each somewhere different.
    Sololuce Datasheets' catalogue_folder is always included too, regardless of
    which brand this call is for — /api/generate hardcodes gen_brand=SOLOLUCE for
    every CAT doc (see its own comment), so a CAT file always lives under the
    SOLOLUCE bucket's catalogue_folder no matter which brand is active when the
    user later opens All Docs. Without this, All Docs would silently stop
    scanning that folder (and so stop showing any Sololuce Datasheet at all)
    the moment the active brand drifts away from SOLOLUCE."""
    s = brand_settings(brand)
    out = []
    for key in ("inv_folder", "do_folder", "qtn2_folder", "pi_folder", "rv_folder", "cn_folder", "catalogue_folder", "expense_folder"):
        f = s.get(key, "")
        if f and f not in out:
            out.append(f)
    sololuce_catalogue = brand_settings("SOLOLUCE").get("catalogue_folder", "")
    if sololuce_catalogue and sololuce_catalogue not in out:
        out.append(sololuce_catalogue)
    return out

def resolve_rel(rel, brand=None):
    """Find which of a brand's configured document folders a relative path
    actually lives under. Returns (folder, abs_path) or (None, None)."""
    for folder in all_doc_folders(brand):
        path = os.path.join(folder, rel)
        if os.path.isdir(folder) and os.path.abspath(path).startswith(os.path.abspath(folder)) and os.path.exists(path):
            return folder, path
    return None, None

_tk_job_queue = None
_tk_worker_lock = threading.Lock()

def _tk_worker():
    """Owns Tkinter for the whole life of the process, on ONE dedicated OS
    thread that never changes. Flask's dev server hands each request to a
    fresh worker thread (confirmed: successive requests land on different
    threads, never the real main thread) — Tcl's notifier is bound to
    whichever specific thread created it, so a Tk() root touched from a
    different thread than the one that made it throws exactly "main
    thread is not in main loop". Routing every dialog call through this
    single long-lived thread via a job queue sidesteps that: Tkinter only
    ever sees the one thread that created its root, no matter which
    request thread asked for the dialog."""
    import tkinter as tk
    root = tk.Tk()
    root.withdraw()
    while True:
        job_fn, reply_q = _tk_job_queue.get()
        try:
            reply_q.put(("ok", job_fn(root)))
        except Exception as e:
            reply_q.put(("error", str(e)))

def _run_on_tk_thread(job_fn, timeout=180):
    """Run job_fn(root) on the dedicated Tkinter thread and block for its
    result. job_fn should show/close whatever dialog it needs and return
    the chosen path (or "" if the user cancelled)."""
    global _tk_job_queue
    with _tk_worker_lock:
        if _tk_job_queue is None:
            _tk_job_queue = queue.Queue()
            threading.Thread(target=_tk_worker, daemon=True, name="tk-dialog-worker").start()
    reply_q = queue.Queue()
    _tk_job_queue.put((job_fn, reply_q))
    status, result = reply_q.get(timeout=timeout)
    if status == "error":
        raise RuntimeError(result)
    return result

# ---------------------------------------------------------------- native folder picker
@app.post("/api/browse")
def browse():
    """Open a native folder (or, for clients_file, file) dialog on the
    machine running this app, and save the chosen path into the current
    brand's settings field that asked for it."""
    field = (request.json or {}).get("field", "")
    if field not in SETTINGS_FIELDS:
        return jsonify({"error": "Unknown settings field."}), 400
    try:
        from tkinter import filedialog
        def job(root):
            root.attributes("-topmost", True)
            if field in FILE_SETTINGS_FIELDS:
                # asksaveasfilename (not askopenfilename) so the user can either
                # pick an existing spreadsheet or type a brand-new filename to
                # create one — both are valid here, unlike a normal "open" dialog.
                p = filedialog.asksaveasfilename(
                    parent=root, title="Choose or create a Clients spreadsheet", defaultextension=".xlsx",
                    filetypes=[("Excel workbook", "*.xlsx")], confirmoverwrite=False)
            else:
                p = filedialog.askdirectory(parent=root, title="Choose a folder")
            root.attributes("-topmost", False)
            return p
        path = _run_on_tk_thread(job)
        if path:
            brand = current_brand()
            cfg = load_cfg()
            old_path = cfg["brand_settings"][brand].get(field, "")
            cfg["brand_settings"][brand][field] = path
            save_cfg(cfg)
            if field == "clients_file":
                _maybe_migrate_clients_to_file(brand, old_path, path)
        return jsonify({"field": field, "value": path or ""})
    except Exception as e:
        return jsonify({"error": f"Could not open dialog: {e}. Paste the path instead."}), 200

@app.get("/api/config")
def get_config():
    return jsonify(load_cfg())

@app.get("/api/settings")
def get_settings():
    return jsonify(brand_settings())

@app.post("/api/settings")
def set_settings():
    data = request.json or {}
    brand = current_brand()
    cfg = load_cfg()
    bucket = cfg["brand_settings"][brand]
    old_clients_file = bucket.get("clients_file", "")
    for key, val in data.items():
        if key not in SETTINGS_FIELDS:
            continue
        val = (val or "").strip()
        if key in FILE_SETTINGS_FIELDS:
            if val and not os.path.isdir(os.path.dirname(val) or "."):
                return jsonify({"error": f"That folder doesn't exist: {os.path.dirname(val)}"}), 400
        elif val and not os.path.isdir(val):
            return jsonify({"error": f"This folder doesn't exist: {val}"}), 400
        bucket[key] = val
    save_cfg(cfg)
    _maybe_migrate_clients_to_file(brand, old_clients_file, bucket.get("clients_file", ""))
    return jsonify(bucket)

@app.get("/api/brands")
def api_brands():
    return jsonify({"brands": [{"code": c, "label": l} for c, l in engine.BRANDS.items()],
                     "current": current_brand()})

@app.post("/api/set-brand")
def api_set_brand():
    code = (request.json or {}).get("brand", "").strip().upper()
    if code not in engine.BRANDS:
        return jsonify({"error": "Unknown brand."}), 400
    lock = session.get("brand_lock")
    if lock and code != lock:
        return jsonify({"error": "Your account is restricted to " + engine.BRANDS.get(lock, lock) + "."}), 403
    cfg = load_cfg(); cfg["brand"] = code; save_cfg(cfg)
    return jsonify({"brand": code})

@app.get("/api/units")
def api_units():
    return jsonify({"units": load_cfg().get("units", [])})

@app.get("/api/cat-next-page")
def api_cat_next_page():
    """A convenience suggestion for the Page Number field on a brand-new
    datasheet — never authoritative (only a Full Catalog Build fixes real
    positions; see catalog_builder.py). Computed fresh from whatever's on
    disk each time rather than trusting a persisted counter, which would
    drift the moment products get grouped/reordered: one past the highest
    Page Number already typed into any real datasheet in the folder, or one
    past the last real build's total page count, whichever is bigger, or 1
    if neither exists yet. no-store so the browser's GET heuristic caching
    can't serve back a stale suggestion from before the last Generate/Build."""
    cfg = load_cfg()
    folder = brand_settings("SOLOLUCE").get("catalogue_folder", "")
    best = 0
    if folder and os.path.isdir(folder):
        for rec in engine.index_folder(folder):
            if rec["type"].upper() != "CAT" or rec["ext"].lower() != "pdf":
                continue
            try:
                pn = int(engine.read_sidecar(rec["path"]).get("page_number") or 0)
            except (TypeError, ValueError):
                pn = 0
            best = max(best, pn)
    last_build = cfg.get("catalog_last_build") or {}
    best = max(best, int(last_build.get("total_pages") or 0))
    resp = jsonify({"next_page": best + 1 if best else 1})
    resp.headers["Cache-Control"] = "no-store"
    return resp

@app.get("/api/cat-spec-labels")
def api_cat_spec_labels():
    return jsonify({"labels": load_cfg().get("cat_spec_labels", [])})

@app.post("/api/cat-spec-labels-add")
def api_cat_spec_labels_add():
    label = (request.json or {}).get("label", "").strip()
    cfg = load_cfg()
    labels = cfg.get("cat_spec_labels", [])
    if label and label not in labels:
        labels.append(label)
        cfg["cat_spec_labels"] = labels
        _log_change(cfg, "add", "cat_spec_labels", MANAGED_STRING_LISTS["cat_spec_labels"], label)
        save_cfg(cfg)
    return jsonify({"labels": labels})

@app.get("/api/cat-series")
def api_cat_series():
    cfg = load_cfg()
    return jsonify({"labels": cfg.get("cat_series_labels", []), "colors": cfg.get("cat_series_colors", {}),
                     "sections": cfg.get("cat_series_sections", {})})

@app.post("/api/cat-series-add")
def api_cat_series_add():
    label = (request.json or {}).get("label", "").strip()
    cfg = load_cfg()
    labels = cfg.get("cat_series_labels", [])
    if label and label not in labels:
        labels.append(label)
        cfg["cat_series_labels"] = labels
        _log_change(cfg, "add", "cat_series_labels", MANAGED_STRING_LISTS["cat_series_labels"], label)
    if label:
        _assign_series_color(cfg, label)  # no-op if it already has one
    save_cfg(cfg)
    return jsonify({"labels": labels, "colors": cfg.get("cat_series_colors", {})})

@app.post("/api/cat-series-move")
def api_cat_series_move():
    """Swap a category with its up/down neighbor in cat_series_labels —
    this array's own order IS the Full Catalog Builder's category display
    order (see catalog_builder.py's group_and_order), so reordering here is
    exactly reordering the eventual printed catalog."""
    data = request.json or {}
    label, direction = (data.get("label") or "").strip(), data.get("direction", "")
    cfg = load_cfg()
    labels = cfg.get("cat_series_labels", [])
    if label not in labels:
        return jsonify({"error": "Unknown category."}), 400
    i = labels.index(label)
    j = i - 1 if direction == "up" else i + 1 if direction == "down" else None
    if j is not None and 0 <= j < len(labels):
        labels[i], labels[j] = labels[j], labels[i]
        cfg["cat_series_labels"] = labels
        save_cfg(cfg)
    return jsonify({"labels": labels})

@app.post("/api/cat-series-color")
def api_cat_series_color():
    """Explicit color override — _assign_series_color only ever picks one
    automatically for a category that has none yet; this is the one place
    a user-chosen hex actually overwrites whatever's there, auto-assigned
    or not."""
    data = request.json or {}
    label = (data.get("label") or "").strip()
    hexval = (data.get("hex") or "").strip()
    cfg = load_cfg()
    if label not in cfg.get("cat_series_labels", []):
        return jsonify({"error": "Unknown category."}), 400
    if not re.match(r"^#[0-9a-fA-F]{6}$", hexval):
        return jsonify({"error": "Invalid color."}), 400
    colors = cfg.setdefault("cat_series_colors", {})
    colors[label] = hexval
    save_cfg(cfg)
    return jsonify({"colors": colors})

@app.post("/api/cat-series-rename")
def api_cat_series_rename():
    """Renames a category in place — the remembered list, its color, and
    its Index Order entry all move to the new name. Critically, also
    rewrites the `series` field in every already-generated CAT product's
    OWN sidecar that still says the old name: a product's category lives
    in that sidecar, not derived from this list, so without this migration
    every existing product tagged with the old name would silently fall out
    into the "no category set" catch-all the next time anything reads
    cat_series_labels. Silently skipped (not an error) if no catalogue
    folder is configured yet — nothing to migrate."""
    data = request.json or {}
    old = (data.get("old_label") or "").strip()
    new = (data.get("new_label") or "").strip()
    if not old or not new:
        return jsonify({"error": "Both names are required."}), 400
    cfg = load_cfg()
    labels = cfg.get("cat_series_labels", [])
    if old not in labels:
        return jsonify({"error": "Unknown category."}), 400
    if new != old and new in labels:
        return jsonify({"error": f'"{new}" already exists.'}), 400
    labels[labels.index(old)] = new
    cfg["cat_series_labels"] = labels
    colors = cfg.setdefault("cat_series_colors", {})
    if old in colors:
        colors[new] = colors.pop(old)
    order = cfg.setdefault("catalog_index_order", {})
    if old in order:
        order[new] = order.pop(old)
    migrated = 0
    if new != old:
        folder = brand_settings("SOLOLUCE").get("catalogue_folder", "")
        if folder and os.path.isdir(folder):
            for rec in engine.index_folder(folder):
                if rec["type"].upper() != "CAT" or rec["ext"].lower() != "pdf":
                    continue
                sidecar = engine.read_sidecar(rec["path"])
                if sidecar.get("series") == old:
                    sidecar["series"] = new
                    engine.save_sidecar(rec["path"], sidecar)
                    migrated += 1
        _log_change(cfg, "rename", "cat_series_labels", MANAGED_STRING_LISTS["cat_series_labels"], f"{old} → {new}")
    save_cfg(cfg)
    return jsonify({"labels": labels, "colors": colors, "migrated_products": migrated})

@app.post("/api/cat-series-sort")
def api_cat_series_sort():
    """mode: "alpha" (A-Z) or "date" (earliest-added first, per the
    audit_log's own "add" entries for this list — a category from before
    audit logging existed, or seeded by default, has no such entry and
    sorts first, in its current relative order, since it necessarily
    predates anything that IS logged). One-shot action, not a persisted
    mode — the result becomes the new manual order, still freely
    draggable afterward via /api/cat-series-move."""
    mode = (request.json or {}).get("mode", "")
    cfg = load_cfg()
    labels = cfg.get("cat_series_labels", [])
    if mode == "alpha":
        labels = sorted(labels, key=lambda l: l.strip().lower())
    elif mode == "date":
        first_added = {}
        for e in cfg.get("audit_log", []):
            if e.get("list_key") == "cat_series_labels" and e.get("action") == "add":
                v = e.get("value")
                if v not in first_added:
                    first_added[v] = e.get("ts", "")
        labels = sorted(labels, key=lambda l: (l in first_added, first_added.get(l, "")))
    else:
        return jsonify({"error": "Unknown sort mode."}), 400
    cfg["cat_series_labels"] = labels
    save_cfg(cfg)
    return jsonify({"labels": labels})

@app.post("/api/cat-series-reorder")
def api_cat_series_reorder():
    """Full-list reorder — the frontend computes the complete new order
    (e.g. after a drag-and-drop within one Outdoor/Indoor group, splicing
    that group's new sub-order back into the other groups' unchanged
    positions) and posts the whole thing; this just validates it's a true
    permutation of the current list (nothing added, dropped, or duplicated)
    before saving. /api/cat-series-move's adjacent swap-with-neighbor stays
    for the ▲▼ arrows; this is for drag-and-drop's arbitrary repositioning."""
    labels = (request.json or {}).get("labels")
    if not isinstance(labels, list):
        return jsonify({"error": "Invalid order."}), 400
    cfg = load_cfg()
    current = cfg.get("cat_series_labels", [])
    if sorted(labels) != sorted(current):
        return jsonify({"error": "That order doesn't match the current category list."}), 400
    cfg["cat_series_labels"] = labels
    save_cfg(cfg)
    return jsonify({"labels": labels})

@app.post("/api/cat-series-randomize-colors")
def api_cat_series_randomize_colors():
    """One-shot action (not a persisted mode, same idea as sort above) —
    regenerates EVERY category's color at once, evenly spread by the golden
    angle (same method _assign_series_color uses for a single new one) from
    a random starting hue each time, so repeat clicks give a different but
    still well-spread palette — no two colors land close enough in hue to
    read as "the same shade". Lightness/saturation stay pinned to the same
    vivid mid-tone band _assign_series_color always uses, so nothing ever
    comes out close to black or white regardless of hue."""
    cfg = load_cfg()
    labels = cfg.get("cat_series_labels", [])
    colors = cfg.setdefault("cat_series_colors", {})
    start = random.uniform(0, 360)
    for i, label in enumerate(labels):
        hue = (start + i * 137.508) % 360
        r, g, b = colorsys.hls_to_rgb(hue / 360, 0.46, 0.62)
        colors[label] = "#%02x%02x%02x" % (round(r * 255), round(g * 255), round(b * 255))
    save_cfg(cfg)
    return jsonify({"colors": colors})

@app.post("/api/cat-series-section")
def api_cat_series_section():
    """Tags a category as Outdoor/Indoor/Striplight (or clears the tag with
    an empty value) — purely an organizational hint for grouping the
    Category Order screen, NOT wired into the actual build (a product's
    real section always comes from ITS OWN Product Type field, same as
    always — see catalog_builder.py's group_and_order). A category with no
    tag yet shows in its own "Unassigned" group until set."""
    data = request.json or {}
    label = (data.get("label") or "").strip()
    section = (data.get("section") or "").strip()
    cfg = load_cfg()
    if label not in cfg.get("cat_series_labels", []):
        return jsonify({"error": "Unknown category."}), 400
    if section and section not in catalog_builder.SECTION_VALUES:
        return jsonify({"error": "Unknown section."}), 400
    sections = cfg.setdefault("cat_series_sections", {})
    if section:
        sections[label] = section
    else:
        sections.pop(label, None)
    save_cfg(cfg)
    return jsonify({"sections": sections})

@app.get("/api/cat-family")
def api_cat_family():
    return jsonify({"labels": load_cfg().get("cat_family_labels", [])})

@app.post("/api/cat-family-add")
def api_cat_family_add():
    label = (request.json or {}).get("label", "").strip()
    cfg = load_cfg()
    labels = cfg.get("cat_family_labels", [])
    if label and label not in labels:
        labels.append(label)
        cfg["cat_family_labels"] = labels
        _log_change(cfg, "add", "cat_family_labels", MANAGED_STRING_LISTS["cat_family_labels"], label)
        save_cfg(cfg)
    return jsonify({"labels": labels})

@app.get("/api/cat-products-search")
def api_cat_products_search():
    """Backs the Family card's linked-products search AND its chip list (the
    latter calls this with family=<name>&q= to get that family's current
    live membership). Always scoped to Sololuce's own catalogue_folder —
    never the unrelated datasheets_folder/Product Finder subsystem the
    Quotation Builder uses."""
    folder = brand_settings("SOLOLUCE").get("catalogue_folder", "")
    if not folder or not os.path.isdir(folder):
        return jsonify({"results": []})
    results = catalog_builder.search_products(
        folder,
        query=request.args.get("q", ""),
        family=request.args.get("family", ""),
        exclude_rel=request.args.get("exclude", ""),
    )
    return jsonify({"results": results})

@app.post("/api/cat-family-link")
def api_cat_family_link():
    """Two-way linking: adding/removing an already-generated datasheet
    to/from a family rewrites THAT product's own saved sidecar immediately —
    the currently-open form's own family value still only saves when the
    user hits Generate, same as every other field on the form. `rel` is a
    folder-relative path from search_products; re-validated against
    Sololuce's own catalogue_folder here (containment + existence) before
    ever writing to disk, since it arrives from the client."""
    data = request.json or {}
    family = (data.get("family") or "").strip()
    add_rels = data.get("add") or []
    remove_rels = data.get("remove") or []
    folder = brand_settings("SOLOLUCE").get("catalogue_folder", "")
    if not folder or not os.path.isdir(folder):
        return jsonify({"error": "Sololuce catalogue folder isn't set."}), 400
    if add_rels and not family:
        return jsonify({"error": "Family name is required to link products."}), 400

    def _resolve(rel):
        path = os.path.join(folder, rel)
        if not os.path.abspath(path).startswith(os.path.abspath(folder)) or not os.path.isfile(path):
            return None
        return path

    for rel in add_rels:
        path = _resolve(rel)
        sidecar = engine.read_sidecar(path) if path else None
        if not sidecar:
            continue
        sidecar["family"] = family
        engine.save_sidecar(path, sidecar)
    for rel in remove_rels:
        path = _resolve(rel)
        sidecar = engine.read_sidecar(path) if path else None
        if not sidecar:
            continue
        sidecar["family"] = ""
        engine.save_sidecar(path, sidecar)
    return jsonify({"ok": True})

# ---------------------------------------------------------------- Full Catalog Builder
# Assembles every generated Sololuce Datasheet into one bound book — see
# catalog_builder.py for the full architecture. Sololuce-only and always
# hardcoded to that brand's own folders, same reasoning as /api/generate's
# gen_brand hardcode for CAT itself: trusting current_brand() here would let
# a stray brand switch quietly point this at the wrong bucket.
_FULL_CATALOG_FILENAME = "Sololuce_Catalog.pdf"

def _full_catalog_output_path():
    folder = brand_settings("SOLOLUCE").get("full_catalog_folder", "")
    return os.path.join(folder, _FULL_CATALOG_FILENAME) if folder else ""

@app.get("/api/full-catalog/summary")
def api_full_catalog_summary():
    folder = brand_settings("SOLOLUCE").get("catalogue_folder", "")
    if not folder or not os.path.isdir(folder):
        return jsonify({"error": "Set the Sololuce Datasheets folder first (see Settings)."}), 400
    return jsonify(catalog_builder.summarize(folder, load_cfg()))

@app.post("/api/full-catalog/build")
def api_full_catalog_build():
    folder = brand_settings("SOLOLUCE").get("catalogue_folder", "")
    if not folder or not os.path.isdir(folder):
        return jsonify({"error": "Set the Sololuce Datasheets folder first (see Settings)."}), 400
    out_folder = brand_settings("SOLOLUCE").get("full_catalog_folder", "")
    if not out_folder:
        return jsonify({"error": "Set the Full Catalog output folder first (see Settings)."}), 400
    if not os.path.isdir(out_folder):
        return jsonify({"error": f"This folder doesn't exist: {out_folder}"}), 400
    try:
        result = catalog_builder.build_full_catalog(
            folder, os.path.join(out_folder, _FULL_CATALOG_FILENAME), load_cfg())
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"Build failed: {e}"}), 500
    cfg = load_cfg()
    cfg["catalog_last_build"] = result
    save_cfg(cfg)
    return jsonify(result)

@app.get("/full-catalog-pdf")
def full_catalog_pdf():
    path = _full_catalog_output_path()
    if not path or not os.path.exists(path):
        return "Not found", 404
    return send_file(path, mimetype="application/pdf")

@app.get("/full-catalog-page")
def full_catalog_page():
    """One rasterized page of the last-built catalogue, for the Full Catalog
    Builder screen's inline page-by-page preview. Deliberately NOT the same
    pattern as the per-document live preview (which rasterizes every page up
    front via engine.to_png_pages) — that's fine for a few-page document but
    would mean rasterizing a whole 800-page book on every preview open, so
    this renders exactly the one page asked for, on demand, straight from
    the already-built PDF on disk."""
    path = _full_catalog_output_path()
    if not path or not os.path.exists(path):
        return "Not found", 404
    try:
        page_num = int(request.args.get("page", "1"))
    except ValueError:
        page_num = 1
    import fitz
    with fitz.open(path) as doc:
        if page_num < 1 or page_num > doc.page_count:
            return "Not found", 404
        pix = doc[page_num - 1].get_pixmap(matrix=fitz.Matrix(170 / 72, 170 / 72))
        png_bytes = pix.tobytes("png")
    resp = Response(png_bytes, mimetype="image/png")
    resp.headers["Cache-Control"] = "no-store"
    return resp

@app.get("/api/full-catalog/section-order")
def api_full_catalog_section_order():
    cfg = load_cfg()
    return jsonify({
        "order": cfg.get("catalog_section_order", list(catalog_builder.SECTION_VALUES)),
        "labels": cfg.get("catalog_section_labels", {}),
        "colors": cfg.get("catalog_section_colors", {}),
    })

@app.post("/api/full-catalog/section-order-move")
def api_full_catalog_section_order_move():
    data = request.json or {}
    label, direction = (data.get("label") or "").strip(), data.get("direction", "")
    cfg = load_cfg()
    order = cfg.get("catalog_section_order", list(catalog_builder.SECTION_VALUES))
    if label not in order:
        return jsonify({"error": "Unknown section."}), 400
    i = order.index(label)
    j = i - 1 if direction == "up" else i + 1 if direction == "down" else None
    if j is not None and 0 <= j < len(order):
        order[i], order[j] = order[j], order[i]
        cfg["catalog_section_order"] = order
        save_cfg(cfg)
    return jsonify({"order": order})

@app.post("/api/full-catalog/section-label")
def api_full_catalog_section_label():
    """Renames a section's DISPLAY label only — the underlying value
    (Outdoor/Indoor/Striplight) stays fixed, since Product Type, Class
    1/2/3 badge auto-assignment, and section grouping all key off that
    literal value throughout the app. What actually prints in the built
    catalogue (top-level Index, per-section "{LABEL} INDEX" header, Last
    Build summary) uses this display label instead — see
    catalog_builder.py's build_full_catalog."""
    data = request.json or {}
    section = (data.get("section") or "").strip()
    label = (data.get("label") or "").strip()
    if section not in catalog_builder.SECTION_VALUES:
        return jsonify({"error": "Unknown section."}), 400
    if not label:
        return jsonify({"error": "Label can't be empty."}), 400
    cfg = load_cfg()
    labels = cfg.setdefault("catalog_section_labels", {})
    labels[section] = label
    save_cfg(cfg)
    return jsonify({"labels": labels})

@app.post("/api/full-catalog/section-color")
def api_full_catalog_section_color():
    data = request.json or {}
    section = (data.get("section") or "").strip()
    hexval = (data.get("hex") or "").strip()
    if section not in catalog_builder.SECTION_VALUES:
        return jsonify({"error": "Unknown section."}), 400
    if not re.match(r"^#[0-9a-fA-F]{6}$", hexval):
        return jsonify({"error": "Invalid color."}), 400
    cfg = load_cfg()
    colors = cfg.setdefault("catalog_section_colors", {})
    colors[section] = hexval
    save_cfg(cfg)
    return jsonify({"colors": colors})

def _save_catalog_extra_pdf(data_url):
    """Decode a data:application/pdf;base64,... URL and write it under
    catalog_builder.CATALOG_EXTRAS_DIR with a fresh uuid4 filename — never a
    name derived from the client-supplied original filename, same
    containment-by-construction reasoning as /api/cat-family-link's rel
    checks (except here there's nothing to validate against, since the app
    itself chooses the filename outright). Validates it actually opens as a
    PDF right away (via fitz) so a bad upload surfaces immediately, not
    confusingly mid-Build. Returns the stored filename (not a full path).
    Raises ValueError on any problem."""
    if "," not in data_url:
        raise ValueError("No file received.")
    _header, b64 = data_url.split(",", 1)
    os.makedirs(catalog_builder.CATALOG_EXTRAS_DIR, exist_ok=True)
    stored = f"{uuid.uuid4().hex}.pdf"
    path = os.path.join(catalog_builder.CATALOG_EXTRAS_DIR, stored)
    try:
        with open(path, "wb") as f:
            f.write(base64.b64decode(b64))
        import fitz
        with fitz.open(path):
            pass  # just confirming it actually parses as a PDF
    except Exception as e:
        if os.path.exists(path):
            os.remove(path)
        raise ValueError(f"Couldn't read that PDF: {e}")
    return stored

def _delete_catalog_extra(stored_as):
    if not stored_as:
        return
    path = os.path.join(catalog_builder.CATALOG_EXTRAS_DIR, stored_as)
    try:
        if os.path.exists(path):
            os.remove(path)
    except OSError:
        pass

@app.get("/api/full-catalog/extras")
def api_full_catalog_extras():
    return jsonify(load_cfg().get("catalog_extras", {"front_matter": [], "family_dividers": {}}))

@app.post("/api/full-catalog/extras/<slot>")
def api_full_catalog_extras_set(slot):
    """slot: cover | ending — the two single fixed slots, always first and
    always last respectively. Introduction/custom pages live in the
    reorderable catalog_extras["front_matter"] list instead, in between
    (see the /api/full-catalog/front-matter* routes below). Uploading a new
    file for an already-filled slot replaces it (old file deleted)."""
    if slot not in ("cover", "ending"):
        return jsonify({"error": "Unknown slot."}), 400
    data = request.json or {}
    try:
        stored = _save_catalog_extra_pdf(data.get("pdf", ""))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    cfg = load_cfg()
    extras = cfg.setdefault("catalog_extras", {})
    extras.setdefault("family_dividers", {})
    old = extras.get(slot)
    extras[slot] = {"filename": (data.get("filename") or "uploaded.pdf").strip(), "stored_as": stored}
    save_cfg(cfg)
    _delete_catalog_extra(old.get("stored_as") if old else None)
    return jsonify({"ok": True, "extras": extras})

@app.post("/api/full-catalog/extras/<slot>/remove")
def api_full_catalog_extras_remove(slot):
    if slot not in ("cover", "ending"):
        return jsonify({"error": "Unknown slot."}), 400
    cfg = load_cfg()
    extras = cfg.setdefault("catalog_extras", {})
    extras.setdefault("family_dividers", {})
    old = extras.pop(slot, None)
    save_cfg(cfg)
    _delete_catalog_extra(old.get("stored_as") if old else None)
    return jsonify({"ok": True, "extras": extras})

@app.post("/api/full-catalog/front-matter")
def api_full_catalog_front_matter_set():
    """Upload/replace a front-matter row's file. body: {id, pdf, filename}.
    id must match an existing row (builtin "introduction", or a custom
    row's own generated id) — this route only ever attaches a file to a
    row that already exists; see front-matter-add for creating one."""
    data = request.json or {}
    item_id = (data.get("id") or "").strip()
    cfg = load_cfg()
    extras = cfg.setdefault("catalog_extras", {})
    items = extras.setdefault("front_matter", [])
    item = next((it for it in items if it.get("id") == item_id), None)
    if not item:
        return jsonify({"error": "Unknown front-matter page."}), 400
    try:
        stored = _save_catalog_extra_pdf(data.get("pdf", ""))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    old_stored = item.get("stored_as")
    item["filename"] = (data.get("filename") or "uploaded.pdf").strip()
    item["stored_as"] = stored
    save_cfg(cfg)
    _delete_catalog_extra(old_stored)
    return jsonify({"ok": True, "extras": extras})

@app.post("/api/full-catalog/front-matter-add")
def api_full_catalog_front_matter_add():
    """Adds a new custom front-matter row, no file yet (upload separately
    via /api/full-catalog/front-matter once it has an id). body: {label}."""
    label = (request.json or {}).get("label", "").strip()
    if not label:
        return jsonify({"error": "Give the page a name first."}), 400
    cfg = load_cfg()
    extras = cfg.setdefault("catalog_extras", {})
    items = extras.setdefault("front_matter", [])
    items.append({"id": uuid.uuid4().hex, "label": label, "builtin": False})
    save_cfg(cfg)
    return jsonify({"ok": True, "extras": extras})

@app.post("/api/full-catalog/front-matter/remove")
def api_full_catalog_front_matter_remove():
    """For the builtin "introduction" row this only clears its uploaded
    file, reverting to "Not set" — the row itself is permanent, same as
    /api/full-catalog/extras/<slot>/remove for Cover/Ending. For a custom
    row it deletes the row entirely. body: {id}."""
    item_id = (request.json or {}).get("id", "").strip()
    cfg = load_cfg()
    extras = cfg.setdefault("catalog_extras", {})
    items = extras.setdefault("front_matter", [])
    item = next((it for it in items if it.get("id") == item_id), None)
    if not item:
        return jsonify({"error": "Unknown front-matter page."}), 400
    old_stored = item.get("stored_as")
    if item.get("builtin"):
        item.pop("filename", None)
        item.pop("stored_as", None)
    else:
        items.remove(item)
    save_cfg(cfg)
    _delete_catalog_extra(old_stored)
    return jsonify({"ok": True, "extras": extras})

@app.post("/api/full-catalog/front-matter-move")
def api_full_catalog_front_matter_move():
    """Swap a front-matter row with its up/down neighbor — same
    swap-with-neighbor pattern as /api/cat-series-move. body: {id, direction}."""
    data = request.json or {}
    item_id, direction = (data.get("id") or "").strip(), data.get("direction", "")
    cfg = load_cfg()
    extras = cfg.setdefault("catalog_extras", {})
    items = extras.setdefault("front_matter", [])
    ids = [it.get("id") for it in items]
    if item_id not in ids:
        return jsonify({"error": "Unknown front-matter page."}), 400
    i = ids.index(item_id)
    j = i - 1 if direction == "up" else i + 1 if direction == "down" else None
    if j is not None and 0 <= j < len(items):
        items[i], items[j] = items[j], items[i]
        save_cfg(cfg)
    return jsonify({"extras": extras})

@app.post("/api/full-catalog/family-divider")
def api_full_catalog_family_divider_set():
    """Per-family version of the same upload slot — tagged by family name
    (from the same cat_family_labels every other Family Name picker in this
    app uses) rather than a fixed key. No upload for a given family simply
    means that family's members cluster together with no divider page in
    front of them (see catalog_builder.py's _resolve_extra)."""
    data = request.json or {}
    family = (data.get("family") or "").strip()
    if not family:
        return jsonify({"error": "Pick a family first."}), 400
    try:
        stored = _save_catalog_extra_pdf(data.get("pdf", ""))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    cfg = load_cfg()
    extras = cfg.setdefault("catalog_extras", {})
    dividers = extras.setdefault("family_dividers", {})
    old = dividers.get(family)
    dividers[family] = {"filename": (data.get("filename") or "uploaded.pdf").strip(), "stored_as": stored}
    save_cfg(cfg)
    _delete_catalog_extra(old.get("stored_as") if old else None)
    return jsonify({"ok": True, "extras": extras})

@app.post("/api/full-catalog/family-divider/remove")
def api_full_catalog_family_divider_remove():
    family = (request.json or {}).get("family", "").strip()
    cfg = load_cfg()
    extras = cfg.setdefault("catalog_extras", {})
    dividers = extras.setdefault("family_dividers", {})
    old = dividers.pop(family, None)
    save_cfg(cfg)
    _delete_catalog_extra(old.get("stored_as") if old else None)
    return jsonify({"ok": True, "extras": extras})

@app.get("/api/full-catalog/index-order")
def api_full_catalog_index_order():
    """Live structure for the Index Order management card: every section
    (in current Section Order) -> every category (in current Category
    Order) -> its products in current Index display order, INCLUDING
    excluded ones (each tagged excluded:true) so the UI can still show and
    re-include them. Built the same way the actual grid render is, via
    catalog_builder.group_and_order + compute_index_rows, so this card
    never disagrees with what a real Build will show."""
    folder = brand_settings("SOLOLUCE").get("catalogue_folder", "")
    if not folder or not os.path.isdir(folder):
        return jsonify({"error": "Set the Sololuce Datasheets folder first (see Settings)."}), 400
    cfg = load_cfg()
    products, warnings = catalog_builder.gather_products(folder)
    sections, group_warnings = catalog_builder.group_and_order(
        products, cfg.get("cat_series_labels", []), cfg.get("catalog_section_order", list(catalog_builder.SECTION_VALUES)))
    out_sections = []
    for sec in sections:
        out_categories = []
        for cat in sec["categories"]:
            rows = catalog_builder.compute_index_rows(cat["products"], cat["label"], cfg)
            out_categories.append({
                "label": cat["label"],
                "products": [{"product_name": p["product_name"], "main_photo": p["main_photo"],
                               "excluded": p["excluded"]} for p in rows],
            })
        out_sections.append({"label": sec["label"], "categories": out_categories})
    return jsonify({"sections": out_sections, "warnings": warnings + group_warnings})

@app.post("/api/full-catalog/index-order-move")
def api_full_catalog_index_order_move():
    """Swap a product with its up/down neighbor within its category's
    Index order — same materialize-then-swap-with-neighbor pattern as
    /api/cat-series-move, except the "current order" has to be computed
    fresh from real generated products first (catalog_index_order may not
    have an entry for this category yet)."""
    data = request.json or {}
    category = (data.get("category") or "").strip()
    product_name = (data.get("product_name") or "").strip()
    direction = data.get("direction", "")
    folder = brand_settings("SOLOLUCE").get("catalogue_folder", "")
    if not folder or not os.path.isdir(folder):
        return jsonify({"error": "Set the Sololuce Datasheets folder first (see Settings)."}), 400
    cfg = load_cfg()
    products, _ = catalog_builder.gather_products(folder)
    cat_products = sorted((p for p in products if (p["category"] or catalog_builder.UNCATEGORIZED_CATEGORY) == category),
                           key=lambda p: p["product_name"].strip().lower())
    rows = catalog_builder.compute_index_rows(cat_products, category, cfg)
    names = [p["product_name"] for p in rows]
    if product_name not in names:
        return jsonify({"error": "Unknown product."}), 400
    i = names.index(product_name)
    j = i - 1 if direction == "up" else i + 1 if direction == "down" else None
    if j is not None and 0 <= j < len(names):
        names[i], names[j] = names[j], names[i]
        cfg.setdefault("catalog_index_order", {})[category] = names
        save_cfg(cfg)
    return jsonify({"order": names})

@app.post("/api/full-catalog/index-exclude")
def api_full_catalog_index_exclude():
    """Toggle a product's membership in catalog_index_excluded — hides/
    shows it in the Index grid only, its own datasheet page is untouched."""
    data = request.json or {}
    product_name = (data.get("product_name") or "").strip()
    excluded = bool(data.get("excluded"))
    if not product_name:
        return jsonify({"error": "Unknown product."}), 400
    cfg = load_cfg()
    items = cfg.setdefault("catalog_index_excluded", [])
    if excluded and product_name not in items:
        items.append(product_name)
    elif not excluded and product_name in items:
        items.remove(product_name)
    save_cfg(cfg)
    return jsonify({"excluded": items})

@app.get("/api/full-catalog/last-build")
def api_full_catalog_last_build():
    return jsonify(load_cfg().get("catalog_last_build", {}))

@app.get("/api/cat-ordering-categories")
def api_cat_ordering_categories():
    return jsonify({"categories": load_cfg().get("cat_ordering_categories", [])})

@app.post("/api/cat-ordering-categories-add")
def api_cat_ordering_categories_add():
    label = (request.json or {}).get("label", "").strip()
    cfg = load_cfg()
    categories = cfg.get("cat_ordering_categories", [])
    if label and label not in categories:
        categories.append(label)
        cfg["cat_ordering_categories"] = categories
        _log_change(cfg, "add", "cat_ordering_categories", MANAGED_STRING_LISTS["cat_ordering_categories"], label)
        save_cfg(cfg)
    return jsonify({"categories": categories})

@app.get("/api/cat-cct-options")
def api_cat_cct_options():
    return jsonify({"options": load_cfg().get("cat_cct_options", [])})

@app.post("/api/cat-cct-options-add")
def api_cat_cct_options_add():
    value = (request.json or {}).get("value", "").strip()
    cfg = load_cfg()
    options = cfg.get("cat_cct_options", [])
    if value and value not in options:
        options.append(value)
        cfg["cat_cct_options"] = options
        _log_change(cfg, "add", "cat_cct_options", MANAGED_STRING_LISTS["cat_cct_options"], value)
        save_cfg(cfg)
    return jsonify({"options": options})

@app.get("/api/cat-controls-options")
def api_cat_controls_options():
    return jsonify({"options": load_cfg().get("cat_controls_options", [])})

@app.post("/api/cat-controls-options-add")
def api_cat_controls_options_add():
    value = (request.json or {}).get("value", "").strip()
    cfg = load_cfg()
    options = cfg.get("cat_controls_options", [])
    if value and value not in options:
        options.append(value)
        cfg["cat_controls_options"] = options
        _log_change(cfg, "add", "cat_controls_options", MANAGED_STRING_LISTS["cat_controls_options"], value)
        save_cfg(cfg)
    return jsonify({"options": options})

@app.get("/api/expense-payment-methods")
def api_expense_payment_methods():
    return jsonify({"options": load_cfg().get("expense_payment_methods", [])})

@app.post("/api/expense-payment-methods-add")
def api_expense_payment_methods_add():
    value = (request.json or {}).get("value", "").strip()
    cfg = load_cfg()
    options = cfg.get("expense_payment_methods", [])
    if value and value not in options:
        options.append(value)
        cfg["expense_payment_methods"] = options
        _log_change(cfg, "add", "expense_payment_methods", MANAGED_STRING_LISTS["expense_payment_methods"], value)
        save_cfg(cfg)
    return jsonify({"options": options})

@app.get("/api/expense-employees")
def api_expense_employees():
    return jsonify({"options": load_cfg().get("expense_employees", [])})

@app.post("/api/expense-employees-add")
def api_expense_employees_add():
    value = (request.json or {}).get("value", "").strip()
    cfg = load_cfg()
    options = cfg.get("expense_employees", [])
    if value and value not in options:
        options.append(value)
        cfg["expense_employees"] = options
        _log_change(cfg, "add", "expense_employees", MANAGED_STRING_LISTS["expense_employees"], value)
        save_cfg(cfg)
    return jsonify({"options": options})

@app.get("/api/expense-categories")
def api_expense_categories():
    return jsonify({"options": load_cfg().get("expense_categories", [])})

@app.post("/api/expense-categories-add")
def api_expense_categories_add():
    value = (request.json or {}).get("value", "").strip()
    cfg = load_cfg()
    options = cfg.get("expense_categories", [])
    if value and value not in options:
        options.append(value)
        cfg["expense_categories"] = options
        _log_change(cfg, "add", "expense_categories", MANAGED_STRING_LISTS["expense_categories"], value)
        save_cfg(cfg)
    return jsonify({"options": options})

@app.get("/api/expense-products")
def api_expense_products():
    return jsonify({"options": load_cfg().get("expense_products", [])})

@app.post("/api/expense-products-add")
def api_expense_products_add():
    value = (request.json or {}).get("value", "").strip()
    cfg = load_cfg()
    options = cfg.get("expense_products", [])
    if value and value not in options:
        options.append(value)
        cfg["expense_products"] = options
        _log_change(cfg, "add", "expense_products", MANAGED_STRING_LISTS["expense_products"], value)
        save_cfg(cfg)
    return jsonify({"options": options})

@app.get("/api/expense-descriptions")
def api_expense_descriptions():
    return jsonify({"options": load_cfg().get("expense_descriptions", [])})

@app.post("/api/expense-descriptions-add")
def api_expense_descriptions_add():
    value = (request.json or {}).get("value", "").strip()
    cfg = load_cfg()
    options = cfg.get("expense_descriptions", [])
    if value and value not in options:
        options.append(value)
        cfg["expense_descriptions"] = options
        _log_change(cfg, "add", "expense_descriptions", MANAGED_STRING_LISTS["expense_descriptions"], value)
        save_cfg(cfg)
    return jsonify({"options": options})

@app.get("/api/cat-voltage-options")
def api_cat_voltage_options():
    return jsonify({"options": load_cfg().get("cat_voltage_options", [])})

@app.post("/api/cat-voltage-options-add")
def api_cat_voltage_options_add():
    value = (request.json or {}).get("value", "").strip()
    cfg = load_cfg()
    options = cfg.get("cat_voltage_options", [])
    if value and value not in options:
        options.append(value)
        cfg["cat_voltage_options"] = options
        _log_change(cfg, "add", "cat_voltage_options", MANAGED_STRING_LISTS["cat_voltage_options"], value)
        save_cfg(cfg)
    return jsonify({"options": options})

@app.get("/api/cat-power-options")
def api_cat_power_options():
    return jsonify({"options": load_cfg().get("cat_power_options", [])})

@app.post("/api/cat-power-options-add")
def api_cat_power_options_add():
    value = (request.json or {}).get("value", "").strip()
    cfg = load_cfg()
    options = cfg.get("cat_power_options", [])
    if value and value not in options:
        options.append(value)
        # Wattages outside the 1-40W baseline are "custom" and capped at 5
        # remembered at once — the oldest one is dropped once a 6th is
        # added, so the dropdown doesn't grow unbounded from one-off/unusual
        # wattages over years of use. The 1-40W baseline itself is never
        # touched by this cap.
        customs = [o for o in options if _norm_label_key(o) not in CAT_POWER_BASELINE_NORM]
        if len(customs) > 5:
            options = [o for o in options if o != customs[0]]
        cfg["cat_power_options"] = options
        _log_change(cfg, "add", "cat_power_options", MANAGED_STRING_LISTS["cat_power_options"], value)
        save_cfg(cfg)
    return jsonify({"options": options})

@app.get("/api/cat-beamangle-options")
def api_cat_beamangle_options():
    return jsonify({"options": load_cfg().get("cat_beamangle_options", [])})

@app.post("/api/cat-beamangle-options-add")
def api_cat_beamangle_options_add():
    value = (request.json or {}).get("value", "").strip()
    cfg = load_cfg()
    options = cfg.get("cat_beamangle_options", [])
    if value and value not in options:
        options.append(value)
        cfg["cat_beamangle_options"] = options
        _log_change(cfg, "add", "cat_beamangle_options", MANAGED_STRING_LISTS["cat_beamangle_options"], value)
        save_cfg(cfg)
    return jsonify({"options": options})

@app.get("/api/cat-size-index")
def api_cat_size_index():
    return jsonify({"sizes": load_cfg().get("cat_size_index", [])})

@app.post("/api/cat-size-index-add")
def api_cat_size_index_add():
    value = (request.json or {}).get("value", "").strip()
    cfg = load_cfg()
    sizes = cfg.get("cat_size_index", [])
    if value and value not in sizes:
        sizes.append(value)
        cfg["cat_size_index"] = sizes
        _log_change(cfg, "add", "cat_size_index", MANAGED_STRING_LISTS["cat_size_index"], value)
        save_cfg(cfg)
    return jsonify({"sizes": sizes})

@app.get("/api/cat-modelno-options")
def api_cat_modelno_options():
    return jsonify({"options": load_cfg().get("cat_modelno_options", [])})

@app.post("/api/cat-modelno-options-add")
def api_cat_modelno_options_add():
    value = (request.json or {}).get("value", "").strip()
    cfg = load_cfg()
    options = cfg.get("cat_modelno_options", [])
    if value and value not in options:
        options.append(value)
        cfg["cat_modelno_options"] = options
        _log_change(cfg, "add", "cat_modelno_options", MANAGED_STRING_LISTS["cat_modelno_options"], value)
        save_cfg(cfg)
    return jsonify({"options": options})

@app.get("/api/cat-cutout-options")
def api_cat_cutout_options():
    return jsonify({"options": load_cfg().get("cat_cutout_options", [])})

@app.post("/api/cat-cutout-options-add")
def api_cat_cutout_options_add():
    value = (request.json or {}).get("value", "").strip()
    cfg = load_cfg()
    options = cfg.get("cat_cutout_options", [])
    if value and value not in options:
        options.append(value)
        cfg["cat_cutout_options"] = options
        _log_change(cfg, "add", "cat_cutout_options", MANAGED_STRING_LISTS["cat_cutout_options"], value)
        save_cfg(cfg)
    return jsonify({"options": options})

# The Ordering Table's "Fill Standard Information" button — which of these
# known dropdown-backed field types it fills (every one of that field's own
# remembered preset options, one per row) is user-editable, not fixed —
# see cat_standard_fill_fields's own comment in load_cfg(). Kept as a
# server-side allowlist (not whatever the client sends) so a stray/renamed
# key can never silently wedge itself into the saved list.
CAT_STANDARD_FILL_KEYS = ("cct", "controls", "voltage", "power", "beamangle", "modelno", "options", "cutout", "size")

def _cat_standard_fill_order(cfg):
    """cat_standard_fill_order is a full permutation of CAT_STANDARD_FILL_KEYS
    — falls back to (and self-heals) the built-in order if the saved list is
    missing a key (an allowlist entry added after the user already saved a
    custom order) or carries a since-removed one."""
    order = cfg.get("cat_standard_fill_order") or list(CAT_STANDARD_FILL_KEYS)
    order = [k for k in order if k in CAT_STANDARD_FILL_KEYS]
    order += [k for k in CAT_STANDARD_FILL_KEYS if k not in order]
    return order

@app.get("/api/cat-standard-fill-fields")
def api_cat_standard_fill_fields():
    cfg = load_cfg()
    return jsonify({
        "fields": cfg.get("cat_standard_fill_fields", ["cct", "controls"]),
        "available": CAT_STANDARD_FILL_KEYS,
        "order": _cat_standard_fill_order(cfg),
    })

@app.post("/api/cat-standard-fill-fields")
def api_cat_standard_fill_fields_set():
    fields = (request.json or {}).get("fields", [])
    if not isinstance(fields, list):
        return jsonify({"error": "fields must be a list."}), 400
    cfg = load_cfg()
    # Canonicalized against the user's OWN saved display order (not the
    # hardcoded CAT_STANDARD_FILL_KEYS tuple) — otherwise every save would
    # silently snap the selected fields back to the built-in sequence,
    # which is exactly what made this unreorderable before.
    order = _cat_standard_fill_order(cfg)
    cleaned = [f for f in order if f in fields]  # also dedupes
    cfg["cat_standard_fill_fields"] = cleaned
    save_cfg(cfg)
    return jsonify({"fields": cleaned})

@app.post("/api/cat-standard-fill-order-move")
def api_cat_standard_fill_order_move():
    """Swap a field with its up/down neighbor in cat_standard_fill_order —
    same swap-with-neighbor pattern as /api/cat-series-move. Reordering here
    reorders both the Configure popover's checklist and the sequence
    fillCatOrdStandardInfo() processes the selected fields in."""
    data = request.json or {}
    key, direction = (data.get("key") or "").strip(), data.get("direction", "")
    if key not in CAT_STANDARD_FILL_KEYS:
        return jsonify({"error": "Unknown field."}), 400
    cfg = load_cfg()
    order = _cat_standard_fill_order(cfg)
    i = order.index(key)
    j = i - 1 if direction == "up" else i + 1 if direction == "down" else None
    if j is not None and 0 <= j < len(order):
        order[i], order[j] = order[j], order[i]
        cfg["cat_standard_fill_order"] = order
        save_cfg(cfg)
    return jsonify({"order": order})

# One level deeper than the two routes above: not just which FIELDS fill and
# in what order, but which of a given field's own saved preset VALUES
# actually get used, and in what sequence — e.g. only 3 of 7 saved CCTs,
# Non-Dimmable first. A field absent from cat_standard_fill_values means
# "use its full current preset list" (the original, still-default
# behavior) — see load_cfg()'s own comment on this key.
@app.get("/api/cat-standard-fill-values")
def api_cat_standard_fill_values():
    return jsonify({"values": load_cfg().get("cat_standard_fill_values", {})})

@app.post("/api/cat-standard-fill-values")
def api_cat_standard_fill_values_set():
    """Replaces one field's entire chosen-values list in one shot (the
    picker UI always has the full candidate list in hand client-side, same
    "resend the whole cleaned list" shape as /api/cat-standard-fill-fields).
    An empty list is stored as-is (so the picker's checkbox states round-
    trip faithfully — unchecking everything reloads as everything
    unchecked, not silently re-checked) but the frontend's own effective-
    options lookup (catStandardFillEffectiveOptions) treats an empty list
    the same as "untouched" and falls back to the full preset list —
    "skip this field's fill entirely" already has its own dedicated
    control (the field's own checkbox in the Fill Standard Information Uses
    popover), so an empty subset here isn't given that same meaning too."""
    data = request.json or {}
    key, values = (data.get("key") or "").strip(), data.get("values", [])
    if key not in CAT_STANDARD_FILL_KEYS or not isinstance(values, list):
        return jsonify({"error": "Invalid field or values."}), 400
    cleaned, seen = [], set()
    for v in values:
        v = str(v)
        if v not in seen:
            seen.add(v)
            cleaned.append(v)
    cfg = load_cfg()
    fv = cfg.get("cat_standard_fill_values", {})
    fv[key] = cleaned
    cfg["cat_standard_fill_values"] = fv
    save_cfg(cfg)
    return jsonify({"values": cleaned})

@app.post("/api/cat-standard-fill-values-move")
def api_cat_standard_fill_values_move():
    """Swap-with-neighbor within one field's OWN chosen-values list — index-
    based (like /api/settings-list-move) rather than value-keyed, since a
    field's chosen list isn't a flat global registry entry the way
    MANAGED_STRING_LISTS' lists are."""
    data = request.json or {}
    key, index, direction = (data.get("key") or "").strip(), data.get("index"), data.get("direction", "")
    if key not in CAT_STANDARD_FILL_KEYS or not isinstance(index, int):
        return jsonify({"error": "Invalid field or index."}), 400
    cfg = load_cfg()
    fv = cfg.get("cat_standard_fill_values", {})
    values = fv.get(key, [])
    if index < 0 or index >= len(values):
        return jsonify({"error": "Value not found."}), 400
    j = index - 1 if direction == "up" else index + 1 if direction == "down" else None
    if j is not None and 0 <= j < len(values):
        values[index], values[j] = values[j], values[index]
        fv[key] = values
        cfg["cat_standard_fill_values"] = fv
        save_cfg(cfg)
    return jsonify({"values": values})

@app.post("/api/cat-standard-fill-values-reset")
def api_cat_standard_fill_values_reset():
    """The one explicit way back to "use the full preset list" — removes the
    field's key entirely rather than storing an empty list, which (see the
    POST route above) deliberately means something different."""
    data = request.json or {}
    key = (data.get("key") or "").strip()
    if key not in CAT_STANDARD_FILL_KEYS:
        return jsonify({"error": "Unknown field."}), 400
    cfg = load_cfg()
    fv = cfg.get("cat_standard_fill_values", {})
    fv.pop(key, None)
    cfg["cat_standard_fill_values"] = fv
    save_cfg(cfg)
    return jsonify({"ok": True})

@app.get("/api/cat-options-options")
def api_cat_options_options():
    return jsonify({"options": load_cfg().get("cat_options_options", [])})

@app.post("/api/cat-ordering-widths")
def api_cat_ordering_widths():
    """Read-only: runs the Ordering Table's own weight algorithm
    (html_engine.build_ordering_table — the exact same code the real PDF
    render uses, including the saved "standard" width fallback) against the
    current form's columns and hands back just the label/weight pairs. The
    sidebar's draggable-width widget calls this for any column that doesn't
    have a manual width yet, so its "current auto width" reading can never
    drift from what the PDF actually renders — single source of truth
    instead of a second copy of the algorithm in JS. This is also what
    makes Reset Widths land on the saved standard rather than raw content-
    sizing once one exists: Reset just clears every column's manual width
    and re-renders, reading this same endpoint."""
    cols = (request.json or {}).get("ordering_columns") or []
    if not isinstance(cols, list):
        return jsonify({"error": "ordering_columns must be a list."}), 400
    default_widths = load_cfg().get("cat_ordering_default_widths", {})
    table = html_engine.build_ordering_table(cols, default_weights=default_widths)
    return jsonify({"labels": [l["text"] for l in table["labels"]], "weights": table["col_weights"]})

@app.post("/api/cat-ordering-default-widths")
def api_cat_ordering_default_widths_set():
    """"Save as Standard" in the Column Widths widget — captures whatever
    weight each column is CURRENTLY showing (manual drag or auto-computed,
    the client works this out via catOrdEffectiveWeights() before calling
    this) and remembers it per LABEL, so the same field starts from this
    width on every future datasheet, not just the one it was set on. Fully
    additive: only ever affects a label once someone explicitly saves it —
    see load_cfg()'s own comment on this config key."""
    weights = (request.json or {}).get("weights")
    if not isinstance(weights, dict):
        return jsonify({"error": "weights must be a {label: weight} object."}), 400
    cleaned = {}
    for label, w in weights.items():
        label = (label or "").strip()
        if label and isinstance(w, (int, float)) and w > 0:
            cleaned[label] = w
    cfg = load_cfg()
    dw = cfg.get("cat_ordering_default_widths", {})
    dw.update(cleaned)
    cfg["cat_ordering_default_widths"] = dw
    save_cfg(cfg)
    return jsonify({"default_widths": dw})

@app.post("/api/cat-options-options-add")
def api_cat_options_options_add():
    value = (request.json or {}).get("value", "").strip()
    cfg = load_cfg()
    options = cfg.get("cat_options_options", [])
    if value and value not in options:
        options.append(value)
        cfg["cat_options_options"] = options
        _log_change(cfg, "add", "cat_options_options", MANAGED_STRING_LISTS["cat_options_options"], value)
        save_cfg(cfg)
    return jsonify({"options": options})

@app.get("/api/cat-spec-values")
def api_cat_spec_values():
    return jsonify({"values": load_cfg().get("cat_spec_values", {})})

@app.post("/api/cat-spec-values-add")
def api_cat_spec_values_add():
    data = request.json or {}
    label = (data.get("label") or "").strip()
    value = (data.get("value") or "").strip()
    key = _norm_label_key(label)
    if not key or not value:
        return jsonify({"error": "Missing label or value."}), 400
    cfg = load_cfg()
    store = cfg.setdefault("cat_spec_values", {})
    values = store.get(key, [])
    if value not in values:
        values.append(value)
        store[key] = values
        _log_change(cfg, "add", "cat_spec_values:" + key, "Technical Spec Values — " + label, value)
        save_cfg(cfg)
    return jsonify({"key": key, "values": values})

@app.get("/api/cat-finish-colors")
def api_cat_finish_colors():
    return jsonify({"colors": load_cfg().get("cat_finish_colors", [])})

@app.post("/api/cat-finish-colors-add")
def api_cat_finish_colors_add():
    data = request.json or {}
    label = (data.get("label") or "").strip()
    hexval = (data.get("hex") or "#ffffff").strip()
    if not label:
        return jsonify({"error": "Enter a color name."}), 400
    cfg = load_cfg()
    colors = cfg.get("cat_finish_colors", [])
    if not any(c.get("label", "").strip().lower() == label.lower() for c in colors):
        colors.append({"label": label, "hex": hexval})
        cfg["cat_finish_colors"] = colors
        _log_change(cfg, "add", "cat_finish_colors", "Finish Colors", label)
        save_cfg(cfg)
    return jsonify({"colors": colors})

@app.post("/api/cat-finish-colors-remove")
def api_cat_finish_colors_remove():
    label = (request.json or {}).get("label", "").strip()
    cfg = load_cfg()
    colors = [c for c in cfg.get("cat_finish_colors", []) if c.get("label") != label]
    cfg["cat_finish_colors"] = colors
    _log_change(cfg, "remove", "cat_finish_colors", "Finish Colors", label)
    save_cfg(cfg)
    return jsonify({"colors": colors})

CAT_BADGES_DIR = os.path.join(engine.DATA_BASE, "static", "cat_badges")

@app.get("/api/cat-badges")
def api_cat_badges():
    return jsonify({"badges": load_cfg().get("cat_badge_library", [])})

@app.post("/api/cat-badges-add")
def api_cat_badges_add():
    """Lets the user grow the badge library themselves — upload any icon
    image plus a name, and it shows up as its own checkbox tile in the Spec
    Badges picker from then on, exactly like the original 39. No per-product
    value overlay to worry about (unlike the app's earlier hand-drawn-icon
    badge system) since the uploaded image is used exactly as given."""
    data = request.json or {}
    label = (data.get("label") or "").strip()
    image_data_url = data.get("image") or ""
    if not label:
        return jsonify({"error": "Enter a name for this badge."}), 400
    if "," not in image_data_url:
        return jsonify({"error": "Choose an image."}), 400
    header, b64 = image_data_url.split(",", 1)
    ext_match = re.search(r"image/(\w+)", header)
    ext = (ext_match.group(1) if ext_match else "png").lower()
    if ext == "jpeg":
        ext = "jpg"
    cfg = load_cfg()
    library = cfg.get("cat_badge_library", [])
    key = engine._slug(label).lower() or "badge"
    base_key = key
    existing_keys = {b.get("key") for b in library}
    i = 2
    while key in existing_keys:
        key = f"{base_key}-{i}"; i += 1
    filename = f"{key}.{ext}"
    os.makedirs(CAT_BADGES_DIR, exist_ok=True)
    try:
        with open(os.path.join(CAT_BADGES_DIR, filename), "wb") as f:
            f.write(base64.b64decode(b64))
    except Exception as e:
        return jsonify({"error": f"Couldn't save that image: {e}"}), 400
    library.append({"key": key, "label": label, "filename": filename})
    cfg["cat_badge_library"] = library
    save_cfg(cfg)
    return jsonify({"badges": library})

@app.post("/api/add-unit")
def api_add_unit():
    unit = (request.json or {}).get("unit", "").strip()
    if not unit:
        return jsonify({"error": "Enter a unit."}), 400
    cfg = load_cfg()
    units = cfg.get("units", [])
    if unit not in units:
        units.append(unit)
        cfg["units"] = units
        _log_change(cfg, "add", "units", MANAGED_STRING_LISTS["units"], unit)
        save_cfg(cfg)
    return jsonify({"units": units})

# Auto-update: checkForAppUpdate() in the page script calls this on launch
# (and periodically while the app stays open) to poll GitHub Releases; see
# update_checker.py for the one-time repo setup this needs.
@app.get("/api/check-update")
def api_check_update():
    return jsonify(update_checker.check_for_update())

# Update Center prefs (Settings top bar, everyone — not admin-only) —
# whether the app checks for updates automatically on launch, and whether
# a found update installs itself without the usual double-click confirm.
# Defaults match the app's original behavior (always checks, never
# auto-installs) so nothing changes for anyone until they opt in.
@app.get("/api/update-prefs")
def api_update_prefs():
    cfg = load_cfg()
    prefs = cfg.get("update_prefs") or {}
    return jsonify({"check_on_start": prefs.get("check_on_start", True), "auto_update": prefs.get("auto_update", False)})

@app.post("/api/update-prefs")
def api_update_prefs_save():
    data = request.json or {}
    cfg = load_cfg()
    cfg["update_prefs"] = {"check_on_start": bool(data.get("check_on_start", True)), "auto_update": bool(data.get("auto_update", False))}
    save_cfg(cfg)
    return jsonify({"ok": True})

# Downloads the newer installer and launches it, then this process exits
# itself shortly after (so the installer isn't stuck trying to close a
# still-running instance of the app it's about to replace). The installer
# only ever touches the [Files] it lists — config.json/drafts/submissions
# survive, same as any manual update — see installer.iss's own comment.
@app.post("/api/apply-update")
def api_apply_update():
    data = request.json or {}
    url = data.get("download_url", "")
    if not url:
        return jsonify({"ok": False, "error": "Missing download_url."}), 400
    try:
        update_checker.download_and_launch_installer(url)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

    def _exit_soon():
        time.sleep(1.0)
        os._exit(0)
    threading.Thread(target=_exit_soon, daemon=True).start()
    return jsonify({"ok": True})

# Generic read/delete pair the Settings "Manage Lists" UI drives, covering
# every list in MANAGED_STRING_LISTS uniformly — one route instead of a
# one-off GET+remove per list. Deletion is by index (not value) so it still
# works correctly if a list ever has two identical-looking entries.
@app.get("/api/settings-list/<list_key>")
def api_settings_list_get(list_key):
    if list_key not in MANAGED_STRING_LISTS:
        return jsonify({"error": "Unknown list."}), 404
    return jsonify({"items": load_cfg().get(list_key, [])})

@app.post("/api/settings-list-remove")
def api_settings_list_remove():
    data = request.json or {}
    list_key = data.get("list_key", "")
    index = data.get("index")
    if list_key not in MANAGED_STRING_LISTS or not isinstance(index, int):
        return jsonify({"error": "Invalid list or index."}), 400
    cfg = load_cfg()
    items = cfg.get(list_key, [])
    if index < 0 or index >= len(items):
        return jsonify({"error": "Item not found."}), 400
    removed = items.pop(index)
    cfg[list_key] = items
    _log_change(cfg, "remove", list_key, MANAGED_STRING_LISTS[list_key], removed)
    save_cfg(cfg)
    return jsonify({"items": items})

# Generic swap-with-neighbor reorder for the same registry — every managed
# list is draggable in the Settings UI (not just delete-able), same
# swap-with-neighbor pattern as /api/cat-series-move and every other Order
# card in this app. Not routed through _log_change: that log is for content
# added/removed, not display-order changes, matching /api/cat-series-move's
# own precedent. cat_size_index used to be excluded here — its position
# once WAS every size's printed "D" number, catalogue-wide, so reordering
# it would have silently repointed already-generated datasheets at the
# wrong dimension. That's no longer true (see cat_size_index's own comment
# a few lines up, and recomputeCatOrdSizeDNumbers on the front end): it's
# a plain suggestion list now, safe to reorder like any other preset list.

@app.post("/api/settings-list-move")
def api_settings_list_move():
    data = request.json or {}
    list_key, index, direction = data.get("list_key", ""), data.get("index"), data.get("direction", "")
    if list_key not in MANAGED_STRING_LISTS or not isinstance(index, int):
        return jsonify({"error": "Invalid list or index."}), 400
    cfg = load_cfg()
    items = cfg.get(list_key, [])
    if index < 0 or index >= len(items):
        return jsonify({"error": "Item not found."}), 400
    j = index - 1 if direction == "up" else index + 1 if direction == "down" else None
    if j is not None and 0 <= j < len(items):
        items[index], items[j] = items[j], items[index]
        cfg[list_key] = items
        save_cfg(cfg)
    return jsonify({"items": items})

@app.get("/api/audit-log")
def api_audit_log():
    log = load_cfg().get("audit_log", [])
    return jsonify({"log": list(reversed(log))})

@app.get("/api/match-photo")
def api_match_photo():
    """Look up a product photo for a line item by matching its description
    text against the product photo catalog folder."""
    desc = request.args.get("desc", "")
    folder = brand_settings().get("product_photos_folder", "")
    photo = engine.match_product_photo(desc, folder)
    return jsonify({"photo": photo})

# ---------------------------------------------------------------- Shared Product Photos (Cloudflare R2)
# See photo_store.py's own top-of-file comment for the full picture. Every
# call here operates against the CURRENTLY ACTIVE brand's product_photos_folder
# (brand_settings()) — same folder /api/match-photo already reads, so a
# synced-down photo is picked up with zero other changes.
@app.get("/api/photostore-config")
def api_photostore_config():
    return jsonify(photo_store.get_public_config())

@app.post("/api/photostore-config")
def api_photostore_config_save():
    """Saves the one Cloudflare key both places: this machine's own
    config.json (used directly here) AND the bundled r2_readonly.json
    (ships to everyone else on the next rebuild+GUPDATE) — see
    photo_store.py's own comment on why there's no separate read-only
    tier anymore. The bundle half is best-effort: it only ever succeeds
    from the admin's own dev checkout (a frozen .exe can't write into
    itself), so a failure there is reported back but doesn't fail the
    whole save — the local (this-machine) key is still saved either way."""
    data = request.json or {}
    photo_store.save_config(data.get("account_id", ""), data.get("access_key_id", ""),
                             data.get("secret_access_key", ""), data.get("bucket", ""))
    bundled = True
    bundle_error = None
    try:
        # Mirror the just-saved config's OWN resolved values (not the raw
        # request body) — guarantees the bundle always matches exactly,
        # secret included, even when the submitted secret was blank
        # ("keep existing") and the bundle file itself had never been
        # seeded with a real one yet (see get_resolved_admin_config()).
        resolved = photo_store.get_resolved_admin_config()
        photo_store.save_readonly_config(resolved.get("account_id", ""), resolved.get("access_key_id", ""),
                                          resolved.get("secret_access_key", ""), resolved.get("bucket", ""))
    except Exception as e:
        bundled = False
        bundle_error = str(e)
    return jsonify({"ok": True, "bundled": bundled, "bundle_error": bundle_error})

@app.get("/api/photostore-status")
def api_photostore_status():
    if not photo_store.is_configured():
        return jsonify({"configured": False})
    try:
        usage = photo_store.get_usage()
        return jsonify({"configured": True, **usage})
    except Exception as e:
        return jsonify({"configured": True, "error": str(e)}), 502

@app.get("/api/photostore-list")
def api_photostore_list():
    try:
        return jsonify({"photos": photo_store.list_photos()})
    except Exception as e:
        return jsonify({"error": str(e)}), 502

@app.get("/api/photostore-fetch")
def api_photostore_fetch():
    """Streams one photo's bytes straight from R2 — used both as the
    gallery grid's <img src> (thumbnails, browser-cached) and, on click,
    the source the picker draws to a canvas to produce the same kind of
    data: URI a local file upload would (see pickCatImage() vs
    openCloudPhotoPicker() in the page script — CAT_IMG[slot].src is
    always a self-contained data URI either way, zero special-casing
    needed anywhere downstream in PDF/xlsx generation)."""
    key = request.args.get("key", "")
    if not key:
        return jsonify({"error": "Missing key."}), 400
    try:
        data, content_type = photo_store.get_photo_bytes(key)
    except Exception as e:
        return jsonify({"error": str(e)}), 502
    resp = Response(data, mimetype=content_type)
    resp.headers["Cache-Control"] = "private, max-age=3600"
    return resp

@app.post("/api/photostore-upload")
def api_photostore_upload():
    """Accepts either loose files or a whole folder (the frontend sends
    each File's webkitRelativePath as its filename when uploading a
    folder — see uploadPhotosToStore()) — so a key here may contain "/"
    for the folder structure, which R2/S3 stores natively as a prefix."""
    files = request.files.getlist("files")
    if not files:
        return jsonify({"ok": False, "error": "No files."}), 400
    try:
        running_usage = photo_store.get_usage()  # one listing for this whole batch, not one per file
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502
    uploaded, errors = [], []
    for f in files:
        raw_name = (f.filename or "").replace("\\", "/")
        # Sanitize: this becomes both an R2 key AND a path under
        # DATA_BASE (the temp file below) — strip any leading slashes and
        # drop ".."/"." segments so it can never escape either.
        parts = [p for p in raw_name.split("/") if p not in ("", ".", "..")]
        if not parts:
            continue
        key = "/".join(parts)
        tmp_path = os.path.join(engine.DATA_BASE, "_photostore_tmp_" + uuid.uuid4().hex + "_" + parts[-1])
        try:
            f.save(tmp_path)
            photo_store.upload_photo(tmp_path, key, running_usage=running_usage)
            uploaded.append(key)
        except Exception as e:
            errors.append(key + ": " + str(e))
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
    return jsonify({"ok": not errors, "uploaded": uploaded, "errors": errors})

@app.post("/api/photostore-delete")
def api_photostore_delete():
    data = request.json or {}
    name = data.get("filename", "")
    if not name:
        return jsonify({"ok": False, "error": "Missing filename."}), 400
    try:
        photo_store.delete_photo(name)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502

@app.post("/api/photostore-sync")
def api_photostore_sync():
    folder = brand_settings().get("product_photos_folder", "")
    if not folder:
        return jsonify({"ok": False, "error": "Set a Product Pictures folder in Settings first."}), 400
    return jsonify(photo_store.sync_down(folder))

@app.get("/api/match-datasheets")
def api_match_datasheets():
    """Find every datasheet matching a line item's description text — a
    product can have more than one (e.g. a plain and a motion-sensor variant)."""
    desc = request.args.get("desc", "")
    folder = brand_settings().get("datasheets_folder", "")
    matches = engine.match_datasheets(desc, folder)
    out = [{"name": m["name"], "rel": os.path.relpath(m["path"], folder).replace(os.sep, "/")} for m in matches]
    return jsonify({"datasheets": out})

@app.get("/api/product-list")
def api_product_list():
    """Every product in the datasheet catalog, for the Build tab's "Find
    Product" finder."""
    folder = brand_settings().get("datasheets_folder", "")
    if not folder or not os.path.isdir(folder):
        return jsonify({"products": []})
    products = engine.list_datasheet_products(folder)
    out = [{"name": p["name"], "rel": os.path.relpath(p["path"], folder).replace(os.sep, "/")} for p in products]
    return jsonify({"products": out})

def _resolve_datasheet_pdf(rel):
    """Resolve a datasheet 'rel' path to a servable PDF, converting a true
    (non-PDF-compatible) .ai via LibreOffice if needed. Returns (pdf_path,
    None) on success, or (None, (body, status)) to return straight to Flask."""
    folder = brand_settings().get("datasheets_folder", "")
    path = os.path.join(folder, rel)
    if (not folder or not os.path.abspath(path).startswith(os.path.abspath(folder))
            or not os.path.exists(path) or not path.lower().endswith((".pdf", ".ai"))):
        return None, ("Not found", 404)
    if path.lower().endswith(".pdf"):
        return path, None
    with open(path, "rb") as f:
        head = f.read(5)
    if head == b"%PDF-":
        return path, None
    os.makedirs(CS_CACHE, exist_ok=True)
    tag = f"ai_{abs(hash(os.path.abspath(path)))}_{int(os.path.getmtime(path))}.pdf"
    cache_path = os.path.join(CS_CACHE, tag)
    if not os.path.exists(cache_path):
        try:
            tmp_pdf = engine.to_pdf(path, CS_CACHE)
            if os.path.abspath(tmp_pdf) != os.path.abspath(cache_path):
                shutil.move(tmp_pdf, cache_path)
        except Exception:
            return None, ("This datasheet is a native Illustrator file without a PDF "
                           "inside it, and it couldn't be converted automatically. "
                           "Open it directly in Illustrator instead.", 415)
    return cache_path, None

@app.get("/datasheet")
def get_datasheet():
    """Serve a datasheet. Illustrator files saved with 'PDF Compatibility'
    (Illustrator's default) are valid PDFs under the hood, so those are
    served as-is. A true non-PDF-compatible .ai gets converted via
    LibreOffice as a fallback; if even that fails, we say so plainly rather
    than serving garbage."""
    pdf_path, err = _resolve_datasheet_pdf(request.args.get("rel", ""))
    if err:
        return err
    return send_file(pdf_path, mimetype="application/pdf")

@app.get("/datasheet-thumb")
def datasheet_thumb():
    """Render a first-page PNG of a datasheet for the in-app Company System
    modal — same treatment as /cs-thumb for documents, so opening a
    datasheet never navigates away from (and loses) the in-progress form."""
    rel = request.args.get("rel", "")
    pdf_path, err = _resolve_datasheet_pdf(rel)
    if err:
        return err
    os.makedirs(CS_CACHE, exist_ok=True)
    tag = f"ds_{abs(hash(os.path.abspath(pdf_path)))}_{int(os.path.getmtime(pdf_path))}.png"
    cache_path = os.path.join(CS_CACHE, tag)
    if not os.path.exists(cache_path):
        try:
            tmp_png = engine.to_png(pdf_path, CS_CACHE)
            if os.path.abspath(tmp_png) != os.path.abspath(cache_path):
                shutil.move(tmp_png, cache_path)
        except Exception as e:
            return f"Couldn't render a preview for this datasheet: {e}", 500
    return send_file(cache_path, mimetype="image/png")

@app.get("/api/product-options")
def api_product_options():
    """Product Builder: parse a matched datasheet's spec/ordering-code
    tables into pickable option lists (Wattage/CCT/Beam Angle/Controls/
    Color/Size) — see engine.extract_product_options for how."""
    pdf_path, err = _resolve_datasheet_pdf(request.args.get("rel", ""))
    if err:
        return err
    return jsonify(engine.extract_product_options(pdf_path))

@app.post("/api/remove-unit")
def api_remove_unit():
    unit = (request.json or {}).get("unit", "").strip()
    cfg = load_cfg()
    units = cfg.get("units", [])
    if unit in units:
        units.remove(unit)
        cfg["units"] = units
        save_cfg(cfg)
    return jsonify({"units": units})

# ---------------------------------------------------------------- clients
@app.get("/api/clients")
def api_clients():
    return jsonify({"clients": load_clients()})

@app.post("/api/clients")
def api_save_client():
    data = request.json or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Enter a client name."}), 400
    brand = current_brand()
    clients = load_clients(brand)
    cid = (data.get("id") or "").strip()
    record = {
        "id": cid or f"c_{int(datetime.datetime.now().timestamp() * 1000)}",
        "name": name,
        "category": (data.get("category") or "").strip(),
        "attn": (data.get("attn") or "").strip(),
        "address": (data.get("address") or "").strip(),
        "po_box": (data.get("po_box") or "").strip(),
        "city": (data.get("city") or "").strip(),
        "country": (data.get("country") or "").strip(),
        "phone": (data.get("phone") or "").strip(),
        "landline": (data.get("landline") or "").strip(),
        "email": (data.get("email") or "").strip(),
        "website": (data.get("website") or "").strip(),
        "trn": (data.get("trn") or "").strip(),
        "notes": (data.get("notes") or "").strip(),
        "logo": data.get("logo") or "",
        "updated": datetime.date.today().isoformat(),
    }
    existing = next((c for c in clients if c["id"] == record["id"]), None)
    if existing:
        clients[clients.index(existing)] = record
    else:
        clients.append(record)
    save_clients(brand, clients)
    return jsonify({"client": record})

@app.post("/api/clients-delete")
def api_delete_client():
    cid = (request.json or {}).get("id", "")
    brand = current_brand()
    clients = [c for c in load_clients(brand) if c["id"] != cid]
    save_clients(brand, clients)
    return jsonify({"clients": clients})

@app.get("/api/drafts")
def api_drafts():
    # A Sololuce Datasheet (CAT) draft is always Sololuce's, full stop — same
    # reasoning as /api/generate's own gen_brand hardcode — so it must stay
    # visible here regardless of which brand is currently active, the same
    # way a generated CAT file always shows in All Docs regardless of brand.
    # Without this, switching brands mid-edit (or an autosave firing after a
    # brand switch) could silently strand a real draft where the user would
    # never see it again.
    brand = current_brand()
    drafts = load_drafts(brand)
    if brand != "SOLOLUCE":
        drafts = drafts + [d for d in load_drafts("SOLOLUCE") if d.get("doc_type") == "CAT"]
    return jsonify({"drafts": drafts})

@app.post("/api/drafts")
def api_save_draft():
    body = request.json or {}
    doc_type = (body.get("doc_type") or "").upper()
    if doc_type not in engine.DOC_TYPES:
        return jsonify({"error": "Unknown document type."}), 400
    data = body.get("data") or {}
    company = (data.get("company") or "").strip()
    if not company:
        return jsonify({"error": "Enter the company name before saving a draft."}), 400
    # Locks a CAT draft to Sololuce's own bucket no matter which brand is
    # active at save time — mirrors /api/generate's gen_brand hardcode.
    # Without this, an autosave firing shortly after a brand switch (e.g. the
    # user just picked a different brand while a CAT draft was still open)
    # would silently file it under the wrong brand instead.
    brand = "SOLOLUCE" if doc_type == "CAT" else current_brand()
    drafts = load_drafts(brand)
    did = (body.get("id") or "").strip()
    label = f"{engine.TYPE_LABEL.get(doc_type, doc_type)} — {company}"
    record = {
        "id": did or f"d_{int(datetime.datetime.now().timestamp() * 1000)}",
        "doc_type": doc_type,
        "label": label,
        "company": company,
        "data": data,
        "updated": datetime.datetime.now().isoformat(timespec="seconds"),
    }
    existing = next((d for d in drafts if d["id"] == record["id"]), None)
    if existing:
        drafts[drafts.index(existing)] = record
    else:
        drafts.append(record)
    save_drafts(brand, drafts)
    return jsonify({"draft": record})

@app.post("/api/drafts-delete")
def api_delete_draft():
    did = (request.json or {}).get("id", "")
    brand = current_brand()
    drafts = [d for d in load_drafts(brand) if d["id"] != did]
    save_drafts(brand, drafts)
    # A CAT draft returned by GET /api/drafts while on a different brand (see
    # above) actually lives in Sololuce's own bucket — delete it there too,
    # or its "Delete" button would silently no-op instead of removing it.
    if brand != "SOLOLUCE":
        sololuce_drafts = load_drafts("SOLOLUCE")
        filtered = [d for d in sololuce_drafts if d["id"] != did]
        if len(filtered) != len(sololuce_drafts):
            save_drafts("SOLOLUCE", filtered)
    return jsonify({"drafts": drafts})

_CLIENT_CONTACT_FIELDS = ("attn", "address", "po_box", "city", "country", "phone", "landline", "email", "website", "trn")

@app.post("/api/clients-import")
def api_clients_import():
    """Scan every configured document folder (QTN/INV/DO/QTN2) for the current
    brand: add a client for every distinct company name found that isn't
    already in the client list, and backfill any of attn/address/phone/
    landline/email/website/trn still blank on both new AND pre-existing
    clients (never overwrites a field that already has a value).
    Two sources, tried in priority order per company:
    1. QTN2 sidecars (`engine.read_sidecar`) — exact, structured, but only
       exist for documents made with this app's new Quotation pipeline.
    2. `engine.extract_legacy_contact()` on that company's own QTN/INV/DO
       xlsx files — a heuristic read of the letterhead text boxes real
       historical files were hand-built with (see engine.py for why cell
       reading alone can't see this). Tries up to 5 of that company's most
       recent xlsx files, stopping early once every field is filled."""
    brand = current_brand()
    folders = all_doc_folders()
    companies = set()
    xlsx_by_company = {}
    attn_by_company, addr_by_company = {}, {}
    for folder in folders:
        if not os.path.isdir(folder):
            continue
        for r in engine.scan_all(folder):
            if r.get("brand") and r["brand"] != brand:
                continue
            label = (r.get("company_label") or "").strip()
            if not label or label == "(root)" or engine.looks_like_non_client_label(label):
                continue
            companies.add(label)
            key = label.lower()
            if r.get("type", "").upper() in engine.HTML_DOC_TYPES and r.get("ext") == "pdf":
                sc = engine.read_sidecar(r["path"])
                if sc.get("customer_attn") and key not in attn_by_company:
                    attn_by_company[key] = sc["customer_attn"]
                if sc.get("customer_address") and key not in addr_by_company:
                    addr_by_company[key] = sc["customer_address"]
            elif r.get("ext") == "xlsx":
                xlsx_by_company.setdefault(key, []).append((r.get("date") or "", r["path"]))

    clients = load_clients(brand)
    existing_by_name = {c["name"].strip().lower(): c for c in clients}

    def needs_enrichment(rec):
        return any(not (rec.get(f) or "").strip() for f in _CLIENT_CONTACT_FIELDS)

    def legacy_contact_for(key, company_name):
        paths = sorted(xlsx_by_company.get(key, []), key=lambda t: t[0], reverse=True)
        merged = {}
        for _, path in paths[:5]:
            data = engine.extract_legacy_contact(path, company_name=company_name)
            for f in _CLIENT_CONTACT_FIELDS:
                if data.get(f) and not merged.get(f):
                    merged[f] = data[f]
            if all(merged.get(f) for f in _CLIENT_CONTACT_FIELDS):
                break
        return merged

    now_ms = int(datetime.datetime.now().timestamp() * 1000)
    imported = enriched = 0
    for label in sorted(companies, key=str.lower):
        key = label.lower()
        record = existing_by_name.get(key)
        if record is None:
            record = {
                "id": f"c_{now_ms}_{imported}", "name": label, "category": "",
                "attn": "", "address": "", "po_box": "", "city": "", "country": "",
                "phone": "", "landline": "", "email": "", "website": "", "trn": "",
                "notes": "", "logo": "", "updated": datetime.date.today().isoformat(),
            }
            clients.append(record)
            existing_by_name[key] = record
            imported += 1
        if not needs_enrichment(record):
            continue
        legacy = legacy_contact_for(key, record["name"])
        changed = False
        if not record["attn"].strip() and attn_by_company.get(key):
            record["attn"] = attn_by_company[key]; changed = True
        if not record["address"].strip() and addr_by_company.get(key):
            record["address"] = addr_by_company[key]; changed = True
        for f in _CLIENT_CONTACT_FIELDS:
            if not (record.get(f) or "").strip() and legacy.get(f):
                record[f] = legacy[f]; changed = True
        if changed:
            record["updated"] = datetime.date.today().isoformat()
            enriched += 1

    save_clients(brand, clients)
    return jsonify({"clients": clients, "imported": imported, "enriched": enriched})

@app.get("/api/clients-export")
def api_clients_export():
    """Excel export of the client database — every client for the current
    brand (Settings' "Export All Client Profiles"), or a single one when
    ?id=... is given (the Clients tab's per-card "Export" button). Same
    workbook layout either way (engine.build_clients_workbook)."""
    brand = current_brand()
    cid = request.args.get("id", "")
    clients = load_clients(brand)
    if cid:
        clients = [c for c in clients if c["id"] == cid]
        if not clients:
            return jsonify({"error": "Client not found."}), 404
    wb = engine.build_clients_workbook(clients)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{engine._slug(clients[0]['name'])}.xlsx" if cid else f"{brand}_Clients.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------------------------------------------------------- indexing / history
@app.get("/api/index")
def api_index():
    """Every file actually found across the QTN/INV/DO folders (recursive),
    grouped by the company subfolder it lives in. Not just convention-named
    files. Scoped to the currently selected brand — untagged legacy docs (made
    before multi-brand support) show up regardless of which brand is picked,
    but docs tagged for a different brand are hidden."""
    folders = all_doc_folders()
    brand = current_brand()
    if not folders:
        return jsonify({"records": [], "companies": [], "note": "No folder set.", "brand": brand})
    recs = []
    attns, addresses = set(), set()
    for folder in folders:
        if not os.path.isdir(folder):
            continue
        for r in engine.scan_all(folder):
            # Sololuce Datasheets are always generated under SOLOLUCE (see
            # all_doc_folders' comment) — they must stay visible in All Docs
            # regardless of whichever brand is currently active elsewhere in
            # the app, since "which brand am I working in right now" and
            # "does a Sololuce Datasheet I already made still exist" are
            # unrelated questions.
            if r.get("brand") and r["brand"] != brand and not (r.get("type", "").upper() == "CAT" and r["brand"] == "SOLOLUCE"):
                continue
            r["rel"] = os.path.relpath(r["path"], folder).replace(os.sep, "/")
            if r.get("type", "").upper() in engine.HTML_DOC_TYPES and r.get("ext") == "pdf":
                sc = engine.read_sidecar(r["path"])
                if sc.get("customer_attn"):
                    attns.add(sc["customer_attn"])
                if sc.get("customer_address"):
                    addresses.add(sc["customer_address"])
                # Status (Draft/Sent/Approved/Revised/None) only exists for
                # QTN2 — the legacy xlsx pipeline has no such concept at all.
                if sc.get("status"):
                    r["status"] = sc["status"]
            r.pop("path", None)
            recs.append(r)
    # A DO/INV counts as "in progress" (highlighted red in All Docs) for as
    # long as its submission hasn't reached "submittal_built" — matched by
    # (type, number) rather than file path, since that's stable regardless
    # of exactly how the path was stored.
    in_progress = set()
    for s in load_submissions(brand):
        if s.get("stage") == "submittal_built":
            continue
        if s.get("do_number"):
            in_progress.add(("DO", int(s["do_number"])))
        if s.get("inv_number"):
            in_progress.add(("INV", int(s["inv_number"])))
    for r in recs:
        # filenames zero-pad the number ("0001"), submissions store it as a
        # plain int — compare numerically, not as raw strings, or every
        # match silently misses.
        num = r.get("number", "")
        if str(num).isdigit() and (r.get("type", "").upper(), int(num)) in in_progress:
            r["in_progress"] = True
    companies = sorted({r["company_label"] for r in recs}, key=str.lower)
    projects = sorted({r["project_label"] for r in recs if r.get("project_label")}, key=str.lower)
    return jsonify({"records": recs, "companies": companies, "projects": projects,
                     "attns": sorted(attns), "addresses": sorted(addresses), "brand": brand})

def _api_doc_from_pdf(rel, path, meta):
    """Edit for a document that only exists as a PDF (no xlsx twin) — either
    an HTML_DOC_TYPE (QTN2/CAT/EXP, PDF-only by design since they're
    Playwright-rendered) or a Delivery Order that only has a PDF (kept as-is
    after physical signature — the one legacy type this business regularly
    keeps PDF-only). Everything else that only has a PDF alongside is noise
    living in the same folder tree (client drawings/LPOs) or already has its
    xlsx twin and never reaches here, so it still falls back to the
    read-only Company System preview."""
    basename = os.path.basename(rel)
    doc_type = meta["type"].upper() if meta else engine._guess_type(basename)
    if doc_type in engine.HTML_DOC_TYPES:
        # The app's own sidecar JSON (saved alongside every generated
        # QTN2/CAT/EXP PDF — see engine.save_sidecar) already has the exact
        # header fields + items/rows used to generate it, no PDF
        # text-scraping needed at all, this is just reading it back. Also
        # used by the Submissions picker so a confirmed Quotation (New
        # Design) can start a submission exactly like a legacy xlsx one.
        sidecar = engine.read_sidecar(path)
        if not sidecar:
            return jsonify({"error": "No saved data found for this document."}), 404
        sidecar = dict(sidecar)
        sidecar["original"] = rel
        sidecar["doc_type"] = doc_type
        sidecar["imported"] = False
        return jsonify(sidecar)
    if doc_type != "DO":
        return jsonify({"error": "Only Delivery Orders can be edited straight from a PDF — "
                                  "this file isn't one."}), 400
    try:
        result = engine.read_do_pdf(path)
    except Exception as e:
        return jsonify({"error": f"Could not read the file: {e}"}), 500
    if not result.get("items"):
        return jsonify({"error": "Couldn't read this PDF automatically — it's likely a scanned/"
                                  "signed copy with no selectable text. Use the \"PDF\" button to "
                                  "open it directly instead."}), 422
    header = result.get("header", {})
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        mtime = 0
    folder_company = rel.split("/")[0] if "/" in rel else ""
    rev = header.get("rev", "")
    data = {"number": header.get("number") or (meta["number"] if meta else engine._guess_number(basename)),
            "rev": int(rev) if rev.isdigit() else (meta["rev"] if meta else 0),
            "company": meta["company_label"] if meta else folder_company.replace("_", " "),
            "project": header.get("project") or (meta["project_label"] if meta else ""),
            "date": (engine._parse_ddmmyyyy(header.get("date", "")) or
                     (meta["date"] if meta else engine._guess_date(basename, mtime))),
            "lpo_number": header.get("lpo_number", ""),
            "original": rel, "doc_type": "DO", "items": result["items"], "imported": True}
    return jsonify(data)

# ---------------------------------------------------------------- All Docs bulk actions (multi-select Delete/Clone/Cut)
# A doc's "canonical" path (xlsx if it has one, else pdf) always has up to
# 4 real sibling files sharing its exact stem: the xlsx/pdf pair itself,
# engine.markdown_path's .md companion, and — for HTML_DOC_TYPES only —
# engine.sidecar_path's .json (the exact data the PDF was rendered from).
# Same file set /api/generate's own `replace` cleanup already touches; this
# is that same logic, just reachable directly instead of only as a side
# effect of overwriting a document under a new name.
def _related_doc_files(path):
    stem = os.path.splitext(path)[0]
    return [f for f in (stem + ".xlsx", stem + ".pdf", stem + ".md", stem + ".json") if os.path.exists(f)]

def _resolve_alldocs_rel(rel, brand):
    """All Docs rows are already resolved against the current brand
    elsewhere (see resolve_rel) — bulk actions take the same rel strings
    those rows already carry, scoped to the brand the request specifies
    (defaults to whichever brand is currently active)."""
    return resolve_rel(rel, brand)

@app.post("/api/alldocs-delete")
def api_alldocs_delete():
    """Permanently deletes every real file (xlsx/pdf/md/json sidecar) for
    each selected document. The frontend gates this behind its own
    confirm-again-to-proceed control — nothing here double-checks, so
    only call this once the user has actually confirmed."""
    data = request.json or {}
    rels = data.get("rels") or []
    brand = current_brand()
    deleted, errors = [], []
    for rel in rels:
        folder, path = _resolve_alldocs_rel(rel, brand)
        if not path:
            errors.append({"rel": rel, "error": "File not found."})
            continue
        try:
            meta = engine.parse_filename(os.path.basename(path))
            for f in _related_doc_files(path):
                os.remove(f)
            deleted.append(rel)
            if meta and meta["type"].upper() == "INV":
                inv_rel = os.path.relpath(os.path.splitext(path)[0] + ".xlsx", folder).replace(os.sep, "/")
                entries = [e for e in load_ledger(brand) if e["rel"] != inv_rel]
                save_ledger(brand, entries)
        except OSError as e:
            errors.append({"rel": rel, "error": str(e)})
    return jsonify({"deleted": deleted, "errors": errors})

@app.post("/api/alldocs-clone")
def api_alldocs_clone():
    """Copies each selected document's real files as-is (same content,
    same company/project/date) under a freshly assigned next number —
    a fast exact duplicate. The clone's PDF/xlsx still shows the
    ORIGINAL's number/date inside until it's edited and regenerated; only
    the filename (and so its All Docs listing) reflects the new number."""
    data = request.json or {}
    rels = data.get("rels") or []
    brand = current_brand()
    cloned, errors = [], []
    for rel in rels:
        folder, path = _resolve_alldocs_rel(rel, brand)
        if not path:
            errors.append({"rel": rel, "error": "File not found."})
            continue
        meta = engine.parse_filename(os.path.basename(path))
        if not meta:
            errors.append({"rel": rel, "error": "Not a recognized document filename."})
            continue
        try:
            new_number = _next_number_for(meta["type"].upper(), brand)
            new_stem = engine.build_filename(meta["type"], new_number, 0, meta["company"].replace("-", " "),
                                              meta["project"].replace("-", " "), meta["date"],
                                              brand=meta.get("brand"), ext="")[:-1]
            new_path_base = os.path.join(os.path.dirname(path), new_stem)
            for f in _related_doc_files(path):
                ext = os.path.splitext(f)[1]
                shutil.copy2(f, new_path_base + ext)
            cloned.append(os.path.relpath(new_path_base + os.path.splitext(path)[1], folder).replace(os.sep, "/"))
        except OSError as e:
            errors.append({"rel": rel, "error": str(e)})
    return jsonify({"cloned": cloned, "errors": errors})

@app.post("/api/alldocs-move")
def api_alldocs_move():
    """Renames each selected document's real files, replacing only the
    company segment of the filename with target_company — since All Docs
    groups strictly by that filename segment (see allDocsGroupKey in the
    frontend), this is what actually "moves" a document into a different
    company/project group. Refuses (per file) rather than overwrites when
    a same-named file already sits at the destination."""
    data = request.json or {}
    rels = data.get("rels") or []
    target_company = (data.get("target_company") or "").strip()
    if not target_company:
        return jsonify({"error": "No destination company given."}), 400
    brand = current_brand()
    moved, errors = [], []
    for rel in rels:
        folder, path = _resolve_alldocs_rel(rel, brand)
        if not path:
            errors.append({"rel": rel, "error": "File not found."})
            continue
        meta = engine.parse_filename(os.path.basename(path))
        if not meta:
            errors.append({"rel": rel, "error": "Not a recognized document filename."})
            continue
        try:
            new_stem = engine.build_filename(meta["type"], meta["number"], meta["rev"], target_company,
                                              meta["project"].replace("-", " "), meta["date"],
                                              brand=meta.get("brand"), ext="")[:-1]
            new_path_base = os.path.join(os.path.dirname(path), new_stem)
            related = _related_doc_files(path)
            if any(os.path.exists(new_path_base + os.path.splitext(f)[1]) for f in related):
                errors.append({"rel": rel, "error": "A document with that number/date already exists for "
                                                      + target_company + "."})
                continue
            old_inv_rel = (os.path.relpath(os.path.splitext(path)[0] + ".xlsx", folder).replace(os.sep, "/")
                            if meta["type"].upper() == "INV" else None)
            for f in related:
                ext = os.path.splitext(f)[1]
                os.rename(f, new_path_base + ext)
            if old_inv_rel:
                new_inv_rel = os.path.relpath(new_path_base + ".xlsx", folder).replace(os.sep, "/")
                entries = load_ledger(brand)
                entry = next((e for e in entries if e["rel"] == old_inv_rel), None)
                if entry:
                    entry["rel"] = new_inv_rel
                    save_ledger(brand, entries)
            moved.append(os.path.relpath(new_path_base + os.path.splitext(path)[1], folder).replace(os.sep, "/"))
        except OSError as e:
            errors.append({"rel": rel, "error": str(e)})
    return jsonify({"moved": moved, "errors": errors})

@app.get("/api/doc")
def api_doc():
    """Load every field of an existing document so it can be fully edited."""
    rel = request.args.get("rel", "")
    _folder, path = resolve_rel(rel)
    ext = os.path.splitext(rel)[1].lower()
    if not path or ext not in (".xlsx", ".pdf"):
        return jsonify({"error": "File not found."}), 404
    meta = engine.parse_filename(os.path.basename(rel))
    if ext == ".pdf":
        return _api_doc_from_pdf(rel, path, meta)
    editable = bool(meta and meta["type"].upper() in engine.FILLERS)
    if editable:
        try:
            extra = engine.read_full_record(path, meta["type"])
        except Exception as e:
            return jsonify({"error": f"Could not read the file: {e}"}), 500
        data = {"number": meta["number"], "rev": meta["rev"],
                "company": meta["company_label"], "project": meta["project_label"],
                "date": meta["date"], "original": rel, "imported": False}
        data.update(extra)          # may add area / qtn_number / lpo_number / type (INV's own "Type" field)
        data["doc_type"] = meta["type"]   # QTN / INV / DO — kept separate so it can't collide with INV's "type" field
        return jsonify(data)
    # Best-effort import: this file wasn't made by the app, so we don't know
    # its exact cell layout up front. We still pull the line items out
    # heuristically (by locating the "Item Description" column), and now
    # generically scan the WHOLE sheet for any labeled header cell ("QTN
    # Number", "LPO No", "Project", "Type", "Rev"...) — same "find the
    # label, take the adjacent value" idea as the PDF DO reader, so it
    # adapts to whatever layout the file actually uses instead of assuming
    # the app's own cell positions. The customer/"To" block is tried two
    # ways: many of this business's real historical files draw it as a
    # floating Excel shape rather than a cell (confirmed by unzipping one —
    # the text lives in xl/drawings/drawingN.xml, invisible to the normal
    # cell API), so that's tried first, falling back to the same cell the
    # app's own template uses. The real doc type comes from the same
    # keyword guesser index_folder/scan_all already use for the All Docs
    # listing (so a legacy Tax Invoice doesn't get silently treated as a
    # Quotation), and the on-disk company subfolder is the fallback company
    # name (same source All Docs uses to label it). Anything genuinely not
    # determinable is left blank, never guessed outright. Generate then
    # saves this as a brand-new file in the app's own template — it doesn't
    # overwrite the original, since the layouts aren't the same document.
    basename = os.path.basename(rel)
    guessed_type = engine._guess_type(basename)
    doc_type = (meta["type"].upper() if meta and meta["type"].upper() in engine.FILLERS
                else guessed_type if guessed_type in engine.FILLERS else "INV")
    try:
        items = engine.read_items_from_doc(path)
    except Exception as e:
        return jsonify({"error": f"Could not read the file: {e}"}), 500
    header = {}
    customer_block = ""
    try:
        wb = load_workbook(path, data_only=True)
        header = engine.read_xlsx_header_labels(wb.active, engine._XLSX_HEADER_LABELS.get(doc_type, {}))
        customer_cell = engine.CUSTOMER_CELL.get(doc_type)
        cell_block = (str(wb.active[customer_cell].value).strip()
                      if customer_cell and wb.active[customer_cell].value else "")
        customer_block = engine.read_xlsx_customer_shape(path) or cell_block
    except Exception as e:
        print(f"open_in_cs: could not read header/customer block from {path}: {e}")
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        mtime = 0
    folder_company = rel.split("/")[0] if "/" in rel else ""
    rev = header.get("rev", "")
    scanned_date = header.get("date", "")
    date_iso = scanned_date if re.match(r"^\d{4}-\d{2}-\d{2}$", scanned_date) else engine._parse_ddmmyyyy(scanned_date)
    # The top-level folder is often a *region* or catch-all bucket ("ABU
    # DHABI QTN", "ADS QUOTATION") holding many different clients' files,
    # not the client itself — confirmed by checking several real files
    # where it was clearly wrong. The customer block's own first line
    # (the actual company name typed into the "To" section) is a far more
    # reliable source when it's available; the folder name is only a
    # last-resort fallback now, not the primary guess.
    company_from_block = customer_block.split("\n")[0].strip() if customer_block else ""
    data = {"number": header.get("number") or (meta["number"] if meta else engine._guess_number(basename)),
            "rev": int(rev) if rev.isdigit() else (meta["rev"] if meta else 0),
            "company": meta["company_label"] if meta else (company_from_block or folder_company.replace("_", " ")),
            "project": header.get("project") or (meta["project_label"] if meta else ""),
            "date": date_iso or (meta["date"] if meta else engine._guess_date(basename, mtime)),
            "qtn_number": header.get("qtn_number", ""), "lpo_number": header.get("lpo_number", ""),
            "type": header.get("type", ""), "area": header.get("area", ""),
            "customer_block": customer_block,
            "original": rel, "doc_type": doc_type, "items": items, "imported": True}
    return jsonify(data)

@app.get("/api/previous")
def api_previous():
    company = request.args.get("company", "")
    dtype = request.args.get("type", "QTN2")
    folder = folder_for(dtype)
    if not folder or not company:
        return jsonify({"previous": []})
    brand = current_brand()
    prev = [p for p in engine.previous_for_company(folder, company, dtype, limit=20) if _brand_matches(p, brand)][:3]
    for p in prev:
        p["items"] = (engine.read_sidecar_items(p["path"]) if dtype.upper() in engine.HTML_DOC_TYPES
                       else engine.read_items_from_doc(p["path"]))
    return jsonify({"previous": prev})

def _next_number_for(dtype, brand):
    folder = folder_for(dtype)
    n = 0
    if folder and os.path.isdir(folder):
        for r in engine.index_folder(folder):
            if r["type"].upper() == dtype and str(r["number"]).isdigit() and _brand_matches(r, brand):
                n = max(n, int(r["number"]))
    return n + 1

@app.get("/api/next-number")
def next_number():
    """Suggest the next sequential number for a type, scoped to the current brand."""
    dtype = request.args.get("type", "QTN2").upper()
    return jsonify({"next": _next_number_for(dtype, current_brand())})

# ---------------------------------------------------------------- generate
@app.post("/api/generate")
def api_generate():
    data = request.json or {}
    dtype = data.get("doc_type", "QTN2").upper()
    # Sololuce Datasheets are always Sololuce's, full stop — never whichever
    # brand happens to be globally active when Generate is clicked (brand is
    # one mutable setting shared by the whole app, not a per-document
    # property). Hardcoding this is what actually *guarantees* a CAT file is
    # tagged and filed under Sololuce, so it only ever shows up in Sololuce's
    # own All Docs — rather than relying on every path that can reach
    # Generate (new build, resumed draft, reopened record) having correctly
    # force-switched the active brand first.
    gen_brand = "SOLOLUCE" if dtype == "CAT" else current_brand()
    folder = folder_for(dtype, brand=gen_brand)
    if not folder or not os.path.isdir(folder):
        err_msg = ("Set the Sololuce Datasheets folder first (see Settings)." if dtype == "CAT"
                   else f"Set your {dtype} documents folder first (see Settings).")
        return jsonify({"error": err_msg}), 400
    replace = (data.get("replace", "") or "").strip()
    try:
        if dtype in engine.HTML_DOC_TYPES:
            stem = engine.build_filename(dtype, data.get("number"), data.get("rev", 0),
                                          data.get("company"), data.get("project", ""),
                                          data.get("date"), brand=gen_brand, ext="")[:-1]
            pdf_path = os.path.join(folder, stem + ".pdf")
            # A shallow copy carries the current Ordering Table "standard"
            # widths into the render without baking that global setting into
            # the saved sidecar below (save_sidecar(pdf_path, data) still
            # gets the untouched original) — see build_ordering_table's own
            # docstring on why a saved draft must never freeze a stale copy
            # of a setting that's meant to keep tracking whatever's current.
            render_data = data
            if dtype == "CAT":
                render_data = dict(data)
                render_data["cat_ordering_default_widths"] = load_cfg().get("cat_ordering_default_widths", {})
            html_engine.RENDERERS[dtype](render_data, pdf_path, brand=gen_brand)
            if dtype == "CAT":
                # Page numbers/tabs are no longer stamped here — the Full
                # Catalog Builder (catalog_builder.py) is the only place that
                # happens now, computed fresh against the real book layout
                # every time it runs. A number burned onto this file the
                # moment it's generated could only ever be a guess (grouping
                # products by category means one new product can shift every
                # later page in an 800-page book) — see catalog_builder.py's
                # own module docstring for the full reasoning. Series color
                # assignment stays here as a defensive no-op-if-it-already-
                # has-one safety net, in case a category ever reaches this
                # point without going through /api/cat-series-add's normal
                # commit flow first (that route is the one real users always
                # go through — typing a custom category commits on blur).
                series_label = (data.get("series") or "").strip()
                if series_label:
                    cfg = load_cfg()
                    _assign_series_color(cfg, series_label)
                    save_cfg(cfg)
            engine.save_sidecar(pdf_path, data)
            # Markdown companion — every field in full, plain readable text,
            # no embedded images — so this document's complete content can be
            # read (and edited) in one pass without opening the PDF or
            # wading through the sidecar's base64 photo blobs. xlsx-based
            # types (INV/DO/...) get theirs for free inside engine.generate()
            # itself; HTML_DOC_TYPES (QTN2/CAT/EXP) don't go through that
            # function, so it's written explicitly here instead.
            badge_library = load_cfg().get("cat_badge_library", []) if dtype == "CAT" else None
            engine.save_markdown(pdf_path, dtype, data, brand=gen_brand, badge_library=badge_library)
            res = {"pdf": pdf_path}
        else:
            res = engine.generate(dtype, data, folder, brand=gen_brand, make_pdf=True)
            if dtype == "INV":
                record_invoice_in_ledger(data, res["xlsx"], current_brand(), folder)
        if replace:
            _rfolder, old_xlsx = resolve_rel(replace)
            if old_xlsx and os.path.abspath(old_xlsx) != os.path.abspath(res.get("xlsx", res.get("pdf"))):
                old_pdf = os.path.splitext(old_xlsx)[0] + ".pdf"
                for f in (old_xlsx, old_pdf):
                    if os.path.exists(f):
                        os.remove(f)
                old_md = engine.markdown_path(old_xlsx)
                if os.path.exists(old_md):
                    os.remove(old_md)
                if dtype in engine.HTML_DOC_TYPES:
                    # Editing an HTML_DOC_TYPE (QTN2/CAT/EXP) into a different
                    # filename (number/company/date changed) leaves its old
                    # JSON sidecar behind pointing at a file that no longer
                    # exists — drop it along with the old xlsx/pdf above,
                    # same reasoning as INV's stale ledger entry below.
                    old_sidecar = engine.sidecar_path(old_pdf)
                    if os.path.exists(old_sidecar):
                        os.remove(old_sidecar)
                if dtype == "INV":
                    # editing renamed the file (different number/company/date) —
                    # drop the stale ledger entry pointing at the now-deleted old one
                    old_rel = os.path.relpath(old_xlsx, folder).replace(os.sep, "/")
                    brand = current_brand()
                    entries = [e for e in load_ledger(brand) if e["rel"] != old_rel]
                    save_ledger(brand, entries)
        rel = os.path.relpath(res.get("pdf", res.get("xlsx")), folder)
        return jsonify({"ok": True, "xlsx": os.path.basename(res.get("xlsx", "")),
                        "pdf": os.path.basename(res.get("pdf", "")),
                        "preview_url": "/preview?f=" + rel})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# ---------------------------------------------------------------- submissions
# QTN (Approved) -> DO + INV generated together immediately, real files but
# flagged "in progress" (red in All Docs) -> scanned/signed DO linked once
# delivery happens -> "build the submittal" bundles QTN+LPO+DO+INV into one
# combined PDF, correcting DO/INV quantities first if the LPO differs from
# what was quoted. Per the user's own sequencing: the LPO only enters the
# picture at the submittal step, not upfront.
SUBMISSION_STAGES = ("in_progress", "delivered", "submittal_built")

@app.get("/api/submissions")
def api_submissions():
    return jsonify({"submissions": load_submissions()})

def _resolve_qtn_pdf(qtn_rel, brand):
    """The submittal merge needs the Quotation as a PDF. QTN2 already saved
    one; a legacy xlsx QTN normally has a .pdf sibling from its own
    Generate — fall back to converting on the fly for the rare case it
    doesn't (e.g. a very old file predating this app's own PDF export)."""
    if not qtn_rel:
        return None
    _folder, path = resolve_rel(qtn_rel, brand)
    if not path:
        return None
    if path.lower().endswith(".pdf"):
        return path
    sibling = os.path.splitext(path)[0] + ".pdf"
    if os.path.exists(sibling):
        return sibling
    try:
        return engine.to_pdf(path, os.path.dirname(path))
    except Exception:
        return None

def _html_desc_to_plain(html):
    """QTN2's own item description box is a contenteditable richbox — real
    HTML (<div>/<br> line breaks), which is exactly right for QTN2's own
    HTML-rendered output. DO and INV are different: engine.py writes their
    description straight into an openpyxl cell (ws.cell(r, 5, ...)), and
    Excel cells can't render HTML at all — the raw tags show up as literal
    text in the generated PDF (e.g. "CELIA<div>IP65 - 3000K -
    10W/m<br>CODE:...</div>"). Converts that HTML back to the plain,
    newline-separated text those xlsx templates actually expect, the same
    shape a hand-typed DO/INV description (plain <textarea>, literal \n)
    already has."""
    if not html:
        return ""
    text = re.sub(r"<br\s*/?>", "\n", html, flags=re.IGNORECASE)
    # both the OPENING and closing tag are line-break boundaries: a
    # contenteditable div's first line of text sits before its own <div>
    # even opens (e.g. "CELIA<div>IP65...<br>CODE:...</div>") — treating
    # only </div> as a break would glue that first line onto the next one.
    text = re.sub(r"</?(div|p)[^>]*>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    from html import unescape
    text = unescape(text)
    return "\n".join(ln.strip() for ln in text.split("\n") if ln.strip())

def _generate_do_and_inv(sub_id, brand, company, project, qtn_number, items,
                          do_number=None, inv_number=None):
    """Generates (or, when numbers are passed in, re-generates in place —
    same filename, so it overwrites rather than creating a new numbered
    document) the DO and INV together from one shared items list, so they
    can never disagree on quantities."""
    do_folder, inv_folder = folder_for("DO"), folder_for("INV")
    if not do_folder or not os.path.isdir(do_folder):
        return None, "Set your Delivery Orders folder first (see Settings)."
    if not inv_folder or not os.path.isdir(inv_folder):
        return None, "Set your Tax Invoices folder first (see Settings)."
    do_number = do_number or _next_number_for("DO", brand)
    inv_number = inv_number or _next_number_for("INV", brand)
    do_items = [{"description": _html_desc_to_plain(it.get("description", "")), "unit": it.get("unit") or "PCS",
                 "lpo_qty": it.get("lpo_qty", it.get("qty", "")), "prev_delivery": "", "delivered": "",
                 "photo": it.get("photo", "")}
                for it in items]
    inv_items = [{"description": _html_desc_to_plain(it.get("description", "")), "unit": it.get("unit") or "PCS",
                  "qty": it.get("lpo_qty", it.get("qty", "")), "price": it.get("price", ""),
                  "photo": it.get("photo", "")}
                 for it in items]
    do_data = {"number": do_number, "rev": 0, "date": datetime.date.today().isoformat(),
               "project": project, "lpo_number": "", "company": company, "items": do_items}
    inv_data = {"number": inv_number, "date": datetime.date.today().isoformat(),
                "qtn_number": qtn_number, "lpo_number": "", "project": project, "type": "",
                "company": company, "items": inv_items, "discount": {},
                "vat": {"enabled": True, "mode": "percent", "value": 5}}
    try:
        do_res = engine.generate("DO", do_data, do_folder, brand=brand, make_pdf=True)
        inv_res = engine.generate("INV", inv_data, inv_folder, brand=brand, make_pdf=True)
        record_invoice_in_ledger(inv_data, inv_res["xlsx"], brand, inv_folder)
    except Exception as e:
        traceback.print_exc()
        return None, f"Could not generate the Delivery Order/Invoice: {e}"
    return {
        "do_number": do_number, "do_rel": os.path.relpath(do_res["xlsx"], do_folder).replace(os.sep, "/"),
        "inv_number": inv_number, "inv_rel": os.path.relpath(inv_res["xlsx"], inv_folder).replace(os.sep, "/"),
    }, None

@app.post("/api/submissions")
def api_submissions_create():
    """Approving a Quotation immediately generates a real DO + Invoice from
    the quoted quantities (per the user's own call — flagged "in progress"/
    red in All Docs rather than held back as a draft), well before the LPO
    or delivery even happen. The LPO and any quantity correction come later,
    at the "build submittal" step."""
    data = request.json or {}
    company = (data.get("company") or "").strip()
    items = data.get("items") or []
    if not company:
        return jsonify({"error": "Missing company."}), 400
    if not items:
        return jsonify({"error": "No line items."}), 400
    brand = current_brand()
    project = (data.get("project") or "").strip()
    qtn_number = data.get("qtn_number", "")
    result, err = _generate_do_and_inv(None, brand, company, project, qtn_number, items)
    if err:
        return jsonify({"error": err}), 400

    sub_id = f"sub_{int(datetime.datetime.now().timestamp()*1000)}"
    now = datetime.datetime.now().isoformat()
    sub = {"id": sub_id, "brand": brand, "created": now, "updated": now,
           "qtn_rel": data.get("qtn_rel", ""), "qtn_number": qtn_number,
           "company": company, "project": project, "items": items,
           "lpo_number": "", "lpo_filename": "", "lpo_saved_name": "",
           "scanned_do_rel": "", "submittal_rel": "", "stage": "in_progress", **result}
    subs = load_submissions(brand)
    subs.append(sub)
    save_submissions(brand, subs)
    return jsonify({"submission": sub})

def _find_submission(sub_id, brand):
    subs = load_submissions(brand)
    return subs, next((s for s in subs if s["id"] == sub_id), None)

@app.post("/api/submissions-link-scanned-do")
def api_submissions_link_scanned_do():
    """Linking the scanned/signed DO is the "delivery confirmed" moment —
    the frontend prompts to build the submittal right after this succeeds."""
    data = request.json or {}
    brand = current_brand()
    subs, sub = _find_submission(data.get("id", ""), brand)
    if not sub:
        return jsonify({"error": "Submission not found."}), 404
    sub["scanned_do_rel"] = data.get("rel", "")
    sub["stage"] = "delivered"
    sub["updated"] = datetime.datetime.now().isoformat()
    save_submissions(brand, subs)
    return jsonify({"submission": sub})

@app.post("/api/submissions-build-submittal")
def api_submissions_build_submittal():
    """The LPO shows up here for the first time — save it, and if its
    confirmed quantities differ from what was quoted, regenerate the DO/INV
    in place (same number/filename) so the final documents match what was
    actually delivered. Then merge Quotation + LPO + scanned DO + Invoice
    into one combined submittal PDF."""
    data = request.json or {}
    brand = current_brand()
    subs, sub = _find_submission(data.get("id", ""), brand)
    if not sub:
        return jsonify({"error": "Submission not found."}), 404
    if not sub.get("scanned_do_rel"):
        return jsonify({"error": "Link the scanned Delivery Order first."}), 400

    items = data.get("items") or sub.get("items", [])
    sub_dir = os.path.join(SUBMISSIONS_DIR, brand, sub["id"])
    os.makedirs(sub_dir, exist_ok=True)

    lpo_data_url = data.get("lpo_file", "")
    if lpo_data_url and "," in lpo_data_url:
        header, b64 = lpo_data_url.split(",", 1)
        mime_match = re.search(r"data:([^;]+)", header)
        ext = mimetypes.guess_extension(mime_match.group(1)) if mime_match else None
        lpo_filename = (data.get("lpo_filename") or "").strip()
        lpo_saved_name = "lpo" + (ext or os.path.splitext(lpo_filename)[1] or ".pdf")
        try:
            with open(os.path.join(sub_dir, lpo_saved_name), "wb") as f:
                f.write(base64.b64decode(b64))
            sub["lpo_filename"], sub["lpo_saved_name"] = lpo_filename, lpo_saved_name
        except Exception as e:
            return jsonify({"error": f"Could not save the LPO file: {e}"}), 500

    changed = any(str(it.get("lpo_qty", it.get("qty", ""))) != str(old.get("lpo_qty", old.get("qty", "")))
                  for it, old in zip(items, sub.get("items", [])))
    if changed or len(items) != len(sub.get("items", [])):
        result, err = _generate_do_and_inv(sub["id"], brand, sub["company"], sub["project"], sub["qtn_number"],
                                            items, do_number=sub["do_number"], inv_number=sub["inv_number"])
        if err:
            return jsonify({"error": err}), 400
        sub.update(result)
    sub["items"] = items

    do_folder, inv_folder = folder_for("DO"), folder_for("INV")
    do_pdf = os.path.join(do_folder, os.path.splitext(sub["do_rel"])[0] + ".pdf") if do_folder else None
    inv_pdf = os.path.join(inv_folder, os.path.splitext(sub["inv_rel"])[0] + ".pdf") if inv_folder else None
    qtn_pdf = _resolve_qtn_pdf(sub.get("qtn_rel", ""), brand)
    lpo_path = (os.path.join(sub_dir, sub["lpo_saved_name"]) if sub.get("lpo_saved_name") else None)
    scanned_do_path = _resolve_scanned_do(sub.get("scanned_do_rel", ""))
    submittal_path = os.path.join(sub_dir, "submittal.pdf")
    built = engine.build_submittal_pdf([qtn_pdf, lpo_path, scanned_do_path, inv_pdf], submittal_path)
    if not built:
        return jsonify({"error": "Could not build the submittal — none of the source files (Quotation/LPO/scanned DO/Invoice) were readable."}), 500
    sub["submittal_rel"] = "submittal.pdf"
    sub["stage"] = "submittal_built"
    sub["updated"] = datetime.datetime.now().isoformat()
    save_submissions(brand, subs)
    return jsonify({"submission": sub})

@app.post("/api/submissions-delete")
def api_submissions_delete():
    data = request.json or {}
    brand = current_brand()
    sub_id = data.get("id", "")
    subs = [s for s in load_submissions(brand) if s["id"] != sub_id]
    save_submissions(brand, subs)
    shutil.rmtree(os.path.join(SUBMISSIONS_DIR, brand, sub_id), ignore_errors=True)
    return jsonify({"submissions": subs})

@app.get("/submission-lpo")
def submission_lpo():
    """Serve the uploaded LPO file back for viewing/downloading."""
    sub_id = request.args.get("id", "")
    brand = current_brand()
    _subs, sub = _find_submission(sub_id, brand)
    if not sub or not sub.get("lpo_saved_name"):
        return "Not found", 404
    path = os.path.join(SUBMISSIONS_DIR, brand, sub_id, sub["lpo_saved_name"])
    if not os.path.exists(path):
        return "Not found", 404
    return send_file(path)

@app.get("/submission-submittal")
def submission_submittal():
    """Serve the final combined submittal PDF back for viewing/downloading."""
    sub_id = request.args.get("id", "")
    brand = current_brand()
    _subs, sub = _find_submission(sub_id, brand)
    if not sub or not sub.get("submittal_rel"):
        return "Not found", 404
    path = os.path.join(SUBMISSIONS_DIR, brand, sub_id, sub["submittal_rel"])
    if not os.path.exists(path):
        return "Not found", 404
    return send_file(path)

# ---------------------------------------------------------------- Statement of Account
@app.get("/api/finance/ledger")
def api_finance_ledger():
    return jsonify({"ledger": load_ledger()})

@app.post("/api/finance/mark-paid")
def api_finance_mark_paid():
    data = request.json or {}
    brand = current_brand()
    entries = load_ledger(brand)
    entry = next((e for e in entries if e["rel"] == data.get("rel")), None)
    if not entry:
        return jsonify({"error": "Invoice not found in the ledger."}), 404
    entry["paid"] = bool(data.get("paid"))
    entry["paid_date"] = (data.get("paid_date") or datetime.date.today().isoformat()) if entry["paid"] else None
    save_ledger(brand, entries)
    return jsonify({"ledger": entries})

@app.post("/api/browse-scanned-do")
def browse_scanned_do():
    """Native file-open dialog rooted at the configured Scanned Delivery
    Orders folder, so linking a scanned DO to a submission is a real file
    pick rather than typing a path by hand."""
    folder = brand_settings().get("scanned_do_folder", "")
    if not folder or not os.path.isdir(folder):
        return jsonify({"error": "Set your Scanned Delivery Orders folder first (see Settings)."}), 400
    try:
        from tkinter import filedialog
        def job(root):
            root.attributes("-topmost", True)
            p = filedialog.askopenfilename(parent=root, title="Choose the scanned Delivery Order", initialdir=folder)
            root.attributes("-topmost", False)
            return p
        path = _run_on_tk_thread(job)
        if not path:
            return jsonify({"rel": ""})
        if os.path.abspath(path).lower().startswith(os.path.abspath(folder).lower()):
            rel = os.path.relpath(path, folder).replace(os.sep, "/")
        else:
            rel = path  # picked outside the configured folder — kept as an absolute path
        return jsonify({"rel": rel, "name": os.path.basename(path)})
    except Exception as e:
        return jsonify({"error": f"Could not open dialog: {e}"}), 200

def _resolve_scanned_do(rel):
    if not rel:
        return None
    if os.path.isabs(rel) and os.path.exists(rel):
        return rel
    folder = brand_settings().get("scanned_do_folder", "")
    path = os.path.join(folder, rel) if folder else None
    return path if path and os.path.exists(path) else None

@app.get("/open-scanned-do")
def open_scanned_do():
    rel = request.args.get("rel", "")
    path = _resolve_scanned_do(rel)
    if not path:
        return "Not found", 404
    import subprocess, sys
    try:
        if sys.platform.startswith("win"): os.startfile(path)
        elif sys.platform == "darwin": subprocess.Popen(["open", path])
        else: subprocess.Popen(["xdg-open", path])
    except Exception as e:
        return jsonify({"ok": False, "error": f"Could not open that file: {e}"})
    return jsonify({"ok": True})

# ---------------------------------------------------------------- Scan Now (WIA scanner)
# Replaces "scan externally with the scanner's own software, then Browse to
# find the file" with one in-app action — see scanner.py's own top-of-file
# comment for the full flow. Every route here is submissions-gated (same
# blocked_tools prefix as the rest of the scanned-DO flow), not admin-only.
@app.get("/api/scanner-list")
def api_scanner_list():
    return jsonify({"scanners": scanner.list_scanners()})

@app.post("/api/scanner-scan-page")
def api_scanner_scan_page():
    data = request.json or {}
    try:
        result = scanner.scan_one_page(device_id=data.get("device_id") or None, session_id=data.get("session_id"))
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502

@app.post("/api/scanner-remove-last-page")
def api_scanner_remove_last_page():
    data = request.json or {}
    return jsonify(scanner.remove_last_page(data.get("session_id", "")))

@app.post("/api/scanner-cancel")
def api_scanner_cancel():
    data = request.json or {}
    scanner.cancel_session(data.get("session_id", ""))
    return jsonify({"ok": True})

@app.post("/api/scanner-finalize")
def api_scanner_finalize():
    """Combines the session's captured pages into one PDF, named after the
    submission it's for, saved straight into the Scanned Delivery Orders
    folder — same {rel, name} shape /api/browse-scanned-do already
    returns, so the frontend's existing submissions-link-scanned-do call
    needs no changes at all."""
    data = request.json or {}
    folder = brand_settings().get("scanned_do_folder", "")
    if not folder:
        return jsonify({"ok": False, "error": "Set your Scanned Delivery Orders folder first (see Settings)."}), 400
    brand = current_brand()
    _subs, sub = _find_submission(data.get("submission_id", ""), brand)
    filename = scanner.build_scan_filename(brand, sub.get("do_number") if sub else "", sub.get("company") if sub else "")
    try:
        path = scanner.finalize_session(data.get("session_id", ""), folder, filename)
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 502
    return jsonify({"ok": True, "rel": os.path.basename(path), "name": os.path.basename(path)})

@app.get("/preview")
def preview():
    rel = request.args.get("f", "")
    _folder, path = resolve_rel(rel)
    if not path:
        return "Not found", 404
    return send_file(path, mimetype="application/pdf")

CS_CACHE = os.path.join(engine.DATA_BASE, "_cs_cache")  # temp render cache, never written into the user's own folder

CS_PREVIEWABLE = (".pdf", ".xlsx", ".xls", ".doc", ".docx")

@app.get("/cs-thumb")
def cs_thumb():
    """Render a first-page PNG thumbnail for in-app ('Company System') viewing
    — never a new browser page, never an OS app. Works uniformly for Excel,
    PDF, and Word by rendering through LibreOffice and caching the result.
    (An embedded native PDF viewer was tried first, but renders as a blank
    black box in this environment — a flat image sidesteps that entirely.)"""
    rel = request.args.get("f", "")
    _folder, path = resolve_rel(rel)
    if not path:
        return "Not found", 404
    ext = os.path.splitext(path)[1].lower()
    if ext not in CS_PREVIEWABLE:
        return "Preview isn't supported for this file type.", 415
    if ext != ".pdf":
        sibling_pdf = os.path.splitext(path)[0] + ".pdf"
        if os.path.exists(sibling_pdf):
            path, ext = sibling_pdf, ".pdf"  # prefer the already-rendered PDF for fidelity
    os.makedirs(CS_CACHE, exist_ok=True)
    tag = f"{abs(hash(os.path.abspath(path)))}_{int(os.path.getmtime(path))}.png"
    cache_path = os.path.join(CS_CACHE, tag)
    if not os.path.exists(cache_path):
        try:
            tmp_png = engine.to_png(path, CS_CACHE)
            if os.path.abspath(tmp_png) != os.path.abspath(cache_path):
                shutil.move(tmp_png, cache_path)
        except Exception as e:
            return f"Couldn't render a preview for this file: {e}", 500
    return send_file(cache_path, mimetype="image/png")

DRAFT_DIR = os.path.join(engine.DATA_BASE, "_cs_draft")
DRAFT_XLSX = os.path.join(DRAFT_DIR, "draft.xlsx")

@app.post("/api/preview-draft")
def api_preview_draft():
    """Render the Build form's current (unsaved) state as a live preview —
    every page of it, for the multi-page preview pane. Never touches the
    user's documents folder — just a throwaway temp render overwritten on
    every call."""
    data = request.json or {}
    dtype = data.get("doc_type", "QTN2").upper()
    if dtype not in engine.FILLERS and dtype not in engine.HTML_DOC_TYPES:
        return jsonify({"error": "Unknown type."}), 400
    draft = dict(data)
    draft["number"] = data.get("number") or "0000"
    draft["date"] = data.get("date") or datetime.date.today().isoformat()
    draft["rev"] = data.get("rev") or "0"
    if dtype == "CAT":
        draft["cat_ordering_default_widths"] = load_cfg().get("cat_ordering_default_widths", {})
    try:
        os.makedirs(DRAFT_DIR, exist_ok=True)
        for f in os.listdir(DRAFT_DIR):
            if f.startswith("draft_") and f.endswith(".png"):
                os.remove(os.path.join(DRAFT_DIR, f))
        # INV/DO are xlsx types (real .xlsx still generated on Generate —
        # see engine.HTML_PDF_DOC_TYPES's own comment) but their PDF, here
        # included, is rendered via html_engine same as the true
        # HTML_DOC_TYPES — otherwise this live preview would show the old
        # plain LibreOffice-converted look while the real Generate button
        # produces the new pixel-fidelity design, visibly disagreeing.
        if dtype in engine.HTML_DOC_TYPES or dtype in engine.HTML_PDF_DOC_TYPES:
            draft_pdf = os.path.join(DRAFT_DIR, "draft.pdf")
            html_engine.RENDERERS[dtype](draft, draft_pdf, brand=current_brand())
            pdf = draft_pdf
        else:
            engine.FILLERS[dtype](draft, DRAFT_XLSX, brand=current_brand())
            pdf = engine.to_pdf(DRAFT_XLSX, DRAFT_DIR)
        pages = engine.to_png_pages(pdf, DRAFT_DIR, "draft")
        return jsonify({"ok": True, "pages": pages})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.get("/draft-preview")
def draft_preview():
    page = request.args.get("page", "1")
    path = os.path.join(DRAFT_DIR, f"draft_{page}.png")
    if not os.path.exists(path):
        return "Not found", 404
    resp = send_file(path, mimetype="image/png")
    resp.headers["Cache-Control"] = "no-store"
    return resp

CAT_IMPORT_DIR = os.path.join(engine.DATA_BASE, "_cs_import")  # scratch, like _cs_cache/_cs_draft — safe to delete
CAT_IMPORT_DPI = 170  # must match whatever dpi resolve_boxes()/recapture_region() scale against

def _cat_import_dir(import_id):
    """Validates import_id looks like a real uuid4 we generated — never trust
    a client-supplied path segment directly (this feeds straight into
    os.path.join for a folder we then read/write)."""
    try:
        uuid.UUID(import_id)
    except (ValueError, TypeError, AttributeError):
        return None
    return os.path.join(CAT_IMPORT_DIR, import_id)

@app.post("/api/cat-import/upload")
def api_cat_import_upload():
    """Sololuce Datasheets' "Import from PDF": save the uploaded vendor PDF
    and rasterize every page (engine.to_png_pages is already source-agnostic
    — see CLAUDE.md's preview-rendering notes — so this works unmodified on
    a third-party PDF, not just this app's own generated drafts)."""
    data = request.json or {}
    data_url = data.get("pdf", "")
    if "," not in data_url:
        return jsonify({"error": "No PDF received."}), 400
    _header, b64 = data_url.split(",", 1)
    import_id = str(uuid.uuid4())
    folder = os.path.join(CAT_IMPORT_DIR, import_id)
    os.makedirs(folder, exist_ok=True)
    pdf_path = os.path.join(folder, "source.pdf")
    try:
        with open(pdf_path, "wb") as f:
            f.write(base64.b64decode(b64))
        pages = engine.to_png_pages(pdf_path, folder, "src", dpi=CAT_IMPORT_DPI)
        return jsonify({"ok": True, "importId": import_id, "pages": pages})
    except Exception as e:
        shutil.rmtree(folder, ignore_errors=True)
        return jsonify({"error": f"Couldn't read that PDF: {e}"}), 400

@app.get("/cat-import-preview")
def cat_import_preview():
    folder = _cat_import_dir(request.args.get("importId", ""))
    if not folder:
        return "Not found", 404
    path = os.path.join(folder, f"src_{request.args.get('page', '1')}.png")
    if not os.path.exists(path):
        return "Not found", 404
    resp = send_file(path, mimetype="image/png")
    resp.headers["Cache-Control"] = "no-store"
    return resp

@app.post("/api/cat-import/extract")
def api_cat_import_extract():
    folder = _cat_import_dir((request.json or {}).get("importId", ""))
    if not folder:
        return jsonify({"error": "Unknown import."}), 404
    pdf_path = os.path.join(folder, "source.pdf")
    try:
        extracted = pdf_extract.extract_datasheet(pdf_path)
        extracted = pdf_extract.resolve_boxes(pdf_path, extracted, CAT_IMPORT_DPI)
        return jsonify({"ok": True, "fields": extracted})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Extraction failed: {e}"}), 500

@app.get("/cat-import-crop")
def cat_import_crop():
    """Plain crop of a box's region, as a real PNG — how a photo/lifestyle/
    diagram box's pixels actually get pulled into the CAT form's image
    fields once the user has it positioned correctly."""
    folder = _cat_import_dir(request.args.get("importId", ""))
    if not folder:
        return "Not found", 404
    try:
        rect = {k: float(request.args[k]) for k in ("x0", "y0", "x1", "y1")}
        png_bytes = pdf_extract.crop_png(
            os.path.join(folder, "source.pdf"), int(request.args.get("page", 1)), rect, CAT_IMPORT_DPI)
    except Exception as e:
        return f"Crop failed: {e}", 400
    return Response(png_bytes, mimetype="image/png")

@app.post("/api/cat-import/recapture")
def api_cat_import_recapture():
    data = request.json or {}
    folder = _cat_import_dir(data.get("importId", ""))
    if not folder:
        return jsonify({"error": "Unknown import."}), 404
    pdf_path = os.path.join(folder, "source.pdf")
    try:
        value = pdf_extract.recapture_region(
            pdf_path, int(data.get("page", 1)), data.get("rect") or {},
            CAT_IMPORT_DPI, data.get("hint", ""))
        return jsonify({"ok": True, "value": value})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Re-capture failed: {e}"}), 500

@app.post("/api/file-op")
def api_file_op():
    """Copy / move / delete a document (and every sibling file sharing its
    stem — e.g. the .xlsx and .pdf pair) for the All Docs cut/copy/paste/delete menu."""
    data = request.json or {}
    op = data.get("op", "")
    rel = data.get("rel", "")
    folder, src_path = resolve_rel(rel)
    if not src_path:
        return jsonify({"error": "File not found."}), 404
    stem = os.path.splitext(src_path)[0]
    siblings = [f for f in glob.glob(glob.escape(stem) + ".*") if os.path.isfile(f)]

    if op == "delete":
        deleted = 0
        for f in siblings:
            try:
                os.remove(f)
                deleted += 1
            except PermissionError:
                # Windows refuses to delete a file that's open in another
                # program (a PDF sitting open in Acrobat/Edge/whatever the
                # OS default viewer is, most commonly — confirmed directly:
                # this exact error, reproduced by leaving a just-generated
                # CAT PDF open in Acrobat then deleting it from All Docs).
                # Left silently unhandled before, this raised all the way
                # up to an unhandled 500 with an HTML error body, which the
                # front end then failed to JSON-parse — so the delete
                # appeared to just do nothing at all, no error shown
                # anywhere, exactly the "I can't interact with it" report
                # that surfaced this. A clear, actionable message instead
                # of a silent/broken one — any sibling already removed
                # before hitting the locked one stays removed (no undo),
                # same as it always would have on a mid-loop crash.
                return jsonify({"error": f"Can't delete {os.path.basename(f)} — it's currently open in another program (a PDF viewer, most likely). Close it there and try again."}), 409
            except OSError as e:
                return jsonify({"error": f"Couldn't delete {os.path.basename(f)}: {e}"}), 500
        return jsonify({"ok": True, "deleted": deleted})

    if op in ("copy", "move"):
        dest_company = (data.get("dest_company") or "").strip()
        dest_dir = os.path.join(folder, dest_company) if dest_company else folder
        if not os.path.abspath(dest_dir).startswith(os.path.abspath(folder)):
            return jsonify({"error": "Invalid destination."}), 400
        if os.path.abspath(dest_dir) == os.path.abspath(os.path.dirname(src_path)):
            return jsonify({"error": "That's already where this file is."}), 400
        os.makedirs(dest_dir, exist_ok=True)
        targets = [(f, os.path.join(dest_dir, os.path.basename(f))) for f in siblings]
        for _f, t in targets:
            if os.path.exists(t):
                return jsonify({"error": f"{os.path.basename(t)} already exists in the destination."}), 409
        for f, t in targets:
            try:
                (shutil.copy2 if op == "copy" else shutil.move)(f, t)
            except PermissionError:
                # Same open-elsewhere lock as delete above, same fix.
                return jsonify({"error": f"Can't {op} {os.path.basename(f)} — it's currently open in another program (a PDF viewer, most likely). Close it there and try again."}), 409
            except OSError as e:
                return jsonify({"error": f"Couldn't {op} {os.path.basename(f)}: {e}"}), 500
        return jsonify({"ok": True, "count": len(targets)})

    return jsonify({"error": "Unknown operation."}), 400

@app.get("/open-file")
def open_file():
    """Open the generated file in the OS default app (Excel)."""
    name = request.args.get("name", "")
    _folder, path = resolve_rel(name)
    if not path:
        return jsonify({"ok": False, "error": "File not found."}), 404
    import subprocess, sys
    try:
        if sys.platform.startswith("win"): os.startfile(path)
        elif sys.platform == "darwin": subprocess.Popen(["open", path])
        else: subprocess.Popen(["xdg-open", path])
    except Exception as e:
        return jsonify({"ok": False, "error": f"Could not open that file: {e}"})
    return jsonify({"ok": True})

@app.get("/")
def home():
    return Response(PAGE, mimetype="text/html")

# ---------------------------------------------------------------- the UI
PAGE = r"""<!DOCTYPE html><html lang=en><head><meta charset=utf-8>
<meta name=viewport content="width=device-width,initial-scale=1">
<title>Office Tool</title>
<script>(function(){try{var t=localStorage.getItem('theme');if(t==='light'||t==='dark')document.documentElement.setAttribute('data-theme',t)}catch(e){}})()</script>
<style>
:root{
  --ink:#1a1d24;--amber:#e2952c;--amber2:#c97d16;--line:#e4e1da;--muted:#74716a;--canvas:#f5f4f0;--tint:#fbf2e3;
  --card-bg:#fff;--border:#cfccc3;--panel-bg:#ecebe6;--surface-2:#faf9f6;--seg-bg:#eceae4;
  --brand-dark:#1a1d24;
  --danger:#b91c28;--danger-bg:#fbe3e3;--danger-line:#f0c4c4;
  --success:#1a7a34;--success-bg:#e3f5e6;--success-line:#c3e6c9;
  --info:#1a4d8f;--info-bg:#e3edfb;
  --warning:#8a5a10;--warning-bg:#fdf3dd;--warning-line:#f0dca8;
  --r-sm:8px;--r-md:12px;--r-lg:16px;--r-xl:20px;
  --shadow-sm:0 1px 2px rgba(20,18,14,.06),0 1px 1px rgba(20,18,14,.04);
  --shadow-md:0 8px 24px rgba(20,18,14,.14);
  --shadow-lg:0 24px 48px rgba(20,18,14,.28);
  --shadow-xl:0 20px 48px rgba(20,18,14,.3);
  --glass-bg:rgba(255,255,255,.72);--glass-border:rgba(20,18,14,.08);
  --glass-blur:blur(20px) saturate(1.7);--glass-rail-bg:rgba(26,29,36,.82);
}
:root[data-theme="dark"]{
  --ink:#eef0f2;--amber:#f0a63e;--amber2:#ffbb5c;--line:#333841;--muted:#9aa1ac;--canvas:#14161a;--tint:#2a2416;
  --card-bg:#1e2126;--border:#3a3f47;--panel-bg:#101215;--surface-2:#262a30;--seg-bg:#262a30;
  --brand-dark:#1a1d24;
  --danger:#ff6b5e;--danger-bg:#3a1e1c;--danger-line:#5c2c28;
  --success:#4ade80;--success-bg:#16301f;--success-line:#1f4a2b;
  --info:#7db2ff;--info-bg:#182a42;
  --warning:#f0c05a;--warning-bg:#3a2e14;--warning-line:#5c481e;
  --shadow-sm:0 1px 2px rgba(0,0,0,.3),0 1px 1px rgba(0,0,0,.2);
  --shadow-md:0 8px 24px rgba(0,0,0,.45);
  --shadow-lg:0 24px 48px rgba(0,0,0,.6);
  --shadow-xl:0 20px 48px rgba(0,0,0,.5);
  --glass-bg:rgba(30,33,38,.72);--glass-border:rgba(255,255,255,.09);
}
@media(prefers-color-scheme:dark){
  :root:not([data-theme="light"]){
    --ink:#eef0f2;--amber:#f0a63e;--amber2:#ffbb5c;--line:#333841;--muted:#9aa1ac;--canvas:#14161a;--tint:#2a2416;
    --card-bg:#1e2126;--border:#3a3f47;--panel-bg:#101215;--surface-2:#262a30;--seg-bg:#262a30;
    --brand-dark:#1a1d24;
    --danger:#ff6b5e;--danger-bg:#3a1e1c;--danger-line:#5c2c28;
    --success:#4ade80;--success-bg:#16301f;--success-line:#1f4a2b;
    --info:#7db2ff;--info-bg:#182a42;
    --warning:#f0c05a;--warning-bg:#3a2e14;--warning-line:#5c481e;
    --shadow-sm:0 1px 2px rgba(0,0,0,.3),0 1px 1px rgba(0,0,0,.2);
    --shadow-md:0 8px 24px rgba(0,0,0,.45);
    --shadow-lg:0 24px 48px rgba(0,0,0,.6);
    --shadow-xl:0 20px 48px rgba(0,0,0,.5);
    --glass-bg:rgba(30,33,38,.72);--glass-border:rgba(255,255,255,.09);
  }
}
*{box-sizing:border-box}body{margin:0;font-family:-apple-system,Segoe UI,Roboto,Arial,sans-serif;background:var(--canvas);color:var(--ink);font-size:14px}
.app{display:flex;min-height:100vh}
/* Scrolls once the nav items outgrow the viewport (every document type is
   its own entry now, so a short window can't fit them all). overflow-x is
   pinned hidden so the column never grows a horizontal bar of its own —
   doubly important now that its width itself animates open on hover. */
.rail{width:64px;background:var(--glass-rail-bg);border-right:1px solid var(--glass-border);color:#fff;display:flex;flex-direction:column;align-items:stretch;padding:16px 8px;gap:4px;position:sticky;top:0;height:100vh;z-index:20;overflow-y:auto;overflow-x:hidden;scrollbar-width:thin;scrollbar-color:rgba(255,255,255,.28) transparent;transition:width .26s cubic-bezier(.32,.08,.24,1),box-shadow .26s ease}
/* Collapsed = icons only. Hovering the bar (or focusing into it via keyboard,
   or having the brand-switch popup pinned open) widens it and reveals labels
   — see .navlabel/.brandswitchlabel below for the label reveal itself. */
.rail:hover,.rail:focus-within,.rail.pinned{width:212px;box-shadow:8px 0 28px rgba(0,0,0,.24)}
.rail::-webkit-scrollbar{width:6px}
.rail::-webkit-scrollbar-track{background:transparent}
.rail::-webkit-scrollbar-thumb{background:rgba(255,255,255,.22);border-radius:3px}
.rail::-webkit-scrollbar-thumb:hover{background:rgba(255,255,255,.38)}
/* Without this the flex column squashes every button to fit instead of
   scrolling, which is what makes the icons/labels collide on short screens. */
.rail>*{flex-shrink:0}
.bulb{width:32px;height:32px;border-radius:50%;background:radial-gradient(circle at 38% 32%,#ffe6b3,var(--amber) 58%,var(--amber2));box-shadow:0 0 16px rgba(226,149,44,.55);transition:background .35s,box-shadow .35s,transform .15s}
.bulb.b-sololuce{background:radial-gradient(circle at 38% 32%,#e6dbff,#8b5cf6 58%,#6427d6);box-shadow:0 0 16px rgba(139,92,246,.55)}
.bulb.b-ads{background:radial-gradient(circle at 38% 32%,#ffd7d9,#e0464f 58%,#ad1b25);box-shadow:0 0 16px rgba(224,70,79,.55)}
.bulb.b-watt{background:radial-gradient(circle at 38% 32%,#cdf9dc,#22b56a 58%,#127a49);box-shadow:0 0 16px rgba(34,181,106,.55)}
.blogo{width:46px;height:30px;border-radius:9px;object-fit:contain;background:#fff;padding:4px;box-shadow:0 0 14px rgba(0,0,0,.4);transition:transform .22s cubic-bezier(.22,.85,.32,1.3),box-shadow .3s}
.blogo.b-artemis{background:#15171b;box-shadow:0 0 14px rgba(226,149,44,.35)}
.brandbtn:hover .blogo{transform:scale(1.05)}
.brandbtn .blogo{margin:0}
.blogomini{width:40px;height:26px;border-radius:7px;object-fit:contain;background:#fff;flex-shrink:0;padding:3px;box-shadow:0 0 8px rgba(0,0,0,.18);transition:transform .2s cubic-bezier(.22,.85,.32,1.3)}
.blogomini.b-artemis{background:#15171b}
.bcard:hover .blogomini{transform:scale(1.05)}
.blogoset{width:40px;height:26px;border-radius:7px;object-fit:contain;background:#fff;padding:3px;box-shadow:0 0 6px rgba(0,0,0,.15)}
.blogoset.b-artemis{background:#15171b}
.rail b{font-size:9px;letter-spacing:.12em;color:#cdc9c0;margin-bottom:12px}
.brandbtn{background:transparent;border:none;color:#fff;cursor:pointer;display:flex;flex-direction:row;align-items:center;justify-content:center;gap:0;width:100%;box-sizing:border-box;padding:10px 0;border-radius:14px;transition:background .15s,transform .15s,padding .24s ease}
.brandbtn:hover{background:rgba(255,255,255,.12)}
.brandbtn:hover .bulb{transform:scale(1.08)}
.brandbtn:active{transform:scale(.96)}
.brandbtn.open{background:rgba(255,255,255,.18)}
.brandbtn .bulb{margin:0;flex-shrink:0}
/* Collapsed: bulb/logo centered, label collapsed to nothing. Expanded (rail
   hovered/focused/pinned): row shifts left and the brand code+chevron fade in
   — see .rail:hover etc. above for what widens the rail itself. */
.rail:hover .brandbtn,.rail:focus-within .brandbtn,.rail.pinned .brandbtn{justify-content:flex-start;gap:12px;padding:10px 13px}
.brandswitchlabel{display:flex;align-items:center;gap:3px;font-size:8.5px;font-weight:700;letter-spacing:.05em;text-transform:uppercase;max-width:0;opacity:0;overflow:hidden;white-space:nowrap;transition:max-width .2s ease,opacity .12s ease}
.rail:hover .brandswitchlabel,.rail:focus-within .brandswitchlabel,.rail.pinned .brandswitchlabel{max-width:140px;opacity:1;transition:max-width .26s ease .05s,opacity .2s ease .1s}
.chev{font-size:7px;transition:transform .22s;display:inline-block}
.brandbtn.open .chev{transform:rotate(180deg)}
/* left offset matches the rail's EXPANDED width (212px)+8px gap: the popup
   only opens via a click on brandbtn, and getting the mouse/focus there
   always means the rail is already expanded (see .pinned below for what
   keeps it that way once the popup is open and the pointer moves off-rail). */
.brandmenu{position:fixed;left:220px;top:90px;width:230px;background:var(--glass-bg);border:1px solid var(--line);border-radius:var(--r-lg);box-shadow:var(--shadow-xl);padding:10px;z-index:150;transform-origin:top left;animation:brandOpen .2s cubic-bezier(.24,.9,.32,1.24)}
.brandmenu.hide{display:none}
@keyframes brandOpen{from{opacity:0;transform:scale(.86) translateY(-10px)}to{opacity:1;transform:scale(1) translateY(0)}}
.bcard{display:flex;align-items:center;gap:12px;padding:9px 11px;border-radius:11px;cursor:pointer;border:1px solid transparent;margin-bottom:5px;animation:bcardIn .3s both;opacity:0}
.bcard:last-child{margin-bottom:0}
.bcard:hover{background:var(--tint);border-color:#f0dfb8}
.bcard.on{background:var(--brand-dark);border-color:var(--brand-dark)}
.bcard.on .bcardname{color:#fff}
.bcard.on .bcardsub{color:#b8b4aa}
.bcard .bulbmini{width:26px;height:26px;border-radius:50%;flex-shrink:0;box-shadow:0 0 8px rgba(0,0,0,.18)}
.bcardname{font-weight:700;font-size:13px;color:var(--ink);line-height:1.3}
.bcardsub{font-size:10px;color:var(--muted);letter-spacing:.03em}
@keyframes bcardIn{from{opacity:0;transform:translateX(-12px)}to{opacity:1;transform:translateX(0)}}
.nav{width:100%;box-sizing:border-box;display:flex;align-items:center;justify-content:center;gap:0;padding:11px 0;border:none;background:transparent;color:#b3afa5;border-radius:13px;font-size:11px;font-weight:700;letter-spacing:.01em;cursor:pointer;transition:background .15s,color .15s,transform .1s,padding .24s ease}
.nav .navicon{width:21px;height:21px;flex-shrink:0;margin:0;transition:transform .2s cubic-bezier(.34,1.56,.64,1)}
.nav:hover .navicon{transform:scale(1.12)}
.nav:hover{background:rgba(255,255,255,.1);color:#fff}
.nav:active{transform:scale(.94)}
.nav.on{background:var(--amber);color:var(--brand-dark)}
.nav.on:hover{background:var(--amber);color:var(--brand-dark)}
/* Collapsed: icon dead-centered, label width-collapsed to invisible. Expanded
   (hover / keyboard focus-within / brand-popup pinned open): row goes
   left-aligned and the label fades+widens in with a slight delay so it
   doesn't feel like it's racing the rail's own width transition. */
.rail:hover .nav,.rail:focus-within .nav,.rail.pinned .nav{justify-content:flex-start;padding:11px 0 11px 15px;gap:13px}
.navlabel{max-width:0;opacity:0;overflow:hidden;white-space:nowrap;transition:max-width .2s ease,opacity .12s ease}
.rail:hover .navlabel,.rail:focus-within .navlabel,.rail.pinned .navlabel{max-width:150px;opacity:1;transition:max-width .28s ease .05s,opacity .22s ease .1s}
@media (prefers-reduced-motion:reduce){
  .rail,.rail:hover,.rail:focus-within,.rail.pinned,.nav,.nav .navicon,.nav:hover .navicon,.navlabel,.rail:hover .navlabel,.rail:focus-within .navlabel,.rail.pinned .navlabel,.brandbtn,.brandswitchlabel,.rail:hover .brandswitchlabel,.rail:focus-within .brandswitchlabel,.rail.pinned .brandswitchlabel{transition-duration:.001ms!important}
}
.main{flex:1;min-width:0}
.watermark{position:fixed;top:0;left:64px;right:0;bottom:0;z-index:-1;display:flex;align-items:center;justify-content:center;pointer-events:none;overflow:hidden}
.watermark img{width:42%;max-width:520px;min-width:260px;opacity:.055;object-fit:contain}
.watermark img.b-artemis{opacity:.16}
.bar{height:56px;background:var(--glass-bg);border-bottom:1px solid var(--glass-border);display:flex;align-items:center;justify-content:space-between;padding:0 20px;position:sticky;top:0;z-index:5}
.bar h1{font-size:19px;font-weight:700;letter-spacing:-.01em;margin:0}
.btn{border:1px solid var(--border);background:var(--card-bg);color:var(--ink);padding:8px 14px;border-radius:8px;font-weight:600;font-size:13px;cursor:pointer;transition:border-color .15s,background .15s}
.btn:hover{border-color:var(--amber)}
.btn.amber{background:var(--amber);border-color:var(--amber);color:var(--brand-dark)}.btn.amber:hover{background:var(--amber2);color:#fff}
.btn.dark{background:var(--brand-dark);color:#fff;border-color:var(--brand-dark)}
/* Which Align/Mask Position button matches the slot's actual current x/y
   (see syncPhotoAdjustAlignHighlight/syncPhotoAdjustMaskAnchorHighlight,
   app.py) — reusing .btn.amber's own accent rather than inventing a new
   color, scoped to just these two grids so it can't affect any other
   plain .btn that happens to pick up an 'on' class elsewhere. */
#photoadjust-align-grid button.on,#photoadjust-mask-anchor-grid button.on{background:var(--amber);border-color:var(--amber);color:var(--brand-dark)}
.wrap{display:grid;grid-template-columns:var(--leftw,540px) 7px minmax(0,1fr);gap:0;height:calc(100vh - 56px);overflow:hidden}
@media(max-width:980px){.wrap{grid-template-columns:1fr;height:auto;overflow:visible}.resizer{display:none}.left,.right{height:auto!important;overflow:visible!important}}
/* Same left-form/right-preview split as .wrap (reuses .left/.right/
   .resizer), and now the same draggable --leftw too — initResizer() is
   scoped per-wrap (via resizer.closest('.wrap,.fcwrap')) so this and
   .wrap's own #resizer can share one CSS var/localStorage key without
   fighting over document.querySelector('.left') picking the wrong one. */
.fcwrap{display:grid;grid-template-columns:var(--leftw,560px) 7px minmax(0,1fr);gap:0;height:calc(100vh - 56px);overflow:hidden}
@media(max-width:980px){.fcwrap{grid-template-columns:1fr;height:auto;overflow:visible}.fcwrap .left,.fcwrap .right{height:auto!important;overflow:visible!important}}
.left{padding:18px 20px;border-right:1px solid var(--line);min-width:0;height:100%;overflow-y:auto}
.resizer{cursor:col-resize;position:relative}
.resizer::after{content:'';position:absolute;top:0;bottom:0;left:3px;width:1px;background:var(--line);transition:background .12s,width .12s,left .12s}
.resizer:hover::after,.resizer.active::after{background:var(--amber);width:3px;left:2px}
.right{padding:18px;background:var(--panel-bg);height:100%;min-width:0;display:flex;flex-direction:column;overflow:hidden}
.card{background:var(--card-bg);border:1px solid var(--line);border-radius:10px;margin-bottom:14px;box-shadow:var(--shadow-sm)}
.ch{padding:11px 14px;border-bottom:1px solid var(--line);font-size:11px;letter-spacing:.07em;text-transform:uppercase;color:var(--muted);font-weight:700;display:flex;justify-content:space-between;align-items:center}
.cb{padding:14px}
/* Sololuce Datasheet's collapsible sections (Basics/Photos/Spec Badges/
   Technical Specifications/Finish Colors/Ordering Table) only grow a drag
   handle when their content actually overflows the default view — a short
   section (e.g. Basics) never gets one, since there's nothing extra to
   reveal. JS (initCatSectionResize) toggles .is-resizable and caps
   max-height at the content's own natural height, so dragging can never
   pull the handle past the real content into empty space (the bug where
   the Ordering Table left a dead gap above Generate PDF). Scoped to
   cat-only cards so no other doc type's cards change. */
.card.cat-only>.cb{overflow:auto;resize:none}
.card.cat-only>.cb.is-resizable{resize:vertical;min-height:120px}
.card.cat-only>.cb.is-resizable::-webkit-resizer{
  background-color:transparent;
  background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='13' height='13'%3E%3Cg stroke='%231f6feb' stroke-width='1.6' stroke-linecap='round'%3E%3Cline x1='11' y1='2' x2='2' y2='11'/%3E%3Cline x1='11' y1='6' x2='6' y2='11'/%3E%3Cline x1='11' y1='10' x2='10' y2='11'/%3E%3C/g%3E%3C/svg%3E");
  background-repeat:no-repeat;
  background-position:center;
  background-size:12px 12px;
}
.seg{display:inline-flex;background:var(--seg-bg);border-radius:9px;padding:3px;flex-wrap:wrap}
.seg button{border:none;background:transparent;padding:7px 13px;border-radius:7px;font-weight:600;color:var(--muted);cursor:pointer;font-size:13px;transition:background .15s,color .15s}
.seg button.on{background:var(--card-bg);color:var(--ink);box-shadow:var(--shadow-sm)}
.dvcb{display:flex;flex-direction:column;gap:10px}
.dvrow{display:flex;align-items:center;gap:10px;flex-wrap:wrap}
.dvcheck{display:flex;align-items:center;gap:6px;font-size:13px;font-weight:600;color:var(--ink);text-transform:none;letter-spacing:0;margin:0;cursor:pointer;white-space:nowrap}
.dvcheck input[type=checkbox]{width:auto;accent-color:var(--amber);cursor:pointer}
.dvseg button{padding:6px 11px;font-size:12px}
.dvrow input:not([type=checkbox]){flex:1;min-width:110px}
.dvrow input:disabled{background:var(--surface-2);color:var(--muted);cursor:not-allowed}
label{display:block;font-size:11px;letter-spacing:.04em;text-transform:uppercase;color:var(--muted);margin:0 0 4px;font-weight:600}
.f{margin-bottom:11px}.g2{display:grid;grid-template-columns:minmax(0,1fr) minmax(0,1fr);gap:10px}
/* minmax(0,1fr), not a flat 1fr 1fr, +.g2>.f{min-width:0} below — same
   "reflow, don't overflow" family of fix as g3's own auto-fit/minmax a
   few lines down (see that rule's own comment for the full story on WHY
   the left form panel needs this at all: it's user-draggable narrow via
   #resizer). g2 keeps its column COUNT fixed at exactly 2 always (unlike
   g3's auto-fit, which drops to fewer columns) — g2 pairs fields that are
   meant to stay side by side (Product Name/Series, Top Left/Top Right,
   Bottom Left/Bottom Right mirroring the printed page's own 2x2 shape),
   and g2 is used far more widely across the whole app than g3's Photos-
   section-only scope, so reflowing it to 1-per-row would be a much
   bigger, riskier behavior change than this narrower fix needs. Confirmed
   as a real bug via an annotated screenshot: at a narrow dragged panel
   width, the "Merge with Bottom Right (wide)" checkbox label — by far the
   longest text in either g2 row, and white-space:nowrap like every
   .dvcheck — had nothing to wrap into under the old flat 1fr 1fr, so its
   own unbreakable min-content width forced the WHOLE grid wider than the
   panel, leaving a blank gap on one side while everything inside still
   read as squeezed/wrapped. minmax(0,1fr) lets the track itself shrink
   past that point; .g2>.f{min-width:0} (same scoping as g3>.f below, not
   a global .f rule, to avoid touching .f's behavior in flex contexts
   elsewhere it's used) is what actually lets the merge label's own text
   wrap onto a 2nd line instead of just overflowing its now-narrower box —
   see that checkbox label's own added white-space:normal for the other
   half of this fix. */
.g2>.f{min-width:0}
/* auto-fit/minmax, not a flat "1fr 1fr 1fr" — g3 is only ever used by the
   Photos section's two 3-across rows (Main/Application/Diagram, then Top
   Left/Top Right/Bottom Left), and the left form panel's own width is
   user-draggable (#resizer, see initResizer — .wrap's --leftw). A fixed
   3-column grid has no floor: dragging the panel narrow just keeps
   squeezing all 3 columns until a column's own content (a slot's name +
   its Placeholder/Reserve checkboxes) can no longer fit on one line and
   has nowhere to go, so it overflows the card sideways instead — the
   panel gets a horizontal scrollbar and the last column's checkbox reads
   truncated ("PLACEHOL…"), confirmed directly by dragging the resizer
   narrow. minmax(140px,1fr) instead gives every column a real floor:
   once 3 of them stop fitting, auto-fit drops to 2 per row, then 1,
   rather than ever shrinking a column below what its own content needs —
   the same "reflow, don't overflow" fix as min-width:0 below, so the
   resizer itself stays fully draggable to any width, it just changes how
   many columns fit instead of breaking. */
.g3{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:8px}
/* Grid items default to min-width:auto (their own content's min-content
   size), NOT 0 — the classic overflow trap already documented elsewhere
   in this codebase (doc-page.js's .frame td/th comment is the same bug in
   a table). Scoped to .g3's own children only, not every .f in the app,
   so this doesn't change any other section's layout. */
.g3>.f{min-width:0}
.setrow{display:flex;gap:8px}.setrow input{flex:1}
input,textarea,select{width:100%;padding:8px 10px;border:1px solid var(--border);border-radius:7px;font-size:13px;font-family:inherit;background:var(--card-bg);color:var(--ink);transition:border-color .15s,box-shadow .15s}
input:focus,textarea:focus,select:focus{outline:none;border-color:var(--amber);box-shadow:0 0 0 3px var(--tint)}
.richbox{width:100%;box-sizing:border-box;padding:8px 10px;border:1px solid var(--border);border-radius:7px;font-size:13px;font-family:inherit;background:var(--card-bg);color:var(--ink);min-height:38px;cursor:text}
.richbox:focus{outline:none;border-color:var(--amber);box-shadow:0 0 0 3px var(--tint)}
.richbox:empty:before{content:attr(data-placeholder);color:var(--muted)}
.richbox.small{min-height:78px}
.richtoolbar{position:fixed;display:none;align-items:center;gap:2px;background:var(--glass-bg);border:1px solid var(--line);border-radius:var(--r-sm);box-shadow:var(--shadow-md);padding:4px;z-index:190;animation:brandOpen .16s cubic-bezier(.24,.9,.32,1.2)}
.rtbtn{border:none;background:transparent;color:var(--ink);width:26px;height:26px;border-radius:6px;font-size:13px;cursor:pointer;display:inline-flex;align-items:center;justify-content:center}
.rtbtn:hover{background:var(--tint)}
.rtbtn.on{background:var(--amber);color:var(--brand-dark)}
.rtsep{width:1px;height:18px;background:var(--line);margin:0 2px}
.rtcolorwrap{position:relative;width:26px;height:26px;border-radius:6px;display:inline-flex;align-items:center;justify-content:center;flex-shrink:0}
.rtcolorwrap:hover{background:var(--tint)}
.rtcoloricon{font-size:15px;pointer-events:none;line-height:1}
.rtcolor{position:absolute;inset:0;width:100%;height:100%;border:none;padding:0;background:none;opacity:0;cursor:pointer}
.rtswatches{display:inline-flex;align-items:center;gap:3px}
.rtswatch{width:17px;height:17px;border-radius:50%;border:1px solid rgba(0,0,0,.18);padding:0;cursor:pointer;flex-shrink:0}
.rtswatch:hover{transform:scale(1.15)}
.autocomplete{position:fixed;background:var(--glass-bg);border:1px solid var(--line);border-radius:var(--r-sm);box-shadow:var(--shadow-md);z-index:190;max-height:220px;overflow:auto;display:none;animation:brandOpen .16s cubic-bezier(.24,.9,.32,1.2)}
.acitem{padding:7px 12px;font-size:13px;cursor:pointer;white-space:nowrap}
.acitem:hover,.acitem.hi{background:var(--tint)}
.clientpicker{position:fixed;display:none;background:var(--glass-bg);border:1px solid var(--line);border-radius:var(--r-md);box-shadow:var(--shadow-xl);padding:10px;z-index:190;width:280px;animation:brandOpen .16s cubic-bezier(.24,.9,.32,1.2)}
.clientpicker input{margin-bottom:8px}
.cplist{max-height:220px;overflow:auto;margin-bottom:8px}
.cpitem{display:flex;align-items:center;gap:8px;padding:7px 8px;border-radius:7px;cursor:pointer}
.cpitem:hover{background:var(--tint)}
.cpph,.cpitem img{width:24px;height:24px;border-radius:50%;object-fit:cover;flex-shrink:0;background:var(--tint);display:flex;align-items:center;justify-content:center;font-size:10.5px;font-weight:700;color:var(--amber2)}
.cpname{font-size:12.5px;font-weight:600;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.countrybtn{display:flex;align-items:center;gap:8px;width:100%;background:var(--card-bg);border:1px solid var(--border);border-radius:8px;padding:9px 11px;font-size:13px;cursor:pointer;text-align:left;color:var(--ink)}
.countrybtn:hover{border-color:var(--amber)}
.countrybtn .flag{font-size:16px}
.countrybtn .ph{color:var(--muted)}
.countrypicker{position:fixed;display:none;background:var(--glass-bg);border:1px solid var(--line);border-radius:var(--r-md);box-shadow:var(--shadow-xl);padding:10px;z-index:230;width:260px;animation:brandOpen .16s cubic-bezier(.24,.9,.32,1.2)}
.countrypicker input{margin-bottom:8px}
.cflag{font-size:15px;width:20px;text-align:center;flex-shrink:0}
.maplocation{border:1px solid var(--line);border-radius:10px;overflow:hidden;margin-top:6px}
.maplocation iframe{width:100%;height:150px;border:none;display:block}
.mapactions{display:flex;gap:8px;padding:8px}
.mapactions button{flex:1;font-size:11.5px;padding:6px 8px}
.mapempty{padding:16px;text-align:center;font-size:12px;color:var(--muted)}
.pbpreview{border:1px solid var(--line);border-radius:8px;background:var(--surface-2);padding:9px 11px;font-size:12.5px;white-space:pre-wrap;line-height:1.5;min-height:20px}
.itemscb{padding:14px 15px}
.itemslist{display:flex;flex-direction:column;gap:18px;margin-bottom:18px}
.itemcard{border:1px solid var(--line);border-radius:10px;padding:11px;background:var(--surface-2)}
.itemcardtop{display:flex;align-items:center;gap:8px;margin-bottom:9px}
.itemcardtop .photobox{width:44px;height:44px;flex-shrink:0}
.itemcardtop .itemphoto,.itemcardtop .phbtn{height:44px}
.itcardtype{width:100px;flex-shrink:0}
.itemcardfind{display:flex;flex-wrap:wrap;align-items:center;gap:6px;margin-bottom:9px}
.ffind{border:1px dashed var(--border);background:var(--card-bg);color:var(--muted);font-size:11.5px;font-weight:600;padding:6px 10px;border-radius:20px;cursor:pointer;font-family:inherit}
.ffind:hover{border-color:var(--amber);color:var(--amber2)}
.itcarddesc{width:100%;min-height:60px;resize:vertical;font-family:inherit;line-height:1.4;margin-bottom:9px;display:block}
.itemcard .richbox{min-height:60px;margin-bottom:9px}
.itemcardmeta{display:flex;gap:8px;flex-wrap:wrap}
.itcardfield{flex:1;min-width:88px}
.itcardfield label{display:block;font-size:9px;text-transform:uppercase;color:var(--muted);font-weight:700;margin-bottom:3px}
.itemcard input,.itemcard select{background:var(--surface-2);border-color:var(--border)}
.additembar{display:block;width:100%;border:2px dashed var(--border);border-radius:10px;background:transparent;color:var(--muted);font-size:13px;font-weight:600;padding:12px;cursor:pointer;transition:border-color .15s,color .15s,background .15s;text-align:center;font-family:inherit}
.additembar:hover{border-color:var(--amber);color:var(--amber2);background:var(--tint)}
.itemcard.selected{border-color:var(--amber);background:var(--tint);box-shadow:0 0 0 1px var(--amber)}
.itemcardsel{width:16px;height:16px;flex-shrink:0;cursor:pointer;accent-color:var(--amber)}
.itemslist:not(.selectmode) .itemcardsel{display:none}
.itemsbulkbar{display:flex;align-items:center;gap:8px;padding:9px 12px;margin-bottom:14px;background:var(--tint);border:1px solid var(--amber);border-radius:8px;font-size:12.5px}
.itemsbulkbar b{color:var(--amber2)}
.itemsbulkbar .sp{flex:1}
.itemsbulkbar button{font-size:12px;padding:6px 11px}
.itemsbulkbar .bulkdel{background:var(--danger-bg);color:var(--danger);border-color:var(--danger-line)}
.itemsbulkbar .bulkdel:hover{background:var(--danger-line)}
.itempastebar{display:block;width:100%;border:2px dashed var(--amber);border-radius:10px;background:transparent;color:var(--amber2);font-size:13px;font-weight:600;padding:12px;cursor:pointer;margin-top:10px;text-align:center;font-family:inherit}
.itempastebar:hover{background:var(--tint)}
.dspill{display:inline-flex;align-items:center;gap:3px;font-size:10.5px;font-weight:600;font-family:inherit;color:var(--amber2);background:var(--tint);border:1px solid #f0dfb8;border-radius:20px;padding:3px 8px;text-decoration:none;white-space:nowrap;cursor:pointer}
.dspill:hover{background:var(--amber);color:var(--brand-dark)}
.dspillicon{padding:3px 7px;font-size:12px}
.rm{border:none;background:var(--danger-bg);color:var(--danger);font-size:15px;font-weight:700;line-height:1;cursor:pointer;padding:4px 7px;border-radius:5px;transition:background .15s}.rm:hover{background:var(--danger-line);color:var(--danger)}
.mlsection{padding:10px 0;border-bottom:1px solid var(--line)}
.mlsection:last-child{border-bottom:none;padding-bottom:0}
.mlsection:first-child{padding-top:0}
.mlhead{display:flex;align-items:center;gap:8px;font-size:12px;font-weight:600;color:var(--ink);cursor:pointer;user-select:none}
.mlhead:hover{color:var(--amber2)}
.mlcount{background:var(--tint);color:var(--amber2);border-radius:20px;padding:1px 8px;font-size:10px;font-weight:700}
.mlchev{margin-left:auto;font-size:10px;transition:transform .15s}
.mlsection.collapsed .mlchev{transform:rotate(-90deg)}
.mlsection.collapsed .mlbody{display:none}
.mlbody{display:flex;flex-wrap:wrap;gap:6px;margin-top:10px}
.mlx{border:none;background:var(--danger-bg);color:var(--danger);font-size:9px;font-weight:700;line-height:1;cursor:pointer;padding:3px 6px;border-radius:10px}
.mlx:hover{background:var(--danger-line)}
.mlempty{font-size:11px;color:var(--muted);font-style:italic}
.mlswatch{width:13px;height:13px;border-radius:3px;border:1px solid var(--line);display:inline-block}
.mlwarn{flex-basis:100%;font-size:11px;color:var(--warning);background:var(--warning-bg);border:1px solid var(--warning-line);border-radius:6px;padding:6px 9px;line-height:1.4}
.alrow{display:flex;gap:8px;padding:6px 0;border-bottom:1px solid var(--line);font-size:11px}
.alrow:last-child{border-bottom:none}
.altime{color:var(--muted);flex:0 0 auto;white-space:nowrap;font-variant-numeric:tabular-nums}
.ordmovebtn{border:none;background:var(--surface-2);color:var(--muted);font-size:8px;line-height:1;cursor:pointer;padding:2px 4px;border-radius:3px;transition:background .15s}.ordmovebtn:hover{background:var(--tint);color:var(--amber2)}.ordmovebtn:disabled{opacity:.3;cursor:default;background:var(--surface-2);color:var(--muted)}.ordmovebtn+.ordmovebtn{margin-top:1px}
.catordfieldicon{border:none;background:none;color:var(--ink);cursor:pointer;padding:3px;border-radius:5px;display:flex;transition:background .12s}.catordfieldicon:hover{background:var(--tint)}.catordfieldicon svg{width:16px;height:16px}
/* Shared drag-to-reorder feedback — every draggable row across Category/
   Section/Index Order, front matter, and the datasheet form's own
   reorderable rows (spec values, ordering table) gets this same class.
   .dragging fades+scales the row actually being picked up (smoothly, via
   the transition, not an instant snap); .dragover-top/.dragover-bottom
   draw an animated amber insert-line on whichever edge of the row the
   cursor is currently over, driven by dragRowOver's cursor-half check —
   see dragRowStart/dragRowOver/dragRowLeave/dragRowEnd below. */
.dragrow{transition:opacity .15s ease,transform .15s ease,box-shadow .12s ease;position:relative}
.dragrow.dragging{opacity:.4;transform:scale(.98)}
.dragrow.dragover-top{box-shadow:inset 0 2px 0 0 var(--amber)}
.dragrow.dragover-bottom{box-shadow:inset 0 -2px 0 0 var(--amber)}
.dragrow.dragover-left{box-shadow:inset 2px 0 0 0 var(--amber)}
.dragrow.dragover-right{box-shadow:inset -2px 0 0 0 var(--amber)}
.draghandle{cursor:grab;color:var(--muted);flex-shrink:0;font-size:13px;line-height:1;transition:color .12s}
.draghandle:active{cursor:grabbing}
.dragrow:hover .draghandle{color:var(--amber2)}
.photobox{position:relative;margin:0;height:78px;background:#e9e7e2;border-radius:7px;display:flex;align-items:center;justify-content:center;padding:3px;box-sizing:border-box}
.photobox.sugg{background:var(--tint);border:1px dashed var(--amber)}
.itemphoto{width:100%;height:100%;object-fit:cover;border-radius:6px;cursor:pointer;border:1px solid var(--border);vertical-align:middle;display:block}
.itemphoto.sugg{border-color:var(--amber)}
.phbtn{position:relative;border:1px solid var(--border);background:var(--card-bg);color:var(--muted);font-size:10.5px;font-weight:700;padding:8px 20px 8px 8px;border-radius:7px;cursor:pointer;width:100%;height:78px;text-align:left}
.phbtn:hover{border-color:var(--amber);color:var(--amber2)}
.phbtn::after{content:'▾';position:absolute;right:7px;top:50%;transform:translateY(-50%);font-size:11px;color:var(--muted)}
.phrm{position:absolute;top:-6px;left:26px;border:none;background:var(--brand-dark);color:#fff;border-radius:50%;width:16px;height:16px;font-size:11px;line-height:1;cursor:pointer;padding:0}
.phrm:hover{background:#b8493a}
.fmphotopreview{display:flex;align-items:center;justify-content:center;padding:8px 12px}
.fmphotopreview img{max-width:150px;max-height:150px;border-radius:8px;border:1px solid var(--line)}
.hist{border:1px solid var(--amber);background:var(--tint);border-radius:10px;margin-top:12px;overflow:hidden}
.hist .hh{padding:9px 13px;display:flex;gap:8px;align-items:center;border-bottom:1px solid #f0e0c2;font-size:12px}
.glow{width:9px;height:9px;border-radius:50%;background:var(--amber);box-shadow:0 0 8px var(--amber)}
.hc{padding:9px 13px;border-bottom:1px solid #f3e6cc;font-size:12px}.hc:last-child{border:none}
.hc .top{display:flex;justify-content:space-between;font-weight:600;margin-bottom:5px}
.hi{display:flex;justify-content:space-between;gap:8px;padding:2px 0;color:#444}
.hi button{border:1px solid var(--border);background:var(--card-bg);color:var(--ink);border-radius:5px;font-size:11px;padding:0 7px;cursor:pointer}
.previewtoolbar{display:flex;align-items:center;justify-content:space-between;gap:10px;margin-bottom:10px;flex-wrap:wrap;flex-shrink:0}
.pvbtns{display:inline-flex;align-items:center;gap:6px}
.pbtn{border:1px solid var(--border);background:var(--card-bg);color:var(--muted);width:26px;height:26px;border-radius:6px;font-size:14px;font-weight:700;cursor:pointer;display:inline-flex;align-items:center;justify-content:center;line-height:1;padding:0}
.pbtn.wide{width:auto;padding:0 11px;font-size:11px;gap:5px;height:26px}
.pbtn:hover{border-color:var(--ink);color:var(--ink)}
.pbtn.on{background:var(--brand-dark);color:#fff;border-color:var(--brand-dark)}
.pbtn:disabled{opacity:.4;cursor:default;pointer-events:none}
.zoomlabel{font-size:11.5px;color:var(--muted);min-width:36px;text-align:center;font-variant-numeric:tabular-nums;display:inline-block}
.previewpane{width:100%;flex:1;min-height:0;border:1px solid var(--line);border-radius:8px;background:var(--canvas);overflow:auto;padding:16px;box-sizing:border-box;position:relative}
.pvpages{display:flex;flex-direction:column;gap:14px;align-items:center;transition:opacity .15s ease}
.pvpages.hide{display:none}
.pvpages.pv-loading{opacity:.45}
.pvrow{display:flex;gap:14px;align-items:flex-start;justify-content:safe center;width:100%}
.pvpage{box-shadow:0 6px 18px rgba(0,0,0,.28);background:#fff;display:block;border-radius:2px;-webkit-user-drag:none;user-select:none;flex-shrink:0;min-width:0}
.previewpane img#previewimg{max-width:100%;height:auto;box-shadow:0 8px 24px rgba(0,0,0,.35);background:#fff;margin:0 auto;display:block;-webkit-user-drag:none;user-select:none}
.previewpane.pan-ready{cursor:grab}
.previewpane.panning{cursor:grabbing}
.previewpane .empty{position:absolute;inset:0;display:flex;align-items:center;justify-content:center;color:var(--border);text-align:center;padding:20px;font-size:13px}
.previewpane .empty.hide{display:none}
.filterbar{display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap;align-items:center}
.filterbar input,.filterbar select{padding:8px 10px}
.clientsgrid{display:grid;grid-template-columns:repeat(auto-fill,minmax(240px,1fr));gap:12px}
.clientcard{position:relative;background:var(--card-bg);border:1px solid var(--line);border-radius:11px;padding:14px;cursor:pointer;box-shadow:var(--shadow-sm);transition:box-shadow .15s,border-color .15s}
.clientcard:hover{border-color:var(--amber);box-shadow:var(--shadow-md)}
.clientcard .cctop{display:flex;gap:10px;align-items:center;margin-bottom:8px;padding-right:26px}
.clientcardedit{position:absolute;top:10px;right:10px;border:1px solid var(--line);background:var(--card-bg);color:var(--muted);width:24px;height:24px;border-radius:7px;font-size:12px;cursor:pointer;display:flex;align-items:center;justify-content:center;padding:0}
.clientcardedit:hover{border-color:var(--amber);color:var(--amber2);background:var(--tint)}
.clientlogo{width:42px;height:42px;border-radius:50%;object-fit:cover;background:var(--tint);flex-shrink:0;border:1px solid var(--line)}
.clientlogo.ph{display:flex;align-items:center;justify-content:center;font-weight:700;color:var(--amber2);font-size:16px}
.clientname{font-weight:700;font-size:13.5px;color:var(--ink);overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.clientmeta{font-size:11.5px;color:var(--muted);line-height:1.5;overflow:hidden;text-overflow:ellipsis;display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical}
.clientactions{display:flex;gap:6px;margin-top:10px}
.clientactions button{flex:1;font-size:11.5px;padding:6px 8px}
.empty-note{color:var(--muted);font-size:13px;text-align:center;padding:40px}
.empty-state{text-align:center;padding:56px 24px;max-width:360px;margin:0 auto}
.empty-state-icon{width:52px;height:52px;margin:0 auto 16px;border-radius:50%;background:var(--tint);display:flex;align-items:center;justify-content:center;color:var(--amber2)}
.empty-state-icon svg{width:24px;height:24px}
.empty-state-title{font-size:14px;font-weight:700;color:var(--ink);margin-bottom:6px}
.empty-state-sub{font-size:12.5px;color:var(--muted);line-height:1.5}
.clientsection{margin-bottom:22px}
.clientsection:last-child{margin-bottom:0}
.clientsectionhead{display:flex;align-items:center;gap:8px;font-size:11px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;color:var(--muted);padding-bottom:7px;margin-bottom:12px;border-bottom:1px solid var(--line);cursor:pointer;user-select:none}
.clientsectionhead:hover{color:var(--ink)}
.clientsectioncount{background:var(--tint);color:var(--amber2);border-radius:20px;padding:1px 8px;font-size:10px;font-weight:700}
.clientsectionchev{margin-left:auto;font-size:10px;transition:transform .15s}
.clientsection.collapsed .clientsectionchev{transform:rotate(-90deg)}
.clientsection.collapsed .clientsgrid{display:none}
.cogroup{position:sticky;top:0;background:var(--canvas);padding:14px 4px 7px;font-size:11px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;color:var(--amber2);z-index:2}
.more{text-align:center;padding:8px;font-size:12px;color:var(--amber2);font-weight:600;cursor:pointer;border:1px dashed #e0c48f;border-radius:var(--r-sm);margin-bottom:8px}
.more:hover{background:var(--tint)}
/* One rounded cluster per company group (iOS grouped-table style) — rows
   inside share the cluster's border/shadow instead of each being its own
   bordered card, separated only by a thin divider. */
.rowgroup{background:var(--card-bg);border:1px solid var(--line);border-radius:var(--r-md);box-shadow:var(--shadow-sm);overflow:hidden;margin-bottom:16px}
.list .row{display:flex;gap:12px;align-items:center;padding:12px 14px;border-bottom:1px solid var(--line);cursor:pointer;transition:background .12s}
.rowgroup .row:last-child{border-bottom:none}
.list .row:hover{background:var(--tint)}
.list .row.inprogress{background:var(--danger-bg)}
.list .row.inprogress:hover{background:var(--danger-bg);filter:brightness(.96)}
.list .row.selected{background:var(--tint);box-shadow:inset 3px 0 0 var(--amber)}
.list .row.selected:hover{background:var(--tint)}
.adsel{width:16px;height:16px;flex-shrink:0;cursor:pointer;accent-color:var(--amber)}
.list:not(.selectmode) .adsel{display:none}
.inprogresspill{background:var(--danger-bg);color:var(--danger)}
.searchwrap{position:relative;display:flex;align-items:center}
.searchwrap .searchicon{position:absolute;left:12px;width:16px;height:16px;color:var(--muted);pointer-events:none}
.filterbar .searchwrap input{padding-left:34px;border-radius:var(--r-xl);background:var(--surface-2)}
.pill{font-size:10px;font-weight:700;padding:3px 8px;border-radius:20px;background:var(--tint);color:var(--amber2)}
.famchip{display:inline-flex;align-items:center;gap:5px;font-size:11.5px;font-weight:600;padding:4px 5px 4px 10px;border-radius:20px;background:var(--tint);color:var(--amber2);margin:0 6px 6px 0}
.famchip button{border:none;background:transparent;cursor:pointer;font-size:13px;line-height:1;color:inherit;padding:3px;border-radius:50%;display:flex;align-items:center;justify-content:center}
.famchip button:hover{background:rgba(0,0,0,.1)}
.rowactions{display:flex;gap:6px;flex-shrink:0}
.subcard{background:var(--card-bg);border:1px solid var(--line);border-radius:11px;padding:14px;margin-bottom:12px;box-shadow:var(--shadow-sm)}
.subtop{display:flex;align-items:center;gap:10px;margin-bottom:8px;flex-wrap:wrap}
.subname{font-weight:700;font-size:13.5px;color:var(--ink)}
.submeta{font-size:11.5px;color:var(--muted)}
.stagebadge{font-size:10px;font-weight:700;padding:3px 9px;border-radius:20px;text-transform:uppercase;letter-spacing:.03em}
.stagebadge.in_progress{background:var(--danger-bg);color:var(--danger)}
.stagebadge.delivered{background:var(--info-bg);color:var(--info)}
.stagebadge.submittal_built{background:var(--success-bg);color:var(--success)}
/* Quotation (New Design) status — colorized everywhere in the app UI (All
   Docs, Build tab, the status-confirm modal) but deliberately NOT in the
   printed document itself, which keeps its own plain navy-outline pill
   (templates_html/quotation.html) unchanged. */
.statuspill{font-size:10px;font-weight:700;padding:3px 9px;border-radius:20px;text-transform:uppercase;letter-spacing:.03em}
.statuspill.Draft{background:var(--seg-bg);color:var(--muted)}
.statuspill.Sent{background:var(--info-bg);color:var(--info)}
.statuspill.Approved{background:var(--success-bg);color:var(--success)}
.statuspill.Revised{background:var(--warning-bg);color:var(--warning)}
#qtn2-statusseg-top button.on[data-s=Draft]{background:var(--muted);color:#fff}
#qtn2-statusseg-top button.on[data-s=Sent]{background:var(--info);color:#fff}
#qtn2-statusseg-top button.on[data-s=Approved]{background:var(--success);color:#fff}
#qtn2-statusseg-top button.on[data-s=Revised]{background:var(--warning);color:#fff}
.substeps{display:flex;gap:8px;flex-wrap:wrap;margin:10px 0}
.substep{border:1px solid var(--line);border-radius:8px;padding:7px 11px;font-size:11.5px;display:flex;align-items:center;gap:6px;background:var(--surface-2)}
.substep b{font-weight:700}
.substep a{color:var(--amber2);text-decoration:none;font-weight:600;cursor:pointer}
.substep a:hover{text-decoration:underline}
.substepdone{color:var(--success)}
.subactions{display:flex;gap:8px;margin-top:10px}
.subactions button{font-size:12px;padding:7px 12px}
.kpirow{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:4px}
.kpitile{background:var(--card-bg);border:1px solid var(--line);border-radius:11px;padding:16px 18px}
.kpilabel{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.04em;color:var(--muted);margin-bottom:6px}
.kpivalue{font-size:24px;font-weight:800;color:var(--ink);font-variant-numeric:tabular-nums}
.chartlegend{display:flex;gap:16px;margin-bottom:10px;font-size:12px;color:var(--muted)}
.chartlegend span{display:inline-flex;align-items:center;gap:6px}
.chartlegend i{width:10px;height:10px;border-radius:2px;display:inline-block}
.chartwrap{position:relative}
.charttooltip{position:absolute;background:var(--brand-dark);color:#fff;font-size:11.5px;padding:6px 10px;border-radius:7px;pointer-events:none;white-space:nowrap;opacity:0;transition:opacity .1s;transform:translate(-50%,-100%);z-index:10}
.charttooltip.show{opacity:1}
.chartbar{cursor:pointer}
.chartbar:hover{opacity:.85}
.companyrow{display:flex;align-items:center;gap:12px;padding:11px 4px;border-bottom:1px solid var(--line)}
.companyrow:last-child{border-bottom:none}
.companyname{font-weight:700;font-size:13px;flex:1}
.companystat{font-size:12.5px;color:var(--muted);width:120px;text-align:right;font-variant-numeric:tabular-nums}
.companybalance{font-size:13px;font-weight:700;width:120px;text-align:right;font-variant-numeric:tabular-nums}
.companyexpand{cursor:pointer;color:var(--amber2);font-size:11px;font-weight:600;flex-shrink:0}
.companyinvoices{padding:0 4px 10px;display:none}
.companyinvoices.open{display:block}
.invoicerow{display:flex;align-items:center;gap:10px;padding:7px 0;font-size:12px;border-top:1px dashed var(--line)}
.invoicerow .muted{flex:1}
.paytogglebtn{font-size:11px;padding:4px 10px;border-radius:14px;border:1px solid var(--line);background:var(--card-bg);cursor:pointer;font-weight:600}
.paytogglebtn.paid{background:var(--success-bg);color:var(--success);border-color:var(--success-line)}
.paytogglebtn.unpaid{background:var(--danger-bg);color:var(--danger);border-color:var(--danger-line)}
.rbtn{border:1px solid var(--border);background:var(--card-bg);color:var(--ink);font-size:11px;font-weight:700;padding:6px 10px;border-radius:7px;cursor:pointer;white-space:nowrap}
.rbtn:hover{border-color:var(--ink)}
.rbtn.cs{background:var(--brand-dark);color:#fff;border-color:var(--brand-dark)}
.rbtn.cs:hover{background:var(--amber2);border-color:var(--amber2)}
.hoverprev{position:fixed;width:40vw;height:56vh;min-width:280px;min-height:220px;background:var(--card-bg);border:1px solid var(--line);border-radius:var(--r-md);box-shadow:var(--shadow-xl);overflow:hidden;pointer-events:none;z-index:80;display:none}
.hoverprev img{width:100%;height:100%;object-fit:contain;background:var(--surface-2)}
.hoverprev .empty,.hoverprev .loading{display:flex;align-items:center;justify-content:center;height:100%;color:var(--muted);font-size:12px;padding:14px;text-align:center}
.csmodal{position:fixed;inset:0;background:rgba(20,18,14,.6);z-index:200;display:flex;flex-direction:column;padding:26px}
.csmodal.hide{display:none}
.csmodalbar{background:var(--card-bg);border-radius:10px 10px 0 0;padding:12px 18px;display:flex;justify-content:space-between;align-items:center;box-shadow:0 -1px 0 var(--line) inset}
.csmodalbar b{font-size:13px}
.csframewrap{flex:1;width:100%;background:#5c584f;border-radius:0 0 10px 10px;overflow:auto;display:flex;align-items:flex-start;justify-content:center;padding:18px}
.csframe{max-width:100%;height:auto;box-shadow:0 8px 24px rgba(0,0,0,.35);background:#fff}
.editmodal{position:fixed;inset:0;background:var(--canvas);z-index:220;display:flex;flex-direction:column}
.editmodal.hide{display:none}
.editmodalbar{background:var(--card-bg);padding:12px 20px;display:flex;justify-content:space-between;align-items:center;box-shadow:0 1px 0 var(--line);flex-shrink:0}
.editmodalbar b{font-size:14px}
.editwarnbanner{background:var(--warning-bg);color:var(--warning);font-size:12px;font-weight:600;padding:8px 20px;border-bottom:1px solid var(--warning-line);flex-shrink:0}
.editmodalbody{flex:1;overflow:auto;min-height:0}
.editmodalbody .wrap{height:100%}
.editmodalbody .right{min-height:100%}
.catimportmodal{position:fixed;inset:0;background:var(--canvas);z-index:230;display:flex;flex-direction:column}
.catimportmodal.hide{display:none}
.catimportbar{background:var(--card-bg);padding:12px 20px;display:flex;justify-content:space-between;align-items:center;box-shadow:0 1px 0 var(--line);flex-shrink:0}
.catimportbar b{font-size:14px}
.catimportbody{flex:1;display:flex;min-height:0}
.catimportleft{flex:1;overflow:auto;padding:20px;background:#5c584f;display:flex;flex-direction:column;align-items:center;gap:16px}
.catimportright{width:380px;flex-shrink:0;overflow:auto;padding:16px;background:var(--card-bg);border-left:1px solid var(--line)}
.impPageWrap{position:relative;background:#fff;box-shadow:0 6px 18px rgba(0,0,0,.28);flex-shrink:0}
.impPageWrap img{display:block;width:100%;user-select:none;-webkit-user-drag:none}
.impBox{position:absolute;border:2px solid #ec6b2f;background:rgba(236,107,47,.14);cursor:move;box-sizing:border-box}
.impBox.selected{border-color:#1f6feb;background:rgba(31,111,235,.18);z-index:5}
.impHandle{position:absolute;width:10px;height:10px;margin:-5px;background:#fff;border:2px solid #ec6b2f;border-radius:50%;box-sizing:border-box}
.impBox.selected .impHandle{border-color:#1f6feb}
.impfield{border:1px solid var(--line);border-radius:8px;padding:10px 12px;margin-bottom:8px;cursor:pointer}
#cat-sec-basics .ch span,#cat-sec-photos .ch span,#cat-sec-badges .ch span,#cat-sec-specs .ch span,#cat-sec-finish .ch span{font-size:12px;color:var(--muted);font-weight:400}
.impfield.selected{border-color:#1f6feb;background:var(--info-bg)}
.impfield .impfieldhead{display:flex;justify-content:space-between;align-items:center;gap:8px;font-size:10.5px;font-weight:700;text-transform:uppercase;letter-spacing:.4px;color:var(--muted);margin-bottom:4px}
.impfield .impfieldval{font-size:12.5px;white-space:pre-line;word-break:break-word}
.impfield .impnoloc{font-size:10.5px;color:var(--danger);margin-top:4px}
.impgrouphead{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:var(--muted);margin:16px 0 6px}
.impgrouphead:first-child{margin-top:0}
.clientmodal{position:fixed;inset:0;background:rgba(20,18,14,.55);z-index:225;display:flex;align-items:center;justify-content:center;padding:20px}
.clientmodal.hide{display:none}
.clientmodalbox{background:var(--glass-bg);border-radius:var(--r-lg);width:100%;max-width:460px;max-height:88vh;overflow:auto;box-shadow:var(--shadow-xl);animation:brandOpen .2s cubic-bezier(.24,.9,.32,1.24)}
.clientmodalbar{padding:16px 20px;border-bottom:1px solid var(--line);display:flex;justify-content:space-between;align-items:center}
.clientmodalbar b{font-size:14px}
.clientmodalbody{padding:20px}
.clientlogopicker{width:76px;height:76px;border-radius:50%;background:var(--tint);border:1.5px dashed #e0c48f;display:flex;align-items:center;justify-content:center;cursor:pointer;margin:0 auto 16px;position:relative;overflow:hidden}
.clientlogopicker img{width:100%;height:100%;object-fit:cover}
.clientlogopicker .lprm{position:absolute;top:1px;right:1px;background:var(--brand-dark);color:#fff;border:none;border-radius:50%;width:18px;height:18px;font-size:11px;cursor:pointer;line-height:1;padding:0}
.filemenu{position:fixed;background:var(--glass-bg);border:1px solid var(--line);border-radius:var(--r-md);box-shadow:var(--shadow-xl);padding:6px;z-index:180;min-width:190px;display:none;animation:brandOpen .16s cubic-bezier(.24,.9,.32,1.2)}
/* Progressive enhancement only — every surface above already has a solid
   translucent background as its real fallback, so browsers without
   backdrop-filter support just keep today's tinted-glass-without-blur look
   instead of anything broken. */
@supports((backdrop-filter:blur(1px)) or (-webkit-backdrop-filter:blur(1px))){
  .rail,.bar,.brandmenu,.filemenu,.clientpicker,.countrypicker,.richtoolbar,.autocomplete,.clientmodalbox{
    -webkit-backdrop-filter:var(--glass-blur);backdrop-filter:var(--glass-blur)
  }
}
.fmi{display:flex;align-items:center;gap:10px;padding:9px 12px;border-radius:8px;cursor:pointer;font-size:13px;font-weight:600;color:var(--ink);white-space:nowrap}
.fmi:hover{background:var(--tint)}
.fmi.danger{color:var(--danger)}
.fmi.danger:hover{background:var(--danger-bg)}
.fmi.disabled{opacity:.4;cursor:default;pointer-events:none}
.fmi .ic{width:16px;text-align:center;font-size:13px;flex-shrink:0}
.fmsep{height:1px;background:var(--line);margin:5px 4px}
.fmtitle{font-size:10px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;color:var(--muted);padding:6px 12px 4px}
.launchergrid{display:grid;grid-template-columns:repeat(4,1fr);gap:22px 8px;margin:0 0 26px}
.launchertile{display:flex;flex-direction:column;align-items:center;gap:9px;border:none;background:none;padding:0;cursor:pointer;font-size:12.5px;font-weight:600;color:var(--ink)}
.launchertile:active .launchertileicon{transform:scale(.92)}
.launchertileicon{width:58px;height:58px;border-radius:var(--r-xl);background:var(--tint);display:flex;align-items:center;justify-content:center;color:var(--amber2);flex-shrink:0;box-shadow:var(--shadow-sm);transition:box-shadow .15s,transform .12s}
.launchertile:hover .launchertileicon{box-shadow:var(--shadow-md);transform:translateY(-2px)}
.launchertileicon svg{width:28px;height:28px}
.mono{font-family:ui-monospace,Menlo,monospace;font-size:12px}
.muted{color:var(--muted)}.hide{display:none}
.toast{position:fixed;bottom:20px;left:50%;transform:translateX(-50%);background:var(--brand-dark);color:#fff;padding:10px 18px;border-radius:9px;font-size:13px;opacity:0;transition:.2s;pointer-events:none}
.toast.show{opacity:1}
.updatedot{position:absolute;top:-2px;right:-3px;width:8px;height:8px;border-radius:50%;background:#e0464f;box-shadow:0 0 0 2px var(--brand-dark);animation:updatePulse 1.8s ease-in-out infinite}
@keyframes updatePulse{0%,100%{transform:scale(1);opacity:1}50%{transform:scale(1.3);opacity:.7}}
.fmupdatever{font-size:12px;color:var(--muted);padding:0 12px 6px}
.fmupdatenotes{font-size:12px;color:var(--ink);padding:0 12px 10px;max-height:140px;overflow-y:auto;white-space:pre-wrap;line-height:1.4}
.loginoverlay{position:fixed;inset:0;z-index:500;display:flex;align-items:center;justify-content:center;background:var(--brand-dark);background-image:radial-gradient(circle at 30% 20%,rgba(226,149,44,.14),transparent 55%)}
.loginoverlay.hide{display:none}
.loginbox{width:320px;background:var(--glass-bg);border:1px solid var(--line);border-radius:var(--r-lg);box-shadow:var(--shadow-xl);padding:28px 26px;animation:brandOpen .25s cubic-bezier(.24,.9,.32,1.24)}
.loginbox h1{margin:0 0 2px;font-size:19px;color:var(--ink)}
.loginbulb{width:36px;height:36px;border-radius:50%;margin:0 auto 14px;background:radial-gradient(circle at 38% 32%,#ffe6b3,var(--amber) 58%,var(--amber2));box-shadow:0 0 16px rgba(226,149,44,.55)}
.usercard{display:flex;align-items:center;gap:10px;padding:9px 11px;border-radius:11px;border:1px solid var(--line);margin-bottom:6px}
.usercard b{font-size:13px;flex:1}
.userbadge{font-size:10px;font-weight:700;letter-spacing:.04em;text-transform:uppercase;padding:2px 7px;border-radius:6px;background:var(--tint);color:var(--muted)}
.userbadge.admin{background:var(--brand-dark);color:#fff}
.usertoolchip{font-size:11px;display:flex;align-items:center;gap:6px}
.cloudphototile{cursor:pointer;border:1px solid var(--line);border-radius:9px;padding:6px;text-align:center;transition:border-color .15s,transform .15s;background:var(--card-bg)}
.cloudphototile:hover{border-color:var(--brand-dark);transform:translateY(-2px)}
.cloudphototile img{width:100%;aspect-ratio:1;object-fit:contain;background:#fff;border-radius:6px;display:block;margin-bottom:5px}
.cloudphototile span{display:block;font-size:10px;color:var(--muted);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.scanpagetile{position:relative;width:64px;height:84px;border:1px solid var(--line);border-radius:6px;overflow:hidden;background:#fff}
.scanpagetile img{width:100%;height:100%;object-fit:cover}
.scanpagetile span{position:absolute;bottom:2px;right:3px;background:rgba(0,0,0,.6);color:#fff;font-size:9px;padding:1px 4px;border-radius:3px}
</style></head><body>
<!-- Shown until /api/current-user confirms a session (or a login succeeds)
     — see checkLogin()/bootApp() near the bottom of the page script. The
     rest of the app (.app below) still renders behind it so nothing has
     to wait on login just to exist in the DOM, but every /api/* call
     401s server-side until logged in (see app.py's before_request), so
     nothing real can happen while this is up. -->
<div class=loginoverlay id=loginoverlay>
  <form class=loginbox onsubmit="return doLogin(event)">
    <div class=loginbulb></div>
    <h1>Office Tool</h1>
    <p class=muted style="margin:0 0 18px;font-size:13px">Sign in to continue</p>
    <div class=f><label>Username</label><input id=login-username autocomplete=username autofocus></div>
    <div class=f><label>Password</label><input id=login-password type=password autocomplete=current-password></div>
    <label class=dvcheck style="text-transform:none;font-size:12.5px;margin:2px 0 10px"><input type=checkbox id=login-remember checked> Remember me on this PC for 30 days</label>
    <p id=login-error class="muted hide" style="color:#e0464f;font-size:12.5px;margin:2px 0 10px"></p>
    <button class="btn dark" style="width:100%" type=submit id=login-submit>Sign In</button>
  </form>
</div>
<div class=app>
 <div class=watermark id=watermark></div>
 <!-- Deliberately a sibling of .rail, not a child: the rail scrolls
      (overflow-y:auto) and also carries a backdrop-filter, which makes it a
      containing block for fixed-position descendants too — so a popup left
      inside it would be clipped at the rail's edge no matter whether it used
      position:absolute or fixed. Living outside, it stays a plain
      viewport-anchored popup, positioned (.brandmenu's left:220px) against
      the rail's expanded width and kept there via .rail.pinned for as long
      as the popup itself is open — see toggleBrandMenu(). -->
 <div class="brandmenu hide" id=brandmenu></div>
 <div class=rail id=rail>
   <button class=brandbtn id=brandbtn onclick="toggleBrandMenu()" title="Switch company">
     <span id=bulb></span>
     <span class=brandswitchlabel><span id=brandbtnlabel>—</span><span class=chev>▾</span></span>
   </button>
   <button class="nav on" id=n-launcher onclick="view('menu')" title="Jump to any tool"><svg class=navicon viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><rect x=3 y=3 width=7 height=7 rx=1.5 /><rect x=14 y=3 width=7 height=7 rx=1.5 /><rect x=3 y=14 width=7 height=7 rx=1.5 /><rect x=14 y=14 width=7 height=7 rx=1.5 /></svg><span class=navlabel>Menu</span></button>
   <button class=nav id=n-qtn2 onclick="openDocType('QTN2')"><svg class=navicon viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M12.5 2H4a2 2 0 0 0-2 2v8.5a2 2 0 0 0 .59 1.41l9.5 9.5a2 2 0 0 0 2.82 0l7.5-7.5a2 2 0 0 0 0-2.82l-9.5-9.5A2 2 0 0 0 12.5 2Z"/><circle cx=7.5 cy=7.5 r=1.5 /></svg><span class=navlabel>Quotation</span></button>
   <button class=nav id=n-inv onclick="openDocType('INV')"><svg class=navicon viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8Z"/><path d="M14 2v6h6"/><path d="m9 15 2 2 4-4"/></svg><span class=navlabel>Tax Invoice</span></button>
   <button class=nav id=n-do onclick="openDocType('DO')"><svg class=navicon viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M21 8a2 2 0 0 0-1-1.73l-7-4a2 2 0 0 0-2 0l-7 4A2 2 0 0 0 3 8v8a2 2 0 0 0 1 1.73l7 4a2 2 0 0 0 2 0l7-4A2 2 0 0 0 21 16Z"/><path d="M3.3 7 12 12l8.7-5"/><path d="M12 22V12"/></svg><span class=navlabel>Delivery Order</span></button>
   <button class=nav id=n-exp onclick="openDocType('EXP')"><svg class=navicon viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M5 3h14a1 1 0 0 1 1 1v17l-3-2-3 2-3-2-3 2-3-2V4a1 1 0 0 1 1-1Z"/><path d="M8 8h8M8 12h5"/></svg><span class=navlabel>Expense Report</span></button>
   <button class="nav hide" id=n-catbuild onclick="openDocType('CAT')"><svg class=navicon viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8Z"/><path d="M14 2v6h6"/><path d="M9 15h6M9 11h3"/></svg><span class=navlabel>Sololuce Datasheets</span></button>
   <button class="nav hide" id=n-fullcatalog onclick="launcherGoFullCatalog()"><svg class=navicon viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M12 2 2 7l10 5 10-5-10-5Z"/><path d="m2 17 10 5 10-5"/><path d="m2 12 10 5 10-5"/></svg><span class=navlabel>Full Catalog</span></button>
   <button class=nav id=n-all onclick="view('all')"><svg class=navicon viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M3 7a2 2 0 0 1 2-2h4l2 2h8a2 2 0 0 1 2 2v8a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2Z"/></svg><span class=navlabel>All Docs</span></button>
   <button class=nav id=n-submissions onclick="view('submissions')"><svg class=navicon viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M9 11l3 3L22 4"/><path d="M21 12v7a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h11"/></svg><span class=navlabel>Submissions</span></button>
   <button class=nav id=n-statement onclick="view('statement')"><svg class=navicon viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><line x1=12 y1=20 x2=12 y2=10 /><line x1=18 y1=20 x2=18 y2=4 /><line x1=6 y1=20 x2=6 y2=16 /></svg><span class=navlabel>Statement</span></button>
   <div style="flex:1"></div>
   <button class=nav id=n-clients onclick="view('clients')"><svg class=navicon viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><circle cx=12 cy=8 r=4 /><path d="M4 21a8 8 0 0 1 16 0"/></svg><span class=navlabel>Clients</span></button>
   <button class=nav id=n-settings onclick="view('settings')"><svg class=navicon viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><circle cx=12 cy=12 r=3 /><path d="M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 1 1-2.83 2.83l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-4 0v-.09A1.65 1.65 0 0 0 9 19.4a1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 1 1-2.83-2.83l.06-.06a1.65 1.65 0 0 0 .33-1.82 1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1 0-4h.09A1.65 1.65 0 0 0 4.6 9a1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 1 1 2.83-2.83l.06.06a1.65 1.65 0 0 0 1.82.33H9a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 4 0v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 1 1 2.83 2.83l-.06.06a1.65 1.65 0 0 0-.33 1.82V9a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 0 4h-.09a1.65 1.65 0 0 0-1.51 1Z"/></svg><span class=navlabel>Settings</span></button>
   <button class=nav id=n-theme onclick="toggleTheme()" title="Switch light/dark theme">
     <svg class=navicon id=themeicon-sun viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><circle cx=12 cy=12 r=4 /><path d="M12 2v2M12 20v2M4.93 4.93l1.41 1.41M17.66 17.66l1.41 1.41M2 12h2M20 12h2M6.34 17.66l-1.41 1.41M19.07 4.93l-1.41 1.41" /></svg>
     <svg class=navicon id=themeicon-moon style="display:none" viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M21 12.79A9 9 0 1 1 11.21 3 7 7 0 0 0 21 12.79Z" /></svg>
     <span id=themelabel class=navlabel>Theme</span>
   </button>
   <!-- Hidden until checkForAppUpdate() (called on launch, then every few
        hours) finds a newer GitHub release than APP_VERSION — see
        update_checker.py. Click opens the shared #filemenu popover with
        version/notes and an install button (same double-click-to-confirm
        pattern used elsewhere in this app — see installUpdate()). -->
   <button class="nav hide" id=n-update onclick="openUpdateMenu(this.getBoundingClientRect())" title="An update is available">
     <span style="position:relative;display:inline-flex">
       <svg class=navicon viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M21 12a9 9 0 1 1-3.02-6.74"/><path d="M21 3v6h-6"/></svg>
       <span class=updatedot></span>
     </span>
     <span class=navlabel>Update</span>
   </button>
   <button class=nav id=n-logout onclick="doLogout()" title="Sign out">
     <svg class=navicon viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M9 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h4"/><path d="M16 17l5-5-5-5"/><path d="M21 12H9"/></svg>
     <span class=navlabel id=logoutlabel>Sign out</span>
   </button>
 </div>
 <div class=main>
  <div class=bar>
    <h1 id=title>Menu</h1>
    <div style="display:flex;gap:8px">
      <!-- Update Center — everyone (not admin-only, per explicit request),
           always visible regardless of whether an update is actually
           available (unlike the rail's own pulsing n-update button, which
           only appears once one's found — this is the always-there entry
           point to check status and change update preferences). -->
      <button type=button class=btn onclick=openUpdateCenter()>Update</button>
      <!-- Global quick-access to Admin Tools (Settings' admin sub-page) —
           lives in the top bar itself (present on every screen, per
           explicit request) rather than buried in a Settings card. Hidden
           by default; only ever shown for role==='admin', see
           applyAccessRestrictions(). -->
      <button type=button class=btn id=admin-tools-btn style="display:none" onclick=openAdminTools()>Admin Tools</button>
    </div>
  </div>

  <!-- MENU (home/landing view — first thing shown on launch) -->
  <div id=v-menu style="padding:24px;max-width:820px;margin:0 auto">
    <div class=fmtitle>Documents</div>
    <div class=launchergrid>
      <button class=launchertile onclick="launcherGoDoc('QTN2')"><span class=launchertileicon><svg viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M12.5 2H4a2 2 0 0 0-2 2v8.5a2 2 0 0 0 .59 1.41l9.5 9.5a2 2 0 0 0 2.82 0l7.5-7.5a2 2 0 0 0 0-2.82l-9.5-9.5A2 2 0 0 0 12.5 2Z"/><circle cx=7.5 cy=7.5 r=1.5 /></svg></span><span>Quotation</span></button>
      <button class=launchertile onclick="launcherGoDoc('INV')"><span class=launchertileicon><svg viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8Z"/><path d="M14 2v6h6"/><path d="m9 15 2 2 4-4"/></svg></span><span>Tax Invoice</span></button>
      <button class=launchertile onclick="launcherGoDoc('DO')"><span class=launchertileicon><svg viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M21 8a2 2 0 0 0-1-1.73l-7-4a2 2 0 0 0-2 0l-7 4A2 2 0 0 0 3 8v8a2 2 0 0 0 1 1.73l7 4a2 2 0 0 0 2 0l7-4A2 2 0 0 0 21 16Z"/><path d="M3.3 7 12 12l8.7-5"/><path d="M12 22V12"/></svg></span><span>Delivery Order</span></button>
      <button class=launchertile onclick="launcherGoDoc('EXP')"><span class=launchertileicon><svg viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M5 3h14a1 1 0 0 1 1 1v17l-3-2-3 2-3-2-3 2-3-2V4a1 1 0 0 1 1-1Z"/><path d="M8 8h8M8 12h5"/></svg></span><span>Expense Report</span></button>
    </div>
    <div class=fmtitle>Sololuce</div>
    <div class=launchergrid>
      <button class=launchertile onclick="launcherGoSololuce()"><span class=launchertileicon><svg viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8Z"/><path d="M14 2v6h6"/><path d="M9 15h6M9 11h3"/></svg></span><span>Sololuce Datasheets</span></button>
      <button class=launchertile onclick="launcherGoFullCatalog()"><span class=launchertileicon><svg viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M12 2 2 7l10 5 10-5-10-5Z"/><path d="m2 17 10 5 10-5"/><path d="m2 12 10 5 10-5"/></svg></span><span>Full Catalog Builder</span></button>
    </div>
    <div class=fmtitle>Records</div>
    <div class=launchergrid>
      <button class=launchertile id=t-all onclick="launcherGo('all')"><span class=launchertileicon><svg viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M3 7a2 2 0 0 1 2-2h4l2 2h8a2 2 0 0 1 2 2v8a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2Z"/></svg></span><span>All Docs</span></button>
      <button class=launchertile id=t-submissions onclick="launcherGo('submissions')"><span class=launchertileicon><svg viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><path d="M9 11l3 3L22 4"/><path d="M21 12v7a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h11"/></svg></span><span>Submissions</span></button>
      <button class=launchertile id=t-statement onclick="launcherGo('statement')"><span class=launchertileicon><svg viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><line x1=12 y1=20 x2=12 y2=10 /><line x1=18 y1=20 x2=18 y2=4 /><line x1=6 y1=20 x2=6 y2=16 /></svg></span><span>Statement</span></button>
      <button class=launchertile id=t-clients onclick="launcherGo('clients')"><span class=launchertileicon><svg viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><circle cx=12 cy=8 r=4 /><path d="M4 21a8 8 0 0 1 16 0"/></svg></span><span>Clients</span></button>
    </div>
    <div class=fmtitle id=t-settings-title>System</div>
    <div class=launchergrid>
      <button class=launchertile id=t-settings onclick="launcherGo('settings')"><span class=launchertileicon><svg viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><circle cx=12 cy=12 r=3 /><path d="M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 1 1-2.83 2.83l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-4 0v-.09A1.65 1.65 0 0 0 9 19.4a1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 1 1-2.83-2.83l.06-.06a1.65 1.65 0 0 0 .33-1.82 1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1 0-4h.09A1.65 1.65 0 0 0 4.6 9a1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 1 1 2.83-2.83l.06.06a1.65 1.65 0 0 0 1.82.33H9a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 4 0v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 1 1 2.83 2.83l-.06.06a1.65 1.65 0 0 0-.33 1.82V9a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 0 4h-.09a1.65 1.65 0 0 0-1.51 1Z"/></svg></span><span>Settings</span></button>
    </div>
  </div>

  <!-- BUILD -->
  <div id=v-build class=hide>
   <div class=wrap id=buildwrap>
    <div class=left>
      <div class=card id=modebar><div class=cb style="display:flex;align-items:center;gap:8px;flex-wrap:wrap">
        <button class="btn hide" id=impbtn style="padding:6px 11px;font-size:12px" onclick="pickCatImportFile()" title="Read a manufacturer's own datasheet PDF and pre-fill this form from it">📄 Import from PDF</button>
        <input type=file id=catimportfile accept=".pdf" style="display:none" onchange="onCatImportFile(this)">
        <button class=btn id=draftsbtn style="padding:6px 11px;font-size:12px" onclick="openDraftsPicker(event)">Drafts</button>
        <button class=btn style="padding:6px 11px;font-size:12px" onclick="startNewDocument()" title="Clear the form and start a brand-new document">+ New</button>
      </div></div>

      <div class=card id=qtn2-statusseg-top><div class=cb>
        <div class=seg>
          <button class=on data-s=Draft onclick="setQtn2Status('Draft')">Draft</button>
          <button data-s=Sent onclick="setQtn2Status('Sent')">Sent</button>
          <button data-s=Approved onclick="setQtn2Status('Approved')">Approved</button>
          <button data-s=Revised onclick="setQtn2Status('Revised')">Revised</button>
          <button data-s=None onclick="setQtn2Status('None')">None</button>
        </div>
      </div></div>

      <div class=card id=headcard><div class=ch>Header</div><div class=cb id=headfields></div></div>

      <div class="card exp-only" id=exp-sec-details>
        <div class=ch>Expense Details</div>
        <div class=cb>
          <div class=f><label>Employee Name</label><div id=exp-employee-wrap></div></div>
          <div class=g2>
            <div class=f><label>Category / Reference</label><div id=exp-category-wrap></div></div>
            <div class=f><label>Currency</label>
              <div class=seg id=exp-currency-seg>
                <button type=button class=on data-c=AED onclick="setExpCurrency('AED')">AED</button>
                <button type=button data-c=USD onclick="setExpCurrency('USD')">Dollar</button>
                <button type=button data-c=__custom__ onclick="setExpCurrency('__custom__')">Custom</button>
              </div>
              <input type=hidden id=exp-currency value=AED>
              <input id=exp-currency-custom class=hide placeholder="e.g. EUR" oninput=onExpCurrencyCustomInput() style="margin-top:8px">
            </div>
          </div>
        </div>
      </div>

      <div class="card exp-only" id=exp-sec-items>
        <div class=ch>Expense Items</div>
        <div class="cb itemscb">
          <div id=explist class=itemslist></div>
          <button type=button class=additembar onclick=addExpRow()>+ Add Expense</button>
        </div>
      </div>

      <div class="card cat-only" id=cat-sec-basics>
        <div class=ch style=cursor:pointer onclick="toggleCatSection('basics')">
          <span style="display:flex;align-items:center;gap:6px">Basics
            <button type=button class=btn onclick="event.stopPropagation();toggleCatSectionInfo('cat-basics-info')" title="What does this do?" style="width:16px;height:16px;padding:0;border-radius:50%;font-size:10px;font-weight:700;line-height:1;display:flex;align-items:center;justify-content:center;text-transform:none;letter-spacing:normal;flex-shrink:0">i</button>
          </span>
          <span id=cat-sec-basics-chev>▾</span>
        </div>
        <div class=cb id=cat-sec-basics-body>
        <div class=g2>
          <div class=f><label>Product Name</label><input id=cat-productname oninput="catUpperCaseInPlace(this);schedulePreview()" onblur="checkCatProductNameDuplicate()" placeholder="e.g. AURA ECO"><div id=cat-productname-warn class="hide" style="color:#b42318;font-size:11px;margin-top:4px"></div></div>
          <div class=f><label>Series / Category</label><div id=cat-series-wrap></div></div>
        </div>
        <div class=g2>
          <div class=f><label>Page Number</label><input id=cat-pagenum type=number min=1 oninput="schedulePreview()" placeholder="1">
            <p class="muted hide cat-basics-info" style="font-size:10px;margin:4px 0 0">Your working estimate of where this datasheet lands in the printed catalog — prefilled from what's already been generated, but you can type over it. Only a guess: the real page number (and this product's colored index tab) is fixed once you run a Full Catalog Build.</p>
          </div>
          <!-- Moved up alongside Page Number (was its own full-width row
               below, leaving this g2's 2nd column empty next to Page
               Number) per explicit request — annotated screenshot pointed
               out the wasted blank grid cell here. -->
          <div class=f><label>Product Type</label>
            <select id=cat-producttype onchange="onCatProductTypeChange();schedulePreview()">
              <option value="">— Select —</option>
              <option value="Outdoor">Outdoor</option>
              <option value="Indoor">Indoor</option>
              <option value="Striplight">Striplight / Neon Flex</option>
            </select>
            <p class="muted hide cat-basics-info" style="font-size:10px;margin:4px 0 0">Drives the Class 1/2/3 badge automatically — Outdoor→Class 1, Indoor→Class 2, Striplight/Neon Flex→Class 3.</p>
          </div>
        </div>
        <!-- Main Product Photo/Application Photo — moved here from the
             Photos section per explicit request ("put it under Product
             Type"). #cat-img-main/#cat-img-lifestyle and their gear-icon
             popovers work identically wherever they sit in the DOM —
             renderCatImages()/openZoneSettings both look these up by id,
             not by any assumption about which section contains them. -->
        <div class=g2>
          <div class=f>
            <label style="margin:0 0 4px;display:flex;align-items:center">Main Product Photo<button type=button class=btn onclick="event.stopPropagation();openZoneSettings('main',this)" title="Placeholder option" style="width:16px;height:16px;padding:0;border-radius:50%;font-size:9px;line-height:1;display:inline-flex;align-items:center;justify-content:center;flex-shrink:0;margin-left:5px">⚙</button></label>
            <div id=cat-img-main></div>
          </div>
          <div class=f>
            <label style="margin:0 0 4px;display:flex;align-items:center">Application Photo<button type=button class=btn onclick="event.stopPropagation();openZoneSettings('lifestyle',this)" title="Placeholder option" style="width:16px;height:16px;padding:0;border-radius:50%;font-size:9px;line-height:1;display:inline-flex;align-items:center;justify-content:center;flex-shrink:0;margin-left:5px">⚙</button></label>
            <div id=cat-img-lifestyle></div>
          </div>
        </div>
        <div class=f><label>Description</label><textarea id=cat-description rows=4 oninput="schedulePreview()" placeholder="Short product description paragraph…"></textarea></div>
        </div>
      </div>

      <div class="card cat-only" id=cat-sec-family>
        <div class=ch style=cursor:pointer onclick="toggleCatSection('family')">
          <span style="display:flex;align-items:center;gap:6px">Family
            <button type=button class=btn onclick="event.stopPropagation();toggleCatSectionInfo('cat-family-info')" title="What does this do?" style="width:16px;height:16px;padding:0;border-radius:50%;font-size:10px;font-weight:700;line-height:1;display:flex;align-items:center;justify-content:center;text-transform:none;letter-spacing:normal;flex-shrink:0">i</button>
          </span>
          <span id=cat-sec-family-chev>▾</span>
        </div>
        <div class=cb id=cat-sec-family-body>
        <div class=f>
          <label class=dvcheck style="text-transform:none;font-size:13px"><input type=checkbox id=cat-family-enabled onchange="onCatFamilyToggle()"> This product is part of a family</label>
          <p class="muted hide cat-family-info" style="font-size:10px;margin:4px 0 0">Family members are clustered together in the printed catalogue, each preceded by its own illustrated divider page — independent of Series/Category, so a family can span multiple categories.</p>
        </div>
        <div id=cat-family-details class=hide>
          <div class=f><label>Family Name</label><div id=cat-family-wrap></div></div>
          <div class=f><label>Linked Products</label>
            <div id=cat-family-chips style="margin-bottom:8px"></div>
            <input id=cat-family-search type=text placeholder="Search generated datasheets by name to add…" oninput="onCatFamilySearch()" autocomplete=off>
            <div id=cat-family-search-results></div>
          </div>
        </div>
        </div>
      </div>

      <div class="card cat-only" id=cat-sec-photos>
        <div class=ch style=cursor:pointer onclick="toggleCatSection('photos')">
          <span style="display:flex;align-items:center;gap:6px">Photos
            <!-- The long "Top Left and Top Right only take up space..."
                 paragraph used to sit permanently at the bottom of this
                 section, costing vertical space on every load whether or
                 not it was ever read. Moved behind this (i) toggle per
                 explicit request — removed from the default view, still
                 one click away, same shape as photoadjust-info-text's own
                 toggle in the Adjust Photo modal (see that button's own
                 comment) and [[feedback_icon_popover_reusable_pattern]]. -->
            <button type=button class=btn onclick="event.stopPropagation();toggleCatSectionInfo('cat-photos-info')" title="What does this do?" style="width:16px;height:16px;padding:0;border-radius:50%;font-size:10px;font-weight:700;line-height:1;display:flex;align-items:center;justify-content:center;text-transform:none;letter-spacing:normal;flex-shrink:0">i</button>
          </span>
          <span id=cat-sec-photos-chev>▾</span>
        </div>
        <div class=cb id=cat-sec-photos-body>
        <!-- Main Product Photo/Application Photo moved out of this section
             entirely — per explicit request ("put it under Product Type")
             they now live in Basics, right below the Page Number/Product
             Type row (see that section).
             The remaining 4 zones are laid out as a real 2x2 TABLE now —
             Top Left/Top Right in one g2 row, Bottom Left/Bottom Right in
             a second g2 row directly below — per explicit request that the
             upload area's own shape MATCH the printed page's own 2x2
             layout (Top Left/Top Right on top, Bottom Left/Bottom Right
             below — see the template's own comment for that grid), rather
             than the previous "Bottom Right alone on its own full-width
             row, then Top Left/Top Right/Bottom Left squeezed into one
             3-column row" arrangement, which didn't read as a 2x2 shape
             at all. Bottom Right (was "Dimension Diagram" — renamed per
             explicit request, see catImgTitle's own comment for the full
             "why"; still the dimension_diagram field under the hood) now
             sits directly opposite Bottom Left, matching the print
             preview's own bottom row exactly. -->
        <p class="muted hide cat-photos-info" style="font-size:11px;margin:0 0 10px">Laid out the same 2x2 shape as the printed page itself — Top Left/Top Right on top, Bottom Left/Bottom Right below. Each zone's gear icon opens its Placeholder checkbox and (where applicable) an optional print caption. Reserve keeps a Top Left/Top Right zone's space even before it has a photo; Merge combines a zone with its neighbor into one wide photo, using the left one's own upload.</p>
        <div class=g2>
          <div class=f>
            <!-- Reserve is the one control THIS row keeps always-visible —
                 label+gear on line 1, Reserve checkbox on line 2, so both
                 columns here share one baseline. -->
            <div style="margin:0 0 4px">
              <label style="margin:0;display:flex;align-items:center">Top Left<button type=button class=btn onclick="event.stopPropagation();openZoneSettings('extra1',this)" title="Placeholder & caption options" style="width:16px;height:16px;padding:0;border-radius:50%;font-size:9px;line-height:1;display:inline-flex;align-items:center;justify-content:center;flex-shrink:0;margin-left:5px">⚙</button></label>
              <label class=dvcheck style="font-size:9.5px;font-weight:600;text-transform:uppercase;letter-spacing:.03em;color:var(--muted);gap:3px;margin-top:3px" title="Keep this zone's space on the sheet even before it has a photo"><input type=checkbox id=cat-img-extra1-show onchange="CAT_IMG.extra1.show=this.checked;schedulePreview()">Reserve</label>
            </div>
            <div id=cat-img-extra1></div>
            <!-- Merge lives on Top Left specifically (not a third, separate
                 toggle) because Top Left's own photo/zoom/pan/mask become
                 the single wide zone's content when merged — Top Right's
                 own upload is simply unused while this is checked, per
                 explicit request for a photo "which will be narrow but
                 long" that a single square-ish zone can't show well. -->
            <!-- white-space:normal overrides .dvcheck's own nowrap, scoped
                 to just this label (not the class globally, which plenty
                 of shorter checkboxes elsewhere still want kept on one
                 line) — "Merge with Top/Bottom Right (wide)" is easily the
                 longest label in this panel, and at a narrow window width
                 nowrap text with nothing to wrap into was forcing the
                 whole .g2 grid wider than its own container instead of
                 letting the label itself break onto a 2nd line — see the
                 .f/.g2 min-width:0/minmax(0,1fr) fix (that class's own
                 comment) for the other half of this bug. -->
            <label class=dvcheck style="font-size:9.5px;font-weight:600;text-transform:uppercase;letter-spacing:.03em;color:var(--muted);gap:3px;margin-top:4px;white-space:normal" title="Combine Top Left and Top Right into one wide zone, using Top Left's own photo — for a panoramic shot too wide/short to fit either square zone alone"><input type=checkbox id=cat-img-extra1-merged onchange="CAT_IMG.extra1.merged=this.checked;renderCatImages();schedulePreview()">Merge with Top Right (wide)</label>
            <!-- Optional, only meaningful once Merge above is on and a
                 photo's been added — explicit request: instead of the
                 merged box always claiming its default height, size it to
                 match the real photo's own shape, freeing up whatever
                 that saves for Bottom Left/Bottom Right below (see
                 autoSizeTopRowToPicture in the template for the actual
                 mechanics). Left visible even when Merge is off (rather
                 than hidden/disabled) so its own state isn't lost if the
                 user unchecks Merge temporarily — same "don't discard a
                 setting just because it's inactive right now" reasoning
                 as Top Right's own upload staying in memory while merged. -->
            <label class=dvcheck style="font-size:9.5px;font-weight:600;text-transform:uppercase;letter-spacing:.03em;color:var(--muted);gap:3px;margin-top:4px;white-space:normal" title="Once merged, size this zone's own box to match the real photo's shape instead of its default height — a naturally short/wide photo then takes up less vertical space, which flows down to make Bottom Left/Bottom Right bigger. No effect unless Merge above is also checked and a photo has been added."><input type=checkbox id=cat-img-extra1-autosize onchange="CAT_IMG.extra1.autosize=this.checked;schedulePreview()">Auto-size to picture (frees space below)</label>
            <!-- Manual override on top of the automatic aspect-ratio-based
                 height — explicit request: auto-size can only ever hug the
                 uploaded PICTURE FILE's own outer pixel dimensions, not
                 whatever's actually drawn inside it — a vendor spec-sheet
                 screenshot with real white margin baked around the product
                 renders auto-sizes to include that margin too (confirmed
                 directly: measured the real generated PDF's own image vs.
                 box geometry and found zero letterboxing — the box already
                 hugs the file exactly; the "extra space" is IN the file).
                 There's no way to detect that from pixel dimensions alone,
                 so instead of trying to be cleverer about it automatically,
                 this just hands the user a direct lever: type a height and
                 it wins outright over the computed one (still clamped to
                 the same 40–220px floor/ceiling the automatic value uses,
                 for the same reason — a typo shouldn't be able to collapse
                 or blow out this row). Blank/0 keeps pure automatic
                 behavior, unchanged from before this existed. -->
            <label style="display:flex;align-items:center;gap:6px;margin-top:4px">
              <span style="font-size:9.5px;font-weight:600;text-transform:uppercase;letter-spacing:.03em;color:var(--muted);white-space:normal">Custom height (px)</span>
              <input type=number min=0 placeholder="auto" id=cat-img-extra1-autosize-h oninput="CAT_IMG.extra1.autosizeH=this.value?parseInt(this.value,10):0;schedulePreview()" style="width:64px;flex:0 0 auto">
            </label>
            </div>
          <div class=f>
            <div style="margin:0 0 4px">
              <label style="margin:0;display:flex;align-items:center">Top Right<button type=button class=btn onclick="event.stopPropagation();openZoneSettings('extra2',this)" title="Placeholder & caption options" style="width:16px;height:16px;padding:0;border-radius:50%;font-size:9px;line-height:1;display:inline-flex;align-items:center;justify-content:center;flex-shrink:0;margin-left:5px">⚙</button></label>
              <label class=dvcheck style="font-size:9.5px;font-weight:600;text-transform:uppercase;letter-spacing:.03em;color:var(--muted);gap:3px;margin-top:3px" title="Keep this zone's space on the sheet even before it has a photo"><input type=checkbox id=cat-img-extra2-show onchange="CAT_IMG.extra2.show=this.checked;schedulePreview()">Reserve</label>
            </div>
            <div id=cat-img-extra2></div></div>
        </div>
        <!-- Bottom Left and Bottom Right — neither has a Reserve checkbox
             of its own (both always show, unlike Top Left/Top Right which
             are optional), so unlike the row above there's no checkbox-
             row-height mismatch to compensate for here at all: no
             invisible spacer needed on either side now that Bottom Left
             is paired with Bottom Right instead of sharing a row with Top
             Left/Top Right (the old 3-column row's real reason for
             needing one in the first place). -->
        <div class=g2 style="margin-top:11px">
          <div class=f>
            <label style="margin:0 0 4px;display:flex;align-items:center">Bottom Left<button type=button class=btn onclick="event.stopPropagation();openZoneSettings('extra3',this)" title="Placeholder & caption options" style="width:16px;height:16px;padding:0;border-radius:50%;font-size:9px;line-height:1;display:inline-flex;align-items:center;justify-content:center;flex-shrink:0;margin-left:5px">⚙</button></label>
            <div id=cat-img-extra3></div>
            <!-- Same merge pattern as Top Left/Top Right above (see that
                 checkbox's own comment) — lives on Bottom Left specifically,
                 using its own photo/zoom/pan/mask as the combined wide
                 zone's content, Bottom Right's own upload simply
                 unused while this is on. -->
            <label class=dvcheck style="font-size:9.5px;font-weight:600;text-transform:uppercase;letter-spacing:.03em;color:var(--muted);gap:3px;margin-top:4px;white-space:normal" title="Combine Bottom Left and Bottom Right into one wide zone, using Bottom Left's own photo — for a panoramic shot too wide/short to fit either square zone alone"><input type=checkbox id=cat-img-extra3-merged onchange="CAT_IMG.extra3.merged=this.checked;renderCatImages();schedulePreview()">Merge with Bottom Right (wide)</label>
          </div>
          <div class=f>
            <label style="margin:0 0 4px;display:flex;align-items:center">Bottom Right<button type=button class=btn onclick="event.stopPropagation();openZoneSettings('diagram',this)" title="Placeholder & caption options" style="width:16px;height:16px;padding:0;border-radius:50%;font-size:9px;line-height:1;display:inline-flex;align-items:center;justify-content:center;flex-shrink:0;margin-left:5px">⚙</button></label>
            <div id=cat-img-diagram></div>
          </div>
        </div>
        </div>
      </div>

      <div class="card cat-only" id=cat-sec-badges>
        <div class=ch style=cursor:pointer onclick="toggleCatSection('badges')">Spec Badges<span id=cat-sec-badges-chev>▾</span></div>
        <div class=cb id=cat-sec-badges-body>
        <div id=cat-badges-rows></div>
        </div>
      </div>


      <div class="clientmodal hide" id=catbadgesmodal>
        <div class=clientmodalbox style="max-width:460px">
          <div class=clientmodalbar><b>Spec Badges</b><button class=btn onclick=closeCatBadgesModal()>Done</button></div>
          <div class=clientmodalbody>
            <div id=cat-badges-modal-rows></div>
          </div>
        </div>
      </div>

      <div class="clientmodal hide" id=catstdfillvaluesmodal>
        <div class=clientmodalbox style="max-width:380px">
          <div class=clientmodalbar><b id=catstdfillvalues-title>Values</b><button class=btn onclick=closeCatStandardFillValuesModal()>Done</button></div>
          <div class=clientmodalbody>
            <div id=catstdfillvalues-rows></div>
          </div>
        </div>
      </div>

      <div class="card cat-only" id=cat-sec-specs>
        <div class=ch style=cursor:pointer onclick="toggleCatSection('specs')">Technical Specifications<span id=cat-sec-specs-chev>▾</span></div>
        <div class=cb id=cat-sec-specs-body>
        <div id=cat-specs-rows></div>
        <datalist id=catspeclabels></datalist>
        <button type=button class=btn style="width:100%;margin-top:2px" onclick=addCatSpec()>+ Add Spec Row</button>
        <div class=f style="margin-top:14px"><label>Note (optional, shown under specs)</label><textarea id=cat-note rows=2 oninput="schedulePreview()" placeholder="e.g. **Note: Sololuce products only employ standard drivers…"></textarea></div>
        </div>
      </div>

      <div class="card cat-only" id=cat-sec-finish>
        <div class=ch style=cursor:pointer onclick="toggleCatSection('finish')">Finish Colors<span id=cat-sec-finish-chev>▾</span></div>
        <div class=cb id=cat-sec-finish-body>
        <div id=cat-finish-rows></div>
        </div>
      </div>

      <div class="card cat-only" id=cat-sec-ordering>
        <div class=ch style=cursor:pointer onclick="toggleCatSection('ordering')">
          <span style="display:flex;align-items:center;gap:6px">Ordering Table
            <button type=button class=btn onclick="event.stopPropagation();toggleCatSectionInfo('cat-ordering-info')" title="What does this do?" style="width:16px;height:16px;padding:0;border-radius:50%;font-size:10px;font-weight:700;line-height:1;display:flex;align-items:center;justify-content:center;text-transform:none;letter-spacing:normal;flex-shrink:0">i</button>
          </span>
          <span id=cat-sec-ordering-chev>▾</span>
        </div>
        <div class=cb id=cat-sec-ordering-body>
        <div class=f><label>Ordering Code Example (auto-generated from the table's first values)</label><input id=cat-ordcode readonly placeholder="Fill in the Ordering Table below to generate this"></div>

        <div class=f><label>Ordering Table</label>
          <p class="muted hide cat-ordering-info" style="font-size:10.5px;margin:0 0 6px">Each row below is one category (Model No., Power, Size…) with its values to the right — click "+ Add Column" to add one at the end, or the thin + between two rows to insert one there — top, middle, or bottom. Each category keeps its own list of values, so one can have more entries than another. Click × to remove a category or value.</p>
          <div style="display:flex;gap:8px;align-items:center;margin-bottom:10px">
            <button type=button class="dspill" onclick="fillCatOrdStandardInfo()" title="Fills each configured field with every one of its own remembered preset values, one per row">Fill Standard Information</button>
            <button type=button class=btn style="font-size:11px;padding:6px 10px" onclick="openCatStandardFillConfig(event)">Configure…</button>
          </div>
          <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:4px;gap:6px">
            <label style="margin:0">Column widths</label>
            <div style="display:flex;gap:6px">
              <button type=button class=btn style="font-size:10px;padding:4px 8px" onclick="autoFitCatOrdColumns()" title="Excel-style autofit: every column left-to-right snaps to exactly what its own longest word/value needs, and the LAST column absorbs whatever's left — it grows or shrinks so the table always still adds up to the full page width">Auto-Fit Columns</button>
              <button type=button class=btn style="font-size:10px;padding:4px 8px" onclick="saveCatOrdWidthsAsStandard()" title="Remember every column's CURRENT width (dragged or not) as the starting point for every future datasheet — still freely adjustable per document afterward">Save as Standard</button>
              <button type=button class=btn style="font-size:10px;padding:4px 8px" onclick="resetCatOrdWidths()" title="Undo any manual dragging on this datasheet — back to the saved standard widths, or each column's own content where no standard is saved">Reset Widths</button>
            </div>
          </div>
          <p class="muted hide cat-ordering-info" style="font-size:10.5px;margin:0 0 6px">Drag a divider to make one column wider and its neighbor narrower — same total table width either way, bounded by the page margins same as the real printed table. "Auto-Fit Columns" is the Excel double-click-to-autofit equivalent: every column but the last snaps to its own content's exact need, left to right, and the last column takes whatever's left over (bigger or smaller). "Save as Standard" remembers the current widths for every future datasheet; "Reset Widths" undoes manual dragging back to that standard. Everything else about the table lives below, unchanged.</p>
          <div id=cat-ord-widths style="overflow-x:auto;margin-bottom:14px"></div>
          <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:4px;gap:6px">
            <label style="margin:0">Row heights</label>
            <button type=button id=cat-ord-alignrows-btn class=btn style="font-size:10px;padding:4px 8px" onclick="toggleCatOrdAlignRows()"></button>
          </div>
          <p class="muted hide cat-ordering-info" style="font-size:10.5px;margin:0 0 10px">A row's height normally comes from its own tallest cell, so a row where every field is short prints visibly shorter than one with a longer wrapped value. "Align Rows" pulls short rows up to match this table's own most common row height — never the other way around, so a row that genuinely needs more room always keeps every line.</p>
          <div id=cat-ord-table style="overflow-x:auto"></div>
        </div>
        </div>
      </div>

      <div class=card id=clientcard><div class=ch>Client / Company <button class="btn" id=clientpickerbtn style="padding:4px 9px;font-size:12px" onclick="openClientPicker(event)">👤 Client…</button></div><div class=cb>
        <div class=f><label>Company (made for)</label>
          <input id=company list=clients oninput="onCompany();schedulePreview()" placeholder="e.g. Resinal Developments">
          <div id=company-rich class="richbox hide" contenteditable data-placeholder="e.g. Resinal Developments" oninput="onCompany();schedulePreview()"></div>
          <datalist id=clients></datalist>
        </div>
        <datalist id=projects></datalist>
        <div id=qtn2-extra>
          <div class=f><label>Attn</label><div id=customer_attn class=richbox contenteditable data-placeholder="e.g. Mr. Rashid Al Marri" oninput=schedulePreview()></div></div>
          <div class=f><label>Address</label><div id=customer_address class=richbox contenteditable data-placeholder="e.g. Al Quoz Industrial Area 1" oninput=schedulePreview()></div></div>
          <div class=g2>
            <div class=f><label>PO Box</label><input id=customer_pobox placeholder="e.g. 12345" oninput=schedulePreview()></div>
            <div class=f><label>City</label><input id=customer_city placeholder="e.g. Dubai" oninput=schedulePreview()></div>
          </div>
          <div class=f><label>Country</label>
            <button type=button class=countrybtn id=customer-countrybtn onclick="openCountryPicker(event,'customer')">
              <span class=flag id=customer-countryflag></span><span id=customer-countrylabel class=ph>Select a country…</span>
            </button>
            <input type=hidden id=customer-country>
          </div>
        </div>
        <div id=histslot></div>
      </div></div>

      <div class=card id=itemscard><div class=ch>Line items</div>
       <div class="cb itemscb">
         <div id=itemsbulkbar class="itemsbulkbar hide">
           <b id=itemsbulkcount></b>
           <span class=sp></span>
           <button type=button class=btn onclick=bulkCloneSelected() title="Duplicate the selected line items">⧉ Clone</button>
           <button type=button class=btn onclick=bulkCutSelected() title="Remove the selected line items and hold them to paste — into this document or a different one">✂ Cut</button>
           <button type=button class="btn bulkdel" onclick=bulkDeleteSelected() title="Delete the selected line items">🗑 Delete</button>
           <button type=button class=btn onclick=exitItemsSelectMode()>Done</button>
         </div>
         <div id=itemslist class=itemslist></div>
         <button type=button class=additembar onclick=addRow()>+ Add Line Item</button>
         <button type=button id=itempastebar class="itempastebar hide" onclick=pasteClipboardItems()></button>
       </div>
      </div>

      <div class=card id=discvat-card><div class=ch>Discount &amp; VAT</div><div class="cb dvcb">
        <div class=dvrow>
          <label class=dvcheck><input type=checkbox id=disc-on onchange="onDiscVatChange()"> Discount</label>
          <div class="seg dvseg" id=disc-modeseg>
            <button class=on data-m=amount onclick="setDiscMode('amount')">Amount / %</button>
            <button data-m=target onclick="setDiscMode('target')">Target Price</button>
          </div>
          <input id=disc-value placeholder="e.g. 10% or 500" oninput=schedulePreview() disabled>
        </div>
        <div class=dvrow>
          <label class=dvcheck><input type=checkbox id=vat-on checked onchange="onDiscVatChange()"> VAT</label>
          <input id=vat-value placeholder="e.g. 5%" value="5%" oninput=schedulePreview()>
        </div>
      </div></div>

      <div class=card id=qtn2-terms-card><div class=ch>Terms &amp; Conditions</div><div class=cb>
        <div class=f><label>Delivery Time</label><div id=terms-delivery-wrap></div></div>
        <div class=f><label>Payment Terms</label>
          <div id=terms-payment-rows></div>
          <button type=button class=btn style="width:100%;margin-top:2px" onclick=addPaymentStage()>+ Add Payment Stage</button>
        </div>
        <div class=f><label>Warranty</label>
          <div class=seg id=terms-warranty-seg>
            <button data-w=3 onclick="setWarranty('3')">3 Years</button>
            <button class=on data-w=5 onclick="setWarranty('5')">5 Years</button>
            <button data-w=7 onclick="setWarranty('7')">7 Years</button>
            <button data-w=10 onclick="setWarranty('10')">10 Years</button>
          </div>
        </div>
      </div></div>

      <div style="display:flex;gap:8px">
        <button class=btn style="flex:0 0 auto" onclick="saveDraftFromForm()" title="Save your progress without generating the final document">Save as Draft</button>
        <button class="btn dark" style="flex:1" id=genbtn onclick=onGenerateClick()>Generate Excel + PDF</button>
      </div>
      <p class="muted" style="font-size:11.5px;margin-top:8px" id=gencaption>Saves both files into your folder, named by the company convention. Preview on the right is the PDF.</p>
    </div>
    <div class=resizer id=resizer></div>
    <div class=right>
      <div class=previewtoolbar>
        <span id=pagecount class=muted style="font-size:11.5px"></span>
        <div class=pvbtns>
          <button class=pbtn id=pm-hand onclick="togglePanTool()" title="Hand tool — drag to move around (or just hold Space)">
            <svg width=13 height=13 viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><polyline points="5 9 2 12 5 15"/><polyline points="9 5 12 2 15 5"/><polyline points="15 19 12 22 9 19"/><polyline points="19 9 22 12 19 15"/><line x1=2 y1=12 x2=22 y2=12 /><line x1=12 y1=2 x2=12 y2=22 /></svg>
          </button>
          <button class="pbtn wide" onclick="fitPreview()" title="Fit the whole page in view, one page at a time — scroll to move page by page">Fit</button>
          <span style="width:4px"></span>
          <button class="pbtn wide" id=pm-single onclick="setPreviewMode('single')" title="Single page">▭ Single</button>
          <button class="pbtn wide" id=pm-double onclick="setPreviewMode('double')" title="Two pages side by side">▭▭ Double</button>
          <span style="width:4px"></span>
          <button class=pbtn onclick="zoomPreview(-10)" title="Zoom out">−</button>
          <span id=zoomlabel class=zoomlabel>100%</span>
          <button class=pbtn onclick="zoomPreview(10)" title="Zoom in">+</button>
        </div>
      </div>
      <div id=previewbox class=previewpane>
        <div id=previewempty class=empty>Rendering live preview…</div>
        <img id=previewimg class=hide>
        <div id=previewpages class="pvpages hide"></div>
      </div>
    </div>
   </div>
  </div>

  <!-- ALL DOCS -->
  <div id=v-all class=hide style="padding:20px;max-width:1100px;margin:0 auto">
    <div class=seg id=alldocstypeseg style="margin-bottom:10px;flex-wrap:wrap">
      <button class=on data-t=all onclick="setAllDocsType('all')">All</button>
      <button data-t=QTN2 onclick="setAllDocsType('QTN2')">Quotation</button>
      <button data-t=INV onclick="setAllDocsType('INV')">Tax Invoice</button>
      <button data-t=DO onclick="setAllDocsType('DO')">Delivery Order</button>
      <button data-t=CAT onclick="setAllDocsType('CAT')">Sololuce Datasheet</button>
      <button data-t=EXP onclick="setAllDocsType('EXP')">Expense Report</button>
      <button data-t=DRAFTS onclick="setAllDocsType('DRAFTS')">Drafts</button>
    </div>
    <div class=filterbar>
      <div class=searchwrap style="flex:1;min-width:200px">
        <svg class=searchicon viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><circle cx=11 cy=11 r=7 /><path d="m21 21-4.3-4.3"/></svg>
        <input id=search placeholder="Search company, project or number…" oninput=renderList()>
      </div>
      <select id=fcompany onchange=onCompanyFilter() style="width:190px"><option value=all>All companies</option></select>
      <select id=frange onchange=onRangePreset() style="width:150px">
        <option value=all>All dates</option>
        <option value=7>Last 7 days</option>
        <option value=30>Last 30 days</option>
        <option value=month>This month</option>
        <option value=year>This year</option>
        <option value=custom>Custom range…</option>
      </select>
      <input id=ffrom type=date onchange=renderList() class=hide style="width:150px">
      <input id=fto type=date onchange=renderList() class=hide style="width:150px">
      <select id=fsort onchange=renderList() style="width:135px"><option value=desc>Newest first</option><option value=asc>Oldest first</option></select>
      <!-- Folder order — which GROUP (company, or for CAT, category) comes
           first, independent of fsort above (which only orders documents
           WITHIN a group). Defaults per type in setAllDocsType (A-Z for
           CAT's category folders, Recent activity for everyone else's
           company folders — the two defaults already in place before this
           was user-facing), but now a real option instead of a hardcoded
           behavior — switchable per explicit request without needing a
           code change to try the other order on any tab. -->
      <!-- "recent" listed first so it's the value in effect if this view
           ever renders before setAllDocsType has run even once (loadIndex
           can be reached straight from view('all') — see that function —
           without going through setAllDocsType) — matches the prior,
           unconditional default every tab had before this was a choice. -->
      <select id=fgroup onchange=onGroupOrderChange() style="width:160px"><option value=recent>Folders: Recent first</option><option value=az>Folders: A-Z</option></select>
      <button class=btn onclick=loadIndex()>Refresh</button>
    </div>
    <div id=alldocsbulkbar class="itemsbulkbar hide">
      <b id=alldocsbulkcount></b>
      <span class=sp></span>
      <button type=button class=btn onclick=bulkCloneAllDocs() title="Duplicate the selected documents as new-numbered copies">⧉ Clone</button>
      <button type=button class=btn onclick=bulkCutAllDocs() title="Mark the selected documents to move into a different company/project group — click Paste here on that group afterward">✂ Cut</button>
      <button type=button class="btn bulkdel" id=alldocsdelbtn onclick=bulkDeleteAllDocs(this) title="Permanently delete the selected documents">🗑 Delete</button>
      <button type=button class=btn onclick=exitAllDocsSelectMode()>Done</button>
    </div>
    <div id=alldocscutbar class="itemsbulkbar hide">
      <b id=alldocscutcount></b>
      <span class=sp></span>
      <span class=muted style="font-size:11.5px">Click "Paste here" on a company group below to move — or</span>
      <button type=button class=btn onclick=clearAllDocsCut()>Cancel Cut</button>
    </div>
    <div id=listbox class=list></div>
  </div>

  <!-- CLIENTS -->
  <div id=v-clients class=hide style="padding:20px;max-width:1000px;margin:0 auto">
    <div class=filterbar>
      <input id=clientsearch placeholder="Search clients…" oninput=renderClientsGrid() style="flex:1;min-width:200px">
      <div class=seg id=clientgroupseg>
        <button class=on data-g=country onclick="setClientGrouping('country')">Country</button>
        <button data-g=city onclick="setClientGrouping('city')">City</button>
      </div>
      <button class="btn" id=clientsimportbtn onclick="importClients()">Import from documents</button>
      <button class="btn dark" onclick="openClientEditor(null)">+ New Client</button>
    </div>
    <div id=clientsgrid></div>
  </div>

  <!-- SUBMISSIONS -->
  <div id=v-submissions class=hide style="padding:20px;max-width:1000px;margin:0 auto">
    <div class=filterbar>
      <input id=subsearch placeholder="Search submissions…" oninput=renderSubmissions() style="flex:1;min-width:200px">
      <button class="btn dark" onclick=openNewSubmission()>+ New Submission</button>
    </div>
    <div id=submissionsgrid></div>
  </div>

  <!-- STATEMENT OF ACCOUNT -->
  <div id=v-statement class=hide style="padding:20px;max-width:1100px;margin:0 auto">
    <div class=kpirow>
      <div class=kpitile><div class=kpilabel>Total Invoiced</div><div class=kpivalue id=kpi-invoiced>AED 0</div></div>
      <div class=kpitile><div class=kpilabel>Total Collected</div><div class=kpivalue id=kpi-collected style="color:var(--success)">AED 0</div></div>
      <div class=kpitile><div class=kpilabel>Total Outstanding</div><div class=kpivalue id=kpi-outstanding style="color:var(--danger)">AED 0</div></div>
    </div>
    <p class=muted style="font-size:11.5px;margin:4px 0 20px">Reflects invoices generated from today onward. Historical invoices are not included, as prior payment records are unavailable.</p>
    <div class=card><div class=ch>Invoiced vs. Collected by Month</div><div class=cb id=statement-chart></div></div>
    <div class=card><div class=ch>Outstanding Client Payments</div><div class=cb>
      <div id=statement-companies></div>
    </div></div>
  </div>

  <!-- SOLOLUCE FULL CATALOG BUILDER -->
  <div id=v-fullcatalog class=hide>
   <div class=fcwrap>
    <div class=left>
    <h2 style="margin:0 0 6px;font-size:18px">Full Catalog Builder</h2>
    <p class=muted style="font-size:12px;margin:0 0 16px">Combines every generated Sololuce Datasheet into one bound, flat PDF catalogue — Index, a divider page per category, then that category's own datasheets (alphabetical, with any product family clustered together), repeated per section — page-numbered and colored-tabbed by category throughout. Anything beyond that (a cover, an introduction, a family's own divider page, an ending) is a file you upload yourself, inserted exactly as given. Nothing here updates automatically — click Build Catalogue whenever you want the book refreshed with whatever's been generated/uploaded since the last build.</p>

    <div class=card><div class=ch>Catalogue Preview</div><div class=cb>
      <div id=fc-summary class=muted style="font-size:12.5px">Loading…</div>
    </div></div>

    <div class=card><div class=ch>Section Order</div><div class=cb>
      <p class=muted style="font-size:11.5px;margin:0 0 8px">The order Outdoor / Indoor / Strip Light &amp; Neon Flex appear in the catalogue. Click a swatch to recolor, click a name to change what actually prints for it.</p>
      <div id=fc-sectionorder></div>
    </div></div>

    <div class=card><div class=ch>Category Order</div><div class=cb>
      <p class=muted style="font-size:11.5px;margin:0 0 8px">The order categories appear within each section — same list as Series/Category everywhere else in the app; adding one here adds it there too. Click a swatch to recolor, click a name to rename (updates any already-generated products too), drag ⠿ to reorder, or tag a Section to group these for easier navigation (grouping only — a product's real section still comes from its own Product Type).</p>
      <div class=setrow style="margin-bottom:8px"><button class=btn style="flex:1" onclick="sortFcCategories('alpha')">A → Z</button><button class=btn style="flex:1" onclick="sortFcCategories('date')">Oldest first</button><button class=btn style="flex:1" onclick="randomizeFcCategoryColors()">🎲 Randomize Colors</button></div>
      <div id=fc-categoryorder></div>
      <div class=setrow style="margin-top:8px"><input id=fc-newcategory placeholder="New category name…"><button class=btn onclick=addFcCategory()>+ Add</button></div>
    </div></div>

    <div class=card><div class=ch>Index Order</div><div class=cb>
      <p class=muted style="font-size:11.5px;margin:0 0 8px">Which products appear in the printed photo-grid Index, and in what order — auto-filled from what's generated, alphabetical until you reorder something below. Hiding a product only removes its Index entry; its own page in the book is untouched.</p>
      <div id=fc-indexorder></div>
    </div></div>

    <div class=card><div class=ch>Cover, Introduction &amp; Ending</div><div class=cb>
      <p class=muted style="font-size:11.5px;margin:0 0 6px">Upload your own already-designed PDF for each — inserted exactly as given, nothing added or rewritten. Leave any unset and the catalogue simply doesn't have that page.</p>
      <p class=muted style="font-size:10.5px;text-transform:uppercase;letter-spacing:.4px;margin:0 0 4px">Always first</p>
      <div id=fc-cover></div>
      <p class=muted style="font-size:10.5px;text-transform:uppercase;letter-spacing:.4px;margin:14px 0 4px;border-top:1px solid var(--border);padding-top:10px">Reorder with the arrows, or add your own custom pages (Warranty, Certifications, Company Profile…)</p>
      <div id=fc-frontmatter></div>
      <div class=setrow style="margin-top:8px"><input id=fc-newfrontmatter placeholder="Custom page name, e.g. Warranty…"><button class=btn onclick=addFcFrontMatter()>+ Add Custom Page</button></div>
      <p class=muted style="font-size:10.5px;text-transform:uppercase;letter-spacing:.4px;margin:14px 0 4px;border-top:1px solid var(--border);padding-top:10px">Always last</p>
      <div id=fc-ending></div>
    </div></div>

    <div class=card><div class=ch>Family Dividers</div><div class=cb>
      <p class=muted style="font-size:11.5px;margin:0 0 6px">Optional, per family — upload a file to introduce that family's products. A family with nothing uploaded still prints its members together, just with no page in front of them.</p>
      <div id=fc-familydividers></div>
    </div></div>

    <div class=card><div class=ch>Output</div><div class=cb>
      <div class=f><label>Save Catalogue To</label><div class=setrow><input id=set-full_catalog_folder placeholder="Folder where the assembled catalogue PDF is saved" onchange=saveFullCatalogFolder()><button class=btn onclick="browseSetting('full_catalog_folder')">Choose…</button></div></div>
      <button class="btn dark" id=fc-buildbtn style="width:100%;margin-top:6px" onclick=buildFullCatalog()>Build Catalogue</button>
      <p class=muted style="font-size:11px;margin:6px 0 0">This can take up to a minute for a large catalogue — the app won't respond to anything else while it runs.</p>
    </div></div>

    <div class=card id=fc-lastbuild-card style="display:none"><div class=ch>Last Build</div><div class=cb>
      <div id=fc-lastbuild></div>
    </div></div>
    </div>
    <div class=resizer id=fcresizer></div>
    <div class=right>
      <div class=previewtoolbar>
        <span id=fc-pagecount class=muted style="font-size:11.5px"></span>
        <div class=pvbtns>
          <button class=pbtn id=fc-pm-hand onclick="fcTogglePanTool()" title="Hand tool — drag to move around (or just hold Space)">
            <svg width=13 height=13 viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round><polyline points="5 9 2 12 5 15"/><polyline points="9 5 12 2 15 5"/><polyline points="15 19 12 22 9 19"/><polyline points="19 9 22 12 19 15"/><line x1=2 y1=12 x2=22 y2=12 /><line x1=12 y1=2 x2=12 y2=22 /></svg>
          </button>
          <button class="pbtn wide" onclick="fcFitPreview()" title="Reset zoom (or Ctrl+scroll to zoom)">Fit</button>
          <span style="width:4px"></span>
          <button class="pbtn wide" id=fc-pm-single onclick="fcSetPreviewMode('single')" title="Single page">▭ Single</button>
          <button class="pbtn wide" id=fc-pm-double onclick="fcSetPreviewMode('double')" title="Two pages side by side">▭▭ Double</button>
          <span style="width:4px"></span>
          <button class=pbtn onclick="fcZoomPreview(-10)" title="Zoom out">−</button>
          <span id=fc-zoomlabel class=zoomlabel>100%</span>
          <button class=pbtn onclick="fcZoomPreview(10)" title="Zoom in">+</button>
          <span style="width:4px"></span>
          <button class=pbtn onclick=fcPrevPage() title="Previous page">‹</button>
          <input id=fc-pagejump type=number min=1 style="width:44px;text-align:center;border:1px solid var(--border);border-radius:6px;height:26px" onchange=fcJumpPage()>
          <button class=pbtn onclick=fcNextPage() title="Next page">›</button>
        </div>
      </div>
      <div id=fc-previewbox class=previewpane>
        <div id=fc-previewempty class=empty>Build the catalogue to preview it here.</div>
        <div id=fc-previewpages class="pvpages hide"></div>
      </div>
    </div>
   </div>
  </div>

  <!-- SETTINGS -->
  <div id=v-settings class=hide style="padding:24px;max-width:640px;margin:0 auto">
   <!-- Two nested sub-pages within Settings itself (not a separate rail
        item) — per explicit request: Admin lives INSIDE Settings, as its
        own page you go into and back out of, appearing only when logged
        in as an admin. showSettingsAdminPanel()/showSettingsMainPanel()
        just toggle which of these two is visible; both stay in the DOM. -->
   <div id=settings-main-panel>
    <div style="display:flex;align-items:center;gap:10px;margin-bottom:16px">
      <span id=set-bulb></span>
      <h2 style="margin:0;font-size:16px">Settings — <span id=set-brandname>Artemis Lightings</span></h2>
    </div>
    <p class=muted style="font-size:12px;margin-top:-10px;margin-bottom:16px">Each company has its own folders — switch brands from the top-left to edit a different one's settings.</p>
    <div class=card><div class=ch>Document folders</div><div class=cb>
      <div class=f><label>Tax Invoices (INV)</label><div class=setrow><input id=set-inv_folder placeholder="e.g. F:\Documents\Invoices"><button class=btn onclick="browseSetting('inv_folder')">Choose…</button></div></div>
      <div class=f><label>Delivery Orders (DO)</label><div class=setrow><input id=set-do_folder placeholder="e.g. F:\Documents\Delivery Orders"><button class=btn onclick="browseSetting('do_folder')">Choose…</button></div></div>
      <div class=f><label>Quotations (PDF)</label><div class=setrow><input id=set-qtn2_folder placeholder="e.g. F:\Documents\Quotations"><button class=btn onclick="browseSetting('qtn2_folder')">Choose…</button></div></div>
      <div class=f><label>Proforma Invoices (PI)</label><div class=setrow><input id=set-pi_folder placeholder="e.g. F:\Documents\Proforma Invoices"><button class=btn onclick="browseSetting('pi_folder')">Choose…</button></div></div>
      <div class=f><label>Payment Receipts (RV)</label><div class=setrow><input id=set-rv_folder placeholder="e.g. F:\Documents\Payment Receipts"><button class=btn onclick="browseSetting('rv_folder')">Choose…</button></div></div>
      <div class=f><label>Credit Notes (CN)</label><div class=setrow><input id=set-cn_folder placeholder="e.g. F:\Documents\Credit Notes"><button class=btn onclick="browseSetting('cn_folder')">Choose…</button></div></div>
      <div class=f><label>Expense Reports</label><div class=setrow><input id=set-expense_folder placeholder="e.g. F:\Documents\Expense Reports"><button class=btn onclick="browseSetting('expense_folder')">Choose…</button></div></div>
      <div class=f><label>Scanned Delivery Orders</label><div class=setrow><input id=set-scanned_do_folder placeholder="e.g. F:\Documents\Scanned DOs"><button class=btn onclick="browseSetting('scanned_do_folder')">Choose…</button></div></div>
      <p class=muted style="font-size:11.5px;margin:0">Where the physically-signed DO ends up after it's scanned back in — used by Submissions to link a delivered order to its invoice.</p>
    </div></div>
    <div class=card><div class=ch>Product catalog</div><div class=cb>
      <div class=f><label>Product Pictures</label><div class=setrow><input id=set-product_photos_folder placeholder="Folder of product PNGs"><button class=btn onclick="browseSetting('product_photos_folder')">Choose…</button></div></div>
      <div class=f><label>Datasheets</label><div class=setrow><input id=set-datasheets_folder placeholder="Folder of product datasheet PDFs"><button class=btn onclick="browseSetting('datasheets_folder')">Choose…</button></div></div>
    </div></div>
    <!-- Free shared photo library (Cloudflare R2 — see photo_store.py).
         No credential fields here at all, by explicit request — the
         connection itself is entirely admin-managed (Admin Tools):
         admins get their own read/write key there, everyone else
         automatically uses a read-only key baked into the app itself
         (photo_store.py's own comment on the two credential tiers). Sync
         just downloads into the Product Pictures folder above, so
         /api/match-photo (used everywhere a line item shows a photo)
         needs zero changes to pick synced photos up. -->
    <div class=card><div class=ch>Shared Product Photos (Cloud)</div><div class=cb>
      <p class=muted style="font-size:11.5px;margin:0 0 8px">Free shared photo library, synced from the cloud into the Product Pictures folder above — every install sees the same photos without a paid server. Connection is set up by your admin; nothing to configure here.</p>
      <div id=photostore-status style="font-size:12.5px" class=muted>Checking status…</div>
      <button class="btn dark" style="width:100%;margin-top:8px" onclick="syncPhotoStore(this)">Sync Now — Download New Photos</button>
    </div></div>
    <div class=card><div class=ch>Product Datasheets (Generated)</div><div class=cb>
      <div class=f><label>Save Sololuce Datasheets To</label><div class=setrow><input id=set-catalogue_folder placeholder="Folder where generated product datasheet PDFs are saved"><button class=btn onclick="browseSetting('catalogue_folder')">Choose…</button></div></div>
      <p class=muted style="font-size:11.5px;margin:0">Separate from the "Datasheets" folder above — that one is the existing lookup catalog scanned for the Quotation builder. This is where the Sololuce Datasheet builder saves the sheets it generates. The catalogue's colored index tabs and page numbers are no longer set here — see Full Catalog Builder, which computes them from the real assembled book.</p>
    </div></div>
    <div class=card><div class=ch>Templates</div><div class=cb>
      <div class=f><label>Templates folder (optional)</label><div class=setrow><input id=set-templates_folder placeholder="Leave blank to use the templates built into the app"><button class=btn onclick="browseSetting('templates_folder')">Choose…</button></div></div>
      <p class=muted style="font-size:11.5px;margin:0">Expects &lt;folder&gt;\&lt;BRAND&gt;\INV.xlsx / DO.xlsx per brand, same layout as the app's bundled templates.</p>
    </div></div>
    <div class=card><div class=ch>Company Profiles</div><div class=cb>
      <div class=f><label>Clients spreadsheet</label><div class=setrow><input id=set-clients_file placeholder="e.g. F:\Documents\Clients.xlsx"><button class=btn onclick="browseSetting('clients_file')">Choose…</button></div></div>
      <p class=muted style="font-size:11.5px;margin:0 0 11px">Your client list lives in this one file, like every other document folder here — point it at an existing spreadsheet to use that, or a new filename to start fresh. Leave blank to keep using the app's built-in storage.</p>
      <button class=btn style="width:100%" onclick="exportAllClients()">Export All Client Profiles (Excel)</button>
    </div></div>
    <button class="btn dark" style="width:100%" onclick=saveSettings()>Save Settings</button>
    <div class=card style="margin-top:14px"><div class=ch>Manage Lists</div><div class=cb>
      <p class=muted style="font-size:11.5px;margin:0 0 4px">Every dropdown across the app that remembers what you type lives here. Removing an item won't change documents already generated with it.</p>
      <div id=managelists-body></div>
    </div></div>
    <div class=card><div class=ch>Activity Log</div><div class=cb>
      <p class=muted style="font-size:11.5px;margin:0 0 4px">Every preset added or removed anywhere in the app, most recent first.</p>
      <div id=auditlog-body style="max-height:320px;overflow:auto"></div>
    </div></div>
   </div>
   <!-- The Admin sub-page — hidden until showSettingsAdminPanel() runs
        (from the entry card above), only ever reachable at all when
        logged in as admin (the entry card itself is admin-only). The two
        tools here are the ones that actually CHANGE the shared bucket
        contents (accounts, the photo library's own files); read-only
        things like R2 connection settings and Sync Now stay on the main
        Settings page since every install, admin or not, needs those. -->
   <div id=settings-admin-panel class=hide>
    <button type=button class=btn style="margin-bottom:16px" onclick="showSettingsMainPanel()">← Back to Settings</button>
    <!-- ONE Cloudflare R2 key, admin-managed — used both by this machine
         directly AND bundled into future builds for everyone else (see
         photo_store.py's own comment). Only this admin-only page can see
         or change it (also enforced at the API level, app.py's
         before_request). -->
    <div class=card><div class=ch>Cloud Storage</div><div class=cb>
      <p class=muted style="font-size:11.5px;margin:0 0 8px" id=ps-hint>Your Cloudflare R2 key — used on this machine directly, and bundled into the app itself so every other install works automatically with nothing for them to set up. Save here, then rebuild + GUPDATE to ship it to everyone.</p>
      <div class=f><label>Account ID</label><input id=ps-account_id placeholder="Cloudflare R2 Account ID"></div>
      <div class=f><label>Bucket name</label><input id=ps-bucket placeholder="e.g. office-tool-photos"></div>
      <div class=f><label>Access Key ID</label><input id=ps-access_key_id></div>
      <div class=f><label id=ps-secret-label>Secret Access Key</label><input id=ps-secret_access_key type=password placeholder="Leave blank to keep the saved key"></div>
      <button class=btn style="width:100%" onclick="savePhotoStoreConfig(this)">Save Cloud Storage Key</button>
      <p class=muted id=ps-bundle-note style="font-size:11px;margin:6px 0 0"></p>
    </div></div>
    <div class=card><div class=ch>Users &amp; Access</div><div class=cb>
      <p class=muted style="font-size:11.5px;margin:0 0 4px">Who can open this app, and — for non-admins — which brand and which tools they're limited to. Changes take effect for that user next time they log in.</p>
      <div id=users-list></div>
      <div class=f style="margin-top:10px"><label>Username</label><input id=user-username></div>
      <div class=f><label id=user-password-label>Password</label><input id=user-password type=password placeholder="Leave blank to keep current password when editing"></div>
      <div class=f><label>Role</label><select id=user-role onchange="renderUserRoleFields()"><option value=user>Limited user</option><option value=admin>Admin (full access)</option></select></div>
      <div id=user-restrict-fields>
        <div class=f><label>Locked to brand</label><select id=user-brand-lock><option value="">— none, sees brand switcher —</option></select></div>
        <div class=f><label>Blocked tools</label><div id=user-blocked-tools style="display:flex;flex-wrap:wrap;gap:10px"></div></div>
      </div>
      <div style="display:flex;gap:8px;margin-top:8px">
        <button class="btn dark" style="flex:1" onclick=saveUserForm()>+ Add / Save User</button>
        <button class=btn onclick=resetUserForm()>Clear</button>
      </div>
      <button class=btn style="width:100%;margin-top:10px" onclick="publishAccounts(this)">Publish Changes to Cloud</button>
      <p class=muted id=users-publish-note style="font-size:11px;margin:6px 0 0"></p>
    </div></div>
    <div class=card><div class=ch>Manage Cloud Photo Library</div><div class=cb>
      <p class=muted style="font-size:11.5px;margin:0 0 8px">Add to or remove from the shared photo library everyone syncs from (Settings > Shared Product Photos). Needs the read/write R2 key — see photo_store.py's own comment on why non-admins get a separate read-only one.</p>
      <div class=f><label>Add individual photos (filename = product code, e.g. STAGNA-MS.png)</label><input type=file id=ps-upload-files multiple accept=".png"></div>
      <button class=btn style="width:100%" onclick="uploadPhotosToStore('ps-upload-files',this)">Upload Files</button>
      <!-- webkitdirectory: lets the native file picker select a whole
           folder (with subfolders) at once — no per-browser 100-file
           limit like the R2 dashboard's own uploader has, and no
           separate tool needed. Each file keeps its relative path (see
           uploadPhotosToStore()) so a folder of subfolders uploads with
           the same structure; the matcher (engine.load_photo_catalog)
           scans recursively so subfolders don't break matching either. -->
      <div class=f style="margin-top:10px"><label>Or add an entire folder at once</label><input type=file id=ps-upload-folder webkitdirectory multiple></div>
      <button class=btn style="width:100%" onclick="uploadPhotosToStore('ps-upload-folder',this)">Upload Folder</button>
      <div id=photostore-upload-progress class="muted hide" style="font-size:11.5px;margin-top:6px"></div>
      <div id=photostore-list style="margin-top:10px;max-height:280px;overflow:auto"></div>
    </div></div>
   </div>
  </div>
 </div>
</div>
<!-- Scan Now — drives a physically-connected scanner via WIA (see
     scanner.py) instead of "scan externally, then Browse to find it".
     Opened from linkScannedDo() (Submissions) in place of the old direct
     browse-scanned-do call; "Choose Existing File Instead" keeps that old
     path available, since not every scan starts from this machine's own
     scanner (e.g. an emailed scan). Deliberately a GLOBAL sibling here —
     not nested inside v-build like an earlier draft had it, which broke
     it: v-build carries .hide outside document-editing views, and
     display:none on an ancestor removes even position:fixed descendants
     from rendering, no matter their own position value. -->
<div class="clientmodal hide" id=scannowmodal>
  <div class=clientmodalbox style="max-width:480px">
    <div class=clientmodalbar><b>Scan Delivery Order</b><button class=btn onclick=closeScanNowModal()>Cancel</button></div>
    <div class=clientmodalbody>
      <div id=scannow-nodevice class="muted hide" style="font-size:12.5px;margin-bottom:10px">No scanner found — check it's connected and turned on, then <a onclick=refreshScannerList() style="cursor:pointer;text-decoration:underline">try again</a>.</div>
      <div id=scannow-picker-wrap class=f style="display:none"><label>Scanner</label><select id=scannow-device></select></div>
      <div id=scannow-status class=muted style="font-size:12.5px;margin:0 0 10px">Checking for a scanner…</div>
      <div id=scannow-pages style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:10px"></div>
      <button type=button class="btn dark" style="width:100%" id=scannow-scan-btn onclick=scanNextPage() disabled>Scan Page</button>
      <div style="display:flex;gap:8px;margin-top:8px">
        <button type=button class=btn style="flex:1" onclick=removeLastScanPage()>Remove Last Page</button>
        <button type=button class="btn dark" style="flex:1" onclick=finalizeScanAndLink()>Save &amp; Link</button>
      </div>
      <p style="text-align:center;margin:14px 0 0"><a onclick=chooseExistingScannedFile() style="cursor:pointer;text-decoration:underline;font-size:12px;color:var(--muted)">Choose an existing file instead…</a></p>
    </div>
  </div>
</div>
<!-- Cloud Photo Library picker — see openCloudPhotoPicker() in the page
     script. Global sibling for the same reason scannowmodal is (see its
     own comment just above) — this used to be nested inside v-build and
     only "worked" by coincidence, since it's currently only ever opened
     from within that view; moved here so it can't silently break the
     moment something opens it from elsewhere. Thumbnails are plain
     <img src="/api/photostore-fetch?..."> (browser-cached); clicking one
     draws it to an offscreen canvas to get the same kind of data: URI a
     local upload produces via FileReader, so CAT_IMG[slot].src never
     needs to know which source a photo came from. -->
<div class="clientmodal hide" id=cloudphotomodal>
  <div class=clientmodalbox style="max-width:640px">
    <div class=clientmodalbar><b>Choose from Cloud Library</b><button class=btn onclick=closeCloudPhotoPicker()>Cancel</button></div>
    <div class=clientmodalbody>
      <input id=cloudphoto-search placeholder="Search by product code…" oninput=renderCloudPhotoGrid() style="width:100%;margin-bottom:12px">
      <div id=cloudphoto-status class=muted style="font-size:12px;margin-bottom:8px"></div>
      <div id=cloudphoto-grid style="display:grid;grid-template-columns:repeat(auto-fill,minmax(96px,1fr));gap:10px"></div>
    </div>
  </div>
</div>
<!-- Update Center — see openUpdateCenter() in the page script. Global
     sibling, same reasoning as scannowmodal/cloudphotomodal just above:
     never nest a modal inside a view container that carries .hide
     outside its own screen. -->
<div class="clientmodal hide" id=updatecentermodal>
  <div class=clientmodalbox style="max-width:420px">
    <div class=clientmodalbar><b>Office Tool Updates</b><button class=btn onclick=closeUpdateCenter()>Close</button></div>
    <div class=clientmodalbody>
      <div style="text-align:center;margin-bottom:14px">
        <div class=muted style="font-size:11px;text-transform:uppercase;letter-spacing:.06em">Current Version</div>
        <div id=uc-version style="font-size:20px;font-weight:700">—</div>
      </div>
      <div id=uc-status class=muted style="font-size:12.5px;text-align:center;margin-bottom:6px">Checking…</div>
      <div id=uc-notes class="muted hide" style="font-size:11.5px;max-height:120px;overflow:auto;margin-bottom:10px;padding:8px;border:1px solid var(--line);border-radius:8px;white-space:pre-wrap"></div>
      <button type=button class=btn style="width:100%" onclick=checkForAppUpdate(true)>Check Now</button>
      <button type=button class="btn dark hide" id=uc-install-btn style="width:100%;margin-top:8px" onclick=installUpdate(this)>Install &amp; Restart</button>
      <div style="border-top:1px solid var(--line);margin:16px 0 12px"></div>
      <label class=dvcheck style="text-transform:none;font-size:12.5px;display:block;margin-bottom:8px"><input type=checkbox id=uc-check-on-start onchange=saveUpdatePrefs()> Check for updates on each start</label>
      <label class=dvcheck style="text-transform:none;font-size:12.5px;display:block"><input type=checkbox id=uc-auto-update onchange=saveUpdatePrefs()> Automatically install updates when found</label>
      <p class=muted style="font-size:11px;margin:8px 0 0">With auto-install on, a found update installs itself (the app briefly restarts) without asking first.</p>
    </div>
  </div>
</div>
<div class=toast id=toast></div>
<div class=hoverprev id=hoverprev></div>
<div class=filemenu id=filemenu></div>
<div class=richtoolbar id=richtoolbar>
  <button type=button class=rtbtn onmousedown="event.preventDefault()" onclick="richCmd('bold')" title="Bold"><b>B</b></button>
  <button type=button class=rtbtn onmousedown="event.preventDefault()" onclick="richCmd('italic')" title="Italic"><i>I</i></button>
  <button type=button class=rtbtn onmousedown="event.preventDefault()" onclick="richCmd('underline')" title="Underline"><u>U</u></button>
  <button type=button class=rtbtn onmousedown="event.preventDefault()" onclick="richCmd('strikeThrough')" title="Strikethrough"><s>S</s></button>
  <span class=rtsep></span>
  <span class=rtcolorwrap title="Text color">
    <span class=rtcoloricon>🎨</span>
    <input type=color class=rtcolor id=richcolor onchange="richCmd('foreColor',this.value);pinRichColor(this.value)" value="#1a1d24">
  </span>
  <span class=rtswatches id=richswatches></span>
  <span class=rtsep></span>
  <button type=button class=rtbtn onmousedown="event.preventDefault()" onclick="richCmd('removeFormat')" title="Clear formatting">✕</button>
</div>
<div class=autocomplete id=autocomplete></div>
<div class=clientpicker id=clientpicker>
  <input id=cp-search placeholder="Search clients…" oninput=renderClientPicker()>
  <div class=cplist id=cp-list></div>
  <button class="btn dark" style="width:100%" onclick=saveCurrentAsClient()>+ Save current info as client</button>
</div>
<div class=countrypicker id=countrypicker>
  <input id=ctp-search placeholder="Search countries…" oninput=renderCountryPicker()>
  <div class=cplist id=ctp-list></div>
</div>
<div class=clientpicker id=draftspicker>
  <input id=dp-search placeholder="Search drafts…" oninput=renderDraftsPicker()>
  <div class=cplist id=dp-list></div>
  <button class="btn dark" style="width:100%" onclick=saveDraftFromForm()>+ Save current as draft</button>
</div>
<div class=clientpicker id=productfinder>
  <input id=pf-search placeholder="Search products…" oninput=renderProductFinder()>
  <div class=cplist id=pf-list></div>
</div>
<div class=clientpicker id=impnewfieldpopover style="width:240px"></div>
<div class=clientpicker id=qtngenmenu style="width:250px">
  <div class=cplist id=qtngen-list></div>
</div>
<div class=clientpicker id=ctxmenu style="width:190px">
  <div class=cplist id=ctxmenu-list></div>
</div>
<div class="csmodal hide" id=csmodal>
  <div class=csmodalbar><b id=csmodaltitle>Company System</b><button class=btn onclick=closeCS()>Close ✕</button></div>
  <div class=csframewrap><img class=csframe id=csframe></div>
</div>
<div class="editmodal hide" id=editmodal>
  <div class=editmodalbar><b id=editmodaltitle>Editing document</b><button class=btn onclick=exitEditMode()>Close ✕</button></div>
  <div class=editwarnbanner>⚠️ You're editing a document that was already created — nothing is saved until you press Generate, and you'll be asked to confirm before it overwrites anything.</div>
  <div class=editmodalbody id=editmodalbody></div>
</div>
<div class="catimportmodal hide" id=catimportmodal>
  <div class=catimportbar>
    <b id=catimporttitle>Import from PDF</b>
    <div style="display:flex;gap:8px;align-items:center">
      <span class=muted id=catimportstatus style="font-size:12px"></span>
      <button class=btn id=impdrawbtn onclick=toggleImpDraw()>+ Draw Box</button>
      <button class="btn dark" onclick=applyCatImport()>Apply to Datasheet</button>
      <button class=btn onclick=closeCatImport()>Close ✕</button>
    </div>
  </div>
  <div class=catimportbody>
    <div class=catimportleft id=catimportleft></div>
    <div class=catimportright id=catimportright></div>
  </div>
</div>
<div class="clientmodal hide" id=resumedraftsmodal>
  <div class=clientmodalbox style="max-width:460px">
    <div class=clientmodalbar><b>Unfinished Drafts</b><button class=btn onclick=closeResumeDrafts()>Not Now</button></div>
    <div class=clientmodalbody>
      <p class=muted style="font-size:12.5px;margin:0 0 10px">Pick up where you left off, or dismiss to start fresh.</p>
      <div class=cplist id=resumedrafts-list></div>
    </div>
  </div>
</div>
<div class="clientmodal hide" id=saveconfirmmodal>
  <div class=clientmodalbox style="max-width:460px">
    <div class=clientmodalbar><b id=saveconfirmtitle>Review changes</b><button class=btn onclick=closeSaveConfirm()>Close ✕</button></div>
    <div class=clientmodalbody id=saveconfirmbody></div>
  </div>
</div>
<div class="clientmodal hide" id=statusgatemodal>
  <div class=clientmodalbox style="max-width:420px">
    <div class=clientmodalbar><b>Confirm Quotation Status</b><button class=btn onclick=closeStatusGate()>Close ✕</button></div>
    <div class=clientmodalbody id=statusgatebody></div>
  </div>
</div>
<div class="clientmodal hide" id=newsubmodal>
  <div class=clientmodalbox style="max-width:560px">
    <div class=clientmodalbar><b id=newsubtitle>New Submission</b><button class=btn onclick=closeNewSubmission()>Close ✕</button></div>
    <div class=clientmodalbody id=newsubbody></div>
  </div>
</div>
<div class="clientmodal hide" id=productbuilder>
  <div class=clientmodalbox style="max-width:420px">
    <div class=clientmodalbar><b id=pb-title>Product Builder</b><button class=btn onclick=closeProductBuilder()>Close ✕</button></div>
    <div class=clientmodalbody id=pb-body></div>
  </div>
</div>
<div class="clientmodal hide" id=clientmodal>
  <div class=clientmodalbox>
    <div class=clientmodalbar><b id=clientmodaltitle>New Client</b><button class=btn onclick=closeClientEditor()>Close ✕</button></div>
    <div class=clientmodalbody>
      <div class=clientlogopicker id=clientlogopicker onclick=pickClientLogo()>
        <span id=clientlogoplaceholder class=muted style="font-size:11px">Add logo</span>
        <img id=clientlogopreview class=hide>
        <button type=button class="lprm hide" id=clientlogorm onclick="event.stopPropagation();removeClientLogo()" title="Remove logo">✕</button>
      </div>
      <input type=file accept=image/* id=clientlogofile class=hide onchange=onClientLogoFile(this)>
      <div class=f><label>Company Name</label><input id=cf-name placeholder="e.g. Resinal Developments"></div>
      <div class=f><label>Section</label><input id=cf-category list=clientcategories placeholder="e.g. Contractor, Consultant, Government…"></div>
      <datalist id=clientcategories></datalist>
      <div class=f><label>Attn</label><input id=cf-attn placeholder="e.g. Mr. Rashid Al Marri"></div>
      <div class=f><label>Address</label><input id=cf-address placeholder="e.g. Al Quoz Industrial Area 1" oninput=refreshClientMapPreview()></div>
      <div class=g2>
        <div class=f><label>PO Box</label><input id=cf-pobox placeholder="e.g. 12345" oninput=refreshClientMapPreview()></div>
        <div class=f><label>City</label><input id=cf-city placeholder="e.g. Dubai" oninput=refreshClientMapPreview()></div>
      </div>
      <div class=f><label>Country</label>
        <button type=button class=countrybtn id=cf-countrybtn onclick="openCountryPicker(event)">
          <span class=flag id=cf-countryflag></span><span id=cf-countrylabel class=ph>Select a country…</span>
        </button>
        <input type=hidden id=cf-country>
      </div>
      <div class=f><label>Location</label>
        <div class=maplocation id=cf-maplocation></div>
      </div>
      <div class=g2>
        <div class=f><label>Phone / Mobile</label><input id=cf-phone placeholder="e.g. +971 50 000 0000"></div>
        <div class=f><label>Landline</label><input id=cf-landline placeholder="e.g. +971 4 000 0000"></div>
      </div>
      <div class=g2>
        <div class=f><label>Email</label><input id=cf-email placeholder="e.g. name@company.com"></div>
        <div class=f><label>Website</label><input id=cf-website placeholder="e.g. www.company.com"></div>
      </div>
      <div class=f><label>Trade License / TRN</label><input id=cf-trn placeholder="e.g. 100294715600003"></div>
      <div class=f><label>Notes</label><textarea id=cf-notes style="min-height:60px"></textarea></div>
      <div style="display:flex;gap:8px;margin-top:6px">
        <button class="btn dark" style="flex:1" onclick=saveClientEditor()>Save Client</button>
        <button class="btn hide" id=cf-delete onclick="deleteClient(currentClientId,this)">Delete</button>
      </div>
    </div>
  </div>
</div>

<div class="clientmodal hide" id=photoadjustmodal>
  <div class=clientmodalbox style="max-width:380px">
    <div class=clientmodalbar>
      <div style="display:flex;align-items:center;gap:6px">
        <b id=photoadjust-title>Adjust Photo</b>
        <!-- "Drag the photo..." used to sit here as a permanent line of text,
             costing vertical space in every single open of this modal even
             though most users only need to read it once. Moved behind this
             small (i) toggle instead — removed from the default view per
             explicit request, but still one click away (never actually
             deleted), same "collapse it, keep it reachable" shape as
             [[feedback_icon_popover_reusable_pattern]] without needing that
             pattern's full name/description/actions popover for a single
             static sentence with no actions of its own. -->
        <button type=button class=btn onclick=togglePhotoAdjustInfo() title="What does this do?" style="width:20px;height:20px;padding:0;border-radius:50%;font-size:11px;font-weight:700;line-height:1;display:flex;align-items:center;justify-content:center;flex-shrink:0">i</button>
      </div>
      <button class=btn onclick=closePhotoAdjust()>Done</button>
    </div>
    <div class=clientmodalbody>
      <p class="muted hide" id=photoadjust-info-text style="font-size:10.5px;margin:0 0 10px">Drag the photo to move it inside the dashed mask — the mask is what actually shows in the PDF.</p>
      <div id=photoadjust-frame style="width:100%;background:#f3f4f6;border-radius:6px;overflow:hidden;display:flex;">
        <div id=photoadjust-mask style="overflow:hidden;position:relative;outline:1.5px dashed #1f6feb;cursor:grab">
          <img id=photoadjust-img style="width:100%;height:100%;object-fit:contain;user-select:none;-webkit-user-drag:none;display:block">
        </div>
      </div>
      <!-- 9-point alignment shortcuts — one click each straight to a corner/
           edge/center pan position, instead of always having to drag by
           hand for an exact edge. Only shown for Dimension Diagram + the 3
           Extra Photo zones (see CAT_IMG_ALIGN_BUTTON_SLOTS), per explicit
           request scoped to those 4 — Main/Application Photo keep drag-only.
           Sets x/y straight to the same 0-100 scale drag already uses
           (photo_cell's own translate() formula in the template already
           guarantees 0/50/100 land flush at each edge/center/edge, never
           short of it or past it, at any zoom — see that macro's own
           comment), so a button press is just a precise, repeatable
           version of the exact same drag a user could already do by hand. -->
      <div class=f id=photoadjust-align-row>
        <label>Align</label>
        <div id=photoadjust-align-grid style="display:grid;grid-template-columns:repeat(3,1fr);gap:4px;max-width:132px">
          <button type=button class=btn data-x=0 data-y=0 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustAlign(0,0)" title="Top left">↖</button>
          <button type=button class=btn data-x=50 data-y=0 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustAlign(50,0)" title="Top center">↑</button>
          <button type=button class=btn data-x=100 data-y=0 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustAlign(100,0)" title="Top right">↗</button>
          <button type=button class=btn data-x=0 data-y=50 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustAlign(0,50)" title="Center left">←</button>
          <button type=button class=btn data-x=50 data-y=50 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustAlign(50,50)" title="Center">•</button>
          <button type=button class=btn data-x=100 data-y=50 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustAlign(100,50)" title="Center right">→</button>
          <button type=button class=btn data-x=0 data-y=100 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustAlign(0,100)" title="Bottom left">↙</button>
          <button type=button class=btn data-x=50 data-y=100 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustAlign(50,100)" title="Bottom center">↓</button>
          <button type=button class=btn data-x=100 data-y=100 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustAlign(100,100)" title="Bottom right">↘</button>
        </div>
      </div>
      <div class=f><label>Zoom <span id=photoadjust-zoom-val class=muted></span></label><input type=range id=photoadjust-zoom min=100 max=300 step=1 style="width:100%" oninput="onPhotoAdjustChange()"></div>
      <div class=f id=photoadjust-mask-row><label>Mask Size <span id=photoadjust-mask-val class=muted></span></label><input type=range id=photoadjust-mask-slider min=40 max=100 step=1 style="width:100%" oninput="onPhotoAdjustChange()"></div>
      <!-- Mask Position — a SECOND alignment grid, easy to confuse with
           Align above it but a genuinely different question: Align pans
           the PHOTO around *inside* the mask; this decides where the mask
           BOX ITSELF sits in its own cell once Mask Size is under 100%.
           Same 4-slot scope as Align (see CAT_IMG_ALIGN_BUTTON_SLOTS),
           shown/hidden together with it in openPhotoAdjust. -->
      <div class=f id=photoadjust-mask-anchor-row>
        <label>Mask Position</label>
        <div id=photoadjust-mask-anchor-grid style="display:grid;grid-template-columns:repeat(3,1fr);gap:4px;max-width:132px">
          <button type=button class=btn data-x=0 data-y=0 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustMaskAnchor(0,0)" title="Top left">↖</button>
          <button type=button class=btn data-x=50 data-y=0 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustMaskAnchor(50,0)" title="Top center">↑</button>
          <button type=button class=btn data-x=100 data-y=0 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustMaskAnchor(100,0)" title="Top right">↗</button>
          <button type=button class=btn data-x=0 data-y=50 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustMaskAnchor(0,50)" title="Center left">←</button>
          <button type=button class=btn data-x=50 data-y=50 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustMaskAnchor(50,50)" title="Center">•</button>
          <button type=button class=btn data-x=100 data-y=50 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustMaskAnchor(100,50)" title="Center right">→</button>
          <button type=button class=btn data-x=0 data-y=100 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustMaskAnchor(0,100)" title="Bottom left">↙</button>
          <button type=button class=btn data-x=50 data-y=100 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustMaskAnchor(50,100)" title="Bottom center">↓</button>
          <button type=button class=btn data-x=100 data-y=100 style="padding:7px 0;font-size:14px" onclick="setPhotoAdjustMaskAnchor(100,100)" title="Bottom right">↘</button>
        </div>
      </div>
      <button type=button class=btn id=photoadjust-standard-btn style="width:100%;margin-bottom:4px" onclick="setPhotoAdjustCatalogueStandard()"><span id=photoadjust-standard-label>Standard Settings</span></button>
      <div style="display:flex;gap:6px;margin-top:4px">
        <button type=button class=btn style="flex:1" onclick="pickCatImage(PHOTO_ADJUST_SLOT)">Change Photo</button>
        <button type=button class=btn title="Choose from Cloud Library" style="width:42px;padding:0;flex:0 0 auto;color:#22c55e;display:flex;align-items:center;justify-content:center" onclick="openCloudPhotoPicker(PHOTO_ADJUST_SLOT)"><svg width=24 height=24 viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=1.8 stroke-linecap=round stroke-linejoin=round><circle cx=12 cy=12 r=9.5/><path d="M2.5 12h19"/><path d="M12 2.5c2.8 3 4.3 6.2 4.3 9.5s-1.5 6.5-4.3 9.5c-2.8-3-4.3-6.2-4.3-9.5S9.2 5.5 12 2.5Z"/><path d="M3.8 7.5h16.4M3.8 16.5h16.4"/></svg></button>
        <button type=button class=btn style="flex:1" onclick="resetPhotoAdjust()">Reset</button>
      </div>
    </div>
  </div>
</div>

<!-- Resumed-draft Generate hitting an already-generated product name — see
     generate()'s own comment (app.py, further down) and
     askCatReplaceConfirm which opens/wires this. Deliberately a real
     modal, not window.confirm() (silently no-ops in this app's embedded
     webview, same reason fmDelete uses its own click-again-to-confirm
     pattern instead) — this needs a real yes/no answer generate() can
     await, and the consequence (permanently overwriting a real generated
     PDF, no undo) is significant enough to spell out in full rather than
     a bare native "OK/Cancel" ever could. -->
<div class="clientmodal hide" id=catreplacemodal>
  <div class=clientmodalbox style="max-width:380px">
    <div class=clientmodalbar><b>Replace Existing Datasheet?</b></div>
    <div class=clientmodalbody>
      <p style="font-size:13px;margin:0 0 16px">A Sololuce Datasheet named "<b id=catreplace-name></b>" has already been generated. Continuing will overwrite that file — its PDF, saved data, and markdown copy — with this draft's version. The previous generated file can't be recovered afterward.</p>
      <div style="display:flex;gap:8px">
        <button type=button class=btn style="flex:1" id=catreplace-cancel-btn>Cancel</button>
        <button type=button class="btn dark" style="flex:1" id=catreplace-confirm-btn>Replace &amp; Generate</button>
      </div>
    </div>
  </div>
</div>
<script>
let TYPE='QTN2', INDEX=[], items=[], EDITING=null, EDITING_DRAFT=null, hoverTimer=null, hoverToken=0, BRAND='ARTEMIS', BRAND_LIST=[], brandMenuOpen=false, CLIPBOARD=null, CTXROW=null, UNITS=['PCS','MTR','PAIRS','ROLLS'];
async function loadUnits(){const r=await fetch('/api/units').then(r=>r.json());if(r.units&&r.units.length)UNITS=r.units}
const $=id=>document.getElementById(id);
function escHtml(s){return (s??'').toString().replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;')}
// Same icon+heading+subtext shape everywhere a list has nothing to show —
// iconPath reuses that section's own sidebar-nav icon, so the empty state
// visually echoes where the user already is instead of a bare line of text.
function emptyStateHtml(iconPath,title,sub){
  return '<div class="empty-state"><div class="empty-state-icon"><svg viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round>'+iconPath+'</svg></div>'+
    '<div class="empty-state-title">'+escHtml(title)+'</div>'+
    '<div class="empty-state-sub">'+sub+'</div></div>'}
const EMPTY_ICON_DOCS='<path d="M3 7a2 2 0 0 1 2-2h4l2 2h8a2 2 0 0 1 2 2v8a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2Z"/>';
const EMPTY_ICON_SUBMISSIONS='<path d="M9 11l3 3L22 4"/><path d="M21 12v7a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h11"/>';
const EMPTY_ICON_CLIENTS='<circle cx=12 cy=8 r=4 /><path d="M4 21a8 8 0 0 1 16 0"/>';

// ---------------------------------------------------------------- rich text (QTN2 free-text boxes)
let richFocused=null, richRange=null;
function initRichText(){
  document.addEventListener('focusin',e=>{
    if(e.target.classList&&e.target.classList.contains('richbox')){richFocused=e.target;showRichToolbar(e.target)}});
  document.addEventListener('focusout',e=>{
    // fires when the richbox loses focus (e.g. to the color picker) *and*
    // when the toolbar's own color <input> later loses focus (e.g. user
    // clicks away after picking) — either way, only actually hide once
    // focus has landed somewhere that's neither the richbox nor the toolbar.
    if(e.target===richFocused||$('richtoolbar').contains(e.target))
      setTimeout(()=>{
        if(document.activeElement!==richFocused&&!$('richtoolbar').contains(document.activeElement))hideRichToolbar()
      },120)});
  // the color picker is a native popup that can steal focus away from the
  // richbox — selectionchange keeps a live snapshot of the caret/selection
  // while focused, so richCmd can restore it even after focus bounced away.
  document.addEventListener('selectionchange',()=>{
    if(richFocused&&document.activeElement===richFocused){
      const sel=window.getSelection();
      if(sel&&sel.rangeCount)richRange=sel.getRangeAt(0).cloneRange()}});
  window.addEventListener('scroll',()=>{if(richFocused&&document.activeElement===richFocused)showRichToolbar(richFocused)},true)}
function showRichToolbar(el){
  const tb=$('richtoolbar');tb.style.display='flex';
  const r=el.getBoundingClientRect(),tr=tb.getBoundingClientRect();
  let top=r.top-tr.height-6;if(top<4)top=r.bottom+6;
  const left=Math.min(Math.max(8,r.left),window.innerWidth-tr.width-8);
  tb.style.top=top+'px';tb.style.left=left+'px'}
function hideRichToolbar(){$('richtoolbar').style.display='none'}
function richCmd(cmd,val){
  if(!richFocused)return;
  richFocused.focus();
  if(richRange){const sel=window.getSelection();sel.removeAllRanges();try{sel.addRange(richRange)}catch(e){}}
  document.execCommand(cmd,false,val);
  richFocused.dispatchEvent(new Event('input',{bubbles:true}))}
function richText(el){const h=(el.innerHTML||'').trim();return(h==='<br>'?'':h)}

// ---------------------------------------------------------------- color memory pins (recently-used colors, for one-click reuse)
let RECENT_COLORS=[];
try{RECENT_COLORS=JSON.parse(localStorage.getItem('cs_recentColors')||'null')}catch(e){}
if(!RECENT_COLORS||!RECENT_COLORS.length)RECENT_COLORS=['#1a1d24','#e2952c','#c0392b','#2a5b8c','#1f7a54'];
function renderRichSwatches(){
  $('richswatches').innerHTML=RECENT_COLORS.map(c=>
    '<button type=button class=rtswatch style="background:'+c+'" onmousedown="event.preventDefault()" onclick="richCmd(\'foreColor\',\''+c+'\')" title="'+c+'"></button>').join('')}
function pinRichColor(c){
  RECENT_COLORS=[c,...RECENT_COLORS.filter(x=>x!==c)].slice(0,6);
  localStorage.setItem('cs_recentColors',JSON.stringify(RECENT_COLORS));
  renderRichSwatches()}

// ---------------------------------------------------------------- custom autocomplete (richboxes can't use native <datalist>)
function attachAutocomplete(el,listGetter){
  let items=[],hiIdx=-1;
  function close(){$('autocomplete').style.display='none';items=[];hiIdx=-1}
  function draw(){
    const box=$('autocomplete');
    if(!items.length){close();return}
    box.innerHTML=items.map((v,i)=>'<div class="acitem'+(i===hiIdx?' hi':'')+'" data-i='+i+'>'+escHtml(v)+'</div>').join('');
    const r=el.getBoundingClientRect();
    box.style.display='block';box.style.left=r.left+'px';box.style.top=(r.bottom+4)+'px';box.style.width=Math.max(200,r.width)+'px';
    box.querySelectorAll('.acitem').forEach(node=>{
      node.onmousedown=e=>e.preventDefault();
      node.onclick=()=>pick(items[+node.dataset.i])})}
  function pick(v){el.textContent=v;el.dispatchEvent(new Event('input',{bubbles:true}));close()}
  el.addEventListener('input',()=>{
    const q=(el.textContent||'').trim().toLowerCase();
    if(!q){close();return}
    items=listGetter().filter(v=>v.toLowerCase().includes(q)&&v.toLowerCase()!==q).slice(0,8);
    hiIdx=-1;draw()});
  el.addEventListener('keydown',e=>{
    if(!items.length)return;
    if(e.key==='ArrowDown'){e.preventDefault();hiIdx=Math.min(hiIdx+1,items.length-1);draw()}
    else if(e.key==='ArrowUp'){e.preventDefault();hiIdx=Math.max(hiIdx-1,0);draw()}
    else if(e.key==='Enter'&&hiIdx>=0){e.preventDefault();pick(items[hiIdx])}
    else if(e.key==='Escape')close()});
  el.addEventListener('blur',()=>setTimeout(close,150))}
const BRAND_LOGOS={ARTEMIS:'/static/logos/artemis.png',SOLOLUCE:'/static/logos/sololuce.png',ADS:'/static/logos/ads.png',WATT:'/static/logos/watt.png'};
const BRAND_WATERMARKS={ARTEMIS:'/static/logos/artemis_watermark.png',SOLOLUCE:'/static/logos/sololuce_watermark.png',ADS:'/static/logos/ads.png',WATT:'/static/logos/watt.png'};
const ICON_CLS={main:['blogo','bulb'],mini:['blogomini','bulbmini'],set:['blogoset','bulb']};
function brandIcon(code,kind){
  const c=code.toLowerCase(),logo=BRAND_LOGOS[code],[imgCls,divCls]=ICON_CLS[kind];
  return logo?'<img class="'+imgCls+' b-'+c+'" src="'+logo+'" alt="">':'<div class="'+divCls+' b-'+c+'"></div>'}
async function loadBrands(){const r=await fetch('/api/brands').then(r=>r.json());
  BRAND=r.current;BRAND_LIST=r.brands||[];applyBrandUI()}
function applyBrandUI(){
  const b=BRAND_LIST.find(x=>x.code===BRAND)||{code:BRAND,label:BRAND};
  $('brandbtnlabel').textContent=b.code;
  $('bulb').innerHTML=brandIcon(b.code,'main');
  const wm=BRAND_WATERMARKS[b.code];
  $('watermark').innerHTML=wm?'<img class="b-'+b.code.toLowerCase()+'" src="'+wm+'" alt="">':'';
  const isSololuce=BRAND==='SOLOLUCE';
  $('n-catbuild').classList.toggle('hide',!isSololuce);
  $('n-fullcatalog').classList.toggle('hide',!isSololuce);
  // Sololuce Datasheets is Sololuce-only — switching to another brand while
  // it's open (or just leaving it active underneath) drops back to the
  // Quotation screen so the now-hidden menu item can't stay "current".
  if(!isSololuce&&TYPE==='CAT')openDocType('QTN2');
  // Same reasoning for the Full Catalog Builder — it isn't a doc-type Build
  // form (no TYPE to check), just check which view is actually showing.
  if(!isSololuce&&!$('v-fullcatalog').classList.contains('hide'))view('menu');
  if(brandMenuOpen)renderBrandMenu()}
function renderBrandMenu(){
  $('brandmenu').innerHTML=BRAND_LIST.map((b,i)=>
    '<div class="bcard'+(b.code===BRAND?' on':'')+'" style="animation-delay:'+(i*0.045)+'s" onclick="pickBrand(\''+b.code+'\')">'+
      brandIcon(b.code,'mini')+
      '<div><div class=bcardname>'+b.label+'</div><div class=bcardsub>'+b.code+'</div></div></div>').join('')}
function toggleBrandMenu(){
  brandMenuOpen=!brandMenuOpen;
  $('brandbtn').classList.toggle('open',brandMenuOpen);
  $('brandmenu').classList.toggle('hide',!brandMenuOpen);
  // Keeps the rail expanded (see .rail.pinned CSS) for as long as the popup
  // is open, even after the mouse/keyboard-focus leaves it for the popup
  // itself — otherwise the rail would snap back to its collapsed width while
  // the popup (a sibling, positioned off the rail's expanded edge) stayed put,
  // leaving a visible gap between the two.
  $('rail').classList.toggle('pinned',brandMenuOpen);
  if(brandMenuOpen)renderBrandMenu()}
function pickBrand(code){toggleBrandMenu();if(code!==BRAND)setBrand(code)}
document.addEventListener('click',e=>{
  if(brandMenuOpen&&!e.target.closest('#brandbtn')&&!e.target.closest('#brandmenu'))toggleBrandMenu()});
// Menu — the home/landing page (v-menu, shown first on launch), so every
// destination is reachable regardless of which brand is currently active
// (Sololuce Datasheets is otherwise invisible in the rail unless already on
// that brand — see launcherGoSololuce below).
function launcherGoDoc(t){openDocType(t)}
function launcherGo(v){view(v)}
// setBrand() is async (it awaits a fetch + UI refresh) — must finish before
// openDocType('CAT') runs, or n-catbuild can still carry .hide when view() fires.
async function launcherGoSololuce(){
  if(BRAND!=='SOLOLUCE')await setBrand('SOLOLUCE');
  openDocType('CAT')}
async function launcherGoFullCatalog(){
  if(BRAND!=='SOLOLUCE')await setBrand('SOLOLUCE');
  view('fullcatalog')}
async function setBrand(code){
  await fetch('/api/set-brand',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({brand:code})});
  BRAND=code;applyBrandUI();
  const b=BRAND_LIST.find(x=>x.code===code);toast('Switched to '+(b?b.label:code));
  if(!EDITING)nextNumber(true);
  onCompany();
  if(!$('v-all').classList.contains('hide'))loadIndex();else loadClients();
  if(!$('v-settings').classList.contains('hide'))loadSettings()}
function toast(m){$('toast').textContent=m;$('toast').classList.add('show');setTimeout(()=>$('toast').classList.remove('show'),2000)}

// ---------------------------------------------------------------- Auto-update
// Polls /api/check-update (GitHub Releases — see update_checker.py) on
// launch (if UPDATE_PREFS.check_on_start) and every few hours the app
// stays open, or on demand from the Update Center (top bar, everyone —
// see openUpdateCenter()). Finding one unhides the rail's pulsing-dot
// Update button; with auto_update on, it also installs itself right
// away instead of waiting for a click. UPDATE_INFO holds the last result
// so openUpdateMenu()/renderUpdateCenter() don't need to refetch.
let UPDATE_INFO=null, UPDATE_PREFS={check_on_start:true,auto_update:false};
async function checkForAppUpdate(manual){
  try{
    const r=await fetch('/api/check-update').then(r=>r.json());
    UPDATE_INFO=r;
    $('n-update').classList.toggle('hide',!r.available);
    renderUpdateCenter();
    if(r.available&&UPDATE_PREFS.auto_update&&!manual){
      toast('Installing update automatically…');
      const fakeBtn=document.createElement('button');fakeBtn.dataset.confirm='1';
      actuallyInstallUpdate(fakeBtn)}
  }catch(e){/* offline or GitHub unreachable — silently skip, try again later */}}
function openUpdateMenu(rect){
  if(!UPDATE_INFO)return;
  const menu=$('filemenu');
  renderUpdateMenu();
  menu.style.display='block';
  const r=rect||{left:0,bottom:0};
  const w=menu.offsetWidth||260,h=menu.offsetHeight||160;
  let x=r.left,y=r.bottom+4;
  if(x+w>window.innerWidth-8)x=window.innerWidth-w-8;
  if(y+h>window.innerHeight-8)y=r.top-h-4;
  menu.style.left=x+'px';menu.style.top=y+'px';
  setTimeout(()=>document.addEventListener('click',closeFileMenu,{once:true}),0)}
function renderUpdateMenu(){
  const u=UPDATE_INFO;if(!u)return;
  $('filemenu').innerHTML=
    '<div class=fmtitle>Update available</div>'+
    '<div class=fmupdatever>v'+escHtml(u.current)+' → <b>v'+escHtml(u.latest)+'</b></div>'+
    (u.notes?'<div class=fmupdatenotes>'+escHtml(u.notes)+'</div>':'')+
    '<button type=button class=btn style="width:100%;margin:2px 12px 8px;width:calc(100% - 24px)" '+
      'onclick="event.stopPropagation();installUpdate(this)">Install &amp; Restart</button>'}
// Same click-to-arm-then-confirm pattern as deleteManageListItem() —
// window.confirm() silently no-ops in this environment (pywebview/WebView2).
function installUpdate(btn){
  if(btn.dataset.confirm!=='1'){
    btn.dataset.confirm='1';btn.textContent='Click again to install & restart';
    setTimeout(()=>{if(btn.dataset.confirm==='1'){btn.dataset.confirm='';btn.textContent='Install & Restart'}},3000);
    return}
  actuallyInstallUpdate(btn)}
async function actuallyInstallUpdate(btn){
  btn.disabled=true;btn.textContent='Downloading…';
  try{
    const r=await fetch('/api/apply-update',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({download_url:UPDATE_INFO.download_url})}).then(r=>r.json());
    if(!r.ok){toast('Update failed: '+(r.error||'unknown error'));btn.disabled=false;btn.textContent='Install & Restart';return}
    btn.textContent='Installing…';
    toast('Installer launching — the app will close and reopen the new version');
  }catch(e){toast('Update failed: '+e.message);btn.disabled=false;btn.textContent='Install & Restart'}}

// Reads the saved prefs before deciding whether to auto-check at all —
// unlike openUpdateCenter() (which always checks, since the user just
// asked), a launch-time check should stay off if "Check for updates on
// each start" was unchecked. Still populates UPDATE_PREFS either way, so
// auto_update is known even if the user later finds an update manually.
async function initUpdateChecking(){
  UPDATE_PREFS=await fetch('/api/update-prefs').then(r=>r.json()).catch(()=>UPDATE_PREFS);
  if(!UPDATE_PREFS.check_on_start)return;
  checkForAppUpdate();
  setInterval(checkForAppUpdate,4*60*60*1000)} // re-check every 4h for a long-running session

// ---- Update Center (top bar, everyone) ----
async function openUpdateCenter(){
  $('updatecentermodal').classList.remove('hide');
  const p=await fetch('/api/update-prefs').then(r=>r.json()).catch(()=>UPDATE_PREFS);
  UPDATE_PREFS=p;
  $('uc-check-on-start').checked=!!p.check_on_start;
  $('uc-auto-update').checked=!!p.auto_update;
  await checkForAppUpdate(true)}
function closeUpdateCenter(){$('updatecentermodal').classList.add('hide')}
function renderUpdateCenter(){
  const u=UPDATE_INFO;if(!u)return;
  $('uc-version').textContent=u.current?('v'+u.current):'—';
  const installBtn=$('uc-install-btn'),notes=$('uc-notes');
  if(u.available){
    $('uc-status').textContent='v'+u.latest+' is available.';
    installBtn.classList.remove('hide');
    installBtn.dataset.confirm='';installBtn.disabled=false;installBtn.textContent='Install & Restart';
    if(u.notes){notes.textContent=u.notes;notes.classList.remove('hide')}else notes.classList.add('hide')
  }else{
    $('uc-status').textContent=u.error?'Could not check for updates.':"You're on the latest version.";
    installBtn.classList.add('hide');notes.classList.add('hide')}}
async function saveUpdatePrefs(){
  UPDATE_PREFS={check_on_start:$('uc-check-on-start').checked,auto_update:$('uc-auto-update').checked};
  await fetch('/api/update-prefs',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(UPDATE_PREFS)});
  toast('Saved')}
// Explicit choice wins and persists (localStorage); with no choice yet, the
// :root media-query block already follows the OS preference on its own —
// this just keeps the toggle's own icon/label in sync with whichever is
// actually in effect, from a page-load "flash of wrong theme" script in
// <head> (which is what actually sets the attribute before first paint;
// this only needs to reflect that resulting state, not set it initially).
function effectiveTheme(){
  return document.documentElement.getAttribute('data-theme')||(window.matchMedia&&window.matchMedia('(prefers-color-scheme:dark)').matches?'dark':'light')}
function syncThemeIcon(){
  const dark=effectiveTheme()==='dark';
  $('themeicon-sun').style.display=dark?'none':'block';
  $('themeicon-moon').style.display=dark?'block':'none';
  $('themelabel').textContent=dark?'Dark':'Light'}
function toggleTheme(){
  const next=effectiveTheme()==='dark'?'light':'dark';
  document.documentElement.setAttribute('data-theme',next);
  try{localStorage.setItem('theme',next)}catch(e){}
  syncThemeIcon()}
syncThemeIcon();
// Every document type is its own sidebar entry with its own view name, all
// sharing the one #v-build DOM underneath — which card set is visible is
// still decided entirely by setType()'s existing .cat-only/.exp-only/hide
// toggles, so the split is purely navigational. DOC_VIEWS is the single
// place the type<->view mapping lives; view() and openDocType() both read
// it rather than hardcoding the list twice.
const DOC_VIEWS={QTN2:'qtn2',INV:'inv',DO:'do',EXP:'exp',CAT:'catbuild'};
const DOC_VIEW_LIST=Object.values(DOC_VIEWS);
function view(v){
  // Defense in depth for a restricted user — the corresponding nav
  // button/tile is already hidden (see applyAccessRestrictions()), this
  // just stops a direct view('settings')-style console call too. The
  // real enforcement is server-side (app.py's before_request).
  if(BLOCKED_TOOLS.includes(v))v='menu';
  const isDoc=DOC_VIEW_LIST.includes(v);
  $('v-menu').classList.toggle('hide',v!='menu');$('v-build').classList.toggle('hide',!isDoc);$('v-all').classList.toggle('hide',v!='all');$('v-clients').classList.toggle('hide',v!='clients');$('v-settings').classList.toggle('hide',v!='settings');$('v-submissions').classList.toggle('hide',v!='submissions');$('v-statement').classList.toggle('hide',v!='statement');$('v-fullcatalog').classList.toggle('hide',v!='fullcatalog');
  $('n-launcher').classList.toggle('on',v=='menu');$('n-all').classList.toggle('on',v=='all');$('n-clients').classList.toggle('on',v=='clients');$('n-settings').classList.toggle('on',v=='settings');$('n-submissions').classList.toggle('on',v=='submissions');$('n-statement').classList.toggle('on',v=='statement');$('n-fullcatalog').classList.toggle('on',v=='fullcatalog');
  DOC_VIEW_LIST.forEach(dv=>$('n-'+dv).classList.toggle('on',v===dv));
  if(v=='menu')$('title').textContent='Menu';if(v=='all')loadIndex();if(v=='clients')loadClientsView();if(v=='settings')loadSettings();if(v=='submissions')loadSubmissions();if(v=='statement')loadStatement();if(v=='fullcatalog'){$('title').textContent='Full Catalog Builder';loadFullCatalogView()}}
// The one way to open a document screen — every rail button and Menu tile
// goes through here. Only resets the form when actually switching type, so
// clicking the nav item you're already on never wipes work in progress
// (same behaviour the old openCatBuild had).
function openDocType(t){view(DOC_VIEWS[t]);if(TYPE!==t)setType(t)}
// The Sololuce Datasheet form is split into collapsible sections (Basics /
// Photos / Spec Badges / Technical Specifications / Finish & Ordering)
// instead of one long card, so a user working on one part can collapse the
// others rather than scrolling past everything every time. All start open
// on a fresh form — collapsing is something the user opts into, not a
// default that could hide a required field they haven't seen yet.
function toggleCatSection(id){
  const body=$('cat-sec-'+id+'-body'),chev=$('cat-sec-'+id+'-chev');
  const collapsed=body.classList.toggle('hide');
  chev.textContent=collapsed?'▸':'▾';
  if(!collapsed)refreshCatSectionResize(body.id)}
// A section only gets a drag handle once its content is actually taller
// than this — a short section (e.g. Basics) never grows one, since there's
// nothing hidden for resizing to reveal. max-height is capped at the
// content's own natural height so dragging can never pull the section
// taller than its real content (previously left a dead gap above Generate
// PDF when the Ordering Table had been dragged oversized).
const CAT_RESIZE_THRESHOLD=340;
function refreshCatSectionResize(id){
  const el=$(id);if(!el||el.classList.contains('hide'))return;
  // scrollHeight always reflects the full content height regardless of any
  // height/max-height currently applied, so a manually-dragged height can
  // be read back and re-clamped rather than wiped out by every edit.
  const manualHeight=el.classList.contains('is-resizable')&&el.style.height?parseFloat(el.style.height):null;
  const natural=el.scrollHeight;
  if(natural>CAT_RESIZE_THRESHOLD){
    el.classList.add('is-resizable');
    el.style.maxHeight=Math.ceil(natural)+'px';
    if(manualHeight!=null)el.style.height=Math.min(manualHeight,natural)+'px'
  }else{
    el.classList.remove('is-resizable');
    el.style.height='';el.style.maxHeight=''}}
(function initCatSectionResize(){
  ['cat-sec-basics-body','cat-sec-photos-body','cat-sec-badges-body','cat-sec-specs-body','cat-sec-finish-body','cat-sec-ordering-body'].forEach(id=>{
    const el=$(id);if(!el)return;
    const run=()=>refreshCatSectionResize(id);
    new MutationObserver(()=>{clearTimeout(el._resizeT);el._resizeT=setTimeout(run,50)}).observe(el,{childList:true,subtree:true,characterData:true});
    run()})})();
const SETTINGS_FIELDS=['inv_folder','do_folder','qtn2_folder','pi_folder','rv_folder','cn_folder','scanned_do_folder','product_photos_folder','datasheets_folder','templates_folder','clients_file','catalogue_folder','expense_folder'];
async function loadSettings(){
  showSettingsMainPanel();  // always land on the main panel, not wherever the admin sub-page was left
  const r=await fetch('/api/settings').then(r=>r.json());
  SETTINGS_FIELDS.forEach(k=>{const el=$('set-'+k);if(el)el.value=r[k]||''});
  const b=BRAND_LIST.find(x=>x.code===BRAND)||{code:BRAND,label:BRAND};
  $('set-brandname').textContent=b.label;
  $('set-bulb').innerHTML=brandIcon(BRAND,'set');
  renderManageLists();renderAuditLog();loadPhotoStoreSettings()}
// Admin Tools — a sub-page nested WITHIN Settings itself (own rail item
// deliberately rejected — see applyAccessRestrictions()'s comment), only
// ever reachable via the entry card that's only visible for role==='admin'.
// Global top-bar button (visible on every screen for role==='admin', see
// the .bar markup) — jumps straight to Settings' Admin sub-page from
// anywhere, no need to go through Settings' own UI first.
function openAdminTools(){view('settings');showSettingsAdminPanel()}
function showSettingsAdminPanel(){
  $('settings-main-panel').classList.add('hide');
  $('settings-admin-panel').classList.remove('hide');
  loadUsersAdmin();loadPhotoStoreList();loadPhotoStoreAdminSettings()}
function showSettingsMainPanel(){
  $('settings-admin-panel').classList.add('hide');
  $('settings-main-panel').classList.remove('hide')}
async function browseSetting(field){
  const r=await fetch('/api/browse',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({field})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  if(r.value)$('set-'+field).value=r.value}
async function saveSettings(){
  const body={};SETTINGS_FIELDS.forEach(k=>{const el=$('set-'+k);if(el)body[k]=el.value.trim()});
  const r=await fetch('/api/settings',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  loadSettings();
  toast('Settings saved')}

// ---------------------------------------------------------------- Shared Product Photos (Cloudflare R2)
// See photo_store.py's own comment for the full design. Credentials here
// are per-machine (this install's own config.json) — every install (admin
// or not) fills this in once with whichever token it was given.
function fmtBytes(n){
  if(n>=1073741824)return (n/1073741824).toFixed(2)+' GB';
  if(n>=1048576)return (n/1048576).toFixed(1)+' MB';
  return Math.round(n/1024)+' KB'}
// General Settings (everyone) — status only, no credential fields at all
// (admin-only now, see Admin Tools) — just whether syncing is working.
async function loadPhotoStoreSettings(){
  refreshPhotoStoreStatus()}
// Admin Tools — both credential tiers. See photo_store.py's own comment:
// "Your Admin Key" is this machine's write-capable key; "Read-Only Key"
// is what gets BUNDLED into future builds for everyone else, not typed
// in by them.
async function loadPhotoStoreAdminSettings(){
  const r=await fetch('/api/photostore-config').then(r=>r.json());
  $('ps-account_id').value=r.account_id||'';
  $('ps-bucket').value=r.bucket||'';
  $('ps-access_key_id').value=r.access_key_id||'';
  $('ps-secret_access_key').value='';
  $('ps-secret-label').textContent=r.has_secret?'Secret Access Key (saved — leave blank to keep it)':'Secret Access Key';
  refreshPhotoStoreStatus()}
// One key, saved to both this machine's own config AND the bundle every
// future build ships (see /api/photostore-config's own comment) — the
// bundle half only actually lands from the admin's dev checkout, so
// report that separately rather than failing the whole save over it.
async function savePhotoStoreConfig(btn){
  btn.disabled=true;btn.textContent='Saving…';
  const r=await fetch('/api/photostore-config',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({
    account_id:$('ps-account_id').value.trim(),bucket:$('ps-bucket').value.trim(),
    access_key_id:$('ps-access_key_id').value.trim(),secret_access_key:$('ps-secret_access_key').value})}).then(r=>r.json());
  btn.disabled=false;btn.textContent='Save Cloud Storage Key';
  $('ps-bundle-note').textContent=r.bundled?'Bundled for future builds too — rebuild + GUPDATE to ship it to everyone.':('Saved for this machine, but not bundled: '+(r.bundle_error||'unknown error'));
  toast('Saved — checking connection…');
  loadPhotoStoreAdminSettings()}
async function refreshPhotoStoreStatus(){
  const el=$('photostore-status');
  const r=await fetch('/api/photostore-status').then(r=>r.json()).catch(()=>({configured:false}));
  if(!r.configured){el.textContent=CURRENT_ROLE==='admin'?'Not set up yet — see Admin Tools.':'Not set up yet — ask your admin.';return}
  if(r.error){el.textContent='Could not connect: '+r.error;return}
  const pct=Math.min(100,Math.round(r.bytes_used/r.limit_bytes*100));
  el.innerHTML=r.count+' photo'+(r.count!==1?'s':'')+' in the shared library — '+fmtBytes(r.bytes_used)+' / 10 GB used'+
    '<div style="height:6px;border-radius:4px;background:var(--tint);margin-top:5px;overflow:hidden">'+
    '<div style="height:100%;width:'+pct+'%;background:'+(pct>90?'var(--danger)':'var(--brand-dark)')+'"></div></div>'}
async function syncPhotoStore(btn){
  btn.disabled=true;btn.textContent='Syncing…';
  const r=await fetch('/api/photostore-sync',{method:'POST'}).then(r=>r.json());
  btn.disabled=false;btn.textContent='Sync Now — Download New Photos';
  toast(r.ok?(r.downloaded?'Downloaded '+r.downloaded+' new photo'+(r.downloaded!==1?'s':''):'Already up to date'):('Sync failed: '+(r.error||'unknown error')));
  refreshPhotoStoreStatus()}
async function loadPhotoStoreList(){
  const r=await fetch('/api/photostore-list').then(r=>r.json()).catch(()=>null);
  if(!r||r.error){$('photostore-list').innerHTML='';return}
  $('photostore-list').innerHTML=r.photos.map(p=>
    '<div class=usercard><b style="font-weight:600;font-size:12.5px">'+escHtml(p.key)+'</b>'+
      '<span class=userbadge>'+fmtBytes(p.size)+'</span>'+
      '<button type=button class=btn style="padding:4px 9px;font-size:11px" onclick="deletePhotoFromStore(\''+escHtml(p.key).replace(/'/g,"\\'")+'\',this)">✕</button></div>').join('')
    || '<p class=muted style="font-size:12px">No photos uploaded yet.</p>'}
function deletePhotoFromStore(name,btn){
  if(btn.dataset.confirm!=='1'){
    btn.dataset.confirm='1';btn.textContent='Sure?';
    setTimeout(()=>{if(btn.dataset.confirm==='1'){btn.dataset.confirm='';btn.textContent='✕'}},2500);
    return}
  actuallyDeletePhotoFromStore(name)}
async function actuallyDeletePhotoFromStore(name){
  const r=await fetch('/api/photostore-delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({filename:name})}).then(r=>r.json());
  if(!r.ok){toast(r.error||'Could not delete');return}
  toast('Removed '+name);loadPhotoStoreList();refreshPhotoStoreStatus()}
// Handles both the plain multi-file picker and the whole-folder picker
// (webkitdirectory) — a folder pick can easily be hundreds of files, so
// this uploads in small batches sequentially (one big multipart request
// for hundreds of images risks timing out or exhausting memory) with a
// progress line, and skips anything that isn't a .png up front since
// that's all engine.load_photo_catalog() ever matches against.
const UPLOAD_BATCH_SIZE=15;
async function uploadPhotosToStore(inputId,btn){
  const input=$(inputId);
  const all=[...input.files];
  const pngs=all.filter(f=>f.name.toLowerCase().endsWith('.png'));
  const skipped=all.length-pngs.length;
  if(!pngs.length){toast(all.length?'None of those were .png files — only .png is matched by the app':'Choose at least one photo first');return}
  const origLabel=btn.textContent;
  btn.disabled=true;
  const prog=$('photostore-upload-progress');prog.classList.remove('hide');
  let uploadedTotal=0,errorList=[];
  for(let i=0;i<pngs.length;i+=UPLOAD_BATCH_SIZE){
    const batch=pngs.slice(i,i+UPLOAD_BATCH_SIZE);
    btn.textContent='Uploading '+Math.min(i+UPLOAD_BATCH_SIZE,pngs.length)+' / '+pngs.length+'…';
    prog.textContent='Uploading '+Math.min(i+UPLOAD_BATCH_SIZE,pngs.length)+' of '+pngs.length+' photos'+(skipped?' ('+skipped+' non-.png file'+(skipped!==1?'s':'')+' skipped)':'');
    const fd=new FormData();
    batch.forEach(f=>fd.append('files',f,f.webkitRelativePath||f.name));
    const r=await fetch('/api/photostore-upload',{method:'POST',body:fd}).then(r=>r.json()).catch(e=>({ok:false,errors:[e.message]}));
    uploadedTotal+=(r.uploaded||[]).length;
    errorList=errorList.concat(r.errors||[]);
  }
  btn.disabled=false;btn.textContent=origLabel;
  prog.classList.add('hide');
  input.value='';
  toast(uploadedTotal+' photo'+(uploadedTotal!==1?'s':'')+' uploaded'+(errorList.length?', '+errorList.length+' failed':'')+(skipped?' — '+skipped+' non-.png skipped':''));
  if(errorList.length)console.warn('Photo upload errors:',errorList);
  loadPhotoStoreList();refreshPhotoStoreStatus()}

// ---------------------------------------------------------------- Full Catalog Builder
// Assembles every generated Sololuce Datasheet into one bound book — see
// catalog_builder.py for the full architecture (why it's a from-scratch
// pass every time, never incremental). This screen is Sololuce-only, like
// the Datasheet builder itself, but is its own top-level view (not part of
// DOC_VIEWS — there's no per-document form here, just settings + a build
// action), so it's wired directly through view()/applyBrandUI() instead.
async function loadFullCatalogView(){
  const [summary,series,sectionOrder,extras,families,lastBuild,settings]=await Promise.all([
    fetch('/api/full-catalog/summary').then(r=>r.json()),
    fetch('/api/cat-series').then(r=>r.json()),
    fetch('/api/full-catalog/section-order').then(r=>r.json()),
    fetch('/api/full-catalog/extras').then(r=>r.json()),
    fetch('/api/cat-family').then(r=>r.json()),
    fetch('/api/full-catalog/last-build').then(r=>r.json()),
    fetch('/api/settings').then(r=>r.json()),
  ]);
  renderFcSummary(summary);
  FC_CATEGORIES=series.labels||[];FC_CATEGORY_COLORS=series.colors||{};FC_CAT_SECTIONS=series.sections||{};
  renderFcCategoryOrder();
  FC_SECTIONS=sectionOrder.order||[];FC_SECTION_LABELS=sectionOrder.labels||{};FC_SECTION_COLORS=sectionOrder.colors||{};
  renderFcSectionOrder();
  loadFcIndexOrder();
  FC_EXTRAS=extras||{};FC_EXTRAS.cover=FC_EXTRAS.cover||{};FC_EXTRAS.front_matter=FC_EXTRAS.front_matter||[];FC_EXTRAS.family_dividers=FC_EXTRAS.family_dividers||{};
  FC_FAMILY_NAMES=families.labels||[];
  renderFcCover();renderFcFrontMatter();renderFcEnding();renderFcFamilyDividers();
  $('set-full_catalog_folder').value=settings.full_catalog_folder||'';
  renderFcLastBuild(lastBuild)}
function renderFcSummary(data){
  if(data.error){$('fc-summary').innerHTML='<span style="color:var(--danger)">'+escHtml(data.error)+'</span>';return}
  let html='<b>'+data.product_count+'</b> product'+(data.product_count!=1?'s':'')+' across <b>'+data.category_count+'</b> categor'+(data.category_count!=1?'ies':'y')+
    ' in <b>'+data.section_count+'</b> section'+(data.section_count!=1?'s':'')+' — at least <b>'+data.estimated_datasheet_pages+'</b> page'+(data.estimated_datasheet_pages!=1?'s':'')+
    ' of datasheets alone (the real total also includes Index/Pre-index pages and anything you\'ve uploaded, known only once you Build).';
  if(data.warnings&&data.warnings.length)html+='<ul style="margin:10px 0 0;padding-left:18px;color:var(--warning)">'+data.warnings.map(w=>'<li>'+escHtml(w)+'</li>').join('')+'</ul>';
  $('fc-summary').innerHTML=html}
// Category Order — grouped by an Outdoor/Indoor/Strip-Neon tag purely for
// navigation (see /api/cat-series-section's docstring — this tag never
// touches the actual build; a product's real section always comes from
// its own Product Type field), each row draggable within its own group
// (HTML5 drag-and-drop, see fcCatDrop — the ▲▼ arrows stay too, as a
// fallback/for precise single-step moves), plus color (native
// <input type=color>, POSTs straight to /api/cat-series-color on change),
// inline rename (click the name, same commit-on-blur/Enter pattern as the
// Family Name/Series custom-entry fields elsewhere in this file, POSTs to
// /api/cat-series-rename which also migrates any already-generated
// product's own sidecar off the old name), delete (double-click-to-confirm,
// same pattern as every other Remove in this app — window.confirm() no-ops
// here — reuses the generic /api/settings-list-remove Manage Lists already
// uses), two one-shot auto-sort buttons (A→Z, Oldest first), and a
// Randomize Colors button — all one-shot actions, not persisted modes;
// the result is still freely draggable/manually adjustable afterward.
// Shared drag-to-reorder plumbing, reused by every draggable list on this
// screen (and, further down, the datasheet form's own reorderable rows).
// dragRowStart/Over/Leave/End just handle the VISUAL feedback (.dragrow's
// CSS transition + the amber insert-line — see the .dragrow rules above);
// the actual reordering happens in dragReorder, which never invents a new
// "set full order" endpoint per list — it works out how many single steps
// separate the dragged row from the drop target, then replays that many
// calls to whatever adjacent-swap move function the list already has
// (awaited sequentially, since each call re-renders and the next step's
// index math needs that render settled first). DRAG_KEY is one shared
// module-level slot since only one drag can ever be in progress at once.
let DRAG_KEY=null;
function dragRowStart(e,key){DRAG_KEY=key;e.dataTransfer.effectAllowed='move';e.currentTarget.classList.add('dragging')}
function dragRowOver(e){
  e.preventDefault();e.dataTransfer.dropEffect='move';
  const row=e.currentTarget,rect=row.getBoundingClientRect();
  const before=(e.clientY-rect.top)<rect.height/2;
  row.classList.toggle('dragover-top',before);
  row.classList.toggle('dragover-bottom',!before)}
function dragRowLeave(e){e.currentTarget.classList.remove('dragover-top','dragover-bottom')}
function dragRowEnd(e){
  e.currentTarget.classList.remove('dragging');
  document.querySelectorAll('.dragover-top,.dragover-bottom,.dragover-left,.dragover-right').forEach(el=>el.classList.remove('dragover-top','dragover-bottom','dragover-left','dragover-right'))}
// Horizontal counterpart of dragRowOver/dragRowLeave — for rows that reorder
// left/right (Ordering Table's field pills) instead of up/down, so the
// amber insert-line reads as before/after in the direction the list
// actually moves rather than reusing the vertical top/bottom cue sideways.
function dragColOver(e){
  e.preventDefault();e.dataTransfer.dropEffect='move';
  const row=e.currentTarget,rect=row.getBoundingClientRect();
  const before=(e.clientX-rect.left)<rect.width/2;
  row.classList.toggle('dragover-left',before);
  row.classList.toggle('dragover-right',!before)}
function dragColLeave(e){e.currentTarget.classList.remove('dragover-left','dragover-right')}
async function dragReorder(targetKey,currentOrder,moveFn){
  if(DRAG_KEY===null||DRAG_KEY===targetKey)return;
  const from=currentOrder.indexOf(DRAG_KEY),to=currentOrder.indexOf(targetKey);
  if(from<0||to<0)return;
  const dir=to>from?'down':'up';
  for(let i=0;i<Math.abs(to-from);i++)await moveFn(DRAG_KEY,dir)}
let FC_CATEGORIES=[],FC_CATEGORY_COLORS={},FC_CAT_SECTIONS={},FC_SECTIONS=[],FC_SECTION_LABELS={},FC_SECTION_COLORS={};
const FC_CAT_GROUP_ORDER=['Indoor','Outdoor','Striplight',''];
const FC_CAT_GROUP_TITLE={Indoor:'Indoor',Outdoor:'Outdoor',Striplight:'Strip Light & Neon Flex','':'Unassigned'};
function fcCategoryGroups(){
  const by={Indoor:[],Outdoor:[],Striplight:[],'':[]};
  FC_CATEGORIES.forEach(label=>{const sec=FC_CAT_SECTIONS[label]||'';(by[sec]||by['']).push(label)});
  return FC_CAT_GROUP_ORDER.filter(k=>by[k].length).map(k=>[k,by[k]])}
function renderFcCategoryOrder(){
  if(!FC_CATEGORIES.length){
    $('fc-categoryorder').innerHTML='<p class=muted style="font-size:12px;margin:0">No categories yet — add one below, or type one into any Sololuce Datasheet\'s Series/Category field.</p>';
    return}
  $('fc-categoryorder').innerHTML=fcCategoryGroups().map(([key,labels])=>{
    const rows=labels.map((label,i)=>{
      const l=label.replace(/'/g,"\\'"),color=FC_CATEGORY_COLORS[label]||'#9a9a9a';
      const gi=FC_CATEGORIES.indexOf(label);
      return '<div class=dragrow draggable=true ondragstart="fcCatDragStart(event,\''+l+'\',\''+key+'\')" ondragover="dragRowOver(event)" ondragleave="dragRowLeave(event)" ondrop="fcCatDrop(event,\''+l+'\',\''+key+'\')" ondragend="dragRowEnd(event)" style="display:flex;align-items:center;gap:6px;padding:6px 0;'+(i?'border-top:1px solid var(--border)':'')+'">'+
        '<span class=draghandle title="Drag to reorder">⠿</span>'+
        '<input type=color value="'+color+'" style="width:20px;height:20px;padding:0;border:1px solid var(--border);border-radius:4px;cursor:pointer;flex-shrink:0" title="Category color" onchange="setFcCategoryColor(\''+l+'\',this.value)">'+
        '<span style="flex:1;font-size:12.5px;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;cursor:text" title="Click to rename" onclick="startFcCategoryRename(this,\''+l+'\')">'+escHtml(label)+'</span>'+
        '<select style="font-size:10px;padding:2px;border:1px solid var(--border);border-radius:4px;background:var(--card-bg);color:var(--muted);width:64px;flex-shrink:0" title="Section (for grouping here only)" onchange="setFcCategorySection(\''+l+'\',this.value)">'+
          '<option value=""'+(key===''?' selected':'')+'>—</option>'+
          '<option value="Indoor"'+(key==='Indoor'?' selected':'')+'>Indoor</option>'+
          '<option value="Outdoor"'+(key==='Outdoor'?' selected':'')+'>Outdoor</option>'+
          '<option value="Striplight"'+(key==='Striplight'?' selected':'')+'>Strip/Neon</option>'+
        '</select>'+
        '<button type=button class=ordmovebtn'+(gi===0?' disabled':'')+' onclick="moveFcCategory(\''+l+'\',\'up\')" title="Move up">▲</button>'+
        '<button type=button class=ordmovebtn'+(gi===FC_CATEGORIES.length-1?' disabled':'')+' onclick="moveFcCategory(\''+l+'\',\'down\')" title="Move down">▼</button>'+
        '<button type=button class=rm onclick="deleteFcCategory('+gi+',this)" title="Delete">×</button>'+
      '</div>'}).join('');
    return '<div style="margin-bottom:10px">'+
      '<div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.4px;color:var(--muted);padding:6px 0 2px" ondragover="fcCatDragOver(event)" ondrop="fcCatDrop(event,null,\''+key+'\')">'+escHtml(FC_CAT_GROUP_TITLE[key])+'</div>'+rows+
    '</div>'}).join('')}
async function moveFcCategory(label,direction){
  const r=await fetch('/api/cat-series-move',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({label,direction})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  FC_CATEGORIES=r.labels||FC_CATEGORIES;renderFcCategoryOrder()}
let FC_CAT_DRAG=null;
function fcCatDragStart(e,label,group){FC_CAT_DRAG={label,group};dragRowStart(e,label)}
async function fcCatDrop(e,targetLabel,group){
  e.preventDefault();
  e.currentTarget.classList.remove('dragover-top','dragover-bottom');
  if(!FC_CAT_DRAG||FC_CAT_DRAG.group!==group)return;  // only reorder within the same group
  const dragged=FC_CAT_DRAG.label;
  if(dragged===targetLabel)return;
  const groupLabels=FC_CATEGORIES.filter(l=>(FC_CAT_SECTIONS[l]||'')===group);
  const from=groupLabels.indexOf(dragged);
  if(from<0)return;
  groupLabels.splice(from,1);
  const to=targetLabel?groupLabels.indexOf(targetLabel):groupLabels.length;
  groupLabels.splice(to<0?groupLabels.length:to,0,dragged);
  let gi=0;
  const newFull=FC_CATEGORIES.map(l=>(FC_CAT_SECTIONS[l]||'')===group?groupLabels[gi++]:l);
  const r=await fetch('/api/cat-series-reorder',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({labels:newFull})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  FC_CATEGORIES=r.labels||newFull;renderFcCategoryOrder()}
async function setFcCategorySection(label,section){
  const r=await fetch('/api/cat-series-section',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({label,section})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  FC_CAT_SECTIONS=r.sections||FC_CAT_SECTIONS;renderFcCategoryOrder()}
async function setFcCategoryColor(label,hex){
  const r=await fetch('/api/cat-series-color',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({label,hex})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  FC_CATEGORY_COLORS=r.colors||FC_CATEGORY_COLORS;renderFcCategoryOrder()}
async function randomizeFcCategoryColors(){
  const r=await fetch('/api/cat-series-randomize-colors',{method:'POST'}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  FC_CATEGORY_COLORS=r.colors||FC_CATEGORY_COLORS;renderFcCategoryOrder()}
function startFcCategoryRename(span,oldLabel){
  const inp=document.createElement('input');
  inp.type='text';inp.value=oldLabel;inp.style.cssText='flex:1;font-size:12.5px;min-width:0;border:1px solid var(--border);border-radius:4px;padding:2px 4px';
  let done=false;
  const commit=async()=>{
    if(done)return;done=true;
    const v=inp.value.trim();
    if(!v||v===oldLabel){renderFcCategoryOrder();return}
    const r=await fetch('/api/cat-series-rename',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({old_label:oldLabel,new_label:v})}).then(r=>r.json());
    if(r.error){alert(r.error);renderFcCategoryOrder();return}
    FC_CATEGORIES=r.labels||FC_CATEGORIES;FC_CATEGORY_COLORS=r.colors||FC_CATEGORY_COLORS;
    renderFcCategoryOrder();
    toast(r.migrated_products?v+' — '+r.migrated_products+' existing product'+(r.migrated_products!=1?'s':'')+' updated':'Renamed')};
  inp.addEventListener('keydown',e=>{
    if(e.key==='Enter'){e.preventDefault();inp.blur()}
    else if(e.key==='Escape'){done=true;renderFcCategoryOrder()}});
  inp.addEventListener('blur',commit);
  span.replaceWith(inp);inp.focus();inp.select()}
async function deleteFcCategory(index,btn){
  if(btn.dataset.confirm!=='1'){btn.dataset.confirm='1';btn.textContent='Sure?';setTimeout(()=>{if(btn.dataset.confirm==='1'){btn.dataset.confirm='';btn.textContent='×'}},2500);return}
  const r=await fetch('/api/settings-list-remove',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({list_key:'cat_series_labels',index})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  FC_CATEGORIES=r.items||FC_CATEGORIES;renderFcCategoryOrder();toast('Deleted')}
async function sortFcCategories(mode){
  const r=await fetch('/api/cat-series-sort',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({mode})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  FC_CATEGORIES=r.labels||FC_CATEGORIES;renderFcCategoryOrder()}
async function addFcCategory(){
  const v=$('fc-newcategory').value.trim();
  if(!v)return;
  const r=await fetch('/api/cat-series-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({label:v})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  FC_CATEGORIES=r.labels||FC_CATEGORIES;FC_CATEGORY_COLORS=r.colors||FC_CATEGORY_COLORS;
  $('fc-newcategory').value='';renderFcCategoryOrder()}
// Index Order — which products show in the photo-grid Index and in what
// order, per category (see catalog_builder.py's compute_index_rows; the
// actual grid render and this card are built from the exact same function
// so they never disagree). Section/Category order themselves reuse the
// two cards above — this one is purely about product-level order +
// include/exclude within whatever category a product already belongs to.
// Collapsed by default per category (its own small Set, not the Manage
// Lists screen's MANAGE_LISTS_COLLAPSED — unrelated screen, kept separate
// on purpose) since a category can run 60+ products.
// Tracks EXPANDED categories (inverted from Manage Lists' own collapsed-
// tracking Set) so a category starts collapsed the first time it's ever
// seen, without needing to pre-seed every category key up front.
let FC_INDEX_DATA=null, FC_INDEX_EXPANDED=new Set();
async function loadFcIndexOrder(){
  const r=await fetch('/api/full-catalog/index-order').then(r=>r.json());
  if(r.error){$('fc-indexorder').innerHTML='<p class=muted style="font-size:12px;margin:0">'+escHtml(r.error)+'</p>';FC_INDEX_DATA=null;return}
  FC_INDEX_DATA=r;renderFcIndexOrder()}
function renderFcIndexOrder(){
  if(!FC_INDEX_DATA){return}
  let html='';
  FC_INDEX_DATA.sections.forEach(sec=>{
    sec.categories.forEach(cat=>{
      const key=sec.label+'|'+cat.label,collapsed=!FC_INDEX_EXPANDED.has(key);
      const hidden=cat.products.filter(p=>p.excluded).length;
      const countLabel=cat.products.length+' product'+(cat.products.length!=1?'s':'')+(hidden?', '+hidden+' hidden':'');
      const rows=cat.products.map((p,i)=>{
        const nameSafe=p.product_name.replace(/'/g,"\\'"),catSafe=cat.label.replace(/'/g,"\\'");
        return '<div class=dragrow draggable=true ondragstart="dragRowStart(event,\''+nameSafe+'\')" ondragover="dragRowOver(event)" ondragleave="dragRowLeave(event)" ondrop="fcIndexDrop(event,\''+catSafe+'\',\''+nameSafe+'\')" ondragend="dragRowEnd(event)" style="display:flex;align-items:center;gap:8px;padding:6px 0;'+(i?'border-top:1px solid var(--border)':'')+(p.excluded?';opacity:.5':'')+'">'+
          '<span class=draghandle title="Drag to reorder">⠿</span>'+
          '<div style="width:30px;height:30px;border-radius:2px;overflow:hidden;background:#e5e5e5;flex-shrink:0;display:flex;align-items:center;justify-content:center;">'+
            (p.main_photo?'<img src="'+p.main_photo+'" style="width:100%;height:100%;object-fit:cover">':'')+
          '</div>'+
          '<span style="flex:1;font-size:12px;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+escHtml(p.product_name)+(p.excluded?' <span class=muted>(hidden)</span>':'')+'</span>'+
          '<button type=button class=ordmovebtn'+(i===0?' disabled':'')+' onclick="moveFcIndexProduct(\''+catSafe+'\',\''+nameSafe+'\',\'up\')" title="Move up">▲</button>'+
          '<button type=button class=ordmovebtn'+(i===cat.products.length-1?' disabled':'')+' onclick="moveFcIndexProduct(\''+catSafe+'\',\''+nameSafe+'\',\'down\')" title="Move down">▼</button>'+
          '<button type=button class=btn style="padding:6px 10px;font-size:11px;flex-shrink:0" onclick="toggleFcIndexExclude(\''+nameSafe+'\','+(!p.excluded)+')">'+(p.excluded?'Show':'Hide')+'</button>'+
        '</div>'}).join('');
      html+='<div class="mlsection'+(collapsed?' collapsed':'')+'" data-key="'+key+'">'+
        '<div class=mlhead onclick="toggleFcIndexCategory(this)">'+escHtml(cat.label)+'<span class=mlcount>'+countLabel+'</span><span class=mlchev>▾</span></div>'+
        '<div class=mlbody><div style="display:flex;flex-direction:column;width:100%">'+rows+'</div></div></div>'})});
  $('fc-indexorder').innerHTML=html||'<p class=muted style="font-size:12px;margin:0">No datasheets generated yet.</p>'}
function toggleFcIndexCategory(head){
  const sec=head.closest('.mlsection'),key=sec.dataset.key;
  sec.classList.toggle('collapsed');
  if(sec.classList.contains('collapsed'))FC_INDEX_EXPANDED.delete(key);else FC_INDEX_EXPANDED.add(key)}
function fcIndexOrderFor(category){
  for(const sec of FC_INDEX_DATA.sections)
    for(const cat of sec.categories)
      if(cat.label===category)return cat.products.map(p=>p.product_name);
  return []}
async function fcIndexDrop(e,category,targetName){
  e.preventDefault();
  e.currentTarget.classList.remove('dragover-top','dragover-bottom');
  await dragReorder(targetName,fcIndexOrderFor(category),(name,dir)=>moveFcIndexProduct(category,name,dir))}
async function moveFcIndexProduct(category,productName,direction){
  const r=await fetch('/api/full-catalog/index-order-move',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({category,product_name:productName,direction})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  await loadFcIndexOrder()}
async function toggleFcIndexExclude(productName,excluded){
  const r=await fetch('/api/full-catalog/index-exclude',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({product_name:productName,excluded})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  await loadFcIndexOrder()}
// Section Order — reorder (existing move arrows) plus, per the underlying
// value staying fixed (Outdoor/Indoor/Striplight are wired into Product
// Type, Class 1/2/3 badge auto-assignment, and grouping throughout the
// app — renaming or deleting the VALUE would break all of that), a
// user-editable DISPLAY label and color instead. No add/delete here —
// there are always exactly the 3 structural values; Category Order is the
// free-growing list.
function renderFcSectionOrder(){
  $('fc-sectionorder').innerHTML=FC_SECTIONS.map((value,i)=>{
    const color=FC_SECTION_COLORS[value]||'#9a9a9a',label=FC_SECTION_LABELS[value]||value;
    return '<div class=dragrow draggable=true ondragstart="dragRowStart(event,\''+value+'\')" ondragover="dragRowOver(event)" ondragleave="dragRowLeave(event)" ondrop="fcSectionDrop(event,\''+value+'\')" ondragend="dragRowEnd(event)" style="display:flex;align-items:center;gap:8px;padding:6px 0;'+(i?'border-top:1px solid var(--border)':'')+'">'+
      '<span class=draghandle title="Drag to reorder">⠿</span>'+
      '<input type=color value="'+color+'" style="width:20px;height:20px;padding:0;border:1px solid var(--border);border-radius:4px;cursor:pointer;flex-shrink:0" title="Section color" onchange="setFcSectionColor(\''+value+'\',this.value)">'+
      '<span class=fcinlinename style="flex:1;font-size:12.5px;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;cursor:text" title="Click to edit display name" onclick="startFcSectionRename(this,\''+value+'\')">'+escHtml(label)+'</span>'+
      '<button type=button class=ordmovebtn'+(i===0?' disabled':'')+' onclick="moveFcSection(\''+value+'\',\'up\')" title="Move up">▲</button>'+
      '<button type=button class=ordmovebtn'+(i===FC_SECTIONS.length-1?' disabled':'')+' onclick="moveFcSection(\''+value+'\',\'down\')" title="Move down">▼</button>'+
    '</div>'}).join('')}
async function fcSectionDrop(e,targetValue){
  e.preventDefault();
  e.currentTarget.classList.remove('dragover-top','dragover-bottom');
  await dragReorder(targetValue,FC_SECTIONS,moveFcSection)}
async function moveFcSection(label,direction){
  const r=await fetch('/api/full-catalog/section-order-move',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({label,direction})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  FC_SECTIONS=r.order||FC_SECTIONS;renderFcSectionOrder()}
async function setFcSectionColor(section,hex){
  const r=await fetch('/api/full-catalog/section-color',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({section,hex})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  FC_SECTION_COLORS=r.colors||FC_SECTION_COLORS;renderFcSectionOrder()}
function startFcSectionRename(span,section){
  const inp=document.createElement('input');
  inp.type='text';inp.value=FC_SECTION_LABELS[section]||section;
  inp.style.cssText='flex:1;font-size:12.5px;min-width:0;border:1px solid var(--border);border-radius:4px;padding:2px 4px';
  let done=false;
  const commit=async()=>{
    if(done)return;done=true;
    const v=inp.value.trim();
    if(!v){renderFcSectionOrder();return}
    const r=await fetch('/api/full-catalog/section-label',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({section,label:v})}).then(r=>r.json());
    if(r.error){alert(r.error);renderFcSectionOrder();return}
    FC_SECTION_LABELS=r.labels||FC_SECTION_LABELS;renderFcSectionOrder();toast('Updated')};
  inp.addEventListener('keydown',e=>{
    if(e.key==='Enter'){e.preventDefault();inp.blur()}
    else if(e.key==='Escape'){done=true;renderFcSectionOrder()}});
  inp.addEventListener('blur',commit);
  span.replaceWith(inp);inp.focus();inp.select()}
// Front Cover is its own single fixed slot, always first — same static
// (non-reorderable) upload/replace/remove flow as Ending, just at the top
// of the card instead of the bottom (a cover can't meaningfully be dragged
// to a non-first position). Introduction and any custom pages the user
// adds live together in one reorderable "front matter" list
// (catalog_extras.front_matter) in between — arrows work exactly like
// Category Order's own move-up/down (moveFcCategory). "Introduction" is
// the list's one permanent builtin row (upload/reorder/clear-the-file
// only, never deletable — see renderFcFrontMatter); anything added via
// "+ Add Custom Page" is a normal row that can also be removed outright.
// Family Dividers further down reuses the same static-row flow too.
// fcExtraRowHtml renders one static row, pickPdfFile does the actual
// file-selection + base64 read (same pattern as pickPhoto() elsewhere in
// this file), and Remove uses this app's existing double-click-to-confirm
// pattern (window.confirm() silently no-ops here — see
// deleteManageListItem for the reference implementation).
let FC_EXTRAS={cover:{},front_matter:[],family_dividers:{}}, FC_FAMILY_NAMES=[];
function fcExtraRowHtml(label,info,isFirst,pickOnclick,removeOnclick){
  const filename=info&&info.filename?info.filename:'';
  return '<div style="display:flex;align-items:center;gap:8px;padding:8px 0;'+(isFirst?'':'border-top:1px solid var(--border)')+'">'+
    '<div style="flex:1;min-width:0">'+
      '<div style="font-size:12.5px;font-weight:600">'+escHtml(label)+'</div>'+
      '<div class=muted style="font-size:11px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+(filename?escHtml(filename):'Not set')+'</div>'+
    '</div>'+
    '<button type=button class=btn style="padding:6px 10px;font-size:11.5px;flex-shrink:0" onclick="'+pickOnclick+'">'+(filename?'Replace':'Choose File…')+'</button>'+
    (filename?'<button type=button class=rm style="flex-shrink:0" onclick="'+removeOnclick+'" title="Remove">×</button>':'')+
  '</div>'}
function renderFcFrontMatter(){
  const items=FC_EXTRAS.front_matter||[];
  $('fc-frontmatter').innerHTML=items.map((item,i)=>{
    const id=item.id.replace(/'/g,"\\'"),filename=item.filename?item.filename:'';
    return '<div class=dragrow draggable=true ondragstart="dragRowStart(event,\''+id+'\')" ondragover="dragRowOver(event)" ondragleave="dragRowLeave(event)" ondrop="fcFrontMatterDrop(event,\''+id+'\')" ondragend="dragRowEnd(event)" style="display:flex;align-items:center;gap:8px;padding:8px 0;'+(i?'border-top:1px solid var(--border)':'')+'">'+
      '<span class=draghandle title="Drag to reorder">⠿</span>'+
      '<div style="flex:1;min-width:0">'+
        '<div style="font-size:12.5px;font-weight:600">'+escHtml(item.label)+'</div>'+
        '<div class=muted style="font-size:11px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+(filename?escHtml(filename):'Not set')+'</div>'+
      '</div>'+
      '<button type=button class=ordmovebtn'+(i===0?' disabled':'')+' onclick="moveFcFrontMatter(\''+id+'\',\'up\')" title="Move up">▲</button>'+
      '<button type=button class=ordmovebtn'+(i===items.length-1?' disabled':'')+' onclick="moveFcFrontMatter(\''+id+'\',\'down\')" title="Move down">▼</button>'+
      '<button type=button class=btn style="padding:6px 10px;font-size:11.5px;flex-shrink:0" onclick="pickFcFrontMatter(\''+id+'\')">'+(filename?'Replace':'Choose File…')+'</button>'+
      (filename?'<button type=button class=rm style="flex-shrink:0" onclick="removeFcFrontMatter(\''+id+'\',this)" title="Remove">×</button>':'')+
    '</div>'}).join('')}
async function fcFrontMatterDrop(e,targetId){
  e.preventDefault();
  e.currentTarget.classList.remove('dragover-top','dragover-bottom');
  await dragReorder(targetId,(FC_EXTRAS.front_matter||[]).map(it=>it.id),moveFcFrontMatter)}
async function moveFcFrontMatter(id,direction){
  const r=await fetch('/api/full-catalog/front-matter-move',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id,direction})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  FC_EXTRAS=r.extras||FC_EXTRAS;renderFcFrontMatter()}
function pickFcFrontMatter(id){
  pickPdfFile(async(dataUrl,filename)=>{
    const r=await fetch('/api/full-catalog/front-matter',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id,pdf:dataUrl,filename})}).then(r=>r.json());
    if(r.error){alert(r.error);return}
    FC_EXTRAS=r.extras;renderFcFrontMatter();toast('Uploaded')})}
async function removeFcFrontMatter(id,btn){
  if(btn.dataset.confirm!=='1'){btn.dataset.confirm='1';btn.textContent='Sure?';setTimeout(()=>{if(btn.dataset.confirm==='1'){btn.dataset.confirm='';btn.textContent='×'}},2500);return}
  const r=await fetch('/api/full-catalog/front-matter/remove',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  FC_EXTRAS=r.extras;renderFcFrontMatter()}
async function addFcFrontMatter(){
  const inp=$('fc-newfrontmatter'),v=inp.value.trim();
  if(!v)return;
  const r=await fetch('/api/full-catalog/front-matter-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({label:v})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  FC_EXTRAS=r.extras;inp.value='';renderFcFrontMatter()}
function renderFcCover(){
  $('fc-cover').innerHTML=fcExtraRowHtml('Front Cover',FC_EXTRAS.cover,true,
    "pickFcExtra('cover')","removeFcExtra('cover',this)")}
function renderFcEnding(){
  $('fc-ending').innerHTML=fcExtraRowHtml('Back Cover / Ending',FC_EXTRAS.ending,true,
    "pickFcExtra('ending')","removeFcExtra('ending',this)")}
function renderFcFamilyDividers(){
  if(!FC_FAMILY_NAMES.length){
    $('fc-familydividers').innerHTML='<p class=muted style="font-size:12px;margin:0">No families yet — mark a product as part of a family on the Datasheet Builder form first.</p>';
    return}
  const dividers=FC_EXTRAS.family_dividers||{};
  $('fc-familydividers').innerHTML=FC_FAMILY_NAMES.map((family,i)=>{
    const safe=family.replace(/'/g,"\\'");
    return fcExtraRowHtml(family,dividers[family],i===0,
      "pickFcFamilyDivider('"+safe+"')","removeFcFamilyDivider('"+safe+"',this)")}).join('')}
function pickPdfFile(onLoaded){
  const inp=document.createElement('input');
  inp.type='file';inp.accept='.pdf,application/pdf';
  inp.onchange=()=>{
    const f=inp.files[0];if(!f)return;
    const reader=new FileReader();
    reader.onload=()=>onLoaded(reader.result,f.name);
    reader.readAsDataURL(f)};
  inp.click()}
function renderFcExtraSlot(slot){if(slot==='cover')renderFcCover();else renderFcEnding()}
function pickFcExtra(slot){
  pickPdfFile(async(dataUrl,filename)=>{
    const r=await fetch('/api/full-catalog/extras/'+slot,{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({pdf:dataUrl,filename})}).then(r=>r.json());
    if(r.error){alert(r.error);return}
    FC_EXTRAS=r.extras;renderFcExtraSlot(slot);toast('Uploaded')})}
async function removeFcExtra(slot,btn){
  if(btn.dataset.confirm!=='1'){btn.dataset.confirm='1';btn.textContent='Sure?';setTimeout(()=>{if(btn.dataset.confirm==='1'){btn.dataset.confirm='';btn.textContent='×'}},2500);return}
  const r=await fetch('/api/full-catalog/extras/'+slot+'/remove',{method:'POST'}).then(r=>r.json());
  FC_EXTRAS=r.extras;renderFcExtraSlot(slot)}
function pickFcFamilyDivider(family){
  pickPdfFile(async(dataUrl,filename)=>{
    const r=await fetch('/api/full-catalog/family-divider',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({family,pdf:dataUrl,filename})}).then(r=>r.json());
    if(r.error){alert(r.error);return}
    FC_EXTRAS=r.extras;renderFcFamilyDividers();toast('Uploaded')})}
async function removeFcFamilyDivider(family,btn){
  if(btn.dataset.confirm!=='1'){btn.dataset.confirm='1';btn.textContent='Sure?';setTimeout(()=>{if(btn.dataset.confirm==='1'){btn.dataset.confirm='';btn.textContent='×'}},2500);return}
  const r=await fetch('/api/full-catalog/family-divider/remove',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({family})}).then(r=>r.json());
  FC_EXTRAS=r.extras;renderFcFamilyDividers()}
async function saveFullCatalogFolder(){
  const v=$('set-full_catalog_folder').value.trim();
  const r=await fetch('/api/settings',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({full_catalog_folder:v})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  toast('Output folder saved')}
let fcBuilding=false;
async function buildFullCatalog(){
  if(fcBuilding)return;
  fcBuilding=true;
  const btn=$('fc-buildbtn'),origText=btn.textContent;
  btn.disabled=true;btn.textContent='Building… this can take up to a minute';
  try{
    const r=await fetch('/api/full-catalog/build',{method:'POST'}).then(r=>r.json());
    if(r.error){alert(r.error);return}
    renderFcLastBuild(r);
    toast('Catalogue built — '+r.total_pages+' pages')
  }finally{
    fcBuilding=false;btn.disabled=false;btn.textContent=origText}}
function renderFcLastBuild(result){
  if(!result||!result.total_pages){$('fc-lastbuild-card').style.display='none';fcSetPreviewTotal(0);return}
  $('fc-lastbuild-card').style.display='';
  let html='<p style="margin:0 0 10px"><b>'+result.total_pages+' pages</b> — built '+escHtml((result.built_at||'').replace('T',' '))+'</p>';
  html+='<div style="font-size:12px">'+result.sections.map(s=>
    '<div style="display:flex;justify-content:space-between;padding:4px 0;border-top:1px solid var(--border)">'+
      '<span>'+escHtml(s.label)+' <span class=muted>('+s.category_count+' categor'+(s.category_count!=1?'ies':'y')+', '+s.product_count+' product'+(s.product_count!=1?'s':'')+')</span></span>'+
      '<span class=num>p.'+s.start_page+'–'+s.end_page+'</span></div>').join('')+'</div>';
  if(result.warnings&&result.warnings.length)html+='<ul style="margin:10px 0 0;padding-left:18px;font-size:11.5px;color:var(--warning)">'+result.warnings.map(w=>'<li>'+escHtml(w)+'</li>').join('')+'</ul>';
  html+='<a href="/full-catalog-pdf" target=_blank class="btn dark" style="width:100%;display:block;text-align:center;box-sizing:border-box;margin-top:12px;text-decoration:none">Open Catalogue PDF</a>';
  $('fc-lastbuild').innerHTML=html;
  fcSetPreviewTotal(result.total_pages)}
// The right-hand preview pane: one or two pages at a time (Single/Double,
// same toggle as the per-document preview), fetched on demand from the
// already-built PDF (/full-catalog-page) — never all pages at once like
// the per-document preview does, since a real catalogue here can run 800+
// pages, so this is a pager (fcPrevPage/fcNextPage/fcJumpPage) rather than
// one long scrolling list. Reuses the exact same .pvpages/.pvrow/.pvpage
// markup + pbtn/zoomlabel CSS as the per-document preview's own toolbar —
// see renderPreviewPages()/computePageWidth() for the pattern this mirrors
// (Double here just means "page N and N+1 side by side", not a real
// odd/even print-spread convention, since a catalogue page's odd/even-ness
// carries no such meaning here).
let FC_PREVIEW_PAGE=1, FC_PREVIEW_TOTAL=0, fcPreviewImgToken=0;
let fcPreviewMode='single', fcPreviewModeAuto=true, fcPreviewZoom=100;
function fcEffectiveMode(){return (fcPreviewMode==='double'&&FC_PREVIEW_TOTAL>1)?'double':'single'}
function fcComputePageWidth(){
  const box=$('fc-previewbox');if(!box)return 600;
  const inner=box.clientWidth-32;
  const perRow=fcEffectiveMode()==='double'?2:1;
  const gap=14;
  let w=(inner-gap*(perRow-1))/perRow;
  w=Math.min(w,perRow===2?620:900);
  return Math.max(160,Math.round(w*(fcPreviewZoom/100)))}
function updateFcPvButtons(){
  const eff=fcEffectiveMode();
  $('fc-pm-single').classList.toggle('on',eff==='single');
  $('fc-pm-double').classList.toggle('on',eff==='double');
  $('fc-pm-double').disabled=FC_PREVIEW_TOTAL<2;
  $('fc-zoomlabel').textContent=fcPreviewZoom+'%'}
function fcRenderPreviewPages(){
  const wrap=$('fc-previewpages');if(!wrap||!FC_PREVIEW_TOTAL)return;
  const eff=fcEffectiveMode();
  wrap.classList.remove('single','double');wrap.classList.add(eff);
  const w=fcComputePageWidth();
  const token=++fcPreviewImgToken;
  const nums=eff==='double'&&FC_PREVIEW_PAGE<FC_PREVIEW_TOTAL?[FC_PREVIEW_PAGE,FC_PREVIEW_PAGE+1]:[FC_PREVIEW_PAGE];
  const empty=$('fc-previewempty');
  wrap.innerHTML='<div class=pvrow>'+nums.map(n=>'<img class=pvpage style="width:'+w+'px" data-page='+n+'>').join('')+'</div>';
  wrap.querySelectorAll('img.pvpage').forEach(img=>{
    img.onload=()=>{if(token!==fcPreviewImgToken)return;wrap.classList.remove('hide');empty.classList.add('hide')};
    img.onerror=()=>{if(token!==fcPreviewImgToken)return;empty.textContent='Could not render this page.';empty.classList.remove('hide')};
    img.src='/full-catalog-page?page='+img.dataset.page+'&t='+Date.now()});
  updateFcPvButtons()}
function fcSetPreviewTotal(total){
  FC_PREVIEW_TOTAL=total;
  const wrap=$('fc-previewpages'),empty=$('fc-previewempty');
  fcPreviewImgToken++;  // invalidate any in-flight fcRenderPreviewPages() load/error callback
  if(!total){
    wrap.classList.add('hide');wrap.innerHTML='';
    empty.textContent='Build the catalogue to preview it here.';empty.classList.remove('hide');
    $('fc-pagecount').textContent='';$('fc-pagejump').value='';
    updateFcPvButtons();
    return}
  if(fcPreviewModeAuto)fcPreviewMode=total>1?'double':'single';
  fcShowPage(1)}
function fcShowPage(n){
  if(!FC_PREVIEW_TOTAL)return;
  FC_PREVIEW_PAGE=Math.max(1,Math.min(n,FC_PREVIEW_TOTAL));
  $('fc-pagejump').value=FC_PREVIEW_PAGE;
  const eff=fcEffectiveMode();
  $('fc-pagecount').textContent=(eff==='double'&&FC_PREVIEW_PAGE<FC_PREVIEW_TOTAL)
    ?'Pages '+FC_PREVIEW_PAGE+'–'+(FC_PREVIEW_PAGE+1)+' of '+FC_PREVIEW_TOTAL
    :'Page '+FC_PREVIEW_PAGE+' of '+FC_PREVIEW_TOTAL;
  $('fc-previewpages').classList.remove('hide');
  $('fc-previewempty').classList.add('hide');
  fcRenderPreviewPages()}
function fcPrevPage(){fcShowPage(FC_PREVIEW_PAGE-(fcEffectiveMode()==='double'?2:1))}
function fcNextPage(){fcShowPage(FC_PREVIEW_PAGE+(fcEffectiveMode()==='double'?2:1))}
function fcJumpPage(){fcShowPage(parseInt($('fc-pagejump').value,10)||1)}
function fcSetPreviewMode(m){fcPreviewMode=m;fcPreviewModeAuto=false;fcShowPage(FC_PREVIEW_PAGE)}
function fcZoomPreview(delta){fcPreviewZoom=Math.max(40,Math.min(300,fcPreviewZoom+delta));fcRenderPreviewPages()}
function fcFitPreview(){fcPreviewZoom=100;fcRenderPreviewPages()}

// ---------------------------------------------------------------- Manage Lists + Activity Log (Settings)
// Every "preset dropdown + Custom… + remember" list (CCT/Controls/Voltage/
// Power/Beam Angle) plus every other plain list that grows over time (Units,
// Spec Labels, Series, Ordering column names, Size Index) shares one delete
// UI here, keyed against the backend's MANAGED_STRING_LISTS registry.
// Finish Colors is the same idea with a bespoke {label,hex} shape and its
// own remove-by-label route, so it gets its own render/delete pair below
// instead of going through the generic by-index route.
// Badge Library is deliberately NOT here — it's backed by real image files
// on disk (not a plain value list) and already has its own dedicated
// checkbox+thumbnail manager in the CAT builder.
const MANAGED_LISTS=[
  ['units','Units'],
  ['cat_spec_labels','Technical Spec Labels'],
  ['cat_series_labels','Series / Category'],
  ['cat_family_labels','Family Names'],
  ['cat_ordering_categories','Ordering Table Column Names'],
  ['cat_modelno_options','Model No. Presets'],
  ['cat_cct_options','CCT Presets'],
  ['cat_controls_options','Controls Presets'],
  ['cat_voltage_options','Input Voltage Presets'],
  ['cat_power_options','Power Presets'],
  ['cat_beamangle_options','Beam Angle Presets'],
  ['cat_cutout_options','Cut Out Presets'],
  ['cat_options_options','Options Presets'],
  ['cat_size_index','Size Presets'],
  ['expense_payment_methods','Payment Methods'],
  ['expense_employees','Employees'],
  ['expense_categories','Expense Categories'],
  ['expense_products','Expense Products'],
  ['expense_descriptions','Expense Descriptions'],
];
const MANAGE_LIST_REFRESH={
  units:loadUnits,cat_spec_labels:loadCatSpecLabels,cat_series_labels:loadCatSeriesLabels,
  cat_family_labels:loadCatFamilyLabels,
  cat_ordering_categories:loadCatOrdCategories,cat_modelno_options:loadCatModelNoOptions,
  cat_cct_options:loadCatCctOptions,
  cat_controls_options:loadCatControlsOptions,cat_voltage_options:loadCatVoltageOptions,
  cat_power_options:loadCatPowerOptions,cat_beamangle_options:loadCatBeamAngleOptions,
  cat_cutout_options:loadCatCutOutOptions,cat_options_options:loadCatOptionsOptions,
  cat_size_index:loadCatSizeIndex,
  expense_payment_methods:loadExpPaymentMethods,
  expense_employees:loadExpEmployees,expense_categories:loadExpCategories,
  expense_products:loadExpProducts,expense_descriptions:loadExpDescriptions,
};
// cat_size_index used to need a visible standing warning here — its
// position once WAS the "D<n>" printed on generated datasheets, so
// deleting/reordering anything but the last entry silently renumbered
// every later size. That's no longer true: each datasheet's own D-numbers
// are computed fresh from that ONE product's own Size column (see the
// front end's recomputeCatOrdSizeDNumbers), so this list is now just a
// plain suggestion list, same "delete = gone from future suggestions,
// touches nothing already generated" contract every other preset list
// here already has — no special warning needed anymore.
const MANAGE_LIST_WARN={};
let MANAGE_LISTS_COLLAPSED=new Set(['units','cat_spec_labels','cat_series_labels','cat_family_labels','cat_ordering_categories','cat_modelno_options','cat_cct_options','cat_controls_options','cat_voltage_options','cat_power_options','cat_beamangle_options','cat_cutout_options','cat_options_options','cat_size_index','cat_finish_colors']);
// Live client-side cache of each list's current items, keyed by list_key —
// lets a drag-drop's dragReorder() steps (and the ◀▶ buttons) resolve
// "index of this value right now" without a round-trip per step, same
// role CAT_STANDARD_FILL_ORDER plays for the Fill Standard Information
// popover. Populated fresh on every renderManageLists() and kept in sync
// by manageListMove()'s own response.
let MANAGE_LIST_ITEMS={};
async function renderManageLists(){
  const body=$('managelists-body');
  if(!body)return;
  const sections=await Promise.all(MANAGED_LISTS.map(async([key,label])=>{
    const r=await fetch('/api/settings-list/'+key).then(r=>r.json());
    MANAGE_LIST_ITEMS[key]=r.items||[];
    return manageListSectionHtml(key,label,r.items||[])}));
  const fc=await fetch('/api/cat-finish-colors').then(r=>r.json());
  sections.push(manageFinishColorSectionHtml(fc.colors||[]));
  body.innerHTML=sections.join('')}
// Every managed list is draggable to reorder (HTML5 drag-and-drop, same
// dragRowStart/dragColOver/dragColLeave/dragRowEnd + dragReorder plumbing
// as every other reorderable list in this app — chips wrap horizontally
// here, so the left/right insert cue (dragColOver) reads correctly instead
// of the vertical top/bottom one), with the ◀▶ buttons kept alongside as a
// precise-single-step fallback, same convention as everywhere else. The
// drag identity is the VALUE, not the array index — indices shift mid-drag
// as steps replay, but /api/settings-list-move is itself index-based (so
// two identical-looking entries still resolve unambiguously — see that
// route's own docstring), so manageListMove() re-resolves "index of this
// value right now" from MANAGE_LIST_ITEMS at the moment each step fires.
// cat_size_index used to be a fixed exception (its position once WAS the
// printed "D" number) — now just a plain suggestion list like every other
// preset here, so it's fully draggable too, no special case left.
const MANAGE_LIST_KEYSEP=':::';
function manageListSectionHtml(key,label,items){
  const collapsed=MANAGE_LISTS_COLLAPSED.has(key);
  const warn=MANAGE_LIST_WARN[key]?'<div class=mlwarn>'+escHtml(MANAGE_LIST_WARN[key])+'</div>':'';
  const chips=items.length?items.map((v,i)=>{
      const text=escHtml(String(v)),vSafe=String(v).replace(/'/g,"\\'").replace(/\\/g,'\\\\');
      return '<span class="dspill dragrow" draggable=true ondragstart="dragRowStart(event,\''+key+MANAGE_LIST_KEYSEP+vSafe+'\')" ondragover="dragColOver(event)" ondragleave="dragColLeave(event)" ondrop="manageListDrop(event,\''+key+'\',\''+vSafe+'\')" ondragend="dragRowEnd(event)">'+
        '<span class=draghandle title="Drag to reorder" onclick="event.stopPropagation()">⠿</span>'+text+
        '<button type=button class=ordmovebtn'+(i===0?' disabled':'')+' onclick="manageListMove(\''+key+'\',\''+vSafe+'\',\'up\')" title="Move earlier">◀</button>'+
        '<button type=button class=ordmovebtn'+(i===items.length-1?' disabled':'')+' onclick="manageListMove(\''+key+'\',\''+vSafe+'\',\'down\')" title="Move later">▶</button>'+
        '<button type=button class=mlx onclick="deleteManageListItem(\''+key+'\','+i+',\''+label+'\',this)">✕</button></span>'}).join('')
    :'<span class=mlempty>Nothing here yet</span>';
  return '<div class="mlsection'+(collapsed?' collapsed':'')+'" data-key="'+key+'">'+
    '<div class=mlhead onclick="toggleManageList(this)">'+escHtml(label)+
      '<span class=mlcount>'+items.length+'</span><span class=mlchev>▾</span></div>'+
    '<div class=mlbody>'+chips+warn+'</div></div>'}
async function manageListDrop(e,key,targetValue){
  e.preventDefault();e.stopPropagation();
  e.currentTarget.classList.remove('dragover-left','dragover-right');
  if(!DRAG_KEY)return;
  const sep=DRAG_KEY.indexOf(MANAGE_LIST_KEYSEP);
  if(sep<0||DRAG_KEY.slice(0,sep)!==key)return;  // cross-list drags aren't a supported gesture
  const currentOrder=(MANAGE_LIST_ITEMS[key]||[]).map(v=>key+MANAGE_LIST_KEYSEP+v);
  await dragReorder(key+MANAGE_LIST_KEYSEP+targetValue,currentOrder,
    (k,dir)=>manageListMove(key,k.slice(k.indexOf(MANAGE_LIST_KEYSEP)+MANAGE_LIST_KEYSEP.length),dir))}
async function manageListMove(key,value,direction){
  const items=MANAGE_LIST_ITEMS[key]||[];
  const index=items.indexOf(value);
  if(index<0)return;
  const r=await fetch('/api/settings-list-move',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({list_key:key,index,direction})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  if(r.items){
    MANAGE_LIST_ITEMS[key]=r.items;
    const refresh=MANAGE_LIST_REFRESH[key];
    if(refresh)await refresh();
    renderManageLists()}}
function manageFinishColorSectionHtml(colors){
  const collapsed=MANAGE_LISTS_COLLAPSED.has('cat_finish_colors');
  const chips=colors.length?colors.map(c=>{
      const safeLabel=escHtml(c.label||'').replace(/'/g,"\\'");
      return '<span class=dspill><span class=mlswatch style="background:'+escHtml(c.hex||'#fff')+'"></span>'+escHtml(c.label||'')+
        '<button type=button class=mlx onclick="deleteFinishColorFromSettings(\''+safeLabel+'\',this)">✕</button></span>'}).join('')
    :'<span class=mlempty>Nothing here yet</span>';
  return '<div class="mlsection'+(collapsed?' collapsed':'')+'" data-key="cat_finish_colors">'+
    '<div class=mlhead onclick="toggleManageList(this)">Finish Colors'+
      '<span class=mlcount>'+colors.length+'</span><span class=mlchev>▾</span></div>'+
    '<div class=mlbody>'+chips+'</div></div>'}
function toggleManageList(head){
  const sec=head.closest('.mlsection'),key=sec.dataset.key;
  sec.classList.toggle('collapsed');
  if(sec.classList.contains('collapsed'))MANAGE_LISTS_COLLAPSED.add(key);else MANAGE_LISTS_COLLAPSED.delete(key)}
// window.confirm() silently no-ops in this environment — same inline
// double-click-to-confirm pattern as deleteClient() elsewhere in this file.
function deleteManageListItem(key,index,label,btn){
  if(btn.dataset.confirm!=='1'){
    btn.dataset.confirm='1';btn.textContent='Sure?';
    setTimeout(()=>{if(btn.dataset.confirm==='1'){btn.dataset.confirm='';btn.textContent='✕'}},2500);
    return}
  actuallyDeleteManageListItem(key,index,label)}
async function actuallyDeleteManageListItem(key,index,label){
  const r=await fetch('/api/settings-list-remove',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({list_key:key,index})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  const refresh=MANAGE_LIST_REFRESH[key];
  if(refresh)await refresh();
  renderManageLists();renderAuditLog();
  toast('Removed from '+label)}
function deleteFinishColorFromSettings(label,btn){
  if(btn.dataset.confirm!=='1'){
    btn.dataset.confirm='1';btn.textContent='Sure?';
    setTimeout(()=>{if(btn.dataset.confirm==='1'){btn.dataset.confirm='';btn.textContent='✕'}},2500);
    return}
  actuallyDeleteFinishColorFromSettings(label)}
async function actuallyDeleteFinishColorFromSettings(label){
  await fetch('/api/cat-finish-colors-remove',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({label})}).then(r=>r.json());
  await loadCatFinishColors();
  renderManageLists();renderAuditLog();
  toast('Removed from Finish Colors')}
const AUDIT_ACTION_LABEL={add:'Added to',remove:'Removed from',rename:'Renamed in'};
async function renderAuditLog(){
  const body=$('auditlog-body');
  if(!body)return;
  const r=await fetch('/api/audit-log').then(r=>r.json());
  const log=r.log||[];
  if(!log.length){body.innerHTML='<p class=mlempty style="margin:0">No changes yet.</p>';return}
  body.innerHTML=log.map(e=>{
    const verb=AUDIT_ACTION_LABEL[e.action]||e.action;
    const value=(e.value&&typeof e.value==='object')?(e.value.label||JSON.stringify(e.value)):e.value;
    return '<div class=alrow><span class=altime>'+escHtml((e.ts||'').replace('T',' '))+'</span>'+
      '<span>'+verb+' <b>'+escHtml(e.list_label||e.list_key||'')+'</b>: '+escHtml(String(value??''))+'</span></div>'
  }).join('')}

const HEAD={
 INV:[['number','INV Number'],['date','Date','date'],['qtn_number','QTN Number'],['lpo_number','LPO Number'],['project','Project'],['type','Type']],
 DO:[['number','DO No'],['date','Date','date'],['project','Project'],['lpo_number','LPO no']],
 QTN2:[['number','QTN Number'],['rev','Rev','0'],['date','Date','date'],['project','Project'],['area','Area']],
 CAT:[['number','Sheet No.','0'],['date','Date','date']],
 EXP:[['number','Report No.','0'],['date','Date','date'],['period_from','Period From (auto)','date','readonly'],['period_to','Period To (auto)','date','readonly']],
};
const COLS={
 INV:[['photo','Photo'],['description','Item Description'],['unit','Unit','PCS'],['qty','Qty'],['price','Price']],
 DO:[['photo','Photo'],['description','Item Description'],['unit','Unit','PCS'],['lpo_qty','LPO Qty'],['prev_delivery','Prev. Delivery'],['delivered','Delivered']],
 QTN2:[['type','Type'],['photo','Photo'],['description','Item Description'],['unit','Unit','PCS'],['qty','Qty'],['price','Price']],
 CAT:[],
 EXP:[],
};
const LABEL={INV:'Tax Invoice',DO:'Delivery Order',QTN2:'Quotation',PI:'Proforma Invoice',RV:'Payment Receipt',CN:'Credit Note',CAT:'Sololuce Datasheet',EXP:'Expense Report'};

// PI/RV/CN (Proforma Invoice, Payment Receipt, Credit Note) are
// HTML_DOC_TYPEs (PDF-only, pixel-fidelity pipeline — see
// engine.HTML_DOC_TYPES) but intentionally NOT in VISIBLE_TYPES/DOC_VIEWS:
// each is generate-only-from-another-document for now, via the
// "Generate ▾" menu on a Quotation/Proforma Invoice/Delivery Order/Tax
// Invoice row (see openQtnGenMenu/runRowGenerate) — no "start a blank one"
// sidebar entry yet.
const HTML_DOC_TYPES=['QTN2','CAT','EXP','PI','RV','CN'];
// Of those, only QTN2/CAT/EXP have a real edit FORM (HEAD/COLS entries or
// a dedicated populate*Form()) — PI/RV/CN have neither, so routing them
// through openDoc() would crash on HEAD[TYPE] being undefined. See the
// isEditableHtmlDocType comment on the All Docs row actions below.
const EDITABLE_HTML_DOC_TYPES=['QTN2','CAT','EXP'];
const VISIBLE_TYPES=['QTN2','INV','DO','CAT','EXP'];
function startNewDocument(){setType(VISIBLE_TYPES.includes(TYPE)?TYPE:'QTN2')}
// Which document type the form is currently showing. Each type is its own
// sidebar entry (see DOC_VIEWS/openDocType) — the rail highlight is handled
// by view(), so this is purely about swapping the form's own cards.
function setType(t,silent){TYPE=t;
  renderHead();
  const isCat=t==='CAT', isExp=t==='EXP';
  $('headcard').classList.toggle('hide',isCat);
  document.querySelectorAll('.cat-only').forEach(el=>el.classList.toggle('hide',!isCat));
  document.querySelectorAll('.exp-only').forEach(el=>el.classList.toggle('hide',!isExp));
  $('impbtn').classList.toggle('hide',!isCat);
  $('clientcard').classList.toggle('hide',isCat||isExp);
  $('itemscard').classList.toggle('hide',isCat||isExp);
  $('discvat-card').classList.toggle('hide',t==='DO'||isCat||isExp);
  $('qtn2-statusseg-top').classList.toggle('hide',t!=='QTN2');
  $('qtn2-terms-card').classList.toggle('hide',t!=='QTN2');
  $('company').classList.toggle('hide',t==='QTN2');
  $('company-rich').classList.toggle('hide',t!=='QTN2');
  const htmlOnly=HTML_DOC_TYPES.includes(t);
  $('genbtn').textContent=htmlOnly?'Generate PDF':'Generate Excel + PDF';
  $('gencaption').textContent=htmlOnly
    ?'Saves a PDF into your folder, named by the company convention. Preview on the right is the same PDF.'
    :'Saves both files into your folder, named by the company convention. Preview on the right is the PDF.';
  if(!silent){EDITING=null;EDITING_DRAFT=null;$('title').textContent='New '+LABEL[t];items=[{}];SELECTED_ITEMS.clear();renderItems();onCompany();nextNumber(true);resetDiscVat();
    setCompanyVal('');$('customer_attn').innerHTML='';$('customer_address').innerHTML='';
    $('customer_pobox').value='';$('customer_city').value='';setCountryValue('customer','');
    setQtn2Status('Draft');resetTerms();
    if(isCat)resetCatForm();
    if(isExp)resetExpForm();
    // Switching tools is a discrete, deliberate action, not a keystroke —
    // skip the normal 700ms typing-debounce and render right away, so the
    // preview pane doesn't sit on the PREVIOUS tool's page for another
    // three-quarters of a second after you've already moved on.
    runPreview()}}
// The standard layout every new Sololuce Datasheet always starts with —
// stays fixed regardless of what custom labels get remembered below (those
// only power the autocomplete suggestions, so a one-off custom spec on one
// product doesn't creep into becoming part of every future product's
// default template).
const CAT_DEFAULT_SPEC_LABELS=['Power','Lifespan','Light Source','Luminare Efficacy','Power Factor','Ambient Temperature','Body Material','Diffuser','Mounting Type','IP Rating','Driver'];
const CAT_DEFAULT_NOTE='**Note: Sololuce products only employ standard drivers, although other driver options are available by request.';
// The Ordering/Variant Table's standard columns — every brand-new datasheet
// starts with these already in place (matching the real catalogue's own
// convention) instead of an empty table the user has to build from scratch
// every time. "Size" defaults to the Ø×H (round-fitting) variant since
// that's the one named first — swap the column to "Size (L×W×H)" via the
// picker for a rectangular fitting, same as any other column.
const CAT_DEFAULT_ORD_LABELS=['Model No.','Power','Size (Ø×H)','Cut Out','CCT','Beam Angle','Input Voltage','Controls','Finish Options','Options','Lumen'];
async function loadCatNextPage(){
  const r=await fetch('/api/cat-next-page').then(r=>r.json());
  $('cat-pagenum').value=r.next_page||1}
function resetCatForm(){
  ['cat-productname','cat-description','cat-ordcode'].forEach(id=>{$(id).value=''});
  loadCatNextPage();
  $('cat-producttype').value='';
  renderCatSeriesField('');
  $('cat-family-enabled').checked=false;
  $('cat-family-details').classList.add('hide');
  renderCatFamilyField('');
  $('cat-note').value=CAT_DEFAULT_NOTE;
  CAT_BADGES=CAT_STANDARD_BADGE_KEYS.map(key=>({key}));
  CAT_SPECS=CAT_DEFAULT_SPEC_LABELS.map(label=>({label,values:['']}));
  CAT_FINISH=[];CAT_COL_SEQ=0;CAT_ORD_COL_CLIPBOARD=null;CAT_ORD_CURRENT_VARIANT=0;CAT_ORD_ALIGN_ROWS=false;
  CAT_ORD_COLS=CAT_DEFAULT_ORD_LABELS.map(label=>({key:'col'+(CAT_COL_SEQ++),label,values:['']}));
  CAT_IMG={main:catImgDefault(),lifestyle:catImgDefault(),diagram:catImgDefault(),extra1:catImgDefault(),extra2:catImgDefault(),extra3:catImgDefault()};
  renderCatBadges();renderCatSpecs();renderCatFinish();renderCatOrdTable();renderCatImages()}
function renderHead(){
  const isQtn2=TYPE==='QTN2';
  $('headfields').innerHTML='<div class=g2>'+HEAD[TYPE].map(f=>{
    if(isQtn2&&(f[0]==='project'||f[0]==='area')){
      return `<div class=f><label>${f[1]}</label><div id="h-${f[0]}" class=richbox contenteditable data-placeholder="e.g. ${f[1]}" oninput="schedulePreview()"></div></div>`}
    const t=f[2]=='date'?'date':'text';const v=f[2]&&f[2]!='date'?f[2]:'';
    const list=f[0]==='project'?' list=projects':'';
    const ro=f[3]==='readonly'?' readonly':'';
    return `<div class=f><label>${f[1]}</label><input id="h-${f[0]}" type=${t} value="${v}"${list}${ro} oninput="schedulePreview()"></div>`}).join('')+'</div>';
  if(HEAD[TYPE].find(f=>f[2]=='date')){const d=$('h-date');if(d&&!d.value)d.value=new Date().toISOString().slice(0,10)}
  if(isQtn2){const pr=$('h-project');if(pr)attachAutocomplete(pr,()=>PROJECTS)}
}

let discMode='amount';
function setDiscMode(m){discMode=m;document.querySelectorAll('#disc-modeseg button').forEach(b=>b.classList.toggle('on',b.dataset.m===m));
  $('disc-value').placeholder=m==='target'?'e.g. 300000':'e.g. 10% or 500';
  schedulePreview()}
function onDiscVatChange(){
  const discOn=$('disc-on').checked;
  $('disc-value').disabled=!discOn;
  document.querySelectorAll('#disc-modeseg button').forEach(b=>b.disabled=!discOn);
  $('vat-value').disabled=!$('vat-on').checked;
  schedulePreview()}
function resetDiscVat(){
  $('disc-on').checked=false;$('disc-value').value='';
  $('vat-on').checked=true;$('vat-value').value='5%';
  discMode='amount';setDiscMode('amount');onDiscVatChange()}
function parsePctOrFixed(str){
  str=(str||'').trim();
  if(!str)return{mode:'fixed',value:0};
  if(str.endsWith('%'))return{mode:'percent',value:parseFloat(str.slice(0,-1))||0};
  return{mode:'fixed',value:parseFloat(str)||0}}
function collectDiscVat(){
  let discount={enabled:false},vat={enabled:false};
  if($('disc-on').checked){
    if(discMode==='target')discount={enabled:true,mode:'target',value:parseFloat($('disc-value').value)||0};
    else{const p=parsePctOrFixed($('disc-value').value);discount={enabled:true,mode:p.mode,value:p.value}}
  }
  if($('vat-on').checked){const p=parsePctOrFixed($('vat-value').value);vat={enabled:true,mode:p.mode,value:p.value}}
  return{discount,vat}}
let qtn2Status='Draft';
function setQtn2Status(s){qtn2Status=s;document.querySelectorAll('#qtn2-statusseg-top button').forEach(b=>b.classList.toggle('on',b.dataset.s===s));schedulePreview()}
function composedCustomerAddress(){
  const base=richText($('customer_address'));
  const pobox=$('customer_pobox').value.trim(),city=$('customer_city').value.trim();
  const country=$('customer-country').value;
  const extra=[pobox?('PO Box '+pobox):'',city,countryName(country)].filter(Boolean).join(', ');
  if(!extra)return base;
  return base?base+'<br>'+escHtml(extra):escHtml(extra)}
function collectQtn2Extra(){
  return{customer_attn:richText($('customer_attn')),
    customer_address:composedCustomerAddress(),   // final, composed-with-pobox/city/country — what the PDF prints
    customer_address_raw:richText($('customer_address')),   // just the Address box's own text — for restoring the field itself (drafts), never re-compose from the composed version or pobox/city/country would double up
    customer_po_box:$('customer_pobox').value.trim(),
    customer_city:$('customer_city').value.trim(),
    customer_country:$('customer-country').value,
    status:qtn2Status,
    terms:collectTerms(),
    terms_ui:{delivery:termsDeliveryValue(),payment:TERMS_PAYMENT,warranty:TERMS_WARRANTY}}}
// ---------------------------------------------------------------- Terms & Conditions (Quotation New Design only)
// Delivery is a preset-range dropdown ("ranged table" per the user's own
// words) with a Custom escape hatch, same pattern as Unit's "Custom…"
// elsewhere in this app. Payment is a variable list of {percent,label}
// stages (2 by default — Advance/Upon Delivery, the existing standard —
// but a 3rd/4th stage can be added for e.g. an interim milestone).
// Warranty is single-select among 3/5/7/10 years (5 is standard).
const TERMS_DELIVERY_PRESETS=['1-2 weeks','2-4 weeks','4-6 weeks','6-8 weeks','8-10 weeks','10-12 weeks'];
let TERMS_PAYMENT=[{percent:50,label:'Advance'},{percent:50,label:'Upon Delivery'}];
let TERMS_WARRANTY='5';
function termsDeliverySelectHtml(value){
  const v=value||'10-12 weeks';
  return '<select id=terms-delivery onchange="onTermsDeliveryChange()">'+
    TERMS_DELIVERY_PRESETS.map(p=>'<option'+(p===v?' selected':'')+'>'+p+'</option>').join('')+
    (v&&!TERMS_DELIVERY_PRESETS.includes(v)?'<option selected>'+escHtml(v)+'</option>':'')+
    '<option value="__custom__">Custom…</option></select>'}
function renderTermsDeliveryDefault(){$('terms-delivery-wrap').innerHTML=termsDeliverySelectHtml()}
function termsDeliveryValue(){const el=$('terms-delivery');return el?el.value:'10-12 weeks'}
function onTermsDeliveryChange(){
  const sel=$('terms-delivery');
  if(sel.value!=='__custom__'){schedulePreview();return}
  sel.outerHTML='<div style="display:flex;gap:6px"><input type=text id=terms-delivery style="flex:1" placeholder="e.g. 6-8 weeks" oninput=schedulePreview()>'+
    '<button type=button class=btn style="padding:0 11px" onclick=revertTermsDelivery() title="Choose from the list instead">▾</button></div>';
  $('terms-delivery').focus()}
function revertTermsDelivery(){$('terms-delivery-wrap').innerHTML=termsDeliverySelectHtml();schedulePreview()}
function renderPaymentStages(){
  $('terms-payment-rows').innerHTML=TERMS_PAYMENT.map((p,i)=>
    '<div style="display:flex;gap:6px;align-items:center;margin-bottom:6px">'+
      '<input type=number min=0 max=100 style="width:64px" value="'+escHtml(String(p.percent??''))+'" oninput="TERMS_PAYMENT['+i+'].percent=this.value;schedulePreview()">'+
      '<span class=muted style="font-size:12px;flex-shrink:0">%</span>'+
      '<input style="flex:1" value="'+escHtml(p.label||'')+'" placeholder="e.g. Advance" oninput="TERMS_PAYMENT['+i+'].label=this.value;schedulePreview()">'+
      (TERMS_PAYMENT.length>1?'<button type=button class=rm onclick="removePaymentStage('+i+')" title="Remove stage">×</button>':'')+
    '</div>').join('')}
function addPaymentStage(){TERMS_PAYMENT.push({percent:0,label:''});renderPaymentStages();schedulePreview()}
function removePaymentStage(i){TERMS_PAYMENT.splice(i,1);renderPaymentStages();schedulePreview()}
function setWarranty(y){TERMS_WARRANTY=y;document.querySelectorAll('#terms-warranty-seg button').forEach(b=>b.classList.toggle('on',b.dataset.w===y));schedulePreview()}
function resetTerms(){
  renderTermsDeliveryDefault();
  TERMS_PAYMENT=[{percent:50,label:'Advance'},{percent:50,label:'Upon Delivery'}];
  renderPaymentStages();
  setWarranty('5')}
function collectTerms(){
  const delivery=termsDeliveryValue();
  const payment=TERMS_PAYMENT.filter(p=>(p.label||'').trim()||Number(p.percent)>0)
    .map(p=>(p.percent||0)+'% '+(p.label||'').trim()).filter(Boolean).join(', ');
  const warranty=TERMS_WARRANTY+' years warranty';
  return {delivery, payment: payment||undefined, warranty}}
// QTN/INV/DO's own xlsx templates only have ONE free-text "To" cell for the
// customer block (not separate name/address/PO-box/etc fields — confirmed
// by inspecting the bundled templates), so everything the same Attn/Address/
// PO Box/City/Country fields hold gets composed into one plain-text (no
// HTML) multi-line block for that cell, company name included since there's
// no separate slot for it the way QTN2's own template has.
function customerBlockForXlsx(){
  const attn=$('customer_attn').textContent.trim();
  const addr=$('customer_address').textContent.trim();
  const pobox=$('customer_pobox').value.trim(),city=$('customer_city').value.trim();
  const country=countryName($('customer-country').value);
  const extra=[pobox?('PO Box '+pobox):'',city,country].filter(Boolean).join(', ');
  return[companyVal(),attn,addr,extra].filter(Boolean).join('\n')}
function isProductBuilt(desc){return /CODE:/i.test(desc||'')}
function itemUnitFieldHtml(i,it,fallback){
  const v=it.unit??(fallback||UNITS[0]||'');
  const opts=UNITS.map(u=>'<option'+(u===v?' selected':'')+'>'+u+'</option>').join('')
    +(v&&!UNITS.includes(v)?'<option selected>'+escHtml(v)+'</option>':'')
    +'<option value="__custom__">Custom…</option>'
    +'<option value="__manage__">Manage units…</option>';
  return '<div class=itcardfield><label>Unit</label><select data-i='+i+' data-k=unit onchange="onUnitChange('+i+',this)">'+opts+'</select></div>'}
function itemCardHtml(it,i,cols,hasType){
  const p=it.photo,sug=it.suggestedPhoto;
  let photoInner;
  if(p){photoInner='<img class=itemphoto src="'+p+'" onclick="pickPhoto('+i+')" title="Click to replace"><button class=phrm onclick="event.stopPropagation();removePhoto('+i+')" title="Remove photo">×</button>'}
  else if(sug){photoInner='<img class="itemphoto sugg" src="'+sug+'" onclick="openPhotoPicker('+i+',event)" title="Suggested photo — click to choose">'}
  else{photoInner='<button class=phbtn onclick="pickPhoto('+i+')" title="Click to add a photo"></button>'}

  const typeHtml=hasType?'<input class=itcardtype placeholder=Type value="'+escHtml(it.type??'')+'" oninput="upd('+i+',\'type\',this.value)">':'';

  const descHtml=TYPE==='QTN2'
    ?'<div class="richbox itcarddesc" contenteditable data-placeholder="Item description" oninput="upd('+i+',\'description\',this.innerHTML)" onblur="matchPhotoForItem('+i+');matchDatasheetsForItem('+i+')">'+(it.description||'')+'</div>'
    :'<textarea class=itcarddesc oninput="upd('+i+',\'description\',this.value)" onblur="matchPhotoForItem('+i+');matchDatasheetsForItem('+i+')">'+escHtml(it.description)+'</textarea>';

  // Once a description already has a CODE: line (the app's own composed
  // format — and the same convention real historical descriptions already
  // used before this feature existed), the primary action for that
  // datasheet is re-opening what's already configured, so the button reads
  // "Edit" instead of "Build". The datasheet-preview action becomes a
  // small icon-only secondary button instead of competing for attention.
  const isBuilt=isProductBuilt(it.description);
  const dsHtml=(it.datasheets||[]).map(d=>
    '<button type=button class=dspill data-i="'+i+'" data-rel="'+escHtml(d.rel)+'" data-name="'+escHtml(d.name)+'" onclick="openProductBuilder(+this.dataset.i,this.dataset.rel,this.dataset.name)" title="'+(isBuilt?'Edit':'Build')+' the item description from this datasheet\'s options">'+
      (isBuilt?'✏️ Edit':'🛠 Build')+' — '+escHtml(d.name)+
    '</button>'+
    '<button type=button class="dspill dspillicon" data-rel="'+escHtml(d.rel)+'" data-name="'+escHtml(d.name)+'" onclick="openDatasheetCS(this.dataset.rel,this.dataset.name)" title="Open the datasheet PDF">📄</button>'
  ).join('');

  const metaHtml=cols.filter(c=>!['photo','description','type','unit'].includes(c[0])).map(c=>
    '<div class=itcardfield><label>'+escHtml(c[1])+'</label><input data-i='+i+' data-k='+c[0]+' value="'+escHtml(it[c[0]]??(c[2]||''))+'" oninput="upd('+i+',\''+c[0]+'\',this.value)"></div>').join('');
  const unitHtml=cols.some(c=>c[0]==='unit')?itemUnitFieldHtml(i,it,cols.find(c=>c[0]==='unit')[2]):'';

  return '<div class="itemcard'+(SELECTED_ITEMS.has(i)?' selected':'')+'" data-i='+i+' oncontextmenu="openItemCtxMenu('+i+',event)">'+
    '<div class=itemcardtop>'+
      '<input type=checkbox class=itemcardsel data-i='+i+(SELECTED_ITEMS.has(i)?' checked':'')+' onchange="toggleItemSelect('+i+',this.checked)" title="Select this line">'+
      '<div class="photobox'+(sug&&!p?' sugg':'')+'">'+photoInner+'</div>'+
      typeHtml+
      '<button class=rm style="margin-left:auto" onclick=delRow('+i+') title="Remove line">×</button>'+
    '</div>'+
    '<div class=itemcardfind>'+
      '<button type=button class=ffind onclick="openProductFinder('+i+',event)">🔍 Find Product…</button>'+dsHtml+
    '</div>'+
    descHtml+
    '<div class=itemcardmeta>'+unitHtml+metaHtml+'</div>'+
  '</div>'}
// ---------------------------------------------------------------- multi-select line items (choose several "boxes" at once, then Clone/Cut/Delete as a batch)
// SELECTED_ITEMS holds indices into the CURRENT items[] array — cleared
// whenever the array's own shape changes in a way indices can't survive
// (delRow, a bulk action, or switching document type/loading a different
// document). ITEM_CLIPBOARD deliberately survives all of that: "cut" in
// one document, switch type or open a different draft, "paste" there —
// the whole point of a clipboard is that it outlives the selection.
let SELECTED_ITEMS=new Set(), ITEM_CLIPBOARD=[], ITEMS_SELECT_MODE=false;
// ---------------------------------------------------------------- shared right-click context menu (line items + All Docs both use this one popover)
function showCtxMenu(ev,itemsHtml){
  ev.preventDefault();ev.stopPropagation();
  const box=$('ctxmenu');
  const x=Math.min(ev.clientX,window.innerWidth-206),y=Math.min(ev.clientY,window.innerHeight-8);
  box.style.left=Math.max(8,x)+'px';box.style.top=Math.max(8,y)+'px';
  box.style.display='block';
  $('ctxmenu-list').innerHTML=itemsHtml}
function closeCtxMenu(){$('ctxmenu').style.display='none'}
document.addEventListener('click',e=>{if($('ctxmenu').style.display==='block'&&!e.target.closest('#ctxmenu'))closeCtxMenu()});
document.addEventListener('contextmenu',e=>{if($('ctxmenu').style.display==='block'&&!e.target.closest('.itemcard')&&!e.target.closest('.list .row'))closeCtxMenu()});
function enterItemsSelectMode(){ITEMS_SELECT_MODE=true;$('itemslist').classList.add('selectmode')}
function exitItemsSelectMode(){ITEMS_SELECT_MODE=false;$('itemslist').classList.remove('selectmode');clearItemSelection()}
// Right-click on a card that's already part of the current selection acts
// on the WHOLE selection (so you can right-click any one of several
// checked cards to bulk-act on all of them); right-click on a card that
// ISN'T selected replaces the selection with just that one — the same
// convention Explorer/Finder use, so a stray right-click never silently
// bulk-deletes cards you didn't mean to touch.
function openItemCtxMenu(i,ev){
  if(!SELECTED_ITEMS.has(i)){SELECTED_ITEMS.clear();SELECTED_ITEMS.add(i)}
  enterItemsSelectMode();
  renderItems();
  const n=SELECTED_ITEMS.size;
  showCtxMenu(ev,
    '<div class=cpitem style="cursor:pointer" onclick="closeCtxMenu();bulkCloneSelected()"><span class=cpname>⧉ Clone'+(n>1?' ('+n+')':'')+'</span></div>'+
    '<div class=cpitem style="cursor:pointer" onclick="closeCtxMenu();bulkCutSelected()"><span class=cpname>✂ Cut'+(n>1?' ('+n+')':'')+'</span></div>'+
    '<div class=cpitem style="cursor:pointer;color:var(--danger)" onclick="closeCtxMenu();bulkDeleteSelected()"><span class=cpname>🗑 Delete'+(n>1?' ('+n+')':'')+'</span></div>')}
function renderItems(){
  const cols=COLS[TYPE];
  const hasType=cols.some(c=>c[0]==='type');
  $('itemslist').innerHTML=items.map((it,i)=>itemCardHtml(it,i,cols,hasType)).join('');
  updateItemsBulkBar();updateItemsPasteBar()}
function toggleItemSelect(i,checked){
  if(checked)SELECTED_ITEMS.add(i);else SELECTED_ITEMS.delete(i);
  const card=document.querySelector('.itemcard[data-i="'+i+'"]');
  if(card)card.classList.toggle('selected',checked);
  updateItemsBulkBar()}
function clearItemSelection(){
  SELECTED_ITEMS.clear();
  document.querySelectorAll('.itemcard.selected').forEach(c=>c.classList.remove('selected'));
  document.querySelectorAll('.itemcardsel').forEach(cb=>cb.checked=false);
  updateItemsBulkBar()}
function updateItemsBulkBar(){
  const bar=$('itemsbulkbar');if(!bar)return;
  const n=SELECTED_ITEMS.size;
  bar.classList.toggle('hide',n===0);
  if(n)$('itemsbulkcount').textContent=n+' line item'+(n>1?'s':'')+' selected'}
function updateItemsPasteBar(){
  const btn=$('itempastebar');if(!btn)return;
  const n=ITEM_CLIPBOARD.length;
  btn.classList.toggle('hide',!n);
  if(n)btn.textContent='📋 Paste '+n+' line item'+(n>1?'s':'')+' — cut from '+(ITEM_CLIPBOARD_SRC||'another document')}
function selectedIndicesSorted(){return[...SELECTED_ITEMS].sort((a,b)=>a-b)}
function bulkDeleteSelected(){
  if(!SELECTED_ITEMS.size)return;
  const idx=selectedIndicesSorted();
  for(let k=idx.length-1;k>=0;k--)items.splice(idx[k],1);
  if(!items.length)items.push({});
  exitItemsSelectMode();
  renderItems();schedulePreview();
  toast(idx.length+' line item'+(idx.length>1?'s':'')+' deleted')}
function bulkCloneSelected(){
  if(!SELECTED_ITEMS.size)return;
  const idx=selectedIndicesSorted();
  // Walk in reverse so each insertion doesn't shift the indices still
  // queued to process — every clone lands directly after its original.
  for(let k=idx.length-1;k>=0;k--){
    const clone=JSON.parse(JSON.stringify(items[idx[k]]));
    delete clone.suggestedPhoto;
    items.splice(idx[k]+1,0,clone)}
  exitItemsSelectMode();
  renderItems();schedulePreview();
  toast(idx.length+' line item'+(idx.length>1?'s':'')+' cloned')}
let ITEM_CLIPBOARD_SRC='';
function bulkCutSelected(){
  if(!SELECTED_ITEMS.size)return;
  const idx=selectedIndicesSorted();
  ITEM_CLIPBOARD=idx.map(i=>{const c=JSON.parse(JSON.stringify(items[i]));delete c.suggestedPhoto;return c});
  ITEM_CLIPBOARD_SRC=LABEL[TYPE]||TYPE;
  for(let k=idx.length-1;k>=0;k--)items.splice(idx[k],1);
  if(!items.length)items.push({});
  exitItemsSelectMode();
  renderItems();schedulePreview();
  toast(idx.length+' line item'+(idx.length>1?'s':'')+' cut — use Paste to move '+(idx.length>1?'them':'it')+' here or into another document')}
function pasteClipboardItems(){
  if(!ITEM_CLIPBOARD.length)return;
  const clones=ITEM_CLIPBOARD.map(it=>JSON.parse(JSON.stringify(it)));
  if(items.length===1&&!items[0].description)items=[];  // drop the one always-present blank row rather than paste after it
  items.push(...clones);
  renderItems();schedulePreview();
  toast(clones.length+' line item'+(clones.length>1?'s':'')+' pasted')}
async function matchPhotoForItem(i){
  const desc=(items[i]?.description||'').trim();
  if(!desc||items[i].photo)return;   // never overwrite a photo that's already set
  const r=await fetch('/api/match-photo?desc='+encodeURIComponent(desc)).then(r=>r.json());
  if(r.photo&&items[i]&&!items[i].photo){items[i].suggestedPhoto=r.photo;renderItems()}}
async function matchDatasheetsForItem(i){
  const desc=(items[i]?.description||'').trim();
  if(!desc)return;
  const r=await fetch('/api/match-datasheets?desc='+encodeURIComponent(desc)).then(r=>r.json());
  if(items[i]){items[i].datasheets=r.datasheets||[];renderItems()}}
function openPhotoPicker(i,ev){
  const rect=ev.currentTarget.getBoundingClientRect();
  const sug=items[i].suggestedPhoto;
  $('filemenu').innerHTML='<div class=fmtitle>Suggested photo</div>'+
    '<div class=fmphotopreview><img src="'+sug+'"></div>'+
    '<div class=fmi onclick="acceptSuggestedPhoto('+i+')"><span class=ic>✓</span>Use this photo</div>'+
    '<div class=fmi onclick="closeFileMenu();pickPhoto('+i+')"><span class=ic>⤴</span>Upload my own instead</div>'+
    '<div class=fmi onclick="dismissSuggestedPhoto('+i+')"><span class=ic>✕</span>Dismiss suggestion</div>';
  const menu=$('filemenu');menu.style.display='block';
  const w=menu.offsetWidth||200,h=menu.offsetHeight||260;
  let x=rect.left,y=rect.bottom+4;
  if(x+w>window.innerWidth-8)x=window.innerWidth-w-8;
  if(y+h>window.innerHeight-8)y=window.innerHeight-h-8;
  menu.style.left=x+'px';menu.style.top=y+'px';
  setTimeout(()=>document.addEventListener('click',closeFileMenu,{once:true}),0)}
function acceptSuggestedPhoto(i){items[i].photo=items[i].suggestedPhoto;delete items[i].suggestedPhoto;closeFileMenu();renderItems();schedulePreview()}
function dismissSuggestedPhoto(i){delete items[i].suggestedPhoto;closeFileMenu();renderItems()}
function pickPhoto(i){
  const inp=document.createElement('input');
  inp.type='file';inp.accept='image/*';
  inp.onchange=()=>{
    const f=inp.files[0];if(!f)return;
    const reader=new FileReader();
    reader.onload=()=>{items[i].photo=reader.result;renderItems();schedulePreview()};
    reader.readAsDataURL(f)};
  inp.click()}
function removePhoto(i){delete items[i].photo;renderItems();schedulePreview()}

// ---------------------------------------------------------------- Sololuce Product Datasheet (CAT)
// Fixed badge vocabulary mirrors html_engine.BADGES (Python) — keep both in
// sync if a new badge type is ever added. needsValue=true badges print a
// per-product value inside the box (e.g. IP "20"); needsValue=false ones are
// pure icon/mark badges (CE, RoHS, etc). 'ip' is picked twice on real sheets
// that show two IP boxes (e.g. IP20 + IP44) — since this is a plain list,
// adding the same badge type more than once just works.
let CAT_BADGES=[];
let CAT_BADGE_LIBRARY=[];
let CAT_BADGE_SEARCH='';
let CAT_SPECS=[{label:'',values:['']}];
let CAT_FINISH=[];
let CAT_IMG={main:catImgDefault(),lifestyle:catImgDefault(),diagram:catImgDefault(),extra1:catImgDefault(),extra2:catImgDefault(),extra3:catImgDefault()};
// Every column in CAT_ORD_COLS is kept "lockstepped" to the same length —
// one shared variant count across the whole table — except Finish Options
// and Lumen, which are computed/derived and keep their own independent
// length (see CAT_ORD_SPECIAL_LABELS below). This replaced an earlier,
// deliberately independent-per-column-length model: the PDF already
// transposes column[c].values[i] into row i for every column (see
// html_engine.render_datasheet_pdf), so position i was always meant to read
// as one coherent product variant — lockstep just makes the input form
// guarantee that instead of leaving it to coincidence.
let CAT_ORD_COLS=[];
let CAT_COL_SEQ=0;
// Ordering Table field clipboard — separately named from the existing
// document-level CLIPBOARD/CTXROW/fmCopy/fmCut/fmPaste (used for whole
// files in All Docs, via /api/file-op) so the two never collide. Pure
// in-memory form state, no backend persistence of its own — a column is
// just {label, values, width, codeOn}, not a disk object to move
// atomically, so Cut removes its source immediately rather than deferring
// to paste-time the way the file clipboard's move does.
let CAT_ORD_COL_CLIPBOARD=null;
// Which variant the one-at-a-time Ordering Table view is currently
// showing (0-indexed) — pure UI state, never saved. Clamped defensively at
// the top of renderCatOrdTable() (variant count can shrink out from under
// it via remove/Import-from-PDF), reset to 0 at the same document-boundary
// points as CAT_ORD_COL_CLIPBOARD above.
let CAT_ORD_CURRENT_VARIANT=0;
// "Align Rows" toggle — per-document (saved on the draft/record like
// ordering_columns itself, NOT a global setting the way
// cat_ordering_default_widths is), so it round-trips through
// populateCatForm(). Recomputed fresh server-side every render (see
// html_engine.py's _ord_row_min_height_pt) rather than a frozen snapshot,
// so it stays correct if the data it's based on changes later.
let CAT_ORD_ALIGN_ROWS=false;
let CAT_SPEC_LABELS=[];
async function loadCatSpecLabels(){
  const r=await fetch('/api/cat-spec-labels').then(r=>r.json());
  if(r.labels)CAT_SPEC_LABELS=r.labels;
  renderCatSpecLabelsDatalist()}
function renderCatSpecLabelsDatalist(){
  const dl=$('catspeclabels');if(dl)dl.innerHTML=CAT_SPEC_LABELS.map(l=>'<option value="'+escHtml(l)+'">').join('')}
async function rememberCatSpecLabel(label){
  label=(label||'').trim();
  if(!label||CAT_SPEC_LABELS.includes(label))return;
  const r=await fetch('/api/cat-spec-labels-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({label})}).then(r=>r.json());
  if(r.labels){CAT_SPEC_LABELS=r.labels;renderCatSpecLabelsDatalist()}}

// Series/Category: a select of remembered labels + "Custom…" that swaps to a
// text input, same pattern as the Unit dropdown (onUnitChange/itemUnitFieldHtml).
let CAT_SERIES_LABELS=[];
async function loadCatSeriesLabels(){
  const r=await fetch('/api/cat-series').then(r=>r.json());
  if(r.labels)CAT_SERIES_LABELS=r.labels;}
function renderCatSeriesField(value){
  const v=value||'';
  const wrap=$('cat-series-wrap');if(!wrap)return;
  const opts='<option value="">— Select —</option>'+
    CAT_SERIES_LABELS.map(l=>'<option'+(l===v?' selected':'')+'>'+escHtml(l)+'</option>').join('')+
    (v&&!CAT_SERIES_LABELS.includes(v)?'<option selected>'+escHtml(v)+'</option>':'')+
    '<option value="__custom__">Custom…</option>';
  wrap.innerHTML='<select id=cat-series onchange="onCatSeriesChange(this)">'+opts+'</select>'}
function onCatSeriesChange(sel){
  if(sel.value!=='__custom__'){schedulePreview();return}
  const wrap=$('cat-series-wrap');
  wrap.innerHTML='<input id=cat-series type=text placeholder="Type a series/category, press Enter…">';
  const inp=$('cat-series');
  inp.focus();
  let done=false;
  const commit=async()=>{
    if(done)return; done=true;
    const v=inp.value.trim();
    if(!v){renderCatSeriesField('');return}
    const r=await fetch('/api/cat-series-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({label:v})}).then(r=>r.json());
    if(r.labels)CAT_SERIES_LABELS=r.labels;
    renderCatSeriesField(v);schedulePreview()};
  inp.addEventListener('keydown',e=>{
    if(e.key==='Enter'){e.preventDefault();inp.blur()}
    else if(e.key==='Escape'){done=true;renderCatSeriesField('')}});
  inp.addEventListener('blur',commit)}

// Family Tree: same dropdown+Custom pattern as Series/Category above, plus a
// live linked-products list scoped to whatever family name is currently
// selected. Unlike Series, a family carries no color and isn't referenced by
// the individual datasheet's own preview at all (it only matters once a Full
// Catalog Build clusters this product with its family behind a divider page)
// — so, deliberately unlike onCatSeriesChange, nothing here calls
// schedulePreview().
let CAT_FAMILY_LABELS=[], CAT_FAMILY_LINKED=[], CAT_FAMILY_SEARCH_TOKEN=0;
async function loadCatFamilyLabels(){
  const r=await fetch('/api/cat-family').then(r=>r.json());
  if(r.labels)CAT_FAMILY_LABELS=r.labels;}
function catFamilyCurrentValue(){
  const el=$('cat-family');if(!el)return '';
  return el.tagName==='SELECT'?(el.value==='__custom__'?'':el.value):el.value.trim()}
function renderCatFamilyField(value){
  const v=value||'';
  const wrap=$('cat-family-wrap');if(!wrap)return;
  const opts='<option value="">— Select —</option>'+
    CAT_FAMILY_LABELS.map(l=>'<option'+(l===v?' selected':'')+'>'+escHtml(l)+'</option>').join('')+
    (v&&!CAT_FAMILY_LABELS.includes(v)?'<option selected>'+escHtml(v)+'</option>':'')+
    '<option value="__custom__">Custom…</option>';
  wrap.innerHTML='<select id=cat-family onchange="onCatFamilyChange(this)">'+opts+'</select>';
  refreshCatFamilyLinked()}
function onCatFamilyChange(sel){
  if(sel.value!=='__custom__'){refreshCatFamilyLinked();return}
  const wrap=$('cat-family-wrap');
  wrap.innerHTML='<input id=cat-family type=text placeholder="Type a family name, press Enter…">';
  const inp=$('cat-family');
  inp.focus();
  let done=false;
  const commit=async()=>{
    if(done)return; done=true;
    const v=inp.value.trim();
    if(!v){renderCatFamilyField('');return}
    const r=await fetch('/api/cat-family-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({label:v})}).then(r=>r.json());
    if(r.labels)CAT_FAMILY_LABELS=r.labels;
    renderCatFamilyField(v)};
  inp.addEventListener('keydown',e=>{
    if(e.key==='Enter'){e.preventDefault();inp.blur()}
    else if(e.key==='Escape'){done=true;renderCatFamilyField('')}});
  inp.addEventListener('blur',commit)}
function onCatFamilyToggle(){
  const on=$('cat-family-enabled').checked;
  $('cat-family-details').classList.toggle('hide',!on);
  if(on)refreshCatFamilyLinked()}
// The chip list is never staged client-side — it's always a fresh read of
// "every OTHER product whose saved family equals this one" (search_products
// with family=), because linking/unlinking (addCatFamilyMember/
// removeCatFamilyMember) writes straight to that other product's own sidecar
// the moment it happens. Re-fetched on every family-name change too, so
// switching Custom text mid-edit never shows a stale/mismatched chip set.
async function refreshCatFamilyLinked(){
  const box=$('cat-family-chips');if(!box)return;
  const search=$('cat-family-search'),results=$('cat-family-search-results');
  if(search)search.value='';
  if(results)results.innerHTML='';
  const fam=catFamilyCurrentValue();
  if(!fam){box.innerHTML='';CAT_FAMILY_LINKED=[];return}
  const selfRel=(TYPE==='CAT'&&EDITING)?EDITING:'';
  const r=await fetch('/api/cat-products-search?family='+encodeURIComponent(fam)+'&exclude='+encodeURIComponent(selfRel)).then(r=>r.json());
  CAT_FAMILY_LINKED=r.results||[];
  renderCatFamilyChips()}
function renderCatFamilyChips(){
  const box=$('cat-family-chips');if(!box)return;
  box.innerHTML=CAT_FAMILY_LINKED.map(p=>
    '<span class=famchip>'+escHtml(p.product_name)+
    '<button type=button data-rel="'+escHtml(p.rel)+'" onclick="removeCatFamilyMember(this.dataset.rel)" title="Remove from family">×</button></span>'
  ).join('')||'<p class=muted style="font-size:11.5px;margin:0">No other products in this family yet — search below to add one.</p>'}
async function onCatFamilySearch(){
  const q=$('cat-family-search').value.trim();
  const results=$('cat-family-search-results');
  if(!q){results.innerHTML='';return}
  const fam=catFamilyCurrentValue();
  const selfRel=(TYPE==='CAT'&&EDITING)?EDITING:'';
  const token=++CAT_FAMILY_SEARCH_TOKEN;
  const r=await fetch('/api/cat-products-search?q='+encodeURIComponent(q)+'&exclude='+encodeURIComponent(selfRel)).then(r=>r.json());
  if(token!==CAT_FAMILY_SEARCH_TOKEN)return;
  const rows=(r.results||[]).filter(p=>p.family!==fam).slice(0,20);
  results.innerHTML=rows.length?rows.map(p=>
    '<div class=cpitem data-name="'+escHtml(p.product_name)+'" data-rel="'+escHtml(p.rel)+'" onclick="addCatFamilyMember(this.dataset.rel,this.dataset.name)">'+
      '<span class=cpname>'+escHtml(p.product_name)+'</span>'+
      (p.family?'<span class=muted style="font-size:10.5px;margin-left:auto;white-space:nowrap">currently: '+escHtml(p.family)+'</span>':'')+
    '</div>').join(''):'<p class="muted" style="font-size:11.5px;padding:4px;margin:0">No matching datasheets.</p>'}
async function addCatFamilyMember(rel,name){
  const fam=catFamilyCurrentValue();
  if(!fam)return;
  await fetch('/api/cat-family-link',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({family:fam,add:[rel]})});
  refreshCatFamilyLinked()}
async function removeCatFamilyMember(rel){
  const fam=catFamilyCurrentValue();
  await fetch('/api/cat-family-link',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({family:fam,remove:[rel]})});
  refreshCatFamilyLinked()}

// Product Name must be unique per Sololuce Datasheet (the same "company"
// slot used by build_filename) — reuses the existing /api/previous lookup
// (same one QTN/INV/DO already use for "last N docs for this company") rather
// than a new endpoint. Only checked for brand-new datasheets (not while
// EDITING an existing one, since a legit revision keeps the same name).
// Returns the existing file's own basename (truthy — usable directly as
// /api/generate's `replace` target) if a duplicate exists, '' otherwise —
// one fetch now answers both "does a duplicate exist" (every existing
// caller, which only ever used this as a boolean, so the type change is
// silently compatible) AND "which file would Replace & Generate actually
// Product names are always capitals in our system — force-uppercase as the
// user types (not just on blur/save) so what's on screen never disagrees
// with what gets stored/printed. Re-applying .value moves the caret to the
// end by default, so save/restore the selection to keep mid-string edits
// (e.g. fixing a typo in the middle of the name) from jumping the cursor.
function catUpperCaseInPlace(el){
  const s=el.selectionStart,e=el.selectionEnd,v=el.value,up=v.toUpperCase();
  if(v===up)return;
  el.value=up;
  try{el.setSelectionRange(s,e)}catch(e){}}
// replace" (generate()'s own resumed-draft path below), instead of two
// separate round-trips for the same lookup.
async function catProductNameExists(name){
  if(!name)return '';
  const r=await fetch('/api/previous?type=CAT&company='+encodeURIComponent(name)).then(r=>r.json());
  const p=r.previous&&r.previous[0];
  return p&&p.path?p.path.split(/[\\/]/).pop():''}
async function checkCatProductNameDuplicate(){
  const warn=$('cat-productname-warn');if(!warn)return;
  const name=$('cat-productname').value.trim();
  if(!name||EDITING){warn.classList.add('hide');return}
  const dup=await catProductNameExists(name);
  warn.classList.toggle('hide',!dup);
  if(dup)warn.textContent='A Sololuce Datasheet named "'+name+'" already exists — product names must be unique.'}
// Resumed-draft Generate hitting an already-generated product name (see
// generate()'s own comment) — asks before replacing, never silently.
// Promise-based rather than a plain callback so generate() can just
// `await` it inline like any other step. #catreplacemodal's buttons are
// wired here (not inline onclick=) so this can be called repeatedly
// without ever double-binding a stale handler from a previous call.
function askCatReplaceConfirm(name){
  return new Promise(resolve=>{
    $('catreplace-name').textContent=name;
    $('catreplacemodal').classList.remove('hide');
    $('catreplace-cancel-btn').onclick=()=>{$('catreplacemodal').classList.add('hide');resolve(false)};
    $('catreplace-confirm-btn').onclick=()=>{$('catreplacemodal').classList.add('hide');resolve(true)}})}

// Badges are real images the user picks from a library (see html_engine.py's
// comment on why — every PNG is fully self-contained, icon+caption+border
// already baked in, so there's no per-product value to type afterward the
// way the app's first hand-drawn-icon badge system needed). The library
// (39 real Sololuce icons + anything the user has added) loads once at
// bootstrap; picking a badge just toggles its key in/out of CAT_BADGES.
async function loadCatBadgeLibrary(){
  const r=await fetch('/api/cat-badges').then(r=>r.json());
  if(r.badges)CAT_BADGE_LIBRARY=r.badges;
  renderCatBadges()}
function catBadgeLabel(key){const b=CAT_BADGE_LIBRARY.find(x=>x.key===key);return b?b.label:key}
function isCatBadgeSelected(key){return CAT_BADGES.some(b=>b.key===key)}
function toggleCatBadge(key){
  if(isCatBadgeSelected(key))CAT_BADGES=CAT_BADGES.filter(b=>b.key!==key);
  else CAT_BADGES.push({key});
  renderCatBadges();schedulePreview()}
function setCatBadgeSearch(v){CAT_BADGE_SEARCH=v;renderCatBadges()}

// Badges that must always be pre-selected on a fresh datasheet, per the
// user's fixed "standard badges" list — kept separate from the
// Class/Indoor-Outdoor/IP badges below, which are auto-selected instead
// (driven by Product Type / the IP Rating spec) rather than always-on.
// RoHS swapped out for the "5 Years" warranty badge per a later explicit
// request (RoHS itself was retired from the library entirely, see
// _retired_badge_keys in load_cfg).
const CAT_STANDARD_BADGE_KEYS=['dispose-product-through-special-recycling','excellent-color-rendering-very-natural-light','emergency-lighting-module-for-power-failure','general-performance-or-summary-rating','sdcm-3','european-conformity-safety-marking','coverage-against-manufacturing-defects'];
// IP number -> library key, built from actually reading each badge PNG
// (several library labels were purely descriptive with no literal "IP65"
// text, so filename/label guessing alone isn't reliable here).
const CAT_IP_BADGE_KEYS={20:'protected-against-vertically-falling-water',44:'protected-against-splashing-water-ingress',54:'dust-protected-water-splash-resistant',65:'dust-tight-protected-against-jets',66:'dust-tight-strong-water-jets',67:'safe-for-temporary-water-immersion',68:'safe-for-continuous-water-immersion'};
const CAT_ALL_CLASS_IO_KEYS=['class-1-fixture','class-2-fixture','class-3-fixture','designed-for-use-inside-buildings-only','designed-for-use-outside-buildings-only'];
// Re-derives the Class 1/2/3 + Indoor/Outdoor badge from Product Type and
// the IP badge from the "IP Rating" spec value, replacing whichever one of
// each group was previously auto-picked — called on every Product Type
// change and every spec edit so the badge list always reflects the current
// form instead of accumulating stale auto-picks.
function recomputeCatAutoBadges(){
  const type=$('cat-producttype')?$('cat-producttype').value:'';
  const AUTO_CLASS_IO={Outdoor:{cls:'class-1-fixture',io:'designed-for-use-outside-buildings-only'},
    Indoor:{cls:'class-2-fixture',io:'designed-for-use-inside-buildings-only'},
    Striplight:{cls:'class-3-fixture',io:null}};
  CAT_BADGES=CAT_BADGES.filter(b=>!CAT_ALL_CLASS_IO_KEYS.includes(b.key));
  const pick=AUTO_CLASS_IO[type];
  if(pick){
    CAT_BADGES.push({key:pick.cls});
    if(pick.io)CAT_BADGES.push({key:pick.io})}
  const allIpKeys=Object.values(CAT_IP_BADGE_KEYS);
  CAT_BADGES=CAT_BADGES.filter(b=>!allIpKeys.includes(b.key));
  // IP Rating can hold several lines (e.g. a product sold in both IP20 and
  // IP65 variants) — match a badge per line, not just the first, deduped in
  // case two lines somehow round to the same rating.
  const ipSpec=CAT_SPECS.find(s=>(s.label||'').trim().toLowerCase()==='ip rating');
  if(ipSpec){
    const seen=new Set();
    (ipSpec.values||[]).forEach(v=>{
      const m=(v||'').match(/(\d{2})/);
      const bk=m&&CAT_IP_BADGE_KEYS[m[1]];
      if(bk&&!seen.has(bk)){seen.add(bk);CAT_BADGES.push({key:bk})}})}}
function onCatProductTypeChange(){recomputeCatAutoBadges();renderCatBadges()}
// Grouped by real-world meaning rather than one flat alphabetical list, so
// the picker reads like a spec sheet's own badge legend — matched by KEY
// (stable across the label renames done for IP/Indoor/Outdoor) rather than
// label text. Any library badge not in a known group (custom-added ones)
// falls into a trailing "Custom" group instead of disappearing.
const CAT_BADGE_CATEGORIES=[
  {name:'Safety & Compliance', keys:['european-conformity-safety-marking','restricts-hazardous-substances-in-electronics','limits-hazardous-substances-in-electronics','fire-rated','coverage-against-manufacturing-defects','dispose-product-through-special-recycling']},
  {name:'IP / Water & Dust Protection', keys:['protected-against-vertically-falling-water','protected-against-splashing-water-ingress','dust-protected-water-splash-resistant','dust-tight-protected-against-jets','dust-tight-strong-water-jets','safe-for-temporary-water-immersion','safe-for-continuous-water-immersion']},
  {name:'Installation Class & Location', keys:['class-1-fixture','class-2-fixture','class-3-fixture','designed-for-use-inside-buildings-only','designed-for-use-outside-buildings-only']},
  {name:'Color Quality', keys:['excellent-color-rendering-very-natural-light','sdcm-3','3-color-access']},
  {name:'Glare Control', keys:['no-direct-eye-exposure']},
  {name:'Controls & Connectivity', keys:['dali-dimmable-system','connected-devices-communicating-via-internet','power-and-data-via-ethernet','emergency-lighting-module-for-power-failure']},
  {name:'Adjustability', keys:['adjustable-beam-angle','adjustable-orientation','adjustable-rotation']},
  {name:'General', keys:['general-performance-or-summary-rating','office-standard','client-supplied-materials-supply-and-installation']},
];
function catBadgeTileHtml(b){
  const on=isCatBadgeSelected(b.key);
  return '<div onclick="toggleCatBadge(\''+b.key+'\')" title="'+escHtml(b.label)+'" style="width:64px;cursor:pointer;text-align:center">'+
    '<div style="position:relative;width:60px;height:60px;border:2px solid '+(on?'#1f6feb':'var(--border)')+';border-radius:6px;overflow:hidden;background:#fff">'+
      '<img src="/static/cat_badges/'+b.filename+'" style="width:100%;height:100%;object-fit:contain" loading=lazy>'+
      (on?'<div style="position:absolute;top:2px;right:2px;width:14px;height:14px;background:#1f6feb;border-radius:50%;color:#fff;font-size:10px;line-height:14px;font-weight:700">✓</div>':'')+
    '</div>'+
    '<div style="font-size:9px;margin-top:2px;line-height:1.2;color:var(--muted);white-space:normal;word-break:break-word">'+escHtml(b.label)+'</div>'+
  '</div>'}
// The full picker (search box + every category's tiles + custom-badge-add
// row) used to render directly inline in the sidebar — for a fitting with
// a dozen+ badges already applied plus every unselected one still shown
// below, that's a LOT of scroll before the next card (Technical
// Specifications) even comes into view. Now it lives in a modal
// (#catbadgesmodal, opened via the compact summary row below), same
// bigger-modal pattern as Photo Adjust (.clientmodal/.clientmodalbox) —
// the sidebar itself only ever shows a one-line "Badges (N) — Edit"
// button plus a small read-only preview strip of whatever's already
// applied, so the card's height stays constant regardless of how many
// badges a product carries.
function renderCatBadgesSummary(){
  const applied=CAT_BADGES.map(b=>CAT_BADGE_LIBRARY.find(x=>x.key===b.key)).filter(Boolean);
  const preview=applied.slice(0,8).map(b=>
    '<img src="/static/cat_badges/'+b.filename+'" title="'+escHtml(b.label)+'" style="width:26px;height:26px;object-fit:contain;border:1px solid var(--border);border-radius:4px;background:#fff;flex-shrink:0">').join('');
  const overflow=applied.length>8?'<span class=muted style="font-size:10.5px;flex-shrink:0">+'+(applied.length-8)+' more</span>':'';
  $('cat-badges-rows').innerHTML=
    '<button type=button class=btn style="width:100%" onclick="openCatBadgesModal()">'+
      'Badges ('+applied.length+' selected) — Edit</button>'+
    (applied.length?'<div style="display:flex;flex-wrap:wrap;gap:5px;margin-top:8px;align-items:center">'+preview+overflow+'</div>':'')}
function openCatBadgesModal(){$('catbadgesmodal').classList.remove('hide');renderCatBadgesPicker()}
function closeCatBadgesModal(){$('catbadgesmodal').classList.add('hide')}
function renderCatBadgesPicker(){
  const q=CAT_BADGE_SEARCH.trim().toLowerCase();
  const knownKeys=CAT_BADGE_CATEGORIES.flatMap(c=>c.keys);
  const groups=CAT_BADGE_CATEGORIES.map(c=>({name:c.name,items:c.keys.map(k=>CAT_BADGE_LIBRARY.find(b=>b.key===k)).filter(Boolean)}));
  const custom=CAT_BADGE_LIBRARY.filter(b=>!knownKeys.includes(b.key));
  if(custom.length)groups.push({name:'Custom',items:custom});
  // Every selected badge is pulled out of its real category into a leading
  // "Applied" group instead of just being sorted first within its own
  // category — so the picker always opens showing exactly what's already
  // on the sheet, in the order they were picked, with the category layout
  // below unchanged apart from those badges being missing from their own
  // section (they're already shown, right above).
  const appliedKeys=new Set(CAT_BADGES.map(b=>b.key));
  const allGroups=[{name:'Applied',items:CAT_BADGES.map(b=>CAT_BADGE_LIBRARY.find(x=>x.key===b.key)).filter(Boolean)}]
    .concat(groups.map(g=>({name:g.name,items:g.items.filter(b=>!appliedKeys.has(b.key))})));
  const sections=allGroups.map(g=>{
    const items=g.items.filter(b=>!q||b.label.toLowerCase().includes(q));
    if(!items.length)return '';
    return '<div style="width:100%;font-size:10px;font-weight:700;letter-spacing:.04em;text-transform:uppercase;color:#767b82;margin:10px 0 6px">'+escHtml(g.name)+'</div>'+
      items.map(catBadgeTileHtml).join('')}).join('');
  $('cat-badges-modal-rows').innerHTML=
    '<input placeholder="Search badges…" value="'+escHtml(CAT_BADGE_SEARCH)+'" oninput="setCatBadgeSearch(this.value)" style="margin-bottom:8px;width:100%;box-sizing:border-box">'+
    '<div style="display:flex;flex-wrap:wrap;gap:8px;padding:2px">'+
      (sections||'<p class="muted" style="font-size:12px;margin:0">No badges match your search.</p>')+
    '</div>'+
    '<div style="display:flex;gap:6px;align-items:center;margin-top:10px">'+
      '<input id=cat-badge-custom-label style="flex:1" placeholder="New badge name, e.g. IP66">'+
      '<button type=button class=btn onclick=pickCatCustomBadgeImage()>+ Add Custom Badge</button>'+
    '</div>'}
// Both the collapsed summary AND the (possibly-closed) modal picker
// refresh together on every data change — cheap enough (same "always
// fully re-render" convention every other list in this app already uses)
// and means the modal is never stale the next time it's opened.
function renderCatBadges(){renderCatBadgesSummary();renderCatBadgesPicker()}
function pickCatCustomBadgeImage(){
  const label=$('cat-badge-custom-label').value.trim();
  if(!label){alert('Enter a name for this badge first.');return}
  const inp=document.createElement('input');
  inp.type='file';inp.accept='image/*';
  inp.onchange=()=>{
    const f=inp.files[0];if(!f)return;
    const reader=new FileReader();
    reader.onload=async()=>{
      const r=await fetch('/api/cat-badges-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({label,image:reader.result})}).then(r=>r.json());
      if(r.error){alert(r.error);return}
      CAT_BADGE_LIBRARY=r.badges;
      const added=CAT_BADGE_LIBRARY.find(b=>b.label===label)||CAT_BADGE_LIBRARY[CAT_BADGE_LIBRARY.length-1];
      if(added&&!isCatBadgeSelected(added.key))CAT_BADGES.push({key:added.key});
      $('cat-badge-custom-label').value='';
      renderCatBadges();schedulePreview()};
    reader.readAsDataURL(f)};
  inp.click()}

// Per-label smart value inputs: most Technical Specification labels are just
// free text, but a handful have a known unit/format the user shouldn't have
// to retype every time — matched by label with spaces/case ignored so
// "Life Span" and "Lifespan" are the same key regardless of exact spelling.
function normSpecLabel(label){return (label||'').trim().toLowerCase().replace(/\s+/g,'')}
const CAT_SPEC_SUFFIX={power:'W',lifespan:'Hrs',luminareefficacy:'lm/W'};
function appendCatSpecSuffix(v,suffix){
  v=(v||'').trim();
  if(!v)return '';
  const re=new RegExp('\\s*'+suffix.replace(/[.*+?^${}()|[\]\\]/g,'\\$&')+'$','i');
  // Non-breaking space between the number and its unit — a regular space
  // there lets the number and unit land on separate lines in the Ordering
  // Table's narrow columns (e.g. "100*50" then an orphaned "mm" below it);
  //   keeps them glued together while still allowing the cell to wrap
  // normally anywhere else if the content is genuinely too long to fit.
  return v.replace(re,'').trim()+' '+suffix}
function stripAmbientTemp(v){const m=(v||'').match(/^±\s*(.*?)\s*°\s*C\s*$/);return m?m[1]:(v||'')}
function formatAmbientTemp(core){core=(core||'').trim();return core?('± '+core+' °C'):''}
function stripIpRatingValue(v){const m=(v||'').match(/^IP\s*(.*)$/i);return m?m[1]:(v||'')}
function formatIpRatingValue(core){core=(core||'').trim();return core?('IP'+core):''}
// Every Technical Spec label remembers whatever values were typed for it —
// keyed by normalized label (cat_spec_values on the backend is a dict, not
// one flat list) since "Body Material" and "Driver" need separate pools, not
// a shared one. A label with nothing remembered yet skips straight to the
// raw input (no dropdown with nothing in it); once at least one value has
// been saved for that label, it becomes a select+Custom… pick, same as
// every Ordering Table column. Light Source/Power Factor used to be a
// hardcoded, non-persistent dropdown (the old CAT_SPEC_DROPDOWNS) — now just
// two more labels in this same system, seeded server-side in load_cfg so
// their old curated options (COB/SMD, >0.9) still show, and any custom
// value typed for them from now on is remembered too.
let CAT_SPEC_VALUES={};
async function loadCatSpecValues(){
  const r=await fetch('/api/cat-spec-values').then(r=>r.json());
  if(r.values)CAT_SPEC_VALUES=r.values}
async function rememberCatSpecValue(label,value){
  value=(value||'').trim();
  const key=normSpecLabel(label);
  if(!key||!value)return;
  const have=CAT_SPEC_VALUES[key]||[];
  if(have.includes(value))return;
  const r=await fetch('/api/cat-spec-values-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({label,value})}).then(r=>r.json());
  if(r.values)CAT_SPEC_VALUES[r.key||key]=r.values}
// Every spec row holds an ARRAY of values (CAT_SPECS[i].values), not one
// string — a product with 7 wattage variants needs Power to show 7 lines,
// not "3, 5, 7, 9, 12, 15, 18 W" crammed onto one (which is exactly the bug
// this fixed: the Ordering Table's Power column already supported multiple
// values, but Technical Specifications' Power row didn't, so syncing them
// either lost data or produced an unreadable comma-joined mess). Rendered
// with \n between lines — templates_html/sololuce_datasheet.html's spec
// value already has white-space:pre-line, so no template change was needed,
// only the builder-side data model and UI.
let CAT_SPEC_VALUE_CUSTOM=new Set();
function catSpecValueCellHtml(i,vi,s){
  const key=normSpecLabel(s.label);
  if(key==='power')return catSpecPowerCellHtml(i,vi,s);
  const known=CAT_SPEC_VALUES[key]||[];
  const v=s.values[vi]||'';
  const ckey=i+':'+vi;
  const customMode=!known.length||CAT_SPEC_VALUE_CUSTOM.has(ckey)||(v&&!known.includes(v));
  if(!customMode){
    const opts='<option value="">— Select —</option>'+
      known.map(o=>'<option value="'+escHtml(o)+'"'+(o===v?' selected':'')+'>'+escHtml(o)+'</option>').join('')+
      '<option value="__custom__">Custom…</option>';
    return '<select style="flex:1" onchange="onCatSpecValueChange('+i+','+vi+',this)">'+opts+'</select>'}
  return catSpecCustomInputHtml(i,vi,s,key)}
// Power collapses to ONE printed line, unlike every other multi-value spec
// (which stays one value per line via the template's white-space:pre-line)
// — a real wattage lineup reads better as "5W, 8W, 11W" than as a tall
// stack, and once there are more than a handful it's more useful as a
// plain range ("3-25W") than as a wall of individual values. Only affects
// the printed/collected value (collectCatData) — the input side still
// edits each wattage as its own line/box, unchanged.
function formatCatSpecPowerValue(values){
  const vals=(values||[]).map(v=>(v||'').trim()).filter(Boolean);
  if(!vals.length)return '';
  if(vals.length<=3)return vals.join(', ');
  const nums=vals.map(v=>{const m=v.match(/[\d.]+/);return m?parseFloat(m[0]):null});
  if(nums.some(n=>n==null))return vals.join(', ');
  const unit=(vals[0].match(/[^\d.]+$/)||[''])[0].trim()||'W';
  return Math.min(...nums)+'-'+Math.max(...nums)+unit}
// Power is the one Technical Spec label that shares its preset list AND its
// current values with the Ordering Table's Power column (see
// syncOrderingPowerFromSpec/syncSpecPowerFromOrdering) instead of having its
// own independent cat_spec_values bucket — same real-world spec, the two
// full arrays are kept mirrored (same values, same order) both ways rather
// than drifting apart.
let CAT_SPEC_POWER_CUSTOM=new Set();
function catSpecPowerCellHtml(i,vi,s){
  const v=s.values[vi]||'';
  const normV=catOrdPowerNorm(v);
  const options=catPowerOptionsOrdered();
  const ckey=i+':'+vi;
  const customMode=CAT_SPEC_POWER_CUSTOM.has(ckey)||(v&&!options.some(o=>catOrdPowerNorm(o)===normV));
  if(customMode)
    return '<textarea rows=1 style="flex:1;resize:vertical;font:inherit;padding:8px 10px;box-sizing:border-box" placeholder="e.g. 15W" oninput="CAT_SPECS['+i+'].values['+vi+']=this.value;schedulePreview()" onblur="onCatSpecPowerCustomBlur('+i+','+vi+',this)">'+escHtml(v)+'</textarea>';
  const opts='<option value="">— Select —</option>'+
    options.map(o=>'<option value="'+escHtml(o)+'"'+(catOrdPowerNorm(o)===normV?' selected':'')+'>'+escHtml(o)+'</option>').join('')+
    '<option value="__custom__">Custom…</option>';
  return '<select style="flex:1" onchange="onCatSpecPowerChange('+i+','+vi+',this)">'+opts+'</select>'}
function onCatSpecPowerChange(i,vi,sel){
  const ckey=i+':'+vi;
  if(sel.value==='__custom__'){
    CAT_SPEC_POWER_CUSTOM.add(ckey);
    CAT_SPECS[i].values[vi]='';
    renderCatSpecs();
    return}
  CAT_SPEC_POWER_CUSTOM.delete(ckey);
  CAT_SPECS[i].values[vi]=sel.value;
  syncOrderingPowerFromSpec(CAT_SPECS[i].values);
  schedulePreview()}
async function onCatSpecPowerCustomBlur(i,vi,el){
  const formatted=appendCatSpecSuffix(el.value,'W');
  CAT_SPECS[i].values[vi]=formatted;
  if(!formatted){CAT_SPEC_POWER_CUSTOM.delete(i+':'+vi);renderCatSpecs();syncOrderingPowerFromSpec(CAT_SPECS[i].values);schedulePreview();return}
  const ckey=i+':'+vi;
  const finish=()=>{
    CAT_SPEC_POWER_CUSTOM.delete(ckey);
    syncOrderingPowerFromSpec(CAT_SPECS[i].values);
    schedulePreview()};
  rememberCatPowerOption(formatted).then(finish)}
function onCatSpecValueChange(i,vi,sel){
  const ckey=i+':'+vi;
  if(sel.value==='__custom__'){
    CAT_SPEC_VALUE_CUSTOM.add(ckey);
    CAT_SPECS[i].values[vi]='';
    renderCatSpecs();
    return}
  CAT_SPEC_VALUE_CUSTOM.delete(ckey);
  CAT_SPECS[i].values[vi]=sel.value;
  renderCatOrdTable();recomputeCatAutoBadges();renderCatBadges();renderCatSpecs();schedulePreview()}
// Same "remember what was typed" pattern as every Ordering Table column: a
// custom value the user actually typed gets saved as a new preset for this
// label, then the row drops back into select mode showing what it just
// learned. Clearing a field to empty ALSO drops back to select mode (empty
// "— Select —") rather than staying stuck in custom/typing mode forever —
// otherwise clicking Custom… was a one-way door with no way back to the
// dropdown short of reloading the page.
async function onCatSpecCustomBlur(i,vi,el){
  const label=CAT_SPECS[i].label;
  const key=normSpecLabel(label);
  let formatted;
  if(key==='ambienttemperature')formatted=formatAmbientTemp(el.value);
  else if(key==='iprating')formatted=formatIpRatingValue(el.value);
  else{
    const suffix=CAT_SPEC_SUFFIX[key];
    formatted=suffix?appendCatSpecSuffix(el.value,suffix):el.value.trim()}
  CAT_SPECS[i].values[vi]=formatted;
  if(!formatted){
    CAT_SPEC_VALUE_CUSTOM.delete(i+':'+vi);
    renderCatSpecs();renderCatOrdTable();recomputeCatAutoBadges();renderCatBadges();schedulePreview();
    return}
  const ckey=i+':'+vi;
  const finish=()=>{
    CAT_SPEC_VALUE_CUSTOM.delete(ckey);
    renderCatSpecs();renderCatOrdTable();recomputeCatAutoBadges();renderCatBadges();schedulePreview()};
  rememberCatSpecValue(label,formatted).then(finish)}
// The raw "typing mode" input for a spec value line — same bespoke ± / IP /
// unit-suffix formatting each label had before, just addressed by (row,
// line) instead of just row now that a row can hold several lines.
function catSpecCustomInputHtml(i,vi,s,key){
  const v=s.values[vi]||'';
  if(key==='ambienttemperature')
    return '<div style="display:flex;align-items:center;gap:5px;flex:1">'+
      '<span style="font-size:11px;color:#767b82">±</span>'+
      '<input style="flex:1;min-width:0" placeholder="e.g. 5" value="'+escHtml(stripAmbientTemp(v))+'" oninput="CAT_SPECS['+i+'].values['+vi+']=formatAmbientTemp(this.value);schedulePreview()" onblur="onCatSpecCustomBlur('+i+','+vi+',this)">'+
      '<span style="font-size:11px;color:#767b82">°C</span>'+
    '</div>';
  if(key==='iprating')
    return '<div style="display:flex;align-items:center;gap:5px;flex:1">'+
      '<span style="font-size:11px;color:#767b82">IP</span>'+
      '<input style="flex:1;min-width:0" placeholder="e.g. 65" value="'+escHtml(stripIpRatingValue(v))+'" oninput="CAT_SPECS['+i+'].values['+vi+']=formatIpRatingValue(this.value);recomputeCatAutoBadges();renderCatBadges();schedulePreview()" onblur="onCatSpecCustomBlur('+i+','+vi+',this)">'+
    '</div>';
  const suffix=CAT_SPEC_SUFFIX[key];
  if(suffix)
    return '<textarea rows=1 style="flex:1;resize:vertical;font:inherit;padding:8px 10px;box-sizing:border-box" placeholder="Value, e.g. 3 - 18" oninput="CAT_SPECS['+i+'].values['+vi+']=this.value;renderCatOrdTable();schedulePreview()" onblur="onCatSpecCustomBlur('+i+','+vi+',this)">'+escHtml(v)+'</textarea>';
  return '<textarea rows=1 style="flex:1;resize:vertical;font:inherit;padding:8px 10px;box-sizing:border-box" placeholder="Value" oninput="CAT_SPECS['+i+'].values['+vi+']=this.value;renderCatOrdTable();recomputeCatAutoBadges();renderCatBadges();schedulePreview()" onblur="onCatSpecCustomBlur('+i+','+vi+',this)">'+escHtml(v)+'</textarea>'}
// Stacks one cell per value vertically (move up/down + remove-line, same
// affordances catOrdValueRowWrap already gives the Ordering Table) plus an
// "+ Add Value" pill below — so any spec that legitimately has more than
// one real-world value (Power, IP Rating, or anything else) gets its own
// line per value instead of being crammed onto one.
function renderCatSpecValueStack(i,s){
  const total=s.values.length;
  const lines=s.values.map((v,vi)=>{
    const cell=catSpecValueCellHtml(i,vi,s);
    const moves=total>1?(
      '<div style="display:flex;flex-direction:column;">'+
        '<button type=button class=ordmovebtn'+(vi===0?' disabled':'')+' onclick="moveCatSpecValue('+i+','+vi+',-1)" title="Move up">▲</button>'+
        '<button type=button class=ordmovebtn'+(vi===total-1?' disabled':'')+' onclick="moveCatSpecValue('+i+','+vi+',1)" title="Move down">▼</button>'+
      '</div>'):'';
    const rm=total>1?'<button type=button class=rm onclick="removeCatSpecValue('+i+','+vi+')" title="Remove this value">×</button>':'';
    const handle=total>1?'<span class=draghandle style="align-self:center" title="Drag to reorder">⠿</span>':'';
    return '<div class=dragrow draggable="'+(total>1)+'" ondragstart="dragRowStart(event,{i:'+i+',vi:'+vi+'})" ondragover="dragRowOver(event)" ondragleave="dragRowLeave(event)" ondrop="specValueDrop(event,'+i+','+vi+')" ondragend="dragRowEnd(event)" style="display:flex;gap:3px;align-items:flex-start;margin-bottom:3px">'+handle+moves+cell+rm+'</div>'}).join('');
  return '<div style="flex:1;min-width:0">'+lines+
    '<button type=button class="dspill" onclick="addCatSpecValue('+i+')">+ Add Value</button>'+
  '</div>'}
function addCatSpecValue(i){
  CAT_SPECS[i].values.push('');
  if(normSpecLabel(CAT_SPECS[i].label)==='power')syncOrderingPowerFromSpec(CAT_SPECS[i].values);
  renderCatSpecs();renderCatOrdTable();recomputeCatAutoBadges();renderCatBadges();schedulePreview()}
function removeCatSpecValue(i,vi){
  const s=CAT_SPECS[i];
  s.values.splice(vi,1);
  if(!s.values.length)s.values.push('');
  if(normSpecLabel(s.label)==='power')syncOrderingPowerFromSpec(s.values);
  renderCatSpecs();renderCatOrdTable();recomputeCatAutoBadges();renderCatBadges();schedulePreview()}
function moveCatSpecValue(i,vi,dir){
  const s=CAT_SPECS[i];
  const nvi=vi+dir;
  if(nvi<0||nvi>=s.values.length)return;
  const tmp=s.values[vi];s.values[vi]=s.values[nvi];s.values[nvi]=tmp;
  if(normSpecLabel(s.label)==='power')syncOrderingPowerFromSpec(s.values);
  renderCatSpecs();schedulePreview()}
function specValueDrop(e,i,targetVi){
  e.preventDefault();
  e.currentTarget.classList.remove('dragover-top','dragover-bottom');
  // Purely local array state (no server round-trip), so this splices
  // directly to the drop position in one step rather than replaying single
  // steps like the Full Catalog Builder lists do — DRAG_KEY here is
  // {i,vi} (see renderCatSpecValueStack), scoped to one row: dragging a
  // value into a DIFFERENT spec row's list wouldn't mean anything.
  if(!DRAG_KEY||typeof DRAG_KEY!=='object'||DRAG_KEY.i!==i||DRAG_KEY.vi===targetVi)return;
  const s=CAT_SPECS[i],v=s.values;
  const item=v.splice(DRAG_KEY.vi,1)[0];
  v.splice(targetVi,0,item);
  if(normSpecLabel(s.label)==='power')syncOrderingPowerFromSpec(v);
  renderCatSpecs();schedulePreview()}
function renderCatSpecs(){
  $('cat-specs-rows').innerHTML=CAT_SPECS.map((s,i)=>
    '<div style="display:flex;gap:6px;align-items:flex-start;margin-bottom:6px">'+
      '<input style="flex:1" list=catspeclabels placeholder="Label, e.g. Power" value="'+escHtml(s.label||'')+'" oninput="CAT_SPECS['+i+'].label=this.value;recomputeCatAutoBadges();renderCatBadges();schedulePreview()" onblur="rememberCatSpecLabel(this.value);renderCatSpecs();schedulePreview()">'+
      renderCatSpecValueStack(i,s)+
      (CAT_SPECS.length>1?'<button type=button class=rm onclick="removeCatSpec('+i+')" title="Remove row">×</button>':'')+
    '</div>').join('')}
function addCatSpec(){CAT_SPECS.push({label:'',values:['']});renderCatSpecs();schedulePreview()}
function removeCatSpec(i){CAT_SPECS.splice(i,1);if(!CAT_SPECS.length)CAT_SPECS.push({label:'',values:['']});renderCatSpecs();schedulePreview()}

let CAT_FINISH_PRESETS=[];
async function loadCatFinishColors(){
  const r=await fetch('/api/cat-finish-colors').then(r=>r.json());
  if(r.colors)CAT_FINISH_PRESETS=r.colors;
  renderCatFinish()}
// Common RAL Classic shades, so a finish name like "RAL 7016" auto-fills the
// right swatch color instead of the user eyeballing it on a generic color
// wheel. This is a curated list of the RAL codes actually used for fixture
// finishes (blacks/whites/greys/metallics + a few common accent colors),
// not the full ~213-color official chart — hex values are the commonly
// published screen approximations, not certified print-accurate swatches.
const CAT_RAL_COLORS=[
  {code:'1013',name:'Oyster White',hex:'#e3d9c6'},{code:'1015',name:'Light Ivory',hex:'#e6d2b5'},
  {code:'1021',name:'Colza Yellow',hex:'#f6b600'},{code:'1023',name:'Traffic Yellow',hex:'#fad201'},
  {code:'2004',name:'Pure Orange',hex:'#e75b12'},{code:'3000',name:'Flame Red',hex:'#af2b1e'},
  {code:'3020',name:'Traffic Red',hex:'#cc0605'},{code:'4005',name:'Blue Lilac',hex:'#6c4675'},
  {code:'5002',name:'Ultramarine Blue',hex:'#20214f'},{code:'5010',name:'Gentian Blue',hex:'#0e294b'},
  {code:'5015',name:'Sky Blue',hex:'#2271b3'},{code:'6005',name:'Moss Green',hex:'#0f4336'},
  {code:'6009',name:'Fir Green',hex:'#27352a'},{code:'6018',name:'Yellow Green',hex:'#57a639'},
  {code:'7001',name:'Silver Grey',hex:'#8a9597'},{code:'7011',name:'Iron Grey',hex:'#434b4d'},
  {code:'7015',name:'Slate Grey',hex:'#434750'},{code:'7016',name:'Anthracite Grey',hex:'#293133'},
  {code:'7021',name:'Black Grey',hex:'#23282b'},{code:'7024',name:'Graphite Grey',hex:'#474a51'},
  {code:'7035',name:'Light Grey',hex:'#d7d7d7'},{code:'7038',name:'Agate Grey',hex:'#b5b8b1'},
  {code:'7040',name:'Window Grey',hex:'#9da3a6'},{code:'7042',name:'Traffic Grey A',hex:'#8f9695'},
  {code:'7043',name:'Traffic Grey B',hex:'#4e5451'},{code:'8014',name:'Sepia Brown',hex:'#43302e'},
  {code:'8017',name:'Chocolate Brown',hex:'#45322e'},{code:'8019',name:'Grey Brown',hex:'#403a3a'},
  {code:'8022',name:'Black Brown',hex:'#212121'},{code:'9001',name:'Cream',hex:'#fdf4e3'},
  {code:'9002',name:'Grey White',hex:'#e7ebda'},{code:'9003',name:'Signal White',hex:'#f4f4f4'},
  {code:'9004',name:'Signal Black',hex:'#282828'},{code:'9005',name:'Jet Black',hex:'#0a0a0a'},
  {code:'9006',name:'White Aluminium',hex:'#a5a8a5'},{code:'9007',name:'Grey Aluminium',hex:'#8f8f8c'},
  {code:'9010',name:'Pure White',hex:'#eff0ea'},{code:'9016',name:'Traffic White',hex:'#f6f6f6'},
  {code:'9017',name:'Traffic Black',hex:'#2a2a2a'},
];
function ralDatalistLabel(r){return 'RAL '+r.code+' '+r.name}
// Matches a typed name against a RAL code (accepts "RAL 7016", "ral7016",
// or bare "7016") and auto-fills the swatch to that shade — the user can
// still override the color picker manually afterward if they want a
// different shade under the same name.
// Only Black/White/Grey show directly on the DSB form — every other saved
// color (any other preset, plus anything custom or RAL the user adds) lives
// in the "More Colors" popup instead, so the main form doesn't accumulate a
// permanently-growing wall of checkboxes as the library grows over time.
const CAT_FINISH_QUICK=['Black','White','Grey'];
function isFinishSelected(label){return CAT_FINISH.some(f=>f.label===label)}
function toggleCatFinish(label){
  const c=CAT_FINISH_PRESETS.find(p=>p.label===label);if(!c)return;
  if(isFinishSelected(label))CAT_FINISH=CAT_FINISH.filter(f=>f.label!==label);
  else CAT_FINISH.push({label:c.label,hex:c.hex});
  renderCatFinish();renderCatFinishMenu();renderCatOrdTable();schedulePreview()}
function finishChipHtml(c){
  const on=isFinishSelected(c.label);
  return '<div onclick="toggleCatFinish(\''+c.label.replace(/'/g,"\\'")+'\')" style="display:flex;align-items:center;gap:6px;padding:5px 9px;border:1px solid '+(on?'#1f6feb':'#d6dbe1')+';border-radius:6px;cursor:pointer;font-size:12.5px;background:'+(on?'#eef4ff':'#fff')+'">'+
    '<span style="width:15px;height:15px;border:1px solid #00000030;border-radius:3px;background:'+c.hex+';display:inline-block;flex-shrink:0"></span>'+
    escHtml(c.label)+(on?' ✓':'')+
  '</div>'}
function renderCatFinish(){
  const quick=CAT_FINISH_QUICK.map(label=>CAT_FINISH_PRESETS.find(c=>c.label===label)).filter(Boolean);
  const extraCount=CAT_FINISH_PRESETS.length-quick.length;
  $('cat-finish-rows').innerHTML=
    '<div style="display:flex;flex-wrap:wrap;gap:6px;align-items:center">'+
      quick.map(finishChipHtml).join('')+
      '<button type=button class=btn style="font-size:12.5px;padding:6px 11px" onclick="openCatFinishMenu(event)">More Colors'+(extraCount>0?' ('+extraCount+')':'')+' ▾</button>'+
    '</div>'}
// "More Colors" reuses the app's shared #filemenu popup (same one used by
// Manage Units etc.) — a list of every saved color that ISN'T one of the
// 3 quick ones, each togglable and removable, plus two distinct entries to
// add a brand-new color: one for a plain custom swatch, one RAL-code-driven
// (auto-fills the swatch from CAT_RAL_COLORS). Both add-flows share the
// same tiny inline form, swapped in over the list view.
let CAT_FINISH_MENU_MODE='list';
function openCatFinishMenu(ev){
  ev.stopPropagation();
  CAT_FINISH_MENU_MODE='list';
  const menu=$('filemenu');
  renderCatFinishMenu();
  menu.style.display='block';
  const rect=ev.currentTarget.getBoundingClientRect();
  const w=menu.offsetWidth||240,h=menu.offsetHeight||240;
  let x=rect.left,y=rect.bottom+4;
  if(x+w>window.innerWidth-8)x=window.innerWidth-w-8;
  if(y+h>window.innerHeight-8)y=window.innerHeight-h-8;
  menu.style.left=x+'px';menu.style.top=y+'px';
  setTimeout(()=>document.addEventListener('click',closeFileMenu,{once:true}),0)}
function renderCatFinishMenu(){
  const menu=$('filemenu');
  if(CAT_FINISH_MENU_MODE==='custom'){
    menu.innerHTML='<div class=fmtitle>Add Custom Color</div>'+
      '<div style="display:flex;gap:6px;align-items:center;padding:6px 10px">'+
        '<input type=color id=cat-finish-new-hex value="#ffffff" style="width:32px;height:32px;padding:0;border:1px solid #d6dbe1;border-radius:5px;cursor:pointer" onclick="event.stopPropagation()">'+
        '<input id=cat-finish-new-label placeholder="Color name, e.g. Bronze" style="flex:1" onclick="event.stopPropagation()">'+
      '</div>'+
      '<div style="display:flex;gap:6px;padding:6px 10px">'+
        '<button type=button class=btn style="flex:1" onclick="event.stopPropagation();catFinishMenuBack()">← Back</button>'+
        '<button type=button class="btn dark" style="flex:1" onclick="event.stopPropagation();confirmAddCatFinish()">+ Add</button>'+
      '</div>';
    setTimeout(()=>{const inp=$('cat-finish-new-label');if(inp)inp.focus()},0);
    return}
  if(CAT_FINISH_MENU_MODE==='ral'){
    menu.innerHTML='<div class=fmtitle>Add RAL Color</div>'+
      '<div style="padding:0 10px 6px">'+
        '<input id=cat-finish-ral-search placeholder="Search RAL code or name…" style="width:100%;box-sizing:border-box" oninput="renderCatFinishRalList()" onclick="event.stopPropagation()">'+
      '</div>'+
      '<div id=cat-finish-ral-list style="max-height:220px;overflow:auto"></div>'+
      '<div style="padding:6px 10px">'+
        '<button type=button class=btn style="width:100%" onclick="event.stopPropagation();catFinishMenuBack()">← Back</button>'+
      '</div>';
    renderCatFinishRalList();
    setTimeout(()=>{const inp=$('cat-finish-ral-search');if(inp)inp.focus()},0);
    return}
  const extra=CAT_FINISH_PRESETS.filter(c=>!CAT_FINISH_QUICK.includes(c.label));
  const rows=extra.map(c=>{
    const on=isFinishSelected(c.label);
    return '<div class=fmi onclick="event.stopPropagation();toggleCatFinish(\''+c.label.replace(/'/g,"\\'")+'\')">'+
      '<span style="width:14px;height:14px;border:1px solid #00000030;border-radius:3px;background:'+c.hex+';display:inline-block;flex-shrink:0"></span>'+
      '<span style="flex:1">'+escHtml(c.label)+(on?' ✓':'')+'</span>'+
      '<span class=ic style=cursor:pointer title="Remove from list" onclick="event.stopPropagation();removeCatFinishPreset(\''+c.label.replace(/'/g,"\\'")+'\')">🗑</span>'+
    '</div>'}).join('');
  menu.innerHTML='<div class=fmtitle>More Colors</div>'+
    (rows||'<div class="fmi disabled">No other colors saved yet</div>')+
    '<div class=fmsep></div>'+
    '<div class=fmi onclick="event.stopPropagation();catFinishMenuAdd(\'custom\')"><span class=ic>🎨</span>Add Custom Color</div>'+
    '<div class=fmi onclick="event.stopPropagation();catFinishMenuAdd(\'ral\')"><span class=ic>▦</span>Add RAL Color</div>'}
function catFinishMenuAdd(mode){CAT_FINISH_MENU_MODE=mode;renderCatFinishMenu()}
function catFinishMenuBack(){CAT_FINISH_MENU_MODE='list';renderCatFinishMenu()}
// RAL picker shows an actual swatch next to every code/name (a native
// <datalist> can't render color boxes, text only), so browsing feels like
// browsing a real color chart instead of guessing from a code number.
function renderCatFinishRalList(){
  const q=($('cat-finish-ral-search')?.value||'').trim().toLowerCase();
  const list=CAT_RAL_COLORS.filter(r=>!q||r.code.includes(q)||r.name.toLowerCase().includes(q));
  $('cat-finish-ral-list').innerHTML=(list.map(r=>
    '<div class=fmi onclick="event.stopPropagation();pickCatFinishRal(\''+r.code+'\')">'+
      '<span style="width:14px;height:14px;border:1px solid #00000030;border-radius:3px;background:'+r.hex+';display:inline-block;flex-shrink:0"></span>'+
      '<span style="flex:1">'+ralDatalistLabel(r)+'</span>'+
    '</div>').join(''))||'<div class="fmi disabled">No match</div>'}
async function pickCatFinishRal(code){
  const r=CAT_RAL_COLORS.find(x=>x.code===code);if(!r)return;
  await confirmAddCatFinishValue(ralDatalistLabel(r),r.hex)}
async function confirmAddCatFinish(){
  const hex=$('cat-finish-new-hex').value||'#ffffff';
  const label=$('cat-finish-new-label').value.trim();
  if(!label){alert('Enter a color name.');return}
  await confirmAddCatFinishValue(label,hex)}
async function confirmAddCatFinishValue(label,hex){
  const r=await fetch('/api/cat-finish-colors-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({label,hex})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  CAT_FINISH_PRESETS=r.colors;
  if(!isFinishSelected(label))CAT_FINISH.push({label,hex});
  CAT_FINISH_MENU_MODE='list';
  renderCatFinishMenu();renderCatFinish();renderCatOrdTable();schedulePreview()}
async function removeCatFinishPreset(label){
  const r=await fetch('/api/cat-finish-colors-remove',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({label})}).then(r=>r.json());
  if(r.colors)CAT_FINISH_PRESETS=r.colors;
  renderCatFinishMenu();renderCatFinish();renderCatOrdTable()}

// Each photo slot stores {src, zoom, x, y, mask} instead of a raw string —
// zoom/x/y drive a CSS translate()+scale() transform (pan/zoom within the
// mask, exactly mirrored in the PDF template so the adjust modal is
// WYSIWYG — see photo_cell's own comment in sololuce_datasheet.html for
// the translate math and why it replaced object-position), and
// mask is the box-inset percentage — user-adjustable for Main Product Photo
// and Dimension Diagram, but fixed at 100 for the Application Photo (see
// CAT_IMG_MASK_LOCKED below) per explicit request that its mask "be fixed
// in its location" — move/zoom stay fully free, only the mask's own size
// isn't. Default is 100 (fills the mask fully from the start) rather than
// the old 70, also per explicit request; x/y=50/50 + zoom=1 centers
// whatever the photo's natural crop is.
// label is only ever used by the 3 "extra zone" slots (see CAT_IMG_EXTRA_SLOTS
// below) — harmless dead field on main/lifestyle/diagram, kept here anyway so
// every slot shares one shape. show is the same story but only meaningful for
// extra1/extra2 — see extra_top_row_active in html_engine.py: a checkbox that
// reserves a zone's space in the printed grid even before it has a photo.
// placeholder: whether an EMPTY slot shows its dashed box + name (e.g.
// "Main Product Photo") or just stays blank — a user-owned checkbox next
// to each slot's own label in the Photos section, defaulting on so a new
// document looks exactly like it always has until the user turns one off.
// Same value renders in the live preview and the real generated PDF alike
// — no separate "hide it for real, but keep showing it while I edit" mode,
// per explicit request to keep this simple: what the checkbox says is what
// you get, everywhere, always.
// maskAnchorX/maskAnchorY: WHERE the mask box itself sits in its cell
// once Mask Size < 100% — only meaningful (and only exposed in the UI)
// for the 4 slots in CAT_IMG_ALIGN_BUTTON_SLOTS below, but every slot
// shares one shape (same reasoning as label/show above), so this is a
// harmless unused pair on main/lifestyle. 100/100 (bottom-right) matches
// the fixed position those 4 slots always used before this was
// configurable, so an old saved product renders identically until the
// user actually touches the new Mask Position buttons.
function catImgDefault(){return {src:'',zoom:1,x:50,y:50,mask:100,maskAnchorX:100,maskAnchorY:100,label:'',show:false,placeholder:true}}
// Slots whose Mask Size slider is hidden — their mask is permanently 100,
// only pan (x/y) and zoom are adjustable.
const CAT_IMG_MASK_LOCKED={lifestyle:true};
// Display name + "does this slot have its own print caption" lookup for
// openZoneSettings' popover below — every one of the 6 photo slots, not
// just the 3 CAT_IMG_EXTRA_SLOTS (diagram has a caption too; main/lifestyle
// don't, see photo_cell_captioned's own scope in the template).
// diagram:'Bottom Right' (was 'Dimension Diagram') — deliberately a FIXED
// position name here, unlike catImgTitle() a bit further down (which
// prefers extra1/2/3's own typed caption when one exists, for the Adjust
// modal's title). This popover's own title needs the fixed zone name
// even when a caption IS set, since the caption text itself is right
// there in the same popover as its own field — showing it twice (once as
// the title, once as the editable value) would read as a display bug,
// not real duplication of catImgTitle's data.
const CAT_IMG_ZONE_NAME={main:'Main Product Photo',lifestyle:'Application Photo',diagram:'Bottom Right',extra1:'Top Left',extra2:'Top Right',extra3:'Bottom Left'};
const CAT_IMG_ZONE_HAS_CAPTION={diagram:true,extra1:true,extra2:true,extra3:true};
// The 3 free zones sharing the printed "Dimension Diagram" grid with the
// (always bottom-right, fixed) diagram itself — each its own generic photo
// slot with a user-typed label instead of a fixed name.
const CAT_IMG_EXTRA_SLOTS=['extra1','extra2','extra3'];
// Slots that get the Adjust modal's 9-point alignment button grid (see
// the modal's own #photoadjust-align-row) — Dimension Diagram + the 3
// Extra Photo zones, per explicit request scoped to just those 4. Main
// Product Photo/Application Photo keep drag-only, no change there.
const CAT_IMG_ALIGN_BUTTON_SLOTS=new Set(['diagram',...CAT_IMG_EXTRA_SLOTS]);
// Grid position each generic extra zone lands in — used both as the
// default (unlabeled) placeholder text and as the Adjust modal's title,
// so it's always clear where a photo will land before typing a label.
const CAT_IMG_EXTRA_POSITION={extra1:'Top Left',extra2:'Top Right',extra3:'Bottom Left'};
// Shared "choose from Cloud Library" trigger icon — a sketchy hand-drawn
// globe (thin stroke, extra latitude lines beyond the plain Feather-icon
// version) rather than a label, colored bright green so it reads as a
// distinct action from the plain-text "Choose image…" button beside it.
const CLOUD_GLOBE_ICON='<svg width=22 height=22 viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=1.8 stroke-linecap=round stroke-linejoin=round><circle cx=12 cy=12 r=9.5/><path d="M2.5 12h19"/><path d="M12 2.5c2.8 3 4.3 6.2 4.3 9.5s-1.5 6.5-4.3 9.5c-2.8-3-4.3-6.2-4.3-9.5S9.2 5.5 12 2.5Z"/><path d="M3.8 7.5h16.4M3.8 16.5h16.4"/></svg>';
function pickCatImage(slot){
  const inp=document.createElement('input');
  inp.type='file';inp.accept='image/*';
  inp.onchange=()=>{
    const f=inp.files[0];if(!f)return;
    const reader=new FileReader();
    reader.onload=()=>{
      // Preserve whatever label/show-checkbox/merged-state/placeholder-
      // checkbox/mask-anchor was already set for this zone — swapping in a
      // new photo shouldn't wipe out what the user already said this zone
      // is for (or un-reserve its space, or un-merge it, or flip its
      // placeholder checkbox back on, or reset where its mask sits — every
      // one of these is this slot's own setting, independent of whether
      // there's currently a photo in it).
      CAT_IMG[slot]=Object.assign(catImgDefault(),{src:reader.result,label:CAT_IMG[slot].label||'',show:CAT_IMG[slot].show||false,merged:CAT_IMG[slot].merged||false,autosize:CAT_IMG[slot].autosize||false,autosizeH:CAT_IMG[slot].autosizeH||0,placeholder:CAT_IMG[slot].placeholder,maskAnchorX:CAT_IMG[slot].maskAnchorX,maskAnchorY:CAT_IMG[slot].maskAnchorY});
      renderCatImages();schedulePreview();
      openPhotoAdjust(slot)};
    reader.readAsDataURL(f)};
  inp.click()}
function removeCatImage(slot){CAT_IMG[slot]=Object.assign(catImgDefault(),{label:CAT_IMG[slot].label||'',show:CAT_IMG[slot].show||false,merged:CAT_IMG[slot].merged||false,autosize:CAT_IMG[slot].autosize||false,autosizeH:CAT_IMG[slot].autosizeH||0,placeholder:CAT_IMG[slot].placeholder,maskAnchorX:CAT_IMG[slot].maskAnchorX,maskAnchorY:CAT_IMG[slot].maskAnchorY});renderCatImages();schedulePreview()}

// ---------------------------------------------------------------- Cloud Photo Library picker
// Lets a user pick a product photo visually out of the shared R2 library
// (see photo_store.py) instead of needing the exact file already on their
// own PC — same end result as pickCatImage() (CAT_IMG[slot].src becomes a
// self-contained data: URI), just sourced from the cloud gallery instead
// of a local file. Every logged-in user can browse/pick (read-only); only
// admins can add to or remove from the library itself (Settings).
let CLOUD_PHOTO_SLOT=null, CLOUD_PHOTO_LIST=null;
async function openCloudPhotoPicker(slot){
  CLOUD_PHOTO_SLOT=slot;
  $('cloudphotomodal').classList.remove('hide');
  $('cloudphoto-search').value='';
  $('cloudphoto-status').textContent='Loading…';
  $('cloudphoto-grid').innerHTML='';
  const r=await fetch('/api/photostore-list').then(r=>r.json()).catch(e=>({error:e.message}));
  if(r.error){$('cloudphoto-status').textContent='Could not load the cloud library: '+r.error;CLOUD_PHOTO_LIST=[];return}
  CLOUD_PHOTO_LIST=r.photos||[];
  renderCloudPhotoGrid()}
function closeCloudPhotoPicker(){$('cloudphotomodal').classList.add('hide');CLOUD_PHOTO_SLOT=null}
function renderCloudPhotoGrid(){
  const q=($('cloudphoto-search').value||'').trim().toLowerCase();
  const list=(CLOUD_PHOTO_LIST||[]).filter(p=>!q||p.key.toLowerCase().includes(q));
  $('cloudphoto-status').textContent=(CLOUD_PHOTO_LIST||[]).length
    ?list.length+' of '+CLOUD_PHOTO_LIST.length+' photo'+(CLOUD_PHOTO_LIST.length!==1?'s':'')
    :'The cloud library is empty — an admin can add photos from Settings > Shared Product Photos.';
  $('cloudphoto-grid').innerHTML=list.map(p=>{
    const name=p.key.split('/').pop().replace(/\.png$/i,'');
    return '<div class=cloudphototile onclick="selectCloudPhoto(\''+p.key.replace(/'/g,"\\'")+'\')" title="'+escHtml(p.key)+'">'+
      '<img src="/api/photostore-fetch?key='+encodeURIComponent(p.key)+'" loading=lazy>'+
      '<span>'+escHtml(name)+'</span></div>'}).join('')}
function selectCloudPhoto(key){
  const slot=CLOUD_PHOTO_SLOT;if(!slot)return;
  $('cloudphoto-status').textContent='Downloading…';
  const img=new Image();
  img.crossOrigin='anonymous';
  img.onload=()=>{
    const canvas=document.createElement('canvas');
    canvas.width=img.naturalWidth;canvas.height=img.naturalHeight;
    canvas.getContext('2d').drawImage(img,0,0);
    const dataUrl=canvas.toDataURL('image/png');
    // Same "preserve everything except the photo itself" merge as
    // pickCatImage()'s own onload — see that function's comment.
    CAT_IMG[slot]=Object.assign(catImgDefault(),{src:dataUrl,label:CAT_IMG[slot].label||'',show:CAT_IMG[slot].show||false,merged:CAT_IMG[slot].merged||false,autosize:CAT_IMG[slot].autosize||false,autosizeH:CAT_IMG[slot].autosizeH||0,placeholder:CAT_IMG[slot].placeholder,maskAnchorX:CAT_IMG[slot].maskAnchorX,maskAnchorY:CAT_IMG[slot].maskAnchorY});
    closeCloudPhotoPicker();
    renderCatImages();schedulePreview();openPhotoAdjust(slot)};
  img.onerror=()=>{$('cloudphoto-status').textContent='Could not download that photo — try again.'};
  img.src='/api/photostore-fetch?key='+encodeURIComponent(key)}
// Compact button per slot (3 fit side-by-side in the .g3 row) — the actual
// "where will this land" position box lives in the live PDF preview itself
// (sololuce_datasheet.html always renders the frame, image or not), not
// here in the input form.
function renderCatImages(){
  const mergedEl=$('cat-img-extra1-merged');
  if(mergedEl)mergedEl.checked=!!CAT_IMG.extra1.merged;
  const autosizeEl=$('cat-img-extra1-autosize');
  if(autosizeEl)autosizeEl.checked=!!CAT_IMG.extra1.autosize;
  const autosizeHEl=$('cat-img-extra1-autosize-h');
  if(autosizeHEl)autosizeHEl.value=CAT_IMG.extra1.autosizeH||'';
  const mergedEl3=$('cat-img-extra3-merged');
  if(mergedEl3)mergedEl3.checked=!!CAT_IMG.extra3.merged;
  ['main','lifestyle','diagram'].concat(CAT_IMG_EXTRA_SLOTS).forEach(slot=>{
    const el=$('cat-img-'+slot);if(!el)return;
    // label is the optional print CAPTION above this zone's own box (diagram
    // + the 3 extra zones only — see photo_cell_captioned in the template),
    // not a stand-in for the box's own empty-state placeholder text
    // anymore (that's always just the fixed grid-position name now).
    // Caption + Placeholder both moved into openZoneSettings' popover below
    // (per explicit "too many checkmarks" simplification request) — their
    // inputs only exist in the DOM at all while that zone's popover happens
    // to be open, so these two lookups are almost always harmless no-ops
    // now (guarded below); CAT_IMG itself, not the DOM, stays the actual
    // source of truth throughout — see openZoneSettings' own comment.
    const labelEl=$('cat-img-'+slot+'-label');
    if(labelEl)labelEl.value=CAT_IMG[slot].label||'';
    const showEl=$('cat-img-'+slot+'-show');
    if(showEl)showEl.checked=!!CAT_IMG[slot].show;
    const placeholderEl=$('cat-img-'+slot+'-placeholder');
    if(placeholderEl)placeholderEl.checked=CAT_IMG[slot].placeholder!==false;
    // Top Right's own picker is replaced by a note while merged — its
    // photo (if it already had one from before merging was turned on)
    // stays untouched in memory, just unused, so unchecking Merge later
    // brings it right back instead of losing the upload.
    if(slot==='extra2'&&CAT_IMG.extra1.merged){
      el.innerHTML='<p class="muted" style="font-size:10.5px;margin:0;padding:8px 0">Merged into Top Left\'s wide photo — uncheck Merge to use this zone on its own again.</p>';
      return}
    // Dimension Diagram's own picker, same story as Top Right's above —
    // replaced by a note while Bottom Left's merge is on; its own photo (if
    // any) stays untouched in memory, just unused, so unchecking Merge
    // brings it right back.
    if(slot==='diagram'&&CAT_IMG.extra3.merged){
      el.innerHTML='<p class="muted" style="font-size:10.5px;margin:0;padding:8px 0">Merged into Bottom Left\'s wide photo — uncheck Merge to use this zone on its own again.</p>';
      return}
    const has=!!CAT_IMG[slot].src;
    el.innerHTML=has
      ?'<div style="display:flex;align-items:center;gap:6px">'+
         '<img src="'+CAT_IMG[slot].src+'" style="width:32px;height:32px;object-fit:cover;border:1px solid #d6dbe1;border-radius:4px;flex-shrink:0">'+
         '<button type=button class=btn style="flex:1;font-size:11px;padding:8px 4px" onclick="openPhotoAdjust(\''+slot+'\')">Adjust</button>'+
         '<button type=button class=rm onclick="removeCatImage(\''+slot+'\')" title="Remove">×</button>'+
       '</div>'
      :'<div style="display:flex;gap:5px">'+
         '<button type=button class=btn style="flex:1;font-size:11px;padding:8px 4px" onclick="pickCatImage(\''+slot+'\')">Choose image…</button>'+
         '<button type=button class=btn title="Choose from Cloud Library" style="width:38px;padding:0;flex:0 0 auto;color:#22c55e;display:flex;align-items:center;justify-content:center" onclick="openCloudPhotoPicker(\''+slot+'\')">'+CLOUD_GLOBE_ICON+'</button>'+
       '</div>'})}

// Photos section decluttering — explicit "so many checkmarks... confusing"
// request. Each zone used to show its Placeholder checkbox (every zone)
// and, for the 4 captioned ones, an always-visible caption text input
// inline, on top of whatever Reserve/Merge checkboxes it also has — up to
// 4 controls per zone x 6 zones. Reserve and Merge stay directly visible
// (they're the ones a user actually reaches for regularly — deciding
// whether a zone holds space, or combining two zones into one wide one);
// Placeholder (rarely touched — it's an on-by-default "how should this
// look before there's a photo" preference) and the optional print caption
// (typed once, then left alone) both move into this small settings
// popover instead, opened via a compact gear icon next to each zone's own
// label. Reuses the app's existing shared #filemenu popup — same
// open/position/close mechanics as openUnitManager and every other
// #filemenu-based popover here (see that function's own comment) — rather
// than inventing a second popover mechanism for one more use of the same
// idea.
// CAT_IMG (not these DOM inputs) has always been the real source of truth
// — the actual save/generate data-gathering function reads CAT_IMG.X
// directly, never $('cat-img-X-placeholder').checked — so it's safe for
// cat-img-{slot}-placeholder/-label to now only exist in the DOM while
// this popover happens to be open for that slot: renderZoneSettingsMenu
// below always initializes them fresh FROM CAT_IMG on open, and their own
// onchange/oninput handlers write straight back INTO CAT_IMG as the user
// edits, exactly like every other slot control already does — nothing
// about how the data flows actually changed, only where its controls live
// on screen.
function renderZoneSettingsMenu(slot){
  const hasCaption=!!CAT_IMG_ZONE_HAS_CAPTION[slot];
  // onclick=stopPropagation on this outer wrapper — openZoneSettings'
  // close-on-next-click listener sits on `document` (same as every other
  // #filemenu opener), so WITHOUT this, the very first click INTO the
  // caption text input would itself count as "the next click" and close
  // the whole popover before the user could type anything — confirmed
  // directly, a click that should have focused the input instead closed
  // the menu out from under it. Every other #filemenu content here avoids
  // this the same way (see removeUnit's own per-button stopPropagation) —
  // just done once at the wrapper level here since this popover's whole
  // content (checkbox AND text input) needs to survive being clicked into,
  // not just one button.
  // Tooltip wording branches on hasCaption — Main Product Photo/
  // Application Photo still print their own fixed name automatically
  // inside an empty box (no typed-caption mechanism exists for them to
  // opt into instead), but the 4 captioned zones no longer do: per
  // explicit request ("I don't want the text to be shown automatically…
  // only [show] if the text was put by the user"), their empty-state box
  // now only ever shows the SAME optional caption typed just below (see
  // photo_cell_captioned's own template comment) — nothing automatic.
  const placeholderHint=hasCaption
    ?'Show a dashed placeholder box when it has no photo yet — turn off to leave it blank instead. No text shows automatically; type an optional caption below to label it.'
    :'Show a dashed placeholder box with this slot\'s name when it has no photo yet — turn off to leave it blank instead';
  let html='<div class=fmtitle>'+CAT_IMG_ZONE_NAME[slot]+'</div>'+
    '<div onclick="event.stopPropagation()" style="padding:2px 12px 10px;'+(hasCaption?'width:220px':'width:180px')+'">'+
    '<label class=dvcheck style="font-size:11px;font-weight:600;text-transform:none;letter-spacing:normal;color:var(--ink);gap:6px;margin-bottom:'+(hasCaption?'10px':'0')+'" title="'+placeholderHint+'"><input type=checkbox id=cat-img-'+slot+'-placeholder '+(CAT_IMG[slot].placeholder!==false?'checked':'')+' onchange="CAT_IMG.'+slot+'.placeholder=this.checked;schedulePreview()">Show placeholder box when empty</label>';
  if(hasCaption)html+=
    '<label style="display:block;font-size:9.5px;font-weight:600;text-transform:uppercase;letter-spacing:.03em;color:var(--muted);margin-bottom:4px">Optional caption above photo</label>'+
    '<input type=text id=cat-img-'+slot+'-label value="'+(CAT_IMG[slot].label||'').replace(/"/g,'&quot;')+'" placeholder="Optional caption above this photo…" style="width:100%;margin-bottom:0" oninput="CAT_IMG.'+slot+'.label=this.value;schedulePreview()">';
  html+='</div>';
  $('filemenu').innerHTML=html}
function openZoneSettings(slot,btn){
  const rect=btn.getBoundingClientRect();
  renderZoneSettingsMenu(slot);
  const menu=$('filemenu');
  menu.style.display='block';
  const w=menu.offsetWidth||220,h=menu.offsetHeight||120;
  let x=rect.left,y=rect.bottom+4;
  if(x+w>window.innerWidth-8)x=window.innerWidth-w-8;
  if(y+h>window.innerHeight-8)y=window.innerHeight-h-8;
  menu.style.left=x+'px';menu.style.top=y+'px';
  // Clicking inside the popover itself (typing a caption, ticking the
  // checkbox) must NOT close it — only a click elsewhere should, same
  // "next document click, but not this one" guard every other #filemenu
  // opener here already uses via the {once:true} listener firing on the
  // NEXT click after this one, never the click that opened it.
  setTimeout(()=>document.addEventListener('click',closeFileMenu,{once:true}),0)}

// "Adjust Photo" modal — zoom (CSS scale), move (CSS translate, dragged
// directly on the preview), and mask size (the box-inset %) per photo
// slot. The preview here uses the *exact same* CSS mechanism (object-
// fit:contain + translate()+scale(), inside a mask div sized to mask%) as
// the real PDF template, so what's shown here is what actually renders —
// no separate crop/bake step, no drift between the two.
// Extra zones share the diagram's own 2x2-grid cell shape EXACTLY — not
// roughly. All 4 (Dimension Diagram + Top Left/Top Right/Bottom Left) sit
// in the identical grid.template-columns:1fr 1fr / gap:8px right-column
// layout (sololuce_datasheet.html — the "active" 2x2 grid when at least one
// of Top Left/Top Right is in use, or the "collapsed" 1x2 grid — just
// Bottom Left + Diagram — when neither is, per that template's own
// comment), and confirmed by direct Playwright measurement (getBoundingClientRect
// on the real generated page) that both grid modes produce the IDENTICAL
// 280x106 cell size — the "active" grid's extra row-gap is exactly offset
// by its own extra row, so cutting from 2 rows to 1 changes nothing about
// any single cell's own shape. 280/106 was previously approximated as 1.88
// for diagram and a completely different 1.5 for the 3 extra zones despite
// all 4 being the SAME cell shape — confirmed as the actual root cause of
// a real "the picture isn't full" report: the modal, cropping to the wrong
// (too-square) guide shape, showed meaningfully MORE of a mismatched-aspect
// photo filling the frame than the real page (the true, more elongated
// 280:106 shape) then actually renders — not a pan/zoom bug, a stale
// cropping-guide constant. main/lifestyle aren't touched here — those two
// use a real CSS aspect-ratio (1/1, 16/9) directly on their own template
// element, confirmed exact via the same measurement, so they were never
// approximations to begin with.
const CAT_IMG_ASPECT={main:1,lifestyle:16/9,diagram:280/106,extra1:280/106,extra2:280/106,extra3:280/106};
// The 4 bottom-right grid zones (Dimension Diagram + the 3 generic Extra
// Photo slots) are often used for line-art/diagrams that need to shrink
// well below their frame rather than fill it, so their zoom floor is
// lower than a fill-the-frame photo's. Application Photo joined this list
// per explicit request too, leaving Main Product Photo as the only slot
// still floored at 100% — zooming out there just reveals blank space
// around a product shot, which was never wanted, only up-zoom (crop in).
const CAT_IMG_ZOOM_MIN={lifestyle:50,diagram:50,extra1:50,extra2:50,extra3:50};
// Every photo slot uses object-fit:contain in the real template
// (sololuce_datasheet.html's photo_cell macro / the Main+Application photo
// blocks right above it — see their own comments), NOT cover: cover crops
// to fill the frame's aspect ratio before zoom is even applied, so content
// could stay clipped no matter how far out the user zooms — confirmed
// directly, a real dimension diagram's "H"/"Ø" callouts never became fully
// visible at any zoom level under cover. This used to be diagram/extra-
// only (main/lifestyle kept cover, on the theory that losing a sliver of
// background isn't a real information loss) but was widened to every slot
// per explicit request — "these settings" (never-crop, contain-based) now
// apply to every placeholder picture alike, not just the line-art ones.
// This modal's own preview (renderPhotoAdjustPreview) always renders
// contain now too, so what's shown here never disagrees with what the
// real PDF renders — same contract as the rest of this modal's CSS
// mechanism, see this file's own comment above openPhotoAdjust.
const CAT_IMG_ALIGN={
  main:{items:'center',justify:'center'},
  lifestyle:{items:'flex-end',justify:'flex-start'},
  diagram:{items:'flex-end',justify:'flex-end'},
  extra1:{items:'flex-end',justify:'flex-end'},
  extra2:{items:'flex-end',justify:'flex-end'},
  extra3:{items:'flex-end',justify:'flex-end'}};
// v (0/50/100, same scale the Align buttons already use for photo pan)
// -> the flex keyword it maps to. Flexbox only ever has 3 real stops per
// axis — there's no continuous "where does a smaller box sit inside a
// bigger one" the way translate() gives x/y for the photo-inside-mask
// case — so 3 is genuinely all there is here, same reasoning as
// photo_cell's own anchor_x/anchor_y in the template.
function maskAnchorToFlex(v){
  return v<=0?'flex-start':(v>=100?'flex-end':'center')}
// #photoadjust-frame's own align-items/justify-content: CAT_IMG_ALIGN's
// static per-slot lookup for every slot EXCEPT the 4 in
// CAT_IMG_ALIGN_BUTTON_SLOTS, which instead read the document's own live
// CAT_IMG[slot].maskAnchorX/Y (the Mask Position buttons' own state) —
// same axis mapping as photo_cell's own anchor_x/anchor_y in the
// template (justify-content=X, align-items=Y, this frame's default row
// flex-direction), so the modal preview can never disagree with the real
// generated PDF about which button means what.
function applyPhotoAdjustFrameAlign(slot){
  const frame=$('photoadjust-frame');
  if(CAT_IMG_ALIGN_BUTTON_SLOTS.has(slot)){
    const st=CAT_IMG[slot];
    frame.style.justifyContent=maskAnchorToFlex(st.maskAnchorX);
    frame.style.alignItems=maskAnchorToFlex(st.maskAnchorY);
    return}
  const align=CAT_IMG_ALIGN[slot];
  frame.style.alignItems=align.items;
  frame.style.justifyContent=align.justify}
// Mask Position buttons — where the mask BOX ITSELF sits in its cell,
// not where the photo pans inside it (that's setPhotoAdjustAlign above).
function setPhotoAdjustMaskAnchor(x,y){
  const slot=PHOTO_ADJUST_SLOT;if(!slot)return;
  const st=CAT_IMG[slot];
  st.maskAnchorX=x;st.maskAnchorY=y;
  applyPhotoAdjustFrameAlign(slot);
  syncPhotoAdjustMaskAnchorHighlight();
  schedulePreview()}
function catImgTitle(slot){
  if(slot==='main')return 'Main Product Photo';
  if(slot==='lifestyle')return 'Application Photo';
  // Was 'Dimension Diagram' — renamed to 'Bottom Right' per explicit
  // request, to read as the 4th corner of a plain TOP LEFT/TOP RIGHT/
  // BOTTOM LEFT/BOTTOM RIGHT 2x2 grid instead of implying a fixed
  // technical-diagram-only purpose. Display name only — the underlying
  // field/variable (dimension_diagram, CAT_IMG.diagram, slot key
  // 'diagram') is unchanged everywhere else, so no saved draft needs
  // migrating.
  if(slot==='diagram')return 'Bottom Right';
  return CAT_IMG[slot].label||CAT_IMG_EXTRA_POSITION[slot]||('Extra Zone '+slot.slice(-1))}
let PHOTO_ADJUST_SLOT=null;
function openPhotoAdjust(slot){
  if(!CAT_IMG[slot].src)return;
  PHOTO_ADJUST_SLOT=slot;
  $('photoadjust-info-text').classList.add('hide');
  $('photoadjust-title').textContent='Adjust '+catImgTitle(slot);
  const frame=$('photoadjust-frame');
  // Merge (extra1/extra3 + their own "Merge with ... Right" checkbox)
  // doubles this zone's real printed width (spanning both the left and
  // right column it replaces) — needs its own special-cased aspect ratio
  // here (568/106, i.e. 2*280+8: the doubled width, over the single row
  // height — CAT_IMG_ASPECT[slot] alone is only ever the single-column
  // shape). This has flip-flopped with the merged box's own real on-page
  // shape across several requests (see sololuce_datasheet.html's own
  // history on this): full-width originally (568/106, correct); briefly
  // 50%-narrower/right-aligned (280/106 — same as unmerged, so the
  // special case was correctly removed for that period); reverted back to
  // full-width per a later explicit correction (an annotated screenshot).
  // That last revert updated the TEMPLATE but missed updating this
  // function to match — a real, confirmed bug of the exact SAME shape
  // this comment already once described ("a photo positioned to look
  // right in the modal did NOT land there in the real document"),
  // re-introduced by omission rather than by a genuine design change this
  // time. Restored below.
  //
  // autosize (extra_photo_1_autosize/extra_photo_3_autosize) compounds
  // this further when it's actually on: the real merged box's height then
  // ISN'T the fixed 106px this 568/106 ratio assumes at all — it's
  // recomputed at print time from the photo's OWN natural aspect ratio
  // (autoSizeTopRowToPicture in the template), so the box becomes exactly
  // the photo's own shape with zero letterbox slack to pan through at
  // zoom:1 — reported directly as "top left is not responding to user
  // alignment". Using the SAME photo's natural aspect ratio here for the
  // modal's own frame (read off the already-loaded thumbnail <img> this
  // panel already renders, not a fresh fetch) makes the modal honest
  // about that: it shows the same zero-slack shape the real PDF will
  // actually use, so a drag genuinely doing nothing at zoom:1 is
  // consistent between modal and output — matches, rather than silently
  // diverges from, the real page. Zooming in past 100% still restores
  // real pan/align in both places alike (see computePhotoOffset — any
  // zoom>1 creates overflow to pan through regardless of box aspect).
  const merged=(slot==='extra1'||slot==='extra3')&&!!CAT_IMG[slot].merged;
  let aspect=CAT_IMG_ASPECT[slot];
  if(merged){
    aspect=568/106;
    if(CAT_IMG[slot].autosize){
      const thumbImg=document.querySelector('#cat-img-'+slot+' img');
      if(thumbImg&&thumbImg.naturalWidth&&thumbImg.naturalHeight)aspect=thumbImg.naturalWidth/thumbImg.naturalHeight}}
  frame.style.aspectRatio=aspect;
  applyPhotoAdjustFrameAlign(slot);
  const st=CAT_IMG[slot];
  // Locked slots (Application Photo) always fill the mask fully — its
  // location/size is fixed, only pan+zoom are adjustable — so the Mask
  // Size row is hidden and the value pinned to 100 rather than left at
  // whatever it happened to be (old data saved before this lock existed).
  const locked=!!CAT_IMG_MASK_LOCKED[slot];
  if(locked)st.mask=100;
  $('photoadjust-mask-row').classList.toggle('hide',locked);
  $('photoadjust-align-row').classList.toggle('hide',!CAT_IMG_ALIGN_BUTTON_SLOTS.has(slot));
  $('photoadjust-mask-anchor-row').classList.toggle('hide',!CAT_IMG_ALIGN_BUTTON_SLOTS.has(slot));
  // "Standard Settings" one-click preset — shown for every photo slot
  // this modal ever opens for (badges never reach this modal at all, so
  // there's no separate exclusion to write). Started as "Catalogue
  // Standard (65%)" for Main Product Photo only, then widened to every
  // slot per explicit request. A locked-mask slot (Application Photo)
  // can't take the size part of that preset at all — its frame is fixed —
  // so its label and action drop the size part and only recenter pan/zoom.
  // Dimension Diagram's own target is 100%, not 65% — see
  // catImgStandardMaskTarget's own comment.
  $('photoadjust-standard-label').textContent=locked?'Standard Settings':'Standard Settings ('+catImgStandardMaskTarget(slot)+'% size)';
  // Set min BEFORE value — an <input type=range> silently clamps its
  // current value up to a newly-raised min, so setting them in the other
  // order would snap a slot with no floor back to 100 right after we
  // just lowered it for the previous slot.
  $('photoadjust-zoom').min=CAT_IMG_ZOOM_MIN[slot]||100;
  $('photoadjust-zoom').value=Math.round(st.zoom*100);
  $('photoadjust-mask-slider').value=st.mask;
  syncPhotoAdjustLabels();
  // Modal must be shown (class hide removed) BEFORE renderPhotoAdjustPreview()
  // runs — that function reads #photoadjust-mask's real getBoundingClientRect()
  // to compute the photo's pan transform, and .clientmodal.hide is display:none,
  // so calling it first (the old order) always measured a 0x0 box: photoPanConsts
  // short-circuits boxW/boxH=0 to {cx:0,cy:0}, so the transform came out as
  // translate(0%,0%) — a centered-looking photo — no matter what the slot's real
  // saved x/y actually was. This silently mismatched the real position every
  // single time the modal opened; previously invisible because nothing else on
  // screen disagreed with it. Now that the Align/Mask Position grids correctly
  // highlight the real saved position (see syncPhotoAdjustAlignHighlight above),
  // the gap between "Align says Top" and a preview photo that isn't actually
  // top-aligned became obvious — reported directly as "after reopening the
  // window, the preview shows wrong."
  $('photoadjustmodal').classList.remove('hide');
  renderPhotoAdjustPreview();
  syncPhotoAdjustAlignHighlight();
  syncPhotoAdjustMaskAnchorHighlight()}
// Neither button grid ever highlighted which position was actually active
// — clicking one only ever left the browser's own transient :focus ring,
// which is gone the instant the modal closes and reopens. The underlying
// x/y (or maskAnchorX/Y) was always saved correctly — reported directly as
// "when user presses align... pressing adjust again the align preview is
// back to standard again", i.e. it LOOKS reset (nothing highlighted) even
// though the real position didn't move. These two walk the button grid
// (data-x/data-y set on each button, see the HTML above) and mark whichever
// one's x/y exactly matches the slot's current value — none match after a
// free-form drag that didn't land exactly on a grid point, which is
// correct: no single button describes that position, so none should look
// selected.
function syncPhotoAdjustAlignHighlight(){
  const slot=PHOTO_ADJUST_SLOT;if(!slot)return;
  const st=CAT_IMG[slot],grid=$('photoadjust-align-grid');if(!grid)return;
  grid.querySelectorAll('button').forEach(b=>b.classList.toggle('on',+b.dataset.x===st.x&&+b.dataset.y===st.y))}
function syncPhotoAdjustMaskAnchorHighlight(){
  const slot=PHOTO_ADJUST_SLOT;if(!slot)return;
  const st=CAT_IMG[slot],grid=$('photoadjust-mask-anchor-grid');if(!grid)return;
  grid.querySelectorAll('button').forEach(b=>b.classList.toggle('on',+b.dataset.x===st.maskAnchorX&&+b.dataset.y===st.maskAnchorY))}
// The "Standard Settings" preset's own target mask size, per slot: a
// locked slot (Application Photo) has no size to change at all — its
// frame is fixed at 100%, see CAT_IMG_MASK_LOCKED. Dimension Diagram
// isn't locked (its Mask Size slider stays visible and user-adjustable)
// but its OWN standard target is 100%, not the 65% every other slot
// lands on — per explicit request: a technical drawing reads best
// filling its whole box, unlike a 65%-cropped product PHOTO. Every other
// slot (Main Product Photo, the 3 Extra Photo zones) keeps 65%.
function catImgStandardMaskTarget(slot){
  if(CAT_IMG_MASK_LOCKED[slot]||slot==='diagram')return 100;
  return 65}
function setPhotoAdjustCatalogueStandard(){
  const slot=PHOTO_ADJUST_SLOT;if(!slot)return;
  const st=CAT_IMG[slot];
  // Recenters pan and resets zoom for every slot — a "standard" crop only
  // means something starting from a known, centered position. The size
  // itself layers on top of that per catImgStandardMaskTarget above,
  // for every slot except a locked one (Application Photo), whose frame
  // is intentionally fixed at 100% and has no size to change here.
  st.zoom=1;st.x=50;st.y=50;
  $('photoadjust-zoom').value=100;
  if(!CAT_IMG_MASK_LOCKED[slot]){st.mask=catImgStandardMaskTarget(slot);$('photoadjust-mask-slider').value=st.mask}
  syncPhotoAdjustLabels();
  renderPhotoAdjustPreview();syncPhotoAdjustAlignHighlight();schedulePreview()}
function closePhotoAdjust(){
  $('photoadjustmodal').classList.add('hide');
  PHOTO_ADJUST_SLOT=null;
  renderCatImages();schedulePreview()}
// The (i) toggle next to the modal title — see the comment on that button
// in the HTML above for why this replaced a permanently-visible line of
// text. Collapsed again on every fresh open (openPhotoAdjust) so it never
// carries a "left open" state from one slot into the next.
function togglePhotoAdjustInfo(){$('photoadjust-info-text').classList.toggle('hide')}
// Generic version of the same "(i) icon collapses this section's own
// permanently-visible explanation paragraph(s) behind a click" pattern —
// see photoadjust-info-text's own toggle above for the original single-
// paragraph case. Data-driven by class rather than a single id so one
// header icon can cover several paragraphs scattered through one section
// (e.g. Ordering Table's column-widths/row-heights notes, or Photos'
// own two) without a separate named function per section. Applied to
// Basics, Family, Photos, and Ordering Table — the only cat-only sections
// that actually carry a static "how this works" paragraph; Spec Badges,
// Technical Specifications, and Finish Colors have none, so no icon there.
function toggleCatSectionInfo(cls){document.querySelectorAll('.'+cls).forEach(p=>p.classList.toggle('hide'))}
// Always shows the slider's own current value next to its label — whichever
// photo (main/lifestyle/diagram) is open, the reading is never left stale
// from a previous slot or a previous drag.
function syncPhotoAdjustLabels(){
  $('photoadjust-zoom-val').textContent=$('photoadjust-zoom').value+'%';
  $('photoadjust-mask-val').textContent=$('photoadjust-mask-slider').value+'%'}
// Shared with the exact same function inlined into the real PDF template
// (sololuce_datasheet.html's own <script> near the end of <body>) — see
// that copy's own comment for the full derivation. Kept as two copies
// (JS in a browser tab here vs. JS re-executed inside Playwright's
// headless page for the PDF) rather than one shared file, same reasoning
// as applyPhotoAdjustFrameAlign/photo_cell's anchor_x/anchor_y already
// being two copies — there's no module system bridging this app.py
// (Flask + embedded <script>) and a Jinja2-rendered template, so keeping
// the two copies textually identical is what "WYSIWYG parity" means in
// practice here. boxW/boxH: the mask box's actual rendered pixel size.
// natW/natH: the photo's own intrinsic pixel size. x/y: 0-100 pan value
// (50=centered). zoom: 1.0-3.0 multiplier.
//
// Old formula was tx=(50-x)*(zoom-1) — correct ONLY when the photo's own
// aspect ratio happens to exactly match the mask's, so contain-fit leaves
// zero letterbox gap in either axis: at zoom=1 there's genuinely nothing
// to pan (the whole photo already fills the mask), so tx=0 was right.
// But whenever the photo's shape DOESN'T match the mask's (the common
// case — a wide reference photo in a squarer zone, say), contain-fit
// leaves an empty gap in one axis even at zoom=1, and the old formula
// left that gap unreachable no matter how the user dragged, since
// (zoom-1)=0 at the default 100% zoom regardless of x/y — confirmed
// directly, dragging visibly did nothing until zooming in first, even
// though the photo had real letterbox room to move into. Fixed by
// scaling the offset against the photo's OWN rendered-at-this-zoom size
// relative to the box (rW/boxW, rH/boxH) instead of assuming they're
// equal: at zoom=1 this ratio is exactly the existing gap (still 0 for a
// matching-aspect photo, preserving the old behavior there exactly), and
// growing zoom smoothly transitions it into the old zoom-created-overflow
// case once the photo's rendered size passes the box's — one formula
// covers both regimes, and the photo is still never cropped at zoom=1
// either way, same contract as before.
// photoPanConsts: the (zoom*rendered/box - 1) term from the tx/ty formula,
// factored out because initPhotoAdjustDrag (below) needs it too — dragging
// has to INVERT this formula (pixel delta -> x/y delta), not just evaluate
// it, and the two need to agree on exactly the same constant or the drag
// and the resulting render would disagree about which x/y a given mouse
// position maps to.
function photoPanConsts(boxW,boxH,natW,natH,zoom){
  if(!boxW||!boxH||!natW||!natH)return{cx:0,cy:0};
  // No safety-margin factor here (a *0.98 shrink briefly lived on this
  // line). It was added to try to protect thin edge-hugging design
  // elements (e.g. a dimension diagram's "H" label) from vanishing at
  // extreme pan positions, but testing then showed it never actually
  // fixed that — even a much bigger margin (15%) didn't bring the label
  // back, so the real cause was something else entirely (a Chromium PDF-
  // export quirk with heavily-downscaled, heavily-transformed images),
  // not reachable by shrinking this fit target. Meanwhile the margin had
  // a real, confirmed cost of its own: it made "Align Top" (or any other
  // flush position) never actually reach the mask's true edge, leaving a
  // permanent ~2% gap no matter what — reported directly as "align it
  // from top, the picture is not actually where I'm putting it". Since it
  // wasn't fixing the thing it was added for, keeping a regression that
  // breaks correct positioning for everyone isn't a reasonable trade —
  // removed. s0 is exact contain-fit again: x/y=0/50/100 land flush with
  // the mask's real edges with zero gap, matching what the Align buttons
  // (and dragging) actually say they're doing.
  const s0=Math.min(boxW/natW,boxH/natH);
  return{cx:(zoom*natW*s0/boxW)-1,cy:(zoom*natH*s0/boxH)-1}}
function computePhotoOffset(boxW,boxH,natW,natH,x,y,zoom){
  const{cx,cy}=photoPanConsts(boxW,boxH,natW,natH,zoom);
  return{tx:(50-x)*cx,ty:(50-y)*cy}}
function renderPhotoAdjustPreview(){
  const slot=PHOTO_ADJUST_SLOT;if(!slot)return;
  const st=CAT_IMG[slot],mask=$('photoadjust-mask'),img=$('photoadjust-img');
  mask.style.width=st.mask+'%';mask.style.height=st.mask+'%';
  img.style.objectFit='contain';
  // A modal-only paint glitch was investigated here (a real, extremely
  // wide merged photo — ~9.6:1 accessories icon strip — painted nothing
  // at all at Align Top, in this modal's live preview specifically) and a
  // small inward-nudge fix was tried and reverted: it didn't actually fix
  // the blank paint (confirmed directly — still blank with a clean 2px
  // margin off the clip edge, and a from-scratch re-encoded copy of the
  // exact same pixels still failed to paint under this transform, while a
  // synthetic solid-color test image at the identical size/transform
  // painted fine) — apparently a Chromium paint/compositing quirk
  // specific to that one real image, not a geometry bug this math can
  // reach. It also had a real cost of its own: a permanent 2px short of
  // true "flush" for every non-centered position on every OTHER photo,
  // the same class of regression the *0.98 fix above was reverted for.
  // Not worth keeping a change that doesn't fix the thing it was for.
  // Confirmed separately (Playwright, pixel-exact) that the REAL
  // generated PDF does NOT have this bug — only this on-screen preview,
  // for this one real photo, at this one exact pan position.
  const apply=()=>{
    const rect=mask.getBoundingClientRect();
    const off=computePhotoOffset(rect.width,rect.height,img.naturalWidth,img.naturalHeight,st.x,st.y,st.zoom);
    img.style.transform='translate('+off.tx+'%,'+off.ty+'%) scale('+st.zoom+')'};
  if(img.src!==st.src){
    // Fresh photo (first open, or Change Photo) — naturalWidth/Height
    // aren't available until it's actually decoded, so wait for load
    // before the real (aspect-ratio-aware) transform can be computed;
    // apply() still runs once immediately with whatever's already there
    // so zoom/mask changes on an already-loaded photo stay instant.
    img.onload=apply;img.src=st.src}
  else if(img.complete&&img.naturalWidth){apply()}
  else{img.onload=apply}}
function onPhotoAdjustChange(){
  const slot=PHOTO_ADJUST_SLOT;if(!slot)return;
  const st=CAT_IMG[slot];
  st.zoom=$('photoadjust-zoom').value/100;
  if(!CAT_IMG_MASK_LOCKED[slot])st.mask=Number($('photoadjust-mask-slider').value);
  syncPhotoAdjustLabels();
  renderPhotoAdjustPreview();schedulePreview()}
function resetPhotoAdjust(){
  const slot=PHOTO_ADJUST_SLOT;if(!slot)return;
  Object.assign(CAT_IMG[slot],{zoom:1,x:50,y:50,mask:100});
  $('photoadjust-zoom').value=100;$('photoadjust-mask-slider').value=100;
  syncPhotoAdjustLabels();
  renderPhotoAdjustPreview();syncPhotoAdjustAlignHighlight();syncPhotoAdjustMaskAnchorHighlight();schedulePreview()}
// One-click 9-point alignment (#photoadjust-align-row's own buttons) —
// jumps straight to x/y=0/50/100, the exact same 0-100 scale a manual
// drag already writes to (see initPhotoAdjustDrag below), just without
// having to eyeball a corner/edge/center by hand. Zoom/mask/src are all
// untouched — this is a pan-only shortcut, same as a drag would be.
function setPhotoAdjustAlign(x,y){
  const slot=PHOTO_ADJUST_SLOT;if(!slot)return;
  const st=CAT_IMG[slot];
  st.x=x;st.y=y;
  renderPhotoAdjustPreview();syncPhotoAdjustAlignHighlight();schedulePreview()}
(function initPhotoAdjustDrag(){
  const mask=$('photoadjust-mask');if(!mask)return;
  let dragging=false,lastX=0,lastY=0;
  mask.addEventListener('mousedown',e=>{
    if(!PHOTO_ADJUST_SLOT)return;
    dragging=true;lastX=e.clientX;lastY=e.clientY;
    mask.style.cursor='grabbing';
    e.preventDefault()});
  window.addEventListener('mousemove',e=>{
    if(!dragging||!PHOTO_ADJUST_SLOT)return;
    // Direct manipulation — the photo follows the cursor 1:1 in actual
    // screen pixels, same as dragging a layer in any image editor: drag
    // right, the photo visibly moves right, by exactly as many pixels as
    // the cursor did. That's NOT the same thing as "increase x" or
    // "decrease x" by a fixed sign — tx/ty's own formula is
    // (50-x)*(zoom*rendered/box-1), and that (...-1) term can be NEGATIVE
    // (the photo's rendered size is SMALLER than the box in this axis —
    // the letterbox-gap case a 100%-zoom, mismatched-aspect photo hits) or
    // POSITIVE (the photo overflows the box — the classic zoomed-in-past-
    // 100% case), and each sign needs the OPPOSITE x-delta to move the
    // same visual direction. A single fixed-sign version (either the
    // original startPx-dxPct, or a flipped startPx+dxPct tried right
    // after) can only ever be correct in ONE of those two regimes and
    // backwards in the other — confirmed directly: startPx+dxPct fixed
    // the letterbox-gap case (the only one tested at the time) but was
    // then reported as newly-backwards once tested on an actually-zoomed
    // photo, which is the overflow regime. Dividing the pixel delta by
    // this SAME (...-1) constant (photoPanConsts, shared with
    // computePhotoOffset above) inverts the formula properly instead of
    // guessing a sign, so it's correct in both regimes at once — confirmed
    // by re-deriving: solving tx_new = tx_start + mouseDeltaPct for x_new
    // in that formula gives exactly x -= deltaPct/const. Tracked frame-to-
    // frame (delta since the LAST mousemove, not since mousedown) rather
    // than accumulated from a fixed start — zoom never changes mid-drag,
    // so both give the same end result, but frame-to-frame never compounds
    // a large single division and needs no separate startPx/startPy state.
    // Near-zero const (the photo exactly fills the box in this axis, at
    // this exact zoom — genuinely nothing to pan into) is left alone
    // rather than divided, which would blow up to a huge/NaN jump.
    const rect=mask.getBoundingClientRect();
    const img=$('photoadjust-img');
    const st=CAT_IMG[PHOTO_ADJUST_SLOT];
    const dxPct=((e.clientX-lastX)/rect.width)*100;
    const dyPct=((e.clientY-lastY)/rect.height)*100;
    lastX=e.clientX;lastY=e.clientY;
    const{cx,cy}=photoPanConsts(rect.width,rect.height,img.naturalWidth,img.naturalHeight,st.zoom);
    if(Math.abs(cx)>1e-6)st.x=Math.max(0,Math.min(100,st.x-dxPct/cx));
    if(Math.abs(cy)>1e-6)st.y=Math.max(0,Math.min(100,st.y-dyPct/cy));
    renderPhotoAdjustPreview();syncPhotoAdjustAlignHighlight()});
  window.addEventListener('mouseup',()=>{
    if(dragging){dragging=false;mask.style.cursor='grab';schedulePreview()}})})();

// Ordering/variant table — the one place in this app with variable *fields*
// (columns) as well as variable variants (rows). Line Items only ever
// varies row count, per a fixed COLS[TYPE] field set — see itemCardHtml
// above; here the field set itself is user-defined too. Each field gets a
// stable internal key at creation time (independent of its editable label)
// so renaming a field header never has to remap every variant's data.
//
// Fields are added via a standard-category picker (real catalogue
// convention: Model No., Power, Size, CCT, Beam Angle, Input Voltage,
// Controls, Finish Options, Options, Lumen — "Size" has two variants,
// Ø×H for round fittings and L×W×H for rectangular ones, kept as two
// separate pickable entries) rather than typed blind, but the label stays
// a free-text input afterward so it can still be tweaked or replaced with
// anything. New fields always append to the end; use a field chip's ◀/▶
// controls (moveCatOrdField) to reposition it afterward.
let CAT_ORD_CATEGORIES=[];
async function loadCatOrdCategories(){
  const r=await fetch('/api/cat-ordering-categories').then(r=>r.json());
  if(r.categories)CAT_ORD_CATEGORIES=r.categories}
function openCatOrdColMenu(ev){
  ev.stopPropagation();
  const rect=ev.currentTarget.getBoundingClientRect();
  const menu=$('filemenu');
  // A pending Cut/Copy shows up here as a Paste option — same "reuse the
  // + Add Field entry point rather than invent new UI" as the rest of this
  // clipboard feature (see CAT_ORD_COL_CLIPBOARD's own comment).
  const pasteRow=CAT_ORD_COL_CLIPBOARD?'<div class=fmi onclick="ordColPaste()"><span class=ic>📋</span>Paste "'+escHtml(CAT_ORD_COL_CLIPBOARD.label||'field')+'"</div><div class=fmsep></div>':'';
  menu.innerHTML=pasteRow+'<div class=fmtitle>Add field</div>'+
    CAT_ORD_CATEGORIES.map((c,i)=>'<div class=fmi onclick="pickCatOrdCategory('+i+')">'+escHtml(c)+'</div>').join('')+
    '<div class=fmsep></div>'+
    '<div class=fmi onclick="pickCatOrdCategory(-1)">Custom…</div>';
  menu.style.display='block';
  const w=menu.offsetWidth||200,h=menu.offsetHeight||200;
  let x=rect.left,y=rect.bottom+4;
  if(x+w>window.innerWidth-8)x=window.innerWidth-w-8;
  if(y+h>window.innerHeight-8)y=window.innerHeight-h-8;
  menu.style.left=x+'px';menu.style.top=y+'px';
  setTimeout(()=>document.addEventListener('click',closeFileMenu,{once:true}),0)}
function pickCatOrdCategory(i){
  const label=i>=0?CAT_ORD_CATEGORIES[i]:'';
  addCatOrdField(label);
  closeFileMenu();
  // Renaming now lives in the field's own click popover (see
  // openCatOrdFieldPopover — the pill itself lost its text input when it
  // became icon-only), so a fresh Custom… field opens straight into that
  // popover's rename mode instead of focusing an input that no longer
  // exists on the pill. Deferred a tick (matches this function's own prior
  // behavior) so it runs after THIS click has fully finished bubbling —
  // opening #filemenu synchronously inside the same click that's still
  // bubbling toward the stale "close on next click" listener from
  // whichever popover was open a moment ago would have it close again
  // immediately, before the user ever sees it.
  if(i<0){
    const ci=CAT_ORD_COLS.length-1;
    setTimeout(()=>{
      const btn=document.querySelectorAll('#cat-ord-fields .catordfieldicon')[ci];
      if(btn){renderCatOrdFieldPopoverAt(btn.getBoundingClientRect(),ci);ordFieldEnterRename(ci)}},0)}}
// Click-to-open info/actions popover for one field pill — full name, a
// short description (CAT_ORD_FIELD_DESC), and Copy/Cut/Delete. Reuses
// #filemenu (small, positioned-near-click, viewport-clamped), the same
// surface Fill Standard Information's own Configure popover uses — this is
// the "keep this tool accessible for other parts of the datasheet later"
// pattern, not a one-off.
function openCatOrdFieldPopover(ev,ci){
  ev.stopPropagation();
  renderCatOrdFieldPopoverAt(ev.currentTarget.getBoundingClientRect(),ci)}
function renderCatOrdFieldPopoverAt(rect,ci){
  const c=CAT_ORD_COLS[ci];
  const key=catOrdFieldIconKey(c);
  const menu=$('filemenu');
  menu.innerHTML=
    '<div class=fmtitle id=catordfieldpopover-title style="cursor:text" title="Click to rename" onclick="event.stopPropagation();ordFieldEnterRename('+ci+')">'+escHtml(c.label||'(untitled field)')+'</div>'+
    '<div style="padding:0 12px 8px;font-size:11px;color:var(--muted)">'+escHtml(CAT_ORD_FIELD_DESC[key])+'</div>'+
    // Same ✓/— toggle the pill's own icon row keeps (a quick at-a-glance
    // scan is still worth it there), just also reachable from here — the
    // click popover is meant to hold every per-field action, not just the
    // ones that lost their pill button when the pill went icon-only.
    '<label class=fmi style="cursor:pointer" onclick="event.stopPropagation()">'+
      '<input type=checkbox id=catordfieldpopover-codeon'+(catOrdCodeOn(c)?' checked':'')+' onchange="ordFieldPopoverToggleCode(event,'+ci+')" style="margin-right:2px">'+
      'Include in Ordering Code Example</label>'+
    '<div class=fmsep></div>'+
    '<div class=fmi onclick="ordColCopy('+ci+')"><span class=ic>⧉</span>Copy</div>'+
    '<div class=fmi onclick="ordColCut('+ci+')"><span class=ic>✂</span>Cut</div>'+
    (CAT_ORD_COLS.length>1?'<div class="fmi danger" onclick="ordColDelete(this,event,'+ci+')"><span class=ic>🗑</span><span class=fmilabel>Delete</span></div>':'');
  menu.style.display='block';
  const w=menu.offsetWidth||200,h=menu.offsetHeight||180;
  let x=rect.left,y=rect.bottom+4;
  if(x+w>window.innerWidth-8)x=window.innerWidth-w-8;
  if(y+h>window.innerHeight-8)y=window.innerHeight-h-8;
  menu.style.left=x+'px';menu.style.top=y+'px';
  setTimeout(()=>document.addEventListener('click',closeFileMenu,{once:true}),0)}
// Re-renders the pill grid (unaffected by the popover being open) and
// keeps the popover's own checkbox in sync with the result — toggleCatOrd
// CodeOn doesn't know this popover exists, so this is the seam that keeps
// the two displays of the same state (pill + popover) from drifting apart.
function ordFieldPopoverToggleCode(ev,ci){
  ev.stopPropagation();
  toggleCatOrdCodeOn(ci);
  const cb=$('catordfieldpopover-codeon');
  if(cb)cb.checked=catOrdCodeOn(CAT_ORD_COLS[ci]);}
function ordFieldEnterRename(ci){
  const c=CAT_ORD_COLS[ci];
  const title=$('catordfieldpopover-title');
  if(!title)return;
  title.outerHTML='<input id=catordfieldpopover-rename class=fmtitle style="width:100%;box-sizing:border-box;border:1px solid var(--border);border-radius:6px" value="'+escHtml(c.label||'')+'" oninput="CAT_ORD_COLS['+ci+'].label=this.value;schedulePreview()" onblur="rememberCatOrdCategory(this.value);renderCatOrdTable();closeFileMenu();schedulePreview()">';
  setTimeout(()=>{const inp=$('catordfieldpopover-rename');if(inp){inp.focus();inp.select()}},0)}
function ordColCopy(ci){
  const c=CAT_ORD_COLS[ci];
  CAT_ORD_COL_CLIPBOARD={label:c.label,values:c.values.slice(),width:c.width,codeOn:c.codeOn};
  closeFileMenu();
  toast('Copied "'+(c.label||'field')+'" — click "+ Add Field" then Paste to add it')}
function ordColCut(ci){
  const c=CAT_ORD_COLS[ci];
  CAT_ORD_COL_CLIPBOARD={label:c.label,values:c.values.slice(),width:c.width,codeOn:c.codeOn};
  closeFileMenu();
  toast('Cut "'+(c.label||'field')+'" — click "+ Add Field" then Paste to place it');
  removeCatOrdCol(ci)}
// Same double-click-to-confirm pattern fmDelete uses in the document-level
// file menu (window.confirm() silently no-ops in this app's runtime) — an
// inline per-button timer, not a shared one, same as deleteManageListItem.
function ordColDelete(el,ev,ci){
  if(el.dataset.confirm!=='1'){
    ev.stopPropagation();
    el.dataset.confirm='1';
    el.querySelector('.fmilabel').textContent='Click again to confirm';
    setTimeout(()=>{if(el.dataset.confirm==='1'){el.dataset.confirm='';const lbl=el.querySelector('.fmilabel');if(lbl)lbl.textContent='Delete'}},2500);
    return}
  closeFileMenu();
  removeCatOrdCol(ci)}
function ordColPaste(){
  if(!CAT_ORD_COL_CLIPBOARD)return;
  const key='col'+(CAT_COL_SEQ++);
  CAT_ORD_COLS.push({key,label:CAT_ORD_COL_CLIPBOARD.label,values:CAT_ORD_COL_CLIPBOARD.values.slice(),width:CAT_ORD_COL_CLIPBOARD.width,codeOn:CAT_ORD_COL_CLIPBOARD.codeOn});
  closeFileMenu();
  renderCatOrdTable();schedulePreview();
  toast('Pasted "'+(CAT_ORD_COL_CLIPBOARD.label||'field')+'"')}
// Always appends — a brand-new field is seeded with one blank value per
// EXISTING variant (not just one), so it starts already lockstepped with
// every other field instead of becoming the one ragged column.
function addCatOrdField(label){
  const key='col'+(CAT_COL_SEQ++);
  CAT_ORD_COLS.push({key,label,values:Array(catOrdVariantCount()).fill('')});
  renderCatOrdTable();schedulePreview()}
async function rememberCatOrdCategory(label){
  label=(label||'').trim();
  if(!label||CAT_ORD_CATEGORIES.includes(label))return;
  const r=await fetch('/api/cat-ordering-categories-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({label})}).then(r=>r.json());
  if(r.categories)CAT_ORD_CATEGORIES=r.categories}
function removeCatOrdCol(i){
  CAT_ORD_COLS.splice(i,1);
  renderCatOrdTable();schedulePreview()}
// Swaps two whole fields (label + entire values array) — repositions a
// field that landed in the wrong spot without losing whatever was already
// typed into it, unlike delete-and-recreate.
function moveCatOrdField(ci,dir){
  const nci=ci+dir;
  if(nci<0||nci>=CAT_ORD_COLS.length)return;
  const tmp=CAT_ORD_COLS[ci];
  CAT_ORD_COLS[ci]=CAT_ORD_COLS[nci];
  CAT_ORD_COLS[nci]=tmp;
  renderCatOrdTable();schedulePreview()}
// Every field in CAT_ORD_COLS is kept "lockstepped" to one shared variant
// count, except the two computed fields (Finish Options/Lumen — see
// recomputeCatSpecialOrdColumns below), which keep their own independent
// length driven by Finish Colors' count / Power's count respectively.
const CAT_ORD_SPECIAL_LABELS=['Finish Options','Lumen'];
function catOrdLockstepCols(){
  return CAT_ORD_COLS.filter(c=>!CAT_ORD_SPECIAL_LABELS.includes((c.label||'').trim()))}
function catOrdVariantCount(){
  const cols=catOrdLockstepCols();
  return cols.length?Math.max(...cols.map(c=>c.values.length)):1}
// Shared by catOrdVariantLabel/catOrdVariantDropdownLabel below — Model
// No. + Power are the two fields that actually identify a real orderable
// SKU, so both variant-naming schemes are built from the same lookup
// rather than each re-finding the columns.
function catOrdVariantIdentity(vi){
  const modelCol=CAT_ORD_COLS.find(c=>isCatOrdModelNoColumn(c.label));
  const powCol=CAT_ORD_COLS.find(c=>isCatOrdPowerColumn(c.label));
  return {
    model:((modelCol&&modelCol.values[vi])||'').trim(),
    pow:((powCol&&powCol.values[vi])||'').trim()}}
// What the Complete button calls a variant — "SLAQU 8W" instead of a
// meaningless "Variant 3" once there's more than a couple of them. Falls
// back to "Variant N" wherever Model No./Power aren't typed yet (a fresh
// blank variant, or a table that doesn't even have those fields) — self-
// upgrades to the real name the moment they're filled in, no separate
// "name this variant" step needed.
function catOrdVariantLabel(vi){
  const id=catOrdVariantIdentity(vi);
  const parts=[id.model,id.pow].filter(Boolean);
  return parts.length?parts.join(' '):'Variant '+(vi+1)}
// The dropdown gets its own, more verbose label — a running position
// number ("No 1", "No 2"...) so it's obvious at a glance where in the
// list you are, plus "Crafting Table" as a fixed place-name prefix (this
// screen's own nickname throughout, printed literally so the dropdown
// reads the same way regardless of which document/screen it's opened from
// later) — ahead of the same Model No./Power identity catOrdVariantLabel
// uses everywhere else.
function catOrdVariantDropdownLabel(vi){
  const id=catOrdVariantIdentity(vi);
  const parts=['No '+(vi+1),'Crafting Table'];
  if(id.model)parts.push(id.model);
  if(id.pow)parts.push(id.pow);
  return parts.join(' - ')}
// Technical Specifications' own generic "+ Add Value"/Remove/Move on its
// Power row (addCatSpecValue etc.) can change the Ordering Table's Power
// column length independently of every control in this table, via
// syncOrderingPowerFromSpec — this pads every OTHER lockstepped field back
// up to match, right before every render, so that kind of edit always
// ripples into a proper new/removed variant everywhere else instead of
// silently letting Power run longer than Size/Cut Out/etc. Never truncates:
// shrinking one field's own value count is not the same action as removing
// a whole variant (removeCatOrdVariant), and shouldn't delete data from
// every other field as a side effect.
function normalizeCatOrdLockstep(){
  const n=catOrdVariantCount();
  catOrdLockstepCols().forEach(c=>{while(c.values.length<n)c.values.push('')})}
// Model No. auto-copies down from the row above whenever a row has picked
// up SOME other field's value but Model No. itself is still blank there —
// covers every edit path (typing, picking from a dropdown, drag-reorder,
// import), not just the dedicated "+ Add Variant" button's own existing
// copy-forward (addCatOrdVariant), since a row can end up with data in it
// through any of those. Never overwrites a Model No. the user actually
// typed for that specific row — only fills a genuinely blank one.
function autoFillModelNoDown(){
  const modelCol=CAT_ORD_COLS.find(c=>isCatOrdModelNoColumn(c.label));
  if(!modelCol)return;
  const ci=CAT_ORD_COLS.indexOf(modelCol);
  const n=catOrdVariantCount();
  for(let vi=0;vi<n;vi++){
    if((modelCol.values[vi]||'').trim())continue;
    if(CAT_ORD_MODELNO_EXPLICIT_EMPTY.has(ci+':'+vi))continue;
    const hasInfo=CAT_ORD_COLS.some(c=>c!==modelCol&&(c.values[vi]||'').trim());
    if(!hasInfo)continue;
    for(let j=vi-1;j>=0;j--){
      if((modelCol.values[j]||'').trim()){modelCol.values[vi]=modelCol.values[j];break}
    }
  }}
// Model No. is the one field that copies forward instead of starting blank
// on a new variant — its code is usually shared/similar across wattage
// variants of the same fixture, so a fresh row starts as a copy of the
// previous one (to tweak) rather than empty. Every other field starts
// blank. Power syncs to Technical Specifications afterward since its array
// just grew; nothing else needs to, since every field grew together.
function addCatOrdVariant(){
  const prev=catOrdVariantCount()-1;
  catOrdLockstepCols().forEach(c=>{
    c.values.push(normSpecLabel(c.label)==='modelno.'?(c.values[prev]||''):'')});
  const pow=findCatOrdPowerCol();
  if(pow)syncSpecPowerFromOrdering(pow.values);
  // A new blank row can't introduce a new distinct size on its own, but
  // recompute anyway — cheap, and keeps this in lockstep with remove/move
  // below rather than relying on "add never needs it" staying true forever.
  recomputeCatOrdSizeDNumbers();
  recomputeOrderingCodeExample();
  renderCatOrdTable();schedulePreview()}
function removeCatOrdVariant(vi){
  if(catOrdVariantCount()<=1)return;
  catOrdLockstepCols().forEach(c=>{
    if(vi<c.values.length)c.values.splice(vi,1);
    if(!c.values.length)c.values.push('')});
  const pow=findCatOrdPowerCol();
  if(pow)syncSpecPowerFromOrdering(pow.values);
  // Removing a row can retire a distinct size entirely (or leave its
  // D-number now belonging to whichever row is left using it first) — the
  // whole column needs fresh numbering, not just a shift.
  recomputeCatOrdSizeDNumbers();
  recomputeOrderingCodeExample();
  renderCatOrdTable();schedulePreview()}
// Manual reorder — swaps an entire variant (every field's value at index vi)
// with its neighbor in one atomic step, so Size/Cut Out/etc. always stay
// matched to whichever Power value they travel with. Numeric auto-sort was
// deliberately removed (see sortCatOrdVariantsBy) since sorting one field in
// isolation would silently decouple it from the rest of its own row.
function moveCatOrdVariant(vi,dir){
  const nvi=vi+dir;
  if(nvi<0||nvi>=catOrdVariantCount())return;
  catOrdLockstepCols().forEach(c=>{
    if(vi>=c.values.length||nvi>=c.values.length)return;
    const tmp=c.values[vi];c.values[vi]=c.values[nvi];c.values[nvi]=tmp});
  const pow=findCatOrdPowerCol();
  if(pow)syncSpecPowerFromOrdering(pow.values);
  // Reordering changes which row counts as the "first" occurrence of a
  // distinct size, which is exactly what decides its D-number here.
  recomputeCatOrdSizeDNumbers();
  recomputeOrderingCodeExample();
  renderCatOrdTable();schedulePreview()}
// Drag-drop counterpart of moveCatOrdField — unlike the arrow version this
// can land anywhere in one move, so it splices the whole field out of
// CAT_ORD_COLS and back in at the drop position (before/after read off
// which half of the target pill the pointer is over) instead of replaying
// adjacent swaps.
function ordFieldDrop(e,targetCi){
  e.preventDefault();
  e.currentTarget.classList.remove('dragover-left','dragover-right');
  if(!DRAG_KEY||DRAG_KEY.kind!=='field')return;
  const ci=DRAG_KEY.ci;
  if(ci===targetCi)return;
  const rect=e.currentTarget.getBoundingClientRect();
  const before=(e.clientX-rect.left)<rect.width/2;
  let insertAt=targetCi+(before?0:1);
  const item=CAT_ORD_COLS.splice(ci,1)[0];
  if(ci<insertAt)insertAt--;
  CAT_ORD_COLS.splice(insertAt,0,item);
  renderCatOrdTable();schedulePreview()}
// (The old ordVariantDrop — drag-drop counterpart to moveCatOrdVariant —
// was removed when the Ordering Table moved to showing one variant at a
// time: with only one variant visible there's no second card to drop onto
// anymore. moveCatOrdVariant's ▲▼ buttons, exposed via
// moveCatOrdVariantCurrent, are now the only reorder path.)
// Shared by every value-cell builder below (generic/CCT/Size/Controls/
// Voltage/Power) instead of each hand-rolling its own flex wrapper. Move/
// remove chrome is per-cell plumbing left over from the old column-major
// layout — the row-major renderer always calls these with total=1,
// canRemove=false (a single cell never shows its own move/remove buttons
// anymore; that's handled once per variant card instead, see
// moveCatOrdVariant/removeCatOrdVariant), so in practice `moves`/`rm` below
// are always empty today. Left general rather than hardcoded to no-op, in
// case a future caller legitimately wants per-cell chrome again.
function catOrdValueRowWrap(ci,vi,total,inner,canRemove){
  const moves=total>1?(
    '<div style="display:flex;flex-direction:column;">'+
      '<button type=button class=ordmovebtn'+(vi===0?' disabled':'')+' onclick="moveCatOrdValue('+ci+','+vi+',-1)" title="Move up">▲</button>'+
      '<button type=button class=ordmovebtn'+(vi===total-1?' disabled':'')+' onclick="moveCatOrdValue('+ci+','+vi+',1)" title="Move down">▼</button>'+
    '</div>'):'';
  const rm=canRemove?'<button type=button class=rm onclick="removeCatOrdValue('+ci+','+vi+')" title="Remove this value">×</button>':'';
  return '<div style="display:flex;gap:3px;align-items:flex-start">'+moves+inner+rm+'</div>'}
function onCatOrdValueInput(ci,vi,val){
  CAT_ORD_COLS[ci].values[vi]=val;
  recomputeOrderingCodeExample();
  schedulePreview()}
// Finish Options and Lumen recompute on blur (not on every keystroke) so
// typing in an unrelated column's textarea never gets its DOM rebuilt out
// from under the cursor mid-word — collectCatData() also recomputes them
// right before Generate/preview, so the saved PDF is always correct even
// if the user never blurs the field they were last typing in.
function onCatOrdValueBlur(){
  recomputeCatSpecialOrdColumns();
  // The oninput handler already scheduled a preview while the user was
  // typing, but that debounced timer can fire (and render) BEFORE blur —
  // e.g. the raw "15" gets rendered, then blur reformats it to "15 W" a
  // moment later with nothing left to trigger a re-render. Scheduling again
  // here means blur-triggered formatting always gets its own preview pass.
  schedulePreview()}
// Same auto-unit behavior as the Technical Specifications smart inputs
// (CAT_SPEC_SUFFIX/appendCatSpecSuffix), matched by column label instead of
// spec label — "Size" matches any column starting with "Size" (e.g. the
// default "Size (Ø×H)") since the unit applies regardless of what's in the
// parentheses.
// Every known Ordering Table column now has its own dropdown-of-presets
// cell (CCT/Size/Controls/Power/Input Voltage/Beam Angle/Model No./Cut Out/
// Options below) — these two maps only still apply to a genuinely custom
// column name the user invents that matches none of the above, which is why
// both are empty by default now. Left in place (rather than deleted) since
// it's still real, harmless infrastructure for that one remaining case.
const CAT_ORD_SUFFIX={};
function catOrdSuffixFor(label){
  return CAT_ORD_SUFFIX[normSpecLabel(label)]||null}
const CAT_ORD_PLACEHOLDERS={};
function catOrdPlaceholderFor(label){
  return CAT_ORD_PLACEHOLDERS[normSpecLabel(label)]||'e.g. Value'}
function catOrdSortableValue(val){
  val=val||'';
  const full=val.includes('\n')?val.split('\n').slice(1).join(' '):val;
  const m=full.match(/[\d.]+/);
  return m?parseFloat(m[0]):null}
// Explicit, on-demand replacement for what used to be an automatic
// per-column sort (removed — silently re-sorting just the field being
// edited would have decoupled it from the rest of its own variant row,
// since a row's fields must all move together to stay matched, e.g. a
// specific Size with the Power it actually belongs to). Picking a field
// here computes ONE permutation from that field's values (ascending
// numeric; non-numeric/blank values sink to the end) and applies the exact
// same permutation to every lockstepped field at once, so nothing decouples
// from anything else. Also reindexes every per-cell "*_CUSTOM" Set (keyed
// ci+':'+vi) so a cell mid-custom-typing doesn't silently point at the
// wrong row after the reorder.
function sortCatOrdVariantsBy(ci){
  const ref=CAT_ORD_COLS[ci];
  if(!ref)return;
  const n=catOrdVariantCount();
  const order=Array.from({length:n},(_,i)=>i);
  order.sort((a,b)=>{
    const na=catOrdSortableValue(ref.values[a]),nb=catOrdSortableValue(ref.values[b]);
    if(na==null&&nb==null)return 0;
    if(na==null)return 1;
    if(nb==null)return -1;
    return na-nb});
  catOrdLockstepCols().forEach(c=>{
    if(c.values.length!==n)return;
    c.values=order.map(i=>c.values[i])});
  const newPosOf=new Array(n);
  order.forEach((oldVi,newVi)=>{newPosOf[oldVi]=newVi});
  [CAT_ORD_SIZE_CUSTOM,CAT_ORD_CCT_CUSTOM,CAT_ORD_CONTROLS_CUSTOM,CAT_ORD_VOLTAGE_CUSTOM,
   CAT_ORD_POWER_CUSTOM,CAT_ORD_BEAMANGLE_CUSTOM,CAT_ORD_MODELNO_CUSTOM,CAT_ORD_OPTIONS_CUSTOM,
   CAT_ORD_CUTOUT_CUSTOM].forEach(set=>{
    [...set].forEach(key=>{
      const sep=key.indexOf(':');
      const kci=key.slice(0,sep),kvi=parseInt(key.slice(sep+1),10);
      if(kvi>=0&&kvi<n){set.delete(key);set.add(kci+':'+newPosOf[kvi])}})});
  const pow=findCatOrdPowerCol();
  if(pow)syncSpecPowerFromOrdering(pow.values);
  // Same reasoning as moveCatOrdVariant — sorting changes which row is the
  // "first" occurrence of each distinct size.
  recomputeCatOrdSizeDNumbers();
  recomputeOrderingCodeExample();
  renderCatOrdTable();schedulePreview()}
function onCatOrdValueUnitBlur(ci,vi,el){
  const label=CAT_ORD_COLS[ci]&&CAT_ORD_COLS[ci].label;
  const suffix=catOrdSuffixFor(label);
  if(suffix){
    const formatted=appendCatSpecSuffix(el.value,suffix);
    el.value=formatted;
    CAT_ORD_COLS[ci].values[vi]=formatted}
  renderCatOrdTable();
  onCatOrdValueBlur()}
// Ordering Code Example is generated, not typed — real Sololuce sheets build
// it by joining each ordering-table column's *code* half (the bold first
// line of a two-line cell) from each column's first value with '-', e.g.
// Model/Power/Size/CCT/... -> "SLAUR-5W-D3-30-...". Recomputed on every
// table edit; left alone (not blanked) while there's no table yet, so an
// imported/typed value isn't wiped out before a table exists.
// Most columns' code half is just the bold first line as typed (Model No.,
// Controls, Finish Options, Size/CCT's own short codes below). Power and
// Beam Angle are the two exceptions — they're plain single-line values with
// a unit suffix, and the code needs a compact form that isn't in the
// suffix: Power keeps the "W" but drops the space ("9 W" -> "9W"), Beam
// Angle drops the ° entirely per the real catalogue's own code convention
// ("100 °" -> "100").
function catOrdCodeSegment(label,firstValue){
  const code=(firstValue||'').split('\n')[0].trim();
  const key=normSpecLabel(label);
  if(key==='power')return code.replace(/\s+/g,'');
  if(key==='beamangle')return code.replace(/°/g,'').replace(/\s+/g,'');
  return code}
// These categories default ON (matches the real catalogue's own code
// convention — Cut Out/Options/Lumen/a fresh custom field don't normally
// belong in the code); every field still gets its own on/off checkmark
// regardless, per explicit request ("the check mark should be on all the
// headers basically") — a field outside this default set just starts
// unchecked instead of not having the control at all, so it CAN be added
// to the code with one click when a particular sheet actually needs that.
const CAT_ORD_CODE_CATEGORIES=['power','cct','beamangle','inputvoltage','controls'];
function isCatOrdCodeableCategory(label){
  const n=normSpecLabel(label);
  if(n.startsWith('modelno'))return true;
  if(n.startsWith('size'))return true;
  if(n.startsWith('finish'))return true;
  return CAT_ORD_CODE_CATEGORIES.includes(n)}
// Effective on/off state for a field's own checkmark — c.codeOn is only
// ever explicitly set once the user actually clicks the toggle; before
// that it falls back to isCatOrdCodeableCategory's own default so existing
// products' generated codes don't change just from this control becoming
// visible everywhere.
function catOrdCodeOn(c){return c.codeOn===undefined?isCatOrdCodeableCategory(c.label):c.codeOn!==false}
function toggleCatOrdCodeOn(ci){
  const c=CAT_ORD_COLS[ci];
  c.codeOn=!catOrdCodeOn(c);
  renderCatOrdTable();schedulePreview()}
function recomputeOrderingCodeExample(){
  if(!CAT_ORD_COLS.length)return;
  const code=CAT_ORD_COLS
    .filter(c=>catOrdCodeOn(c))
    .map(c=>catOrdCodeSegment(c.label,(c.values&&c.values[0])||''))
    .filter(Boolean).join('-');
  $('cat-ordcode').value=code}
// "Finish Options" and "Lumen" are computed columns, not typed — matched by
// exact column label, so renaming either column away from these names just
// turns it back into a normal free-text column (an accepted trade-off, not
// a bug: the special behavior is tied to what the column is *called*).
function catFinishShortCode(label){
  const known={'Black':'BK','White':'WH','Grey':'GY'};
  if(known[label])return known[label];
  const m=(label||'').match(/^RAL\s*(\d{3,4})/i);
  if(m)return 'RAL'+m[1];
  return (label||'').split(/\s+/).map(w=>w[0]||'').join('').toUpperCase().slice(0,3)||'?'}
function catFinishFullText(label){
  const m=(label||'').match(/^RAL\s*\d{3,4}\s*(.*)$/i);
  if(m&&m[1].trim())return m[1].trim().toUpperCase();
  return (label||'').toUpperCase()}
function recomputeCatSpecialOrdColumns(){
  const fin=CAT_ORD_COLS.find(c=>(c.label||'').trim()==='Finish Options');
  if(fin){
    fin.values=CAT_FINISH.length
      ?CAT_FINISH.map(f=>catFinishShortCode(f.label)+'\n'+catFinishFullText(f.label))
      :['']}
  const lum=CAT_ORD_COLS.find(c=>(c.label||'').trim()==='Lumen');
  if(lum){
    const pow=CAT_ORD_COLS.find(c=>(c.label||'').trim()==='Power');
    const spec=CAT_SPECS.find(s=>(s.label||'').trim().toLowerCase()==='luminare efficacy');
    // A fitting can have more than one efficacy figure (different quality/
    // driver grades of the same fixture, not tied to any particular
    // wattage tier), so every row's OWN Power gets multiplied against
    // EVERY real efficacy value — not just whichever one happened to share
    // that row's index (the earlier convention: row 0 only ever got
    // efficacy value 0, every other row silently fell back to the LAST
    // efficacy value, so with e.g. 8 Power rows and only 2 efficacy
    // values, 7 of the 8 rows all showed the same one figure). Each row's
    // Lumen cell instead stacks one line per grade — "500 lm\n600 lm" —
    // the same '\n'-joined multi-line convention CCT/Controls/Size already
    // use, so it prints as extra lines in that one cell (narrow), never as
    // extra rows.
    const effs=(spec&&spec.values||[]).map(v=>{
      const m=(v||'').match(/[\d.]+/);
      return m?parseFloat(m[0]):null}).filter(e=>e!=null);
    if(pow&&pow.values.length&&effs.length){
      lum.values=pow.values.map(v=>{
        const pm=(v||'').match(/[\d.]+/);
        if(!pm)return '';
        const p=parseFloat(pm[0]);
        return effs.map(eff=>Math.round(p*eff)+' lm').join('\n')})}
    else lum.values=['']}}
// Size gets a short "D{n}" index (D for Dimension) instead of the raw
// dimensions in the code — same two-line "shortcode\nfull spec" storage as
// CCT, except the index isn't computable from the text alone (it's
// arbitrary, e.g. "120x60mm"). cat_size_index is a catalogue-wide
// persisted list of every size text ever typed, used ONLY to power the
// suggestion dropdown below (so a size used on an earlier product is a
// one-click pick here too, no code/alias shown next to it — see
// catOrdSizeCellHtml) — it is explicitly NOT where the D-number comes
// from anymore. D-numbers used to be that list's own global position (so
// "120x60mm" was always D-whatever-slot-it-first-claimed, catalogue-wide,
// even on a totally different product), but that's wrong for how these
// codes are actually used: each datasheet's own D1/D2/D3... is specific
// to THAT product, always starting at D1 for whichever size it uses
// first, regardless of what D-number that same size text happens to
// carry on some other product's sheet — see recomputeCatOrdSizeDNumbers
// below, which is the actual source of every row's D-number now.
let CAT_SIZE_INDEX=[];
async function loadCatSizeIndex(){
  const r=await fetch('/api/cat-size-index').then(r=>r.json());
  if(r.sizes)CAT_SIZE_INDEX=r.sizes}
// Registers fullSpec in the catalogue-wide suggestion list if it isn't
// there yet — no return value worth using anymore (see the block comment
// above): a size's position in this shared list no longer has anything to
// do with its D-number on any given datasheet.
async function rememberCatSizeOption(fullSpec){
  if(CAT_SIZE_INDEX.includes(fullSpec))return;
  const r=await fetch('/api/cat-size-index-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({value:fullSpec})}).then(r=>r.json());
  if(r.sizes)CAT_SIZE_INDEX=r.sizes}
function isCatOrdSizeColumn(label){return normSpecLabel(label).startsWith('size')}
function findCatOrdSizeCol(){return CAT_ORD_COLS.find(c=>isCatOrdSizeColumn(c.label))}
function sizeFullSpecOf(val){
  const nl=(val||'').indexOf('\n');
  return nl>=0?val.slice(nl+1):(val||'')}
function stripCatOrdSizeMm(fullSpec){return (fullSpec||'').replace(/\s*mm$/i,'').trim()}
// The actual D-number source for every row of THIS product's own Size
// column — recomputed from scratch, whole column at once, every time any
// row's size changes (edited, picked, cleared) or a variant row is added/
// removed/reordered, since inserting or deleting a distinct size anywhere
// can shift every LATER distinct size's number, not just the row that
// changed. Walks the column top to bottom, hands out D1/D2/D3... in the
// order each DISTINCT size text is first seen, and a later row reusing an
// already-seen size just gets that same number back — matching how a real
// printed sheet's D-codes work (one physical size = one D-number, for
// this product only, however many variant rows share it). Empty rows are
// skipped entirely — they consume no D-number and leave a gap for
// nothing. The stored size TEXT is never touched here, only the "D{n}\n"
// prefix ahead of it.
function recomputeCatOrdSizeDNumbers(){
  const col=findCatOrdSizeCol();
  if(!col)return;
  const seen=[];
  col.values=col.values.map(v=>{
    const fullSpec=sizeFullSpecOf(v);
    if(!fullSpec)return '';
    const norm=stripCatOrdSizeMm(fullSpec);
    let idx=seen.indexOf(norm);
    if(idx<0){seen.push(norm);idx=seen.length-1}
    return 'D'+(idx+1)+'\n'+fullSpec})}
let CAT_ORD_SIZE_CUSTOM=new Set();
// Size already had a full "remember what was typed" registry (CAT_SIZE_INDEX/
// rememberCatSizeOption above) — it just never had a dropdown surfacing it.
// This adds the same select-or-Custom… shell as every other column on top
// of that existing registry, so a previously-typed size becomes a one-click
// pick instead of the user having to retype the exact same text to reuse
// it — plain size text only, no "(D{n})" alias next to it (that number is
// this PRODUCT's own, decided by recomputeCatOrdSizeDNumbers above, not a
// property of the size itself worth advertising in a cross-product list).
function catOrdSizeCellHtml(ci,vi,val,canRemove,total){
  const raw=stripCatOrdSizeMm(sizeFullSpecOf(val));
  const key=ci+':'+vi;
  const customMode=!CAT_SIZE_INDEX.length||CAT_ORD_SIZE_CUSTOM.has(key)||(raw&&!CAT_SIZE_INDEX.some(o=>stripCatOrdSizeMm(o)===raw));
  if(customMode)
    return catOrdValueRowWrap(ci,vi,total,
      '<textarea rows=1 style="width:110px;font-size:11px;resize:vertical;box-sizing:border-box" placeholder="e.g. 120×60" oninput="onCatOrdValueInput('+ci+','+vi+',this.value)" onblur="onCatOrdSizeBlur('+ci+','+vi+',this)">'+escHtml(raw)+'</textarea>',
      canRemove);
  const opts='<option value="">— Select —</option>'+
    CAT_SIZE_INDEX.map(o=>{
      const or=stripCatOrdSizeMm(o);
      return '<option value="'+escHtml(o)+'"'+(or===raw?' selected':'')+'>'+escHtml(or)+'</option>'}).join('')+
    '<option value="__empty__">Empty</option>'+
    '<option value="__custom__">Custom…</option>';
  return catOrdValueRowWrap(ci,vi,total,
    '<select style="width:110px;font-size:11px" onchange="onCatOrdSizeChange('+ci+','+vi+',this)">'+opts+'</select>',
    canRemove)}
function onCatOrdSizeChange(ci,vi,sel){
  const key=ci+':'+vi;
  if(sel.value==='__custom__'){
    CAT_ORD_SIZE_CUSTOM.add(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    return}
  if(sel.value==='__empty__'){
    CAT_ORD_SIZE_CUSTOM.delete(key);
    CAT_ORD_COLS[ci].values[vi]='';
    recomputeCatOrdSizeDNumbers();
    renderCatOrdTable();
    recomputeOrderingCodeExample();
    schedulePreview();
    return}
  CAT_ORD_SIZE_CUSTOM.delete(key);
  // Store the raw size text with no D-prefix yet — recomputeCatOrdSizeDNumbers
  // fills in the correct LOCAL D-number for every row of this column right
  // after, this one included (sizeFullSpecOf treats a prefix-less value as
  // its own full spec, so this is a safe, valid intermediate state).
  CAT_ORD_COLS[ci].values[vi]=sel.value;
  recomputeCatOrdSizeDNumbers();
  renderCatOrdTable();
  recomputeOrderingCodeExample();
  schedulePreview()}
async function onCatOrdSizeBlur(ci,vi,el){
  const key=ci+':'+vi;
  const raw=el.value.trim();
  if(!raw){CAT_ORD_SIZE_CUSTOM.delete(key);CAT_ORD_COLS[ci].values[vi]='';recomputeCatOrdSizeDNumbers();renderCatOrdTable();onCatOrdValueBlur();return}
  const fullSpec=appendCatSpecSuffix(raw,'mm');
  await rememberCatSizeOption(fullSpec);
  CAT_ORD_SIZE_CUSTOM.delete(key);
  CAT_ORD_COLS[ci].values[vi]=fullSpec;
  recomputeCatOrdSizeDNumbers();
  renderCatOrdTable();
  onCatOrdValueBlur()}
// CCT gets a dropdown of the standard color temperatures instead of free
// text, same "Custom…" escape hatch as the Technical Specifications
// dropdowns — matched by column label (isCatOrdCctColumn), same as the
// suffix map above.
let CAT_ORD_CCT_OPTIONS=['2700','3000','3500','4000','5000','6000','6500'];
async function loadCatCctOptions(){
  const r=await fetch('/api/cat-cct-options').then(r=>r.json());
  if(r.options)CAT_ORD_CCT_OPTIONS=r.options}
async function rememberCatCctOption(value){
  value=(value||'').trim();
  if(!value)return;
  // Same trailing-K/whitespace-insensitive compare as the cell builder's own
  // digits extraction (full.replace(/\s*K$/i,'')) — see rememberCatPowerOption.
  const norm=value.replace(/\s*K$/i,'').trim();
  if(CAT_ORD_CCT_OPTIONS.some(o=>o.replace(/\s*K$/i,'').trim()===norm))return;
  const r=await fetch('/api/cat-cct-options-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({value})}).then(r=>r.json());
  if(r.options)CAT_ORD_CCT_OPTIONS=r.options}
let CAT_ORD_CCT_CUSTOM=new Set();
function isCatOrdCctColumn(label){return normSpecLabel(label)==='cct'}
// Stored as a two-line "shortcode\nfull spec" value — same convention as
// Finish Options/Model No. — so the Ordering Code Example's generic "first
// line of each column" join picks up just "27" for real, and the table
// cell itself shows "27" bold with "2700 K" underneath, per the real
// catalogue's own code (SLCER-9W-D1-27-100-24-ND-WH).
function cctFullSpecOf(val){
  const nl=(val||'').indexOf('\n');
  return nl>=0?val.slice(nl+1):(val||'')}
function cctShortCode(fullSpec){
  const m=(fullSpec||'').match(/(\d+)/);
  return m?String(Math.round(parseInt(m[1],10)/100)):''}
function catOrdCctCellHtml(ci,vi,val,canRemove,total){
  const full=cctFullSpecOf(val).trim();
  const digits=full.replace(/\s*K$/i,'').trim();
  const key=ci+':'+vi;
  const customMode=CAT_ORD_CCT_CUSTOM.has(key)||(full&&!CAT_ORD_CCT_OPTIONS.includes(digits));
  if(customMode)
    return catOrdValueRowWrap(ci,vi,total,
      '<textarea rows=1 style="width:110px;font-size:11px;resize:vertical;box-sizing:border-box" placeholder="e.g. 3450K" oninput="onCatOrdValueInput('+ci+','+vi+',this.value)" onblur="onCatOrdCctCustomBlur('+ci+','+vi+',this)">'+escHtml(full)+'</textarea>',
      canRemove);
  const opts='<option value="">— Select —</option>'+
    CAT_ORD_CCT_OPTIONS.map(o=>'<option value="'+o+' K"'+(digits===o?' selected':'')+'>'+o+' K</option>').join('')+
    '<option value="__empty__">Empty</option>'+
    '<option value="__custom__">Custom…</option>';
  return catOrdValueRowWrap(ci,vi,total,
    '<select style="width:110px;font-size:11px" onchange="onCatOrdCctChange('+ci+','+vi+',this)">'+opts+'</select>',
    canRemove)}
function onCatOrdCctChange(ci,vi,sel){
  const key=ci+':'+vi;
  if(sel.value==='__custom__'){
    CAT_ORD_CCT_CUSTOM.add(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    return}
  if(sel.value==='__empty__'){
    CAT_ORD_CCT_CUSTOM.delete(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    recomputeOrderingCodeExample();
    schedulePreview();
    return}
  CAT_ORD_CCT_CUSTOM.delete(key);
  const fullSpec=sel.value;
  CAT_ORD_COLS[ci].values[vi]=cctShortCode(fullSpec)+'\n'+fullSpec;
  renderCatOrdTable();
  recomputeOrderingCodeExample();
  schedulePreview()}
// A custom CCT the user actually typed (not the "still empty, just chose
// Custom…" case) gets remembered as a new preset — once saved it drops
// back into dropdown mode showing the value it just learned, same
// "remember what was typed" pattern as Series/Category and Spec Labels.
function onCatOrdCctCustomBlur(ci,vi,el){
  const raw=el.value.trim();
  const key=ci+':'+vi;
  if(!raw){CAT_ORD_CCT_CUSTOM.delete(key);CAT_ORD_COLS[ci].values[vi]='';renderCatOrdTable();onCatOrdValueBlur();return}
  const fullSpec=appendCatSpecSuffix(raw,'K');
  const digits=fullSpec.replace(/\s*K$/i,'').trim();
  const finish=()=>{
    CAT_ORD_CCT_CUSTOM.delete(key);
    CAT_ORD_COLS[ci].values[vi]=cctShortCode(fullSpec)+'\n'+fullSpec;
    renderCatOrdTable();
    onCatOrdValueBlur()};
  if(digits&&/^\d+$/.test(digits))rememberCatCctOption(digits).then(finish);
  else finish()}
// Controls gets the same dropdown-of-presets-plus-Custom… treatment as CCT,
// stored as the same two-line "shortcode\nfull spec" convention (Controls is
// one of the code-contributing categories — CAT_ORD_CODE_CATEGORIES above —
// so its code half needs to stay compact, e.g. "ND" not "Non-Dimmable", to
// match the real catalogue's own terse dash-joined code style).
let CAT_ORD_CONTROLS_OPTIONS=['DALI','0-10V','Phase Dim','Non-Dimmable'];
async function loadCatControlsOptions(){
  const r=await fetch('/api/cat-controls-options').then(r=>r.json());
  if(r.options)CAT_ORD_CONTROLS_OPTIONS=r.options}
async function rememberCatControlsOption(value){
  value=(value||'').trim();
  if(!value)return;
  // Whitespace-insensitive compare, same pattern as rememberCatPowerOption —
  // keeps this consistent with its sibling remember*Option functions even
  // though Controls values aren't currently unit-suffixed.
  const norm=value.replace(/\s+/g,'');
  if(CAT_ORD_CONTROLS_OPTIONS.some(o=>o.replace(/\s+/g,'')===norm))return;
  const r=await fetch('/api/cat-controls-options-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({value})}).then(r=>r.json());
  if(r.options)CAT_ORD_CONTROLS_OPTIONS=r.options}
let CAT_ORD_CONTROLS_CUSTOM=new Set();
function isCatOrdControlsColumn(label){return normSpecLabel(label)==='controls'}
function controlsFullSpecOf(val){
  const nl=(val||'').indexOf('\n');
  return nl>=0?val.slice(nl+1):(val||'')}
// Known compact codes for the built-in presets (matching the real catalogue's
// terse style, e.g. "...-ND-WH"); an unrecognized custom value falls back to
// the same initials-based shortener as catFinishShortCode.
function controlsShortCode(full){
  const known={'DALI':'DALI','0-10V':'010V','Phase Dim':'PD','Non-Dimmable':'ND'};
  if(known[full])return known[full];
  return (full||'').split(/\s+/).map(w=>w[0]||'').join('').toUpperCase().slice(0,4)||'?'}
function catOrdControlsCellHtml(ci,vi,val,canRemove,total){
  const full=controlsFullSpecOf(val).trim();
  const key=ci+':'+vi;
  const customMode=CAT_ORD_CONTROLS_CUSTOM.has(key)||(full&&!CAT_ORD_CONTROLS_OPTIONS.includes(full));
  if(customMode)
    return catOrdValueRowWrap(ci,vi,total,
      '<textarea rows=1 style="width:110px;font-size:11px;resize:vertical;box-sizing:border-box" placeholder="e.g. Push Dim" oninput="onCatOrdValueInput('+ci+','+vi+',this.value)" onblur="onCatOrdControlsCustomBlur('+ci+','+vi+',this)">'+escHtml(full)+'</textarea>',
      canRemove);
  const opts='<option value="">— Select —</option>'+
    CAT_ORD_CONTROLS_OPTIONS.map(o=>'<option value="'+escHtml(o)+'"'+(full===o?' selected':'')+'>'+escHtml(o)+'</option>').join('')+
    '<option value="__empty__">Empty</option>'+
    '<option value="__custom__">Custom…</option>';
  return catOrdValueRowWrap(ci,vi,total,
    '<select style="width:110px;font-size:11px" onchange="onCatOrdControlsChange('+ci+','+vi+',this)">'+opts+'</select>',
    canRemove)}
function onCatOrdControlsChange(ci,vi,sel){
  const key=ci+':'+vi;
  if(sel.value==='__custom__'){
    CAT_ORD_CONTROLS_CUSTOM.add(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    return}
  if(sel.value==='__empty__'){
    CAT_ORD_CONTROLS_CUSTOM.delete(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    recomputeOrderingCodeExample();
    schedulePreview();
    return}
  CAT_ORD_CONTROLS_CUSTOM.delete(key);
  const fullSpec=sel.value;
  CAT_ORD_COLS[ci].values[vi]=controlsShortCode(fullSpec)+'\n'+fullSpec;
  recomputeOrderingCodeExample();
  schedulePreview()}
// Same "remember what was typed" pattern as CCT/Series/Spec Labels — a
// custom Controls value the user actually typed gets saved as a new preset,
// then drops back into dropdown mode showing what it just learned.
function onCatOrdControlsCustomBlur(ci,vi,el){
  const raw=el.value.trim();
  const key=ci+':'+vi;
  if(!raw){CAT_ORD_CONTROLS_CUSTOM.delete(key);CAT_ORD_COLS[ci].values[vi]='';renderCatOrdTable();onCatOrdValueBlur();return}
  const finish=()=>{
    CAT_ORD_CONTROLS_CUSTOM.delete(key);
    CAT_ORD_COLS[ci].values[vi]=controlsShortCode(raw)+'\n'+raw;
    renderCatOrdTable();
    onCatOrdValueBlur()};
  rememberCatControlsOption(raw).then(finish)}
// Input Voltage gets the same dropdown-of-presets-plus-Custom… treatment as
// CCT/Controls, stored as the same two-line "shortcode\nfull spec"
// convention — Input Voltage is a code-contributing category too
// (CAT_ORD_CODE_CATEGORIES above), and the real catalogue's own code uses
// the bare number ("24" for "24V DC"), not the full spec.
let CAT_ORD_VOLTAGE_OPTIONS=['12V DC','24V DC','48V DC','100-240V','110-120V','220-240V','380-415V'];
async function loadCatVoltageOptions(){
  const r=await fetch('/api/cat-voltage-options').then(r=>r.json());
  if(r.options)CAT_ORD_VOLTAGE_OPTIONS=r.options}
async function rememberCatVoltageOption(value){
  value=(value||'').trim();
  if(!value)return;
  // Reuses voltageShortCode — the same unit-token-stripping compare the cell
  // builder's own customMode detection already uses — so a manually-typed
  // value that the dropdown would already recognize as an existing preset
  // isn't remembered as a spurious duplicate (see rememberCatPowerOption).
  const norm=voltageShortCode(value);
  if(CAT_ORD_VOLTAGE_OPTIONS.some(o=>voltageShortCode(o)===norm))return;
  const r=await fetch('/api/cat-voltage-options-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({value})}).then(r=>r.json());
  if(r.options)CAT_ORD_VOLTAGE_OPTIONS=r.options}
let CAT_ORD_VOLTAGE_CUSTOM=new Set();
function isCatOrdVoltageColumn(label){return normSpecLabel(label)==='inputvoltage'}
function voltageFullSpecOf(val){
  const nl=(val||'').indexOf('\n');
  return nl>=0?val.slice(nl+1):(val||'')}
// The code half drops the unit ("220-240V" -> "220-240", "24V DC" -> "24")
// rather than an initials-shortener like Controls — voltage's own digits
// already read as a compact code in the real catalogue, no abbreviation
// needed.
function voltageShortCode(full){
  return (full||'').replace(/\s*(V\s*DC|V\s*AC|VDC|VAC|DC|AC|V)\s*$/i,'').trim()||'?'}
function catOrdVoltageCellHtml(ci,vi,val,canRemove,total){
  const full=voltageFullSpecOf(val).trim();
  const key=ci+':'+vi;
  // Compared via voltageShortCode (strips a trailing unit token) rather than
  // an exact string match — pre-existing data typed before this dropdown
  // existed is missing the "V"/"DC" unit entirely (e.g. "220-240" instead of
  // "220-240V"), and would otherwise never match its own preset.
  const normFull=voltageShortCode(full);
  const customMode=CAT_ORD_VOLTAGE_CUSTOM.has(key)||(full&&!CAT_ORD_VOLTAGE_OPTIONS.some(o=>voltageShortCode(o)===normFull));
  if(customMode)
    return catOrdValueRowWrap(ci,vi,total,
      '<textarea rows=1 style="width:110px;font-size:11px;resize:vertical;box-sizing:border-box" placeholder="e.g. 36V DC" oninput="onCatOrdValueInput('+ci+','+vi+',this.value)" onblur="onCatOrdVoltageCustomBlur('+ci+','+vi+',this)">'+escHtml(full)+'</textarea>',
      canRemove);
  const opts='<option value="">— Select —</option>'+
    CAT_ORD_VOLTAGE_OPTIONS.map(o=>'<option value="'+escHtml(o)+'"'+(voltageShortCode(o)===normFull?' selected':'')+'>'+escHtml(o)+'</option>').join('')+
    '<option value="__empty__">Empty</option>'+
    '<option value="__custom__">Custom…</option>';
  return catOrdValueRowWrap(ci,vi,total,
    '<select style="width:110px;font-size:11px" onchange="onCatOrdVoltageChange('+ci+','+vi+',this)">'+opts+'</select>',
    canRemove)}
function onCatOrdVoltageChange(ci,vi,sel){
  const key=ci+':'+vi;
  if(sel.value==='__custom__'){
    CAT_ORD_VOLTAGE_CUSTOM.add(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    return}
  if(sel.value==='__empty__'){
    CAT_ORD_VOLTAGE_CUSTOM.delete(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    recomputeOrderingCodeExample();
    schedulePreview();
    return}
  CAT_ORD_VOLTAGE_CUSTOM.delete(key);
  const fullSpec=sel.value;
  CAT_ORD_COLS[ci].values[vi]=voltageShortCode(fullSpec)+'\n'+fullSpec;
  renderCatOrdTable();
  recomputeOrderingCodeExample();
  schedulePreview()}
// Same "remember what was typed" pattern as CCT/Controls — a custom voltage
// value the user actually typed gets saved as a new preset, then drops back
// into dropdown mode showing what it just learned.
function onCatOrdVoltageCustomBlur(ci,vi,el){
  const raw=el.value.trim();
  const key=ci+':'+vi;
  if(!raw){CAT_ORD_VOLTAGE_CUSTOM.delete(key);CAT_ORD_COLS[ci].values[vi]='';renderCatOrdTable();onCatOrdValueBlur();return}
  const finish=()=>{
    CAT_ORD_VOLTAGE_CUSTOM.delete(key);
    CAT_ORD_COLS[ci].values[vi]=voltageShortCode(raw)+'\n'+raw;
    renderCatOrdTable();
    onCatOrdValueBlur()};
  rememberCatVoltageOption(raw).then(finish)}
// Power/Wattage gets the same dropdown-of-presets-plus-Custom… treatment,
// but unlike CCT/Controls/Voltage it stays a FLAT single-line value ("12W",
// not "12\n12W") — the real catalogue's own code keeps Power's unit
// ("...-9W-..."), so there's no separate shortcode worth splitting out, and
// keeping it flat means the existing numeric-sort/Lumen-derivation/Model No.
// sync logic (all written against a plain "12W" string) needs no changes.
let CAT_ORD_POWER_OPTIONS=Array.from({length:40},(_,i)=>(i+1)+'W');
async function loadCatPowerOptions(){
  const r=await fetch('/api/cat-power-options').then(r=>r.json());
  if(r.options)CAT_ORD_POWER_OPTIONS=r.options}
async function rememberCatPowerOption(value){
  value=(value||'').trim();
  if(!value)return;
  const norm=catOrdPowerNorm(value);
  // Normalized compare (not plain .includes()) — a custom-typed "9" glues to
  // "9\xa0W" (non-breaking space) via appendCatSpecSuffix, which would
  // otherwise never string-match the compact seeded "9W" and get remembered
  // as a spurious duplicate every time someone types instead of picking it.
  if(CAT_ORD_POWER_OPTIONS.some(o=>catOrdPowerNorm(o)===norm))return;
  const r=await fetch('/api/cat-power-options-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({value})}).then(r=>r.json());
  if(r.options)CAT_ORD_POWER_OPTIONS=r.options}
let CAT_ORD_POWER_CUSTOM=new Set();
// 5W/10W/15W/20W always show first in the Power dropdown, per explicit
// request — everything else (the rest of the 1-40W baseline, plus any
// custom wattage) follows in its existing order. Purely a display-order
// concern — doesn't change what's actually stored in CAT_ORD_POWER_OPTIONS.
const CAT_POWER_PINNED=['5W','10W','15W','20W'];
function catPowerOptionsOrdered(){
  const pinnedNorm=CAT_POWER_PINNED.map(catOrdPowerNorm);
  const rest=CAT_ORD_POWER_OPTIONS.filter(o=>!pinnedNorm.includes(catOrdPowerNorm(o)));
  return CAT_POWER_PINNED.concat(rest)}
function isCatOrdPowerColumn(label){return normSpecLabel(label)==='power'}
// Compares with all whitespace stripped (regular AND non-breaking — old
// values formatted by appendCatSpecSuffix glue the number+unit with a
// non-breaking space, e.g. "3\xa0W", while a freshly seeded/typed preset is
// the bare compact form "3W") so pre-existing data that's really just a
// known wattage still shows as a clean dropdown pick instead of always
// falling into custom mode over an invisible-character mismatch.
function catOrdPowerNorm(s){return (s||'').replace(/\s+/g,'')}
function catOrdPowerCellHtml(ci,vi,val,canRemove,total){
  const full=(val||'').trim();
  const normFull=catOrdPowerNorm(full);
  const key=ci+':'+vi;
  const customMode=CAT_ORD_POWER_CUSTOM.has(key)||(full&&!CAT_ORD_POWER_OPTIONS.some(o=>catOrdPowerNorm(o)===normFull));
  if(customMode)
    return catOrdValueRowWrap(ci,vi,total,
      '<textarea rows=1 style="width:110px;font-size:11px;resize:vertical;box-sizing:border-box" placeholder="e.g. 15W" oninput="onCatOrdValueInput('+ci+','+vi+',this.value)" onblur="onCatOrdPowerCustomBlur('+ci+','+vi+',this)">'+escHtml(full)+'</textarea>',
      canRemove);
  const opts='<option value="">— Select —</option>'+
    catPowerOptionsOrdered().map(o=>'<option value="'+escHtml(o)+'"'+(catOrdPowerNorm(o)===normFull?' selected':'')+'>'+escHtml(o)+'</option>').join('')+
    '<option value="__empty__">Empty</option>'+
    '<option value="__custom__">Custom…</option>';
  return catOrdValueRowWrap(ci,vi,total,
    '<select style="width:110px;font-size:11px" onchange="onCatOrdPowerChange('+ci+','+vi+',this)">'+opts+'</select>',
    canRemove)}
function onCatOrdPowerChange(ci,vi,sel){
  const key=ci+':'+vi;
  if(sel.value==='__custom__'){
    CAT_ORD_POWER_CUSTOM.add(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    return}
  if(sel.value==='__empty__'){
    CAT_ORD_POWER_CUSTOM.delete(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    recomputeOrderingCodeExample();
    syncSpecPowerFromOrdering(CAT_ORD_COLS[ci].values);
    schedulePreview();
    return}
  CAT_ORD_POWER_CUSTOM.delete(key);
  CAT_ORD_COLS[ci].values[vi]=sel.value;
  renderCatOrdTable();
  recomputeOrderingCodeExample();
  syncSpecPowerFromOrdering(CAT_ORD_COLS[ci].values);
  schedulePreview()}
// Same "remember what was typed" pattern as CCT/Controls/Voltage — a custom
// wattage the user actually typed gets saved as a new preset, then drops
// back into dropdown mode showing what it just learned.
function onCatOrdPowerCustomBlur(ci,vi,el){
  const key=ci+':'+vi;
  const formatted=appendCatSpecSuffix(el.value,'W');
  if(!formatted){CAT_ORD_POWER_CUSTOM.delete(key);CAT_ORD_COLS[ci].values[vi]='';renderCatOrdTable();syncSpecPowerFromOrdering(CAT_ORD_COLS[ci].values);onCatOrdValueBlur();return}
  const finish=()=>{
    CAT_ORD_POWER_CUSTOM.delete(key);
    CAT_ORD_COLS[ci].values[vi]=formatted;
    renderCatOrdTable();
    syncSpecPowerFromOrdering(CAT_ORD_COLS[ci].values);
    onCatOrdValueBlur()};
  rememberCatPowerOption(formatted).then(finish)}
// Technical Specifications' "Power" row and the Ordering Table's "Power"
// column are the same real-world spec, not two independent ones — both hold
// a full array of values and are kept as exact mirrors of each other (same
// values, same order) rather than Tech Spec only ever showing one of
// several real wattages. Neither side auto-sorts anymore (see
// sortCatOrdVariantsBy — sorting is now an explicit, on-demand, whole-row
// action, not something either side of this sync does silently), so
// whichever side just changed simply copies straight across; existing
// drafts saved before this feature existed are NOT reconciled on load —
// sync only takes effect from the next edit onward, so nothing gets
// silently overwritten.
function findCatOrdPowerCol(){return CAT_ORD_COLS.find(c=>isCatOrdPowerColumn(c.label))}
function findCatSpecPowerIndex(){return CAT_SPECS.findIndex(s=>normSpecLabel(s.label)==='power')}
function syncSpecPowerFromOrdering(values){
  const i=findCatSpecPowerIndex();
  if(i<0)return;
  CAT_SPECS[i].values=values.slice();
  renderCatSpecs()}
function syncOrderingPowerFromSpec(values){
  const c=findCatOrdPowerCol();
  if(!c)return;
  c.values=values.slice();
  renderCatOrdTable();
  recomputeOrderingCodeExample()}
// Beam Angle gets the same dropdown-of-presets-plus-Custom… treatment as
// Power, staying a flat single-line value ("100°", not "100\n100°") —
// catOrdCodeSegment already strips the ° and any whitespace from whichever
// column's first value is used as its code, so there's no shortcode half
// worth splitting out here either.
let CAT_ORD_BEAMANGLE_OPTIONS=['15°','24°','36°','38°','45°','60°','90°','100°','120°'];
async function loadCatBeamAngleOptions(){
  const r=await fetch('/api/cat-beamangle-options').then(r=>r.json());
  if(r.options)CAT_ORD_BEAMANGLE_OPTIONS=r.options}
async function rememberCatBeamAngleOption(value){
  value=(value||'').trim();
  if(!value)return;
  // Reuses catOrdBeamAngleNorm — the same whitespace-insensitive compare the
  // cell builder's own customMode detection already uses (see
  // rememberCatPowerOption for the original fix of this bug).
  const norm=catOrdBeamAngleNorm(value);
  if(CAT_ORD_BEAMANGLE_OPTIONS.some(o=>catOrdBeamAngleNorm(o)===norm))return;
  const r=await fetch('/api/cat-beamangle-options-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({value})}).then(r=>r.json());
  if(r.options)CAT_ORD_BEAMANGLE_OPTIONS=r.options}
let CAT_ORD_BEAMANGLE_CUSTOM=new Set();
function isCatOrdBeamAngleColumn(label){return normSpecLabel(label)==='beamangle'}
// Same whitespace-insensitive compare as Power (catOrdPowerNorm) — reused
// here under its own name rather than a shared helper, matching how each
// dropdown column in this file already owns its compare function.
function catOrdBeamAngleNorm(s){return (s||'').replace(/\s+/g,'')}
function catOrdBeamAngleCellHtml(ci,vi,val,canRemove,total){
  const full=(val||'').trim();
  const normFull=catOrdBeamAngleNorm(full);
  const key=ci+':'+vi;
  const customMode=CAT_ORD_BEAMANGLE_CUSTOM.has(key)||(full&&!CAT_ORD_BEAMANGLE_OPTIONS.some(o=>catOrdBeamAngleNorm(o)===normFull));
  if(customMode)
    return catOrdValueRowWrap(ci,vi,total,
      '<textarea rows=1 style="width:110px;font-size:11px;resize:vertical;box-sizing:border-box" placeholder="e.g. 100°" oninput="onCatOrdValueInput('+ci+','+vi+',this.value)" onblur="onCatOrdBeamAngleCustomBlur('+ci+','+vi+',this)">'+escHtml(full)+'</textarea>',
      canRemove);
  const opts='<option value="">— Select —</option>'+
    CAT_ORD_BEAMANGLE_OPTIONS.map(o=>'<option value="'+escHtml(o)+'"'+(catOrdBeamAngleNorm(o)===normFull?' selected':'')+'>'+escHtml(o)+'</option>').join('')+
    '<option value="__empty__">Empty</option>'+
    '<option value="__custom__">Custom…</option>';
  return catOrdValueRowWrap(ci,vi,total,
    '<select style="width:110px;font-size:11px" onchange="onCatOrdBeamAngleChange('+ci+','+vi+',this)">'+opts+'</select>',
    canRemove)}
function onCatOrdBeamAngleChange(ci,vi,sel){
  const key=ci+':'+vi;
  if(sel.value==='__custom__'){
    CAT_ORD_BEAMANGLE_CUSTOM.add(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    return}
  if(sel.value==='__empty__'){
    CAT_ORD_BEAMANGLE_CUSTOM.delete(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    recomputeOrderingCodeExample();
    schedulePreview();
    return}
  CAT_ORD_BEAMANGLE_CUSTOM.delete(key);
  CAT_ORD_COLS[ci].values[vi]=sel.value;
  renderCatOrdTable();
  recomputeOrderingCodeExample();
  schedulePreview()}
// Same "remember what was typed" pattern as Power/CCT/Controls/Voltage — a
// custom beam angle the user actually typed gets saved as a new preset,
// then drops back into dropdown mode showing what it just learned.
function onCatOrdBeamAngleCustomBlur(ci,vi,el){
  const key=ci+':'+vi;
  const formatted=appendCatSpecSuffix(el.value,'°');
  if(!formatted){CAT_ORD_BEAMANGLE_CUSTOM.delete(key);CAT_ORD_COLS[ci].values[vi]='';renderCatOrdTable();onCatOrdValueBlur();return}
  const finish=()=>{
    CAT_ORD_BEAMANGLE_CUSTOM.delete(key);
    CAT_ORD_COLS[ci].values[vi]=formatted;
    renderCatOrdTable();
    onCatOrdValueBlur()};
  rememberCatBeamAngleOption(formatted).then(finish)}
// Model No./Options/Cut Out are the last 3 free-text columns to get the
// same dropdown-of-remembered-values treatment — no curated seed data for
// any of these (unlike Beam Angle), so each starts empty and shows the raw
// input directly until the user has typed at least one value; only once
// something's actually been remembered does the dropdown appear, per the
// user's own framing ("remember what I fill in, AFTER keep it in the
// dropdown" — nothing to offer back before that first entry exists).
let CAT_ORD_MODELNO_OPTIONS=[];
async function loadCatModelNoOptions(){
  const r=await fetch('/api/cat-modelno-options').then(r=>r.json());
  if(r.options)CAT_ORD_MODELNO_OPTIONS=r.options}
async function rememberCatModelNoOption(value){
  value=(value||'').trim();
  if(!value||CAT_ORD_MODELNO_OPTIONS.includes(value))return;
  const r=await fetch('/api/cat-modelno-options-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({value})}).then(r=>r.json());
  if(r.options)CAT_ORD_MODELNO_OPTIONS=r.options}
let CAT_ORD_MODELNO_CUSTOM=new Set();
function isCatOrdModelNoColumn(label){return normSpecLabel(label)==='modelno.'}
function catOrdModelNoCellHtml(ci,vi,val,canRemove,total){
  const full=(val||'').trim();
  const key=ci+':'+vi;
  const customMode=!CAT_ORD_MODELNO_OPTIONS.length||CAT_ORD_MODELNO_CUSTOM.has(key)||(full&&!CAT_ORD_MODELNO_OPTIONS.includes(full));
  if(customMode)
    return catOrdValueRowWrap(ci,vi,total,
      '<textarea rows=1 style="width:110px;font-size:11px;resize:vertical;box-sizing:border-box" placeholder="e.g. SLAUR" oninput="onCatOrdValueInput('+ci+','+vi+',this.value)" onblur="onCatOrdModelNoCustomBlur('+ci+','+vi+',this)">'+escHtml(full)+'</textarea>',
      canRemove);
  const opts='<option value="">— Select —</option>'+
    CAT_ORD_MODELNO_OPTIONS.map(o=>'<option value="'+escHtml(o)+'"'+(o===full?' selected':'')+'>'+escHtml(o)+'</option>').join('')+
    '<option value="__empty__">Empty</option>'+
    '<option value="__custom__">Custom…</option>';
  return catOrdValueRowWrap(ci,vi,total,
    '<select style="width:110px;font-size:11px" onchange="onCatOrdModelNoChange('+ci+','+vi+',this)">'+opts+'</select>',
    canRemove)}
// Explicitly marking Model No. Empty for one specific variant has to stick
// — otherwise autoFillModelNoDown (which copies the row above's Model No.
// into any blank-but-has-other-data row) would immediately overwrite the
// user's own deliberate choice right back to non-empty on the next render.
let CAT_ORD_MODELNO_EXPLICIT_EMPTY=new Set();
function onCatOrdModelNoChange(ci,vi,sel){
  const key=ci+':'+vi;
  if(sel.value==='__custom__'){
    CAT_ORD_MODELNO_CUSTOM.add(key);
    CAT_ORD_MODELNO_EXPLICIT_EMPTY.delete(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    return}
  if(sel.value==='__empty__'){
    CAT_ORD_MODELNO_CUSTOM.delete(key);
    CAT_ORD_MODELNO_EXPLICIT_EMPTY.add(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    recomputeOrderingCodeExample();
    schedulePreview();
    return}
  CAT_ORD_MODELNO_CUSTOM.delete(key);
  CAT_ORD_MODELNO_EXPLICIT_EMPTY.delete(key);
  CAT_ORD_COLS[ci].values[vi]=sel.value;
  renderCatOrdTable();
  recomputeOrderingCodeExample();
  schedulePreview()}
function onCatOrdModelNoCustomBlur(ci,vi,el){
  const key=ci+':'+vi;
  const value=el.value.trim();
  if(!value){CAT_ORD_MODELNO_CUSTOM.delete(key);CAT_ORD_COLS[ci].values[vi]='';renderCatOrdTable();onCatOrdValueBlur();return}
  const finish=()=>{
    CAT_ORD_MODELNO_CUSTOM.delete(key);
    CAT_ORD_COLS[ci].values[vi]=value;
    renderCatOrdTable();
    onCatOrdValueBlur()};
  rememberCatModelNoOption(value).then(finish)}
let CAT_ORD_OPTIONS_OPTIONS=[];
async function loadCatOptionsOptions(){
  const r=await fetch('/api/cat-options-options').then(r=>r.json());
  if(r.options)CAT_ORD_OPTIONS_OPTIONS=r.options}
async function rememberCatOptionsOption(value){
  value=(value||'').trim();
  if(!value||CAT_ORD_OPTIONS_OPTIONS.includes(value))return;
  const r=await fetch('/api/cat-options-options-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({value})}).then(r=>r.json());
  if(r.options)CAT_ORD_OPTIONS_OPTIONS=r.options}
let CAT_ORD_OPTIONS_CUSTOM=new Set();
function isCatOrdOptionsColumn(label){return normSpecLabel(label)==='options'}
function catOrdOptionsCellHtml(ci,vi,val,canRemove,total){
  const full=(val||'').trim();
  const key=ci+':'+vi;
  const customMode=!CAT_ORD_OPTIONS_OPTIONS.length||CAT_ORD_OPTIONS_CUSTOM.has(key)||(full&&!CAT_ORD_OPTIONS_OPTIONS.includes(full));
  if(customMode)
    return catOrdValueRowWrap(ci,vi,total,
      '<textarea rows=1 style="width:110px;font-size:11px;resize:vertical;box-sizing:border-box" placeholder="e.g. Emergency Kit" oninput="onCatOrdValueInput('+ci+','+vi+',this.value)" onblur="onCatOrdOptionsCustomBlur('+ci+','+vi+',this)">'+escHtml(full)+'</textarea>',
      canRemove);
  const opts='<option value="">— Select —</option>'+
    CAT_ORD_OPTIONS_OPTIONS.map(o=>'<option value="'+escHtml(o)+'"'+(o===full?' selected':'')+'>'+escHtml(o)+'</option>').join('')+
    '<option value="__empty__">Empty</option>'+
    '<option value="__custom__">Custom…</option>';
  return catOrdValueRowWrap(ci,vi,total,
    '<select style="width:110px;font-size:11px" onchange="onCatOrdOptionsChange('+ci+','+vi+',this)">'+opts+'</select>',
    canRemove)}
function onCatOrdOptionsChange(ci,vi,sel){
  const key=ci+':'+vi;
  if(sel.value==='__custom__'){
    CAT_ORD_OPTIONS_CUSTOM.add(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    return}
  if(sel.value==='__empty__'){
    CAT_ORD_OPTIONS_CUSTOM.delete(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    recomputeOrderingCodeExample();
    schedulePreview();
    return}
  CAT_ORD_OPTIONS_CUSTOM.delete(key);
  CAT_ORD_COLS[ci].values[vi]=sel.value;
  renderCatOrdTable();
  recomputeOrderingCodeExample();
  schedulePreview()}
function onCatOrdOptionsCustomBlur(ci,vi,el){
  const key=ci+':'+vi;
  const value=el.value.trim();
  if(!value){CAT_ORD_OPTIONS_CUSTOM.delete(key);CAT_ORD_COLS[ci].values[vi]='';renderCatOrdTable();onCatOrdValueBlur();return}
  const finish=()=>{
    CAT_ORD_OPTIONS_CUSTOM.delete(key);
    CAT_ORD_COLS[ci].values[vi]=value;
    renderCatOrdTable();
    onCatOrdValueBlur()};
  rememberCatOptionsOption(value).then(finish)}
let CAT_ORD_CUTOUT_OPTIONS=[];
async function loadCatCutOutOptions(){
  const r=await fetch('/api/cat-cutout-options').then(r=>r.json());
  if(r.options)CAT_ORD_CUTOUT_OPTIONS=r.options}
async function rememberCatCutOutOption(value){
  value=(value||'').trim();
  if(!value||CAT_ORD_CUTOUT_OPTIONS.includes(value))return;
  const r=await fetch('/api/cat-cutout-options-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({value})}).then(r=>r.json());
  if(r.options)CAT_ORD_CUTOUT_OPTIONS=r.options}
let CAT_ORD_CUTOUT_CUSTOM=new Set();
function isCatOrdCutOutColumn(label){return normSpecLabel(label)==='cutout'}
function catOrdCutOutNorm(s){return (s||'').replace(/\s+/g,'')}
function catOrdCutOutCellHtml(ci,vi,val,canRemove,total){
  const full=(val||'').trim();
  const normFull=catOrdCutOutNorm(full);
  const key=ci+':'+vi;
  const customMode=!CAT_ORD_CUTOUT_OPTIONS.length||CAT_ORD_CUTOUT_CUSTOM.has(key)||(full&&!CAT_ORD_CUTOUT_OPTIONS.some(o=>catOrdCutOutNorm(o)===normFull));
  if(customMode)
    return catOrdValueRowWrap(ci,vi,total,
      '<textarea rows=1 style="width:110px;font-size:11px;resize:vertical;box-sizing:border-box" placeholder="e.g. 150mm" oninput="onCatOrdValueInput('+ci+','+vi+',this.value)" onblur="onCatOrdCutOutCustomBlur('+ci+','+vi+',this)">'+escHtml(full)+'</textarea>',
      canRemove);
  const opts='<option value="">— Select —</option>'+
    CAT_ORD_CUTOUT_OPTIONS.map(o=>'<option value="'+escHtml(o)+'"'+(catOrdCutOutNorm(o)===normFull?' selected':'')+'>'+escHtml(o)+'</option>').join('')+
    '<option value="__empty__">Empty</option>'+
    '<option value="__custom__">Custom…</option>';
  return catOrdValueRowWrap(ci,vi,total,
    '<select style="width:110px;font-size:11px" onchange="onCatOrdCutOutChange('+ci+','+vi+',this)">'+opts+'</select>',
    canRemove)}
function onCatOrdCutOutChange(ci,vi,sel){
  const key=ci+':'+vi;
  if(sel.value==='__custom__'){
    CAT_ORD_CUTOUT_CUSTOM.add(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    return}
  if(sel.value==='__empty__'){
    CAT_ORD_CUTOUT_CUSTOM.delete(key);
    CAT_ORD_COLS[ci].values[vi]='';
    renderCatOrdTable();
    recomputeOrderingCodeExample();
    schedulePreview();
    return}
  CAT_ORD_CUTOUT_CUSTOM.delete(key);
  CAT_ORD_COLS[ci].values[vi]=sel.value;
  renderCatOrdTable();
  recomputeOrderingCodeExample();
  schedulePreview()}
function onCatOrdCutOutCustomBlur(ci,vi,el){
  const key=ci+':'+vi;
  const formatted=appendCatSpecSuffix(el.value,'mm');
  if(!formatted){CAT_ORD_CUTOUT_CUSTOM.delete(key);CAT_ORD_COLS[ci].values[vi]='';renderCatOrdTable();onCatOrdValueBlur();return}
  const finish=()=>{
    CAT_ORD_CUTOUT_CUSTOM.delete(key);
    CAT_ORD_COLS[ci].values[vi]=formatted;
    renderCatOrdTable();
    onCatOrdValueBlur()};
  rememberCatCutOutOption(formatted).then(finish)}

// ---------------------------------------------------------------- Fill Standard Information
// One click fills a whole field's worth of rows from that field's own
// remembered preset list (e.g. every CCT the catalogue offers, one per
// row) instead of adding N rows and picking each value by hand. Which
// fields it touches is itself user-editable (see CAT_STANDARD_FILL_FIELDS/
// /api/cat-standard-fill-fields) — defaults to CCT+Controls per explicit
// request, but "programmable" means any of the other dropdown-backed
// fields can be added via the Configure popover without needing new code.
// Each entry mirrors that field's own cell builder's value convention
// (two-line "shortcode\nfull spec" for CCT/Controls/Voltage/Size, flat for
// everything else) so a filled row looks identical to one the user picked
// by hand from that same dropdown.
const CAT_STANDARD_FILL_DEFS={
  cct:{label:'CCT',match:c=>isCatOrdCctColumn(c.label),options:()=>CAT_ORD_CCT_OPTIONS,
    toValue:o=>cctShortCode(o+' K')+'\n'+o+' K'},
  controls:{label:'Controls',match:c=>isCatOrdControlsColumn(c.label),options:()=>CAT_ORD_CONTROLS_OPTIONS,
    toValue:o=>controlsShortCode(o)+'\n'+o},
  voltage:{label:'Input Voltage',match:c=>isCatOrdVoltageColumn(c.label),options:()=>CAT_ORD_VOLTAGE_OPTIONS,
    toValue:o=>voltageShortCode(o)+'\n'+o},
  power:{label:'Power',match:c=>isCatOrdPowerColumn(c.label),options:()=>catPowerOptionsOrdered(),
    toValue:o=>o},
  beamangle:{label:'Beam Angle',match:c=>isCatOrdBeamAngleColumn(c.label),options:()=>CAT_ORD_BEAMANGLE_OPTIONS,
    toValue:o=>o},
  modelno:{label:'Model No.',match:c=>isCatOrdModelNoColumn(c.label),options:()=>CAT_ORD_MODELNO_OPTIONS,
    toValue:o=>o},
  options:{label:'Options',match:c=>isCatOrdOptionsColumn(c.label),options:()=>CAT_ORD_OPTIONS_OPTIONS,
    toValue:o=>o},
  cutout:{label:'Cut Out',match:c=>isCatOrdCutOutColumn(c.label),options:()=>CAT_ORD_CUTOUT_OPTIONS,
    toValue:o=>o},
  // Plain size text, no "D{n}\n" prefix baked in here — fillCatOrdStandardInfo
  // below calls recomputeCatOrdSizeDNumbers() right after filling every
  // field, same single source of truth for D-numbers as a manually-edited
  // cell uses (see that function's own comment), rather than a second,
  // separate "assign D{idx+1} from this list's own order" rule living here
  // that could drift from it.
  size:{label:'Size',match:c=>isCatOrdSizeColumn(c.label),options:()=>CAT_SIZE_INDEX,
    toValue:o=>o},
};
let CAT_STANDARD_FILL_FIELDS=['cct','controls'];
// Display/processing order for ALL 9 possible fields, selected or not —
// drives both the Configure popover's row order and the order
// fillCatOrdStandardInfo() walks the selected subset in. Defaults to
// CAT_STANDARD_FILL_DEFS' own key order until the user rearranges it.
let CAT_STANDARD_FILL_ORDER=Object.keys(CAT_STANDARD_FILL_DEFS);
async function loadCatStandardFillFields(){
  const r=await fetch('/api/cat-standard-fill-fields').then(r=>r.json());
  if(r.fields)CAT_STANDARD_FILL_FIELDS=r.fields;
  if(r.order)CAT_STANDARD_FILL_ORDER=r.order}
// One level deeper than CAT_STANDARD_FILL_ORDER above (which fields fill,
// and in what order) — for each field, which SUBSET of its own saved
// preset values actually get used when filling, and in what sequence (e.g.
// only 3 of 7 saved CCTs, Non-Dimmable first). A key absent here means "use
// the field's full current preset list" — see catStandardFillEffectiveOptions,
// the one seam that keeps every untouched field behaving exactly as before
// this existed.
let CAT_STANDARD_FILL_VALUES={};
async function loadCatStandardFillValues(){
  const r=await fetch('/api/cat-standard-fill-values').then(r=>r.json());
  if(r.values)CAT_STANDARD_FILL_VALUES=r.values}
function catStandardFillEffectiveOptions(key){
  const custom=CAT_STANDARD_FILL_VALUES[key];
  return (custom&&custom.length)?custom:CAT_STANDARD_FILL_DEFS[key].options()}
// Overwrites (doesn't merge with) whatever's already in each configured
// field's own column — this is a quick-start shortcut for a table that's
// still mostly blank, not a background sync; re-running it after manual
// edits is expected to reset that field back to the full standard list.
// Any OTHER lockstepped column (Model No. included, if not itself one of
// the configured fields) gets padded to match via renderCatOrdTable's own
// normalizeCatOrdLockstep/autoFillModelNoDown — a Model No. typed once on
// row 1 automatically copies down to every new row this creates.
async function fillCatOrdStandardInfo(){
  const active=CAT_STANDARD_FILL_ORDER.filter(k=>CAT_STANDARD_FILL_FIELDS.includes(k)&&CAT_STANDARD_FILL_DEFS[k]);
  if(!active.length){toast('Nothing configured yet — click Configure to choose which fields to fill');return}
  let filled=0,missing=[];
  active.forEach(key=>{
    const def=CAT_STANDARD_FILL_DEFS[key];
    const col=CAT_ORD_COLS.find(def.match);
    if(!col){missing.push(def.label);return}
    const opts=catStandardFillEffectiveOptions(key);
    if(!opts.length){missing.push(def.label+' (no presets saved yet)');return}
    col.values=opts.map((o,idx)=>def.toValue(o,idx));
    filled++});
  if(!filled){toast('None of the configured fields are in this table yet: '+missing.join(', '));return}
  // Size's own D-numbers always come from recomputeCatOrdSizeDNumbers, never
  // baked directly into toValue above — see that function's own comment.
  recomputeCatOrdSizeDNumbers();
  renderCatOrdTable();recomputeOrderingCodeExample();schedulePreview();
  toast('Filled standard information for '+filled+' field'+(filled>1?'s':'')+(missing.length?' — skipped: '+missing.join(', '):''))}
// Draggable (HTML5 drag-and-drop, same dragRowStart/Over/Leave/End +
// dragReorder plumbing every other reorderable list in this app uses —
// see the Category Order card above) with the ▲▼ buttons kept alongside
// as a precise-single-step fallback, same convention as everywhere else.
function renderCatStandardFillMenu(){
  const menu=$('filemenu');
  menu.innerHTML='<div class=fmtitle>Fill Standard Information uses</div>'+
    CAT_STANDARD_FILL_ORDER.map((key,i)=>{
      const def=CAT_STANDARD_FILL_DEFS[key];
      if(!def)return '';
      const on=CAT_STANDARD_FILL_FIELDS.includes(key);
      const custom=CAT_STANDARD_FILL_VALUES[key];
      // Customized = the user has picked a specific subset/order for this
      // field's own values (see catStandardFillEffectiveOptions) — shown as
      // a small "N/M" badge so the popover itself communicates which
      // fields deviate from "use everything saved," without needing to
      // open the values modal to find out.
      const badge=(custom&&custom.length)?'<span style="font-size:9px;color:var(--muted);flex-shrink:0">'+custom.length+'/'+def.options().length+'</span>':'';
      return '<div class="fmi dragrow" draggable=true ondragstart="dragRowStart(event,\''+key+'\')" ondragover="dragRowOver(event)" ondragleave="dragRowLeave(event)" ondrop="catStandardFillDrop(event,\''+key+'\')" ondragend="dragRowEnd(event)" style="display:flex;align-items:center;gap:6px">'+
        '<span class=draghandle title="Drag to reorder" onclick="event.stopPropagation()">⠿</span>'+
        '<span onclick="toggleCatStandardFillField(event,\''+key+'\')" style="flex:1;display:flex;align-items:center;gap:8px;cursor:pointer;min-width:0">'+
        '<span style="width:14px;text-align:center;color:'+(on?'var(--info)':'var(--muted)')+'">'+(on?'✓':'—')+'</span>'+
        '<span style="overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+def.label+'</span></span>'+
        badge+
        '<button type=button class=ordmovebtn onclick="event.stopPropagation();openCatStandardFillValuesModal(event,\''+key+'\')" title="Choose which '+def.label+' values fill, and in what order">⚙</button>'+
        '<span style="display:flex;flex-direction:column">'+
        '<button type=button class=ordmovebtn'+(i===0?' disabled':'')+' onclick="moveCatStandardFillKey(event,\''+key+'\',\'up\')" title="Move up">▲</button>'+
        '<button type=button class=ordmovebtn'+(i===CAT_STANDARD_FILL_ORDER.length-1?' disabled':'')+' onclick="moveCatStandardFillKey(event,\''+key+'\',\'down\')" title="Move down">▼</button>'+
        '</span></div>'}).join('')}
async function catStandardFillDrop(e,targetKey){
  e.preventDefault();e.stopPropagation();
  e.currentTarget.classList.remove('dragover-top','dragover-bottom');
  await dragReorder(targetKey,CAT_STANDARD_FILL_ORDER,catStandardFillMove)}
async function catStandardFillMove(key,direction){
  const r=await fetch('/api/cat-standard-fill-order-move',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({key,direction})}).then(r=>r.json());
  if(r.order){CAT_STANDARD_FILL_ORDER=r.order;renderCatStandardFillMenu()}}
// stopPropagation so the ▲▼ buttons don't close the popover on the same
// click (same fix as toggleCatStandardFillField below) — drag-and-drop
// never fires a document click in the first place, so catStandardFillDrop
// needs no such guard of its own.
function moveCatStandardFillKey(ev,key,direction){
  ev.stopPropagation();
  return catStandardFillMove(key,direction)}
function openCatStandardFillConfig(ev){
  ev.stopPropagation();
  const rect=ev.currentTarget.getBoundingClientRect();
  renderCatStandardFillMenu();
  const menu=$('filemenu');
  menu.style.display='block';
  const w=menu.offsetWidth||220,h=menu.offsetHeight||280;
  let x=rect.left,y=rect.bottom+4;
  if(x+w>window.innerWidth-8)x=window.innerWidth-w-8;
  if(y+h>window.innerHeight-8)y=window.innerHeight-h-8;
  menu.style.left=x+'px';menu.style.top=y+'px';
  setTimeout(()=>document.addEventListener('click',closeFileMenu,{once:true}),0)}
// stopPropagation so ticking several boxes in a row doesn't close the
// popover after the first click (same fix as the file-menu Delete button
// — the shared #filemenu closes itself on the next document click).
async function toggleCatStandardFillField(ev,key){
  ev.stopPropagation();
  const idx=CAT_STANDARD_FILL_FIELDS.indexOf(key);
  if(idx>=0)CAT_STANDARD_FILL_FIELDS.splice(idx,1);
  else CAT_STANDARD_FILL_FIELDS.push(key);
  renderCatStandardFillMenu();
  const r=await fetch('/api/cat-standard-fill-fields',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({fields:CAT_STANDARD_FILL_FIELDS})}).then(r=>r.json());
  if(r.fields)CAT_STANDARD_FILL_FIELDS=r.fields}

// ---------------------------------------------------------------- Fill Standard Information — per-field value subset + order
// One level deeper than the field-level order above: for one field (e.g.
// CCT), which of its own saved preset VALUES actually get used when
// filling, and in what sequence. Lives in its own modal (.clientmodal, same
// pattern as #catbadgesmodal) rather than nested inside the #filemenu
// popover — that popover's sizing assumes one flat, short content block,
// and a field like Power can have 40+ saved presets to choose from.
function openCatStandardFillValuesModal(ev,key){
  ev.stopPropagation();
  closeFileMenu();
  $('catstdfillvaluesmodal').classList.remove('hide');
  renderCatStandardFillValuesModal(key)}
function closeCatStandardFillValuesModal(){$('catstdfillvaluesmodal').classList.add('hide')}
function renderCatStandardFillValuesModal(key){
  const def=CAT_STANDARD_FILL_DEFS[key];
  $('catstdfillvalues-title').textContent=def.label+' — values to fill';
  const all=def.options();
  const touched=CAT_STANDARD_FILL_VALUES[key]!==undefined;
  const selected=(touched?CAT_STANDARD_FILL_VALUES[key]:all).slice();
  const selSet=new Set(selected);
  const esc=v=>String(v).replace(/'/g,"\\'").replace(/\\/g,'\\\\');
  const chosenRows=selected.filter(v=>all.includes(v)).map((v,i)=>
    '<div class="fmi dragrow" draggable=true ondragstart="dragRowStart(event,\''+esc(v)+'\')" ondragover="dragRowOver(event)" ondragleave="dragRowLeave(event)" ondrop="catStandardFillValueDrop(event,\''+key+'\',\''+esc(v)+'\')" ondragend="dragRowEnd(event)" style="display:flex;align-items:center;gap:6px">'+
      '<span class=draghandle title="Drag to reorder">⠿</span>'+
      '<label style="flex:1;display:flex;align-items:center;gap:8px;cursor:pointer;min-width:0">'+
        '<input type=checkbox checked onchange="toggleCatStandardFillValue(\''+key+'\',\''+esc(v)+'\')">'+
        '<span style="overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+escHtml(v)+'</span></label>'+
      '<span style="display:flex;flex-direction:column">'+
      '<button type=button class=ordmovebtn'+(i===0?' disabled':'')+' onclick="moveCatStandardFillValue(\''+key+'\',\''+esc(v)+'\',\'up\')" title="Move up">▲</button>'+
      '<button type=button class=ordmovebtn'+(i===selected.length-1?' disabled':'')+' onclick="moveCatStandardFillValue(\''+key+'\',\''+esc(v)+'\',\'down\')" title="Move down">▼</button>'+
      '</span></div>').join('');
  const rest=all.filter(v=>!selSet.has(v));
  const restRows=rest.map(v=>
    '<div class=fmi style="display:flex;align-items:center;gap:6px">'+
      '<label style="flex:1;display:flex;align-items:center;gap:8px;cursor:pointer;min-width:0">'+
        '<input type=checkbox onchange="toggleCatStandardFillValue(\''+key+'\',\''+esc(v)+'\')">'+
        '<span style="overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:var(--muted)">'+escHtml(v)+'</span></label></div>').join('');
  $('catstdfillvalues-rows').innerHTML=
    '<button type=button class=btn style="width:100%;margin-bottom:10px"'+(touched?'':' disabled')+' onclick="resetCatStandardFillValues(\''+key+'\')">Use full list ('+all.length+')</button>'+
    (chosenRows||'<p class="muted" style="font-size:11px;margin:0 0 6px">Nothing selected — this field is skipped when filling.</p>')+
    (rest.length?'<div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.04em;color:var(--muted);margin:12px 0 4px">Not included</div>'+restRows:'')}
// An untouched field has no entry in CAT_STANDARD_FILL_VALUES at all
// (meaning "use the full list" — see catStandardFillEffectiveOptions). The
// first toggle/move on a field seeds its current full preset list into
// place first, same lazy-materialize-then-edit convention used elsewhere
// in this app (e.g. the Full Catalog Builder's Index Order) — without it,
// an index/value lookup against an untouched field would have nothing real
// on the backend to operate on yet.
async function catStandardFillEnsureMaterialized(key){
  if(CAT_STANDARD_FILL_VALUES[key]!==undefined)return;
  const full=CAT_STANDARD_FILL_DEFS[key].options().slice();
  const r=await fetch('/api/cat-standard-fill-values',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({key,values:full})}).then(r=>r.json());
  if(r.values)CAT_STANDARD_FILL_VALUES[key]=r.values}
async function toggleCatStandardFillValue(key,value){
  await catStandardFillEnsureMaterialized(key);
  const current=CAT_STANDARD_FILL_VALUES[key].slice();
  const idx=current.indexOf(value);
  if(idx>=0)current.splice(idx,1);else current.push(value);
  const r=await fetch('/api/cat-standard-fill-values',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({key,values:current})}).then(r=>r.json());
  if(r.values)CAT_STANDARD_FILL_VALUES[key]=r.values;
  renderCatStandardFillValuesModal(key);renderCatStandardFillMenu()}
async function moveCatStandardFillValue(key,value,direction){
  await catStandardFillEnsureMaterialized(key);
  const index=CAT_STANDARD_FILL_VALUES[key].indexOf(value);
  if(index<0)return;
  const r=await fetch('/api/cat-standard-fill-values-move',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({key,index,direction})}).then(r=>r.json());
  if(r.values){CAT_STANDARD_FILL_VALUES[key]=r.values;renderCatStandardFillValuesModal(key);renderCatStandardFillMenu()}}
async function catStandardFillValueDrop(e,key,targetValue){
  e.preventDefault();e.stopPropagation();
  e.currentTarget.classList.remove('dragover-top','dragover-bottom');
  await catStandardFillEnsureMaterialized(key);
  await dragReorder(targetValue,CAT_STANDARD_FILL_VALUES[key],(v,dir)=>moveCatStandardFillValue(key,v,dir))}
async function resetCatStandardFillValues(key){
  const r=await fetch('/api/cat-standard-fill-values-reset',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({key})}).then(r=>r.json());
  if(r.ok){delete CAT_STANDARD_FILL_VALUES[key];renderCatStandardFillValuesModal(key);renderCatStandardFillMenu()}}

function catOrdValueInput(ci,vi,val,readonly,canRemove,label,total){
  if(readonly){
    const lines=(val||'').split('\n');
    return '<div style="min-width:92px;padding:5px 6px;font-size:10.5px;background:var(--surface-2);border:1px solid var(--border);border-radius:4px">'+
      '<div style="font-weight:700">'+escHtml(lines[0]||'—')+'</div>'+
      lines.slice(1).map(l=>'<div style="color:var(--muted)">'+escHtml(l)+'</div>').join('')+
    '</div>'}
  if(isCatOrdModelNoColumn(label))return catOrdModelNoCellHtml(ci,vi,val,canRemove,total);
  if(isCatOrdCctColumn(label))return catOrdCctCellHtml(ci,vi,val,canRemove,total);
  if(isCatOrdSizeColumn(label))return catOrdSizeCellHtml(ci,vi,val,canRemove,total);
  if(isCatOrdControlsColumn(label))return catOrdControlsCellHtml(ci,vi,val,canRemove,total);
  if(isCatOrdVoltageColumn(label))return catOrdVoltageCellHtml(ci,vi,val,canRemove,total);
  if(isCatOrdPowerColumn(label))return catOrdPowerCellHtml(ci,vi,val,canRemove,total);
  if(isCatOrdBeamAngleColumn(label))return catOrdBeamAngleCellHtml(ci,vi,val,canRemove,total);
  if(isCatOrdCutOutColumn(label))return catOrdCutOutCellHtml(ci,vi,val,canRemove,total);
  if(isCatOrdOptionsColumn(label))return catOrdOptionsCellHtml(ci,vi,val,canRemove,total);
  return catOrdValueRowWrap(ci,vi,total,
    '<textarea rows=1 style="width:110px;font-size:11px;resize:vertical;box-sizing:border-box" placeholder="'+escHtml(catOrdPlaceholderFor(label))+'" oninput="onCatOrdValueInput('+ci+','+vi+',this.value)" onblur="onCatOrdValueUnitBlur('+ci+','+vi+',this)">'+escHtml(val||'')+'</textarea>',
    canRemove)}
// Icon + short-description maps for the field pills below (see
// renderCatOrdTable's fieldsHtml), one per known field type plus a
// "generic" fallback for anything custom-named (this app has no icon
// library — same inline-stroke-SVG convention as the nav rail tiles and
// EMPTY_ICON_* above: viewBox 0 0 24, stroke=currentColor, path-only).
// Scoped ONLY to this one pill grid, per explicit request — not a general
// "icon-ify everything" change.
const CAT_ORD_FIELD_ICON={
  cct:'<path d="M12 2v10"/><circle cx="12" cy="17" r="4"/>',
  controls:'<rect x="4" y="10" width="16" height="8" rx="2"/><circle cx="9" cy="14" r="1.5"/><circle cx="15" cy="14" r="1.5"/>',
  voltage:'<path d="M13 2 4 14h6l-1 8 9-12h-6z"/>',
  power:'<path d="M13 2 4 14h6l-1 8 9-12h-6z" fill="currentColor" stroke="none"/>',
  beamangle:'<path d="M12 3v18"/><path d="M12 3 5 19"/><path d="M12 3 19 19"/>',
  modelno:'<path d="M20.5 7.3 12 2 3.5 7.3v9.4L12 22l8.5-5.3Z"/><path d="M12 22V12"/><path d="m3.5 7.3 8.5 5 8.5-5"/>',
  options:'<path d="M9 3H5a2 2 0 0 0-2 2v4"/><path d="M15 3h4a2 2 0 0 1 2 2v4"/><path d="M9 21H5a2 2 0 0 1-2-2v-4"/><path d="M15 21h4a2 2 0 0 0 2-2v-4"/>',
  cutout:'<rect x="4" y="4" width="16" height="16" rx="2" stroke-dasharray="3 3"/>',
  size:'<path d="M3 17 17 3"/><path d="M3 8V3h5"/><path d="M16 21h5v-5"/>',
  finish:'<circle cx="12" cy="12" r="9"/><path d="M12 3a9 9 0 0 0 0 18c1.5 0 2-1 2-2s-.5-1.5-.5-2.5S14 12 15.5 12 18 10.5 18 9a9 9 0 0 0-6-6Z"/>',
  lumen:'<circle cx="12" cy="12" r="4"/><path d="M12 2v2M12 20v2M4.9 4.9l1.4 1.4M17.7 17.7l1.4 1.4M2 12h2M20 12h2M4.9 19.1l1.4-1.4M17.7 6.3l1.4-1.4"/>',
  generic:'<path d="M4 4h16v16H4z"/><path d="M4 9h16"/>',
};
const CAT_ORD_FIELD_DESC={
  cct:'Light color temperature, in Kelvin',
  controls:'Dimming or control protocol used',
  voltage:'Electrical input voltage rating',
  power:'Wattage draw of this variant',
  beamangle:'Spread angle of the light',
  modelno:'Ordering code for this variant',
  options:'Extra selectable accessories or features',
  cutout:'Ceiling or surface cutout dimension',
  size:'Physical dimensions of the fixture',
  finish:'Computed list of finish colors',
  lumen:'Computed lumen output per wattage',
  generic:'Custom field you named yourself',
};
function catOrdFieldIconKey(c){
  const label=(c.label||'').trim();
  if(label==='Finish Options')return 'finish';
  if(label==='Lumen')return 'lumen';
  if(isCatOrdCctColumn(label))return 'cct';
  if(isCatOrdControlsColumn(label))return 'controls';
  if(isCatOrdVoltageColumn(label))return 'voltage';
  if(isCatOrdPowerColumn(label))return 'power';
  if(isCatOrdBeamAngleColumn(label))return 'beamangle';
  if(isCatOrdModelNoColumn(label))return 'modelno';
  if(isCatOrdOptionsColumn(label))return 'options';
  if(isCatOrdCutOutColumn(label))return 'cutout';
  if(isCatOrdSizeColumn(label))return 'size';
  return 'generic'}
// Row-major: a "field" chips strip up top for managing which fields exist
// (add/rename/remove/reorder), then one card per Variant listing every
// field's value at that position together — instead of the old column-
// major layout (one horizontal row per field, each an independent stack of
// values with no visual link to any other field's boxes). Mirrors how the
// PDF already reads this data (html_engine.render_datasheet_pdf transposes
// column[c].values[i] into row i for every column) so the form now shows
// the same variant-by-variant correspondence while editing, not just in
// the generated document. None of the 9 bespoke per-field cell builders
// (catOrd<X>CellHtml/onCatOrd<X>Change/onCatOrd<X>CustomBlur) change at
// all — only this loop and the canRemove=false,total=1 forced on every
// call, since a single cell no longer shows its own move/remove chrome
// (that's one control per Variant card instead, see
// moveCatOrdVariant/removeCatOrdVariant).
function renderCatOrdTable(){
  normalizeCatOrdLockstep();
  recomputeCatSpecialOrdColumns();
  autoFillModelNoDown();
  recomputeOrderingCodeExample();
  if(!CAT_ORD_COLS.length){
    $('cat-ord-table').innerHTML='<button type=button class=btn style="width:100%" onclick="openCatOrdColMenu(event)">+ Add Field</button>';
    renderCatOrdWidthEditor();
    renderCatOrdAlignRowsBtn();
    return}
  const n=catOrdVariantCount();
  // Variant count can shrink out from under the pointer (remove, Import-
  // from-PDF replacing the whole table) — always land on a real variant.
  if(CAT_ORD_CURRENT_VARIANT>=n)CAT_ORD_CURRENT_VARIANT=Math.max(0,n-1);
  const fieldsHtml='<div id="cat-ord-fields" style="display:flex;flex-wrap:wrap;gap:6px;margin-bottom:14px">'+
    CAT_ORD_COLS.map((c,ci)=>{
      const codeOn=catOrdCodeOn(c);
      const codeBtn='<button type=button onclick="toggleCatOrdCodeOn('+ci+')" title="'+(codeOn?'Included':'Not included')+' in the Ordering Code Example" style="border:none;background:none;cursor:pointer;padding:2px;font-size:10px;line-height:1;color:'+(codeOn?'var(--info)':'var(--muted)')+'">'+(codeOn?'✓':'—')+'</button>';
      const fieldDraggable=CAT_ORD_COLS.length>1;
      const iconKey=catOrdFieldIconKey(c);
      // Icon-only pill (drag handle + icon + the frequent ✓/— code-example
      // toggle) — everything else (rename, reorder precision, delete, copy,
      // cut) moved into the click popover, see openCatOrdFieldPopover.
      // Drag-and-drop (ordFieldDrop) already fully replaces the old ◀▶
      // buttons — it can land a field anywhere in one move, so nothing is
      // lost by dropping the arrows from the pill itself.
      return '<div class=dragrow draggable="'+fieldDraggable+'" ondragstart="dragRowStart(event,{kind:\'field\',ci:'+ci+'})" ondragover="dragColOver(event)" ondragleave="dragColLeave(event)" ondrop="ordFieldDrop(event,'+ci+')" ondragend="dragRowEnd(event)" style="display:flex;align-items:center;gap:2px;background:var(--surface-2);border:1px solid var(--border);border-radius:8px;padding:4px 5px">'+
        (fieldDraggable?'<span class=draghandle title="Drag to reorder">⠿</span>':'')+
        '<button type=button class=catordfieldicon onclick="openCatOrdFieldPopover(event,'+ci+')" title="'+escHtml(c.label||'(untitled field)')+'">'+
          '<svg viewBox="0 0 24 24" fill=none stroke=currentColor stroke-width=2 stroke-linecap=round stroke-linejoin=round>'+CAT_ORD_FIELD_ICON[iconKey]+'</svg></button>'+
        codeBtn+
      '</div>'}).join('')+
    '<button type=button class=btn style="font-size:11px;padding:6px 10px" onclick="openCatOrdColMenu(event)">+ Add Field</button>'+
  '</div>';
  // One variant shown at a time (Technical Specifications' own sidebar
  // layout, label-left/value-right full-width rows, not the old wrap of
  // small boxes) — jump between variants via the dropdown, or step through
  // sequentially with Complete. Drag-reordering variants (ordVariantDrop)
  // doesn't apply anymore: with only one variant visible there's no second
  // card to drop onto, so ▲▼ (already fully functional, already existing)
  // is the reorder mechanism here — a deliberate, narrow exception to this
  // app's usual "reorder tools must be draggable" rule, specific to this
  // view no longer showing more than one item at once.
  const vi=CAT_ORD_CURRENT_VARIANT;
  const navHtml='<div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;margin-bottom:10px">'+
    '<select style="font-size:12px;padding:6px 8px;border-radius:6px;font-weight:600" onchange="setCatOrdCurrentVariant(parseInt(this.value,10))">'+
      Array.from({length:n},(_,i)=>'<option value="'+i+'"'+(i===vi?' selected':'')+'>'+escHtml(catOrdVariantDropdownLabel(i))+'</option>').join('')+
    '</select>'+
    '<button type=button class=ordmovebtn'+(vi===0?' disabled':'')+' onclick="moveCatOrdVariantCurrent(-1)" title="Move this variant earlier">▲</button>'+
    '<button type=button class=ordmovebtn'+(vi===n-1?' disabled':'')+' onclick="moveCatOrdVariantCurrent(1)" title="Move this variant later">▼</button>'+
    (n>1?'<button type=button class=rm onclick="removeCatOrdVariantCurrent()" title="Remove this variant">×</button>':'')+
    '<div style="flex:1"></div>'+
    catOrdSortHtml()+
    '<button type=button class="dspill" onclick="completeCatOrdVariant()">'+escHtml(catOrdVariantLabel(vi))+' Complete →</button>'+
  '</div>';
  const rowsHtml=CAT_ORD_COLS.map((c,ci)=>{
    const isSpecial=CAT_ORD_SPECIAL_LABELS.includes((c.label||'').trim());
    return '<div style="display:flex;gap:10px;align-items:flex-start;padding:7px 0;border-bottom:1px solid var(--line)">'+
      '<div style="flex:0 0 120px;font-size:11px;color:var(--muted);padding-top:6px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'+escHtml(c.label||'(untitled)')+'</div>'+
      '<div style="flex:1;min-width:0">'+catOrdValueInput(ci,vi,c.values[vi]||'',isSpecial,false,c.label,1)+'</div>'+
    '</div>'}).join('');
  $('cat-ord-table').innerHTML=fieldsHtml+navHtml+rowsHtml+
    '<div style="margin-top:10px">'+
      '<button type=button class=btn onclick="addCatOrdVariantAndFocus()">+ Add Variant</button>'+
    '</div>';
  renderCatOrdWidthEditor();
  renderCatOrdAlignRowsBtn()}
// Shared by the variant-nav header — unchanged from its old spot below the
// cards, just re-homed.
function catOrdSortHtml(){
  const sortableCols=catOrdLockstepCols().filter(c=>c.values.some(v=>catOrdSortableValue(v)!=null));
  if(!sortableCols.length)return '';
  return '<select style="font-size:11px;padding:6px 8px;border-radius:6px" onchange="if(this.value!==\'\'){sortCatOrdVariantsBy(parseInt(this.value,10))}this.value=\'\'">'+
    '<option value="">Sort by…</option>'+
    sortableCols.map(c=>'<option value="'+CAT_ORD_COLS.indexOf(c)+'">'+escHtml(c.label||'(untitled)')+'</option>').join('')+
  '</select>'}
function setCatOrdCurrentVariant(vi){
  if(vi<0||vi>=catOrdVariantCount())return;
  CAT_ORD_CURRENT_VARIANT=vi;
  renderCatOrdTable()}
// Keeps the pointer on the SAME variant it was just showing, now
// repositioned — not left pointing at whatever swapped into this slot.
function moveCatOrdVariantCurrent(dir){
  const vi=CAT_ORD_CURRENT_VARIANT;
  const nvi=vi+dir;
  if(nvi<0||nvi>=catOrdVariantCount())return;
  CAT_ORD_CURRENT_VARIANT=nvi;
  moveCatOrdVariant(vi,dir)}
function removeCatOrdVariantCurrent(){
  removeCatOrdVariant(CAT_ORD_CURRENT_VARIANT);
  if(CAT_ORD_CURRENT_VARIANT>=catOrdVariantCount())CAT_ORD_CURRENT_VARIANT=Math.max(0,catOrdVariantCount()-1)}
function addCatOrdVariantAndFocus(){
  addCatOrdVariant();
  CAT_ORD_CURRENT_VARIANT=catOrdVariantCount()-1;
  renderCatOrdTable()}
// "Complete" is the sequential-fill-in workflow this whole redesign is
// for: on any variant but the last, it's just "go to the next one, I'm
// done with this one" (no new data created). On the last variant it folds
// in "+ Add Variant" — creating a fresh blank one (Model No. copied
// forward, same as addCatOrdVariant always did) and landing on it —
// so finishing your newest variant and starting the next is one click.
function completeCatOrdVariant(){
  const n=catOrdVariantCount();
  if(CAT_ORD_CURRENT_VARIANT<n-1){
    CAT_ORD_CURRENT_VARIANT++;
    renderCatOrdTable();schedulePreview()
  }else{
    addCatOrdVariantAndFocus()}}

// Draggable column-width widget for the Ordering Table — a live mini-header
// (same orange/white/bold look as the real printed one) sized from each
// column's effective weight, with a drag handle on every internal boundary.
// Deliberately lives here in the sidebar, not overlaid on the PDF preview
// pane: that pane is a flat raster PNG of whichever page the table lands on
// (/draft-preview?page=N, see renderPreviewPages()), single or side-by-side
// in Double view, at 40-300% zoom — real drag handles pinned to a moving
// target like that would be constantly fighting page/zoom/layout drift.
// This widget is real interactive DOM instead, styled to read as a stand-in
// for the printed header, with the real PDF preview confirming the result
// live underneath exactly like every other field already does.
let CAT_ORD_AUTO_WEIGHTS=[];
let CAT_ORD_AUTO_WEIGHTS_TIMER=null;
let CAT_ORD_DRAG=null;
// A column's effective weight is its own manual .width if the user has
// already dragged it, else the last-known auto weight from the server
// (html_engine.build_ordering_table — the exact function the real PDF uses,
// see /api/cat-ordering-widths) so this widget's starting layout is never
// out of sync with what actually prints. Falls back to a plain default
// (6, the algorithm's own rough middle-of-the-road value) only for the
// instant before that first fetch resolves.
function catOrdEffectiveWeights(){
  return CAT_ORD_COLS.map((c,i)=>(typeof c.width==='number'&&c.width>0)?c.width:(CAT_ORD_AUTO_WEIGHTS[i]||6))}
function renderCatOrdWidthEditor(fromAutoFetch){
  const wrap=$('cat-ord-widths');if(!wrap)return;
  if(!CAT_ORD_COLS.length){wrap.innerHTML='';return}
  const weights=catOrdEffectiveWeights();
  const total=weights.reduce((a,b)=>a+b,0)||1;
  const n=CAT_ORD_COLS.length;
  wrap.innerHTML='<div id=cat-ord-widths-row style="display:flex;min-width:'+(n*46)+'px;border-radius:6px;overflow:hidden;border:1px solid #c4551e">'+
    CAT_ORD_COLS.map((c,i)=>
      '<div class=catordwbox style="flex:0 0 '+(weights[i]/total*100)+'%;min-width:0;box-sizing:border-box;background:#ec6b2f;color:#fff;font-size:9.5px;font-weight:700;padding:6px 7px;text-align:center;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;position:relative;'+(i<n-1?'border-right:1px solid rgba(255,255,255,.35);':'')+'" title="'+escHtml(c.label||'(untitled)')+'">'+
        escHtml(c.label||'(untitled)')+
        (i<n-1?'<div onmousedown="catOrdWidthDragStart(event,'+i+')" style="position:absolute;top:0;right:-4px;width:8px;height:100%;cursor:ew-resize;z-index:2"></div>':'')+
      '</div>').join('')+
  '</div>';
  if(!fromAutoFetch)refreshCatOrdAutoWidths()}
function refreshCatOrdAutoWidths(){
  // Every column already manually sized — nothing auto left to look up.
  if(CAT_ORD_COLS.every(c=>typeof c.width==='number'&&c.width>0))return;
  clearTimeout(CAT_ORD_AUTO_WEIGHTS_TIMER);
  CAT_ORD_AUTO_WEIGHTS_TIMER=setTimeout(async()=>{
    const cols=CAT_ORD_COLS.map(c=>({label:c.label,values:c.values}));
    try{
      const r=await fetch('/api/cat-ordering-widths',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({ordering_columns:cols})}).then(r=>r.json());
      if(r.weights){CAT_ORD_AUTO_WEIGHTS=r.weights;renderCatOrdWidthEditor(true)}
    }catch(e){}},
  300)}
function catOrdWidthDragStart(ev,i){
  ev.preventDefault();ev.stopPropagation();
  const boxes=document.querySelectorAll('#cat-ord-widths-row .catordwbox');
  const boxA=boxes[i],boxB=boxes[i+1];
  if(!boxA||!boxB)return;
  const weights=catOrdEffectiveWeights();
  CAT_ORD_DRAG={i,startX:ev.clientX,widthA:boxA.getBoundingClientRect().width,widthB:boxB.getBoundingClientRect().width,weightA:weights[i],weightB:weights[i+1],boxA,boxB};
  document.addEventListener('mousemove',catOrdWidthDragMove);
  document.addEventListener('mouseup',catOrdWidthDragEnd)}
function catOrdWidthDragMove(ev){
  const d=CAT_ORD_DRAG;if(!d)return;
  // Deliberately just a generic px floor, not each column's own
  // content-derived safe width — column sizing stays fully user/data
  // driven (per explicit request), the same freedom the drag widget
  // always had. The header's own text no longer needs this as a safety
  // net either: min-width:0 on the header cell (sololuce_datasheet.html)
  // already makes it wrap onto a new line as its column narrows instead
  // of overflowing, for as long as there's an actual word-boundary to
  // wrap at — a single unbreakable word forced narrower than itself is
  // an inherent, accepted edge of a manual drag, not something this
  // widget tries to prevent.
  const MIN=30; // px floor per column, so neither side can be dragged to nothing
  let delta=ev.clientX-d.startX;
  delta=Math.max(MIN-d.widthA,Math.min(d.widthB-MIN,delta));
  const pixelsPerWeight=(d.widthA+d.widthB)/(d.weightA+d.weightB);
  d.liveWeightA=(d.widthA+delta)/pixelsPerWeight;
  d.liveWeightB=(d.widthB-delta)/pixelsPerWeight;
  // Only these two weights change — their sum stays exactly d.weightA+
  // d.weightB, so the shared total (and therefore every OTHER column's own
  // share of it) is provably unaffected by this drag.
  const weights=catOrdEffectiveWeights();
  weights[d.i]=d.liveWeightA;weights[d.i+1]=d.liveWeightB;
  const total=weights.reduce((a,b)=>a+b,0)||1;
  d.boxA.style.flex='0 0 '+(d.liveWeightA/total*100)+'%';
  d.boxB.style.flex='0 0 '+(d.liveWeightB/total*100)+'%';
  schedulePreview()}
function catOrdWidthDragEnd(){
  const d=CAT_ORD_DRAG;if(!d)return;
  document.removeEventListener('mousemove',catOrdWidthDragMove);
  document.removeEventListener('mouseup',catOrdWidthDragEnd);
  if(d.liveWeightA!=null){
    CAT_ORD_COLS[d.i].width=Math.round(d.liveWeightA*100)/100;
    CAT_ORD_COLS[d.i+1].width=Math.round(d.liveWeightB*100)/100}
  CAT_ORD_DRAG=null;
  renderCatOrdWidthEditor();
  schedulePreview()}
function resetCatOrdWidths(){
  CAT_ORD_COLS.forEach(c=>{c.width=null});
  renderCatOrdWidthEditor();
  schedulePreview()}
// Excel-style "double-click a column border to autofit" — as a button
// instead of a gesture, and adapted for a table whose total width is
// FIXED (it must always exactly span the page margins, unlike an Excel
// sheet that's free to grow). Autofit can't give every column its own
// exact ideal width AND keep that fixed total, so it works left to
// right: every column except the LAST snaps straight to its own
// content's exact safe weight (the same algorithm that sizes the real
// printed PDF — html_engine.py's build_ordering_table, fetched fresh
// here rather than trusting a possibly-stale cached value), and the
// last column absorbs whatever weight is left over so the grand total
// stays exactly what it was — it grows or shrinks depending on whether
// the other columns needed more or less room than they had. A one-time
// snap (same as Excel's double-click), not a live binding — every
// column, including the last, stays freely drag-adjustable afterward.
async function autoFitCatOrdColumns(){
  const n=CAT_ORD_COLS.length;
  if(n<2)return;
  const total=catOrdEffectiveWeights().reduce((a,b)=>a+b,0)||1;
  const cols=CAT_ORD_COLS.map(c=>({label:c.label,values:c.values}));
  let ideal;
  try{
    const r=await fetch('/api/cat-ordering-widths',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({ordering_columns:cols})}).then(r=>r.json());
    ideal=r.weights
  }catch(e){}
  if(!ideal||ideal.length!==n){toast('Could not compute column sizes — try again');return}
  CAT_ORD_AUTO_WEIGHTS=ideal;
  let sumIdeal=0;
  for(let i=0;i<n-1;i++){
    CAT_ORD_COLS[i].width=Math.round(ideal[i]*100)/100;
    sumIdeal+=ideal[i]}
  const lastWeight=Math.max(4,total-sumIdeal); // 4 = the algorithm's own general floor, so the last column can shrink to fit but never collapse to nothing
  CAT_ORD_COLS[n-1].width=Math.round(lastWeight*100)/100;
  renderCatOrdWidthEditor();
  schedulePreview();
  toast('Columns auto-fit to content, left to right')}
// "Align Rows" — per explicit request, a toggle (not a one-time snap like
// Auto-Fit Columns above) since what counts as this table's own "standard"
// row height can genuinely change later (more rows added, a value edited
// into needing an extra line) and should keep tracking that, not freeze
// whatever was true the moment it was clicked. The actual height math
// lives server-side (html_engine.py's _ord_row_min_height_pt) so the
// live preview and the real PDF can never disagree — this just flips the
// boolean sent with the rest of the document and re-renders the button's
// own label/state.
function toggleCatOrdAlignRows(){
  CAT_ORD_ALIGN_ROWS=!CAT_ORD_ALIGN_ROWS;
  renderCatOrdAlignRowsBtn();
  schedulePreview()}
function renderCatOrdAlignRowsBtn(){
  const btn=$('cat-ord-alignrows-btn');if(!btn)return;
  btn.textContent=CAT_ORD_ALIGN_ROWS?'Rows Aligned ✓':'Align Rows';
  btn.title=CAT_ORD_ALIGN_ROWS
    ?'On: short rows are padded up to this table\'s own most common row height. A row that genuinely needs more room always keeps every line — click to turn off'
    :'Pulls short rows up to match this table\'s own most common row height, so the printed table doesn\'t look uneven — never shrinks a row that genuinely needs more room';
  btn.style.background=CAT_ORD_ALIGN_ROWS?'var(--info)':'';
  btn.style.color=CAT_ORD_ALIGN_ROWS?'#fff':'';
  btn.style.borderColor=CAT_ORD_ALIGN_ROWS?'var(--info)':''}
// Captures whatever width each column is CURRENTLY showing — a manual
// drag if the user made one, else the last-known auto weight, i.e.
// exactly catOrdEffectiveWeights()'s own definition of "current width" —
// and remembers it per LABEL server-side (cat_ordering_default_widths) so
// every future datasheet's same field starts from this same width, not
// just this one document's. Keyed by label (not column position) since
// that's what carries across different datasheets' own column sets.
async function saveCatOrdWidthsAsStandard(){
  const weights=catOrdEffectiveWeights();
  const byLabel={};
  CAT_ORD_COLS.forEach((c,i)=>{
    const label=(c.label||'').trim();
    if(label)byLabel[label]=weights[i]});
  const r=await fetch('/api/cat-ordering-default-widths',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({weights:byLabel})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  toast('Saved as the standard column widths — new datasheets start from this')}

function collectCatData(){
  // Recomputed here (not just on blur) so Generate/the live preview always
  // reflects the current Power/Luminare Efficacy/Finish Colors even if the
  // user is still mid-edit in a field and never blurred it.
  recomputeCatSpecialOrdColumns();
  // Belt-and-braces uppercase (the input already force-uppercases as you type
  // via catUpperCaseInPlace) — covers values that never went through that
  // oninput handler, e.g. Import-from-PDF prefill or an older draft/generated
  // file saved before this rule existed.
  const productName=$('cat-productname').value.trim().toUpperCase();
  const series=$('cat-series').value.trim();
  // Unchecking "part of a family" keeps whatever name is still typed in the
  // DOM (so re-checking the box later doesn't lose it) but saves family:'' —
  // an explicit un-set, not a linked-products unlink (those are their own
  // immediate action via removeCatFamilyMember).
  const family=($('cat-family-enabled')&&$('cat-family-enabled').checked)?catFamilyCurrentValue():'';
  return {
    company:productName, project:series,
    product_name:productName, series, family,
    product_type:$('cat-producttype').value,
    page_number:parseInt($('cat-pagenum').value,10)||1,
    description:$('cat-description').value.trim(),
    note:$('cat-note').value.trim(),
    ordering_code_example:$('cat-ordcode').value.trim(),
    main_photo:CAT_IMG.main.src, main_photo_zoom:CAT_IMG.main.zoom, main_photo_x:CAT_IMG.main.x, main_photo_y:CAT_IMG.main.y, main_photo_mask:CAT_IMG.main.mask, main_photo_placeholder:CAT_IMG.main.placeholder!==false,
    lifestyle_photo:CAT_IMG.lifestyle.src, lifestyle_photo_zoom:CAT_IMG.lifestyle.zoom, lifestyle_photo_x:CAT_IMG.lifestyle.x, lifestyle_photo_y:CAT_IMG.lifestyle.y, lifestyle_photo_mask:CAT_IMG.lifestyle.mask, lifestyle_photo_placeholder:CAT_IMG.lifestyle.placeholder!==false,
    dimension_diagram:CAT_IMG.diagram.src, dimension_diagram_zoom:CAT_IMG.diagram.zoom, dimension_diagram_x:CAT_IMG.diagram.x, dimension_diagram_y:CAT_IMG.diagram.y, dimension_diagram_mask:CAT_IMG.diagram.mask, dimension_diagram_mask_anchor_x:CAT_IMG.diagram.maskAnchorX, dimension_diagram_mask_anchor_y:CAT_IMG.diagram.maskAnchorY, dimension_diagram_label:CAT_IMG.diagram.label, dimension_diagram_placeholder:CAT_IMG.diagram.placeholder!==false,
    extra_photo_1:CAT_IMG.extra1.src, extra_photo_1_zoom:CAT_IMG.extra1.zoom, extra_photo_1_x:CAT_IMG.extra1.x, extra_photo_1_y:CAT_IMG.extra1.y, extra_photo_1_mask:CAT_IMG.extra1.mask, extra_photo_1_mask_anchor_x:CAT_IMG.extra1.maskAnchorX, extra_photo_1_mask_anchor_y:CAT_IMG.extra1.maskAnchorY, extra_photo_1_label:CAT_IMG.extra1.label, extra_photo_1_show:!!CAT_IMG.extra1.show, extra_photo_1_merged:!!CAT_IMG.extra1.merged, extra_photo_1_autosize:!!CAT_IMG.extra1.autosize, extra_photo_1_placeholder:CAT_IMG.extra1.placeholder!==false,
    extra_photo_2:CAT_IMG.extra2.src, extra_photo_2_zoom:CAT_IMG.extra2.zoom, extra_photo_2_x:CAT_IMG.extra2.x, extra_photo_2_y:CAT_IMG.extra2.y, extra_photo_2_mask:CAT_IMG.extra2.mask, extra_photo_2_mask_anchor_x:CAT_IMG.extra2.maskAnchorX, extra_photo_2_mask_anchor_y:CAT_IMG.extra2.maskAnchorY, extra_photo_2_label:CAT_IMG.extra2.label, extra_photo_2_show:!!CAT_IMG.extra2.show, extra_photo_2_placeholder:CAT_IMG.extra2.placeholder!==false,
    extra_photo_3:CAT_IMG.extra3.src, extra_photo_3_zoom:CAT_IMG.extra3.zoom, extra_photo_3_x:CAT_IMG.extra3.x, extra_photo_3_y:CAT_IMG.extra3.y, extra_photo_3_mask:CAT_IMG.extra3.mask, extra_photo_3_mask_anchor_x:CAT_IMG.extra3.maskAnchorX, extra_photo_3_mask_anchor_y:CAT_IMG.extra3.maskAnchorY, extra_photo_3_label:CAT_IMG.extra3.label, extra_photo_3_merged:!!CAT_IMG.extra3.merged, extra_photo_3_placeholder:CAT_IMG.extra3.placeholder!==false,
    badges:CAT_BADGES.filter(b=>b.key),
    // Luminare Efficacy deliberately stays '\n'-joined here (the normal,
    // lossless multi-value convention) rather than compact-joined like
    // Power — Power can safely afford to lose its per-value breakdown on
    // this side because its real values live a second time, in full, in
    // ordering_columns (see syncSpecPowerFromOrdering); Efficacy has no such
    // backup copy anywhere else, so compact-joining it HERE would silently
    // and permanently collapse a 2-tier efficacy fitting down to one number
    // the moment the draft is saved and reloaded (populateCatForm splits on
    // '\n', so a comma/range-joined string comes back as a single value).
    // The compact "like Power" print styling the user asked for is instead
    // applied read-only at PDF-render time — see html_engine.py's
    // render_datasheet_pdf, which never writes back to the saved draft.
    specs:CAT_SPECS.filter(s=>(s.label||'').trim()||(s.values||[]).some(v=>(v||'').trim())).map(s=>({label:s.label,value:normSpecLabel(s.label)==='power'?formatCatSpecPowerValue(s.values):(s.values||['']).join('\n')})),
    finish_colors:CAT_FINISH,
    ordering_columns:CAT_ORD_COLS.map(c=>({label:c.label,values:c.values,width:c.width||undefined})),
    ordering_align_rows:!!CAT_ORD_ALIGN_ROWS,
  }}

// ---------------------------------------------------------------- Import from PDF (Sololuce Datasheets only)
// Reads an arbitrary manufacturer PDF via local heuristics (pdf_extract.py —
// no AI/API key, no cost, no internet call) and pre-fills the CAT form from
// it. Never saves/generates by itself — the user
// reviews everything in the overlay below, then "Apply to Datasheet" copies
// only the accepted rows into the exact same CAT_* globals/renderCat*()
// functions the manual-entry form already uses, and the user still hits the
// real Generate button themselves afterward.
// pdf_extract.py's badge detection still speaks the old fixed-concept
// vocabulary (ip/cri/ugr/rohs/ce/...), but the badge picker is now a real
// image library with no fixed vocabulary — this is a best-effort concept ->
// library-label keyword match, so a detected "RoHS" mention still
// auto-applies to whichever real library badge is about hazardous
// substances, without the two systems needing to agree on a shared key set.
const IMPORT_BADGE_KEYWORDS={
  ce:'conformity', rohs:'hazardous', ground:'class', weee:'recycling',
  house:'buildings', ip:'protected', energy:'performance',
  cri:'color rendering', ugr:'glare', warranty:'manufacturing defects',
  dali:'dali', em:'emergency', sdcm:'color', dimmable:'dali',
};
function findLibraryBadgeForConcept(conceptKey){
  const kw=(IMPORT_BADGE_KEYWORDS[conceptKey]||conceptKey).toLowerCase();
  return CAT_BADGE_LIBRARY.find(b=>b.label.toLowerCase().includes(kw));
}
let IMP_ID=null, IMP_PAGES=0, IMP_FIELDS=[], IMP_SELECTED=null, IMP_DRAW=false, IMP_ZOOM=100, impFieldSeq=0, impDrag=null, impPendingBox=null;

function pickCatImportFile(){$('catimportfile').click()}
function onCatImportFile(input){
  const f=input.files[0];if(!f)return;
  input.value='';
  const reader=new FileReader();
  reader.onload=async()=>{
    openCatImportModal();
    $('catimportstatus').textContent='Uploading…';
    const r=await fetch('/api/cat-import/upload',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({pdf:reader.result})}).then(r=>r.json());
    if(r.error){$('catimportstatus').textContent='';alert(r.error);closeCatImport();return}
    IMP_ID=r.importId;IMP_PAGES=r.pages;IMP_FIELDS=[];IMP_SELECTED=null;
    renderImpPages();renderImpFields();
    runCatExtraction()};
  reader.readAsDataURL(f)}

function openCatImportModal(){$('catimportmodal').classList.remove('hide')}
function closeCatImport(){
  $('catimportmodal').classList.add('hide');
  IMP_ID=null;IMP_PAGES=0;IMP_FIELDS=[];IMP_SELECTED=null;IMP_DRAW=false;
  $('catimportleft').innerHTML='';$('catimportright').innerHTML='';$('catimportstatus').textContent=''}

function renderImpPages(){
  let html='';
  for(let i=1;i<=IMP_PAGES;i++){
    html+='<div class=impPageWrap data-page="'+i+'" style="width:'+Math.round(720*IMP_ZOOM/100)+'px">'+
      '<img onload="positionImpBoxesForPage('+i+')" onmousedown="onImpPageMouseDown(event,'+i+')" src="/cat-import-preview?importId='+IMP_ID+'&page='+i+'">'+
      '<div class=impboxlayer id="impboxlayer-'+i+'" style="position:absolute;inset:0;pointer-events:none"></div>'+
    '</div>'}
  $('catimportleft').innerHTML=html}

async function runCatExtraction(){
  $('catimportstatus').textContent='Scanning the PDF…';
  const r=await fetch('/api/cat-import/extract',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({importId:IMP_ID})}).then(r=>r.json());
  $('catimportstatus').textContent='';
  if(r.error){alert(r.error);return}
  IMP_FIELDS=normalizeExtractedFields(r.fields);
  renderImpFields();renderImpBoxes()}

function normalizeExtractedFields(f){
  const out=[];
  const addScalar=(key,label)=>{
    if(!f[key])return;
    out.push({id:'f'+(impFieldSeq++),group:'Header',label,value:f[key],page:f[key+'_source_page']||null,box:f[key+'_box']||null,target:{kind:'scalar',field:key},accepted:true});
  };
  addScalar('product_name','Product Name');
  addScalar('series','Series / Category');
  addScalar('description','Description');
  addScalar('ordering_code_example','Ordering Code Example');
  (f.specs||[]).forEach(s=>out.push({id:'f'+(impFieldSeq++),group:'Technical Specifications',label:s.label||'Spec',value:s.value||'',page:s.source_page||null,box:s.box||null,target:{kind:'spec'},raw:s,accepted:true}));
  (f.badges||[]).forEach(b=>{
    const matched=findLibraryBadgeForConcept(b.key);
    out.push({id:'f'+(impFieldSeq++),group:'Spec Badges',
      label:matched?matched.label:(b.key+' (no matching badge image in your library)'),
      value:b.value||'',page:b.source_page||null,box:b.box||null,
      target:{kind:'badge'},raw:{libraryKey:matched?matched.key:null},accepted:!!matched})});
  (f.finish_colors||[]).forEach(c=>out.push({id:'f'+(impFieldSeq++),group:'Finish Colors',label:c.label||'Color',value:c.hex_guess||'',page:c.source_page||null,box:c.box||null,target:{kind:'finish'},raw:c,accepted:true}));
  if(f.ordering_columns&&f.ordering_columns.length)out.push({id:'f'+(impFieldSeq++),group:'Ordering Table',
    label:'Ordering Table ('+f.ordering_columns.length+' cols × '+(f.ordering_rows||[]).length+' rows)',
    value:f.ordering_columns.join(' · '),page:f.ordering_table_source_page||null,box:f.ordering_table_box||null,
    target:{kind:'ordering'},raw:{columns:f.ordering_columns,rows:f.ordering_rows||[]},accepted:true});
  (f.photo_candidates||[]).forEach(p=>out.push({id:'f'+(impFieldSeq++),group:'Images',
    label:p.role.charAt(0).toUpperCase()+p.role.slice(1)+' Photo',value:'Page '+p.page,page:p.page||null,box:p.box||null,
    target:{kind:'photo',role:p.role},raw:p,accepted:true}));
  return out}

function renderImpFields(){
  const groups=[];
  IMP_FIELDS.forEach(f=>{if(!groups.includes(f.group))groups.push(f.group)});
  $('catimportright').innerHTML=groups.length?groups.map(g=>
    '<div class=impgrouphead>'+g+'</div>'+IMP_FIELDS.filter(f=>f.group===g).map(impFieldHtml).join('')
  ).join(''):'<p class="muted" style="font-size:12px">No fields extracted yet.</p>'}
function impFieldHtml(f){
  return '<div class="impfield'+(IMP_SELECTED===f.id?' selected':'')+'" data-id="'+f.id+'" onclick="selectImpField(\''+f.id+'\')">'+
    '<div class=impfieldhead><span>'+escHtml(f.label)+'</span>'+
      '<label style="font-weight:400;text-transform:none;font-size:11px;display:flex;gap:4px;align-items:center">'+
        '<input type=checkbox '+(f.accepted?'checked':'')+' onclick="event.stopPropagation();toggleImpAccept(\''+f.id+'\')"> use</label></div>'+
    '<div class=impfieldval>'+escHtml(f.value||'')+'</div>'+
    (f.page&&!f.box?'<div class=impnoloc>Found on page '+f.page+' — exact location not found, value only.</div>':'')+
    (!f.page?'<div class=impnoloc>Not found in this PDF.</div>':'')+
  '</div>'}
function toggleImpAccept(id){const f=IMP_FIELDS.find(x=>x.id===id);if(f)f.accepted=!f.accepted;renderImpFields()}
function selectImpField(id){
  IMP_SELECTED=(IMP_SELECTED===id?null:id);
  renderImpFields();renderImpBoxes();
  const f=IMP_FIELDS.find(x=>x.id===id);
  if(f&&f.box){const wrap=document.querySelector('.impPageWrap[data-page="'+f.box.page+'"]');if(wrap)wrap.scrollIntoView({behavior:'smooth',block:'center'})}}

// Real placeholder-box-vs-preview-page size mismatch, confirmed directly:
// every one of these 6 call sites read scale from wrap.dataset.scale — a
// value only ever WRITTEN in positionImpBoxesForPage, itself only called
// from the page <img>'s own onload. Anything that needed a box position
// BEFORE that onload had fired for a given page (extraction finishing
// before a large 170dpi page image finishes loading is the realistic
// case — the image is far bigger than the extraction API call, so on a
// slow connection or a big multi-page PDF, image load losing that race is
// completely plausible, not just a theoretical corner case) read
// dataset.scale as undefined, and `parseFloat(undefined)||1` silently
// fell back to scale:1 — i.e. treated raw natural-pixel box coordinates
// (at CAT_IMPORT_DPI=170) as if they were already CSS/display pixels.
// Confirmed reproducing this directly: a box at natural x0=100 rendered
// at left:100px instead of the correct ~51px (the page image displays at
// 720px CSS width vs ~1405px natural width for an A4 page at 170dpi, a
// ~0.51 scale) — very nearly 2x too far right/large, exactly the kind of
// "placeholder isn't the same size as the preview page" mismatch this was
// reported as. It's self-correcting the MOMENT the image's onload actually
// fires (positionImpBoxesForPage re-renders with the real scale) — so the
// bug is a transient wrong-then-right flash rather than a permanently
// broken page, but a real, confirmed bug either way, not a false alarm.
// Fix: never trust a cached, possibly-still-unset dataset attribute — read
// the scale fresh off the actual <img> element every time something needs
// it, and treat "the image hasn't loaded yet" (naturalWidth still 0) as
// "there is no valid scale, don't draw anything" rather than silently
// guessing 1. dataset.scale itself is no longer written or read anywhere.
function impScaleFor(wrap){
  const img=wrap&&wrap.querySelector('img');
  if(!img||!img.naturalWidth)return null;
  return img.clientWidth/img.naturalWidth}
function positionImpBoxesForPage(pageNo){
  renderImpBoxesForPage(pageNo)}
function renderImpBoxes(){for(let i=1;i<=IMP_PAGES;i++)renderImpBoxesForPage(i)}
function renderImpBoxesForPage(pageNo){
  const layer=$('impboxlayer-'+pageNo);const wrap=document.querySelector('.impPageWrap[data-page="'+pageNo+'"]');
  if(!layer||!wrap)return;
  const scale=impScaleFor(wrap);
  if(scale===null)return; // image not loaded yet — leave whatever was there (nothing, on first call) rather than draw at a guessed/wrong scale
  const boxes=IMP_FIELDS.filter(f=>f.box&&f.box.page===pageNo);
  layer.innerHTML=boxes.map(f=>{
    const b=f.box,sel=IMP_SELECTED===f.id;
    const left=b.x0*scale,top=b.y0*scale,w=(b.x1-b.x0)*scale,h=(b.y1-b.y0)*scale;
    return '<div class="impBox'+(sel?' selected':'')+'" data-id="'+f.id+'" style="left:'+left+'px;top:'+top+'px;width:'+w+'px;height:'+h+'px;pointer-events:auto" onmousedown="onImpBoxMouseDown(event,\''+f.id+'\')" title="'+escHtml(f.label)+'">'+
      (sel?['nw','ne','sw','se'].map(pos=>'<div class="impHandle" style="top:'+(pos[0]==='n'?'0':'100%')+';left:'+(pos[1]==='w'?'0':'100%')+'" onmousedown="event.stopPropagation();onImpHandleMouseDown(event,\''+f.id+'\',\''+pos+'\')"></div>').join(''):'')+
    '</div>'}).join('')}

function onImpBoxMouseDown(ev,id){
  ev.preventDefault();
  selectImpField(id);
  const f=IMP_FIELDS.find(x=>x.id===id);if(!f||!f.box)return;
  const wrap=document.querySelector('.impPageWrap[data-page="'+f.box.page+'"]');
  impDrag={mode:'move',id,startX:ev.clientX,startY:ev.clientY,orig:{...f.box},scale:impScaleFor(wrap)||1}}
function onImpHandleMouseDown(ev,id,pos){
  ev.preventDefault();
  const f=IMP_FIELDS.find(x=>x.id===id);if(!f||!f.box)return;
  const wrap=document.querySelector('.impPageWrap[data-page="'+f.box.page+'"]');
  impDrag={mode:'resize',id,pos,startX:ev.clientX,startY:ev.clientY,orig:{...f.box},scale:impScaleFor(wrap)||1}}
window.addEventListener('mousemove',e=>{
  if(!impDrag)return;
  const f=IMP_FIELDS.find(x=>x.id===impDrag.id);if(!f)return;
  const dx=(e.clientX-impDrag.startX)/impDrag.scale, dy=(e.clientY-impDrag.startY)/impDrag.scale;
  const o=impDrag.orig;
  if(impDrag.mode==='move'){
    f.box={page:o.page,x0:o.x0+dx,y0:o.y0+dy,x1:o.x1+dx,y1:o.y1+dy};
  }else{
    let x0=o.x0,y0=o.y0,x1=o.x1,y1=o.y1;
    if(impDrag.pos.includes('n'))y0=o.y0+dy;else y1=o.y1+dy;
    if(impDrag.pos.includes('w'))x0=o.x0+dx;else x1=o.x1+dx;
    f.box={page:o.page,x0:Math.min(x0,x1-10),y0:Math.min(y0,y1-10),x1:Math.max(x1,x0+10),y1:Math.max(y1,y0+10)};
  }
  renderImpBoxesForPage(f.box.page)});
window.addEventListener('mouseup',()=>{impDrag=null});

function toggleImpDraw(){IMP_DRAW=!IMP_DRAW;$('impdrawbtn').classList.toggle('on',IMP_DRAW)}
function onImpPageMouseDown(ev,pageNo){
  if(!IMP_DRAW||ev.target.closest('.impBox'))return;
  ev.preventDefault();
  const wrap=document.querySelector('.impPageWrap[data-page="'+pageNo+'"]');
  const rect=wrap.getBoundingClientRect();
  const scale=impScaleFor(wrap)||1;
  const startX=(ev.clientX-rect.left)/scale, startY=(ev.clientY-rect.top)/scale;
  const move=e2=>{
    const curX=(e2.clientX-rect.left)/scale, curY=(e2.clientY-rect.top)/scale;
    drawTempBox(pageNo,{x0:Math.min(startX,curX),y0:Math.min(startY,curY),x1:Math.max(startX,curX),y1:Math.max(startY,curY)})};
  const up=e2=>{
    window.removeEventListener('mousemove',move);window.removeEventListener('mouseup',up);
    const curX=(e2.clientX-rect.left)/scale, curY=(e2.clientY-rect.top)/scale;
    const box={page:pageNo,x0:Math.min(startX,curX),y0:Math.min(startY,curY),x1:Math.max(startX,curX),y1:Math.max(startY,curY)};
    clearTempBox();
    if(box.x1-box.x0>8&&box.y1-box.y0>8)openImpNewFieldPicker(box,e2.clientX,e2.clientY)};
  window.addEventListener('mousemove',move);window.addEventListener('mouseup',up)}
function drawTempBox(pageNo,box){
  const layer=$('impboxlayer-'+pageNo);
  const wrap=document.querySelector('.impPageWrap[data-page="'+pageNo+'"]');
  const scale=impScaleFor(wrap)||1;
  let el=document.getElementById('imp-temp-box');
  if(!el){el=document.createElement('div');el.id='imp-temp-box';el.className='impBox selected';el.style.pointerEvents='none';layer.appendChild(el)}
  el.style.left=(box.x0*scale)+'px';el.style.top=(box.y0*scale)+'px';el.style.width=((box.x1-box.x0)*scale)+'px';el.style.height=((box.y1-box.y0)*scale)+'px'}
function clearTempBox(){const el=document.getElementById('imp-temp-box');if(el)el.remove()}

function openImpNewFieldPicker(box,clientX,clientY){
  impPendingBox=box;
  const pop=$('impnewfieldpopover');
  pop.innerHTML='<div style="font-weight:700;font-size:12px;margin-bottom:6px">What is this?</div>'+
    '<select id=impnewfieldkind style="width:100%;margin-bottom:8px">'+
      '<option value="spec">New Spec Row</option>'+
      '<option value="scalar:product_name">Product Name</option>'+
      '<option value="scalar:series">Series / Category</option>'+
      '<option value="scalar:description">Description</option>'+
      '<option value="scalar:ordering_code_example">Ordering Code Example</option>'+
      '<option value="photo:main">Main Product Photo</option>'+
      '<option value="photo:lifestyle">Application Photo</option>'+
      '<option value="photo:diagram">Bottom Right</option>'+
    '</select>'+
    '<div style="display:flex;gap:6px"><button class="btn dark" style="flex:1" onclick=confirmImpNewField()>Add</button><button class=btn style="flex:1" onclick=cancelImpNewField()>Cancel</button></div>';
  pop.style.left=Math.max(8,Math.min(clientX,window.innerWidth-260))+'px';
  pop.style.top=Math.max(8,Math.min(clientY,window.innerHeight-160))+'px';
  pop.style.display='block'}
function cancelImpNewField(){$('impnewfieldpopover').style.display='none';impPendingBox=null;IMP_DRAW=false;$('impdrawbtn').classList.remove('on')}
async function confirmImpNewField(){
  const kind=$('impnewfieldkind').value;
  const box=impPendingBox;
  $('impnewfieldpopover').style.display='none';
  IMP_DRAW=false;$('impdrawbtn').classList.remove('on');
  if(kind.startsWith('photo:')){
    const role=kind.split(':')[1];
    const url='/cat-import-crop?importId='+IMP_ID+'&page='+box.page+'&x0='+box.x0+'&y0='+box.y0+'&x1='+box.x1+'&y1='+box.y1;
    IMP_FIELDS.push({id:'f'+(impFieldSeq++),group:'Images',label:(role.charAt(0).toUpperCase()+role.slice(1))+' Photo (manual)',
      value:'Cropped region',page:box.page,box,target:{kind:'photo',role},raw:{cropUrl:url},accepted:true});
    renderImpFields();renderImpBoxes();return}
  $('catimportstatus').textContent='Reading selection…';
  const r=await fetch('/api/cat-import/recapture',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({importId:IMP_ID,page:box.page,rect:box,hint:kind})}).then(r=>r.json());
  $('catimportstatus').textContent='';
  const value=r.value||'';
  let group='Technical Specifications',label='New value',target={kind:'spec'},raw=null;
  if(kind.startsWith('scalar:')){
    const field=kind.split(':')[1];
    label=({product_name:'Product Name',series:'Series / Category',description:'Description',ordering_code_example:'Ordering Code Example'})[field];
    group='Header';target={kind:'scalar',field}}
  IMP_FIELDS.push({id:'f'+(impFieldSeq++),group,label,value,page:box.page,box,target,raw,accepted:true});
  renderImpFields();renderImpBoxes()}

function applyCatImport(){
  if(!IMP_FIELDS.length){closeCatImport();return}
  const accepted=IMP_FIELDS.filter(f=>f.accepted);
  const skipped=[];
  accepted.forEach(f=>{
    if(f.target.kind==='scalar'){
      if(f.target.field==='series'){renderCatSeriesField(f.value||'');return}
      const elId={product_name:'cat-productname',description:'cat-description',ordering_code_example:'cat-ordcode'}[f.target.field];
      if(elId)$(elId).value=f.value||'';
    }else if(f.target.kind==='spec'){
      // A brand-new CAT form starts pre-filled with 11 blank standard rows
      // (see resetCatForm/CAT_DEFAULT_SPEC_LABELS) — fill a matching blank
      // row by label instead of always pushing a new one, so importing
      // doesn't leave 11 empty duplicate rows sitting above the real data.
      const label=(f.raw&&f.raw.label)||f.label||'';
      const blank=CAT_SPECS.find(s=>!(s.values&&s.values.some(v=>(v||'').trim()))&&(s.label||'').trim().toLowerCase()===label.trim().toLowerCase());
      if(blank)blank.values=[f.value||''];
      else CAT_SPECS.push({label,values:[f.value||'']});
    }else if(f.target.kind==='badge'){
      const libKey=f.raw&&f.raw.libraryKey;
      if(!libKey||isCatBadgeSelected(libKey)){if(!libKey)skipped.push(f.label);return}
      CAT_BADGES.push({key:libKey});
    }else if(f.target.kind==='finish'){
      CAT_FINISH.push({hex:(f.raw&&f.raw.hex_guess)||'#ffffff',label:(f.raw&&f.raw.label)||f.label||''});
    }else if(f.target.kind==='ordering'){
      const cols=(f.raw&&f.raw.columns)||[],rows=(f.raw&&f.raw.rows)||[];
      CAT_COL_SEQ=0;
      CAT_ORD_COLS=cols.map((label,ci)=>({key:'col'+(CAT_COL_SEQ++),label,
        values:rows.length?rows.map(cells=>cells[ci]||''):['']}));
    }else if(f.target.kind==='photo'){
      const url=(f.raw&&f.raw.cropUrl)||(f.box?('/cat-import-crop?importId='+IMP_ID+'&page='+f.box.page+'&x0='+f.box.x0+'&y0='+f.box.y0+'&x1='+f.box.x1+'&y1='+f.box.y1):null);
      if(!url){skipped.push(f.label);return}
      fetch(url).then(r=>r.blob()).then(blob=>{
        const reader=new FileReader();
        reader.onload=()=>{CAT_IMG[f.target.role]=Object.assign(catImgDefault(),{src:reader.result});renderCatImages();schedulePreview()};
        reader.readAsDataURL(blob)});
    }
  });
  if(!CAT_SPECS.length)CAT_SPECS=[{label:'',values:['']}];
  renderCatBadges();renderCatSpecs();renderCatFinish();renderCatOrdTable();
  schedulePreview();
  closeCatImport();
  let msg='Applied '+accepted.length+' field'+(accepted.length===1?'':'s')+' from the PDF — review before Generate';
  if(skipped.length)msg+=' (skipped: '+skipped.join(', ')+')';
  toast(msg)}

function onUnitChange(i,sel){
  if(sel.value==='__manage__'){const rect=sel.getBoundingClientRect();renderItems();openUnitManager(rect);return}
  if(sel.value!=='__custom__'){upd(i,'unit',sel.value);return}
  const td=sel.parentElement;
  td.innerHTML='<input type=text placeholder="Type a unit, press Enter…">';
  const inp=td.querySelector('input');
  inp.focus();
  let done=false;
  const commit=async()=>{
    if(done)return; done=true;
    const v=inp.value.trim();
    if(!v){renderItems();return}
    const r=await fetch('/api/add-unit',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({unit:v})}).then(r=>r.json());
    if(r.units)UNITS=r.units;
    upd(i,'unit',v);renderItems()};
  inp.addEventListener('keydown',e=>{
    if(e.key==='Enter'){e.preventDefault();inp.blur()}
    else if(e.key==='Escape'){done=true;renderItems()}});
  inp.addEventListener('blur',commit)}

function openUnitManager(rect){
  const menu=$('filemenu');
  renderUnitManager();
  menu.style.display='block';
  const r=rect||{left:0,bottom:0};
  const w=menu.offsetWidth||200,h=menu.offsetHeight||200;
  let x=r.left,y=r.bottom+4;
  if(x+w>window.innerWidth-8)x=window.innerWidth-w-8;
  if(y+h>window.innerHeight-8)y=window.innerHeight-h-8;
  menu.style.left=x+'px';menu.style.top=y+'px';
  setTimeout(()=>document.addEventListener('click',closeFileMenu,{once:true}),0)}
function renderUnitManager(){
  $('filemenu').innerHTML='<div class=fmtitle>Manage units</div>'+
    (UNITS.length?UNITS.map(u=>'<div class=fmi><span style=flex:1>'+u+'</span><span class=ic style=cursor:pointer onclick="event.stopPropagation();removeUnit(\''+u.replace(/'/g,"\\'")+'\')">🗑</span></div>').join('')
      :'<div class="fmi disabled">No units yet</div>')}
async function removeUnit(u){
  const r=await fetch('/api/remove-unit',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({unit:u})}).then(r=>r.json());
  if(r.units)UNITS=r.units;
  renderUnitManager();renderItems()}
function upd(i,k,v){items[i][k]=v;schedulePreview()}
function addRow(){items.push({});renderItems();schedulePreview()}
function delRow(i){items.splice(i,1);if(!items.length)items.push({});SELECTED_ITEMS.clear();renderItems();schedulePreview()}

// ---------------------------------------------------------------- Expense Report line items
// Own array/renderer rather than reusing the shared items[]/renderItems() —
// the shape (date/product/description/payment method/amount) doesn't match
// the shared type/unit/qty/price columns those were built around, same
// reasoning CAT's Ordering Table has its own separate columns.
let EXP_ITEMS=[{}], EXP_PAYMENT_METHODS=['CASH','CARD','BANK TRANSFER'],
  EXP_EMPLOYEES=['Edgar Kagramanyan','Suraj Mathews','Lea Galleato'],
  EXP_CATEGORIES=['PETROL','SALIK','MATERIAL','REPAIR','MAINTENANCE','PARKING','OFFICE SUPPLIES','TRAVEL','ACCOMMODATION','MISCELLANEOUS'],
  EXP_PRODUCTS=['Petrol','Salik','Material','Repair'],
  EXP_DESCRIPTIONS=['VEHICLE 36533','CCT CHANGE','PURCHASE FOR COMPANY'];
async function loadExpPaymentMethods(){const r=await fetch('/api/expense-payment-methods').then(r=>r.json());if(r.options&&r.options.length)EXP_PAYMENT_METHODS=r.options}
async function loadExpEmployees(){const r=await fetch('/api/expense-employees').then(r=>r.json());if(r.options&&r.options.length)EXP_EMPLOYEES=r.options}
async function loadExpCategories(){const r=await fetch('/api/expense-categories').then(r=>r.json());if(r.options&&r.options.length)EXP_CATEGORIES=r.options}
async function loadExpProducts(){const r=await fetch('/api/expense-products').then(r=>r.json());if(r.options&&r.options.length)EXP_PRODUCTS=r.options}
async function loadExpDescriptions(){const r=await fetch('/api/expense-descriptions').then(r=>r.json());if(r.options&&r.options.length)EXP_DESCRIPTIONS=r.options}
// Employee Name / Category — single header-level fields (not per-row), but
// the same "preset select + Custom… swaps to a text input + saved on blur"
// pattern as everywhere else. #exp-employee-wrap/#exp-category-wrap are
// stable containers that only ever get re-rendered into (never replaced
// themselves), while the actual value-holding element — a <select> normally,
// a plain <input> mid-custom-entry — always keeps the id collectExpData()
// reads (#exp-employee/#exp-category), whichever it currently is.
function expEmployeeFieldHtml(v){
  v=v||'';
  const opts='<option value="">— Select —</option>'+
    EXP_EMPLOYEES.map(m=>'<option'+(m===v?' selected':'')+'>'+escHtml(m)+'</option>').join('')
    +(v&&!EXP_EMPLOYEES.includes(v)?'<option selected>'+escHtml(v)+'</option>':'')
    +'<option value="__custom__">Custom…</option>';
  return '<select id=exp-employee onchange="onExpEmployeeChange(this)">'+opts+'</select>'}
function renderExpEmployeeField(v){$('exp-employee-wrap').innerHTML=expEmployeeFieldHtml(v)}
function onExpEmployeeChange(sel){
  if(sel.value!=='__custom__'){schedulePreview();return}
  $('exp-employee-wrap').innerHTML='<input id=exp-employee type=text placeholder="Type a name, press Enter…">';
  const inp=$('exp-employee');inp.focus();
  let done=false;
  const commit=async()=>{
    if(done)return; done=true;
    const v=inp.value.trim();
    if(!v){renderExpEmployeeField('');return}
    const r=await fetch('/api/expense-employees-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({value:v})}).then(r=>r.json());
    if(r.options)EXP_EMPLOYEES=r.options;
    renderExpEmployeeField(v);schedulePreview()};
  inp.addEventListener('keydown',e=>{
    if(e.key==='Enter'){e.preventDefault();inp.blur()}
    else if(e.key==='Escape'){done=true;renderExpEmployeeField('')}});
  inp.addEventListener('blur',commit)}
function expCategoryFieldHtml(v){
  v=v||'';
  const opts='<option value="">— Select —</option>'+
    EXP_CATEGORIES.map(m=>'<option'+(m===v?' selected':'')+'>'+escHtml(m)+'</option>').join('')
    +(v&&!EXP_CATEGORIES.includes(v)?'<option selected>'+escHtml(v)+'</option>':'')
    +'<option value="__custom__">Custom…</option>';
  return '<select id=exp-category onchange="onExpCategoryChange(this)">'+opts+'</select>'}
function renderExpCategoryField(v){$('exp-category-wrap').innerHTML=expCategoryFieldHtml(v)}
function onExpCategoryChange(sel){
  if(sel.value!=='__custom__'){schedulePreview();return}
  $('exp-category-wrap').innerHTML='<input id=exp-category type=text placeholder="Type a category, press Enter…">';
  const inp=$('exp-category');inp.focus();
  let done=false;
  const commit=async()=>{
    if(done)return; done=true;
    const v=inp.value.trim();
    if(!v){renderExpCategoryField('');return}
    const r=await fetch('/api/expense-categories-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({value:v})}).then(r=>r.json());
    if(r.options)EXP_CATEGORIES=r.options;
    renderExpCategoryField(v);schedulePreview()};
  inp.addEventListener('keydown',e=>{
    if(e.key==='Enter'){e.preventDefault();inp.blur()}
    else if(e.key==='Escape'){done=true;renderExpCategoryField('')}});
  inp.addEventListener('blur',commit)}
// Currency — a 3-way segmented control (AED/Dollar/Custom) rather than a
// dropdown, per explicit request. #exp-currency (hidden) stays the single
// source of truth collectExpData() already reads; the visible seg buttons
// and the conditional custom textbox both just drive that hidden value.
function setExpCurrency(c){
  document.querySelectorAll('#exp-currency-seg button').forEach(b=>b.classList.toggle('on',b.dataset.c===c));
  if(c==='__custom__'){
    $('exp-currency-custom').classList.remove('hide');
    $('exp-currency').value=$('exp-currency-custom').value.trim();
    $('exp-currency-custom').focus();
  }else{
    $('exp-currency-custom').classList.add('hide');
    $('exp-currency').value=c;
  }
  schedulePreview()}
function onExpCurrencyCustomInput(){$('exp-currency').value=$('exp-currency-custom').value.trim();schedulePreview()}
// Restores the seg control from a saved currency value (draft/reopen) —
// AED/USD light up their own segment, anything else is treated as a
// previously-typed custom code.
function restoreExpCurrency(v){
  v=(v||'AED').trim();
  if(v==='AED'||v==='USD'){$('exp-currency-custom').value='';setExpCurrency(v)}
  else{$('exp-currency-custom').value=v;setExpCurrency('__custom__')}}
function expPaymentMethodFieldHtml(i,it){
  const v=it.payment_method||'';
  const opts='<option value="">— Select —</option>'+
    EXP_PAYMENT_METHODS.map(m=>'<option'+(m===v?' selected':'')+'>'+escHtml(m)+'</option>').join('')
    +(v&&!EXP_PAYMENT_METHODS.includes(v)?'<option selected>'+escHtml(v)+'</option>':'')
    +'<option value="__custom__">Custom…</option>';
  return '<div class=itcardfield><label>Payment Method</label><select data-i='+i+' onchange="onExpPaymentChange('+i+',this)">'+opts+'</select></div>'}
// Same "swap the cell for a text input, save on blur" pattern as
// onUnitChange() — no native prompt()/confirm() (they silently no-op here).
function onExpPaymentChange(i,sel){
  if(sel.value!=='__custom__'){updExp(i,'payment_method',sel.value);return}
  const cell=sel.parentElement;
  cell.innerHTML='<label>Payment Method</label><input type=text placeholder="Type a method, press Enter…">';
  const inp=cell.querySelector('input');
  inp.focus();
  let done=false;
  const commit=async()=>{
    if(done)return; done=true;
    const v=inp.value.trim();
    if(!v){renderExpItems();return}
    const r=await fetch('/api/expense-payment-methods-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({value:v})}).then(r=>r.json());
    if(r.options)EXP_PAYMENT_METHODS=r.options;
    updExp(i,'payment_method',v);renderExpItems()};
  inp.addEventListener('keydown',e=>{
    if(e.key==='Enter'){e.preventDefault();inp.blur()}
    else if(e.key==='Escape'){done=true;renderExpItems()}});
  inp.addEventListener('blur',commit)}
function expProductFieldHtml(i,it){
  const v=it.product||'';
  const opts='<option value="">— Select —</option>'+
    EXP_PRODUCTS.map(m=>'<option'+(m===v?' selected':'')+'>'+escHtml(m)+'</option>').join('')
    +(v&&!EXP_PRODUCTS.includes(v)?'<option selected>'+escHtml(v)+'</option>':'')
    +'<option value="__custom__">Custom…</option>';
  return '<div class=itcardfield><label>Product</label><select data-i='+i+' onchange="onExpProductChange('+i+',this)">'+opts+'</select></div>'}
function onExpProductChange(i,sel){
  if(sel.value!=='__custom__'){updExp(i,'product',sel.value);return}
  const cell=sel.parentElement;
  cell.innerHTML='<label>Product</label><input type=text placeholder="Type a product, press Enter…">';
  const inp=cell.querySelector('input');
  inp.focus();
  let done=false;
  const commit=async()=>{
    if(done)return; done=true;
    const v=inp.value.trim();
    if(!v){renderExpItems();return}
    const r=await fetch('/api/expense-products-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({value:v})}).then(r=>r.json());
    if(r.options)EXP_PRODUCTS=r.options;
    updExp(i,'product',v);renderExpItems()};
  inp.addEventListener('keydown',e=>{
    if(e.key==='Enter'){e.preventDefault();inp.blur()}
    else if(e.key==='Escape'){done=true;renderExpItems()}});
  inp.addEventListener('blur',commit)}
function expDescriptionFieldHtml(i,it){
  const v=it.description||'';
  const opts='<option value="">— Select —</option>'+
    EXP_DESCRIPTIONS.map(m=>'<option'+(m===v?' selected':'')+'>'+escHtml(m)+'</option>').join('')
    +(v&&!EXP_DESCRIPTIONS.includes(v)?'<option selected>'+escHtml(v)+'</option>':'')
    +'<option value="__custom__">Custom…</option>';
  return '<div class="itcardfield" style="flex:2;min-width:160px"><label>Description</label><select data-i='+i+' onchange="onExpDescriptionChange('+i+',this)">'+opts+'</select></div>'}
function onExpDescriptionChange(i,sel){
  if(sel.value!=='__custom__'){updExp(i,'description',sel.value);return}
  const cell=sel.parentElement;
  cell.innerHTML='<label>Description</label><input type=text placeholder="Type a description, press Enter…">';
  const inp=cell.querySelector('input');
  inp.focus();
  let done=false;
  const commit=async()=>{
    if(done)return; done=true;
    const v=inp.value.trim();
    if(!v){renderExpItems();return}
    const r=await fetch('/api/expense-descriptions-add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({value:v})}).then(r=>r.json());
    if(r.options)EXP_DESCRIPTIONS=r.options;
    updExp(i,'description',v);renderExpItems()};
  inp.addEventListener('keydown',e=>{
    if(e.key==='Enter'){e.preventDefault();inp.blur()}
    else if(e.key==='Escape'){done=true;renderExpItems()}});
  inp.addEventListener('blur',commit)}
// Only the month/day are ever picked — the year is always the current one,
// per explicit request, so it's never even shown as an editable control.
const EXP_MONTHS=['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
const EXP_CURRENT_YEAR=new Date().getFullYear();
function expDateFieldHtml(i,it){
  const parts=(it.date||'').split('-');
  const month=parts.length===3?parts[1]:'', day=parts.length===3?String(parseInt(parts[2],10)):'';
  const opts='<option value="">—</option>'+EXP_MONTHS.map((m,mi)=>{
    const mv=String(mi+1).padStart(2,'0');
    return '<option value="'+mv+'"'+(mv===month?' selected':'')+'>'+m+'</option>'}).join('');
  return '<div class=itcardfield style="min-width:120px"><label>Date ('+EXP_CURRENT_YEAR+')</label>'+
    '<div class=expdatewrap data-i='+i+' style="display:flex;gap:4px">'+
      '<select class=expmonth onchange="onExpDateChange(this)" style="flex:1">'+opts+'</select>'+
      '<input class=expday type=number min=1 max=31 placeholder=DD value="'+escHtml(day)+'" onchange="onExpDateChange(this)" style="width:50px">'+
    '</div></div>'}
// Composes month+day into a full ISO date using the fixed current year, but
// only commits (and triggers the sort-by-date re-render) once BOTH parts are
// actually set — otherwise picking just the month would immediately sort an
// still-incomplete row to the bottom before the day's even been touched.
function onExpDateChange(el){
  const wrap=el.closest('.expdatewrap'),i=parseInt(wrap.dataset.i,10);
  const month=wrap.querySelector('.expmonth').value;
  const dayRaw=wrap.querySelector('.expday').value;
  const day=dayRaw?String(parseInt(dayRaw,10)).padStart(2,'0'):'';
  const bothSet=month&&day, bothBlank=!month&&!day;
  EXP_ITEMS[i].date=bothSet?(EXP_CURRENT_YEAR+'-'+month+'-'+day):'';
  if(bothSet||bothBlank){sortExpItemsByDate();renderExpItems();recomputeExpPeriod()}
  schedulePreview()}
function expRowHtml(it,i){
  return '<div class=itemcard>'+
    '<div class=itemcardtop>'+
      '<div style="font-weight:700;font-size:12px;color:var(--muted)">#'+(i+1)+'</div>'+
      '<button class=rm style="margin-left:auto" onclick=removeExpRow('+i+') title="Remove line">×</button>'+
    '</div>'+
    '<div class=itemcardmeta>'+
      expDateFieldHtml(i,it)+
      expProductFieldHtml(i,it)+
      expDescriptionFieldHtml(i,it)+
      expPaymentMethodFieldHtml(i,it)+
      '<div class=itcardfield><label>Amount</label><input type=number step=0.01 data-i='+i+' value="'+escHtml(it.amount??'')+'" placeholder="0.00" oninput="updExp('+i+',\'amount\',this.value)"></div>'+
    '</div>'+
  '</div>'}
function renderExpItems(){$('explist').innerHTML=EXP_ITEMS.map((it,i)=>expRowHtml(it,i)).join('')}
// Period From/To are read-only (see HEAD.EXP's 'readonly' flag) — the report's
// period is just whatever range its own line items actually span, not an
// independently-typed value, so it's always derived rather than asked for.
// ISO date strings ('YYYY-MM-DD') sort correctly as plain strings, so no
// Date-object parsing is needed here.
function recomputeExpPeriod(){
  const dates=EXP_ITEMS.map(r=>r.date).filter(Boolean).sort();
  const from=$('h-period_from'),to=$('h-period_to');
  if(from)from.value=dates.length?dates[0]:'';
  if(to)to.value=dates.length?dates[dates.length-1]:''}
// Oldest-to-newest, per explicit request — a row with no date yet (still
// mid-entry, or freshly added) sorts to the bottom rather than jumping to
// the top, so it doesn't get lost above rows that already have real dates.
function sortExpItemsByDate(){
  EXP_ITEMS.sort((a,b)=>{
    if(!a.date&&!b.date)return 0;
    if(!a.date)return 1;
    if(!b.date)return -1;
    return a.date.localeCompare(b.date)})}
// Category stays editable (dropdown + Custom…, per earlier request) but is
// kept in sync with whichever Products are actually in the line items —
// only fires on a Product change, and skipped while the user is mid-typing
// a custom category, so it never fights an unrelated in-progress edit.
function recomputeExpCategory(){
  if(document.getElementById('exp-category')?.tagName==='INPUT')return;
  const products=[...new Set(EXP_ITEMS.map(r=>r.product).filter(Boolean))];
  renderExpCategoryField(products.map(p=>p.toUpperCase()).join(', '))}
function addExpRow(){EXP_ITEMS.push({});renderExpItems();schedulePreview()}
function removeExpRow(i){EXP_ITEMS.splice(i,1);if(!EXP_ITEMS.length)EXP_ITEMS.push({});renderExpItems();recomputeExpPeriod();recomputeExpCategory();schedulePreview()}
function updExp(i,k,v){
  // Date changes go through onExpDateChange() instead (see expDateFieldHtml) —
  // it needs to know when BOTH month and day are set before recomputing
  // anything, which a single generic key/value setter can't express.
  EXP_ITEMS[i][k]=v;
  if(k==='product')recomputeExpCategory();
  schedulePreview()}
function resetExpForm(){
  EXP_ITEMS=[{}];
  renderExpEmployeeField('');
  renderExpCategoryField('');
  $('exp-currency-custom').value='';
  setExpCurrency('AED');
  renderExpItems();
  recomputeExpPeriod()}
function collectExpData(){
  const from=$('h-period_from')?$('h-period_from').value:'',to=$('h-period_to')?$('h-period_to').value:'';
  return {
    company:$('exp-employee').value.trim(),
    project:(from||to)?(from+(from&&to?'_to_':'')+to):'',
    category:$('exp-category').value.trim(),
    currency:$('exp-currency').value.trim()||'AED',
    rows:EXP_ITEMS.filter(r=>r.product||r.description||r.amount)
  }}

const FOLDER_KEYS={INV:'inv_folder',DO:'do_folder',QTN2:'qtn2_folder',PI:'pi_folder',RV:'rv_folder',CN:'cn_folder',CAT:'catalogue_folder',EXP:'expense_folder'};
async function loadCfg(){const c=await fetch('/api/config').then(r=>r.json());
  if(c[FOLDER_KEYS[TYPE]]){nextNumber();onCompany()}}
async function nextNumber(force){const r=await fetch('/api/next-number?type='+TYPE).then(r=>r.json());const el=$('h-number');if(el&&(force||!el.value))el.value=String(r.next).padStart(4,'0')}

function companyVal(){return TYPE==='QTN2'?richText($('company-rich')):$('company').value.trim()}
function setCompanyVal(v){$('company').value=v;$('company-rich').textContent=v}
// Company is the link between Build and the Clients database: typing or
// picking a name that exactly matches a saved client (case-insensitive)
// auto-fills Attn/Address from that record, same as the dedicated "Client…"
// picker button does — but never overwrites text the user already typed,
// so this can't clobber a manual edit mid-way through filling the form.
function setBuildAddressExtras(c){
  $('customer_pobox').value=c.po_box||'';
  $('customer_city').value=c.city||'';
  setCountryValue('customer',c.country||'')}
function fillFromClientIfBlank(name){
  if(!name)return;
  const c=CLIENT_RECORDS.find(x=>x.name.trim().toLowerCase()===name.trim().toLowerCase());
  if(!c)return;
  const attnBox=$('customer_attn'),addrBox=$('customer_address');
  if(attnBox&&!attnBox.textContent.trim()&&c.attn){attnBox.textContent=c.attn}
  if(addrBox&&!addrBox.textContent.trim()&&c.address){addrBox.textContent=c.address}
  if(!$('customer_pobox').value.trim()&&!$('customer_city').value.trim()&&!$('customer-country').value)setBuildAddressExtras(c);
}
async function onCompany(){const co=companyVal();$('histslot').innerHTML='';
  fillFromClientIfBlank(co);
  if(!co)return;
  const r=await fetch('/api/previous?type='+TYPE+'&company='+encodeURIComponent(co)).then(r=>r.json());
  if(!r.previous||!r.previous.length)return;
  $('histslot').innerHTML='<div class=hist><div class=hh><span class=glow></span><b>Last '+r.previous.length+' '+LABEL[TYPE].toLowerCase()+'(s) for '+co+'</b></div>'+
    r.previous.map(p=>'<div class=hc><div class=top><span class=mono>'+p.type+' '+p.number+' · R'+p.rev+'</span><span class=muted>'+p.date+'</span></div>'+
      (p.items||[]).slice(0,6).map(it=>'<div class=hi><span>'+(it.description||'')+'</span></div>').join('')+'</div>').join('')+'</div>'}

let previewImgToken=0;
function setPreviewImage(src){
  const img=$('previewimg'),empty=$('previewempty'),pages=$('previewpages');
  if(!img)return;
  const token=++previewImgToken;
  previewPages=0;pages.classList.add('hide');pages.innerHTML='';updatePvButtons();
  // Guarded by `token`: onload/onerror stay attached to this <img> element
  // across calls, so a slow/late-arriving one from an OLDER setPreviewImage()
  // call could otherwise fire after a newer showPreviewPages() already
  // rendered the correct pages — silently re-showing "Could not render a
  // preview." underneath a perfectly good preview. Only the most recent
  // call's callbacks are allowed to touch the DOM.
  img.onload=()=>{if(token!==previewImgToken)return;img.classList.remove('hide');empty.classList.add('hide')};
  img.onerror=()=>{if(token!==previewImgToken)return;img.classList.add('hide');empty.textContent='Could not render a preview.';empty.classList.remove('hide')};
  img.src=src}

let previewPages=0, previewMode='double', previewModeAuto=true, previewZoom=100, previewCacheBust=0;
// Fit mode — a genuinely different viewing mode from the zoom-based one
// below, not just a zoom preset. Reported directly: "Fit" only ever fit
// the page's WIDTH into the pane (computePageWidth's own zoom math has no
// idea how tall the pane actually is), so a full A4 page routinely still
// needed vertical scrolling right after clicking the button whose whole
// job is "make the document fully visible". previewFitMode:true switches
// renderPreviewPages over to computeFitPageWidth (fits BOTH dimensions,
// see that function) and to showing exactly one page at a time
// (previewFitIndex, 1-based) instead of the normal continuous scroll of
// every page stacked/grouped in .pvpages — per the same request, "when
// the user scrolls... it will move page by page not like scrolling".
// Left false by default so nothing about the pre-existing zoom/scroll
// behavior changes for anyone who never touches the Fit button; any
// manual zoom or Single/Double click (zoomPreview/setPreviewMode below)
// explicitly drops back out of it, since both mean "I want the old
// zoom+scroll behavior now", not "keep fitting, but also zoom".
let previewFitMode=false, previewFitIndex=1;
// A4 portrait height/width (297mm/210mm) — every doc-page template this
// app renders declares size="a4" (confirmed by grepping every
// templates_html/*.html for <doc-page size=), so one constant covers all
// of them; a future non-A4 template would need this made per-document.
const PREVIEW_PAGE_ASPECT=297/210;
function effectivePreviewMode(){return (previewMode==='double'&&previewPages>1)?'double':'single'}
function previewPerRow(){return effectivePreviewMode()==='double'?2:1}
// Fits ONE page fully inside the pane on both axes — computePageWidth
// below only ever solves for width (then lets height fall out of the
// page's own aspect ratio, unbounded), which is exactly the bug Fit
// mode exists to fix. Whichever axis is more constraining wins: a wide-
// but-short pane caps on height, a narrow-but-tall one caps on width.
function computeFitPageWidth(){
  const box=$('previewbox');if(!box)return 600;
  const padding=32; // matches .previewpane's 16px padding on every side
  const wFit=box.clientWidth-padding;
  const hFit=(box.clientHeight-padding)/PREVIEW_PAGE_ASPECT;
  return Math.max(160,Math.round(Math.min(wFit,hFit)))}
function computePageWidth(){
  const box=$('previewbox');if(!box)return 600;
  const inner=box.clientWidth-32;
  const perRow=previewPerRow();
  const gap=14;
  let w=(inner-gap*(perRow-1))/perRow;
  w=Math.min(w,perRow===2?620:900);
  return Math.max(160,Math.round(w*(previewZoom/100)))}
function updatePvButtons(){
  const eff=effectivePreviewMode();
  // Fit mode always reads as Single (it only ever shows one page), and
  // its own zoom is whatever computeFitPageWidth lands on — not a
  // percentage the +/- buttons understand — so the label says "Fit"
  // instead of a number, and pagecount switches to a "Page N of M"
  // position readout (there's no other way to tell which page you're on
  // once scrolling no longer moves the page, just swaps it).
  $('pm-single').classList.toggle('on',previewFitMode||eff==='single');
  $('pm-double').classList.toggle('on',!previewFitMode&&eff==='double');
  $('pm-double').disabled=previewPages<2;
  $('zoomlabel').textContent=previewFitMode?'Fit':(previewZoom+'%');
  $('pagecount').textContent=previewPages?(previewFitMode?('Page '+previewFitIndex+' of '+previewPages):(previewPages+(previewPages===1?' page':' pages'))):''}
// Pages are grouped into explicit .pvrow divs (perRow images each) rather
// than letting .pvpages flex-wrap them freely — with free-wrapping, how
// many pages land on one visual row depended on each page's zoomed pixel
// width vs the container's width, so zooming out far enough always
// eventually let extra pages squeeze onto a row regardless of the
// Single/Double selection. Grouping in markup instead means row membership
// is structural, not a width coincidence, so it holds at any zoom level.
function renderPreviewPages(){
  const wrap=$('previewpages');if(!wrap)return;
  wrap.classList.remove('single','double');
  if(!previewPages){wrap.innerHTML='';updatePvButtons();return}
  if(previewFitMode){
    // One page, fit to both dimensions, no grouping needed — see
    // computeFitPageWidth and the wheel listener below (previewFitStep)
    // for the "scroll moves page by page" half of this.
    wrap.classList.add('single');
    previewFitIndex=Math.max(1,Math.min(previewPages,previewFitIndex));
    const fw=computeFitPageWidth();
    wrap.innerHTML='<div class=pvrow><img class=pvpage style="width:'+fw+'px" src="/draft-preview?page='+previewFitIndex+'&t='+previewCacheBust+'"></div>';
    updatePvButtons();
    return}
  wrap.classList.add(effectivePreviewMode());
  const w=computePageWidth();
  const perRow=previewPerRow();
  let html='';
  for(let i=1;i<=previewPages;i+=perRow){
    html+='<div class=pvrow>';
    for(let p=i;p<Math.min(i+perRow,previewPages+1);p++)html+='<img class=pvpage style="width:'+w+'px" src="/draft-preview?page='+p+'&t='+previewCacheBust+'">';
    html+='</div>'}
  wrap.innerHTML=html;
  updatePvButtons()}
// Both explicitly drop out of Fit mode — asking for a zoom percentage or
// a Single/Double layout is asking for the old zoom+scroll behavior back,
// not "keep fitting AND change this too".
function setPreviewMode(m){previewFitMode=false;previewMode=m;previewModeAuto=false;renderPreviewPages()}
function zoomPreview(delta){previewFitMode=false;previewZoom=Math.max(40,Math.min(300,previewZoom+delta));renderPreviewPages()}
function fitPreview(){previewFitMode=true;renderPreviewPages()}
// Wheel-triggered page step while Fit mode is active — see previewFitMode's
// own comment above. previewbox has nothing to natively scroll in this
// mode anyway (one page, fully visible on both axes by construction), so
// hijacking every non-Ctrl wheel event here doesn't fight the browser for
// anything; Ctrl+scroll still reaches initPanBox's own zoom handler
// further down (unaffected — this listener only ever preventDefaults a
// plain wheel).
function previewFitStep(dir){
  if(!previewPages)return;
  previewFitIndex=Math.max(1,Math.min(previewPages,previewFitIndex+dir));
  renderPreviewPages()}

// Hand-tool pan (drag to scroll, hold Space to pan temporarily, Ctrl+scroll
// to zoom) — a self-contained factory, one instance per preview pane, so
// the per-document preview and the Full Catalog Builder's own preview each
// get an independent hand-tool toggle without fighting over shared state.
// Registers its own window-level listeners rather than a shared registry;
// the box.offsetParent===null check keeps a hidden pane's Space/mousemove
// handlers from reacting while some other view is the one actually on
// screen (views stay in the DOM, just .hide-toggled, when not active).
function initPanBox(boxId, handBtnId, onWheelZoom){
  const box=$(boxId);if(!box)return null;
  let panToolOn=false, spacePanning=false, isPanning=false, panStartX=0, panStartY=0, panScrollLeft=0, panScrollTop=0;
  function updateCursor(){box.classList.toggle('pan-ready',panToolOn||spacePanning)}
  function toggle(){panToolOn=!panToolOn;$(handBtnId).classList.toggle('on',panToolOn);updateCursor()}
  box.addEventListener('mousedown',e=>{
    if(!(panToolOn||spacePanning)||e.button!==0)return;
    isPanning=true;panStartX=e.clientX;panStartY=e.clientY;
    panScrollLeft=box.scrollLeft;panScrollTop=box.scrollTop;
    box.classList.add('panning');document.body.style.userSelect='none';
    e.preventDefault()});
  window.addEventListener('mousemove',e=>{
    if(!isPanning)return;
    box.scrollLeft=panScrollLeft-(e.clientX-panStartX);
    box.scrollTop=panScrollTop-(e.clientY-panStartY)});
  window.addEventListener('mouseup',()=>{
    if(!isPanning)return;
    isPanning=false;box.classList.remove('panning');document.body.style.userSelect=''});
  window.addEventListener('keydown',e=>{
    if(e.code!=='Space'||spacePanning||box.offsetParent===null)return;
    const el=document.activeElement;
    if(/INPUT|TEXTAREA|SELECT/.test(el.tagName)||el.isContentEditable)return;
    spacePanning=true;updateCursor();e.preventDefault()});
  window.addEventListener('keyup',e=>{
    if(e.code!=='Space')return;
    spacePanning=false;if(!isPanning)updateCursor()});
  if(onWheelZoom)box.addEventListener('wheel',e=>{
    if(!e.ctrlKey)return;
    e.preventDefault();onWheelZoom(e.deltaY<0?10:-10)},{passive:false});
  return {toggle}}
const buildPan=initPanBox('previewbox','pm-hand',zoomPreview);
function togglePanTool(){if(buildPan)buildPan.toggle()}
const fcPan=initPanBox('fc-previewbox','fc-pm-hand',fcZoomPreview);
function fcTogglePanTool(){if(fcPan)fcPan.toggle()}
// previewFitStep's own trigger — see that function's comment. Bound here
// (own listener, not folded into initPanBox's) since it fires on every
// plain wheel tick while Fit mode is on, unconditionally, unlike
// initPanBox's Ctrl-gated zoom handler on the same element.
(function(){
  const box=$('previewbox');if(!box)return;
  box.addEventListener('wheel',e=>{
    if(!previewFitMode||e.ctrlKey)return;
    e.preventDefault();
    previewFitStep(e.deltaY<0?-1:1)},{passive:false})
})();
function showPreviewPages(n){
  const img=$('previewimg'),empty=$('previewempty'),pages=$('previewpages');
  previewImgToken++;  // invalidate any in-flight setPreviewImage() callback
  img.classList.add('hide');img.src='';
  previewPages=n;previewCacheBust=Date.now();
  if(previewFitIndex>n)previewFitIndex=Math.max(1,n);
  if(previewModeAuto)previewMode=n>1?'double':'single';
  empty.classList.add('hide');
  pages.classList.remove('hide');
  renderPreviewPages()}

// Switching tools (or resuming a draft) mid-render used to be able to show
// the WRONG tool's preview, or "freeze" on a stale one: runPreview() reads
// TYPE fresh at send-time, so each request's payload is always correct for
// whatever was active the moment it was sent — but the server (Flask's dev
// server here, single-threaded/un-pooled — see the plain app.run() at the
// bottom of this file) processes requests strictly one at a time, so a slow
// render for a tool the user has since clicked away from can still be
// sitting in flight when a newer, faster one for the CURRENT tool finishes
// and renders first. Without a guard, that slow response lands last and
// unconditionally overwrites the correct preview with a stale one. previewSeq
// is bumped every time a preview is requested; only the response that still
// matches the latest bump is ever allowed to touch the screen — any older
// one is silently dropped, whatever order the responses actually arrive in.
let previewTimer=null, previewSeq=0;
function schedulePreview(){clearTimeout(previewTimer);previewTimer=setTimeout(runPreview,700)}
function fieldVal(el){return el.classList.contains('richbox')?richText(el):el.value}
function setFieldVal(el,v){if(el.classList.contains('richbox'))el.innerHTML=v??'';else el.value=v??el.value}
async function runPreview(){
  clearTimeout(previewTimer);
  const mySeq=++previewSeq;
  // Dim (not clear) whatever's already on screen while the new one renders —
  // clearing it here would flash an empty pane on every keystroke-triggered
  // re-render, and leaving it fully undimmed with no rendering happening
  // looks identical to the render being stuck/ignored, which was half of
  // what made the underlying bug so confusing to notice.
  const pages=$('previewpages');
  const wasShowingPages=pages&&!pages.classList.contains('hide');
  if(wasShowingPages)pages.classList.add('pv-loading');
  const data={doc_type:TYPE,items:items.filter(x=>x.description),company:companyVal()};
  HEAD[TYPE].forEach(f=>{const el=$('h-'+f[0]);if(el)data[f[0]]=fieldVal(el)});
  if(TYPE!=='DO')Object.assign(data,collectDiscVat());
  if(TYPE==='QTN2')Object.assign(data,collectQtn2Extra());else data.customer_block=customerBlockForXlsx();
  if(TYPE==='CAT')Object.assign(data,collectCatData());
  if(TYPE==='EXP')Object.assign(data,collectExpData());
  let r;
  try{
    r=await fetch('/api/preview-draft',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(data)}).then(r=>r.json());
  }catch(e){
    if(mySeq===previewSeq&&pages)pages.classList.remove('pv-loading');
    return}
  if(mySeq!==previewSeq)return;  // superseded by a newer request — never apply a stale response
  if(pages)pages.classList.remove('pv-loading');
  if(!r.error)showPreviewPages(r.pages||1)}

function collectDocData(){
  const data={doc_type:TYPE,items:items.filter(x=>x.description),company:companyVal()};
  HEAD[TYPE].forEach(f=>data[f[0]]=fieldVal($('h-'+f[0])));
  if(TYPE!=='DO')Object.assign(data,collectDiscVat());
  if(TYPE==='QTN2')Object.assign(data,collectQtn2Extra());else data.customer_block=customerBlockForXlsx();
  if(TYPE==='CAT')Object.assign(data,collectCatData());
  if(TYPE==='EXP')Object.assign(data,collectExpData());
  return data}
async function generate(){
  const data=collectDocData();data.replace=EDITING||'';
  if(!data.company){alert(TYPE==='CAT'?'Enter the product name.':TYPE==='EXP'?'Enter the employee name.':'Enter the company name.');return null}
  if(!data.number){alert('Enter a document number.');return null}
  if(TYPE==='CAT'&&!EDITING){
    const existingRel=await catProductNameExists(data.company);
    if(existingRel){
      // A product with this exact name has already been generated before.
      // From a brand-new/unsaved form this is almost always an accidental
      // name collision — names must stay unique — so it's still a hard
      // block, same as always. From a RESUMED DRAFT, though, it's just as
      // likely the real intent: the user picked this draft back up
      // specifically to regenerate/update that same product. Instead of a
      // dead-end "choose a different name" there too, offer to replace the
      // existing file — but only after an explicit confirm modal, never
      // silently: this permanently overwrites a real generated PDF (+ its
      // saved sidecar/markdown), no undo, no Recycle Bin (os.remove — see
      // /api/file-op's own comment on that same fact). window.confirm()
      // silently no-ops in this app's embedded webview (see fmDelete's own
      // comment), hence a real modal (#catreplacemodal) instead.
      if(!EDITING_DRAFT){
        alert('A Sololuce Datasheet named "'+data.company+'" already exists. Product names must be unique — choose a different name.');
        return null}
      const ok=await askCatReplaceConfirm(data.company);
      if(!ok)return null;
      data.replace=existingRel}}
  toast('Generating…');
  const r=await fetch('/api/generate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(data)}).then(r=>r.json());
  if(r.error){alert(r.error);return r}
  EDITING=r.xlsx||null;
  // a draft's job is done once it becomes a real generated document
  if(EDITING_DRAFT){deleteDraft(EDITING_DRAFT);EDITING_DRAFT=null}
  setPreviewImage('/cs-thumb?f='+encodeURIComponent(r.pdf||r.xlsx));
  toast('Saved '+(r.xlsx?r.xlsx+' + PDF':r.pdf));
  loadClients();
  return {...r,data}}

// ---------------------------------------------------------------- drafts (save in-progress work on ANY doc type before Generate)
let DRAFTS=[];
async function saveDraftFromForm(){
  const data=collectDocData();
  if(!data.company){alert('Enter the company name first.');return}
  const body={id:EDITING_DRAFT||'',doc_type:TYPE,data};
  const r=await fetch('/api/drafts',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  EDITING_DRAFT=r.draft.id;
  closeDraftsPicker();
  toast('Draft saved — reopen it anytime from the Drafts button')}
async function openDraftsPicker(ev){
  ev.stopPropagation();
  const box=$('draftspicker');
  if(box.style.display==='block'){box.style.display='none';return}
  const r=ev.currentTarget.getBoundingClientRect();
  const left=Math.min(Math.round(r.left),window.innerWidth-296);
  box.style.left=Math.max(8,left)+'px';box.style.top=(Math.round(r.bottom)+6)+'px';
  box.style.display='block';
  $('dp-search').value='';
  const res=await fetch('/api/drafts').then(r=>r.json());
  DRAFTS=res.drafts||[];
  renderDraftsPicker();
  $('dp-search').focus()}
function closeDraftsPicker(){$('draftspicker').style.display='none'}
document.addEventListener('click',e=>{
  if($('draftspicker').style.display==='block'&&!e.target.closest('#draftspicker')&&!e.target.closest('#draftsbtn'))closeDraftsPicker()});
function renderDraftsPicker(){
  const q=($('dp-search').value||'').trim().toLowerCase();
  const rows=DRAFTS.filter(d=>!q||(d.label||'').toLowerCase().includes(q)).sort((a,b)=>(b.updated||'').localeCompare(a.updated||''));
  $('dp-list').innerHTML=rows.length?rows.map(d=>
    '<div class=cpitem style="justify-content:space-between" onclick="loadDraft(\''+d.id+'\')">'+
      '<span class=cpname>'+escHtml(d.label)+'</span>'+
      '<button class=rm onclick="event.stopPropagation();deleteDraft(\''+d.id+'\')" title="Delete draft">✕</button>'+
    '</div>'
  ).join(''):'<p class="muted" style="font-size:12px;padding:6px 4px;margin:0">No drafts saved yet.</p>'}
async function deleteDraft(id){
  const r=await fetch('/api/drafts-delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})}).then(r=>r.json());
  DRAFTS=r.drafts||[];
  renderDraftsPicker();
  if(EDITING_DRAFT===id)EDITING_DRAFT=null}
// Autosave-on-leave: if the user switches away from the tab or closes the
// window mid-edit, silently persist the current form (any doc type) as a
// draft, so nothing is ever lost by just walking away. Skipped entirely if
// there's nothing meaningful to save yet (no company/product name), same
// gate as the manual Save-as-Draft button.
// visibilitychange fires when the tab is merely switched away (the page is
// still fully alive), so a normal fetch is used there — its response can
// still update EDITING_DRAFT. pagehide fires for an actual close/navigate,
// where a normal fetch isn't reliably given time to complete, so that one
// uses sendBeacon instead, which is purpose-built to survive it.
async function autosaveDraft(){
  const data=collectDocData();
  if(!data.company)return;
  const body={id:EDITING_DRAFT||'',doc_type:TYPE,data};
  const r=await fetch('/api/drafts',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)}).then(r=>r.json());
  if(r.draft)EDITING_DRAFT=r.draft.id}
function autosaveDraftBeacon(){
  const data=collectDocData();
  if(!data.company)return;
  const body={id:EDITING_DRAFT||'',doc_type:TYPE,data};
  navigator.sendBeacon('/api/drafts',new Blob([JSON.stringify(body)],{type:'application/json'}))}
document.addEventListener('visibilitychange',()=>{if(document.visibilityState==='hidden')autosaveDraft()});
window.addEventListener('pagehide',autosaveDraftBeacon);
// Shown once on launch (see checkUnfinishedDraftsOnLaunch in the init chain
// below) if the current brand has any saved drafts — autosaved or manual,
// same list either way. Uses the .clientmodal centered-dialog pattern (not
// #draftspicker's anchored popover) since there's no triggering button click
// to anchor off of at page-load time.
let RESUME_DRAFTS=[];
function renderResumeDraftsList(){
  RESUME_DRAFTS.sort((a,b)=>(b.updated||'').localeCompare(a.updated||''));
  $('resumedrafts-list').innerHTML=RESUME_DRAFTS.map(d=>
    '<div class=cpitem style="justify-content:space-between" onclick="resumeDraft(\''+d.id+'\')">'+
      '<span class=cpname>'+escHtml(d.label)+'</span>'+
      '<button class=rm onclick="event.stopPropagation();deleteDraftFromResume(\''+d.id+'\')" title="Delete draft">✕</button>'+
    '</div>'
  ).join('')}
async function checkUnfinishedDraftsOnLaunch(){
  const res=await fetch('/api/drafts').then(r=>r.json());
  RESUME_DRAFTS=res.drafts||[];
  if(!RESUME_DRAFTS.length)return;
  renderResumeDraftsList();
  $('resumedraftsmodal').classList.remove('hide')}
function resumeDraft(id){$('resumedraftsmodal').classList.add('hide');DRAFTS=RESUME_DRAFTS;loadDraft(id)}
async function deleteDraftFromResume(id){
  const r=await fetch('/api/drafts-delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})}).then(r=>r.json());
  RESUME_DRAFTS=r.drafts||[];
  if(!RESUME_DRAFTS.length){closeResumeDrafts();return}
  renderResumeDraftsList()}
function closeResumeDrafts(){$('resumedraftsmodal').classList.add('hide')}
function restoreDiscVat(discount,vat){
  discount=discount||{enabled:false};vat=vat||{enabled:false};
  $('disc-on').checked=!!discount.enabled;
  if(discount.enabled&&discount.mode==='target'){discMode='target';$('disc-value').value=discount.value||''}
  else{discMode='amount';$('disc-value').value=discount.enabled?(discount.mode==='percent'?(discount.value+'%'):String(discount.value)):''}
  setDiscMode(discMode);
  $('vat-on').checked=!!vat.enabled;
  $('vat-value').value=vat.enabled?(vat.mode==='percent'?(vat.value+'%'):String(vat.value)):'5%';
  onDiscVatChange()}
// The three functions below populate the Build form from a plain data blob
// shaped exactly like collectDocData()'s output — shared by loadDraft() (a
// saved in-progress draft) and openDoc() (a real generated document's own
// JSON sidecar, opened for editing from All Docs), since both hand the form
// the same shape of data and need the same restore logic. Split by doc type
// because CAT and EXP's field shapes (badges/specs/ordering table; expense
// rows/employee/currency) share nothing with the generic HEAD+items model
// QTN2/INV/DO/legacy-QTN use.
function populateCatForm(data){
  // Mirrors collectCatData()'s shape exactly (the save side already captures
  // all of this). HEAD.CAT (Sheet No./Date) first — collectCatData() itself
  // doesn't cover these (they're merged in separately by collectDocData()),
  // so without this an opened/resumed sheet keeps whatever number happened
  // to already be in the field rather than the one it was actually saved
  // with — silently re-Generating as a different-numbered file instead of
  // overwriting the original.
  HEAD.CAT.forEach(f=>{const el=$('h-'+f[0]);if(el)setFieldVal(el,data[f[0]]||'')});
  // .toUpperCase() here too, so a draft/document saved before this rule
  // existed shows correctly-cased on reopen, not just on the next save.
  $('cat-productname').value=(data.product_name||data.company||'').toUpperCase();
  renderCatSeriesField(data.series||data.project||'');
  $('cat-family-enabled').checked=!!(data.family||'').trim();
  $('cat-family-details').classList.toggle('hide',!$('cat-family-enabled').checked);
  renderCatFamilyField(data.family||'');
  $('cat-producttype').value=data.product_type||'';
  if(data.page_number)$('cat-pagenum').value=data.page_number;else loadCatNextPage();
  $('cat-description').value=data.description||'';
  $('cat-note').value=data.note||CAT_DEFAULT_NOTE;
  $('cat-ordcode').value=data.ordering_code_example||'';
  CAT_BADGES=(data.badges&&data.badges.length)?data.badges:CAT_STANDARD_BADGE_KEYS.map(key=>({key}));
  CAT_SPECS=(data.specs&&data.specs.length)
    ?data.specs.map(s=>({label:s.label,values:(s.value||'').split('\n')}))
    :CAT_DEFAULT_SPEC_LABELS.map(label=>({label,values:['']}));
  CAT_FINISH=data.finish_colors||[];
  CAT_COL_SEQ=0;CAT_ORD_COL_CLIPBOARD=null;CAT_ORD_CURRENT_VARIANT=0;CAT_ORD_ALIGN_ROWS=!!data.ordering_align_rows;
  CAT_ORD_COLS=(data.ordering_columns&&data.ordering_columns.length)
    ?data.ordering_columns.map(c=>({key:'col'+(CAT_COL_SEQ++),label:c.label,values:(c.values&&c.values.length)?c.values:[''],width:(typeof c.width==='number'&&c.width>0)?c.width:null}))
    :CAT_DEFAULT_ORD_LABELS.map(label=>({key:'col'+(CAT_COL_SEQ++),label,values:['']}));
  CAT_IMG={
    // placeholder: data.X_placeholder??true — a saved draft from before this
    // checkbox existed has no such field at all (undefined), which must
    // read back as "on" (today's long-standing look), same default
    // _photo_ctx uses server-side — only an explicit `false` turns it off.
    main:{src:data.main_photo||'',zoom:data.main_photo_zoom||1,x:data.main_photo_x??50,y:data.main_photo_y??50,mask:data.main_photo_mask??100,placeholder:data.main_photo_placeholder??true},
    // Application Photo's mask is locked at 100 (see CAT_IMG_MASK_LOCKED) —
    // ignores whatever a pre-lock document had saved, same reasoning as
    // openPhotoAdjust forcing it live.
    lifestyle:{src:data.lifestyle_photo||'',zoom:data.lifestyle_photo_zoom||1,x:data.lifestyle_photo_x??50,y:data.lifestyle_photo_y??50,mask:100,placeholder:data.lifestyle_photo_placeholder??true},
    diagram:{src:data.dimension_diagram||'',zoom:data.dimension_diagram_zoom||1,x:data.dimension_diagram_x??50,y:data.dimension_diagram_y??50,mask:data.dimension_diagram_mask??100,maskAnchorX:data.dimension_diagram_mask_anchor_x??100,maskAnchorY:data.dimension_diagram_mask_anchor_y??100,label:data.dimension_diagram_label||'',placeholder:data.dimension_diagram_placeholder??true},
    extra1:{src:data.extra_photo_1||'',zoom:data.extra_photo_1_zoom||1,x:data.extra_photo_1_x??50,y:data.extra_photo_1_y??50,mask:data.extra_photo_1_mask??100,maskAnchorX:data.extra_photo_1_mask_anchor_x??100,maskAnchorY:data.extra_photo_1_mask_anchor_y??100,label:data.extra_photo_1_label||'',show:!!data.extra_photo_1_show,merged:!!data.extra_photo_1_merged,autosize:!!data.extra_photo_1_autosize,placeholder:data.extra_photo_1_placeholder??true},
    extra2:{src:data.extra_photo_2||'',zoom:data.extra_photo_2_zoom||1,x:data.extra_photo_2_x??50,y:data.extra_photo_2_y??50,mask:data.extra_photo_2_mask??100,maskAnchorX:data.extra_photo_2_mask_anchor_x??100,maskAnchorY:data.extra_photo_2_mask_anchor_y??100,label:data.extra_photo_2_label||'',show:!!data.extra_photo_2_show,placeholder:data.extra_photo_2_placeholder??true},
    extra3:{src:data.extra_photo_3||'',zoom:data.extra_photo_3_zoom||1,x:data.extra_photo_3_x??50,y:data.extra_photo_3_y??50,mask:data.extra_photo_3_mask??100,maskAnchorX:data.extra_photo_3_mask_anchor_x??100,maskAnchorY:data.extra_photo_3_mask_anchor_y??100,label:data.extra_photo_3_label||'',merged:!!data.extra_photo_3_merged,placeholder:data.extra_photo_3_placeholder??true}
  };
  renderCatBadges();renderCatSpecs();renderCatFinish();renderCatOrdTable();renderCatImages()}
function populateExpForm(data){
  // Mirrors collectExpData()'s shape — its own function rather than folding
  // into the generic HEAD+items path below, same reasoning as CAT above.
  HEAD.EXP.forEach(f=>{const el=$('h-'+f[0]);if(el)setFieldVal(el,data[f[0]]||'')});
  renderExpEmployeeField(data.company||'');
  // Category is preserved as-saved (not recomputed) — it can legitimately be
  // unrelated to any one row's Product, unlike Period below.
  renderExpCategoryField(data.category||'');
  restoreExpCurrency(data.currency);
  EXP_ITEMS=(data.rows&&data.rows.length)?data.rows:[{}];
  // Both self-heal data saved before either became automatic, same reasoning
  // throughout: items are the source of truth, not whatever was separately
  // saved alongside them.
  sortExpItemsByDate();
  renderExpItems();
  recomputeExpPeriod()}
function populateGenericDocForm(doc_type,data){
  HEAD[doc_type].forEach(f=>{const el=$('h-'+f[0]);if(el)setFieldVal(el,data[f[0]]||'')});
  setCompanyVal(data.company||'');
  items=(data.items&&data.items.length)?data.items:[{}];
  SELECTED_ITEMS.clear();
  renderItems();onCompany();
  items.forEach((it,i)=>{if(!it.photo)matchPhotoForItem(i);matchDatasheetsForItem(i)});
  if(doc_type==='QTN2'){
    $('customer_attn').innerHTML=data.customer_attn||'';
    $('customer_address').innerHTML=data.customer_address_raw||'';
    $('customer_pobox').value=data.customer_po_box||'';
    $('customer_city').value=data.customer_city||'';
    setCountryValue('customer',data.customer_country||'');
    setQtn2Status(data.status||'Draft');
    const tui=data.terms_ui||{};
    renderTermsDeliveryDefault();
    if(tui.delivery)$('terms-delivery-wrap').innerHTML=termsDeliverySelectHtml(tui.delivery);
    TERMS_PAYMENT=(tui.payment&&tui.payment.length)?tui.payment:[{percent:50,label:'Advance'},{percent:50,label:'Upon Delivery'}];
    renderPaymentStages();
    setWarranty(tui.warranty||'5');
  }else{
    // customerBlockForXlsx() always composes company first — strip it back
    // off so re-Generating doesn't duplicate it.
    $('customer_attn').innerHTML='';
    let block=data.customer_block||'';
    const lines=block.split('\n');
    if(lines[0]&&lines[0].trim().toLowerCase()===(data.company||'').trim().toLowerCase())block=lines.slice(1).join('\n');
    $('customer_address').innerHTML=escHtml(block).replace(/\n/g,'<br>');
    $('customer_pobox').value='';$('customer_city').value='';setCountryValue('customer','');
  }
  if(doc_type!=='DO')restoreDiscVat(data.discount,data.vat)}
function loadDraft(id){
  const d=DRAFTS.find(x=>x.id===id);if(!d)return;
  closeDraftsPicker();
  const data=d.data||{};
  view(DOC_VIEWS[d.doc_type]||DOC_VIEWS.QTN2);
  setType(d.doc_type,true);
  EDITING=null;EDITING_DRAFT=d.id;
  if(d.doc_type==='CAT')populateCatForm(data);
  else if(d.doc_type==='EXP')populateExpForm(data);
  else populateGenericDocForm(d.doc_type,data);
  $('title').textContent='Draft: '+LABEL[d.doc_type]+(data.company?' — '+data.company:'');
  runPreview();  // resuming a draft is a discrete switch too — see setType's own runPreview() call for why this skips the typing-debounce
  toast('Loaded draft — Generate to save it as a real document')}

// Opening an existing document ("Open in CS") is a distinct mode from
// building a brand-new one — it takes over the screen in its own
// full-window overlay (#editmodal) rather than hijacking the Build tab,
// so there's never a dead-end where the type tabs show nothing selected
// and there's no way back to a normal new-document state. Closing it just
// restores whatever view (All Docs, Build, etc.) was showing underneath.
let EDIT_MODE=false;
let EDIT_SNAPSHOT=null;
function enterDocEditMode(label){
  EDIT_MODE=true;
  $('editmodaltitle').textContent=label;
  $('editmodalbody').appendChild($('buildwrap'));
  $('modebar').classList.add('hide');
  $('editmodal').classList.remove('hide');
  // Snapshot AFTER the caller has finished populating every field, so the
  // diff shown before saving only reflects the user's own edits — not the
  // load itself. Deep-cloned: collectDocData()'s items array is filtered
  // from the live `items` array but holds the SAME item objects, which
  // upd() mutates in place — without cloning, the "original" snapshot
  // would silently drift to match every edit as it happens.
  EDIT_SNAPSHOT=JSON.parse(JSON.stringify(collectDocData()))}
function exitEditMode(){
  EDIT_MODE=false;
  EDIT_SNAPSHOT=null;
  $('v-build').insertBefore($('buildwrap'),$('v-build').firstChild);
  $('modebar').classList.remove('hide');
  $('editmodal').classList.add('hide');
  setType('QTN2')}
function onGenerateClick(){
  if(EDIT_MODE){requestSave();return}
  if(TYPE==='QTN2'){openStatusGate();return}
  generate()}
// ---------------------------------------------------------------- mandatory status confirmation before generating a Quotation (New Design)
const QTN2_STATUSES=[['Draft','Still being worked on'],['Sent','Sent to the client, awaiting response'],
  ['Approved','Client has approved — ready to deliver'],['Revised','Client requested changes'],
  ['None','No status tracking needed']];
const STATUS_COLOR={Draft:'var(--muted)',Sent:'var(--info)',Approved:'var(--success)',Revised:'var(--warning)',None:'var(--border)'};
function openStatusGate(){
  $('statusgatebody').innerHTML=
    '<p class=muted style="font-size:12.5px;margin:0 0 12px">Confirm this quotation\'s status before generating — approving it offers to auto-generate the Delivery Order and Invoice too.</p>'+
    QTN2_STATUSES.map(([s,desc])=>
      '<button type=button class="btn'+(s===qtn2Status?' dark':'')+'" style="width:100%;text-align:left;margin-bottom:7px;display:block;padding:10px 12px;border-left:4px solid '+STATUS_COLOR[s]+'" onclick="confirmStatusGate(\''+s+'\')"><b>'+s+'</b><br><span class="muted" style="font-size:11px;font-weight:400">'+desc+'</span></button>'
    ).join('');
  $('statusgatemodal').classList.remove('hide')}
function closeStatusGate(){$('statusgatemodal').classList.add('hide')}
async function confirmStatusGate(status){
  setQtn2Status(status);
  if(status==='Approved'){openApprovedPrompt();return}
  closeStatusGate();
  await generate()}
function openApprovedPrompt(){
  $('statusgatebody').innerHTML=
    '<p style="font-size:13px;margin:0 0 16px">This quotation is <b>Approved</b>. Generate the Delivery Order and Invoice now too? Both are created for real immediately (using the quoted quantities) and flagged <b>In Progress</b> (highlighted red in All Docs) until the delivery is confirmed and the submittal is built.</p>'+
    '<div style="display:flex;gap:8px">'+
      '<button class=btn style="flex:1" onclick=finishApprovedFlow(false)>Not now</button>'+
      '<button class="btn dark" style="flex:1" onclick=finishApprovedFlow(true)>Yes, generate both</button>'+
    '</div>'}
async function finishApprovedFlow(autoGen){
  closeStatusGate();
  const r=await generate();
  if(autoGen&&r&&!r.error)await autoGenerateDoInv(r)}
async function autoGenerateDoInv(genResult){
  const rel=genResult.pdf||genResult.xlsx;
  const d=genResult.data||{};
  toast('Generating Delivery Order and Invoice…');
  const body={qtn_rel:rel,qtn_number:d.number||'',company:d.company||'',project:d.project||'',
    items:(d.items||[]).map(it=>({...it}))};
  const res=await fetch('/api/submissions',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)}).then(r=>r.json());
  if(res.error){alert(res.error);return}
  toast('Delivery Order '+res.submission.do_number+' and Invoice '+res.submission.inv_number+' generated — marked In Progress')}
// ---------------------------------------------------------------- two-step confirmation before overwriting an existing document
function diffSimpleFields(orig,cur,fieldLabel,changes){
  Object.keys(fieldLabel).forEach(k=>{
    const a=(orig[k]??'').toString().trim(),b=(cur[k]??'').toString().trim();
    if(a!==b)changes.push(fieldLabel[k]+': "'+(a||'—')+'" → "'+(b||'—')+'"')})}
function diffDocData(orig,cur){
  const changes=[];
  // CAT/EXP have their own field shapes entirely (badges/ordering table;
  // expense rows/employee/currency) — same reasoning as populateCatForm()/
  // populateExpForm() needing their own restore logic instead of the
  // generic HEAD+items one below.
  if(TYPE==='CAT'){
    diffSimpleFields(orig,cur,{number:'Sheet No.',date:'Date',company:'Product Name',series:'Series/Category',
      product_type:'Product Type',page_number:'Page Number',description:'Description',note:'Note',
      ordering_code_example:'Ordering Code Example'},changes);
    if(JSON.stringify(orig.badges||[])!==JSON.stringify(cur.badges||[]))changes.push('Spec badges changed');
    if(JSON.stringify(orig.specs||[])!==JSON.stringify(cur.specs||[]))changes.push('Technical specifications changed');
    if(JSON.stringify(orig.finish_colors||[])!==JSON.stringify(cur.finish_colors||[]))changes.push('Finish colors changed');
    if(JSON.stringify(orig.ordering_columns||[])!==JSON.stringify(cur.ordering_columns||[]))changes.push('Ordering table changed');
    if(orig.main_photo!==cur.main_photo)changes.push('Main product photo changed');
    if(orig.lifestyle_photo!==cur.lifestyle_photo)changes.push('Application photo changed');
    if(orig.dimension_diagram!==cur.dimension_diagram)changes.push('Dimension diagram changed');
    return changes}
  if(TYPE==='EXP'){
    diffSimpleFields(orig,cur,{number:'Report No.',date:'Date',company:'Employee',
      category:'Category / Reference',currency:'Currency'},changes);
    const or_=orig.rows||[],cr=cur.rows||[];
    if(or_.length!==cr.length)changes.push('Expense items: '+or_.length+' → '+cr.length+' row(s)');
    for(let i=0;i<Math.min(or_.length,cr.length);i++){
      if(JSON.stringify(or_[i])!==JSON.stringify(cr[i]))changes.push('Expense item '+(i+1)+' changed');
    }
    return changes}
  diffSimpleFields(orig,cur,{number:'Number',rev:'Rev',date:'Date',project:'Project',area:'Area',
    qtn_number:'QTN Number',lpo_number:'LPO Number',type:'Type',company:'Company',
    customer_block:'Customer details'},changes);
  if(TYPE!=='DO'){
    if(JSON.stringify(orig.discount||{})!==JSON.stringify(cur.discount||{}))changes.push('Discount changed');
    if(JSON.stringify(orig.vat||{})!==JSON.stringify(cur.vat||{}))changes.push('VAT changed');
  }
  const oi=orig.items||[],ci=cur.items||[];
  if(oi.length!==ci.length)changes.push('Line items: '+oi.length+' → '+ci.length+' item(s)');
  for(let i=0;i<Math.min(oi.length,ci.length);i++){
    if(JSON.stringify(oi[i])!==JSON.stringify(ci[i]))changes.push('Item '+(i+1)+' changed');
  }
  return changes}
function requestSave(){
  const cur=collectDocData();
  if(!cur.company){alert('Enter the company name.');return}
  if(!cur.number){alert('Enter a document number.');return}
  const changes=EDIT_SNAPSHOT?diffDocData(EDIT_SNAPSHOT,cur):[];
  $('saveconfirmtitle').textContent='Review changes';
  $('saveconfirmbody').innerHTML=
    '<p class=muted style="font-size:12.5px;margin:0 0 10px">You\'re editing a document that was already created. Here\'s what changed:</p>'+
    (changes.length
      ?'<ul style="margin:0 0 16px;padding-left:18px;font-size:12.5px;line-height:1.6">'+changes.map(c=>'<li>'+escHtml(c)+'</li>').join('')+'</ul>'
      :'<p class=muted style="font-size:12.5px;margin:0 0 16px">No changes detected.</p>')+
    '<button class="btn dark" style="width:100%" onclick=showSaveConfirmStep2()>Continue to Save</button>';
  $('saveconfirmmodal').classList.remove('hide')}
function showSaveConfirmStep2(){
  $('saveconfirmtitle').textContent='Confirm';
  const overwrite=!!EDITING;
  $('saveconfirmbody').innerHTML=
    '<p style="font-size:13px;font-weight:700;color:var(--danger);margin:0 0 16px">'+
    (overwrite
      ?'Are you sure you want to save these changes? This will overwrite the original file and cannot be undone.'
      :'Are you sure you want to save these changes? This will save a new document — the original file\'s layout doesn\'t match the app\'s own template, so it won\'t be modified.')+
    '</p>'+
    '<div style="display:flex;gap:8px">'+
      '<button class=btn style="flex:1" onclick=closeSaveConfirm()>Cancel</button>'+
      '<button class=btn style="flex:1;background:var(--danger);color:#fff;border-color:var(--danger)" onclick=confirmAndGenerate()>Yes, Save</button>'+
    '</div>'}
function closeSaveConfirm(){$('saveconfirmmodal').classList.add('hide')}
async function confirmAndGenerate(){
  closeSaveConfirm();
  await generate();
  EDIT_SNAPSHOT=JSON.parse(JSON.stringify(collectDocData()))}

async function openDoc(rel){
  const r=await fetch('/api/doc?rel='+encodeURIComponent(rel)).then(r=>r.json());
  if(r.error){alert(r.error);return}
  setType(r.doc_type,true);
  EDITING=r.imported?null:r.original;   // imported docs save as a new file, not an overwrite
  EDITING_DRAFT=null;
  const label=(r.imported?'Import ':'Edit ')+LABEL[r.doc_type]+(r.number?' '+r.number:'');
  // CAT/EXP's own JSON sidecar is already shaped exactly like a saved draft
  // (see collectCatData()/collectExpData()) — same restore functions as
  // loadDraft(), rather than the generic HEAD+items path below which doesn't
  // know about badges/ordering tables or expense rows/employee/currency at
  // all. Best-effort PDF text-scraping (DO/legacy imports) and QTN2's own
  // exact sidecar both still share that generic path — see
  // populateGenericDocForm() for the "one raw customer_block blob vs QTN2's
  // richer attn/address/pobox/city/country/status/terms fields" split.
  if(r.doc_type==='CAT')populateCatForm(r);
  else if(r.doc_type==='EXP')populateExpForm(r);
  else populateGenericDocForm(r.doc_type,r);
  // Show the real original file untouched until the user actually edits
  // something — don't switch to the app's own live-rendered template up
  // front (existing oninput/onchange handlers already call schedulePreview()
  // the moment anything is changed, so this only affects the initial view).
  setPreviewImage('/cs-thumb?f='+encodeURIComponent(rel));
  enterDocEditMode(label);
  toast(r.imported
    ? 'Imported line items from '+rel.split('/').pop()+' — fill in the header details and press Generate to save it'
    : 'Loaded '+rel.split('/').pop()+' for editing')}

async function openNative(rel){
  toast('Opening '+rel.split('/').pop()+'…');
  const r=await fetch('/open-file?name='+encodeURIComponent(rel)).then(r=>r.json());
  if(!r.ok)toast(r.error||'Could not open that file.')}

function openCS(rel){
  $('csmodaltitle').textContent='Company System — '+rel.split('/').pop()+' (page 1 preview — use Excel/PDF above to open the full file)';
  $('csframe').src='/cs-thumb?f='+encodeURIComponent(rel);
  $('csmodal').classList.remove('hide')}
function closeCS(){$('csmodal').classList.add('hide');$('csframe').src=''}
function openDatasheetCS(rel,name){
  $('csmodaltitle').textContent='Datasheet — '+(name||rel.split('/').pop())+' (page 1 preview)';
  $('csframe').src='/datasheet-thumb?rel='+encodeURIComponent(rel);
  $('csmodal').classList.remove('hide')}

// ---------------------------------------------------------------- Product Builder (compose an item description from a datasheet's own option tables)
let PB_ITEM=-1,PB_OPTS=null,PB_CUSTOM={},PB_ROWS={};
async function openProductBuilder(i,rel,name){
  PB_ITEM=i;PB_OPTS=null;PB_CUSTOM={};PB_ROWS={};
  const isBuilt=isProductBuilt(items[i]&&items[i].description);
  $('pb-title').textContent=(isBuilt?'Edit':'Build')+' — '+(name||rel.split('/').pop());
  $('pb-body').innerHTML='<p class="muted" style="font-size:12px;padding:8px 0">Reading the datasheet…</p>';
  $('productbuilder').classList.remove('hide');
  const r=await fetch('/api/product-options?rel='+encodeURIComponent(rel)).then(r=>r.json());
  if(PB_ITEM!==i)return;   // user closed/switched before this resolved
  PB_OPTS=r;
  renderProductBuilderBody()}
function closeProductBuilder(){$('productbuilder').classList.add('hide');PB_ITEM=-1;PB_OPTS=null;PB_CUSTOM={};PB_ROWS={}}
function pbSelectRow(id,label,list,fmt){
  if(!list||!list.length)return'';
  PB_ROWS[id]={label,list,fmt};
  return'<div class=f>'+pbSelectInner(id)+'</div>'}
function pbSelectInner(id){
  const{label,list,fmt}=PB_ROWS[id];
  return'<label>'+escHtml(label)+'</label><select id="'+id+'" onchange="onPbSelectChange(\''+id+'\')">'+
    list.map((v,i)=>'<option value='+i+'>'+escHtml(fmt?fmt(v):v)+'</option>').join('')+
    '<option value="__custom__">Custom…</option></select>'}
function onPbSelectChange(id){
  const sel=$(id);
  if(sel.value!=='__custom__'){delete PB_CUSTOM[id];updateProductBuilderPreview();return}
  const wrap=sel.closest('.f');
  wrap.innerHTML='<label>'+escHtml(PB_ROWS[id].label)+'</label>'+
    '<div style="display:flex;gap:6px">'+
      '<input type=text id="'+id+'" style="flex:1">'+
      '<button type=button class=btn style="padding:0 11px" onclick="revertPbCustom(\''+id+'\')" title="Choose from the list instead">▾</button>'+
    '</div>';
  const inp=$(id);inp.focus();
  inp.addEventListener('input',()=>{PB_CUSTOM[id]=inp.value;updateProductBuilderPreview()})}
function revertPbCustom(id){
  delete PB_CUSTOM[id];
  $(id).closest('.f').innerHTML=pbSelectInner(id);
  updateProductBuilderPreview()}
function renderProductBuilderBody(){
  const o=PB_OPTS||{};
  if(!o.model){
    $('pb-body').innerHTML='<p class="muted" style="font-size:12.5px;padding:8px 0">Couldn\'t find any option tables in this datasheet — you can still open it directly and type the description by hand.</p>';
    return}
  const rows=[
    pbSelectRow('pb-wattage','Wattage',o.wattage),
    pbSelectRow('pb-cct','Color Temperature',o.cct,v=>v[1]),
    pbSelectRow('pb-beam','Beam Angle',o.beam_angle,v=>v[1]),
    pbSelectRow('pb-size','Size',o.size),
    pbSelectRow('pb-color','Finish Color',o.color,v=>v[1]),
    pbSelectRow('pb-controls','Controls',o.controls,v=>v[1]),
  ].join('');
  const ipRow=o.ip_rating?'<div class=f><label>IP Rating</label><input value="'+escHtml(o.ip_rating)+'" disabled></div>':'';
  $('pb-body').innerHTML=rows+ipRow+
    '<div class=f><label>Description preview</label><div class=pbpreview id=pb-preview></div></div>'+
    '<button class="btn dark" style="width:100%;margin-top:6px" onclick=insertProductBuilderDescription()>Insert into Description</button>';
  updateProductBuilderPreview()}
// A picked option can be a bare string (Wattage/Size: "12W"), a [code,label]
// pair (CCT/Beam/Controls/Color: ["30","3000K"]), or a custom-typed value
// (PB_CUSTOM) — pbLabel/pbCode normalize all three so composeProductBuilderText
// doesn't need to know which one it's holding.
function pbSelected(id,list){
  if(PB_CUSTOM[id]!==undefined)return{custom:PB_CUSTOM[id]};
  const el=$(id);if(!el||!list)return null;
  return list[+el.value]}
function pbLabel(v){if(v==null)return'';if(typeof v==='object'&&!Array.isArray(v))return v.custom;return Array.isArray(v)?v[1]:v}
function pbCode(v){if(v==null)return'';if(typeof v==='object'&&!Array.isArray(v))return v.custom;return Array.isArray(v)?v[0]:v}
function composeProductBuilderText(){
  const o=PB_OPTS||{};
  const watt=pbSelected('pb-wattage',o.wattage);
  const cct=pbSelected('pb-cct',o.cct);
  const beam=pbSelected('pb-beam',o.beam_angle);
  const size=pbSelected('pb-size',o.size);
  const color=pbSelected('pb-color',o.color);
  const controls=pbSelected('pb-controls',o.controls);
  const firstLine=[o.model,pbLabel(watt),pbLabel(cct),pbLabel(beam),o.ip_rating||'',pbLabel(color)]
    .filter(Boolean).join(' ');
  const codeParts=[o.model,pbCode(watt),pbCode(size),pbCode(cct),pbCode(beam),pbCode(controls),pbCode(color)]
    .filter(Boolean);
  const code=codeParts.length>1?'\nCODE:'+codeParts.join('-'):'';
  return firstLine+code}
function updateProductBuilderPreview(){$('pb-preview').textContent=composeProductBuilderText()}
function insertProductBuilderDescription(){
  if(PB_ITEM<0||!items[PB_ITEM])return;
  const text=composeProductBuilderText();
  // QTN2's description cell is a contenteditable richbox (needs HTML line
  // breaks); every other type is a plain <textarea> (a literal "\n" is
  // already a real line break there) — same distinction collectQtn2Extra/
  // renderItems already draw elsewhere in this file.
  items[PB_ITEM].description=TYPE==='QTN2'?escHtml(text).replace(/\n/g,'<br>'):text;
  renderItems();schedulePreview();
  closeProductBuilder();
  matchPhotoForItem(PB_ITEM);matchDatasheetsForItem(PB_ITEM);
  toast('Description built from the datasheet')}

// ---------------------------------------------------------------- Product Finder ("find the product first, then build it" — Item Description's real job)
let PRODUCTS=[],PF_ITEM=-1;
async function openProductFinder(i,ev){
  ev.stopPropagation();
  PF_ITEM=i;
  const box=$('productfinder');
  const r=ev.currentTarget.getBoundingClientRect();
  const left=Math.min(Math.round(r.left),window.innerWidth-296);
  box.style.left=Math.max(8,left)+'px';box.style.top=(Math.round(r.bottom)+6)+'px';
  box.style.display='block';
  $('pf-search').value='';
  if(!PRODUCTS.length){const res=await fetch('/api/product-list').then(r=>r.json());PRODUCTS=res.products||[]}
  renderProductFinder();
  $('pf-search').focus()}
function closeProductFinder(){$('productfinder').style.display='none';PF_ITEM=-1}
document.addEventListener('click',e=>{
  if($('productfinder').style.display==='block'&&!e.target.closest('#productfinder')&&!e.target.closest('.ffind'))closeProductFinder()});
function renderProductFinder(){
  const q=($('pf-search').value||'').trim().toLowerCase();
  const rows=PRODUCTS.filter(p=>!q||p.name.toLowerCase().includes(q)).slice(0,40);
  $('pf-list').innerHTML=rows.length?rows.map(p=>
    '<div class=cpitem data-name="'+escHtml(p.name)+'" data-rel="'+escHtml(p.rel)+'" onclick="pickProduct(this.dataset.name,this.dataset.rel)"><span class=cpname>'+escHtml(p.name)+'</span></div>'
  ).join(''):'<p class="muted" style="font-size:12px;padding:6px 4px;margin:0">'+(PRODUCTS.length?'No matching results.':'No datasheets folder is configured for this brand, or the folder is empty.')+'</p>'}
function pickProduct(name,rel){
  const i=PF_ITEM;
  closeProductFinder();
  if(i<0||!items[i])return;
  items[i].datasheets=[{name,rel}];
  // never clobber text the user already typed — only drop the product name
  // in as a starting point if the description is still blank
  if(!(items[i].description||'').replace(/<[^>]*>/g,'').trim()){
    items[i].description=TYPE==='QTN2'?escHtml(name):name}
  renderItems();schedulePreview();
  openProductBuilder(i,rel,name)}

function openFileMenu(ev,rel,company,label){
  ev.stopPropagation();
  CTXROW={rel,company,label};
  const menu=$('filemenu');
  renderFileMenu();
  menu.style.display='block';
  const w=menu.offsetWidth||190,h=menu.offsetHeight||180;
  let x=ev.clientX,y=ev.clientY;
  if(x+w>window.innerWidth-8)x=window.innerWidth-w-8;
  if(y+h>window.innerHeight-8)y=window.innerHeight-h-8;
  menu.style.left=x+'px';menu.style.top=y+'px';
  setTimeout(()=>document.addEventListener('click',closeFileMenu,{once:true}),0)}
function closeFileMenu(){$('filemenu').style.display='none'}
function renderFileMenu(){
  const hasClip=!!CLIPBOARD;
  $('filemenu').innerHTML=
    '<div class=fmi onclick="fmCopy()"><span class=ic>⧉</span>Copy</div>'+
    '<div class=fmi onclick="fmCut()"><span class=ic>✂</span>Cut</div>'+
    '<div class="fmi'+(hasClip?'':' disabled')+'" onclick="fmPaste()"><span class=ic>📋</span>'+(hasClip?'Paste "'+CLIPBOARD.label+'"':'Paste')+'</div>'+
    '<div class=fmsep></div>'+
    '<div class="fmi danger" onclick="fmDelete(this,event)"><span class=ic>🗑</span><span class=fmilabel>Delete</span></div>'}
function fmCopy(){CLIPBOARD={rel:CTXROW.rel,mode:'copy',label:CTXROW.label};toast('Copied "'+CTXROW.label+'" — click Paste on the target row')}
function fmCut(){CLIPBOARD={rel:CTXROW.rel,mode:'cut',label:CTXROW.label};toast('Cut "'+CTXROW.label+'" — click Paste on the target row')}
async function fmPaste(){
  if(!CLIPBOARD)return;
  const destCompany=CTXROW.company,mode=CLIPBOARD.mode,srcRel=CLIPBOARD.rel;
  const r=await fetch('/api/file-op',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({op:mode==='cut'?'move':'copy',rel:srcRel,dest_company:destCompany})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  if(mode==='cut')CLIPBOARD=null;
  toast('Pasted');loadIndex()}
// window.confirm() silently no-ops in this app's embedded webview (see
// fmDelete's old behavior — the "are you sure" dialog never showed and never
// returned true, so Delete looked completely dead) — same click-again-to-
// confirm pattern used everywhere else in the app (e.g. deleteFcCategory)
// instead of a native dialog. Both clicks stopPropagation so the arming
// click doesn't trip openFileMenu's own close-on-any-click listener before
// the second click can land.
let FM_DELETE_TIMER=null;
async function fmDelete(el,ev){
  if(el.dataset.confirm!=='1'){
    ev.stopPropagation();
    el.dataset.confirm='1';
    el.querySelector('.fmilabel').textContent='Click again to confirm';
    clearTimeout(FM_DELETE_TIMER);
    FM_DELETE_TIMER=setTimeout(()=>{
      el.dataset.confirm='';
      const lbl=el.querySelector('.fmilabel');if(lbl)lbl.textContent='Delete'
    },2500);
    return}
  ev.stopPropagation();
  clearTimeout(FM_DELETE_TIMER);
  const label=CTXROW.label,rel=CTXROW.rel;
  closeFileMenu();
  const r=await fetch('/api/file-op',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({op:'delete',rel})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  toast('Deleted "'+label+'"');loadIndex()}

function showPrev(ev,rel,canPreview){
  clearTimeout(hoverTimer);
  hoverTimer=setTimeout(()=>{
    const box=$('hoverprev'),myToken=++hoverToken;
    if(canPreview){
      box.innerHTML='<div class=loading>Loading preview…</div>';
      const img=new Image();
      img.onload=()=>{if(myToken!==hoverToken)return;box.innerHTML='';box.appendChild(img)};
      img.onerror=()=>{if(myToken!==hoverToken)return;box.innerHTML='<div class=empty>Couldn\'t render a preview for this file.</div>'};
      img.src='/cs-thumb?f='+encodeURIComponent(rel);
    }else{
      box.innerHTML='<div class=empty>No preview available yet for this file.</div>';
    }
    box.style.display='block';movePrev(ev)},150)}
function movePrev(ev){
  const box=$('hoverprev');if(box.style.display!=='block')return;
  const pad=18,w=box.offsetWidth,h=box.offsetHeight;
  let x=ev.clientX+pad,y=ev.clientY+pad;
  if(x+w>window.innerWidth-8)x=ev.clientX-w-pad;
  if(y+h>window.innerHeight-8)y=window.innerHeight-h-8;
  box.style.left=x+'px';box.style.top=y+'px'}
function hidePrev(){clearTimeout(hoverTimer);hoverToken++;$('hoverprev').style.display='none';$('hoverprev').innerHTML=''}

let CLIENTS=[],PROJECTS=[],ATTNS=[],ADDRESSES=[];
function fillDatalist(id,vals){$(id).innerHTML=(vals||[]).map(v=>'<option value="'+escHtml(v)+'">').join('')}
// Build's Company/Attn/Address prediction merges two sources: CLIENTS/ATTNS/
// ADDRESSES from document history (/api/index — every company ever typed
// into a generated doc) and CLIENT_RECORDS from the Clients database
// (/api/clients — the curated address book, incl. anything only ever added
// there and never yet used in a document). Neither alone is "the" list —
// a brand-new client hasn't appeared in a document yet, and old documents
// may reference companies that were never turned into a Clients-tab record.
function allCompanyNames(){
  const set=new Set(CLIENTS);
  CLIENT_RECORDS.forEach(c=>{if(c.name)set.add(c.name)});
  return [...set].sort((a,b)=>a.localeCompare(b))}
function allAttns(){
  const set=new Set(ATTNS);
  CLIENT_RECORDS.forEach(c=>{if(c.attn)set.add(c.attn)});
  return [...set]}
function allAddresses(){
  const set=new Set(ADDRESSES);
  CLIENT_RECORDS.forEach(c=>{if(c.address)set.add(c.address)});
  return [...set]}
function refreshClientNamesDatalist(){fillDatalist('clients',allCompanyNames())}
async function loadClients(){
  const r=await fetch('/api/index').then(r=>r.json());
  CLIENTS=r.companies||[];PROJECTS=r.projects||[];ATTNS=r.attns||[];ADDRESSES=r.addresses||[];
  refreshClientNamesDatalist();
  fillDatalist('projects',PROJECTS);
  return r}
async function loadIndex(){const r=await loadClients();INDEX=r.records||[];
  const cl=r.companies||[...new Set(INDEX.map(x=>x.company_label))];
  const curCo=$('fcompany').value;
  $('fcompany').innerHTML='<option value=all>All companies</option>'+cl.map(c=>'<option'+(c===curCo?' selected':'')+'>'+c+'</option>').join('');
  if(ALLDOCS_TYPE==='DRAFTS'){const dr=await fetch('/api/drafts').then(r=>r.json());ALLDOCS_DRAFTS=dr.drafts||[]}
  renderList()}

// ---------------------------------------------------------------- countries (client Country field: flag + searchable picker)
const COUNTRIES=[["AE","United Arab Emirates"],["AF","Afghanistan"],["AL","Albania"],["DZ","Algeria"],
["AD","Andorra"],["AO","Angola"],["AG","Antigua and Barbuda"],["AR","Argentina"],["AM","Armenia"],
["AU","Australia"],["AT","Austria"],["AZ","Azerbaijan"],["BS","Bahamas"],["BH","Bahrain"],["BD","Bangladesh"],
["BB","Barbados"],["BY","Belarus"],["BE","Belgium"],["BZ","Belize"],["BJ","Benin"],["BT","Bhutan"],
["BO","Bolivia"],["BA","Bosnia and Herzegovina"],["BW","Botswana"],["BR","Brazil"],["BN","Brunei"],
["BG","Bulgaria"],["BF","Burkina Faso"],["BI","Burundi"],["KH","Cambodia"],["CM","Cameroon"],["CA","Canada"],
["CV","Cabo Verde"],["CF","Central African Republic"],["TD","Chad"],["CL","Chile"],["CN","China"],
["CO","Colombia"],["KM","Comoros"],["CG","Congo"],["CD","Congo (DRC)"],["CR","Costa Rica"],["HR","Croatia"],
["CU","Cuba"],["CY","Cyprus"],["CZ","Czechia"],["DK","Denmark"],["DJ","Djibouti"],["DM","Dominica"],
["DO","Dominican Republic"],["EC","Ecuador"],["EG","Egypt"],["SV","El Salvador"],["GQ","Equatorial Guinea"],
["ER","Eritrea"],["EE","Estonia"],["SZ","Eswatini"],["ET","Ethiopia"],["FJ","Fiji"],["FI","Finland"],
["FR","France"],["GA","Gabon"],["GM","Gambia"],["GE","Georgia"],["DE","Germany"],["GH","Ghana"],
["GR","Greece"],["GD","Grenada"],["GT","Guatemala"],["GN","Guinea"],["GW","Guinea-Bissau"],["GY","Guyana"],
["HT","Haiti"],["HN","Honduras"],["HK","Hong Kong"],["HU","Hungary"],["IS","Iceland"],["IN","India"],
["ID","Indonesia"],["IR","Iran"],["IQ","Iraq"],["IE","Ireland"],["IL","Israel"],["IT","Italy"],
["JM","Jamaica"],["JP","Japan"],["JO","Jordan"],["KZ","Kazakhstan"],["KE","Kenya"],["KI","Kiribati"],
["KW","Kuwait"],["KG","Kyrgyzstan"],["LA","Laos"],["LV","Latvia"],["LB","Lebanon"],["LS","Lesotho"],
["LR","Liberia"],["LY","Libya"],["LI","Liechtenstein"],["LT","Lithuania"],["LU","Luxembourg"],
["MO","Macao"],["MG","Madagascar"],["MW","Malawi"],["MY","Malaysia"],["MV","Maldives"],["ML","Mali"],
["MT","Malta"],["MR","Mauritania"],["MU","Mauritius"],["MX","Mexico"],["MD","Moldova"],["MC","Monaco"],
["MN","Mongolia"],["ME","Montenegro"],["MA","Morocco"],["MZ","Mozambique"],["MM","Myanmar"],
["NA","Namibia"],["NP","Nepal"],["NL","Netherlands"],["NZ","New Zealand"],["NI","Nicaragua"],
["NE","Niger"],["NG","Nigeria"],["MK","North Macedonia"],["NO","Norway"],["OM","Oman"],["PK","Pakistan"],
["PA","Panama"],["PG","Papua New Guinea"],["PY","Paraguay"],["PE","Peru"],["PH","Philippines"],
["PL","Poland"],["PT","Portugal"],["QA","Qatar"],["RO","Romania"],["RU","Russia"],["RW","Rwanda"],
["SA","Saudi Arabia"],["SN","Senegal"],["RS","Serbia"],["SC","Seychelles"],["SL","Sierra Leone"],
["SG","Singapore"],["SK","Slovakia"],["SI","Slovenia"],["SO","Somalia"],["ZA","South Africa"],
["KR","South Korea"],["SS","South Sudan"],["ES","Spain"],["LK","Sri Lanka"],["SD","Sudan"],
["SR","Suriname"],["SE","Sweden"],["CH","Switzerland"],["SY","Syria"],["TW","Taiwan"],["TJ","Tajikistan"],
["TZ","Tanzania"],["TH","Thailand"],["TL","Timor-Leste"],["TG","Togo"],["TO","Tonga"],
["TT","Trinidad and Tobago"],["TN","Tunisia"],["TR","Turkey"],["TM","Turkmenistan"],["UG","Uganda"],
["UA","Ukraine"],["GB","United Kingdom"],["US","United States"],["UY","Uruguay"],["UZ","Uzbekistan"],
["VU","Vanuatu"],["VA","Vatican City"],["VE","Venezuela"],["VN","Vietnam"],["YE","Yemen"],
["ZM","Zambia"],["ZW","Zimbabwe"]];
function countryName(code){const c=COUNTRIES.find(x=>x[0]===code);return c?c[1]:''}
function countryFlag(code){
  if(!code||code.length!==2)return '';
  return String.fromCodePoint(...[...code.toUpperCase()].map(ch=>127397+ch.charCodeAt(0)))}

// ---------------------------------------------------------------- client database (full details, logos — the Clients tab)
let CLIENT_RECORDS=[];
async function loadClientsView(){
  const r=await fetch('/api/clients').then(r=>r.json());
  CLIENT_RECORDS=r.clients||[];
  refreshClientNamesDatalist();
  renderClientsGrid()}
async function importClients(){
  const btn=$('clientsimportbtn');
  const orig=btn.textContent;btn.disabled=true;btn.textContent='Scanning documents…';
  try{
    const r=await fetch('/api/clients-import',{method:'POST'}).then(r=>r.json());
    CLIENT_RECORDS=r.clients||[];
    refreshClientNamesDatalist();
    renderClientsGrid();
    const parts=[];
    if(r.imported)parts.push(r.imported+' new client'+(r.imported===1?'':'s'));
    if(r.enriched)parts.push('filled in details for '+r.enriched);
    toast(parts.length?('Imported from documents: '+parts.join(', ')+'.')
                      :'Nothing new found — everything is already in your list.');
  }finally{btn.disabled=false;btn.textContent=orig}}
function clientInitial(name){return escHtml((name||'?').trim().charAt(0).toUpperCase())}
function clientCardHtml(c){
  const logo=c.logo?'<img class=clientlogo src="'+c.logo+'">':'<div class="clientlogo ph">'+clientInitial(c.name)+'</div>';
  const meta=[c.attn,c.address].filter(Boolean).join(' · ');
  return '<div class=clientcard onclick="openClientEditor(\''+c.id+'\')">'+
    '<button class=clientcardedit title="Edit client" onclick="event.stopPropagation();openClientEditor(\''+c.id+'\')">✎</button>'+
    '<div class=cctop>'+logo+'<div style="min-width:0"><div class=clientname title="'+escHtml(c.name)+'">'+escHtml(c.name)+'</div></div></div>'+
    (meta?'<div class=clientmeta>'+escHtml(meta)+'</div>':'')+
    '<div class=clientactions onclick="event.stopPropagation()">'+
      '<button class=btn onclick="useClientInBuild(\''+c.id+'\')">Use in Build</button>'+
      '<button class=btn onclick="exportClient(\''+c.id+'\')">Export</button>'+
      '<button class=btn onclick="deleteClient(\''+c.id+'\',this)">Delete</button>'+
    '</div></div>'}
// Clients are grouped by Country or City so a growing address book stays
// navigable — collapsed state persists only for this page load
// (COLLAPSED_SECTIONS), not saved anywhere, since it's just a viewing
// convenience.
let COLLAPSED_SECTIONS=new Set();
let CLIENT_GROUPING='country';
function setClientGrouping(g){
  CLIENT_GROUPING=g;
  document.querySelectorAll('#clientgroupseg button').forEach(b=>b.classList.toggle('on',b.dataset.g===g));
  renderClientsGrid()}
function renderClientsGrid(){
  const q=($('clientsearch').value||'').trim().toLowerCase();
  const rows=CLIENT_RECORDS.filter(c=>!q||[c.name,c.attn,c.address,c.country,c.city].some(v=>(v||'').toLowerCase().includes(q)));
  if(!rows.length){$('clientsgrid').innerHTML=CLIENT_RECORDS.length
    ?emptyStateHtml(EMPTY_ICON_CLIENTS,'No Matches Found','No clients match the current search criteria.')
    :emptyStateHtml(EMPTY_ICON_CLIENTS,'No Clients on File','Select <b>+ New Client</b> above to add the first client record for this brand.');return}
  const byCity=CLIENT_GROUPING==='city';
  const unassignedLabel='Unspecified';
  const groups={},labels={};
  rows.forEach(c=>{
    const label=(byCity?c.city:c.country||'').trim()||unassignedLabel;
    const key=label;
    labels[key]=label;(groups[key]=groups[key]||[]).push(c)});
  const keys=Object.keys(groups).sort((a,b)=>{
    if(labels[a]===labels[b])return 0;
    if(labels[a]===unassignedLabel)return 1;
    if(labels[b]===unassignedLabel)return -1;
    return labels[a].localeCompare(labels[b])});
  const prefix=byCity?'city:':'country:';
  $('clientsgrid').innerHTML=keys.map(key=>{
    const list=groups[key].sort((a,b)=>a.name.localeCompare(b.name));
    const collapsed=COLLAPSED_SECTIONS.has(prefix+key);
    return '<div class="clientsection'+(collapsed?' collapsed':'')+'" data-cat="'+escHtml(prefix+key)+'">'+
      '<div class=clientsectionhead onclick="toggleClientSection(this)">'+escHtml(labels[key])+
        '<span class=clientsectioncount>'+list.length+'</span><span class=clientsectionchev>▾</span></div>'+
      '<div class=clientsgrid>'+list.map(clientCardHtml).join('')+'</div></div>'}).join('')}
function toggleClientSection(head){
  const sec=head.closest('.clientsection'),cat=sec.dataset.cat;
  sec.classList.toggle('collapsed');
  if(sec.classList.contains('collapsed'))COLLAPSED_SECTIONS.add(cat);else COLLAPSED_SECTIONS.delete(cat)}
function deleteClient(id,btn){
  // window.confirm() silently no-ops in this environment — inline
  // double-click-to-confirm instead, same pattern as elsewhere in this app.
  if(btn.dataset.confirm!=='1'){
    btn.dataset.confirm='1';btn.textContent='Confirm?';
    setTimeout(()=>{if(btn.dataset.confirm==='1'){btn.dataset.confirm='';btn.textContent='Delete'}},2500);
    return}
  actuallyDeleteClient(id)}
async function actuallyDeleteClient(id){
  const r=await fetch('/api/clients-delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})}).then(r=>r.json());
  CLIENT_RECORDS=r.clients||[];refreshClientNamesDatalist();renderClientsGrid();toast('Client deleted');
  if(currentClientId===id)closeClientEditor()}
function useClientInBuild(id){
  const c=CLIENT_RECORDS.find(x=>x.id===id);if(!c)return;
  view(DOC_VIEWS.QTN2);setType('QTN2');
  setCompanyVal(c.name);$('customer_attn').textContent=c.attn||'';$('customer_address').textContent=c.address||'';
  setBuildAddressExtras(c);
  schedulePreview();toast('Loaded '+c.name+' into the Quotation form')}
function exportClient(id){window.location.href='/api/clients-export?id='+encodeURIComponent(id)}
function exportAllClients(){window.location.href='/api/clients-export'}

// ---------------------------------------------------------------- submissions (QTN -> LPO -> DO -> scanned DO -> INV)
let SUBMISSIONS=[];
const STAGE_LABEL={in_progress:'In Progress',delivered:'Delivered',submittal_built:'Submittal Built'};
async function loadSubmissions(){
  const r=await fetch('/api/submissions').then(r=>r.json());
  SUBMISSIONS=r.submissions||[];
  renderSubmissions()}
function renderSubmissions(){
  const q=($('subsearch').value||'').trim().toLowerCase();
  const rows=SUBMISSIONS.filter(s=>!q||(s.company||'').toLowerCase().includes(q)||(s.qtn_number||'').includes(q)||(s.do_number||'').toString().includes(q))
    .sort((a,b)=>(b.updated||'').localeCompare(a.updated||''));
  if(!rows.length){$('submissionsgrid').innerHTML=SUBMISSIONS.length
    ?emptyStateHtml(EMPTY_ICON_SUBMISSIONS,'No Matches Found','No submissions match the current search criteria.')
    :emptyStateHtml(EMPTY_ICON_SUBMISSIONS,'No Submissions Yet','Approving a quotation will prompt you to start one automatically, or select <b>+ New Submission</b> to create one manually.');return}
  $('submissionsgrid').innerHTML=rows.map(subCardHtml).join('')}
function subCardHtml(s){
  const steps=[];
  steps.push('<div class=substep><b>QTN</b> '+escHtml(s.qtn_number||'—')+
    (s.qtn_rel?' <a onclick="openNative(\''+escHtml(s.qtn_rel).replace(/'/g,"\\'")+'\')">open</a>':'')+'</div>');
  steps.push('<div class=substep substepdone><b>DO</b> '+escHtml(String(s.do_number||''))+
    (s.do_rel?' <a onclick="openNative(\''+escHtml(s.do_rel).replace(/'/g,"\\'")+'\')">open</a>':'')+'</div>');
  steps.push('<div class=substep substepdone><b>INV</b> '+escHtml(String(s.inv_number||''))+
    (s.inv_rel?' <a onclick="openNative(\''+escHtml(s.inv_rel).replace(/'/g,"\\'")+'\')">open</a>':'')+'</div>');
  steps.push('<div class=substep'+(s.scanned_do_rel?' substepdone':'')+'><b>Scanned DO</b> '+
    (s.scanned_do_rel?'<a onclick="window.open(\'/open-scanned-do?rel='+encodeURIComponent(s.scanned_do_rel)+'\')">view</a>':'<span class=muted>not linked</span>')+'</div>');
  steps.push('<div class=substep'+(s.lpo_saved_name?' substepdone':'')+'><b>LPO</b> '+
    (s.lpo_saved_name?('<a onclick="window.open(\'/submission-lpo?id='+encodeURIComponent(s.id)+'\')">'+escHtml(s.lpo_filename||'view')+'</a>'):'<span class=muted>not attached yet</span>')+'</div>');
  steps.push('<div class=substep'+(s.submittal_rel?' substepdone':'')+'><b>Submittal</b> '+
    (s.submittal_rel?'<a onclick="window.open(\'/submission-submittal?id='+encodeURIComponent(s.id)+'\')">view</a>':'<span class=muted>not built</span>')+'</div>');
  let actions='';
  if(s.stage==='in_progress')actions='<button class="btn dark" onclick="linkScannedDo(\''+s.id+'\')">Import Scanned DO…</button>';
  else if(s.stage==='delivered')actions='<button class="btn dark" onclick="linkScannedDo(\''+s.id+'\')">Re-link Scanned DO</button><button class="btn dark" onclick="openBuildSubmittal(\''+s.id+'\')">Build Submittal…</button>';
  actions+='<button class=btn onclick="deleteSubmission(\''+s.id+'\',this)">Delete</button>';
  return '<div class=subcard>'+
    '<div class=subtop><span class=subname>'+escHtml(s.company||'')+'</span>'+
    '<span class=stagebadge '+s.stage+'>'+(STAGE_LABEL[s.stage]||s.stage)+'</span>'+
    '<span class=submeta>'+(s.project?escHtml(s.project)+' · ':'')+(s.updated||'').slice(0,10)+'</span></div>'+
    '<div class=substeps>'+steps.join('')+'</div>'+
    '<div class=subactions>'+actions+'</div>'+
  '</div>'}
function deleteSubmission(id,btn){
  if(btn.dataset.confirm!=='1'){
    btn.dataset.confirm='1';btn.textContent='Confirm?';
    setTimeout(()=>{if(btn.dataset.confirm==='1'){btn.dataset.confirm='';btn.textContent='Delete'}},2500);
    return}
  actuallyDeleteSubmission(id)}
async function actuallyDeleteSubmission(id){
  const r=await fetch('/api/submissions-delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})}).then(r=>r.json());
  SUBMISSIONS=r.submissions||[];renderSubmissions();toast('Submission deleted')}
// Scan Now — see scanner.py's own top-of-file comment for the full flow.
// linkScannedDo() now opens that modal instead of going straight to a
// file picker; chooseExistingScannedFile() (reachable from inside the
// modal) is the old direct-browse path, kept as a fallback for a scan
// that didn't originate from this machine's own scanner (e.g. emailed in).
let SCANNOW_SUB_ID=null, SCANNOW_SESSION=null, SCANNOW_SCANNING=false;
async function _applyScannedDoLink(id,rel){
  const r2=await fetch('/api/submissions-link-scanned-do',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id,rel})}).then(r=>r.json());
  if(r2.error){alert(r2.error);return}
  await loadSubmissions();
  // Linking the scanned/signed DO is "delivery confirmed" — per the user's
  // own description, that's the moment to ask about building the submittal.
  // window.confirm() is a silent no-op in this environment (documented
  // elsewhere in this file), so the "ask" is the modal itself opening —
  // its own Cancel button is the "not now" answer.
  toast('Scanned Delivery Order linked');
  openBuildSubmittal(id)}
async function chooseExistingScannedFile(){
  const id=SCANNOW_SUB_ID;
  closeScanNowModal();
  const r=await fetch('/api/browse-scanned-do',{method:'POST'}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  if(!r.rel)return;
  _applyScannedDoLink(id,r.rel)}
async function linkScannedDo(id){
  SCANNOW_SUB_ID=id;SCANNOW_SESSION=null;
  $('scannowmodal').classList.remove('hide');
  $('scannow-pages').innerHTML='';
  $('scannow-scan-btn').disabled=true;
  $('scannow-status').textContent='Checking for a scanner…';
  $('scannow-nodevice').classList.add('hide');
  $('scannow-picker-wrap').style.display='none';
  refreshScannerList()}
function closeScanNowModal(){
  $('scannowmodal').classList.add('hide');
  if(SCANNOW_SESSION)fetch('/api/scanner-cancel',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({session_id:SCANNOW_SESSION})});
  SCANNOW_SESSION=null;SCANNOW_SUB_ID=null}
async function refreshScannerList(){
  $('scannow-nodevice').classList.add('hide');
  const r=await fetch('/api/scanner-list').then(r=>r.json()).catch(()=>({scanners:[]}));
  const scanners=r.scanners||[];
  if(!scanners.length){
    $('scannow-status').textContent='';
    $('scannow-nodevice').classList.remove('hide');
    $('scannow-scan-btn').disabled=true;
    return}
  $('scannow-picker-wrap').style.display=scanners.length>1?'':'none';
  $('scannow-device').innerHTML=scanners.map(s=>'<option value="'+escHtml(s.id)+'">'+escHtml(s.name)+'</option>').join('');
  $('scannow-status').textContent=scanners.length>1?'Choose a scanner, then Scan Page.':'Ready — '+scanners[0].name;
  $('scannow-scan-btn').disabled=false}
async function scanNextPage(){
  if(SCANNOW_SCANNING)return;
  SCANNOW_SCANNING=true;
  const btn=$('scannow-scan-btn');
  btn.disabled=true;btn.textContent='Scanning…';
  $('scannow-status').textContent='Scanning — please wait…';
  const deviceSel=$('scannow-device');
  const device_id=deviceSel.options.length?deviceSel.value:undefined;
  const r=await fetch('/api/scanner-scan-page',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({session_id:SCANNOW_SESSION,device_id})}).then(r=>r.json()).catch(e=>({ok:false,error:e.message}));
  btn.disabled=false;btn.textContent='Scan Page';
  SCANNOW_SCANNING=false;
  if(!r.ok){$('scannow-status').textContent='Scan failed: '+(r.error||'unknown error');return}
  SCANNOW_SESSION=r.session_id;
  $('scannow-status').textContent=r.page_count+' page'+(r.page_count!==1?'s':'')+' scanned — Scan Page again to add another, or Save & Link when done.';
  const div=document.createElement('div');
  div.className='scanpagetile';
  div.innerHTML='<img src="'+r.preview+'"><span>'+r.page_count+'</span>';
  $('scannow-pages').appendChild(div)}
async function removeLastScanPage(){
  if(!SCANNOW_SESSION)return;
  const r=await fetch('/api/scanner-remove-last-page',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({session_id:SCANNOW_SESSION})}).then(r=>r.json());
  const tiles=$('scannow-pages');
  if(tiles.lastChild)tiles.removeChild(tiles.lastChild);
  $('scannow-status').textContent=r.page_count+' page'+(r.page_count!==1?'s':'')+' scanned.'}
async function finalizeScanAndLink(){
  if(!SCANNOW_SESSION){toast('Scan at least one page first');return}
  const id=SCANNOW_SUB_ID;
  $('scannow-status').textContent='Saving…';
  const r=await fetch('/api/scanner-finalize',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({session_id:SCANNOW_SESSION,submission_id:id})}).then(r=>r.json());
  if(!r.ok){$('scannow-status').textContent='Could not save: '+(r.error||'unknown error');return}
  SCANNOW_SESSION=null;
  closeScanNowModal();
  _applyScannedDoLink(id,r.rel)}

// -------- build submittal: attach the LPO (first appearance in the flow), confirm/adjust quantities, merge into one PDF
let SUBMITTAL_SUB=null, SUBMITTAL_ITEMS=[], SUBMITTAL_LPO_FILE='', SUBMITTAL_LPO_NAME='';
function openBuildSubmittal(id){
  SUBMITTAL_SUB=SUBMISSIONS.find(s=>s.id===id);
  if(!SUBMITTAL_SUB)return;
  SUBMITTAL_ITEMS=(SUBMITTAL_SUB.items||[]).map(it=>({...it}));
  SUBMITTAL_LPO_FILE='';SUBMITTAL_LPO_NAME='';
  renderBuildSubmittal();
  $('newsubmodal').classList.remove('hide')}
function closeBuildSubmittal(){$('newsubmodal').classList.add('hide')}
function renderBuildSubmittal(){
  $('newsubtitle').textContent='Build Submittal — '+escHtml(SUBMITTAL_SUB.company||'');
  const rows=SUBMITTAL_ITEMS.map((it,i)=>{
    const orig=(SUBMITTAL_SUB.items||[])[i]||{};
    const origQty=orig.lpo_qty??orig.qty??'';
    const over=origQty&&Number(it.lpo_qty)>Number(origQty);
    return '<div class="f" style="display:flex;gap:8px;align-items:flex-end;margin-bottom:8px">'+
      '<div style="flex:1"><label style="font-size:10px">Item '+(i+1)+'</label>'+
        '<div style="font-size:12.5px;padding:8px 10px;background:var(--surface-2);border:1px solid var(--border);border-radius:8px">'+escHtml((it.description||'').replace(/<[^>]*>/g,' '))+'</div></div>'+
      '<div style="width:90px"><label style="font-size:10px">Quoted Qty</label><input value="'+escHtml(String(origQty))+'" disabled></div>'+
      '<div style="width:90px"><label style="font-size:10px">Delivered Qty</label><input value="'+escHtml(String(it.lpo_qty??''))+'" oninput="SUBMITTAL_ITEMS['+i+'].lpo_qty=this.value;renderBuildSubmittal()"'+(over?' style="border-color:var(--danger)"':'')+'></div>'+
    '</div>'+(over?'<p style="font-size:11px;color:var(--danger);margin:-4px 0 8px">⚠️ exceeds quoted qty</p>':'')}).join('');
  $('newsubbody').innerHTML=
    '<p class=muted style="font-size:12px;margin:0 0 12px">Attach the customer\'s LPO and confirm the delivered quantities — if anything differs from what was quoted, the Delivery Order and Invoice already on file will be corrected to match before the submittal is built.</p>'+
    rows+
    '<div class=g2 style="margin-top:12px">'+
      '<div class=f><label>LPO Number</label><input id=submittal-lpo-number value="'+escHtml(SUBMITTAL_SUB.lpo_number||'')+'" placeholder="e.g. LTTR-PO-00241-1"></div>'+
      '<div class=f><label>LPO File</label><input type=file id=submittal-lpo-file accept=".pdf,.jpg,.jpeg,.png" onchange=onSubmittalLpoFile(this)></div>'+
    '</div>'+
    '<div style="display:flex;gap:8px;margin-top:14px">'+
      '<button class=btn style="flex:1" onclick=closeBuildSubmittal()>Cancel</button>'+
      '<button class="btn dark" style="flex:2" onclick=doBuildSubmittal()>Build Submittal PDF</button>'+
    '</div>'}
function onSubmittalLpoFile(input){
  const f=input.files[0];if(!f)return;
  SUBMITTAL_LPO_NAME=f.name;
  const reader=new FileReader();
  reader.onload=()=>{SUBMITTAL_LPO_FILE=reader.result};
  reader.readAsDataURL(f)}
async function doBuildSubmittal(){
  toast('Building submittal…');
  const body={id:SUBMITTAL_SUB.id,items:SUBMITTAL_ITEMS,
    lpo_number:($('submittal-lpo-number').value||'').trim(),
    lpo_file:SUBMITTAL_LPO_FILE,lpo_filename:SUBMITTAL_LPO_NAME};
  const r=await fetch('/api/submissions-build-submittal',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  closeBuildSubmittal();await loadSubmissions();toast('Submittal built')}

// ---------------------------------------------------------------- Statement of Account
// Deliberately only ever populated from today forward (see finance ledger
// backend comment) — there's no reliable payment history for the
// thousands of real pre-existing invoices, so this never touches them.
let LEDGER=[];
const CHART_INVOICED='var(--info)', CHART_COLLECTED='var(--success)';
function money(n){return 'AED '+Math.round(n||0).toLocaleString()}
async function loadStatement(){
  const r=await fetch('/api/finance/ledger').then(r=>r.json());
  LEDGER=r.ledger||[];
  renderStatement()}
function renderStatement(){
  const totalInvoiced=LEDGER.reduce((s,e)=>s+e.total,0);
  const totalCollected=LEDGER.filter(e=>e.paid).reduce((s,e)=>s+e.total,0);
  const totalOutstanding=totalInvoiced-totalCollected;
  $('kpi-invoiced').textContent=money(totalInvoiced);
  $('kpi-collected').textContent=money(totalCollected);
  $('kpi-outstanding').textContent=money(totalOutstanding);
  renderStatementChart();
  renderStatementCompanies()}
function monthKey(dateStr){return (dateStr||'').slice(0,7)} // YYYY-MM
function monthLabel(key){
  const [y,m]=key.split('-');
  const names=['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
  return names[Number(m)-1]+' '+y}
function renderStatementChart(){
  const box=$('statement-chart');
  if(!LEDGER.length){
    box.innerHTML='<p class="muted" style="font-size:12.5px;padding:8px 0">No invoices recorded yet. This report will populate as new invoices are generated.</p>';
    return}
  const byMonth={};
  LEDGER.forEach(e=>{
    const k=monthKey(e.date);if(!k)return;
    byMonth[k]=byMonth[k]||{invoiced:0,collected:0};
    byMonth[k].invoiced+=e.total;
    if(e.paid)byMonth[k].collected+=e.total});
  const months=Object.keys(byMonth).sort();
  const maxVal=Math.max(1,...months.map(k=>byMonth[k].invoiced));
  const w=Math.max(560,months.length*90), h=220, padL=54, padB=30, padT=10;
  const plotH=h-padB-padT, plotW=w-padL-12;
  const groupW=plotW/months.length, barW=Math.min(26,groupW/3);
  const yTicks=4;
  let gridSvg='';
  for(let i=0;i<=yTicks;i++){
    const v=maxVal*i/yTicks, y=padT+plotH-(plotH*i/yTicks);
    gridSvg+='<line x1='+padL+' y1='+y.toFixed(1)+' x2='+w+' y2='+y.toFixed(1)+' stroke="#e4e1da" stroke-width="1" />';
    gridSvg+='<text x='+(padL-8)+' y='+(y+4).toFixed(1)+' text-anchor=end font-size=10 fill="#74716a">'+Math.round(v).toLocaleString()+'</text>'}
  let bars='';
  months.forEach((k,i)=>{
    const gx=padL+i*groupW;
    const d=byMonth[k];
    const invH=plotH*(d.invoiced/maxVal), colH=plotH*(d.collected/maxVal);
    const bx1=gx+groupW/2-barW-3, bx2=gx+groupW/2+3;
    bars+='<rect class=chartbar data-tip="'+monthLabel(k)+' — Invoiced '+money(d.invoiced)+'" x='+bx1.toFixed(1)+' y='+(padT+plotH-invH).toFixed(1)+' width='+barW+' height='+invH.toFixed(1)+' rx=3 fill="'+CHART_INVOICED+'"></rect>';
    bars+='<rect class=chartbar data-tip="'+monthLabel(k)+' — Collected '+money(d.collected)+'" x='+bx2.toFixed(1)+' y='+(padT+plotH-colH).toFixed(1)+' width='+barW+' height='+colH.toFixed(1)+' rx=3 fill="'+CHART_COLLECTED+'"></rect>';
    bars+='<text x='+(gx+groupW/2)+' y='+(h-8)+' text-anchor=middle font-size=10.5 fill="#74716a">'+monthLabel(k)+'</text>'});
  box.innerHTML=
    '<div class=chartlegend><span><i style="background:'+CHART_INVOICED+'"></i>Invoiced</span><span><i style="background:'+CHART_COLLECTED+'"></i>Collected</span></div>'+
    '<div class=chartwrap><svg viewBox="0 0 '+w+' '+h+'" style="width:100%;height:auto;max-width:'+w+'px" id=statement-svg>'+gridSvg+bars+'</svg><div class=charttooltip id=statement-tip></div></div>';
  const tip=$('statement-tip');
  $('statement-svg').querySelectorAll('.chartbar').forEach(el=>{
    el.addEventListener('mousemove',ev=>{
      const rect=$('statement-svg').getBoundingClientRect();
      tip.textContent=el.dataset.tip;
      tip.style.left=(ev.clientX-rect.left)+'px';
      tip.style.top=(ev.clientY-rect.top-10)+'px';
      tip.classList.add('show')});
    el.addEventListener('mouseleave',()=>tip.classList.remove('show'))})}
function renderStatementCompanies(){
  const box=$('statement-companies');
  if(!LEDGER.length){box.innerHTML='<p class="muted" style="font-size:12.5px;padding:8px 0">No outstanding balances to report.</p>';return}
  const byCompany={};
  LEDGER.forEach(e=>{
    const c=e.company||'(no company)';
    byCompany[c]=byCompany[c]||{invoiced:0,collected:0,items:[]};
    byCompany[c].invoiced+=e.total;
    if(e.paid)byCompany[c].collected+=e.total;
    byCompany[c].items.push(e)});
  const companies=Object.keys(byCompany).sort((a,b)=>(byCompany[b].invoiced-byCompany[b].collected)-(byCompany[a].invoiced-byCompany[a].collected));
  box.innerHTML=companies.map(c=>{
    const d=byCompany[c], balance=d.invoiced-d.collected;
    const cid='co-'+c.replace(/[^a-z0-9]/gi,'');
    return '<div>'+
      '<div class=companyrow onclick="toggleCompanyExpand(\''+cid+'\')">'+
        '<span class=companyexpand id="'+cid+'-chev">▸</span>'+
        '<span class=companyname>'+escHtml(c)+'</span>'+
        '<span class=companystat>Invoiced '+money(d.invoiced)+'</span>'+
        '<span class=companystat>Collected '+money(d.collected)+'</span>'+
        '<span class=companybalance style="color:'+(balance>0?'var(--danger)':'var(--success)')+'">'+money(balance)+'</span>'+
      '</div>'+
      '<div class=companyinvoices id="'+cid+'">'+
        d.items.sort((a,b)=>(b.date||'').localeCompare(a.date||'')).map(e=>
          '<div class=invoicerow>'+
            '<span class=muted>INV '+escHtml(e.number||'')+' · '+escHtml(e.date||'')+' · '+money(e.total)+'</span>'+
            '<button class="paytogglebtn '+(e.paid?'paid':'unpaid')+'" onclick="event.stopPropagation();toggleInvoicePaid(\''+escHtml(e.rel).replace(/'/g,"\\'")+'\')">'+(e.paid?'✓ Paid':'Mark Paid')+'</button>'+
          '</div>').join('')+
      '</div>'+
    '</div>'}).join('')}
function toggleCompanyExpand(cid){
  const el=$(cid),chev=$(cid+'-chev');
  const open=el.classList.toggle('open');
  chev.textContent=open?'▾':'▸'}
async function toggleInvoicePaid(rel){
  const entry=LEDGER.find(e=>e.rel===rel);if(!entry)return;
  const r=await fetch('/api/finance/mark-paid',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({rel,paid:!entry.paid})}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  LEDGER=r.ledger||[];
  renderStatement()}

// -------- new submission (manual entry point): step 1 pick a quotation, step 2 confirm + generate DO/INV
let NEWSUB_QTN=null, NEWSUB_ITEMS=[];
async function openNewSubmission(){
  NEWSUB_QTN=null;NEWSUB_ITEMS=[];
  $('newsubtitle').textContent='New Submission — pick a quotation';
  $('newsubbody').innerHTML='<p class=muted style="font-size:12px;padding:8px 0">Loading quotations…</p>';
  $('newsubmodal').classList.remove('hide');
  if(!INDEX.length)await loadIndex();
  $('newsubbody').innerHTML=
    '<input id=newsub-qtn-search placeholder="Search quotations by company, number or project…" style="width:100%;margin-bottom:12px" oninput=renderNewSubQtnList()>'+
    '<div id=newsub-qtn-list style="max-height:360px;overflow:auto"></div>';
  renderNewSubQtnList()}
function closeNewSubmission(){$('newsubmodal').classList.add('hide')}
function renderNewSubQtnList(){
  const q=($('newsub-qtn-search').value||'').trim().toLowerCase();
  const rows=(INDEX||[]).filter(r=>(r.type==='QTN'&&r.ext==='xlsx')||(r.type==='QTN2'&&r.ext==='pdf'))
    .filter(r=>!q||(r.company_label||'').toLowerCase().includes(q)||(r.number||'').includes(q)||(r.project_label||'').toLowerCase().includes(q))
    .slice(0,60);
  $('newsub-qtn-list').innerHTML=rows.length?rows.map(r=>
    '<div class=cpitem style="cursor:pointer" onclick="pickNewSubQtn(\''+escHtml(r.rel).replace(/'/g,"\\'")+'\')">'+
      '<span class=cpname>'+escHtml(r.company_label||'')+' — '+escHtml(r.type)+' '+escHtml(r.number||'')+
      (r.project_label?' · '+escHtml(r.project_label):'')+'</span></div>'
  ).join(''):'<p class="muted" style="font-size:12px;padding:6px 4px">No quotations found.</p>'}
async function pickNewSubQtn(rel){
  const r=await fetch('/api/doc?rel='+encodeURIComponent(rel)).then(r=>r.json());
  if(r.error){alert(r.error);return}
  NEWSUB_QTN=r;
  NEWSUB_ITEMS=(r.items||[]).filter(it=>it.description).map(it=>({...it,lpo_qty:it.qty??''}));
  renderNewSubStep2()}
function renderNewSubStep2(){
  const r=NEWSUB_QTN;
  $('newsubtitle').textContent='New Submission — '+escHtml(r.company||'');
  const rows=NEWSUB_ITEMS.map((it,i)=>
    '<div style="font-size:12.5px;padding:8px 10px;background:var(--surface-2);border:1px solid var(--border);border-radius:8px;margin-bottom:6px">'+
      escHtml((it.description||'').replace(/<[^>]*>/g,' '))+
      '<span class=muted style="float:right">Qty '+escHtml(String(it.qty??''))+'</span></div>').join('');
  $('newsubbody').innerHTML=
    '<p class=muted style="font-size:12px;margin:0 0 12px">This generates a real Delivery Order and Invoice right away, using these quoted quantities — flagged In Progress until delivery is confirmed. The customer\'s LPO comes later, when you build the submittal.</p>'+
    rows+
    '<div style="display:flex;gap:8px;margin-top:14px">'+
      '<button class=btn style="flex:1" onclick=openNewSubmission()>← Back</button>'+
      '<button class="btn dark" style="flex:2" onclick=createSubmission()>Generate DO &amp; Invoice</button>'+
    '</div>'}
async function createSubmission(){
  const r=NEWSUB_QTN;
  const body={qtn_rel:r.original||'',qtn_number:r.number||'',company:r.company||'',project:r.project||'',items:NEWSUB_ITEMS};
  toast('Generating Delivery Order and Invoice…');
  const res=await fetch('/api/submissions',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)}).then(r=>r.json());
  if(res.error){alert(res.error);return}
  closeNewSubmission();await loadSubmissions();toast('Delivery Order '+res.submission.do_number+' and Invoice '+res.submission.inv_number+' generated')}

// -------- All Docs row "Generate ▾" — same underlying capability as New
// Submission above, minus the two-step modal: pick a document directly
// from any sales/billing row instead of re-finding it in a picker.
// Reuses /api/submissions for DO+INV (so Submissions tracking — scanned-DO
// linking, building the final submittal — stays correct either way) and
// /api/generate for everything else (PI/RV/CN are plain HTML_DOC_TYPEs,
// the standalone Tax Invoice off a DO row is a plain FILLERS/xlsx type —
// same paths the rest of the app already uses for each).
// Each row type's own standard "next document" — the same UAE-trading-
// industry chain the Proforma Invoice decision was researched against:
// Quotation/Proforma Invoice → Delivery Order + Tax Invoice; Delivery
// Order → Tax Invoice (catch-up, in case one was never paired); Tax
// Invoice → Payment Receipt or Credit Note. Sololuce Datasheet (CAT) and
// Expense Report (EXP) have no entry — neither is part of this chain.
const ROW_GEN_OPTIONS={
  QTN:[['DO_INV','Delivery Order + Tax Invoice'],['PI','Proforma Invoice']],
  QTN2:[['DO_INV','Delivery Order + Tax Invoice'],['PI','Proforma Invoice']],
  PI:[['DO_INV','Delivery Order + Tax Invoice']],
  DO:[['INV_ONLY','Tax Invoice']],
  INV:[['RV','Payment Receipt'],['CN','Credit Note']],
};
let ROWGEN_REL='',ROWGEN_TYPE='';
function openRowGenMenu(rel,docType,ev){
  ev.stopPropagation();
  ROWGEN_REL=rel;ROWGEN_TYPE=docType;
  const box=$('qtngenmenu');
  const r=ev.currentTarget.getBoundingClientRect();
  const left=Math.min(Math.round(r.left),window.innerWidth-266);
  box.style.left=Math.max(8,left)+'px';box.style.top=(Math.round(r.bottom)+6)+'px';
  box.style.display='block';
  $('qtngen-list').innerHTML=(ROW_GEN_OPTIONS[docType]||[]).map(([kind,label])=>
    '<div class=cpitem style="cursor:pointer" onclick="selectRowGenKind(\''+kind+'\')"><span class=cpname>'+label+'</span></div>').join('')}
function closeQtnGenMenu(){$('qtngenmenu').style.display='none'}
document.addEventListener('click',e=>{
  if($('qtngenmenu').style.display==='block'&&!e.target.closest('#qtngenmenu')&&!e.target.closest('.qtngenbtn'))closeQtnGenMenu()});
function selectRowGenKind(kind){
  if(kind==='CN'){
    // Credit Note is the one action that genuinely can't run with zero
    // input — which items and why aren't inferable from the invoice alone
    // (see render_credit_note_pdf's own docstring) — so the popover swaps
    // its menu for a small reason field instead of generating immediately.
    $('qtngen-list').innerHTML=
      '<div style="padding:10px 12px">'+
        '<label style="font-size:11px;font-weight:600;color:var(--muted);display:block;margin-bottom:6px">Reason for credit</label>'+
        '<textarea id=cn-reason rows=3 style="width:100%;box-sizing:border-box;margin-bottom:8px" placeholder="e.g. Return of 5 units — supplied surplus to site requirement"></textarea>'+
        '<button class="btn dark" style="width:100%" onclick="runRowGenerate(\'CN\')">Generate Credit Note</button>'+
      '</div>';
    return}
  closeQtnGenMenu();runRowGenerate(kind)}
async function runRowGenerate(kind){
  const rel=ROWGEN_REL,docType=ROWGEN_TYPE;
  const cnReasonEl=$('cn-reason'),cnReason=cnReasonEl?cnReasonEl.value.trim():'';
  if(kind==='CN'){
    if(!cnReason){alert('Enter a reason for the credit.');return}
    closeQtnGenMenu()}
  const r=await fetch('/api/doc?rel='+encodeURIComponent(rel)).then(r=>r.json());
  if(r.error){alert(r.error);return}
  const items=(r.items||[]).filter(it=>it.description).map(it=>({...it,lpo_qty:it.qty??it.lpo_qty??''}));
  // Legacy xlsx types (QTN/DO/INV) already return one composed
  // customer_block; QTN2 returns separate attn/address fields (its own
  // richer customer card) — same distinction customerBlockForXlsx() draws
  // elsewhere in this file.
  const customerBlock=r.customer_block||[r.customer_attn?('Attn: '+r.customer_attn):'',r.customer_address||''].filter(Boolean).join('\n');
  const today=new Date(),iso=d=>d.toISOString().slice(0,10);

  if(kind==='DO_INV'){
    toast('Generating Delivery Order and Invoice…');
    // Sourced from a Proforma Invoice, the QTN reference the new DO/INV
    // should carry is the ORIGINAL quotation the PI itself was generated
    // from (r.qtn_number) — not the PI's own number.
    const qtnNumber=docType==='PI'?(r.qtn_number||''):(r.number||'');
    const body={qtn_rel:docType==='PI'?(r.qtn_rel||''):(r.original||rel),qtn_number:qtnNumber,company:r.company||'',project:r.project||'',items};
    const res=await fetch('/api/submissions',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)}).then(r=>r.json());
    if(res.error){alert(res.error);return}
    await loadIndex();
    toast('Delivery Order '+res.submission.do_number+' and Invoice '+res.submission.inv_number+' generated — marked In Progress');
    return}

  if(kind==='PI'){
    toast('Generating Proforma Invoice…');
    const nextR=await fetch('/api/next-number?type=PI').then(r=>r.json());
    const validUntil=new Date(today.getTime()+14*86400000);
    const data={doc_type:'PI',number:String(nextR.next).padStart(4,'0'),date:iso(today),
      qtn_number:r.number||'',qtn_rel:r.original||rel,valid_until:iso(validUntil),company:r.company||'',project:r.project||'',
      customer_block:customerBlock,items,vat:r.vat||{enabled:true,mode:'percent',value:5}};
    const res=await fetch('/api/generate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(data)}).then(r=>r.json());
    if(res.error){alert(res.error);return}
    await loadIndex();
    toast('Proforma Invoice '+data.number+' generated');
    return}

  if(kind==='INV_ONLY'){
    toast('Generating Tax Invoice…');
    const nextR=await fetch('/api/next-number?type=INV').then(r=>r.json());
    const invItems=items.map(it=>({description:it.description,unit:it.unit||'PCS',
      qty:it.delivered||it.lpo_qty||it.qty||'',price:it.price||''}));
    const data={doc_type:'INV',number:String(nextR.next).padStart(4,'0'),date:iso(today),
      lpo_number:r.lpo_number||'',project:r.project||'',company:r.company||'',customer_block:customerBlock,
      items:invItems,discount:{},vat:{enabled:true,mode:'percent',value:5}};
    const res=await fetch('/api/generate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(data)}).then(r=>r.json());
    if(res.error){alert(res.error);return}
    await loadIndex();
    toast('Tax Invoice generated');
    return}

  if(kind==='RV'){
    toast('Generating Payment Receipt…');
    const nextR=await fetch('/api/next-number?type=RV').then(r=>r.json());
    const data={doc_type:'RV',number:String(nextR.next).padStart(4,'0'),date:iso(today),
      payment_method:'',reference:'',company:r.company||'',project:r.project||'',customer_block:customerBlock,
      items,vat:{enabled:true,mode:'percent',value:5},invoice_number:r.number||'',invoice_date:r.date||''};
    const res=await fetch('/api/generate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(data)}).then(r=>r.json());
    if(res.error){alert(res.error);return}
    await loadIndex();
    toast('Payment Receipt generated');
    return}

  if(kind==='CN'){
    toast('Generating Credit Note…');
    const nextR=await fetch('/api/next-number?type=CN').then(r=>r.json());
    const data={doc_type:'CN',number:String(nextR.next).padStart(4,'0'),date:iso(today),
      against_invoice:r.number||'',invoice_date:r.date||'',company:r.company||'',project:r.project||'',
      customer_block:customerBlock,items,vat:{enabled:true,mode:'percent',value:5},reason:cnReason};
    const res=await fetch('/api/generate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(data)}).then(r=>r.json());
    if(res.error){alert(res.error);return}
    await loadIndex();
    toast('Credit Note generated');
    return}
}

let currentClientId=null, currentClientLogo='';
function openClientEditor(id){
  const c=id?CLIENT_RECORDS.find(x=>x.id===id):null;
  currentClientId=id||null;currentClientLogo=c?(c.logo||''):'';
  $('clientmodaltitle').textContent=c?'Edit Client':'New Client';
  $('cf-name').value=c?c.name:'';
  $('cf-category').value=c?(c.category||''):'';
  $('cf-attn').value=c?c.attn:'';
  $('cf-address').value=c?c.address:'';
  $('cf-pobox').value=c?(c.po_box||''):'';
  $('cf-city').value=c?(c.city||''):'';
  $('cf-phone').value=c?(c.phone||''):'';
  $('cf-landline').value=c?(c.landline||''):'';
  $('cf-email').value=c?(c.email||''):'';
  $('cf-website').value=c?(c.website||''):'';
  $('cf-trn').value=c?(c.trn||''):'';
  $('cf-notes').value=c?(c.notes||''):'';
  setClientCountry(c?(c.country||''):'');
  fillDatalist('clientcategories',[...new Set(CLIENT_RECORDS.map(x=>x.category).filter(Boolean))].sort((a,b)=>a.localeCompare(b)));
  updateClientLogoPreview();
  refreshClientMapPreview();
  const delBtn=$('cf-delete');
  delBtn.classList.toggle('hide',!c);delBtn.textContent='Delete';delBtn.dataset.confirm='';
  $('clientmodal').classList.remove('hide')}
function closeClientEditor(){$('clientmodal').classList.add('hide')}
function updateClientLogoPreview(){
  const img=$('clientlogopreview'),ph=$('clientlogoplaceholder'),rm=$('clientlogorm');
  if(currentClientLogo){img.src=currentClientLogo;img.classList.remove('hide');ph.classList.add('hide');rm.classList.remove('hide')}
  else{img.classList.add('hide');ph.classList.remove('hide');rm.classList.add('hide')}}
function pickClientLogo(){$('clientlogofile').click()}
function onClientLogoFile(input){
  const file=input.files&&input.files[0];if(!file)return;
  const reader=new FileReader();
  reader.onload=()=>{currentClientLogo=reader.result;updateClientLogoPreview()};
  reader.readAsDataURL(file);input.value=''}
function removeClientLogo(){currentClientLogo='';updateClientLogoPreview()}
async function saveClientEditor(){
  const name=$('cf-name').value.trim();
  if(!name){alert('Enter a company name.');return}
  const body={id:currentClientId||'',name,category:$('cf-category').value.trim(),attn:$('cf-attn').value.trim(),address:$('cf-address').value.trim(),
    po_box:$('cf-pobox').value.trim(),city:$('cf-city').value.trim(),country:$('cf-country').value,
    phone:$('cf-phone').value.trim(),landline:$('cf-landline').value.trim(),email:$('cf-email').value.trim(),
    website:$('cf-website').value.trim(),trn:$('cf-trn').value.trim(),notes:$('cf-notes').value.trim(),logo:currentClientLogo};
  const r=await fetch('/api/clients',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  closeClientEditor();await loadClientsView();toast('Client saved')}

// ---------------------------------------------------------------- country picker + Google Maps location (client editor)
// Country picker is shared between the Client editor (prefix "cf") and
// Build's Client/Company card (prefix "customer") — countryPickerTarget
// tracks which one opened it, same idea as richFocused for the rich-text
// toolbar. Every target's fields follow <prefix>-country/-countryflag/
// -countrylabel/-countrybtn.
let countryPickerTarget='cf';
function setCountryValue(prefix,code){
  $(prefix+'-country').value=code||'';
  const flag=$(prefix+'-countryflag'),label=$(prefix+'-countrylabel');
  if(code){flag.textContent=countryFlag(code);label.textContent=countryName(code);label.classList.remove('ph')}
  else{flag.textContent='';label.textContent='Select a country…';label.classList.add('ph')}}
function setClientCountry(code){setCountryValue('cf',code)}
function pickCountry(code){
  setCountryValue(countryPickerTarget,code);
  closeCountryPicker();
  if(countryPickerTarget==='cf')refreshClientMapPreview();else schedulePreview()}
function openCountryPicker(ev,target){
  ev.stopPropagation();
  countryPickerTarget=target||'cf';
  const box=$('countrypicker');
  if(box.style.display==='block'){box.style.display='none';return}
  const r=ev.currentTarget.getBoundingClientRect();
  const left=Math.min(Math.round(r.left),window.innerWidth-276);
  box.style.left=Math.max(8,left)+'px';box.style.top=(Math.round(r.bottom)+6)+'px';
  box.style.display='block';
  $('ctp-search').value='';renderCountryPicker();$('ctp-search').focus()}
function closeCountryPicker(){$('countrypicker').style.display='none'}
document.addEventListener('click',e=>{
  if($('countrypicker').style.display==='block'&&!e.target.closest('#countrypicker')&&!e.target.closest('.countrybtn'))closeCountryPicker()});
function renderCountryPicker(){
  const q=($('ctp-search').value||'').trim().toLowerCase();
  const rows=COUNTRIES.filter(c=>!q||c[1].toLowerCase().includes(q)).slice(0,40);
  $('ctp-list').innerHTML=rows.length?rows.map(c=>
    '<div class=cpitem onclick="pickCountry(\''+c[0]+'\')"><span class=cflag>'+countryFlag(c[0])+'</span><span class=cpname>'+escHtml(c[1])+'</span></div>'
  ).join(''):'<p class="muted" style="font-size:12px;padding:6px 4px;margin:0">No matching results.</p>'}
function clientMapsQuery(){
  const parts=[$('cf-address').value.trim(),$('cf-pobox').value.trim()?('PO Box '+$('cf-pobox').value.trim()):'',
    $('cf-city').value.trim(),countryName($('cf-country').value)].filter(Boolean);
  return parts.join(', ')}
function clientMapsUrl(q){return 'https://www.google.com/maps/search/?api=1&query='+encodeURIComponent(q)}
function refreshClientMapPreview(){
  const q=clientMapsQuery();
  const box=$('cf-maplocation');
  if(!q){box.innerHTML='<p class=mapempty>Add an address, city or country to see its location.</p>';return}
  const embed='https://www.google.com/maps?q='+encodeURIComponent(q)+'&output=embed';
  box.innerHTML='<iframe src="'+embed+'" loading=lazy></iframe>'+
    '<div class=mapactions>'+
      '<button class=btn onclick="copyMapsLink()">Copy Maps Link</button>'+
      '<a class=btn style="text-align:center;text-decoration:none" href="'+clientMapsUrl(q)+'" target=_blank>Open in Google Maps</a>'+
    '</div>'}
function copyMapsLink(){
  const url=clientMapsUrl(clientMapsQuery());
  navigator.clipboard.writeText(url).then(()=>toast('Maps link copied — paste it into WhatsApp or anywhere else')).catch(()=>toast('Could not copy — long-press the link to copy manually'))}

// ---------------------------------------------------------------- Build-tab quick client picker (top of the Client/Company card)
async function openClientPicker(ev){
  ev.stopPropagation();
  const box=$('clientpicker');
  if(box.style.display==='block'){box.style.display='none';return}
  const r=ev.currentTarget.getBoundingClientRect();
  const left=Math.min(Math.round(r.left),window.innerWidth-296);
  box.style.left=Math.max(8,left)+'px';box.style.top=(Math.round(r.bottom)+6)+'px';
  box.style.display='block';
  $('cp-search').value='';
  if(!CLIENT_RECORDS.length){const r2=await fetch('/api/clients').then(r=>r.json());CLIENT_RECORDS=r2.clients||[];refreshClientNamesDatalist()}
  renderClientPicker();
  $('cp-search').focus()}
function closeClientPicker(){$('clientpicker').style.display='none'}
document.addEventListener('click',e=>{
  if($('clientpicker').style.display==='block'&&!e.target.closest('#clientpicker')&&!e.target.closest('#clientpickerbtn'))closeClientPicker()});
function renderClientPicker(){
  const q=($('cp-search').value||'').trim().toLowerCase();
  const rows=CLIENT_RECORDS.filter(c=>!q||c.name.toLowerCase().includes(q)).sort((a,b)=>a.name.localeCompare(b.name)).slice(0,30);
  $('cp-list').innerHTML=rows.length?rows.map(c=>{
    const ph=c.logo?'<img src="'+c.logo+'">':'<div class=cpph>'+clientInitial(c.name)+'</div>';
    return '<div class=cpitem onclick="applyClientPick(\''+c.id+'\')">'+ph+'<span class=cpname>'+escHtml(c.name)+'</span></div>'
  }).join(''):'<p class="muted" style="font-size:12px;padding:6px 4px;margin:0">'+(CLIENT_RECORDS.length?'No matching results.':'No client records saved yet.')+'</p>'}
function applyClientPick(id){
  const c=CLIENT_RECORDS.find(x=>x.id===id);if(!c)return;
  setCompanyVal(c.name);
  $('customer_attn').textContent=c.attn||'';$('customer_address').textContent=c.address||'';setBuildAddressExtras(c);
  onCompany();schedulePreview();closeClientPicker();toast('Loaded '+c.name)}
async function saveCurrentAsClient(){
  const name=companyVal().replace(/<[^>]*>/g,'').trim();
  if(!name){alert('Type a company name first.');return}
  const existing=CLIENT_RECORDS.find(c=>c.name.toLowerCase()===name.toLowerCase());
  // Only overwrite attn/address (the fields Build actually edits) — every
  // other client field falls back to whatever the existing record already
  // has, so this quick-save path can't silently wipe category/phone/
  // landline/website/trn/notes/logo that were set elsewhere (Clients tab,
  // or the "Import from documents" backfill).
  const body={id:existing?existing.id:'',name,
    category:existing?existing.category:'',
    attn:richText($('customer_attn')),
    address:richText($('customer_address')),
    po_box:$('customer_pobox').value.trim(),
    city:$('customer_city').value.trim(),
    country:$('customer-country').value,
    phone:existing?existing.phone:'',landline:existing?existing.landline:'',
    email:existing?existing.email:'',website:existing?existing.website:'',
    trn:existing?existing.trn:'',notes:existing?existing.notes:'',
    logo:existing?existing.logo:''};
  const r=await fetch('/api/clients',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)}).then(r=>r.json());
  if(r.error){alert(r.error);return}
  const r2=await fetch('/api/clients').then(r=>r.json());CLIENT_RECORDS=r2.clients||[];refreshClientNamesDatalist();
  closeClientPicker();toast(existing?'Client updated':'Client saved')}

const CAP=6, expanded=new Set();
function onCompanyFilter(){const v=$('fcompany').value;if(v!=='all')expanded.add(v);renderList()}
function expandCo(name){expanded.add(name);renderList()}
function onRangePreset(){
  const v=$('frange').value,custom=v==='custom';
  $('ffrom').classList.toggle('hide',!custom);$('fto').classList.toggle('hide',!custom);
  if(!custom){
    let from='';const today=new Date();
    if(v==='7'){const d=new Date();d.setDate(d.getDate()-7);from=d.toISOString().slice(0,10)}
    else if(v==='30'){const d=new Date();d.setDate(d.getDate()-30);from=d.toISOString().slice(0,10)}
    else if(v==='month'){from=today.toISOString().slice(0,7)+'-01'}
    else if(v==='year'){from=today.getFullYear()+'-01-01'}
    $('ffrom').value=from;$('fto').value=''}
  renderList()}
let ALLDOCS_TYPE='all', ALLDOCS_DRAFTS=[];
// Drafts get their own tab rather than being interleaved into the
// date-sorted document list — a draft's only timestamp is "last saved"
// (updated), not a business document date, so mixing the two into one
// sorted list would conflate two different meanings of "date".
async function setAllDocsType(t){
  ALLDOCS_TYPE=t;
  document.querySelectorAll('#alldocstypeseg button').forEach(b=>b.classList.toggle('on',b.dataset.t===t));
  const isDrafts=t==='DRAFTS';
  ['fcompany','frange','fgroup'].forEach(id=>$(id).classList.toggle('hide',isDrafts));
  // Folder order's own starting point per tab — A-Z for CAT's category
  // folders (a fixed, known set — see allDocsGroupKey's own comment),
  // Recent first for every other type's company folders (the original,
  // still-standard behavior). Just the DEFAULT now, not a hardcoded rule
  // — fgroup's own onchange (see below) lets it be switched either way on
  // any tab, per explicit request to make this a real option.
  if(!isDrafts)$('fgroup').value=(t==='CAT')?'az':'recent';
  updateFsortVisibility();
  if(isDrafts){
    $('ffrom').classList.add('hide');$('fto').classList.add('hide');
    const res=await fetch('/api/drafts').then(r=>r.json());
    ALLDOCS_DRAFTS=res.drafts||[]}
  renderList()}
// fsort (date newest/oldest) only means something once rows are actually
// ordered by date — in A-Z mode they're ordered by name instead (see
// renderList), so fsort has nothing left to control there and stays
// hidden rather than sitting visible but silently ignored.
function updateFsortVisibility(){
  $('fsort').classList.toggle('hide', ALLDOCS_TYPE==='DRAFTS'||$('fgroup').value==='az')}
function onGroupOrderChange(){updateFsortVisibility();renderList()}
// CAT (Sololuce Datasheet) records group by category/series instead of
// product name in All Docs — every other doc type keeps grouping by its
// own company field, unchanged, per explicit request scoped to Sololuce
// Datasheets only. A CAT record's "company" slot IS the product name and
// its "project" slot IS the series/category (see collectCatData's own
// company:productName, project:series, which build_filename bakes
// straight into the saved filename — see engine.parse_filename), so this
// just swaps which of those two already-existing fields plays which role
// for CAT specifically, on top of the exact same grouped-list rendering
// every other type already uses below.
function allDocsGroupKey(r){return r.type==='CAT'?(r.project_label||r.company_label):r.company_label}

// ---------------------------------------------------------------- All Docs multi-select (choose several real saved documents, then Clone/Cut/Delete as a batch)
// ALLDOCS_SELECTED holds rel strings (not indices — All Docs re-sorts/
// re-groups on every filter change, so a rel is the only stable key).
// ALLDOCS_CUT is a small separate clipboard of rels marked to move: unlike
// the line-item clipboard, "paste" here has a required destination (which
// company/project group), so cutting alone doesn't do anything until
// pasteAllDocsHere() is actually clicked on a group.
let ALLDOCS_SELECTED=new Set(), ALLDOCS_CUT=[], ALLDOCS_SELECT_MODE=false;
function enterAllDocsSelectMode(){ALLDOCS_SELECT_MODE=true;$('listbox').classList.add('selectmode')}
function exitAllDocsSelectMode(){ALLDOCS_SELECT_MODE=false;$('listbox').classList.remove('selectmode');clearAllDocsSelection()}
// Same Explorer/Finder convention as openItemCtxMenu: right-click on an
// already-selected row acts on the whole selection; right-click on an
// unselected one replaces the selection with just that row.
function openAllDocsCtxMenu(rel,ev){
  if(!ALLDOCS_SELECTED.has(rel)){ALLDOCS_SELECTED.clear();ALLDOCS_SELECTED.add(rel)}
  enterAllDocsSelectMode();
  renderList();
  const n=ALLDOCS_SELECTED.size;
  const row=INDEX.find(x=>x.rel===rel);
  const isCat=row&&row.type==='CAT';
  showCtxMenu(ev,
    '<div class=cpitem style="cursor:pointer" onclick="closeCtxMenu();bulkCloneAllDocs()"><span class=cpname>⧉ Clone'+(n>1?' ('+n+')':'')+'</span></div>'+
    (isCat?'':'<div class=cpitem style="cursor:pointer" onclick="closeCtxMenu();bulkCutAllDocs()"><span class=cpname>✂ Cut'+(n>1?' ('+n+')':'')+'</span></div>')+
    // Delete stays a two-step confirm even from here — the context menu
    // hands off to the bulk bar's own guarded Delete button (see
    // bulkDeleteAllDocs) rather than deleting real files on one click.
    '<div class=cpitem style="cursor:pointer;color:var(--danger)" onclick="closeCtxMenu();toast(\'Click Delete in the toolbar to confirm\')"><span class=cpname>🗑 Delete'+(n>1?' ('+n+')':'')+'…</span></div>')}
function toggleAllDocsSelect(rel,checked){
  if(checked)ALLDOCS_SELECTED.add(rel);else ALLDOCS_SELECTED.delete(rel);
  const row=document.querySelector('.row[data-rel="'+rel+'"]');
  if(row)row.classList.toggle('selected',checked);
  updateAllDocsBulkBar()}
function clearAllDocsSelection(){
  ALLDOCS_SELECTED.clear();
  document.querySelectorAll('.list .row.selected').forEach(r=>r.classList.remove('selected'));
  document.querySelectorAll('.adsel').forEach(cb=>cb.checked=false);
  updateAllDocsBulkBar()}
function updateAllDocsBulkBar(){
  const bar=$('alldocsbulkbar');if(!bar)return;
  const n=ALLDOCS_SELECTED.size;
  bar.classList.toggle('hide',n===0);
  if(n)$('alldocsbulkcount').textContent=n+' document'+(n>1?'s':'')+' selected';
  const delBtn=$('alldocsdelbtn');
  if(delBtn){delBtn.dataset.confirm='';delBtn.innerHTML='🗑 Delete'}}
function updateAllDocsCutBar(){
  const bar=$('alldocscutbar');if(!bar)return;
  const n=ALLDOCS_CUT.length;
  bar.classList.toggle('hide',n===0);
  if(n)$('alldocscutcount').textContent=n+' document'+(n>1?'s':'')+' cut';}
async function bulkCloneAllDocs(){
  if(!ALLDOCS_SELECTED.size)return;
  const rels=[...ALLDOCS_SELECTED];
  toast('Cloning '+rels.length+' document'+(rels.length>1?'s':'')+'…');
  const res=await fetch('/api/alldocs-clone',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({rels})}).then(r=>r.json());
  exitAllDocsSelectMode();
  await loadIndex();
  if(res.errors&&res.errors.length)alert('Some documents could not be cloned:\n'+res.errors.map(e=>e.rel+': '+e.error).join('\n'));
  else toast(res.cloned.length+' document'+(res.cloned.length>1?'s':'')+' cloned')}
function bulkDeleteAllDocs(btn){
  // Same click-again-to-confirm pattern used elsewhere in this app (see
  // e.g. actuallyDeleteFinishColorFromSettings's neighbor) instead of a
  // native confirm() dialog, which this app's own environment silently
  // no-ops — a real risk to leave unguarded given this permanently
  // deletes real business documents.
  if(btn.dataset.confirm!=='1'){
    btn.dataset.confirm='1';btn.textContent='Sure? Click again';
    setTimeout(()=>{if(btn.dataset.confirm==='1'){btn.dataset.confirm='';btn.innerHTML='🗑 Delete'}},3000);
    return}
  btn.dataset.confirm='';
  runBulkDeleteAllDocs()}
async function runBulkDeleteAllDocs(){
  const rels=[...ALLDOCS_SELECTED];
  toast('Deleting '+rels.length+' document'+(rels.length>1?'s':'')+'…');
  const res=await fetch('/api/alldocs-delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({rels})}).then(r=>r.json());
  exitAllDocsSelectMode();
  await loadIndex();
  if(res.errors&&res.errors.length)alert('Some documents could not be deleted:\n'+res.errors.map(e=>e.rel+': '+e.error).join('\n'));
  else toast(res.deleted.length+' document'+(res.deleted.length>1?'s':'')+' deleted')}
let ALLDOCS_CUT_GROUPS=new Set();
function bulkCutAllDocs(){
  if(!ALLDOCS_SELECTED.size)return;
  const all=[...ALLDOCS_SELECTED];
  const catRels=all.filter(rel=>{const r=INDEX.find(x=>x.rel===rel);return r&&r.type==='CAT'});
  ALLDOCS_CUT=all.filter(rel=>!catRels.includes(rel));
  // Remembered so Paste-here can hide itself on a cut item's own current
  // group — pasting a document back where it already sits isn't a move,
  // it's a same-path collision that would only surface as a confusing
  // "already exists" error.
  ALLDOCS_CUT_GROUPS=new Set(ALLDOCS_CUT.map(rel=>allDocsGroupKey(INDEX.find(x=>x.rel===rel))));
  exitAllDocsSelectMode();
  renderList();updateAllDocsCutBar();
  if(catRels.length)toast((ALLDOCS_CUT.length)+' document'+(ALLDOCS_CUT.length===1?'':'s')+' cut — '+catRels.length+' Sololuce Datasheet'+(catRels.length>1?'s':'')+' skipped (grouped by project, not company)');
  else toast(ALLDOCS_CUT.length+' document'+(ALLDOCS_CUT.length>1?'s':'')+' cut — click "Paste here" on a company/project group to move '+(ALLDOCS_CUT.length>1?'them':'it'))}
function clearAllDocsCut(){ALLDOCS_CUT=[];ALLDOCS_CUT_GROUPS=new Set();renderList();updateAllDocsCutBar()}
async function pasteAllDocsHere(targetCompany){
  if(!ALLDOCS_CUT.length)return;
  const rels=ALLDOCS_CUT;
  toast('Moving '+rels.length+' document'+(rels.length>1?'s':'')+' to '+targetCompany+'…');
  const res=await fetch('/api/alldocs-move',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({rels,target_company:targetCompany})}).then(r=>r.json());
  ALLDOCS_CUT=[];ALLDOCS_CUT_GROUPS=new Set();
  updateAllDocsCutBar();
  await loadIndex();
  if(res.errors&&res.errors.length)alert('Some documents could not be moved:\n'+res.errors.map(e=>e.rel+': '+e.error).join('\n'));
  else toast(res.moved.length+' document'+(res.moved.length>1?'s':'')+' moved to '+targetCompany)}
function allDocsRowLabel(r){return r.type==='CAT'?(r.company_label||r.project_label):(r.project_label||r.company_label)}
function renderList(){
  if(ALLDOCS_TYPE==='DRAFTS')return renderDraftRows();
  const q=($('search').value||'').toLowerCase(),ft=ALLDOCS_TYPE,fc=$('fcompany').value,
    from=$('ffrom').value,to=$('fto').value,sortDir=$('fsort').value,groupOrder=$('fgroup').value;
  // "Quotation" matches both QTN2 (current) and QTN (legacy Excel) records —
  // one filter tab for what's conceptually one category, even though the
  // two are generated/persisted by entirely different pipelines under the hood.
  // fcompany still filters by the actual company/product field (r.company_label)
  // regardless of type — a separate concern from allDocsGroupKey's grouping/
  // display swap above, so it's untouched here.
  let rows=INDEX.filter(r=>(ft=='all'||r.type==ft||(ft==='QTN2'&&r.type==='QTN'))&&(fc=='all'||r.company_label==fc)
    &&(!q||(r.company_label||'').toLowerCase().includes(q)||(r.number||'').includes(q)||(r.project_label||'').toLowerCase().includes(q))
    &&(!from||r.date>=from)&&(!to||r.date<=to));
  if(!rows.length){$('listbox').innerHTML=emptyStateHtml(EMPTY_ICON_DOCS,'No Documents Found','No documents match the selected filters. Try expanding the date range or clearing the search.');return}
  // fgroup ("Folders: A-Z" / "Folders: Recent first") — a real switchable
  // option (was a hardcoded CAT-only rule at first, per explicit follow-up
  // request to make it a real option instead) controlling BOTH levels at
  // once: which document within a folder comes first, and which folder
  // itself comes first.
  //   A-Z: documents sort by their own name — allDocsRowLabel, which for
  //   CAT IS the product name (see that function's own comment: CAT's
  //   "company" field holds the product name) — per explicit follow-up
  //   ("the sorting alphabetical should be reading the name of the
  //   Product"), not by date at all; folders (categories for CAT,
  //   companies for everyone else) then also sort A-Z by that same key.
  //   Recent first: the original behavior, unchanged — documents sort by
  //   date (fsort controls newest/oldest), folders order by whichever has
  //   the most recent activity (first appearance under that date sort).
  if(groupOrder==='az'){
    rows.sort((a,b)=>(allDocsRowLabel(a)||'').localeCompare(allDocsRowLabel(b)||''));
  } else {
    rows.sort((a,b)=>a.date<b.date?1:a.date>b.date?-1:0);          // newest first
    if(sortDir==='asc')rows.reverse();
  }
  const groupRank={};
  if(groupOrder==='az'){
    const groups=[...new Set(rows.map(allDocsGroupKey))].sort((a,b)=>a.localeCompare(b));
    groups.forEach((g,i)=>groupRank[g]=i);
  } else {
    rows.forEach(r=>{const g=allDocsGroupKey(r);if(!(g in groupRank))groupRank[g]=Object.keys(groupRank).length});
  }
  rows.sort((a,b)=>groupRank[allDocsGroupKey(a)]-groupRank[allDocsGroupKey(b)]);  // stable: keeps the row order set above within each group
  const totals={};rows.forEach(r=>{const g=allDocsGroupKey(r);totals[g]=(totals[g]||0)+1});
  // Rows for one group render inside a single rounded .rowgroup cluster
  // (iOS grouped-table style) rather than each being its own bordered card —
  // opened lazily on the first visible row of a group, closed when the group
  // changes or when the "+N more" cap is hit, so the "+N more" prompt itself
  // sits outside/below the cluster as its own distinct affordance.
  let html='',lastCo=null,coCount=0,groupOpen=false;
  rows.forEach(r=>{
    const g=allDocsGroupKey(r);
    if(g!==lastCo){
      if(groupOpen){html+='</div>';groupOpen=false}
      // CAT groups by project, and its "company" field actually holds the
      // product name (see allDocsGroupKey's own special-case) — pasting a
      // cut document there would rename a PRODUCT, not organize by
      // customer, so Paste-here never shows for a CAT group.
      const pasteHere=(ALLDOCS_CUT.length&&r.type!=='CAT'&&!ALLDOCS_CUT_GROUPS.has(g))
        ?'<button type=button class=btn style="margin-left:10px;font-size:11px;padding:4px 10px;font-weight:600;text-transform:none;letter-spacing:0" onclick="event.stopPropagation();pasteAllDocsHere(\''+g.replace(/'/g,"\\'")+'\')">📋 Paste '+ALLDOCS_CUT.length+' here</button>':'';
      html+='<div class=cogroup>'+g+' <span class=muted style="font-weight:400;text-transform:none;letter-spacing:0">('+totals[g]+')</span>'+pasteHere+'</div>';
      lastCo=g;coCount=0}
    coCount++;
    if(!q&&!expanded.has(g)&&coCount>CAP){
      if(coCount===CAP+1){
        if(groupOpen){html+='</div>';groupOpen=false}
        const co=g.replace(/'/g,"\\'");
        html+='<div class=more onclick="expandCo(\''+co+'\')">+'+(totals[g]-CAP)+' more for '+g+' — click to show all</div>'}
      return}
    if(!groupOpen){html+='<div class=rowgroup>';groupOpen=true}
    const formats=r.formats||[],hasXlsx=formats.includes('xlsx'),hasPdf=formats.includes('pdf'),
      hasWord=formats.includes('doc')||formats.includes('docx'),canPreview=hasPdf||hasXlsx||hasWord,
      rel=r.rel.replace(/'/g,"\\'"),base=rel.replace(/\.[^./\\]+$/,''),
      xlsxRel=base+'.xlsx',pdfRel=base+'.pdf',wordRel=base+'.'+(formats.includes('docx')?'docx':'doc'),
      label=(allDocsRowLabel(r)||'').replace(/'/g,"\\'"),
      folderCompany=(r.folder_company||'').replace(/'/g,"\\'"),
      // Delivery Orders are the one type this business regularly keeps as
      // PDF-only (kept as-is after physical signature) — a best-effort PDF
      // text-scrape edit (see _api_doc_from_pdf). QTN2/CAT/EXP are also
      // PDF-only by design (Playwright-rendered, no xlsx twin) but don't need
      // any scraping — each already saves an exact JSON sidecar alongside
      // its PDF on Generate (engine.save_sidecar), so they get the same
      // real edit path, just reading that back instead.
      isDoPdf=!hasXlsx&&hasPdf&&(r.type||'').toUpperCase()==='DO',
      // isHtmlDocType decides "does /api/doc return real sidecar data" (true
      // for all of QTN2/CAT/EXP/PI/RV/CN); isEditableHtmlDocType decides
      // "does the Build form actually know how to DISPLAY that data" (only
      // QTN2/CAT/EXP have a HEAD/COLS entry or dedicated populate*Form —
      // PI/RV/CN are generate-only for now, see HTML_DOC_TYPES's own
      // comment, so routing them through openDoc() would crash on
      // HEAD[TYPE] being undefined; they get the same read-only preview any
      // other non-editable type gets instead).
      isHtmlDocType=HTML_DOC_TYPES.includes((r.type||'').toUpperCase()),
      isEditableHtmlDocType=EDITABLE_HTML_DOC_TYPES.includes((r.type||'').toUpperCase());
    let actions='';
    if(hasXlsx)actions+='<button class=rbtn onclick="event.stopPropagation();openNative(\''+xlsxRel+'\')" title="Open in Excel">Excel</button>';
    if(hasWord)actions+='<button class=rbtn onclick="event.stopPropagation();openNative(\''+wordRel+'\')" title="Open in Word">Word</button>';
    if(hasPdf)actions+='<button class=rbtn onclick="event.stopPropagation();openNative(\''+pdfRel+'\')" title="Open in your PDF reader">PDF</button>';
    actions+=hasXlsx
      ? '<button class="rbtn cs" onclick="event.stopPropagation();openDoc(\''+xlsxRel+'\')" title="Edit in Company System">Open in CS</button>'
      : (isDoPdf||isEditableHtmlDocType)
        ? '<button class="rbtn cs" onclick="event.stopPropagation();openDoc(\''+pdfRel+'\')" title="Edit in Company System">Open in CS</button>'
        : '<button class="rbtn cs" onclick="event.stopPropagation();openCS(\''+rel+'\')" title="Preview in Company System">Open in CS</button>';
    // Every sales/billing document can spin off the standard next document
    // in its own chain (Quotation/Proforma Invoice → Delivery Order + Tax
    // Invoice; Delivery Order → Tax Invoice; Tax Invoice → Payment Receipt
    // or Credit Note) — see ROW_GEN_OPTIONS/openRowGenMenu. Deliberately
    // excludes Sololuce Datasheet (CAT) and Expense Report (EXP): neither
    // is part of this sales-document chain, so they have no entry in
    // ROW_GEN_OPTIONS and this stays empty/hidden for them.
    const rowGenType=(r.type||'').toUpperCase();
    if(ROW_GEN_OPTIONS[rowGenType])
      actions+='<button class="rbtn qtngenbtn" onclick="openRowGenMenu(\''+rel+'\',\''+rowGenType+'\',event)" title="Generate the next document in this chain">Generate ▾</button>';
    // QTN2 shows the same short "QTN" badge as legacy QTN — they're one
    // merged "Quotation" filter now, so the badge shouldn't visibly split
    // back into two different-looking codes within it.
    const pillText=r.type==='QTN2'?'QTN':(r.type||'DOC');
    html+='<div class="row'+(r.in_progress?' inprogress':'')+(ALLDOCS_SELECTED.has(rel)?' selected':'')+'" data-rel="'+rel+'" data-type="'+(r.type||'')+'" onclick="openFileMenu(event,\''+rel+'\',\''+folderCompany+'\',\''+label+'\')" oncontextmenu="openAllDocsCtxMenu(\''+rel+'\',event)" onmouseenter="showPrev(event,\''+rel+'\','+canPreview+')" onmousemove=movePrev(event) onmouseleave=hidePrev()>'+
      '<input type=checkbox class=adsel onclick="event.stopPropagation()" onchange="toggleAllDocsSelect(\''+rel+'\',this.checked)"'+(ALLDOCS_SELECTED.has(rel)?' checked':'')+'>'+
      '<span class=pill>'+pillText+'</span><span class=mono style=width:70px>'+(r.number||'—')+'</span>'+
      '<span style=flex:1>'+allDocsRowLabel(r)+'</span>'+
      (r.status&&r.status!=='None'?'<span class="statuspill '+r.status+'">'+r.status+'</span>':'')+
      (r.in_progress?'<span class="pill inprogresspill">In Progress</span>':'')+
      '<span class=muted style=width:100px>'+r.date+'</span>'+
      '<span class=rowactions>'+actions+'</span></div>'
  });
  if(groupOpen)html+='</div>';
  $('listbox').innerHTML=html}
function renderDraftRows(){
  const q=($('search').value||'').toLowerCase();
  const rows=ALLDOCS_DRAFTS.filter(d=>!q||(d.label||'').toLowerCase().includes(q)||(d.company||'').toLowerCase().includes(q))
    .sort((a,b)=>(b.updated||'').localeCompare(a.updated||''));
  if(!rows.length){$('listbox').innerHTML=emptyStateHtml(EMPTY_ICON_DOCS,'No Drafts','Unfinished work saves here automatically — pick up where you left off anytime.');return}
  $('listbox').innerHTML='<div class=rowgroup>'+rows.map(d=>
    '<div class=row onclick="resumeDraftFromAllDocs(\''+d.id+'\')">'+
      '<span class=pill>'+(d.doc_type==='QTN2'?'QTN':d.doc_type)+'</span>'+
      '<span style=flex:1>'+escHtml(d.label)+'</span>'+
      '<span class=muted style=width:150px>Saved '+escHtml((d.updated||'').slice(0,16).replace('T',' '))+'</span>'+
      '<span class=rowactions><button class=rbtn onclick="event.stopPropagation();deleteDraftFromAllDocs(\''+d.id+'\')" title="Delete draft">Delete</button></span>'+
    '</div>'
  ).join('')+'</div>'}
function resumeDraftFromAllDocs(id){DRAFTS=ALLDOCS_DRAFTS;loadDraft(id)}
async function deleteDraftFromAllDocs(id){
  const r=await fetch('/api/drafts-delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})}).then(r=>r.json());
  ALLDOCS_DRAFTS=r.drafts||[];
  if(EDITING_DRAFT===id)EDITING_DRAFT=null;
  renderList()}

// Everything that loads real data or hits an /api/* endpoint is gated
// behind login now (see app.py's before_request) — deferred into
// bootApp(), called only after checkLogin()/doLogin() confirms a session.
// initResizer()'s own calls stay outside: they only wire up drag behavior
// on static DOM elements, no API calls, safe to run immediately.
let LOGGED_IN=false, CURRENT_USER=null, CURRENT_ROLE=null, BRAND_LOCK=null, BLOCKED_TOOLS=[];
function bootApp(){
  loadUnits().then(()=>{setType('QTN2');$('title').textContent='Menu'});loadCfg();loadBrands();loadClients();loadClientsView();checkUnfinishedDraftsOnLaunch();
  initUpdateChecking(); // respects the "check for updates on each start" preference — see Update Center
  loadCatSpecLabels();loadCatFinishColors();loadCatBadgeLibrary();loadCatSeriesLabels();loadCatFamilyLabels();loadCatOrdCategories();loadCatModelNoOptions();loadCatCctOptions();loadCatControlsOptions();loadCatVoltageOptions();loadCatPowerOptions();loadCatBeamAngleOptions();loadCatCutOutOptions();loadCatOptionsOptions();loadCatSizeIndex();loadCatSpecValues();loadCatStandardFillFields();loadCatStandardFillValues();loadExpPaymentMethods();loadExpEmployees();loadExpCategories();loadExpProducts();loadExpDescriptions();
  renderCatBadges();renderCatSpecs();renderCatFinish();renderCatImages();renderCatOrdTable();renderExpItems();renderExpEmployeeField('');renderExpCategoryField('');
  initRichText();
  renderRichSwatches();
  attachAutocomplete($('company-rich'),allCompanyNames);
  attachAutocomplete($('customer_attn'),allAttns);
  attachAutocomplete($('customer_address'),allAddresses);
}

// ---------------------------------------------------------------- Login / accounts
async function checkLogin(){
  const r=await fetch('/api/current-user').then(r=>r.json()).catch(()=>({logged_in:false}));
  if(r.logged_in){applySession(r);$('loginoverlay').classList.add('hide');bootApp()}
  else{$('loginoverlay').classList.remove('hide');setTimeout(()=>$('login-username').focus(),50)}
}
function applySession(u){
  LOGGED_IN=true;CURRENT_USER=u.username;CURRENT_ROLE=u.role;BRAND_LOCK=u.brand_lock;BLOCKED_TOOLS=u.blocked_tools||[];
  applyAccessRestrictions()}
function applyAccessRestrictions(){
  // Brand lock: hide the brand switcher entirely and pin BRAND once
  // loadBrands()/applyBrandUI() run — see their own code for where BRAND
  // gets set from cfg (the server already forced cfg.brand to the lock at
  // login, see /api/login).
  $('brandbtn').style.display=BRAND_LOCK?'none':'';
  // Blocked tools: hide the corresponding rail nav button. Server-side
  // enforcement (app.py's before_request) is what actually matters —
  // this is just so a restricted user isn't shown a button that 403s.
  const navByTool={settings:['n-settings','t-settings','t-settings-title'],clients:['n-clients','t-clients'],submissions:['n-submissions','t-submissions'],statement:['n-statement','t-statement'],alldocs:['n-all','t-all']};
  Object.entries(navByTool).forEach(([tool,ids])=>ids.forEach(id=>{
    const el=$(id);if(el)el.classList.toggle('hide',BLOCKED_TOOLS.includes(tool))}));
  // Admin Tools is a sub-page WITHIN Settings (not its own rail item) —
  // just the entry card's visibility, per explicit request. Always land
  // back on the main Settings panel on login, in case a previous session
  // left the admin sub-page showing.
  $('admin-tools-btn').style.display=CURRENT_ROLE==='admin'?'':'none';
  showSettingsMainPanel();
  $('logoutlabel').textContent='Sign out ('+CURRENT_USER+')'}
async function doLogin(ev){
  ev.preventDefault();
  const btn=$('login-submit'),err=$('login-error');
  btn.disabled=true;btn.textContent='Signing in…';err.classList.add('hide');
  try{
    const r=await fetch('/api/login',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({username:$('login-username').value,password:$('login-password').value,remember:$('login-remember').checked})}).then(r=>r.json());
    if(!r.ok){err.textContent=r.error||'Sign in failed';err.classList.remove('hide');btn.disabled=false;btn.textContent='Sign In';return false}
    $('login-password').value='';
    applySession(r);$('loginoverlay').classList.add('hide');bootApp()
  }catch(e){err.textContent='Could not reach the app — try again';err.classList.remove('hide');btn.disabled=false;btn.textContent='Sign In'}
  return false}
async function doLogout(){
  await fetch('/api/logout',{method:'POST'}).catch(()=>{});
  location.reload()}

// ---- Admin: Users & Access (Settings) ----
let USERS_LIST=[],USERS_BLOCKABLE=[],USERS_BRANDS=[],EDITING_USERNAME=null;
const TOOL_LABEL={settings:'Settings',clients:'Clients',submissions:'Submissions',statement:'Statement',alldocs:'All Docs'};
async function loadUsersAdmin(){
  const r=await fetch('/api/accounts').then(r=>r.json()).catch(()=>null);
  if(!r)return;
  USERS_LIST=r.users;USERS_BLOCKABLE=r.blockable_tools;USERS_BRANDS=r.brands;
  const sel=$('user-brand-lock');
  sel.innerHTML='<option value="">— none, sees brand switcher —</option>'+
    USERS_BRANDS.map(b=>'<option value="'+b+'">'+escHtml(engineBrandLabel(b))+'</option>').join('');
  $('user-blocked-tools').innerHTML=USERS_BLOCKABLE.map(t=>
    '<label class=usertoolchip><input type=checkbox value="'+t+'" class=user-tool-cb> '+TOOL_LABEL[t]+'</label>').join('');
  renderUsersList()}
function engineBrandLabel(code){
  const b=BRAND_LIST.find(x=>x.code===code);return b?b.label:code}
function renderUsersList(){
  $('users-list').innerHTML=USERS_LIST.map(u=>
    '<div class=usercard><b>'+escHtml(u.username)+'</b>'+
      '<span class="userbadge'+(u.role==='admin'?' admin':'')+'">'+(u.role==='admin'?'Admin':(u.brand_lock?escHtml(engineBrandLabel(u.brand_lock)):'All brands'))+'</span>'+
      '<button type=button class=btn style="padding:4px 9px;font-size:11px" onclick="editUser(\''+escHtml(u.username).replace(/'/g,"\\'")+'\')">Edit</button>'+
      '<button type=button class=btn style="padding:4px 9px;font-size:11px" onclick="deleteUserConfirm(\''+escHtml(u.username).replace(/'/g,"\\'")+'\',this)">✕</button></div>').join('')
    || '<p class=muted style="font-size:12px">No users yet.</p>'}
function editUser(username){
  const u=USERS_LIST.find(x=>x.username===username);if(!u)return;
  EDITING_USERNAME=username;
  $('user-username').value=u.username;$('user-username').disabled=true;
  $('user-password').value='';$('user-password-label').textContent='New password (leave blank to keep current)';
  $('user-role').value=u.role;
  $('user-brand-lock').value=u.brand_lock||'';
  document.querySelectorAll('.user-tool-cb').forEach(cb=>cb.checked=(u.blocked_tools||[]).includes(cb.value));
  renderUserRoleFields()}
function resetUserForm(){
  EDITING_USERNAME=null;
  $('user-username').value='';$('user-username').disabled=false;
  $('user-password').value='';$('user-password-label').textContent='Password';
  $('user-role').value='user';$('user-brand-lock').value='';
  document.querySelectorAll('.user-tool-cb').forEach(cb=>cb.checked=false);
  renderUserRoleFields()}
function renderUserRoleFields(){
  $('user-restrict-fields').style.display=$('user-role').value==='admin'?'none':''}
async function saveUserForm(){
  const username=$('user-username').value.trim();
  if(!username){toast('Username required');return}
  const blocked=[...document.querySelectorAll('.user-tool-cb:checked')].map(cb=>cb.value);
  const body={username,password:$('user-password').value||undefined,role:$('user-role').value,
    brand_lock:$('user-brand-lock').value||null,blocked_tools:blocked};
  const r=await fetch('/api/accounts-save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)}).then(r=>r.json());
  if(!r.ok){toast(r.error||'Could not save user');return}
  toast('Saved '+username+' — remember to Publish so other installs see it');
  resetUserForm();loadUsersAdmin()}
function deleteUserConfirm(username,btn){
  if(btn.dataset.confirm!=='1'){
    btn.dataset.confirm='1';btn.textContent='Sure?';
    setTimeout(()=>{if(btn.dataset.confirm==='1'){btn.dataset.confirm='';btn.textContent='✕'}},2500);
    return}
  actuallyDeleteUser(username)}
async function actuallyDeleteUser(username){
  const r=await fetch('/api/accounts-delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({username})}).then(r=>r.json());
  if(!r.ok){toast(r.error||'Could not delete user');return}
  toast('Removed '+username+' — remember to Publish');loadUsersAdmin()}
async function publishAccounts(btn){
  btn.disabled=true;btn.textContent='Publishing…';
  const r=await fetch('/api/accounts-publish',{method:'POST'}).then(r=>r.json());
  btn.disabled=false;btn.textContent='Publish Changes to Cloud';
  $('users-publish-note').textContent=r.ok?'Published — other installs pick this up automatically the next time anyone logs in.':('Could not publish: '+(r.error||'unknown error'));
  toast(r.ok?'Published to GitHub':'Publish failed — see note below the button')}

// Draggable left/right divider — shared by every .wrap/.fcwrap screen in
// the app (Document Builder's #resizer, Full Catalog Builder's own
// #fcresizer), all reading/writing the SAME --leftw CSS var + cs_leftw
// localStorage key so the chosen split width carries over between them.
// Scoped via resizer.closest('.wrap,.fcwrap') rather than a bare
// document.querySelector('.left') — both wraps coexist in the DOM at
// once (hidden, not removed, when their view isn't active), so a global
// selector would always resolve to whichever .left happens to come first
// in source order regardless of which screen is actually being dragged.
function initResizer(resizerId, onDrag){
  const resizer=$(resizerId);if(!resizer)return;
  const wrap=resizer.closest('.wrap,.fcwrap');if(!wrap)return;
  const root=document.documentElement;
  const saved=localStorage.getItem('cs_leftw');
  if(saved)root.style.setProperty('--leftw',saved+'px');
  let dragging=false,startX=0,startW=0;
  resizer.addEventListener('mousedown',e=>{
    dragging=true;startX=e.clientX;
    startW=wrap.querySelector('.left').getBoundingClientRect().width;
    resizer.classList.add('active');document.body.style.userSelect='none';
    e.preventDefault()});
  window.addEventListener('mousemove',e=>{
    if(!dragging)return;
    const w=Math.max(360,Math.min(900,startW+(e.clientX-startX)));
    root.style.setProperty('--leftw',w+'px');
    if(onDrag)onDrag()});
  window.addEventListener('mouseup',()=>{
    if(!dragging)return;
    dragging=false;resizer.classList.remove('active');document.body.style.userSelect='';
    localStorage.setItem('cs_leftw',Math.round(wrap.querySelector('.left').getBoundingClientRect().width))});
}
initResizer('resizer', renderPreviewPages);
initResizer('fcresizer', fcRenderPreviewPages);
window.addEventListener('resize',()=>{renderPreviewPages();fcRenderPreviewPages()});
checkLogin();
</script></body></html>"""

# Best-effort sync of accounts.json from R2 — see accounts.py's own
# top-of-file comment. Deliberately placed here, at true end-of-module
# (not right after photo_store.configure() further up), because load_cfg()
# itself references module-level constants (CAT_STANDARD_FILL_KEYS etc.)
# that aren't defined yet earlier in this file — calling it any sooner
# raises NameError before the module even finishes importing. Never
# blocks/crashes startup if R2 isn't configured yet or the machine is
# offline; login still works off the last cached copy (or the hardcoded
# single-admin default).
accounts.refresh_from_cloud()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    url = f"http://127.0.0.1:{port}"
    try:
        # pywebview: a real native window (uses Windows' built-in WebView2 —
        # already installed on Windows 10/11, nothing extra to download) —
        # no address bar, no tabs, no "which browser did this open in".
        # Flask has to run in a background thread here since webview.start()
        # takes over the main thread (required on some platforms) and blocks
        # until the window is closed; that thread is daemon=True so closing
        # the window ends the whole process, same as closing a normal app.
        import webview

        def _run_server():
            app.run(port=port, debug=False, use_reloader=False)
        threading.Thread(target=_run_server, daemon=True).start()
        webview.create_window("Office Tool", url, width=1400, height=900, min_size=(1000, 650))
        # pywebview defaults private_mode=True (incognito-style — wipes ALL
        # cookies the moment the window closes), which silently broke the
        # "Remember me for 30 days" login checkbox: the server-side session
        # was always correct, but the client itself never kept the cookie
        # past a single run. storage_path pins WebView2's persistent
        # profile to this install's own data folder (see engine.DATA_BASE)
        # so it survives reinstalls/updates the same way config.json does.
        webview.start(private_mode=False, storage_path=os.path.join(engine.DATA_BASE, "webview_data"))
    except ImportError:
        # pywebview isn't installed yet — falls back to the original
        # browser-tab launch so the app still runs either way. Run
        # `pip install -r requirements.txt` (pywebview is in there now) for
        # the native-window experience instead of a browser tab.
        import webbrowser
        print("\n  Office Tool running at " + url + "\n  (Press Ctrl+C to stop)\n"
              "\n  Tip: run `pip install pywebview` for a native app window instead of a browser tab.\n")
        threading.Timer(1.2, lambda: webbrowser.open(url)).start()
        app.run(port=port, debug=False)
